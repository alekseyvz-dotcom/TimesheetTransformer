# settings_manager.py
import os
import json
import hmac
import base64
import hashlib
import os as _os
import configparser
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
from pathlib import Path
from typing import Any, Dict, Optional, List
from urllib.parse import urlparse, parse_qs

import psycopg2
from psycopg2 import pool
from psycopg2.extras import RealDictCursor

from openpyxl import load_workbook
from datetime import datetime, date
from typing import Optional as Opt  # чтобы не путаться с уже использованным Optional

# ------------------------- Логика работы с пулом соединений -------------------------
db_connection_pool = None

def set_db_pool(pool):
    """Функция для установки пула соединений извне."""
    global db_connection_pool
    db_connection_pool = pool

def release_db_connection(conn):
    """Возвращает соединение обратно в пул."""
    if db_connection_pool:
        db_connection_pool.putconn(conn)

def get_db_connection():
    """Получает соединение из установленного пула."""
    if db_connection_pool is None:
         raise RuntimeError("Пул соединений не был установлен в settings_manager из главного приложения.")
    return db_connection_pool.getconn()

# ------------------------- Утилиты (копии для разрыва зависимостей) -------------------------
def month_name_ru(month: int) -> str:
    names = [
        "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
        "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"
    ]
    if 1 <= month <= 12:
        return names[month - 1]
    return str(month)
    
# ------------------------- Константы -------------------------
CONFIG_FILE = "tabel_config.ini"
CONFIG_SECTION_PATHS = "Paths"
CONFIG_SECTION_UI = "UI"
CONFIG_SECTION_INTEGR = "Integrations"

KEY_SPR = "spravochnik_path"
KEY_OUTPUT_DIR = "output_dir"
KEY_MEALS_ORDERS_DIR = "meals_orders_dir"

KEY_EXPORT_PWD = "export_password"
KEY_PLANNING_PASSWORD = "planning_password"
KEY_SELECTED_DEP = "selected_department"

KEY_ORDERS_MODE = "orders_mode"
KEY_ORDERS_WEBHOOK_URL = "orders_webhook_url"
KEY_PLANNING_ENABLED = "planning_enabled"
KEY_DRIVER_DEPARTMENTS = "driver_departments"

KEY_MEALS_MODE = "meals_mode"
KEY_MEALS_WEBHOOK_URL = "meals_webhook_url"
KEY_MEALS_WEBHOOK_TOKEN = "meals_webhook_token"
KEY_MEALS_PLANNING_ENABLED = "meals_planning_enabled"
KEY_MEALS_PLANNING_PASSWORD = "meals_planning_password"

SETTINGS_FILENAME = "settings.dat"
APP_SECRET = "KIwcVIWqzrPoBzrlTdN1lvnTcpX7sikf"


def exe_dir() -> Path:
    if getattr(__import__("sys"), "frozen", False):
        return Path(__import__("sys").executable).resolve().parent
    return Path(__file__).resolve().parent


SETTINGS_PATH = exe_dir() / SETTINGS_FILENAME
INI_PATH = exe_dir() / CONFIG_FILE

# ------------------------- ШИФРОВАНИЕ (без изменений) -------------------------

def _is_windows() -> bool:
    import platform
    return platform.system().lower().startswith("win")

def _dpapi_protect(data: bytes) -> bytes:
    import ctypes
    from ctypes import wintypes
    class DATA_BLOB(ctypes.Structure):
        _fields_ = [("cbData", wintypes.DWORD), ("pbData", ctypes.POINTER(ctypes.c_char))]
    CryptProtectData = ctypes.windll.crypt32.CryptProtectData
    CryptProtectData.argtypes = [ctypes.POINTER(DATA_BLOB), wintypes.LPWSTR, ctypes.POINTER(DATA_BLOB), ctypes.c_void_p, ctypes.c_void_p, wintypes.DWORD, ctypes.POINTER(DATA_BLOB)]
    CryptProtectData.restype = wintypes.BOOL
    in_blob = DATA_BLOB(cbData=len(data), pbData=ctypes.cast(ctypes.create_string_buffer(data), ctypes.POINTER(ctypes.c_char)))
    out_blob = DATA_BLOB()
    if not CryptProtectData(ctypes.byref(in_blob), None, None, None, None, 0, ctypes.byref(out_blob)):
        raise RuntimeError("DPAPI protect failed")
    try:
        return ctypes.string_at(out_blob.pbData, out_blob.cbData)
    finally:
        ctypes.windll.kernel32.LocalFree(out_blob.pbData)

def _dpapi_unprotect(data: bytes) -> bytes:
    import ctypes
    from ctypes import wintypes
    class DATA_BLOB(ctypes.Structure):
        _fields_ = [("cbData", wintypes.DWORD), ("pbData", ctypes.POINTER(ctypes.c_char))]
    CryptUnprotectData = ctypes.windll.crypt32.CryptUnprotectData
    CryptUnprotectData.argtypes = [ctypes.POINTER(DATA_BLOB), ctypes.POINTER(wintypes.LPWSTR), ctypes.POINTER(DATA_BLOB), ctypes.c_void_p, ctypes.c_void_p, wintypes.DWORD, ctypes.POINTER(DATA_BLOB)]
    CryptUnprotectData.restype = wintypes.BOOL
    in_blob = DATA_BLOB(cbData=len(data), pbData=ctypes.cast(ctypes.create_string_buffer(data), ctypes.POINTER(ctypes.c_char)))
    out_blob = DATA_BLOB()
    if not CryptUnprotectData(ctypes.byref(in_blob), None, None, None, None, 0, ctypes.byref(out_blob)):
        raise RuntimeError("DPAPI unprotect failed")
    try:
        return ctypes.string_at(out_blob.pbData, out_blob.cbData)
    finally:
        ctypes.windll.kernel32.LocalFree(out_blob.pbData)

def _fallback_key() -> bytes:
    return hashlib.sha256(APP_SECRET.encode("utf-8")).digest()

def _fallback_encrypt(data: bytes) -> bytes:
    key = _fallback_key()
    mac = hmac.new(key, data, hashlib.sha256).digest()
    return base64.b64encode(mac + data)

def _fallback_decrypt(packed: bytes) -> bytes:
    raw = base64.b64decode(packed)
    key = _fallback_key()
    mac = raw[:32]
    data = raw[32:]
    if not hmac.compare_digest(mac, hmac.new(key, data, hashlib.sha256).digest()):
        raise RuntimeError("Settings integrity check failed")
    return data

def _encrypt_dict(d: Dict[str, Any]) -> bytes:
    data = json.dumps(d, ensure_ascii=False, separators=(",", ":"), sort_keys=True).encode("utf-8")
    if _is_windows():
        try:
            return b"WDP1" + _dpapi_protect(data)
        except Exception:
            pass
    return b"FBK1" + _fallback_encrypt(data)

def _decrypt_dict(blob: bytes) -> Dict[str, Any]:
    if not blob: return {}
    try:
        if blob.startswith(b"WDP1"): return json.loads(_dpapi_unprotect(blob[4:]).decode("utf-8"))
        if blob.startswith(b"FBK1"): return json.loads(_fallback_decrypt(blob[4:]).decode("utf-8"))
        return json.loads(blob.decode("utf-8", errors="replace"))
    except Exception:
        return {}

# ---------------- ХРАНИЛИЩЕ НАСТРОЕК ----------------

_defaults: Dict[str, Dict[str, Any]] = {
    "Paths": {
        KEY_SPR: str(exe_dir() / "Справочник.xlsx"),
        KEY_OUTPUT_DIR: str(exe_dir() / "Объектные_табели"),
        KEY_MEALS_ORDERS_DIR: str(exe_dir() / "Заявки_питание"),
    },
    "DB": {
        "provider": "postgres",
        "database_url": "postgresql://myappuser:QweRty123!change@185.55.58.31:5432/myappdb?sslmode=disable",
        "sqlite_path": str(exe_dir() / "app_data.sqlite3"),
        "sslmode": "require",
    },
    "UI": {
        KEY_SELECTED_DEP: "Все",
    },
    "Integrations": {
        KEY_EXPORT_PWD: "2025",
        KEY_PLANNING_PASSWORD: "2025",
        KEY_ORDERS_MODE: "webhook",
        KEY_ORDERS_WEBHOOK_URL: "",
        KEY_PLANNING_ENABLED: "false",
        KEY_DRIVER_DEPARTMENTS: "",
        KEY_MEALS_MODE: "webhook",
        KEY_MEALS_WEBHOOK_URL: "",
        KEY_MEALS_WEBHOOK_TOKEN: "",
        KEY_MEALS_PLANNING_ENABLED: "true",
        KEY_MEALS_PLANNING_PASSWORD: "2025",
    },
}

_store: Dict[str, Dict[str, Any]] = {}


def _ensure_sections():
    for sec, vals in _defaults.items():
        _store.setdefault(sec, {})
        for k, v in vals.items():
            if k not in _store[sec]:
                _store[sec][k] = v


def load_settings():
    global _store
    if SETTINGS_PATH.exists():
        try:
            raw = SETTINGS_PATH.read_bytes()
            _store = _decrypt_dict(raw)
            if not isinstance(_store, dict):
                _store = {}
        except Exception:
            _store = {}
    else:
        _store = {}
    _ensure_sections()
    if not SETTINGS_PATH.exists():
        save_settings()


def save_settings():
    _ensure_sections()
    blob = _encrypt_dict(_store)
    SETTINGS_PATH.write_bytes(blob)


def migrate_from_ini_or_create():
    global _store
    cfg = configparser.ConfigParser()
    if INI_PATH.exists():
        try:
            cfg.read(INI_PATH, encoding="utf-8")
        except Exception:
            pass
    _store = json.loads(json.dumps(_defaults))
    for sec in _store.keys():
        if cfg.has_section(sec):
            for key in _store[sec].keys():
                val = cfg.get(sec, key, fallback=_store[sec][key])
                _store[sec][key] = val
    save_settings()


# Публичные API

def ensure_config():
    load_settings()


class _ProxyConfig:
    def get(self, section: str, key: str, fallback: Opt[str] = None) -> str:
        try:
            v = _store.get(section, {}).get(key, None)
            if v is None or v == "":
                return fallback if fallback is not None else ""
            return str(v)
        except Exception:
            return fallback if fallback is not None else ""


def read_config() -> _ProxyConfig:
    ensure_config()
    return _ProxyConfig()


def write_config(_cfg=None):
    save_settings()

def get_departments_list() -> List[Dict]:
    """Список подразделений для комбобоксов."""
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT id, name FROM departments ORDER BY name;")
            return list(cur.fetchall())
    finally:
        if conn:
            release_db_connection(conn)

def get_spr_path_from_config() -> Path:
    ensure_config()
    raw = _store["Paths"].get(KEY_SPR) or _defaults["Paths"][KEY_SPR]
    return Path(os.path.expandvars(str(raw)))


def get_output_dir_from_config() -> Path:
    ensure_config()
    raw = _store["Paths"].get(KEY_OUTPUT_DIR) or _defaults["Paths"][KEY_OUTPUT_DIR]
    return Path(os.path.expandvars(str(raw)))


def get_meals_orders_dir_from_config() -> Path:
    ensure_config()
    raw = _store["Paths"].get(KEY_MEALS_ORDERS_DIR) or _defaults["Paths"][KEY_MEALS_ORDERS_DIR]
    return Path(os.path.expandvars(str(raw)))


def get_export_password_from_config() -> str:
    ensure_config()
    return str(_store["Integrations"].get(KEY_EXPORT_PWD, _defaults["Integrations"][KEY_EXPORT_PWD]))


def get_selected_department_from_config() -> str:
    ensure_config()
    return str(_store["UI"].get(KEY_SELECTED_DEP, _defaults["UI"][KEY_SELECTED_DEP]))


def set_selected_department_in_config(dep: str):
    ensure_config()
    _store["UI"][KEY_SELECTED_DEP] = dep or "Все"
    save_settings()


def get_db_provider() -> str:
    ensure_config()
    return str(_store["DB"].get("provider", _defaults["DB"]["provider"]))


def get_database_url() -> str:
    ensure_config()
    return str(_store["DB"].get("database_url", ""))


def get_sqlite_path() -> str:
    ensure_config()
    return str(_store["DB"].get("sqlite_path", _defaults["DB"]["sqlite_path"]))


def get_db_sslmode() -> str:
    ensure_config()
    return str(_store["DB"].get("sslmode", _defaults["DB"]["sslmode"]))


def get_meals_mode_from_config() -> str:
    ensure_config()
    return str(
        _store["Integrations"].get(KEY_MEALS_MODE, _defaults["Integrations"][KEY_MEALS_MODE])
    ).strip().lower()


def set_meals_mode_in_config(mode: str):
    ensure_config()
    _store["Integrations"][KEY_MEALS_MODE] = mode or "webhook"
    save_settings()


def get_meals_webhook_url_from_config() -> str:
    ensure_config()
    return str(_store["Integrations"].get(KEY_MEALS_WEBHOOK_URL, _defaults["Integrations"][KEY_MEALS_WEBHOOK_URL]))


def set_meals_webhook_url_in_config(url: str):
    ensure_config()
    _store["Integrations"][KEY_MEALS_WEBHOOK_URL] = url or ""
    save_settings()


def get_meals_webhook_token_from_config() -> str:
    ensure_config()
    return str(_store["Integrations"].get(KEY_MEALS_WEBHOOK_TOKEN, _defaults["Integrations"][KEY_MEALS_WEBHOOK_TOKEN]))


def get_meals_planning_enabled_from_config() -> bool:
    ensure_config()
    v = str(
        _store["Integrations"].get(
            KEY_MEALS_PLANNING_ENABLED,
            _defaults["Integrations"][KEY_MEALS_PLANNING_ENABLED],
        )
    ).strip().lower()
    return v in ("1", "true", "yes", "on")


def set_meals_planning_enabled_in_config(enabled: bool):
    ensure_config()
    _store["Integrations"][KEY_MEALS_PLANNING_ENABLED] = "true" if enabled else "false"
    save_settings()


def get_meals_planning_password_from_config() -> str:
    ensure_config()
    return str(
        _store["Integrations"].get(
            KEY_MEALS_PLANNING_PASSWORD,
            _defaults["Integrations"][KEY_MEALS_PLANNING_PASSWORD],
        )
    )


def set_meals_planning_password_in_config(pwd: str):
    ensure_config()
    _store["Integrations"][KEY_MEALS_PLANNING_PASSWORD] = pwd or ""
    save_settings()


def set_meals_webhook_token_in_config(tok: str):
    ensure_config()
    _store["Integrations"][KEY_MEALS_WEBHOOK_TOKEN] = tok or ""
    save_settings()

# ---------------- ИМПОРТ СОТРУДНИКОВ/ОБЪЕКТОВ ИЗ EXCEL ----------------

def _s_val(val) -> str:
    """Вспомогательная функция приведения значений ячеек к строке."""
    if val is None: return ""
    if isinstance(val, float) and val.is_integer(): val = int(val)
    return str(val).strip()


def import_employees_from_excel(path: Path) -> int:
    """
    Импортирует сотрудников из Excel-файла (как 'ШТАТ на ...').
    Обновляет таблицы departments и employees.
    """
    if not path.exists(): raise FileNotFoundError(f"Файл не найден: {path}")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    hdr = [_s_val(c).lower() for c in header_row]

    def col_idx(substr: str) -> Opt[int]:
        substr = substr.lower()
        for i, h in enumerate(hdr):
            if substr in h: return i
        return None

    idx_tbn = col_idx("табельный номер")
    idx_fio = col_idx("сотрудник")
    idx_pos = col_idx("должность")
    idx_dep = col_idx("подразделение")
    idx_dismissal = col_idx("увольн")

    if idx_fio is None or idx_tbn is None:
        raise RuntimeError("Не найдены обязательные колонки 'Табельный номер' и/или 'Сотрудник'")

    conn = None
    processed = 0
    try:
        conn = get_db_connection()
        with conn:
            with conn.cursor() as cur:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row: continue
                    fio = _s_val(row[idx_fio]) if idx_fio < len(row) else ""
                    tbn = _s_val(row[idx_tbn]) if idx_tbn < len(row) else ""
                    pos = _s_val(row[idx_pos]) if idx_pos is not None and idx_pos < len(row) else ""
                    dep_name = _s_val(row[idx_dep]) if idx_dep is not None and idx_dep < len(row) else ""
                    dismissal_raw = row[idx_dismissal] if idx_dismissal is not None and idx_dismissal < len(row) else None
                    if not fio and not tbn: continue
                    is_fired = bool(dismissal_raw and _s_val(dismissal_raw))
                    department_id = None
                    if dep_name:
                        cur.execute("SELECT id FROM departments WHERE name = %s", (dep_name,))
                        r = cur.fetchone()
                        if r: department_id = r[0]
                        else:
                            cur.execute("INSERT INTO departments (name) VALUES (%s) RETURNING id", (dep_name,))
                            department_id = cur.fetchone()[0]

                    cur.execute("SELECT id FROM employees WHERE tbn = %s", (tbn,)) if tbn else cur.execute("SELECT id FROM employees WHERE fio = %s", (fio,))
                    r = cur.fetchone()
                    if r:
                        cur.execute("UPDATE employees SET fio = %s, tbn = %s, position = %s, department_id = %s, is_fired = %s WHERE id = %s", (fio or None, tbn or None, pos or None, department_id, is_fired, r[0]))
                    else:
                        cur.execute("INSERT INTO employees (fio, tbn, position, department_id, is_fired) VALUES (%s, %s, %s, %s, %s)", (fio or None, tbn or None, pos or None, department_id, is_fired))
                    processed += 1
    finally:
        if conn:
            release_db_connection(conn)
    return processed

def import_objects_from_excel(path: Path) -> int:
    """
    Импортирует/обновляет объекты из Excel 'Справочник программ и объектов...'.
    """
    if not path.exists():
        raise FileNotFoundError(f"Файл не найден: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header_row_idx = None
    header_row = []

    # Ищем строку заголовка по ключевой фразе
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=40, values_only=True), start=1):
        if row and any("id (код) номер объекта" in _s_val(c).lower() for c in row):
            header_row_idx, header_row = i, [_s_val(c).lower() for c in row]
            break

    if header_row_idx is None:
        raise RuntimeError("Не найдена строка заголовка с колонкой 'ID (код) номер объекта'")

    def col_idx(substr: str) -> Opt[int]:
        substr = substr.lower()
        for i, h in enumerate(header_row):
            if substr in h:
                return i
        return None

    # Получаем индексы всех нужных колонок
    idx_excel_id = col_idx("id (код) номер объекта")
    idx_year = col_idx("год")
    idx_program = col_idx("наименование программы")
    idx_customer = col_idx("наименование заказчика")
    idx_addr = col_idx("адрес")
    idx_contract_no = col_idx("№ договора")
    idx_contract_date = col_idx("дата договора")
    idx_short_name = col_idx("сокращенное наименование объекта")
    idx_department_name = col_idx("подразделение")
    idx_contract_type = col_idx("тип договора")

    if idx_excel_id is None or idx_addr is None:
        raise RuntimeError("Не найдены обязательные колонки 'ID (код) номер объекта' и/или 'Адрес объекта'")

    conn = None
    processed = 0
    try:
        conn = get_db_connection()
        with conn:
            with conn.cursor() as cur:
                for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                    # ... (код чтения данных из строки row остается без изменений) ...

                    excel_id = _s_val(row[idx_excel_id]) if idx_excel_id < len(row) else ""
                    if not excel_id: continue

                    year = _s_val(row[idx_year]) if idx_year is not None and idx_year < len(row) else ""
                    program = _s_val(row[idx_program]) if idx_program is not None and idx_program < len(row) else ""
                    customer_name = _s_val(row[idx_customer]) if idx_customer is not None and idx_customer < len(row) else ""
                    address = _s_val(row[idx_addr]) if idx_addr < len(row) else ""
                    contract_number = _s_val(row[idx_contract_no]) if idx_contract_no is not None and idx_contract_no < len(row) else ""
                    
                    contract_date_raw = row[idx_contract_date] if idx_contract_date is not None and idx_contract_date < len(row) else None
                    contract_date_val: Opt[date] = None
                    if isinstance(contract_date_raw, datetime):
                        contract_date_val = contract_date_raw.date()
                    
                    short_name = _s_val(row[idx_short_name]) if idx_short_name is not None and idx_short_name < len(row) else ""
                    executor_department = _s_val(row[idx_department_name]) if idx_department_name is not None and idx_department_name < len(row) else ""
                    contract_type = _s_val(row[idx_contract_type]) if idx_contract_type is not None and idx_contract_type < len(row) else ""
                    
                    # ИСПРАВЛЕНО: имя таблицы на 'objects'
                    cur.execute("SELECT id FROM objects WHERE excel_id = %s", (excel_id,))
                    existing_record = cur.fetchone()

                    if existing_record:
                        db_id = existing_record[0]
                        # ИСПРАВЛЕНО: имя таблицы на 'objects'
                        cur.execute(
                            """
                            UPDATE objects SET
                                address = %s, year = %s, program_name = %s, customer_name = %s, 
                                contract_number = %s, contract_date = %s, short_name = %s, 
                                executor_department = %s, contract_type = %s, updated_at = NOW()
                            WHERE id = %s
                            """,
                            (address, year or None, program or None, customer_name or None, contract_number or None, 
                             contract_date_val, short_name or None, executor_department or None, contract_type or None, db_id)
                        )
                    else:
                        # ИСПРАВЛЕНО: имя таблицы на 'objects'
                        cur.execute(
                            """
                            INSERT INTO objects 
                                (excel_id, address, year, program_name, customer_name, contract_number, contract_date, 
                                short_name, executor_department, contract_type, status)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            """,
                            (excel_id, address, year or None, program or None, customer_name or None, contract_number or None,
                            contract_date_val, short_name or None, executor_department or None, contract_type or None, 'Новый')
                        )
                    processed += 1
    finally:
        if conn:
            release_db_connection(conn)
    return processed

def _hash_password(password: str, salt: Optional[bytes] = None) -> str:
    """Хеширует пароль с использованием PBKDF2."""
    if salt is None:
        salt = _os.urandom(16)
    iterations = 260000
    dk = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, iterations)
    return f"pbkdf2_sha256${iterations}${salt.hex()}${dk.hex()}"

def _verify_password(password: str, stored_hash: str) -> bool:
    """Проверяет пароль по хешу."""
    try:
        if stored_hash.startswith("pbkdf2_sha256$"):
            _, it_str, salt_hex, hash_hex = stored_hash.split("$", 3)
            iterations = int(it_str)
            salt = bytes.fromhex(salt_hex)
            dk = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, iterations)
            return dk.hex() == hash_hex
        return password == stored_hash
    except Exception:
        return False

def get_permissions_catalog() -> List[Dict]:
    """Справочник всех прав (для окна выдачи прав)."""
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT code, title, group_name
                FROM public.app_permissions
                ORDER BY group_name, title;
                """
            )
            return list(cur.fetchall())
    finally:
        if conn:
            release_db_connection(conn)


def get_user_permissions(user_id: int) -> set[str]:
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute(
                "SELECT perm_code FROM public.app_user_permissions WHERE user_id = %s",
                (user_id,),
            )
            return {r[0] for r in cur.fetchall()}
    finally:
        if conn:
            release_db_connection(conn)


def set_user_permissions(user_id: int, perm_codes: List[str]):
    """Полная перезапись прав пользователя."""
    perm_codes = sorted({(p or "").strip() for p in perm_codes if p and str(p).strip()})
    conn = None
    try:
        conn = get_db_connection()
        with conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM public.app_user_permissions WHERE user_id = %s", (user_id,))
                if perm_codes:
                    cur.executemany(
                        """
                        INSERT INTO public.app_user_permissions(user_id, perm_code)
                        VALUES (%s, %s)
                        ON CONFLICT DO NOTHING
                        """,
                        [(user_id, p) for p in perm_codes],
                    )
    finally:
        if conn:
            release_db_connection(conn)


def grant_default_permissions(user_id: int):
    """Минимальные права новому пользователю."""
    conn = None
    try:
        conn = get_db_connection()
        with conn, conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO public.app_user_permissions(user_id, perm_code)
                VALUES (%s, %s)
                ON CONFLICT DO NOTHING
                """,
                (user_id, "page.home"),
            )
            # если у вас есть settings-permission для админов, не включаем его по умолчанию
    finally:
        if conn:
            release_db_connection(conn)

def get_roles_list() -> List[Dict]:
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT code, name FROM roles ORDER BY name;")
            return list(cur.fetchall())
    finally:
        if conn:
            release_db_connection(conn)

def get_app_users() -> List[Dict]:
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT u.id,
                       u.username,
                       u.full_name,
                       u."role",
                       u.is_active,
                       u.department_id,
                       d.name AS department_name
                FROM app_users u
                LEFT JOIN departments d ON d.id = u.department_id
                ORDER BY u.username;
                """
            )
            return list(cur.fetchall())
    finally:
        if conn:
            release_db_connection(conn)

def create_app_user(
    username: str,
    password: str,
    full_name: str,
    role_code: str,
    is_active: bool = True,
    department_id: Optional[int] = None,
):
    username, full_name, role_code = (
        (username or "").strip(),
        (full_name or "").strip(),
        (role_code or "").strip().lower(),
    )
    if not username:
        raise ValueError("Логин не может быть пустым")
    if not password:
        raise ValueError("Пароль не может быть пустым")
    if not role_code:
        raise ValueError("Роль не указана")

    pwd_hash = _hash_password(password)
    conn = None
    try:
        conn = get_db_connection()
        with conn, conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO app_users (username, password_hash, is_active, full_name, "role", department_id)
                VALUES (%s, %s, %s, %s, %s, %s)
                RETURNING id
                """,
                (username, pwd_hash, is_active, full_name, role_code, department_id),
            )
            new_id = cur.fetchone()[0]
        return new_id
    finally:
        if conn:
            release_db_connection(conn)

def update_app_user(
    user_id: int,
    username: str,
    full_name: str,
    role_code: str,
    is_active: bool,
    new_password: Optional[str] = None,
    department_id: Optional[int] = None,
):
    username, full_name, role_code = (
        (username or "").strip(),
        (full_name or "").strip(),
        (role_code or "").strip().lower(),
    )
    if not username:
        raise ValueError("Логин не может быть пустым")
    if not role_code:
        raise ValueError("Роль не указана")

    pwd_hash = _hash_password(new_password) if new_password else None

    conn = None
    try:
        conn = get_db_connection()
        with conn, conn.cursor() as cur:
            if pwd_hash:
                cur.execute(
                    """
                    UPDATE app_users
                    SET username = %s,
                        full_name = %s,
                        "role" = %s,
                        is_active = %s,
                        password_hash = %s,
                        department_id = %s
                    WHERE id = %s
                    """,
                    (username, full_name, role_code, is_active, pwd_hash, department_id, user_id),
                )
            else:
                cur.execute(
                    """
                    UPDATE app_users
                    SET username = %s,
                        full_name = %s,
                        "role" = %s,
                        is_active = %s,
                        department_id = %s
                    WHERE id = %s
                    """,
                    (username, full_name, role_code, is_active, department_id, user_id),
                )
    finally:
        if conn:
            release_db_connection(conn)

def delete_app_user(user_id: int):
    conn = None
    try:
        conn = get_db_connection()
        with conn, conn.cursor() as cur:
            cur.execute("DELETE FROM app_users WHERE id = %s", (user_id,))
    finally:
        if conn:
            release_db_connection(conn)

# ---------------- UI ДЛЯ ПОЛЬЗОВАТЕЛЕЙ ----------------

_vars: Dict[str, Dict[str, Any]] = {}


def _add_context_menu(widget: tk.Widget):
    menu = tk.Menu(widget, tearoff=0)
    menu.add_command(label="Вырезать", command=lambda: widget.event_generate("<<Cut>>"))
    menu.add_command(label="Копировать", command=lambda: widget.event_generate("<<Copy>>"))
    menu.add_command(label="Вставить", command=lambda: widget.event_generate("<<Paste>>"))
    menu.add_separator()
    menu.add_command(label="Выделить всё", command=lambda: widget.event_generate("<<SelectAll>>"))

    def show_menu(event):
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    widget.bind("<Button-3>", show_menu)


class CreateUserDialog(simpledialog.Dialog):
    def __init__(self, parent):
        self.result = None
        super().__init__(parent, title="Создать пользователя")

    def body(self, master):
        tk.Label(master, text="Логин:").grid(row=0, column=0, sticky="e", padx=(0, 6), pady=4)
        self.ent_username = ttk.Entry(master, width=26)
        self.ent_username.grid(row=0, column=1, sticky="w", pady=4)

        tk.Label(master, text="ФИО:").grid(row=1, column=0, sticky="e", padx=(0, 6), pady=4)
        self.ent_fullname = ttk.Entry(master, width=26)
        self.ent_fullname.grid(row=1, column=1, sticky="w", pady=4)

        tk.Label(master, text="Пароль:").grid(row=2, column=0, sticky="e", padx=(0, 6), pady=4)
        self.ent_pwd = ttk.Entry(master, width=26, show="*")
        self.ent_pwd.grid(row=2, column=1, sticky="w", pady=4)

        tk.Label(master, text="Роль:").grid(row=3, column=0, sticky="e", padx=(0, 6), pady=4)
        try:
            roles = get_roles_list()
        except Exception as e:
            messagebox.showerror("Роли", f"Ошибка чтения списка ролей:\n{e}", parent=self)
            roles = []

        self.role_map = {f'{r["name"]} ({r["code"]})': r["code"] for r in roles}
        self.cmb_role = ttk.Combobox(master, state="readonly", width=24, values=list(self.role_map.keys()))
        self.cmb_role.grid(row=3, column=1, sticky="w", pady=4)
        if self.role_map:
            self.cmb_role.current(0)

        # --- Подразделение ---
        tk.Label(master, text="Подразделение:").grid(row=4, column=0, sticky="e", padx=(0, 6), pady=4)
        try:
            deps = get_departments_list()
        except Exception as e:
            messagebox.showerror("Подразделения", f"Ошибка чтения списка подразделений:\n{e}", parent=self)
            deps = []

        self.dep_map = {"(нет)": None}
        for d in deps:
            self.dep_map[d["name"]] = d["id"]

        self.cmb_dep = ttk.Combobox(master, state="readonly", width=24, values=list(self.dep_map.keys()))
        self.cmb_dep.grid(row=4, column=1, sticky="w", pady=4)
        if self.dep_map:
            self.cmb_dep.set("(нет)")

        return self.ent_username

    def validate(self):
        u = self.ent_username.get().strip()
        p = self.ent_pwd.get().strip()
        if not u:
            messagebox.showwarning("Создать пользователя", "Укажите логин.", parent=self)
            return False
        if not p:
            messagebox.showwarning("Создать пользователя", "Укажите пароль.", parent=self)
            return False
        if not self.cmb_role.get():
            messagebox.showwarning("Создать пользователя", "Выберите роль.", parent=self)
            return False
        return True

    def apply(self):
        sel_dep = self.cmb_dep.get()
        self.result = {
            "username": self.ent_username.get().strip(),
            "full_name": self.ent_fullname.get().strip(),
            "password": self.ent_pwd.get().strip(),
            "role": self.role_map[self.cmb_role.get()],
            "department_id": self.dep_map.get(sel_dep),
        }

class EditUserDialog(simpledialog.Dialog):
    def __init__(self, parent, user: dict):
        self.user = user
        self.result = None
        super().__init__(parent, title=f"Редактировать пользователя: {user.get('username', '')}")

    def body(self, master):
        tk.Label(master, text="Логин:").grid(row=0, column=0, sticky="e", padx=(0, 6), pady=4)
        self.ent_username = ttk.Entry(master, width=26)
        self.ent_username.grid(row=0, column=1, sticky="w", pady=4)
        self.ent_username.insert(0, self.user.get("username", ""))

        tk.Label(master, text="ФИО:").grid(row=1, column=0, sticky="e", padx=(0, 6), pady=4)
        self.ent_fullname = ttk.Entry(master, width=26)
        self.ent_fullname.grid(row=1, column=1, sticky="w", pady=4)
        self.ent_fullname.insert(0, self.user.get("full_name") or "")

        tk.Label(master, text="Новый пароль (оставьте пустым, чтобы не менять):").grid(
            row=2, column=0, sticky="e", padx=(0, 6), pady=4
        )
        self.ent_pwd = ttk.Entry(master, width=26, show="*")
        self.ent_pwd.grid(row=2, column=1, sticky="w", pady=4)

        tk.Label(master, text="Роль:").grid(row=3, column=0, sticky="e", padx=(0, 6), pady=4)
        try:
            roles = get_roles_list()
        except Exception as e:
            messagebox.showerror("Роли", f"Ошибка чтения списка ролей:\n{e}", parent=self)
            roles = []

        self.role_map = {f'{r["name"]} ({r["code"]})': r["code"] for r in roles}
        self.cmb_role = ttk.Combobox(master, state="readonly", width=24, values=list(self.role_map.keys()))
        self.cmb_role.grid(row=3, column=1, sticky="w", pady=4)

        current_role_code = (self.user.get("role") or "").lower()
        idx = 0
        for i, (label, code) in enumerate(self.role_map.items()):
            if code.lower() == current_role_code:
                idx = i
                break
        if self.role_map:
            self.cmb_role.current(idx)

        # --- Подразделение ---
        tk.Label(master, text="Подразделение:").grid(row=4, column=0, sticky="e", padx=(0, 6), pady=4)
        try:
            deps = get_departments_list()
        except Exception as e:
            messagebox.showerror("Подразделения", f"Ошибка чтения списка подразделений:\n{e}", parent=self)
            deps = []

        self.dep_map = {"(нет)": None}
        for d in deps:
            self.dep_map[d["name"]] = d["id"]

        self.cmb_dep = ttk.Combobox(master, state="readonly", width=24, values=list(self.dep_map.keys()))
        self.cmb_dep.grid(row=4, column=1, sticky="w", pady=4)

        # выбрать текущее подразделение, если есть
        current_dep_name = self.user.get("department_name")
        if current_dep_name and current_dep_name in self.dep_map:
            self.cmb_dep.set(current_dep_name)
        else:
            self.cmb_dep.set("(нет)")

        tk.Label(master, text="Активен:").grid(row=5, column=0, sticky="e", padx=(0, 6), pady=4)
        self.var_active = tk.BooleanVar(value=bool(self.user.get("is_active")))
        self.cb_active = ttk.Checkbutton(master, variable=self.var_active)
        self.cb_active.grid(row=5, column=1, sticky="w", pady=4)

        return self.ent_username

    def validate(self):
        u = self.ent_username.get().strip()
        if not u:
            messagebox.showwarning("Редактировать пользователя", "Укажите логин.", parent=self)
            return False
        if not self.cmb_role.get():
            messagebox.showwarning("Редактировать пользователя", "Выберите роль.", parent=self)
            return False
        return True

    def apply(self):
        sel_dep = self.cmb_dep.get()
        self.result = {
            "id": self.user["id"],
            "username": self.ent_username.get().strip(),
            "full_name": self.ent_fullname.get().strip(),
            "password": self.ent_pwd.get().strip(),  # может быть пустым
            "role": self.role_map[self.cmb_role.get()],
            "is_active": self.var_active.get(),
            "department_id": self.dep_map.get(sel_dep),
        }

class UsersPage(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        self._build_ui()
        self.reload_users()

    def _build_ui(self):
        top = tk.Frame(self)
        top.pack(fill="x", padx=8, pady=8)

        ttk.Button(top, text="Создать пользователя", command=self._on_create_user).pack(side="left")
        ttk.Button(top, text="Права...", command=self._on_permissions).pack(side="left", padx=4)
        ttk.Button(top, text="Изменить", command=self._on_edit_user).pack(side="left", padx=4)
        ttk.Button(top, text="Удалить", command=self._on_delete_user).pack(side="left")

        self.tree = ttk.Treeview(
            self,
            columns=("id", "username", "full_name", "role", "department", "is_active"),
            show="headings",
            height=12,
        )
        self.tree.heading("id", text="ID")
        self.tree.heading("username", text="Логин")
        self.tree.heading("full_name", text="ФИО")
        self.tree.heading("role", text="Роль")
        self.tree.heading("department", text="Подразделение")
        self.tree.heading("is_active", text="Активен")

        self.tree.column("id", width=40, anchor="center")
        self.tree.column("username", width=120)
        self.tree.column("full_name", width=200)
        self.tree.column("role", width=90)
        self.tree.column("department", width=160)
        self.tree.column("is_active", width=70, anchor="center")

        self.tree.pack(fill="both", expand=True, padx=8, pady=(0, 8))

    def _on_permissions(self):
        user = self._get_selected_user()
        if not user:
            return
        EditUserPermissionsDialog(self, user_id=user["id"], username=user["username"])


    def reload_users(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        try:
            users = get_app_users()
        except Exception as e:
            messagebox.showerror("Пользователи", f"Ошибка загрузки списка:\n{e}", parent=self)
            return

        for u in users:
            self.tree.insert(
                "",
                "end",
                values=(
                    u["id"],
                    u["username"],
                    u.get("full_name") or "",
                    u.get("role") or "",
                    u.get("department_name") or "",
                    "Да" if u.get("is_active") else "Нет",
                ),
            )

    def _get_selected_user(self) -> Optional[dict]:
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Пользователи", "Выберите пользователя в списке.", parent=self)
            return None
        item_id = sel[0]
        values = self.tree.item(item_id, "values")
        # порядок: id, username, full_name, role, department_name, is_active
        user = {
            "id": int(values[0]),
            "username": values[1],
            "full_name": values[2],
            "role": values[3],
            "department_name": values[4],
            "is_active": (values[5] == "Да"),
        }
        return user

    def _on_edit_user(self):
        user = self._get_selected_user()
        if not user:
            return

        dlg = EditUserDialog(self, user)
        if not dlg.result:
            return

        data = dlg.result
        try:
            update_app_user(
                user_id=data["id"],
                username=data["username"],
                full_name=data["full_name"],
                role_code=data["role"],
                is_active=data["is_active"],
                new_password=data["password"] or None,
                department_id=data.get("department_id"),
            )
        except Exception as e:
            messagebox.showerror("Редактировать пользователя", f"Ошибка сохранения:\n{e}", parent=self)
            return

        self.reload_users()
        messagebox.showinfo("Редактировать пользователя", "Изменения сохранены.", parent=self)


    def _on_create_user(self):
        dlg = CreateUserDialog(self)
        if not dlg.result:
            return
        data = dlg.result
        try:
            new_id = create_app_user(
                username=data["username"],
                password=data["password"],
                full_name=data["full_name"],
                role_code=data["role"],
                is_active=True,
                department_id=data.get("department_id"),
            )
            grant_default_permissions(new_id)  # минимум (например page.home)
        except Exception as e:
            messagebox.showerror("Создать пользователя", f"Ошибка создания пользователя:\n{e}", parent=self)
            return
    
        self.reload_users()
        # Сразу открыть права — лучший UX
        EditUserPermissionsDialog(self, user_id=new_id, username=data["username"])

    def _on_delete_user(self):
        user = self._get_selected_user()
        if not user:
            return

        if not messagebox.askyesno(
            "Удалить пользователя",
            f"Удалить пользователя '{user['username']}'?",
            parent=self,
        ):
            return

        try:
            delete_app_user(user["id"])
        except Exception as e:
            messagebox.showerror("Удалить пользователя", f"Ошибка удаления:\n{e}", parent=self)
            return

        self.reload_users()
        messagebox.showinfo("Удалить пользователя", "Пользователь удалён.", parent=self)

class EditUserPermissionsDialog(tk.Toplevel):
    def __init__(self, parent, user_id: int, username: str):
        super().__init__(parent)
        self.title(f"Права доступа — {username}")
        self.geometry("760x520")
        self.minsize(720, 480)

        self.user_id = user_id
        self.username = username

        try:
            self.catalog = get_permissions_catalog()
            self.current = get_user_permissions(user_id)
        except Exception as e:
            messagebox.showerror("Права", f"Ошибка загрузки прав:\n{e}", parent=self)
            self.destroy()
            return

        # code -> (meta, var, widget)
        self._vars: Dict[str, tk.BooleanVar] = {}
        self._cb_widgets: Dict[str, ttk.Checkbutton] = {}
        self._meta: Dict[str, Dict] = {}

        # UI top: поиск + кнопки
        top = ttk.Frame(self)
        top.pack(fill="x", padx=10, pady=10)

        ttk.Label(top, text="Поиск:").pack(side="left")
        self.search_var = tk.StringVar(value="")
        ent = ttk.Entry(top, textvariable=self.search_var, width=40)
        ent.pack(side="left", padx=(6, 10))
        ent.focus_set()

        ttk.Button(top, text="Выбрать всё (видимое)", command=self._select_all_visible).pack(side="left", padx=4)
        ttk.Button(top, text="Снять всё (видимое)", command=self._clear_all_visible).pack(side="left", padx=4)

        # Notebook по группам
        self.nb = ttk.Notebook(self)
        self.nb.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        # Группируем права по group_name
        groups: Dict[str, List[Dict]] = {}
        for p in self.catalog:
            g = (p.get("group_name") or "").strip() or "Другое"
            groups.setdefault(g, []).append(p)

        # Строим вкладки
        self._tabs: Dict[str, ttk.Frame] = {}
        for group_name in sorted(groups.keys()):
            tab = ttk.Frame(self.nb)
            tab.pack(fill="both", expand=True)
            self.nb.add(tab, text=group_name)
            self._tabs[group_name] = tab

            # Вкладка: прокручиваемая область
            canvas = tk.Canvas(tab, highlightthickness=0)
            vsb = ttk.Scrollbar(tab, orient="vertical", command=canvas.yview)
            canvas.configure(yscrollcommand=vsb.set)

            vsb.pack(side="right", fill="y")
            canvas.pack(side="left", fill="both", expand=True)

            inner = ttk.Frame(canvas)
            inner_id = canvas.create_window((0, 0), window=inner, anchor="nw")

            def _on_configure(event, c=canvas, inner_frame=inner):
                c.configure(scrollregion=c.bbox("all"))

            def _on_canvas_configure(event, c=canvas, win_id=inner_id):
                c.itemconfigure(win_id, width=event.width)

            inner.bind("<Configure>", _on_configure)
            canvas.bind("<Configure>", _on_canvas_configure)

            # Кнопки для вкладки
            bar = ttk.Frame(inner)
            bar.grid(row=0, column=0, sticky="ew", padx=6, pady=(6, 8))
            bar.columnconfigure(0, weight=1)

            ttk.Label(bar, text="").grid(row=0, column=0, sticky="w")
            ttk.Button(bar, text="Выбрать всё", command=lambda g=group_name: self._select_group(g)).grid(row=0, column=1, padx=4)
            ttk.Button(bar, text="Снять всё", command=lambda g=group_name: self._clear_group(g)).grid(row=0, column=2, padx=4)

            # Чекбоксы
            row = 1
            for p in groups[group_name]:
                code = p["code"]
                title = p["title"]
                var = tk.BooleanVar(value=(code in self.current))
                self._vars[code] = var
                self._meta[code] = p

                cb = ttk.Checkbutton(inner, text=title, variable=var)
                cb.grid(row=row, column=0, sticky="w", padx=10, pady=2)
                self._cb_widgets[code] = cb
                row += 1

        # Нижние кнопки
        bottom = ttk.Frame(self)
        bottom.pack(fill="x", padx=10, pady=(0, 10))
        ttk.Button(bottom, text="Сохранить", command=self._save).pack(side="right", padx=4)
        ttk.Button(bottom, text="Отмена", command=self.destroy).pack(side="right")

        # Поиск
        self.search_var.trace_add("write", lambda *_: self._apply_filter())
        self._apply_filter()

        self.transient(parent)
        self.grab_set()

    def _visible_codes(self) -> List[str]:
        needle = self.search_var.get().strip().lower()
        if not needle:
            return list(self._vars.keys())
        out = []
        for code, p in self._meta.items():
            hay = f"{p.get('title','')} {p.get('code','')} {p.get('group_name','')}".lower()
            if needle in hay:
                out.append(code)
        return out

    def _apply_filter(self):
        needle = self.search_var.get().strip().lower()
        for code, cb in self._cb_widgets.items():
            p = self._meta[code]
            hay = f"{p.get('title','')} {p.get('code','')} {p.get('group_name','')}".lower()
            show = (not needle) or (needle in hay)
            cb.grid_remove() if not show else cb.grid()

    def _select_all_visible(self):
        for code in self._visible_codes():
            self._vars[code].set(True)

    def _clear_all_visible(self):
        for code in self._visible_codes():
            self._vars[code].set(False)

    def _select_group(self, group_name: str):
        for code, p in self._meta.items():
            g = (p.get("group_name") or "").strip() or "Другое"
            if g == group_name:
                self._vars[code].set(True)

    def _clear_group(self, group_name: str):
        for code, p in self._meta.items():
            g = (p.get("group_name") or "").strip() or "Другое"
            if g == group_name:
                self._vars[code].set(False)

    def _save(self):
        selected = [code for code, var in self._vars.items() if var.get()]
        try:
            set_user_permissions(self.user_id, selected)
        except Exception as e:
            messagebox.showerror("Права", f"Ошибка сохранения:\n{e}", parent=self)
            return
        messagebox.showinfo("Права", "Права сохранены.", parent=self)
        self.destroy()

# ---------------- UI ОКНО НАСТРОЕК ----------------

def open_settings_window(parent: tk.Tk):
    ensure_config()

    win = tk.Toplevel(parent)
    win.title("Настройки")
    win.resizable(False, False)
    nb = ttk.Notebook(win)
    nb.pack(fill="both", expand=True, padx=10, pady=10)

    # Настройки папок
    tab_paths = ttk.Frame(nb)
    nb.add(tab_paths, text="Настройки папок")
    _mk_entry_with_btn(tab_paths, "Справочник (xlsx):", "Paths", KEY_SPR, is_dir=False, row=0)
    _mk_entry_with_btn(tab_paths, "Папка табелей:", "Paths", KEY_OUTPUT_DIR, is_dir=True, row=1)
    _mk_entry_with_btn(tab_paths, "Папка заявок на питание:", "Paths", KEY_MEALS_ORDERS_DIR, is_dir=True, row=2)

    # Основное
    tab_ui = ttk.Frame(nb)
    nb.add(tab_ui, text="Основное")
    _mk_entry(tab_ui, "Подразделение по умолчанию:", "UI", KEY_SELECTED_DEP, row=0, width=40)
    _mk_entry(
        tab_ui,
        "Подразделения водителей:",
        "Integrations",
        KEY_DRIVER_DEPARTMENTS,
        row=1,
        width=64,
    )

    # База данных
    tab_db = ttk.Frame(nb)
    nb.add(tab_db, text="База данных")

    ttk.Label(tab_db, text="Провайдер:").grid(row=0, column=0, sticky="e", padx=(6, 6), pady=4)
    provider_var = tk.StringVar(value=str(_store["DB"].get("provider", "sqlite")))
    cmb_provider = ttk.Combobox(
        tab_db, textvariable=provider_var, state="readonly", width=18, values=["sqlite", "postgres", "mysql"]
    )
    cmb_provider.grid(row=0, column=1, sticky="w", padx=(0, 6), pady=4)
    _vars.setdefault("DB", {})["provider"] = provider_var

    ttk.Label(tab_db, text="Строка подключения (DATABASE_URL):").grid(
        row=1, column=0, sticky="e", padx=(6, 6), pady=4
    )
    v_url = tk.StringVar(value=str(_store["DB"].get("database_url", "")))
    ent_url = ttk.Entry(tab_db, textvariable=v_url, width=64)
    ent_url.grid(row=1, column=1, sticky="w", padx=(0, 6), pady=4, columnspan=2)
    _add_context_menu(ent_url)
    _vars.setdefault("DB", {})["database_url"] = v_url

    ttk.Label(tab_db, text="SQLite файл:").grid(row=2, column=0, sticky="e", padx=(6, 6), pady=4)
    v_sqlite = tk.StringVar(value=str(_store["DB"].get("sqlite_path", _defaults["DB"]["sqlite_path"])))
    ent_sqlite = ttk.Entry(tab_db, textvariable=v_sqlite, width=56)
    ent_sqlite.grid(row=2, column=1, sticky="w", padx=(0, 6), pady=4)
    _add_context_menu(ent_sqlite)

    def browse_sqlite():
        p = filedialog.asksaveasfilename(
            title="Файл SQLite",
            defaultextension=".sqlite3",
            filetypes=[("SQLite DB", "*.sqlite3 *.db"), ("Все файлы", "*.*")],
        )
        if p:
            v_sqlite.set(p)

    ttk.Button(tab_db, text="...", width=3, command=browse_sqlite).grid(row=2, column=2, sticky="w")
    _vars.setdefault("DB", {})["sqlite_path"] = v_sqlite

    ttk.Label(tab_db, text="SSL mode (Postgres):").grid(row=3, column=0, sticky="e", padx=(6, 6), pady=4)
    v_ssl = tk.StringVar(value=str(_store["DB"].get("sslmode", "require")))
    cmb_ssl = ttk.Combobox(
        tab_db, textvariable=v_ssl, state="readonly", width=18, values=["require", "verify-full", "prefer", "disable"]
    )
    cmb_ssl.grid(row=3, column=1, sticky="w", padx=(0, 6), pady=4)
    _vars.setdefault("DB", {})["sslmode"] = v_ssl

    def _toggle_db_fields(*_):
        prov = provider_var.get()
        if prov == "sqlite":
            ent_url.configure(state="disabled")
            ent_sqlite.configure(state="normal")
        else:
            ent_url.configure(state="normal")
            ent_sqlite.configure(state="disabled")

    provider_var.trace_add("write", _toggle_db_fields)
    _toggle_db_fields()

    # Пользователи
    tab_users = ttk.Frame(nb)
    nb.add(tab_users, text="Пользователи")
    users_page = UsersPage(tab_users)
    users_page.pack(fill="both", expand=True)

    # Данные (импорт)
    tab_data = ttk.Frame(nb)
    nb.add(tab_data, text="Данные")

    row_idx = 0
    ttk.Label(
        tab_data,
        text="Загрузка сотрудников из Excel в базу данных:\n"
        "Ожидается файл со штатным расписанием\n"
        "(колонки: 'Табельный номер (с префиксами)', 'Сотрудник', "
        "'Должность', 'Подразделение', 'Дата увольнения').",
    ).grid(row=row_idx, column=0, columnspan=3, sticky="w", padx=6, pady=(6, 4))
    row_idx += 1

    def on_import_employees():
        file_path = filedialog.askopenfilename(
            parent=win,
            title="Выберите файл со штатным расписанием",
            filetypes=[("Excel", "*.xlsx;*.xls"), ("Все файлы", "*.*")],
        )
        if not file_path:
            return
        try:
            cnt = import_employees_from_excel(Path(file_path))
            messagebox.showinfo(
                "Импорт сотрудников",
                f"Импорт завершён.\nОбработано записей: {cnt}",
                parent=win,
            )
        except Exception as e:
            messagebox.showerror("Импорт сотрудников", f"Ошибка при импорте:\n{e}", parent=win)

    ttk.Button(tab_data, text="Загрузить сотрудников из Excel...", command=on_import_employees).grid(
        row=row_idx, column=0, sticky="w", padx=6, pady=(0, 8)
    )
    row_idx += 1

    ttk.Separator(tab_data, orient="horizontal").grid(
        row=row_idx, column=0, columnspan=3, sticky="ew", padx=6, pady=4
    )
    row_idx += 1

    ttk.Label(
        tab_data,
        text="Загрузка объектов из Excel в базу данных:\n"
        "Ожидается файл 'Справочник программ и объектов' "
        "(колонки: ID объекта, год, программа, заказчик, адрес, № договора,\n"
        "дата договора, сокращённое наименование объекта, подразделение исполнителя, тип договора).",
    ).grid(row=row_idx, column=0, columnspan=3, sticky="w", padx=6, pady=(6, 4))
    row_idx += 1

    def on_import_objects():
        file_path = filedialog.askopenfilename(
            parent=win,
            title="Выберите файл справочника объектов",
            filetypes=[("Excel", "*.xlsx;*.xls"), ("Все файлы", "*.*")],
        )
        if not file_path:
            return
        try:
            cnt = import_objects_from_excel(Path(file_path))
            messagebox.showinfo(
                "Импорт объектов",
                f"Импорт завершён.\nОбработано записей: {cnt}",
                parent=win,
            )
        except Exception as e:
            messagebox.showerror("Импорт объектов", f"Ошибка при импорте:\n{e}", parent=win)

    ttk.Button(tab_data, text="Загрузить объекты из Excel...", command=on_import_objects).grid(
        row=row_idx, column=0, sticky="w", padx=6, pady=(0, 8)
    )

    # Кнопки
    btns = ttk.Frame(win)
    btns.pack(fill="x", padx=10, pady=(0, 10))
    ttk.Button(
        btns,
        text="Сохранить",
        command=lambda: (_save_from_vars(win), messagebox.showinfo("Настройки", "Сохранено")),
    ).pack(side="right", padx=4)
    ttk.Button(btns, text="Отмена", command=win.destroy).pack(side="right")

    # Центрирование окна
    try:
        win.update_idletasks()
        px, py = parent.winfo_rootx(), parent.winfo_rooty()
        pw, ph = parent.winfo_width(), parent.winfo_height()
        sw, sh = win.winfo_width(), win.winfo_height()
        win.geometry(f"+{px + (pw - sw) // 2}+{py + (ph - sh) // 2}")
    except Exception:
        pass


def _mk_entry(parent, label, section, key, row, width=30, show=None):
    ttk.Label(parent, text=label).grid(row=row, column=0, sticky="e", padx=(6, 6), pady=4)
    v = tk.StringVar(value=str(_store.get(section, {}).get(key, _defaults[section][key])))
    ent = ttk.Entry(parent, textvariable=v, width=width, show=show)
    ent.grid(row=row, column=1, sticky="w", padx=(0, 6), pady=4)
    _add_context_menu(ent)
    _vars.setdefault(section, {})[key] = v


def _mk_check(parent, label, section, key, row):
    ttk.Label(parent, text=label).grid(row=row, column=0, sticky="e", padx=(6, 6), pady=4)
    cur = (
        str(_store.get(section, {}).get(key, _defaults[section][key]))
        .strip()
        .lower()
        in ("1", "true", "yes", "on")
    )
    v = tk.BooleanVar(value=cur)
    cb = ttk.Checkbutton(parent, variable=v)
    cb.grid(row=row, column=1, sticky="w", padx=(0, 6), pady=4)
    _vars.setdefault(section, {})[key] = v


def _mk_entry_with_btn(parent, label, section, key, is_dir, row):
    ttk.Label(parent, text=label).grid(row=row, column=0, sticky="e", padx=(6, 6), pady=4)
    v = tk.StringVar(value=str(_store.get(section, {}).get(key, _defaults[section][key])))
    ent = ttk.Entry(parent, textvariable=v, width=60)
    ent.grid(row=row, column=1, sticky="w", padx=(0, 6), pady=4)
    _add_context_menu(ent)

    def browse():
        if is_dir:
            d = filedialog.askdirectory(title="Выбор папки")
            if d:
                v.set(d)
        else:
            p = filedialog.askopenfilename(
                title="Выбор файла",
                filetypes=[("Excel", "*.xlsx;*.xls"), ("Все файлы", "*.*")],
            )
            if p:
                v.set(p)

    ttk.Button(parent, text="...", width=3, command=browse).grid(row=row, column=2, sticky="w")
    _vars.setdefault(section, {})[key] = v


def _save_from_vars(win):
    for sec, kv in _vars.items():
        for k, var in kv.items():
            if isinstance(var, tk.BooleanVar):
                _store[sec][k] = "true" if var.get() else "false"
            else:
                _store[sec][k] = str(var.get())
    save_settings()
    try:
        win.destroy()
    except Exception:
        pass
