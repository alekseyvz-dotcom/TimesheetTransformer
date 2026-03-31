from __future__ import annotations

import logging
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional, Sequence, Tuple

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter

from virtual_timesheet_grid import VirtualTimesheetGrid

from timesheet_common import (
    MAX_HOURS_PER_DAY,
    TS_COLORS,
    best_fio_match_with_score,
    calc_row_totals,
    calc_rows_summary,
    compute_day_summary_from_events,
    deduplicate_timesheet_rows,
    ensure_current_month_date,
    find_suspicious_cells,
    format_hours_for_cell,
    format_summary_value,
    make_row_key,
    month_days,
    month_name_ru,
    normalize_hours_list,
    normalize_row_record,
    normalize_spaces,
    normalize_tbn,
    parse_timesheet_cell,
    parse_hours_value,
    read_skud_events_from_xlsx,
    safe_filename,
    validate_rows_before_save,
)
from timesheet_db import (
    db_cursor,
    find_duplicate_employees_for_timesheet,
    find_timesheet_header_id,
    load_all_timesheet_headers,
    load_brigadier_assignments_for_department,
    load_brigadier_names_for_department,
    load_brigadiers_map_for_header,
    load_employees_from_db,
    load_objects_short_for_timesheet,
    load_timesheet_full_by_header_id,
    load_timesheet_rows_by_header_id,
    load_timesheet_rows_for_copy_from_db,
    load_timesheet_rows_from_db,
    load_user_timesheet_headers,
    replace_timesheet_rows,
    set_db_pool,
    upsert_timesheet_header,
)
from timesheet_dialogs import (
    AutoCompleteCombobox,
    BatchAddDialog,
    CopyFromDialog,
    HoursFillDialog,
    SelectDateDialog,
    SelectEmployeesDialog,
    SelectObjectIdDialog,
    SkudMappingReviewDialog,
    SuspiciousHoursWarningDialog,
    TimeForSelectedDialog,
)

logger = logging.getLogger(__name__)

try:
    from settings_manager import (
        get_output_dir_from_config,
        get_selected_department_from_config,
        set_selected_department_in_config,
    )
except Exception:
    get_output_dir_from_config = None
    get_selected_department_from_config = None
    set_selected_department_in_config = None

# Приводим цвета модуля к новой светлой оболочке
TS_COLORS = dict(TS_COLORS)
TS_COLORS["bg"] = "#edf1f5"
TS_COLORS["panel"] = "#f7f9fb"
TS_COLORS["accent_light"] = "#eef3f8"
TS_COLORS.setdefault("border", "#c9d3df")
TS_COLORS.setdefault("accent", "#2f74c0")
TS_COLORS.setdefault("warning", "#c97a20")
TS_COLORS.setdefault("btn_save_bg", "#2f74c0")
TS_COLORS.setdefault("btn_save_fg", "#ffffff")


def _set_timesheet_tab_title(app_ref, key: str, header: Dict[str, Any]):
    """
    Обновляет заголовок вкладки после открытия конкретного табеля.
    """
    try:
        year = int(header.get("year") or 0)
        month = int(header.get("month") or 0)
        month_ru = month_name_ru(month) if 1 <= month <= 12 else str(month or "")
        oid = normalize_spaces(header.get("object_id") or "")
        addr = normalize_spaces(header.get("object_addr") or "")
        dep = normalize_spaces(header.get("department") or "")

        if oid and addr:
            title = f"[{oid}] {month_ru} {year}"
        elif addr:
            title = f"{month_ru} {year}"
        else:
            title = f"Табель {month_ru} {year}"

        if dep:
            title = f"{title} · {dep}"

        if hasattr(app_ref, "_tab_titles"):
            app_ref._tab_titles[key] = title

        frame = getattr(app_ref, "_tab_frames", {}).get(key)
        notebook = getattr(app_ref, "notebook", None)
        if frame is not None and notebook is not None:
            notebook.tab(frame, text=title)
    except Exception:
        logger.exception("Не удалось обновить заголовок вкладки табеля")

PRINT_TITLE_FILL = PatternFill("solid", fgColor="DCE6F1")
PRINT_META_FILL = PatternFill("solid", fgColor="EAF2F8")
PRINT_HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
PRINT_TOTAL_FILL = PatternFill("solid", fgColor="E2F0D9")

THIN_BLACK = Side(style="thin", color="000000")
MEDIUM_BLACK = Side(style="medium", color="000000")

BORDER_THIN = Border(left=THIN_BLACK, right=THIN_BLACK, top=THIN_BLACK, bottom=THIN_BLACK)
BORDER_EMPTY = Border()


def _excel_safe_value(value: Any) -> Any:
    return "" if value is None else value


def _apply_print_style(
    cell,
    *,
    bold: bool = False,
    size: int = 9,
    h: str = "center",
    v: str = "center",
    wrap: bool = True,
    border: Border = BORDER_THIN,
    fill=None,
):
    cell.font = Font(name="Segoe UI", size=size, bold=bold)
    cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    cell.border = border
    if fill is not None:
        cell.fill = fill


def _set_range_fill_and_border(ws, row1: int, col1: int, row2: int, col2: int, *, fill=None, bold: bool = False):
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            cell = ws.cell(r, c)
            _apply_print_style(cell, bold=bold, fill=fill)


def _set_outer_medium_border(ws, row1: int, col1: int, row2: int, col2: int):
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            cell = ws.cell(r, c)
            cell.border = Border(
                left=MEDIUM_BLACK if c == col1 else cell.border.left,
                right=MEDIUM_BLACK if c == col2 else cell.border.right,
                top=MEDIUM_BLACK if r == row1 else cell.border.top,
                bottom=MEDIUM_BLACK if r == row2 else cell.border.bottom,
            )


def _setup_print_sheet_params(ws, *, last_col_letter: str, last_row: int, title_rows: str = "$1:$7"):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    ws.page_margins = PageMargins(
        left=0.25,
        right=0.25,
        top=0.4,
        bottom=0.45,
        header=0.2,
        footer=0.2,
    )

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "E8"
    ws.print_title_rows = title_rows
    ws.print_area = f"A1:{last_col_letter}{last_row}"

    ws.oddHeader.center.text = "&\"Segoe UI,Bold\"&12 Табель учета рабочего времени"
    ws.oddFooter.left.text = "&\"Segoe UI\"&8 Сформировано автоматически"
    ws.oddFooter.center.text = "&\"Segoe UI\"&8 Страница &[Page] из &N"
    ws.oddFooter.right.text = f"&\"Segoe UI\"&8 {datetime.now().strftime('%d.%m.%Y %H:%M')}"


def build_printable_timesheet_sheet(
    ws,
    *,
    year: int,
    month: int,
    object_addr: str,
    object_id: str,
    department: str,
    rows: list[dict[str, Any]],
    brig_map: dict[str, str] | None = None,
    prepared_by: str = "",
):
    brig_map = brig_map or {}

    month_ru = month_name_ru(month) if 1 <= month <= 12 else str(month)
    days_in_month = month_days(year, month)

    fixed_headers = ["№", "ФИО", "Таб. №", "Бригадир"]
    day_headers = [str(i) for i in range(1, days_in_month + 1)]
    total_headers = ["Дни", "Часы", "Ночные", "Пер. день", "Пер. ночь"]
    headers = fixed_headers + day_headers + total_headers

    total_cols = len(headers)
    last_col_letter = get_column_letter(total_cols)

    # --- Заголовок документа ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws["A1"]
    c.value = "ТАБЕЛЬ УЧЕТА РАБОЧЕГО ВРЕМЕНИ"
    _apply_print_style(c, bold=True, size=14, h="center", border=BORDER_EMPTY, fill=PRINT_TITLE_FILL)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    c = ws["A2"]
    c.value = f"за {month_ru} {year} г."
    _apply_print_style(c, bold=True, size=11, h="center", border=BORDER_EMPTY)

    meta_split = min(12, total_cols)
    right_meta_start = meta_split + 1

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=meta_split)
    c = ws["A3"]
    c.value = f"Объект: {object_addr or '-'}"
    _apply_print_style(c, bold=True, h="left", fill=PRINT_META_FILL)

    if right_meta_start <= total_cols:
        ws.merge_cells(start_row=3, start_column=right_meta_start, end_row=3, end_column=total_cols)
        c = ws.cell(3, right_meta_start)
        c.value = f"ID объекта: {object_id or '-'}"
        _apply_print_style(c, bold=True, h="left", fill=PRINT_META_FILL)

    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=meta_split)
    c = ws["A4"]
    c.value = f"Подразделение: {department or '-'}"
    _apply_print_style(c, bold=True, h="left", fill=PRINT_META_FILL)

    if right_meta_start <= total_cols:
        ws.merge_cells(start_row=4, start_column=right_meta_start, end_row=4, end_column=total_cols)
        c = ws.cell(4, right_meta_start)
        c.value = f"Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        _apply_print_style(c, h="left", fill=PRINT_META_FILL)

    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=total_cols)
    c = ws["A5"]
    c.value = "Условные обозначения в ячейках дней: часы или буквенные коды отсутствий/режимов работы."
    _apply_print_style(c, size=8, h="left", border=BORDER_EMPTY)

    # --- Размеры строк ---
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 16

    # --- Заголовок таблицы ---
    header_row = 7
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(header_row, col_idx, title)
        _apply_print_style(cell, bold=True, fill=PRINT_HEADER_FILL)

    ws.row_dimensions[header_row].height = 30

    # --- Ширины колонок ---
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 11
    ws.column_dimensions["D"].width = 24

    first_day_col = 5
    last_day_col = first_day_col + days_in_month - 1
    for col_idx in range(first_day_col, last_day_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 4.2

    totals_start_col = last_day_col + 1
    ws.column_dimensions[get_column_letter(totals_start_col)].width = 8
    ws.column_dimensions[get_column_letter(totals_start_col + 1)].width = 10
    ws.column_dimensions[get_column_letter(totals_start_col + 2)].width = 10
    ws.column_dimensions[get_column_letter(totals_start_col + 3)].width = 11
    ws.column_dimensions[get_column_letter(totals_start_col + 4)].width = 11

    # --- Данные ---
    current_row = header_row + 1

    normalized_rows: list[dict[str, Any]] = []
    for rec in rows:
        norm = normalize_row_record(rec, year, month)
        norm["hours"] = normalize_hours_list(norm.get("hours"), year, month)
        norm["_totals"] = norm.get("_totals") or calc_row_totals(norm["hours"], year, month)
        normalized_rows.append(norm)

    for idx, rec in enumerate(normalized_rows, start=1):
        fio = normalize_spaces(rec.get("fio") or "")
        tbn = normalize_tbn(rec.get("tbn"))
        hours = rec.get("hours") or []
        totals = rec.get("_totals") or {}
        brig_fio = brig_map.get(tbn, "") if tbn else ""

        row_values = [
            idx,
            fio,
            tbn,
            brig_fio,
            *[_excel_safe_value(v) for v in hours[:days_in_month]],
            _excel_safe_value(format_summary_value(totals.get("days"))),
            _excel_safe_value(format_summary_value(totals.get("hours"))),
            _excel_safe_value(format_summary_value(totals.get("night_hours"))),
            _excel_safe_value(format_summary_value(totals.get("ot_day"))),
            _excel_safe_value(format_summary_value(totals.get("ot_night"))),
        ]

        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(current_row, col_idx, value)
            if col_idx in (2, 4):
                _apply_print_style(cell, h="left")
            else:
                _apply_print_style(cell, h="center")

        ws.row_dimensions[current_row].height = 21
        current_row += 1

    # --- Итого ---
    summary = calc_rows_summary(normalized_rows, year, month)

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    total_cell = ws.cell(current_row, 1)
    total_cell.value = "ИТОГО"
    _apply_print_style(total_cell, bold=True, fill=PRINT_TOTAL_FILL)

    for col_idx in range(5, totals_start_col):
        cell = ws.cell(current_row, col_idx, "")
        _apply_print_style(cell, bold=True, fill=PRINT_TOTAL_FILL)

    summary_values = [
        format_summary_value(summary.get("days")),
        format_summary_value(summary.get("hours")),
        format_summary_value(summary.get("night_hours")),
        format_summary_value(summary.get("ot_day")),
        format_summary_value(summary.get("ot_night")),
    ]
    for offset, value in enumerate(summary_values):
        cell = ws.cell(current_row, totals_start_col + offset, value)
        _apply_print_style(cell, bold=True, fill=PRINT_TOTAL_FILL)

    table_last_row = current_row

    # --- Рамка таблицы ---
    _set_outer_medium_border(ws, header_row, 1, table_last_row, total_cols)

    # --- Подписи ---
    sign_row = table_last_row + 3

    left_sign_end = min(12, total_cols)
    right_sign_start = left_sign_end + 1

    ws.merge_cells(start_row=sign_row, start_column=1, end_row=sign_row, end_column=left_sign_end)
    c = ws.cell(sign_row, 1)
    c.value = f"Составил: {prepared_by or '__________________'}    Подпись: __________________"
    _apply_print_style(c, h="left", border=BORDER_EMPTY)

    if right_sign_start <= total_cols:
        ws.merge_cells(start_row=sign_row, start_column=right_sign_start, end_row=sign_row, end_column=total_cols)
        c = ws.cell(sign_row, right_sign_start)
        c.value = "Дата: __________________"
        _apply_print_style(c, h="left", border=BORDER_EMPTY)

    ws.merge_cells(start_row=sign_row + 1, start_column=1, end_row=sign_row + 1, end_column=left_sign_end)
    c = ws.cell(sign_row + 1, 1)
    c.value = "Проверил: _________________________________    Подпись: __________________"
    _apply_print_style(c, h="left", border=BORDER_EMPTY)

    if right_sign_start <= total_cols:
        ws.merge_cells(start_row=sign_row + 1, start_column=right_sign_start, end_row=sign_row + 1, end_column=total_cols)
        c = ws.cell(sign_row + 1, right_sign_start)
        c.value = "Дата: __________________"
        _apply_print_style(c, h="left", border=BORDER_EMPTY)

    ws.merge_cells(start_row=sign_row + 2, start_column=1, end_row=sign_row + 2, end_column=left_sign_end)
    c = ws.cell(sign_row + 2, 1)
    c.value = "Утвердил: ________________________________    Подпись: __________________"
    _apply_print_style(c, h="left", border=BORDER_EMPTY)

    if right_sign_start <= total_cols:
        ws.merge_cells(start_row=sign_row + 2, start_column=right_sign_start, end_row=sign_row + 2, end_column=total_cols)
        c = ws.cell(sign_row + 2, right_sign_start)
        c.value = "Дата: __________________"
        _apply_print_style(c, h="left", border=BORDER_EMPTY)

    # --- Параметры печати ---
    _setup_print_sheet_params(
        ws,
        last_col_letter=last_col_letter,
        last_row=sign_row + 2,
        title_rows="$1:$7",
    )

# ============================================================
# Основная страница табеля
# ============================================================

class TimesheetPage(tk.Frame):
    COLPX = {"fio": 220, "tbn": 100, "day": 36, "days": 52, "hours": 58, "ot_day": 58, "ot_night": 58, "del": 66}
    MIN_FIO_PX = 140
    MAX_FIO_PX = 280

    def __init__(
        self,
        master,
        app_ref,
        init_object_id: Optional[str] = None,
        init_object_addr: Optional[str] = None,
        init_department: Optional[str] = None,
        init_year: Optional[int] = None,
        init_month: Optional[int] = None,
        read_only: bool = False,
        owner_user_id: Optional[int] = None,
        init_header_id: Optional[int] = None,
    ):
        super().__init__(master, bg=TS_COLORS["bg"])

        self.app_ref = app_ref
        self.read_only = bool(read_only)
        self.owner_user_id = owner_user_id
        self._active_header_id: Optional[int] = int(init_header_id) if init_header_id else None

        self._init_object_id = normalize_spaces(init_object_id or "")
        self._init_object_addr = normalize_spaces(init_object_addr or "")
        self._init_department = normalize_spaces(init_department or "")
        self._init_year = init_year
        self._init_month = init_month

        self._initializing = True
        self._suppress_object_id_dialog = False
        self._suppress_events = False

        self._dirty = False
        self._loaded_context: dict[str, Any] = {}

        self._auto_save_job = None
        self._auto_save_delay_ms = 8000
        self._fit_job = None
        self._filter_job = None

        self.out_dir = self._resolve_output_dir()
        self.out_dir.mkdir(parents=True, exist_ok=True)

        self.employees: List[Tuple[str, str, str, str]] = []
        self.objects_full: List[Tuple[str, str, str]] = []

        self.emp_names: List[str] = []
        self.departments: List[str] = ["Все"]
        self.address_options: List[str] = []

        self._fio_to_employees: dict[str, list[tuple[str, str, str, str]]] = {}
        self._employee_display_to_data: dict[str, tuple[str, str, str, str]] = {}
        self.allowed_fio_names: set[str] = set()

        self.model_rows_all: List[Dict[str, Any]] = []
        self.model_rows: List[Dict[str, Any]] = []
        self._selected_row_keys: set[str] = set()

        self.var_filter = tk.StringVar()
        self.var_brigadier = tk.StringVar(value="Все")
        self._brig_assign: dict[str, str | None] = {}
        self._brig_names: dict[str, str] = {}

        self._load_spr_data_from_db()
        self._build_ui()
        self._init_ts_values()

        self.bind("<Configure>", self._on_window_configure)
        self.after(150, self._auto_fit_columns)

    # --------------------------------------------------------
    # Lifecycle
    # --------------------------------------------------------

    def destroy(self):
        for attr in ("_auto_save_job", "_fit_job", "_filter_job"):
            job = getattr(self, attr, None)
            if job is not None:
                try:
                    self.after_cancel(job)
                except Exception:
                    pass
                setattr(self, attr, None)
        super().destroy()

    def _resolve_output_dir(self) -> Path:
        try:
            if get_output_dir_from_config:
                value = get_output_dir_from_config()
                if value:
                    return Path(value)
        except Exception:
            logger.exception("Не удалось получить каталог output из конфигурации")
        return Path("./output")

    # --------------------------------------------------------
    # Helpers
    # --------------------------------------------------------

def _update_selected_count(self):
    try:
        count = len(self._selected_row_keys)
        self.lbl_selected_count.config(text=f"Выбрано строк: {count}")
    except Exception:
        pass
    
    def _safe_current_user_id(self) -> Optional[int]:
        if self.owner_user_id:
            return int(self.owner_user_id)
        try:
            user = getattr(self.app_ref, "current_user", None) or {}
            uid = user.get("id")
            return int(uid) if uid else None
        except Exception:
            return None

    def _safe_get_year(self) -> int:
        try:
            year = int(self.spn_year.get())
            if 2000 <= year <= 2100:
                return year
        except Exception:
            pass
        return datetime.now().year

    def get_year_month(self) -> Tuple[int, int]:
        year = self._safe_get_year()
        month = self.cmb_month.current() + 1
        if not (1 <= month <= 12):
            month = datetime.now().month
        return year, month

    def _capture_current_context(self) -> dict[str, Any]:
        year, month = self.get_year_month()
        return {
            "department": normalize_spaces(self.cmb_department.get() or ""),
            "year": year,
            "month": month,
            "object_addr": normalize_spaces(self.cmb_address.get() or ""),
            "object_id": normalize_spaces(self.cmb_object_id.get() or ""),
            "header_id": self._active_header_id,
        }

    def _restore_controls_to_loaded_context(self):
        if not self._loaded_context:
            return

        self._suppress_events = True
        try:
            dep = self._loaded_context.get("department", "")
            year = int(self._loaded_context.get("year") or datetime.now().year)
            month = int(self._loaded_context.get("month") or datetime.now().month)
            addr = self._loaded_context.get("object_addr", "") or ""
            oid = self._loaded_context.get("object_id", "") or ""

            self._ensure_department_option(dep)
            self.cmb_department.set(dep or "Все")

            self.spn_year.delete(0, "end")
            self.spn_year.insert(0, str(year))

            if 1 <= month <= 12:
                self.cmb_month.current(month - 1)

            self._ensure_address_option(addr)
            self.cmb_address.set(addr)
            self._sync_object_id_values_silent()

            if oid:
                values = list(self.cmb_object_id.cget("values") or [])
                if oid not in values:
                    values.append(oid)
                    self.cmb_object_id.config(values=values)
                self.cmb_object_id.set(oid)
            else:
                self.cmb_object_id.set("")
        finally:
            self._suppress_events = False

    def _set_status_text(self, text: str, fg: str = "#bbdefb"):
        try:
            self.lbl_auto_save.config(text=text, fg=fg)
        except Exception:
            pass

    def _mark_dirty(self):
        if self.read_only:
            return
        self._dirty = True
        self._set_status_text("Есть несохранённые изменения", fg="#ffe082")

    def _mark_saved(self, auto: bool):
        self._dirty = False
        now = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
        if auto:
            self._set_status_text(f"Последнее авто‑сохранение: {now}", fg="#bbdefb")
        else:
            self._set_status_text(f"Сохранено: {now}", fg="#bbdefb")

    def _mark_save_error(self, text: str):
        self._set_status_text(text, fg="#ffccbc")

    def _confirm_leave_with_unsaved(self) -> bool:
        if self.read_only or not self._dirty:
            return True

        answer = messagebox.askyesnocancel(
            "Несохранённые изменения",
            "Есть несохранённые изменения.\n\nСохранить перед переключением?",
            parent=self,
        )
        if answer is None:
            return False

        if answer is True:
            return self._save_all_internal(show_messages=True, is_auto=False)

        return True

    def _resolve_employee_display(self, fio: str, tbn: str) -> str:
        fio = normalize_spaces(fio)
        tbn = normalize_tbn(tbn)
        return f"{fio} [{tbn}]" if tbn else fio

    def _find_unique_employee_by_fio(
        self,
        fio: str,
        department: Optional[str] = None,
    ) -> Optional[Tuple[str, str, str, str]]:
        fio_norm = normalize_spaces(fio)
        dep_norm = normalize_spaces(department or "")
        matches = []

        for emp in self.employees:
            emp_fio, emp_tbn, emp_pos, emp_dep = emp
            if normalize_spaces(emp_fio) != fio_norm:
                continue
            if dep_norm and dep_norm != "Все" and normalize_spaces(emp_dep) != dep_norm:
                continue
            matches.append(emp)

        return matches[0] if len(matches) == 1 else None

    def _find_unique_row_by_fio(self, fio: str) -> Optional[Dict[str, Any]]:
        fio_norm = normalize_spaces(fio)
        matches = [row for row in self.model_rows_all if normalize_spaces(row.get("fio") or "") == fio_norm]
        return matches[0] if len(matches) == 1 else None

    def _grid_selected(self) -> set[int]:
        if hasattr(self, "grid"):
            try:
                return self.grid.get_selected_indices()
            except Exception:
                return set()
        return set()

    def _grid_refresh(self, rows_changed: bool = False):
        if not hasattr(self, "grid"):
            return
        try:
            if rows_changed:
                self.grid.set_rows(self.model_rows)
            else:
                self.grid.refresh()
        except Exception:
            logger.exception("Ошибка обновления грида")

    def _schedule_auto_save(self):
        if self.read_only:
            return
        if self._auto_save_job is not None:
            try:
                self.after_cancel(self._auto_save_job)
            except Exception:
                pass
            self._auto_save_job = None
        self._auto_save_job = self.after(self._auto_save_delay_ms, self._auto_save_callback)

    def _auto_save_callback(self):
        self._auto_save_job = None
        self._save_all_internal(show_messages=False, is_auto=True)

    def _row_key(self, rec: Dict[str, Any]) -> str:
        return make_row_key(rec.get("fio", ""), rec.get("tbn", ""))
    
    def _remember_grid_selection(self):
        """
        Запоминает выделение текущих видимых строк в self._selected_row_keys
        через устойчивые ключи (ФИО + таб.№).
        """
        if not hasattr(self, "grid"):
            return
    
        try:
            selected = self.grid.get_selected_indices()
        except Exception:
            selected = set()
    
        for idx in selected:
            if 0 <= idx < len(self.model_rows):
                rec = self.model_rows[idx]
                self._selected_row_keys.add(self._row_key(rec))
    
    def _restore_grid_selection(self):
        """
        Восстанавливает выделение на текущем self.model_rows
        по ранее сохранённым ключам.
        """
        if not hasattr(self, "grid"):
            return
    
        indices: set[int] = set()
        alive_keys: set[str] = set()
    
        for idx, rec in enumerate(self.model_rows):
            key = self._row_key(rec)
            if key in self._selected_row_keys:
                indices.add(idx)
                alive_keys.add(key)
    
        # Оставляем только реально существующие выбранные строки
        # (если кого-то удалили из табеля — он уйдёт из selection)
        self._selected_row_keys &= {self._row_key(rec) for rec in self.model_rows_all}
    
        try:
            self.grid.set_selected_indices(indices)
        except Exception:
            pass
    
        self._update_selected_count()

    def _on_grid_selection_changed(self, selected_indices: set[int]):
        """
        Обновляет сохранённый набор выбранных сотрудников
        при обычных кликах пользователя по строкам грида.
        """
        visible_keys = {self._row_key(rec) for rec in self.model_rows}
    
        # Удаляем из сохранённого набора только ключи текущего видимого списка,
        # дальше добавим актуальные заново
        self._selected_row_keys -= visible_keys
    
        for idx in selected_indices:
            if 0 <= idx < len(self.model_rows):
                self._selected_row_keys.add(self._row_key(self.model_rows[idx]))
    
        self._update_selected_count()

    # --------------------------------------------------------
    # Справочники
    # --------------------------------------------------------

    def _load_spr_data_from_db(self):
        self.employees = load_employees_from_db()
        self.objects_full = load_objects_short_for_timesheet()

        self._fio_to_employees.clear()
        for emp in self.employees:
            fio = normalize_spaces(emp[0])
            self._fio_to_employees.setdefault(fio, []).append(emp)

        self.emp_names = sorted({normalize_spaces(fio) for fio, _, _, _ in self.employees if normalize_spaces(fio)})

        deps = sorted(
            {
                normalize_spaces(dep)
                for _, _, _, dep in self.employees
                if normalize_spaces(dep)
            }
        )
        self.departments = ["Все"] + deps

        self.address_options = sorted(
            {
                normalize_spaces(addr)
                for _, addr, _ in self.objects_full
                if normalize_spaces(addr)
            }
        )

    def _ensure_department_option(self, dep: str):
        dep = normalize_spaces(dep)
        if not dep:
            return
        values = list(self.cmb_department.cget("values") or self.departments or ["Все"])
        if dep not in values:
            values.append(dep)
            values = sorted(set(values), key=lambda x: (x != "Все", x))
            self.cmb_department.config(values=values)

    def _ensure_address_option(self, addr: str):
        addr = normalize_spaces(addr)
        if not addr:
            return
        if addr not in self.address_options:
            self.address_options.append(addr)
            self.address_options = sorted(set(self.address_options))
            self.cmb_address.set_completion_list(self.address_options)

    def _refresh_employee_selector_for_department(self, dep_sel: str):
        dep_sel = normalize_spaces(dep_sel)
        self._employee_display_to_data.clear()

        if dep_sel == "Все":
            allowed = list(self.employees)
        else:
            allowed = [emp for emp in self.employees if normalize_spaces(emp[3]) == dep_sel]

        for fio, tbn, pos, dep in allowed:
            display = self._resolve_employee_display(fio, tbn)
            self._employee_display_to_data[display] = (fio, tbn, pos, dep)

        self.allowed_fio_names = {normalize_spaces(fio) for fio, _, _, _ in allowed if normalize_spaces(fio)}

    # --------------------------------------------------------
    # UI
    # --------------------------------------------------------

    def _build_ui(self):
        self._build_ts_header()
        self._build_ts_top_form()
        self._build_ts_toolbar()
        self._build_ts_filter_bar()
        self._build_ts_grid()
        self._build_ts_bottom()

    def _ts_lbl(self, parent, text: str, row: int, col: int = 0, required: bool = False, **grid_kw):
        display = f"{text}{'  *' if required else ''}:"
        fg = TS_COLORS["warning"] if required else "#333"
        tk.Label(
            parent,
            text=display,
            font=("Segoe UI", 9),
            bg=TS_COLORS["panel"],
            fg=fg,
            anchor="e",
        ).grid(row=row, column=col, sticky="e", padx=(0, 6), pady=3, **grid_kw)

    def _build_ts_header(self):
        hdr = tk.Frame(self, bg=TS_COLORS["accent"], pady=4)
        hdr.pack(fill="x")

        title = "👁 Просмотр табеля" if self.read_only else "📋 Объектный табель"
        tk.Label(
            hdr,
            text=title,
            font=("Segoe UI", 12, "bold"),
            bg=TS_COLORS["accent"],
            fg="white",
            padx=12,
        ).pack(side="left")

        self.lbl_auto_save = tk.Label(
            hdr,
            text="Авто‑сохранение: нет",
            font=("Segoe UI", 8),
            bg=TS_COLORS["accent"],
            fg="#bbdefb",
            padx=10,
        )
        self.lbl_auto_save.pack(side="right")

    def _build_ts_top_form(self):
        outer = tk.Frame(self, bg=TS_COLORS["bg"])
        outer.pack(fill="x", padx=10, pady=(6, 2))
        outer.grid_columnconfigure(0, weight=1)
        outer.grid_columnconfigure(1, weight=2)

        self._build_ts_period_panel(outer)
        self._build_ts_object_panel(outer)

    def _build_ts_period_panel(self, parent):
        pnl = tk.LabelFrame(
            parent,
            text=" 📅 Период и подразделение ",
            font=("Segoe UI", 9, "bold"),
            bg=TS_COLORS["panel"],
            fg=TS_COLORS["accent"],
            relief="groove",
            bd=1,
            padx=10,
            pady=8,
        )
        pnl.grid(row=0, column=0, sticky="nsew", padx=(0, 4), pady=2)
        pnl.grid_columnconfigure(1, weight=1)

        row = 0
        self._ts_lbl(pnl, "Подразделение", row, required=True)
        self.cmb_department = ttk.Combobox(pnl, state="readonly", values=self.departments, width=36)
        self.cmb_department.grid(row=row, column=1, columnspan=3, sticky="ew", pady=3)
        self.cmb_department.bind("<<ComboboxSelected>>", lambda _e: self._on_department_select())
        row += 1

        self._ts_lbl(pnl, "Месяц", row, required=True)
        self.cmb_month = ttk.Combobox(
            pnl,
            state="readonly",
            width=13,
            values=[month_name_ru(i) for i in range(1, 13)],
        )
        self.cmb_month.grid(row=row, column=1, sticky="w", pady=3)
        self.cmb_month.current(datetime.now().month - 1)
        self.cmb_month.bind("<<ComboboxSelected>>", lambda _e: self._on_period_change())

        tk.Label(
            pnl,
            text="Год  *:",
            font=("Segoe UI", 9),
            bg=TS_COLORS["panel"],
            fg=TS_COLORS["warning"],
            anchor="e",
        ).grid(row=row, column=2, sticky="e", padx=(12, 6), pady=3)

        self.spn_year = tk.Spinbox(pnl, from_=2000, to=2100, width=6, command=self._on_period_change)
        self.spn_year.grid(row=row, column=3, sticky="w", pady=3)
        self.spn_year.bind("<FocusOut>", lambda _e: self._on_period_change())

    def _build_ts_object_panel(self, parent):
        pnl = tk.LabelFrame(
            parent,
            text=" 📍 Объект ",
            font=("Segoe UI", 9, "bold"),
            bg=TS_COLORS["panel"],
            fg=TS_COLORS["accent"],
            relief="groove",
            bd=1,
            padx=10,
            pady=8,
        )
        pnl.grid(row=0, column=1, sticky="nsew", padx=(4, 0), pady=2)
        pnl.grid_columnconfigure(1, weight=1)

        row = 0
        self._ts_lbl(pnl, "Адрес объекта", row, required=True)
        self.cmb_address = AutoCompleteCombobox(pnl, width=42, font=("Segoe UI", 9))
        self.cmb_address.set_completion_list(self.address_options)
        self.cmb_address.grid(row=row, column=1, columnspan=3, sticky="ew", pady=3)
        self.cmb_address.bind("<<ComboboxSelected>>", lambda _e: self._on_address_select())
        self.cmb_address.bind("<Return>", lambda _e: self._on_address_select())
        row += 1

        self._ts_lbl(pnl, "ID объекта", row)
        id_frame = tk.Frame(pnl, bg=TS_COLORS["panel"])
        id_frame.grid(row=row, column=1, columnspan=3, sticky="ew", pady=3)

        self.cmb_object_id = ttk.Combobox(id_frame, state="readonly", values=[], width=22)
        self.cmb_object_id.pack(side="left")
        self.cmb_object_id.bind("<<ComboboxSelected>>", lambda _e: self._on_object_id_select())

        tk.Label(
            id_frame,
            text="← подставляется автоматически по адресу",
            font=("Segoe UI", 7),
            fg="#888",
            bg=TS_COLORS["panel"],
        ).pack(side="left", padx=8)

    def _ts_btn(self, parent, text: str, cmd, side="left", padx=3, pady=0, width=None):
        b = ttk.Button(parent, text=text, command=cmd, width=width)
        b.pack(side=side, padx=padx, pady=pady)
        return b

    def _build_ts_toolbar(self):
        border_color = TS_COLORS.get("border", "#c9d3df")
    
        bar = tk.Frame(self, bg=TS_COLORS["accent_light"], relief="flat")
        bar.pack(fill="x", padx=10, pady=(4, 0))
    
        left = tk.Frame(bar, bg=TS_COLORS["accent_light"])
        left.pack(side="left", fill="x", expand=True)
    
        actions = tk.Frame(bar, bg=TS_COLORS["accent_light"])
        actions.pack(side="right", anchor="n", padx=(10, 4), pady=4)
    
        # --- Верхний ряд ---
        row1 = tk.Frame(left, bg=TS_COLORS["accent_light"])
        row1.pack(fill="x", pady=(4, 2))
    
        self._ts_btn(row1, "Добавить подразделение", self.add_department_all, side="left", padx=(4, 3))
        self._ts_btn(row1, "Выбрать сотрудников", self.add_department_partial, side="left", padx=3)
    
        tk.Frame(row1, bg=border_color, width=1, height=24).pack(side="left", padx=8, fill="y")
    
        self._ts_btn(row1, "Время выбранным", self.fill_time_selected, side="left", padx=3)
        self._ts_btn(row1, "Часы всем", self.fill_hours_all, side="left", padx=3)
        self._ts_btn(row1, "Очистить часы", self.clear_all_rows, side="left", padx=3)
        self._ts_btn(row1, "Снять выделение", self.clear_selection, side="left", padx=3)
    
        # --- Нижний ряд ---
        row2 = tk.Frame(left, bg=TS_COLORS["accent_light"])
        row2.pack(fill="x", pady=(0, 4))
    
        self._ts_btn(row2, "Импорт Excel", self.import_from_excel, side="left", padx=(4, 3))
        self._ts_btn(row2, "Копировать из месяца", self.copy_from_month, side="left", padx=3)
        self._ts_btn(row2, "Загрузить СКУД", self.import_from_skud, side="left", padx=3)
    
        # --- Правый блок действий ---
        btn_save = tk.Button(
            actions,
            text="Сохранить",
            font=("Segoe UI", 9, "bold"),
            bg=TS_COLORS["btn_save_bg"],
            fg=TS_COLORS["btn_save_fg"],
            activebackground="#0d47a1",
            activeforeground="white",
            relief="flat",
            cursor="hand2",
            padx=14,
            pady=4,
            command=self.save_all,
            width=14,
        )
        btn_save.pack(fill="x", pady=(0, 4))
        btn_save.bind("<Enter>", lambda _e: btn_save.config(bg="#0d47a1"))
        btn_save.bind("<Leave>", lambda _e: btn_save.config(bg=TS_COLORS["btn_save_bg"]))
    
        self._btn_export_ref = ttk.Button(
            actions,
            text="Выгрузить Excel",
            command=self.export_current_timesheet_to_excel,
            width=16,
        )
        self._btn_export_ref.pack(fill="x")
    
        self._toolbar_frame = bar
        self._btn_save_ref = btn_save
    
        if self.read_only:
            for container in (row1, row2):
                for child in container.winfo_children():
                    try:
                        child.configure(state="disabled")
                    except Exception:
                        pass
    
            try:
                btn_save.configure(state="disabled")
            except Exception:
                pass

    def _build_ts_filter_bar(self):
        border_color = TS_COLORS.get("border", "#c9d3df")
    
        bar = tk.Frame(self, bg=TS_COLORS["bg"], pady=2)
        bar.pack(fill="x", padx=10, pady=(4, 0))
    
        tk.Label(
            bar,
            text="🔍 Поиск (ФИО / таб. №):",
            font=("Segoe UI", 9),
            bg=TS_COLORS["bg"],
        ).pack(side="left")
    
        ent_filter = ttk.Entry(bar, textvariable=self.var_filter, width=36)
        ent_filter.pack(side="left", padx=(4, 8))
    
        def on_filter_click(_event):
            if hasattr(self, "grid") and self.grid is not None:
                try:
                    self.grid.close_editor(commit=True)
                except Exception:
                    pass
            self.after_idle(lambda: ent_filter.focus_set())
    
        ent_filter.bind("<Button-1>", on_filter_click)
    
        ttk.Button(bar, text="Очистить", command=self._clear_filter).pack(side="left")
    
        tk.Frame(bar, bg=border_color, width=1).pack(side="left", fill="y", padx=12)
    
        tk.Label(
            bar,
            text="Бригадир:",
            font=("Segoe UI", 9),
            bg=TS_COLORS["bg"],
        ).pack(side="left")
    
        self.cmb_brigadier = ttk.Combobox(
            bar,
            state="readonly",
            width=30,
            values=["Все"],
            textvariable=self.var_brigadier,
        )
        self.cmb_brigadier.pack(side="left", padx=(4, 0))
        self.cmb_brigadier.bind("<<ComboboxSelected>>", lambda _e: self._apply_filter())
    
        def _on_filter_key(_e=None):
            try:
                if self._filter_job is not None:
                    self.after_cancel(self._filter_job)
            except Exception:
                pass
            self._filter_job = self.after(120, self._apply_filter)
    
        ent_filter.bind("<KeyRelease>", _on_filter_key)

    def _build_ts_grid(self):
        main_frame = tk.Frame(
            self,
            bg=TS_COLORS["panel"],
            highlightbackground=TS_COLORS.get("border", "#c9d3df"),
            highlightthickness=1,
            bd=0,
        )
        main_frame.pack(fill="both", expand=True, padx=10, pady=(4, 4))
        main_frame.grid_rowconfigure(0, weight=1)
        main_frame.grid_columnconfigure(0, weight=1)
    
        self.grid = VirtualTimesheetGrid(
            main_frame,
            get_year_month=self.get_year_month,
            on_change=self._on_cell_changed,
            on_delete_row=self._grid_delete_row,
            on_selection_change=self._on_grid_selection_changed,
            row_height=22,
            colpx=self.COLPX,
            read_only=self.read_only,
        )
        self.grid.grid(row=0, column=0, sticky="nsew")
    
        # Для notebook / вкладок нужно несколько отложенных refresh
        self.after(60, self.grid.refresh)
        self.after(180, self.grid.refresh)


    def _build_ts_bottom(self):
        bottom = tk.Frame(self, bg=TS_COLORS["accent_light"], pady=5)
        bottom.pack(fill="x", padx=10, pady=(0, 8))

        self.lbl_object_total = tk.Label(
            bottom,
            text="Сотрудников: 0 | Дней: 0 | Часов: 0",
            font=("Segoe UI", 9, "bold"),
            fg=TS_COLORS["accent"],
            bg=TS_COLORS["accent_light"],
        )
        self.lbl_object_total.pack(side="left", padx=10)

        self.lbl_selected_count = tk.Label(
            bottom,
            text="Выбрано строк: 0",
            font=("Segoe UI", 9, "bold"),
            fg=TS_COLORS["warning"],
            bg=TS_COLORS["accent_light"],
        )
        self.lbl_selected_count.pack(side="left", padx=(20, 10))

        if self.read_only:
            tk.Label(
                bottom,
                text="👁 Режим просмотра — изменения недоступны",
                font=("Segoe UI", 9, "italic"),
                fg=TS_COLORS["warning"],
                bg=TS_COLORS["accent_light"],
            ).pack(side="right", padx=10)

    # --------------------------------------------------------
    # Инициализация
    # --------------------------------------------------------

    def _init_ts_values(self):
        self._suppress_events = True
        try:
            deps = self.departments or ["Все"]

            saved_dep = None
            try:
                if get_selected_department_from_config:
                    saved_dep = get_selected_department_from_config()
            except Exception:
                logger.exception("Ошибка чтения подразделения из конфигурации")

            init_dep = self._init_department or saved_dep or (deps[0] if deps else "Все")
            self._ensure_department_option(init_dep)
            self.cmb_department.set(init_dep if init_dep in (self.cmb_department.cget("values") or deps) else deps[0])

            self.spn_year.delete(0, "end")
            self.spn_year.insert(0, str(self._init_year or datetime.now().year))

            if self._init_month and 1 <= self._init_month <= 12:
                self.cmb_month.current(self._init_month - 1)
            else:
                self.cmb_month.current(datetime.now().month - 1)

            if self._active_header_id:
                header = load_timesheet_full_by_header_id(self._active_header_id)
                if header:
                    if self.owner_user_id is None and header.get("user_id"):
                        self.owner_user_id = int(header["user_id"])

                    hist_dep = normalize_spaces(header.get("department") or "")
                    hist_addr = normalize_spaces(header.get("object_addr") or "")
                    hist_oid = normalize_spaces(header.get("object_id") or "")
                    hist_year = int(header.get("year") or self._safe_get_year())
                    hist_month = int(header.get("month") or (self.cmb_month.current() + 1))

                    self._ensure_department_option(hist_dep)
                    self.cmb_department.set(hist_dep or "Все")

                    self.spn_year.delete(0, "end")
                    self.spn_year.insert(0, str(hist_year))

                    if 1 <= hist_month <= 12:
                        self.cmb_month.current(hist_month - 1)

                    self._ensure_address_option(hist_addr)
                    self.cmb_address.set(hist_addr)

                    self._sync_object_id_values_silent()
                    if hist_oid:
                        values = list(self.cmb_object_id.cget("values") or [])
                        if hist_oid not in values:
                            values.append(hist_oid)
                            self.cmb_object_id.config(values=values)
                        self.cmb_object_id.set(hist_oid)

            else:
                if self._init_object_addr:
                    self._ensure_address_option(self._init_object_addr)
                    self.cmb_address.set(self._init_object_addr)

                if self._init_object_id:
                    self._sync_object_id_values_silent()
                    values = list(self.cmb_object_id.cget("values") or [])
                    if self._init_object_id not in values:
                        values.append(self._init_object_id)
                        self.cmb_object_id.config(values=values)
                    self.cmb_object_id.set(self._init_object_id)

            self._refresh_employee_selector_for_department(normalize_spaces(self.cmb_department.get() or "Все"))
        finally:
            self._suppress_events = False
            self._initializing = False

        self._load_existing_rows()
        self._set_status_text("Авто‑сохранение: нет", fg="#bbdefb")

    # --------------------------------------------------------
    # Контекст / переключение
    # --------------------------------------------------------

    def _sync_object_id_values_silent(self):
        addr = normalize_spaces(self.cmb_address.get() or "")
        objects_for_addr = [
            (code, a, short_name)
            for (code, a, short_name) in self.objects_full
            if normalize_spaces(a) == addr
        ]
        if not objects_for_addr:
            self.cmb_object_id.config(state="normal", values=[])
            self.cmb_object_id.set("")
            return

        ids = sorted({normalize_spaces(code) for code, _, _ in objects_for_addr if normalize_spaces(code)})
        cur = normalize_spaces(self.cmb_object_id.get() or "")

        self.cmb_object_id.config(state="readonly", values=ids)
        if cur and cur in ids:
            return
        if len(ids) == 1:
            self.cmb_object_id.set(ids[0])
        else:
            if cur not in ids:
                self.cmb_object_id.set("")

    def _on_period_change(self):
        if self._initializing or self._suppress_events:
            return

        try:
            self.grid.close_editor(commit=True)
        except Exception:
            pass

        if not self._confirm_leave_with_unsaved():
            self._restore_controls_to_loaded_context()
            return

        self._active_header_id = None
        self._load_existing_rows()

    def _on_department_select(self):
        if self._initializing or self._suppress_events:
            return

        if not self._confirm_leave_with_unsaved():
            self._restore_controls_to_loaded_context()
            return

        dep_sel = normalize_spaces(self.cmb_department.get() or "Все")
        try:
            if set_selected_department_in_config:
                set_selected_department_in_config(dep_sel)
        except Exception:
            logger.exception("Не удалось сохранить выбранное подразделение в конфигурацию")

        self._refresh_employee_selector_for_department(dep_sel)
        self._active_header_id = None
        self._load_existing_rows()

    def _on_address_change(self, ask_user: bool = True):
        addr = normalize_spaces(self.cmb_address.get() or "")
        objects_for_addr = [
            (normalize_spaces(code), normalize_spaces(a), normalize_spaces(short_name))
            for (code, a, short_name) in self.objects_full
            if normalize_spaces(a) == addr
        ]

        if not objects_for_addr:
            self.cmb_object_id.config(state="normal", values=[])
            self.cmb_object_id.set("")
            return

        ids = sorted({code for code, _, _ in objects_for_addr if code})
        current_oid = normalize_spaces(self.cmb_object_id.get() or "")
        self.cmb_object_id.config(state="readonly", values=ids)

        if current_oid and current_oid in ids:
            return

        if len(ids) == 1:
            self.cmb_object_id.set(ids[0])
            return

        if not ask_user or getattr(self, "_suppress_object_id_dialog", False):
            self.cmb_object_id.set("")
            return

        dlg = SelectObjectIdDialog(self, objects_for_addr, addr)
        self.wait_window(dlg)

        selected_id = normalize_spaces(dlg.result or "")
        if selected_id and selected_id in ids:
            self.cmb_object_id.set(selected_id)
        else:
            self.cmb_object_id.set("")

    def _on_address_select(self):
        if self._initializing or self._suppress_events:
            return

        if not self._confirm_leave_with_unsaved():
            self._restore_controls_to_loaded_context()
            return

        self._active_header_id = None
        self._on_address_change(ask_user=True)
        self._load_existing_rows()

    def _on_object_id_select(self):
        if self._initializing or self._suppress_events:
            return

        if not self._confirm_leave_with_unsaved():
            self._restore_controls_to_loaded_context()
            return

        self._active_header_id = None
        self._load_existing_rows()

    def _clear_filter(self):
        try:
            self.var_filter.set("")
            self.var_brigadier.set("Все")
        except Exception:
            pass
        self._apply_filter()

    def _parse_selected_brigadier_tbn(self) -> str | None:
        s = normalize_spaces(self.var_brigadier.get() or "Все")
        if s == "Все":
            return None
        if s == "Без бригадира":
            return ""
        if s.endswith(")") and "(" in s:
            tbn = s[s.rfind("(") + 1 : -1].strip()
            return normalize_tbn(tbn) or None
        return None

    # --------------------------------------------------------
    # Загрузка табеля
    # --------------------------------------------------------

    def _reload_brigadier_filter_data(self):
        dep = normalize_spaces(self.cmb_department.get() or "")
        if not dep or dep == "Все":
            self._brig_assign = {}
            self._brig_names = {}
            try:
                self.cmb_brigadier.configure(values=["Все"])
                self.var_brigadier.set("Все")
            except Exception:
                pass
            return

        try:
            self._brig_assign = load_brigadier_assignments_for_department(dep)
            self._brig_names = load_brigadier_names_for_department(dep)
        except Exception:
            logger.exception("Ошибка загрузки фильтра бригадиров")
            self._brig_assign = {}
            self._brig_names = {}
            try:
                self.cmb_brigadier.configure(values=["Все"])
                self.var_brigadier.set("Все")
            except Exception:
                pass
            return

        options = ["Все", "Без бригадира"]
        pairs = []
        for tbn, fio in self._brig_names.items():
            fio = normalize_spaces(fio)
            if fio:
                pairs.append((fio.lower(), f"{fio} ({tbn})"))

        for _k, label in sorted(pairs):
            options.append(label)

        cur = normalize_spaces(self.var_brigadier.get() or "Все")
        self.cmb_brigadier.configure(values=options)
        if cur not in options:
            self.var_brigadier.set("Все")

    def _show_suspicious_warning_on_load(self, suspicious: List[Dict[str, Any]]):
        lines = []
        for item in suspicious[:10]:
            fio = item.get("fio", "")
            day = item.get("day", "?")
            raw = item.get("raw", "")
            parsed = item.get("parsed")
            parsed_str = f"{parsed:.2f}" if isinstance(parsed, (int, float)) else "?"
            lines.append(f"  • {fio}, день {day}: '{raw}' → {parsed_str} ч.")

        msg = (
            f"⚠️ Обнаружено {len(suspicious)} подозрительных значений\n"
            f"(более {MAX_HOURS_PER_DAY} часов в сутки).\n\n"
            f"Возможно, пропущена точка/запятая (825 → 8.25).\n\n"
        )
        if lines:
            msg += "Примеры:\n" + "\n".join(lines)
            if len(suspicious) > 10:
                msg += f"\n  ... и ещё {len(suspicious) - 10}"

        messagebox.showwarning("⚠️ Подозрительные значения", msg, parent=self)

    def _load_existing_rows(self):
        self.model_rows_all.clear()
        self.model_rows = self.model_rows_all
        self._selected_row_keys.clear()
        self._update_selected_count()
        try:
            self.grid.set_selected_indices(set())
        except Exception:
            pass

        addr = normalize_spaces(self.cmb_address.get() or "")
        oid = normalize_spaces(self.cmb_object_id.get() or "")
        year, month = self.get_year_month()
        dep = normalize_spaces(self.cmb_department.get() or "")

        self._reload_brigadier_filter_data()

        if dep == "Все":
            self.grid.set_rows(self.model_rows)
            self._recalc_object_total()
            self._loaded_context = self._capture_current_context()
            self._dirty = False
            return

        try:
            loaded_rows: List[Dict[str, Any]] = []

            if self._active_header_id:
                full = load_timesheet_full_by_header_id(self._active_header_id)
                if full:
                    hist_dep = normalize_spaces(full.get("department") or dep)
                    hist_addr = normalize_spaces(full.get("object_addr") or addr)
                    hist_oid = normalize_spaces(full.get("object_id") or oid)
                    hist_year = int(full.get("year") or year)
                    hist_month = int(full.get("month") or month)

                    self._suppress_events = True
                    try:
                        self._ensure_department_option(hist_dep)
                        self.cmb_department.set(hist_dep or "Все")
                        self.spn_year.delete(0, "end")
                        self.spn_year.insert(0, str(hist_year))
                        if 1 <= hist_month <= 12:
                            self.cmb_month.current(hist_month - 1)

                        self._ensure_address_option(hist_addr)
                        self.cmb_address.set(hist_addr)
                        self._sync_object_id_values_silent()
                        if hist_oid:
                            values = list(self.cmb_object_id.cget("values") or [])
                            if hist_oid not in values:
                                values.append(hist_oid)
                                self.cmb_object_id.config(values=values)
                            self.cmb_object_id.set(hist_oid)
                    finally:
                        self._suppress_events = False

                    if self.owner_user_id is None and full.get("user_id"):
                        self.owner_user_id = int(full["user_id"])

                    loaded_rows = [
                        normalize_row_record(row, hist_year, hist_month)
                        for row in (full.get("rows") or [])
                    ]
                    year, month = hist_year, hist_month
                else:
                    self._active_header_id = None

            if not self._active_header_id:
                user_id = self._safe_current_user_id()
                if user_id:
                    loaded_rows = [
                        normalize_row_record(row, year, month)
                        for row in load_timesheet_rows_from_db(
                            object_id=oid or None,
                            object_addr=addr,
                            department=dep,
                            year=year,
                            month=month,
                            user_id=user_id,
                        )
                    ]

                    resolved_header_id = find_timesheet_header_id(
                        object_id=oid or None,
                        object_addr=addr,
                        department=dep,
                        year=year,
                        month=month,
                        user_id=user_id,
                    )
                    self._active_header_id = resolved_header_id

            self.model_rows_all.extend(loaded_rows)
            self._recalc_all_row_totals()
            self._apply_filter()

            suspicious = find_suspicious_cells(self.model_rows_all, year, month)
            if suspicious:
                self.after(250, lambda s=suspicious: self._show_suspicious_warning_on_load(s))

            self._loaded_context = self._capture_current_context()
            self._loaded_context["header_id"] = self._active_header_id
            self._dirty = False
        except Exception as e:
            logger.exception("Ошибка загрузки табеля из БД")
            messagebox.showerror("Загрузка", f"Не удалось загрузить табель из БД:\n{e}", parent=self)
            self.grid.set_rows(self.model_rows)
            self._recalc_object_total()

    # --------------------------------------------------------
    # Фильтр / итоги
    # --------------------------------------------------------

    def _apply_filter(self):
        self._filter_job = None
    
        # Сначала запоминаем текущее выделение видимых строк
        self._remember_grid_selection()
    
        q = normalize_spaces(self.var_filter.get() or "").lower()
        brig_tbn_sel = self._parse_selected_brigadier_tbn()
    
        filtered: List[Dict[str, Any]] = []
        for rec in self.model_rows_all:
            fio = normalize_spaces(rec.get("fio") or "")
            tbn = normalize_tbn(rec.get("tbn"))
    
            if brig_tbn_sel is not None:
                assigned_brig_tbn = normalize_tbn(self._brig_assign.get(tbn) or "")
                if brig_tbn_sel == "":
                    if assigned_brig_tbn:
                        continue
                else:
                    if assigned_brig_tbn != brig_tbn_sel:
                        continue
    
            if q:
                if q not in fio.lower() and q not in tbn.lower():
                    continue
    
            filtered.append(rec)
    
        self.model_rows = filtered
    
        self.grid.set_rows(self.model_rows)
    
        # После перестройки списка восстанавливаем выделение
        self._restore_grid_selection()
        self._recalc_object_total()

    def _recalc_row_totals_for_rec(self, rec: Dict[str, Any]):
        year, month = self.get_year_month()
        rec["hours"] = normalize_hours_list(rec.get("hours"), year, month)
        rec["_totals"] = calc_row_totals(rec["hours"], year, month)

    def _recalc_all_row_totals(self):
        year, month = self.get_year_month()
        for rec in self.model_rows_all:
            rec["hours"] = normalize_hours_list(rec.get("hours"), year, month)
            rec["_totals"] = calc_row_totals(rec["hours"], year, month)

    def _recalc_object_total(self):
        year, month = self.get_year_month()
        summary = calc_rows_summary(self.model_rows_all, year, month)
        txt = (
            f"Сотрудников: {summary['employees']}  |  "
            f"Дней: {summary['days']}  |  "
            f"Часов: {format_summary_value(summary['hours'])}  |  "
            f"Ночных: {format_summary_value(summary['night_hours'])}  |  "
            f"Пер. день: {format_summary_value(summary['ot_day'])}  |  "
            f"Пер. ночь: {format_summary_value(summary['ot_night'])}"
        )
        try:
            self.lbl_object_total.config(text=txt)
        except Exception:
            pass

    # --------------------------------------------------------
    # Изменения грида
    # --------------------------------------------------------

    def _on_cell_changed(self, row_index: int, day_index: int):
        if 0 <= row_index < len(self.model_rows):
            rec = self.model_rows[row_index]
            self._recalc_row_totals_for_rec(rec)

            hours_list = rec.get("hours") or []
            if 0 <= day_index < len(hours_list):
                raw_value = hours_list[day_index]
                parsed = parse_timesheet_cell(raw_value)
                if parsed.suspicious:
                    parsed_str = f"{parsed.total_hours:.2f}" if parsed.total_hours is not None else "?"
                    fio = rec.get("fio", "")
                    messagebox.showwarning(
                        "⚠️ Подозрительное значение",
                        f"Сотрудник: {fio}\n"
                        f"День: {day_index + 1}\n"
                        f"Значение: '{raw_value}' → {parsed_str} часов\n\n"
                        f"В сутках максимум {MAX_HOURS_PER_DAY} часов.\n"
                        f"Возможно, пропущена точка/запятая?\n"
                        f"(например: 825 → 8.25)",
                        parent=self,
                    )

        self._grid_refresh(rows_changed=False)
        self._recalc_object_total()
        self._mark_dirty()
        self._schedule_auto_save()

    def _grid_delete_row(self, row_index: int):
        if self.read_only:
            return
        if not (0 <= row_index < len(self.model_rows)):
            return

        rec = self.model_rows[row_index]
        try:
            self.model_rows_all.remove(rec)
        except ValueError:
            return

        self._recalc_all_row_totals()
        self._apply_filter()
        self._mark_dirty()
        self._schedule_auto_save()

    # --------------------------------------------------------
    # Операции с сотрудниками / часами
    # --------------------------------------------------------

    def add_row(self):
        if self.read_only:
            return
        messagebox.showinfo(
            "Объектный табель",
            "Ручное добавление сотрудника отключено.\nИспользуйте добавление из подразделения.",
            parent=self,
        )

    def add_department_all(self):
        if self.read_only:
            return

        dep_sel = normalize_spaces(self.cmb_department.get() or "Все")
        if dep_sel == "Все":
            candidates = list(self.employees)
            if not candidates:
                messagebox.showinfo("Объектный табель", "Справочник сотрудников пуст.", parent=self)
                return
            if not messagebox.askyesno(
                "Добавить всех",
                f"Добавить в табель всех сотрудников ({len(candidates)})?",
                parent=self,
            ):
                return
        else:
            candidates = [e for e in self.employees if normalize_spaces(e[3]) == dep_sel]
            if not candidates:
                messagebox.showinfo("Объектный табель", f"В подразделении «{dep_sel}» нет сотрудников.", parent=self)
                return

        existing = {make_row_key(r.get("fio", ""), r.get("tbn", "")) for r in self.model_rows_all}
        added_count = 0

        dlg = BatchAddDialog(self, total=len(candidates), title="Добавление сотрудников")

        def process_batch():
            nonlocal added_count
            try:
                for fio, tbn, _pos, _dep in candidates:
                    if dlg.cancelled:
                        break
                    key = make_row_key(fio, tbn)
                    if key not in existing:
                        self.model_rows_all.append({"fio": fio, "tbn": tbn, "hours": [None] * 31})
                        existing.add(key)
                        added_count += 1
                    dlg.step()
            finally:
                dlg.close()

            self._recalc_all_row_totals()
            self._apply_filter()
            if added_count > 0:
                self._mark_dirty()
                self._schedule_auto_save()
                messagebox.showinfo("Объектный табель", f"Добавлено новых сотрудников: {added_count}", parent=self)
            else:
                messagebox.showinfo("Объектный табель", "Все сотрудники уже есть в списке.", parent=self)

        self.after(50, process_batch)

    def add_department_partial(self):
        if self.read_only:
            return

        dep_sel = normalize_spaces(self.cmb_department.get() or "Все")
        if not self.employees:
            messagebox.showinfo("Объектный табель", "Справочник сотрудников пуст.", parent=self)
            return

        dlg = SelectEmployeesDialog(self, self.employees, dep_sel)
        self.wait_window(dlg)

        if dlg.result is None:
            return

        selected_emps = dlg.result
        if not selected_emps:
            return

        existing = {make_row_key(r.get("fio", ""), r.get("tbn", "")) for r in self.model_rows_all}
        added_count = 0

        for fio, tbn, _pos, _dep in selected_emps:
            key = make_row_key(fio, tbn)
            if key in existing:
                continue
            self.model_rows_all.append({"fio": fio, "tbn": tbn, "hours": [None] * 31})
            existing.add(key)
            added_count += 1

        self._recalc_all_row_totals()
        self._apply_filter()

        if added_count > 0:
            self._mark_dirty()
            self._schedule_auto_save()
            messagebox.showinfo("Объектный табель", f"Добавлено сотрудников: {added_count}", parent=self)
        else:
            messagebox.showinfo("Объектный табель", "Все выбранные сотрудники уже есть в табеле.", parent=self)

    def clear_selection(self):
        self._selected_row_keys.clear()
        if hasattr(self, "grid"):
            try:
                self.grid.set_selected_indices(set())
            except Exception:
                pass
        self._update_selected_count()

    def fill_time_selected(self):
        if self.read_only:
            return
        if not self.model_rows_all:
            messagebox.showinfo("Время для выделенных", "Список сотрудников пуст.", parent=self)
            return
    
        # На всякий случай подтянуть текущее видимое выделение в общее хранилище
        self._remember_grid_selection()
    
        selected_keys = set(self._selected_row_keys)
        if not selected_keys:
            messagebox.showinfo("Время для выделенных", "Не выбрано ни одного сотрудника.", parent=self)
            return
    
        year, month = self.get_year_month()
        dlg = TimeForSelectedDialog(self, month_days(year, month))
        if dlg.result is None:
            return
    
        day_from = dlg.result["from"]
        day_to = dlg.result["to"]
        value = dlg.result["value"]
    
        changed_count = 0
        for rec in self.model_rows_all:
            if self._row_key(rec) not in selected_keys:
                continue
    
            hours_list = normalize_hours_list(rec.get("hours"), year, month)
            for d in range(day_from, day_to + 1):
                hours_list[d - 1] = value
            rec["hours"] = hours_list
            self._recalc_row_totals_for_rec(rec)
            changed_count += 1
    
        self._apply_filter()
        self._mark_dirty()
        self._schedule_auto_save()
    
        msg_val = "очищены" if value is None else f"установлены в '{value}'"
        msg_days = f"для дня {day_from}" if day_from == day_to else f"для дней {day_from}–{day_to}"
        messagebox.showinfo(
            "Время для выделенных",
            f"Значения {msg_val} {msg_days} у {changed_count} выделенных сотрудников.",
            parent=self,
        )

    def fill_hours_all(self):
        if self.read_only:
            return
        if not self.model_rows_all:
            messagebox.showinfo("Проставить часы", "Список сотрудников пуст.", parent=self)
            return

        year, month = self.get_year_month()
        max_day = month_days(year, month)

        dlg = HoursFillDialog(self, max_day)
        if not dlg.result:
            return

        day = dlg.result["day"]
        if not (1 <= day <= max_day):
            messagebox.showwarning("Проставить часы", f"В этом месяце нет дня №{day}.", parent=self)
            return

        day_idx = day - 1
        is_clear = bool(dlg.result.get("clear", False))
        hours_val_str = None if is_clear else format_hours_for_cell(float(dlg.result["hours"]))

        for rec in self.model_rows_all:
            hours = normalize_hours_list(rec.get("hours"), year, month)
            hours[day_idx] = hours_val_str
            rec["hours"] = hours
            rec["_totals"] = calc_row_totals(hours, year, month)

        self._apply_filter()
        self._mark_dirty()
        self._schedule_auto_save()

        if is_clear:
            messagebox.showinfo("Проставить часы", f"День {day} очищен у всех сотрудников.", parent=self)
        else:
            messagebox.showinfo(
                "Проставить часы",
                f"Часы '{hours_val_str}' проставлены в день {day} всем сотрудникам.",
                parent=self,
            )

    def clear_all_rows(self):
        if self.read_only or not self.model_rows_all:
            return
        if not messagebox.askyesno(
            "Очистка табеля",
            "Вы уверены, что хотите очистить все часы у всех сотрудников?\n\nСами сотрудники останутся в списке.",
            parent=self,
        ):
            return

        year, month = self.get_year_month()
        for rec in self.model_rows_all:
            rec["hours"] = normalize_hours_list([None] * 31, year, month)
            rec["_totals"] = calc_row_totals(rec["hours"], year, month)

        self._apply_filter()
        self._mark_dirty()
        self._schedule_auto_save()
        messagebox.showinfo("Очистка", "Все часы были стерты.", parent=self)

    # --------------------------------------------------------
    # Импорт СКУД
    # --------------------------------------------------------

    def import_from_skud(self):
        if self.read_only:
            return

        current_dep = normalize_spaces(self.cmb_department.get() or "")
        if current_dep == "Все":
            messagebox.showwarning("СКУД", "Выберите конкретное подразделение (не 'Все').", parent=self)
            return

        year, month = self.get_year_month()
        dlg_date = SelectDateDialog(self, init_date=date(year, month, 1))
        selected_date = dlg_date.result
        if selected_date is None:
            return

        if not ensure_current_month_date(selected_date, year, month):
            messagebox.showwarning(
                "СКУД",
                f"Выбрана дата {selected_date.strftime('%d.%m.%Y')}, "
                f"но открыт табель за {month_name_ru(month)} {year}.\n"
                "Выберите дату из текущего месяца.",
                parent=self,
            )
            return

        path = filedialog.askopenfilename(
            parent=self,
            title="Выберите Excel-отчёт СКУД",
            filetypes=[("Excel", "*.xlsx"), ("Все", "*.*")],
        )
        if not path:
            return

        try:
            events = read_skud_events_from_xlsx(path)
            summary_by_skud_fio, problems = compute_day_summary_from_events(events, target_date=selected_date)

            if not summary_by_skud_fio and not problems:
                messagebox.showinfo("СКУД", "В отчёте нет событий с ФИО на выбранную дату.", parent=self)
                return

            candidates = sorted(self.allowed_fio_names) if self.allowed_fio_names else self.emp_names

            mapping_rows: List[Dict[str, Any]] = []
            for skud_fio, info in sorted(summary_by_skud_fio.items(), key=lambda x: x[0].lower()):
                best, score = best_fio_match_with_score(skud_fio, candidates)
                hours_val = info.get("hours_rounded")
                apply_default = bool(best) and (score >= 0.90) and isinstance(hours_val, int) and hours_val > 0

                mapping_rows.append(
                    {
                        "skud_fio": skud_fio,
                        "matched_fio": best or "",
                        "score": score,
                        "hours_rounded": hours_val,
                        "minutes": info.get("minutes"),
                        "first_in": info.get("first_in"),
                        "last_out": info.get("last_out"),
                        "count_in": info.get("count_in"),
                        "count_out": info.get("count_out"),
                        "apply": apply_default,
                    }
                )

            dlg = SkudMappingReviewDialog(self, rows=mapping_rows, problems=problems)
            self.wait_window(dlg)
            if not dlg.result or not dlg.result.get("apply"):
                return

            chosen = dlg.result.get("rows") or []
            if not chosen:
                messagebox.showinfo("СКУД", "Ничего не выбрано для применения.", parent=self)
                return

            day_idx = selected_date.day - 1
            applied = 0
            added = 0
            skipped = 0
            ambiguous = 0

            for item in chosen:
                matched_fio = normalize_spaces(item.get("matched_fio") or "")
                if not matched_fio:
                    skipped += 1
                    continue

                hours_val = item.get("hours_rounded")
                if not isinstance(hours_val, int) or hours_val <= 0:
                    skipped += 1
                    continue

                rec = self._find_unique_row_by_fio(matched_fio)
                if rec is None:
                    emp = self._find_unique_employee_by_fio(matched_fio, current_dep)
                    if emp is None:
                        ambiguous += 1
                        continue
                    fio, tbn, _pos, _dep = emp
                    rec = {"fio": fio, "tbn": tbn or "", "hours": [None] * 31}
                    self.model_rows_all.append(rec)
                    added += 1

                hours = normalize_hours_list(rec.get("hours"), year, month)
                hours[day_idx] = format_hours_for_cell(hours_val)
                rec["hours"] = hours
                rec["_totals"] = calc_row_totals(hours, year, month)
                applied += 1

            self._apply_filter()
            if applied > 0 or added > 0:
                self._mark_dirty()
                self._schedule_auto_save()

            messagebox.showinfo(
                "СКУД",
                "Готово.\n"
                f"Применено строк: {applied}\n"
                f"Добавлено сотрудников: {added}\n"
                f"Проблем (нет входа/выхода / аномалии): {len(problems)}\n"
                f"Пропущено строк: {skipped}\n"
                f"Неоднозначных ФИО: {ambiguous}",
                parent=self,
            )

        except Exception as e:
            logger.exception("Ошибка импорта СКУД")
            messagebox.showerror("СКУД", f"Ошибка при загрузке СКУД:\n{e}", parent=self)

    # --------------------------------------------------------
    # Импорт Excel / копирование
    # --------------------------------------------------------

    def _get_import_sheet(self, wb) -> Any:
        if "Табель" not in wb.sheetnames:
            raise RuntimeError("В файле не найден лист 'Табель'.")

        ws = wb["Табель"]
        if normalize_spaces(str(ws.cell(1, 1).value or "")) != "ID объекта":
            raise RuntimeError("Лист 'Табель' имеет неверный формат.")
        return ws

    def import_from_excel(self):
        if self.read_only:
            return

        addr = normalize_spaces(self.cmb_address.get() or "")
        oid = normalize_spaces(self.cmb_object_id.get() or "")
        year, month = self.get_year_month()
        current_dep = normalize_spaces(self.cmb_department.get() or "")

        if current_dep == "Все":
            messagebox.showwarning("Импорт", "Выберите конкретное подразделение (не 'Все').", parent=self)
            return

        if not addr and not oid:
            messagebox.showwarning("Импорт", "Укажите адрес/ID объекта и период.", parent=self)
            return

        path = filedialog.askopenfilename(
            parent=self,
            title="Выберите Excel-файл табеля",
            filetypes=[("Excel", "*.xlsx")],
        )
        if not path:
            return

        try:
            wb = load_workbook(path, data_only=True)
            ws = self._get_import_sheet(wb)

            imported: List[Dict[str, Any]] = []

            for r in range(2, ws.max_row + 1):
                row_month = int(ws.cell(r, 3).value or 0)
                row_year = int(ws.cell(r, 4).value or 0)
                if row_month != month or row_year != year:
                    continue

                row_oid = normalize_spaces(str(ws.cell(r, 1).value or ""))
                row_addr = normalize_spaces(str(ws.cell(r, 2).value or ""))
                row_dep = normalize_spaces(str(ws.cell(r, 7).value or ""))

                if oid:
                    if row_oid != oid:
                        continue
                else:
                    if row_addr != addr:
                        continue

                if row_dep != current_dep:
                    continue

                fio = normalize_spaces(str(ws.cell(r, 5).value or ""))
                tbn = normalize_tbn(ws.cell(r, 6).value)
                if not fio and not tbn:
                    continue

                hours_raw = [normalize_spaces(str(ws.cell(r, c).value or "")) or None for c in range(8, 8 + 31)]
                imported.append(normalize_row_record({"fio": fio, "tbn": tbn, "hours": hours_raw}, year, month))

            if not imported:
                messagebox.showinfo("Импорт", "Подходящих строк не найдено.", parent=self)
                return

            imported = deduplicate_timesheet_rows(imported, year, month)

            replace_mode = (
                messagebox.askyesno("Импорт", "Заменить текущий список?", parent=self)
                if self.model_rows_all
                else True
            )
            if replace_mode:
                self.model_rows_all.clear()
                self.clear_selection()

            existing = {make_row_key(r.get("fio", ""), r.get("tbn", "")) for r in self.model_rows_all}
            added = 0
            for rec in imported:
                key = make_row_key(rec.get("fio", ""), rec.get("tbn", ""))
                if key in existing:
                    continue
                self.model_rows_all.append(rec)
                existing.add(key)
                added += 1

            self._recalc_all_row_totals()
            self._apply_filter()
            self._mark_dirty()
            self._schedule_auto_save()

            messagebox.showinfo("Импорт", f"Импортировано {added} новых сотрудников.", parent=self)

        except Exception as e:
            logger.exception("Ошибка импорта табеля из Excel")
            messagebox.showerror("Импорт", f"Ошибка чтения файла:\n{e}", parent=self)

    def copy_from_month(self):
        if self.read_only:
            return

        addr = normalize_spaces(self.cmb_address.get() or "")
        oid = normalize_spaces(self.cmb_object_id.get() or "")
        current_dep = normalize_spaces(self.cmb_department.get() or "")

        if not addr and not oid:
            messagebox.showwarning("Копирование", "Укажите адрес/ID объекта.", parent=self)
            return

        if current_dep == "Все":
            messagebox.showwarning("Копирование", "Выберите конкретное подразделение (не 'Все').", parent=self)
            return

        user_id = self._safe_current_user_id()
        if not user_id:
            messagebox.showerror("Копирование", "Не удалось определить пользователя.", parent=self)
            return

        cy, cm = self.get_year_month()
        dlg = CopyFromDialog(self, init_year=cy if cm > 1 else cy - 1, init_month=cm - 1 if cm > 1 else 12)
        if not dlg.result:
            return

        src_y = dlg.result["year"]
        src_m = dlg.result["month"]
        with_hours = bool(dlg.result["with_hours"])
        mode = dlg.result["mode"]

        try:
            found_rows = load_timesheet_rows_for_copy_from_db(
                object_id=oid or None,
                object_addr=addr,
                department=current_dep,
                year=src_y,
                month=src_m,
                user_id=user_id,
                with_hours=with_hours,
            )

            if not found_rows:
                messagebox.showinfo(
                    "Копирование",
                    "В БД не найден табель-источник для выбранного месяца/объекта/подразделения.\n"
                    "Проверьте, что в прошлом месяце табель был сохранён в БД.",
                    parent=self,
                )
                return

            current_year, current_month = self.get_year_month()
            found_rows = deduplicate_timesheet_rows(found_rows, current_year, current_month)

            if mode == "replace":
                self.model_rows_all.clear()
                self.clear_selection()

            existing = {make_row_key(r.get("fio", ""), r.get("tbn", "")) for r in self.model_rows_all}
            added = 0
            for rec in found_rows:
                key = make_row_key(rec.get("fio", ""), rec.get("tbn", ""))
                if key in existing:
                    continue
                self.model_rows_all.append(rec)
                existing.add(key)
                added += 1

            self._recalc_all_row_totals()
            self._apply_filter()
            self._mark_dirty()
            self._schedule_auto_save()
            messagebox.showinfo("Копирование", f"Скопировано {added} сотрудников.", parent=self)

        except Exception as e:
            logger.exception("Ошибка копирования из месяца")
            messagebox.showerror("Копирование", f"Ошибка при копировании из БД:\n{e}", parent=self)

    def _get_brigadier_map_for_current_export(self) -> dict[str, str]:
        if self._active_header_id:
            try:
                return load_brigadiers_map_for_header(int(self._active_header_id))
            except Exception:
                logger.exception("Не удалось загрузить карту бригадиров для текущего header_id")

        result: dict[str, str] = {}
        dep = normalize_spaces(self.cmb_department.get() or "")
        if not dep or dep == "Все":
            return result

        try:
            if not self._brig_assign or not self._brig_names:
                self._reload_brigadier_filter_data()

            for emp_tbn, brig_tbn in self._brig_assign.items():
                emp_tbn_norm = normalize_tbn(emp_tbn)
                brig_tbn_norm = normalize_tbn(brig_tbn)
                if not emp_tbn_norm:
                    continue

                brig_fio = normalize_spaces(self._brig_names.get(brig_tbn_norm, "")) if brig_tbn_norm else ""
                result[emp_tbn_norm] = brig_fio
        except Exception:
            logger.exception("Не удалось собрать карту бригадиров для экспорта текущего табеля")

        return result

    def export_current_timesheet_to_excel(self):
        try:
            if hasattr(self, "grid"):
                self.grid.close_editor(commit=True)
        except Exception:
            pass
    
        year, month = self.get_year_month()
        addr = normalize_spaces(self.cmb_address.get() or "")
        oid = normalize_spaces(self.cmb_object_id.get() or "")
        dep = normalize_spaces(self.cmb_department.get() or "")
    
        if not self.model_rows_all:
            messagebox.showinfo("Выгрузка", "В табеле нет строк для выгрузки.", parent=self)
            return
    
        obj_part = oid or addr or "без_объекта"
        dep_part = dep or "без_подразделения"
    
        path = filedialog.asksaveasfilename(
            parent=self,
            title="Сохранить табель в Excel",
            defaultextension=".xlsx",
            initialfile=f"Табель_{safe_filename(obj_part)}_{safe_filename(dep_part)}_{year}_{month:02d}.xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Все", "*.*")],
        )
        if not path:
            return
    
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Табель"
    
            brig_map = self._get_brigadier_map_for_current_export()
    
            prepared_by = ""
            try:
                user = getattr(self.app_ref, "current_user", None) or {}
                prepared_by = normalize_spaces(user.get("full_name") or user.get("username") or "")
            except Exception:
                prepared_by = ""
    
            build_printable_timesheet_sheet(
                ws,
                year=year,
                month=month,
                object_addr=addr,
                object_id=oid,
                department=dep,
                rows=self.model_rows_all,
                brig_map=brig_map,
                prepared_by=prepared_by,
            )
    
            wb.save(path)
    
            messagebox.showinfo(
                "Выгрузка",
                f"Готово.\nСтрок: {len(self.model_rows_all)}\nФайл: {path}",
                parent=self,
            )
        except Exception as e:
            logger.exception("Ошибка выгрузки текущего табеля в Excel")
            messagebox.showerror("Выгрузка", f"Ошибка выгрузки:\n{e}", parent=self)

    # --------------------------------------------------------
    # Сохранение
    # --------------------------------------------------------

    def _current_file_path(self) -> Optional[Path]:
        addr = normalize_spaces(self.cmb_address.get() or "")
        oid = normalize_spaces(self.cmb_object_id.get() or "")
        dep = normalize_spaces(self.cmb_department.get() or "")
        if not addr and not oid:
            return None

        year, month = self.get_year_month()
        id_part = safe_filename(oid) if oid else safe_filename(addr)
        dep_part = safe_filename(dep) if dep and dep != "Все" else "ВсеПодразделения"
        return self.out_dir / f"Объектный_табель_{id_part}_{dep_part}_{year}_{month:02d}.xlsx"

    def _ensure_export_sheet(self, wb) -> Any:
        if "Табель" in wb.sheetnames:
            ws = wb["Табель"]
            hdr_first = normalize_spaces(str(ws.cell(1, 1).value or ""))
            if hdr_first == "ID объекта":
                return ws

            base = "Табель_OLD"
            i = 1
            new_name = base
            while new_name in wb.sheetnames:
                i += 1
                new_name = f"{base}{i}"
            ws.title = new_name

        ws2 = wb.create_sheet("Табель")
        hdr = (
            ["ID объекта", "Адрес", "Месяц", "Год", "ФИО", "Табельный №", "Подразделение"]
            + [str(i) for i in range(1, 32)]
            + ["Итого дней", "Итого часов по табелю", "В т.ч. ночных", "Переработка день", "Переработка ночь"]
        )
        ws2.append(hdr)
        ws2.column_dimensions["A"].width = 14
        ws2.column_dimensions["B"].width = 40
        ws2.column_dimensions["C"].width = 10
        ws2.column_dimensions["D"].width = 8
        ws2.column_dimensions["E"].width = 28
        ws2.column_dimensions["F"].width = 14
        ws2.column_dimensions["G"].width = 20
        for i in range(8, 8 + 31):
            ws2.column_dimensions[get_column_letter(i)].width = 6
        ws2.column_dimensions[get_column_letter(39)].width = 10
        ws2.column_dimensions[get_column_letter(40)].width = 18
        ws2.column_dimensions[get_column_letter(41)].width = 16
        ws2.column_dimensions[get_column_letter(42)].width = 14
        ws2.column_dimensions[get_column_letter(43)].width = 14
        ws2.freeze_panes = "A2"
        return ws2

    def _save_backup_excel(
        self,
        fpath: Path,
        object_id: str,
        object_addr: str,
        department: str,
        year: int,
        month: int,
    ):
        fpath.parent.mkdir(parents=True, exist_ok=True)

        wb = load_workbook(fpath) if fpath.exists() else Workbook()
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
            wb.remove(wb.active)

        ws = self._ensure_export_sheet(wb)

        oid_norm = normalize_spaces(object_id or "")
        addr_norm = normalize_spaces(object_addr or "")
        dep_norm = normalize_spaces(department or "")

        to_del = []
        for r in range(2, ws.max_row + 1):
            row_oid = normalize_spaces(str(ws.cell(r, 1).value or ""))
            row_addr = normalize_spaces(str(ws.cell(r, 2).value or ""))
            row_month = int(ws.cell(r, 3).value or 0)
            row_year = int(ws.cell(r, 4).value or 0)
            row_dep = normalize_spaces(str(ws.cell(r, 7).value or ""))

            if row_oid == oid_norm and row_addr == addr_norm and row_month == month and row_year == year and row_dep == dep_norm:
                to_del.append(r)

        for r in reversed(to_del):
            ws.delete_rows(r, 1)

        for rec in self.model_rows_all:
            fio = normalize_spaces(rec.get("fio") or "")
            tbn = normalize_tbn(rec.get("tbn"))
            hours_list = normalize_hours_list(rec.get("hours"), year, month)
            totals = calc_row_totals(hours_list, year, month)

            row_values = [oid_norm, addr_norm, month, year, fio, tbn, dep_norm] + hours_list + [
                totals["days"] or None,
                totals["hours"] or None,
                totals["night_hours"] or None,
                totals["ot_day"] or None,
                totals["ot_night"] or None,
            ]
            ws.append(row_values)

        temp_path: Optional[Path] = None
        try:
            with tempfile.NamedTemporaryFile(
                suffix=".xlsx",
                delete=False,
                dir=str(fpath.parent),
            ) as tmp:
                temp_path = Path(tmp.name)

            wb.save(temp_path)
            temp_path.replace(fpath)
        finally:
            if temp_path and temp_path.exists():
                try:
                    temp_path.unlink(missing_ok=True)
                except Exception:
                    pass

    def _save_all_internal(self, show_messages: bool, is_auto: bool = False) -> bool:
        if self.read_only:
            if show_messages:
                messagebox.showinfo("Объектный табель", "Сохранение недоступно в режиме просмотра.", parent=self)
            return False

        try:
            self.grid.close_editor(commit=True)
        except Exception:
            pass

        def fail(msg: str, level: str = "warning") -> bool:
            if show_messages:
                if level == "error":
                    messagebox.showerror("Сохранение", msg, parent=self)
                else:
                    messagebox.showwarning("Сохранение", msg, parent=self)
            if is_auto:
                self._mark_save_error(msg if len(msg) < 80 else "Ошибка авто‑сохранения")
            return False

        object_addr = normalize_spaces(self.cmb_address.get() or "")
        object_id = normalize_spaces(self.cmb_object_id.get() or "")
        year, month = self.get_year_month()
        department = normalize_spaces(self.cmb_department.get() or "")

        if department == "Все":
            return fail("Для сохранения выберите конкретное подразделение.")

        user_id = self._safe_current_user_id()
        if not user_id:
            return fail("Не удалось определить пользователя.", level="error")

        self._sync_object_id_values_silent()
        object_addr = normalize_spaces(self.cmb_address.get() or "")
        object_id = normalize_spaces(self.cmb_object_id.get() or "")

        if not object_addr:
            return fail("Не задан адрес объекта. Выберите адрес из списка.")

        if self.address_options and object_addr not in self.address_options:
            self.cmb_object_id.set("")
            return fail("Адрес объекта введён вручную и не найден в справочнике.\nВыберите адрес из списка.")

        objects_for_addr = [
            (normalize_spaces(code), normalize_spaces(a), normalize_spaces(short_name))
            for (code, a, short_name) in self.objects_full
            if normalize_spaces(a) == object_addr
        ]
        ids_for_addr = sorted({code for code, _, _ in objects_for_addr if code})
        if len(ids_for_addr) > 1 and not object_id:
            return fail("По выбранному адресу найдено несколько объектов.\nСначала выберите корректный ID объекта.")

        if object_id and ids_for_addr and object_id not in ids_for_addr:
            return fail("Выбранный ID объекта не соответствует адресу.\nИсправьте адрес или ID.")

        for rec in self.model_rows_all:
            rec["hours"] = normalize_hours_list(rec.get("hours"), year, month)
            rec["_totals"] = calc_row_totals(rec["hours"], year, month)

        validation_errors = validate_rows_before_save(self.model_rows_all, year, month)
        if validation_errors:
            preview = "\n".join(f"• {e}" for e in validation_errors[:15])
            if len(validation_errors) > 15:
                preview += f"\n• ... и ещё {len(validation_errors) - 15}"
            return fail(f"Табель содержит ошибки:\n\n{preview}", level="error")

        suspicious = find_suspicious_cells(self.model_rows_all, year, month)
        if suspicious and not is_auto and show_messages:
            dlg = SuspiciousHoursWarningDialog(self, suspicious, context="сохранением")
            self.wait_window(dlg)
            if not dlg.result:
                return False
        elif suspicious and is_auto:
            logger.warning(
                "Авто‑сохранение: обнаружено %s подозрительных значений часов (> %s)",
                len(suspicious),
                MAX_HOURS_PER_DAY,
            )

        employees_for_check = []
        for rec in self.model_rows_all:
            fio = normalize_spaces(rec.get("fio") or "")
            tbn = normalize_tbn(rec.get("tbn"))
            if fio or tbn:
                employees_for_check.append((fio, tbn))

        try:
            duplicates = find_duplicate_employees_for_timesheet(
                object_id=object_id or None,
                object_addr=object_addr,
                department=department,
                year=year,
                month=month,
                user_id=user_id,
                employees=employees_for_check,
            )
        except Exception as e:
            logger.exception("Ошибка проверки дублей сотрудников между табелями")
            return fail(f"Ошибка при проверке дублей сотрудников:\n{e}", level="error")

        if duplicates:
            lines = []
            for d in duplicates[:20]:
                emp_fio = d.get("fio") or ""
                emp_tbn = d.get("tbn") or ""
                uname = d.get("full_name") or d.get("username") or f"id={d.get('user_id')}"
                lines.append(f"- {emp_fio} (таб.№ {emp_tbn}) — уже есть в табеле пользователя {uname}")

            msg = (
                "Найдены сотрудники, которые уже есть в табелях других пользователей "
                "по этому объекту/подразделению/месяцу:\n\n"
                + "\n".join(lines)
            )
            if len(duplicates) > 20:
                msg += f"\n\n... и ещё {len(duplicates) - 20}"
            msg += "\n\nСохранение отменено. Удалите этих сотрудников из табеля."
            return fail(msg)

        try:
            header_id = upsert_timesheet_header(
                object_id=object_id or None,
                object_addr=object_addr,
                department=department,
                year=year,
                month=month,
                user_id=user_id,
            )
            replace_timesheet_rows(header_id, self.model_rows_all, year, month)
            self._active_header_id = header_id
        except Exception as e:
            logger.exception("Ошибка сохранения табеля в БД")
            return fail(f"Ошибка сохранения в БД:\n{e}", level="error")

        try:
            fpath = self._current_file_path()
            if fpath:
                self._save_backup_excel(
                    fpath=fpath,
                    object_id=object_id,
                    object_addr=object_addr,
                    department=department,
                    year=year,
                    month=month,
                )
        except Exception as e:
            logger.exception("Ошибка резервного сохранения в Excel")
            if show_messages:
                messagebox.showwarning(
                    "Сохранение",
                    f"В БД табель сохранён, но ошибка при записи в Excel:\n{e}",
                    parent=self,
                )

        self._loaded_context = self._capture_current_context()
        self._loaded_context["header_id"] = self._active_header_id
        self._mark_saved(auto=is_auto)

        if show_messages and not is_auto:
            fpath = self._current_file_path()
            if fpath:
                messagebox.showinfo("Сохранение", f"Табель сохранён в БД и в файл:\n{fpath}", parent=self)
            else:
                messagebox.showinfo("Сохранение", "Табель сохранён в БД.", parent=self)

        return True

    def save_all(self):
        self._save_all_internal(show_messages=True, is_auto=False)

    # --------------------------------------------------------
    # Автоподгонка колонок
    # --------------------------------------------------------

    def _content_total_width(self, fio_px: Optional[int] = None) -> int:
        px = self.COLPX.copy()
        if fio_px is not None:
            px["fio"] = fio_px
    
        return (
            px["fio"]
            + px["tbn"]
            + 31 * px["day"]
            + px["days"]
            + px.get("hours", 58)
            + px.get("ot_day", px.get("hours", 58))
            + px.get("ot_night", px.get("hours", 58))
            + px["del"]
        )

    def _auto_fit_columns(self):
        try:
            viewport = self.grid.body.winfo_width() if hasattr(self, "grid") else 0
        except Exception:
            viewport = 0

        if viewport <= 1:
            self.after(120, self._auto_fit_columns)
            return

        total = self._content_total_width()
        new_fio = self.COLPX["fio"]
        if total < viewport:
            new_fio = min(self.MAX_FIO_PX, self.COLPX["fio"] + (viewport - total))

        if int(new_fio) != int(self.COLPX["fio"]):
            self.COLPX["fio"] = int(new_fio)
            try:
                if hasattr(self, "grid"):
                    self.grid.COLPX = self.COLPX
                    self.grid._build_columns()
                    self.grid._draw_header()
                    self._grid_refresh(rows_changed=True)
            except Exception:
                logger.exception("Ошибка автоподгонки колонок")

    def _on_window_configure(self, _evt):
        try:
            if self._fit_job is not None:
                self.after_cancel(self._fit_job)
        except Exception:
            pass
        self._fit_job = self.after(150, self._auto_fit_columns)

# ============================================================
# Мои табели
# ============================================================

class MyTimesheetsPage(tk.Frame):
    def __init__(self, master, app_ref):
        super().__init__(master, bg=TS_COLORS["bg"])
        self.app_ref = app_ref

        self.tree = None
        self._headers: List[Dict[str, Any]] = []

        self.var_year = tk.StringVar(value=str(datetime.now().year))
        self.var_month = tk.StringVar(value="Все")
        self.var_dep = tk.StringVar()
        self.var_obj_addr = tk.StringVar()

        self._build_ui()
        self._load_data()

    def _build_ui(self):
        hdr = tk.Frame(self, bg=TS_COLORS["accent"], pady=4)
        hdr.pack(fill="x")
        tk.Label(
            hdr,
            text="📂 Мои табели",
            font=("Segoe UI", 12, "bold"),
            bg=TS_COLORS["accent"],
            fg="white",
            padx=12,
        ).pack(side="left")

        filter_pnl = tk.LabelFrame(
            self,
            text=" 🔍 Фильтры ",
            font=("Segoe UI", 9, "bold"),
            bg=TS_COLORS["panel"],
            fg=TS_COLORS["accent"],
            relief="groove",
            bd=1,
            padx=10,
            pady=8,
        )
        filter_pnl.pack(fill="x", padx=10, pady=(8, 4))

        tk.Label(filter_pnl, text="Год:", font=("Segoe UI", 9), bg=TS_COLORS["panel"]).grid(
            row=0, column=0, sticky="e", padx=(0, 6), pady=3
        )
        tk.Spinbox(
            filter_pnl,
            from_=2000,
            to=2100,
            width=7,
            textvariable=self.var_year,
            font=("Segoe UI", 9),
        ).grid(row=0, column=1, sticky="w", pady=3)

        tk.Label(filter_pnl, text="Месяц:", font=("Segoe UI", 9), bg=TS_COLORS["panel"]).grid(
            row=0, column=2, sticky="e", padx=(16, 6), pady=3
        )
        cmb_month = ttk.Combobox(
            filter_pnl,
            state="readonly",
            width=14,
            textvariable=self.var_month,
            values=["Все"] + [month_name_ru(i) for i in range(1, 13)],
        )
        cmb_month.grid(row=0, column=3, sticky="w", pady=3)

        tk.Label(filter_pnl, text="Подразделение:", font=("Segoe UI", 9), bg=TS_COLORS["panel"]).grid(
            row=1, column=0, sticky="e", padx=(0, 6), pady=3
        )
        ttk.Entry(filter_pnl, width=26, textvariable=self.var_dep, font=("Segoe UI", 9)).grid(
            row=1, column=1, sticky="w", pady=3
        )

        tk.Label(filter_pnl, text="Объект (адрес):", font=("Segoe UI", 9), bg=TS_COLORS["panel"]).grid(
            row=1, column=2, sticky="e", padx=(16, 6), pady=3
        )
        ttk.Entry(filter_pnl, width=36, textvariable=self.var_obj_addr, font=("Segoe UI", 9)).grid(
            row=1, column=3, sticky="w", pady=3
        )

        btn_frame = tk.Frame(filter_pnl, bg=TS_COLORS["panel"])
        btn_frame.grid(row=0, column=4, rowspan=2, sticky="e", padx=(20, 0))

        tk.Button(
            btn_frame,
            text="🔄 Применить",
            font=("Segoe UI", 9, "bold"),
            bg=TS_COLORS["btn_save_bg"],
            fg=TS_COLORS["btn_save_fg"],
            activebackground="#0d47a1",
            activeforeground="white",
            relief="flat",
            cursor="hand2",
            padx=10,
            pady=4,
            command=self._load_data,
        ).pack(fill="x", pady=(0, 4))

        ttk.Button(btn_frame, text="Сбросить фильтры", command=self._reset_filters).pack(fill="x", pady=(0, 4))
        ttk.Button(btn_frame, text="📊 Выгрузить в Excel", command=self._export_to_excel).pack(fill="x")

        filter_pnl.grid_columnconfigure(3, weight=1)

        tbl_frame = tk.LabelFrame(
            self,
            text=" 📋 Список табелей ",
            font=("Segoe UI", 9, "bold"),
            bg=TS_COLORS["panel"],
            fg=TS_COLORS["accent"],
            relief="groove",
            bd=1,
        )
        tbl_frame.pack(fill="both", expand=True, padx=10, pady=(4, 4))

        cols = ("year", "month", "object", "department", "updated_at")
        self.tree = ttk.Treeview(tbl_frame, columns=cols, show="headings", selectmode="browse")

        heads = {
            "year": ("Год", 70, "center"),
            "month": ("Месяц", 110, "center"),
            "object": ("Объект", 380, "w"),
            "department": ("Подразделение", 210, "w"),
            "updated_at": ("Обновлён", 150, "center"),
        }
        for col, (text, width, anchor) in heads.items():
            self.tree.heading(col, text=text)
            self.tree.column(col, width=width, anchor=anchor, stretch=(col == "object"))

        vsb = ttk.Scrollbar(tbl_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        self.tree.bind("<Double-1>", self._on_open)
        self.tree.bind("<Return>", self._on_open)

        bottom = tk.Frame(self, bg=TS_COLORS["accent_light"], pady=5)
        bottom.pack(fill="x", padx=10, pady=(0, 8))

        self.lbl_count = tk.Label(
            bottom,
            text="Табелей: 0",
            font=("Segoe UI", 9, "bold"),
            fg=TS_COLORS["accent"],
            bg=TS_COLORS["accent_light"],
        )
        self.lbl_count.pack(side="left", padx=10)

        tk.Label(
            bottom,
            text="Двойной щелчок или Enter — открыть табель для редактирования",
            font=("Segoe UI", 9, "italic"),
            fg="#555",
            bg=TS_COLORS["accent_light"],
        ).pack(side="right", padx=10)

    def _reset_filters(self):
        self.var_year.set(str(datetime.now().year))
        self.var_month.set("Все")
        self.var_dep.set("")
        self.var_obj_addr.set("")
        self._load_data()

    def _load_data(self):
        self.tree.delete(*self.tree.get_children())
        self._headers.clear()

        user = getattr(self.app_ref, "current_user", None) or {}
        user_id = user.get("id")
        if not user_id:
            messagebox.showwarning("Мои табели", "Не определён текущий пользователь.", parent=self)
            return

        try:
            year = int(normalize_spaces(self.var_year.get() or ""))
        except Exception:
            year = None

        month = None
        month_name = normalize_spaces(self.var_month.get() or "")
        if month_name and month_name != "Все":
            try:
                month = [month_name_ru(i) for i in range(1, 13)].index(month_name) + 1
            except ValueError:
                month = None

        dep = normalize_spaces(self.var_dep.get() or "") or None
        addr_substr = normalize_spaces(self.var_obj_addr.get() or "") or None

        try:
            headers = load_user_timesheet_headers(user_id, year, month, dep, addr_substr)
        except Exception as e:
            messagebox.showerror("Мои табели", f"Ошибка загрузки из БД:\n{e}", parent=self)
            return

        self._headers = headers

        for h in headers:
            month_ru = month_name_ru(h["month"]) if 1 <= h["month"] <= 12 else str(h["month"])
            obj_display = h["object_addr"] or ""
            if h.get("object_id"):
                obj_display = f"[{h['object_id']}] {obj_display}"

            upd = h.get("updated_at")
            upd_str = upd.strftime("%d.%m.%Y %H:%M") if isinstance(upd, datetime) else ""

            self.tree.insert(
                "",
                "end",
                iid=str(h["id"]),
                values=(h["year"], month_ru, obj_display, h.get("department") or "", upd_str),
            )

        self.lbl_count.config(text=f"Табелей: {len(headers)}")

    def _export_to_excel(self):
        if not self._headers:
            messagebox.showinfo("Экспорт", "Нет данных для выгрузки.", parent=self)
            return

        path = filedialog.asksaveasfilename(
            parent=self,
            title="Сохранить мои табели в Excel",
            defaultextension=".xlsx",
            initialfile=f"Мои_табели_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Все", "*.*")],
        )
        if not path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Мои табели"

            header = (
                ["Год", "Месяц", "Адрес", "ID объекта", "Подразделение", "ФИО", "Табельный №", "ФИО бригадира"]
                + [str(i) for i in range(1, 32)]
                + ["Итого дней", "Итого часов", "В т.ч. ночных", "Переработка день", "Переработка ночь"]
            )
            ws.append(header)

            widths = [6, 10, 40, 14, 22, 28, 12, 28] + [6] * 31 + [10, 12, 16, 16, 16]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

            total_rows = 0
            for h in self._headers:
                rows_data = load_timesheet_rows_by_header_id(int(h["id"]))
                brig_map = load_brigadiers_map_for_header(int(h["id"]))

                for r in rows_data:
                    tbn = normalize_tbn(r.get("tbn"))
                    brig_fio = brig_map.get(tbn, "") if tbn else ""

                    ws.append(
                        [
                            h["year"],
                            h["month"],
                            h.get("object_addr", ""),
                            h.get("object_id", ""),
                            h.get("department", ""),
                            r.get("fio", ""),
                            r.get("tbn", ""),
                            brig_fio,
                        ]
                        + (r.get("hours_raw") or [None] * 31)
                        + [
                            r.get("total_days"),
                            r.get("total_hours"),
                            r.get("night_hours"),
                            r.get("overtime_day"),
                            r.get("overtime_night"),
                        ]
                    )
                    total_rows += 1

            wb.save(path)
            messagebox.showinfo("Экспорт", f"Готово.\nСтрок: {total_rows}\nФайл: {path}", parent=self)
        except Exception as e:
            logger.exception("Ошибка экспорта моих табелей")
            messagebox.showerror("Экспорт", f"Ошибка:\n{e}", parent=self)

    def _get_selected_header(self) -> Optional[Dict[str, Any]]:
        sel = self.tree.selection()
        if not sel:
            return None
        try:
            hid = int(sel[0])
            return next((h for h in self._headers if int(h["id"]) == hid), None)
        except Exception:
            return None

    def _on_open(self, event=None):
        h = self._get_selected_header()
        if not h:
            return
    
        key = f"timesheet_{int(h['id'])}"
        builder = lambda parent: TimesheetPage(
            parent,
            app_ref=self.app_ref,
            init_header_id=int(h["id"]),
            init_object_id=h.get("object_id"),
            init_object_addr=h.get("object_addr"),
            init_department=h.get("department"),
            init_year=h.get("year"),
            init_month=h.get("month"),
            read_only=False,
            owner_user_id=h.get("user_id") or (getattr(self.app_ref, "current_user", None) or {}).get("id"),
        )
    
        if hasattr(self.app_ref, "open_page_in_tab"):
            self.app_ref.open_page_in_tab(key, builder)
        else:
            self.app_ref._show_page(key, builder)
    
        _set_timesheet_tab_title(self.app_ref, key, h)

# ============================================================
# Реестр табелей
# ============================================================

class TimesheetRegistryPage(tk.Frame):
    def __init__(self, master, app_ref):
        super().__init__(master, bg=TS_COLORS["bg"])
        self.app_ref = app_ref

        self.tree = None
        self._headers: List[Dict[str, Any]] = []
        self._all_departments: List[str] = []

        self.var_year = tk.StringVar(value=str(datetime.now().year))
        self.var_month = tk.StringVar(value="Все")
        self.var_dep = tk.StringVar(value="Все")
        self.var_obj_addr = tk.StringVar()
        self.var_obj_id = tk.StringVar()

        self._filter_job = None

        self._build_ui()
        self._load_departments()
        self._load_data()

    def _build_ui(self):
        hdr = tk.Frame(self, bg=TS_COLORS["accent"], pady=6)
        hdr.pack(fill="x")
        tk.Label(
            hdr,
            text="📊 Реестр табелей",
            font=("Segoe UI", 12, "bold"),
            bg=TS_COLORS["accent"],
            fg="white",
            padx=12,
        ).pack(side="left")

        filter_pnl = tk.LabelFrame(
            self,
            text=" 🔍 Фильтры ",
            font=("Segoe UI", 9, "bold"),
            bg=TS_COLORS["panel"],
            fg=TS_COLORS["accent"],
            relief="groove",
            bd=1,
            padx=10,
            pady=8,
        )
        filter_pnl.pack(fill="x", padx=10, pady=(8, 4))

        fields_frame = tk.Frame(filter_pnl, bg=TS_COLORS["panel"])
        fields_frame.pack(side="left", fill="x", expand=True)

        btn_frame = tk.Frame(filter_pnl, bg=TS_COLORS["panel"])
        btn_frame.pack(side="right", padx=(20, 0))

        row0 = tk.Frame(fields_frame, bg=TS_COLORS["panel"])
        row0.pack(fill="x", pady=(0, 4))

        tk.Label(row0, text="Год:", font=("Segoe UI", 9), bg=TS_COLORS["panel"]).pack(side="left", padx=(0, 4))
        tk.Spinbox(row0, from_=2000, to=2100, width=6, textvariable=self.var_year, font=("Segoe UI", 9)).pack(side="left")
        self.var_year.trace_add("write", self._on_year_changed)

        tk.Label(row0, text="Месяц:", font=("Segoe UI", 9), bg=TS_COLORS["panel"]).pack(side="left", padx=(16, 4))
        cmb_month = ttk.Combobox(
            row0,
            state="readonly",
            width=12,
            textvariable=self.var_month,
            values=["Все"] + [month_name_ru(i) for i in range(1, 13)],
        )
        cmb_month.pack(side="left")
        cmb_month.bind("<<ComboboxSelected>>", lambda _e: self._load_data())

        tk.Label(row0, text="Подразделение:", font=("Segoe UI", 9), bg=TS_COLORS["panel"]).pack(side="left", padx=(16, 4))
        self._cmb_dep = ttk.Combobox(row0, state="readonly", width=28, textvariable=self.var_dep, values=["Все"])
        self._cmb_dep.pack(side="left")
        self._cmb_dep.bind("<<ComboboxSelected>>", lambda _e: self._load_data())

        row1 = tk.Frame(fields_frame, bg=TS_COLORS["panel"])
        row1.pack(fill="x")

        tk.Label(row1, text="Объект (адрес):", font=("Segoe UI", 9), bg=TS_COLORS["panel"]).pack(side="left", padx=(0, 4))
        ttk.Entry(row1, width=40, textvariable=self.var_obj_addr, font=("Segoe UI", 9)).pack(side="left")
        self.var_obj_addr.trace_add("write", self._on_text_filter_changed)

        tk.Label(row1, text="ID объекта:", font=("Segoe UI", 9), bg=TS_COLORS["panel"]).pack(side="left", padx=(16, 4))
        ttk.Entry(row1, width=16, textvariable=self.var_obj_id, font=("Segoe UI", 9)).pack(side="left")
        self.var_obj_id.trace_add("write", self._on_text_filter_changed)

        tk.Button(
            btn_frame,
            text="🔄 Применить",
            font=("Segoe UI", 9, "bold"),
            bg=TS_COLORS["btn_save_bg"],
            fg=TS_COLORS["btn_save_fg"],
            activebackground="#0d47a1",
            activeforeground="white",
            relief="flat",
            cursor="hand2",
            padx=10,
            pady=4,
            command=self._load_data,
        ).pack(fill="x", pady=(0, 4))

        ttk.Button(btn_frame, text="Сбросить фильтры", command=self._reset_filters).pack(fill="x", pady=(0, 4))
        ttk.Button(btn_frame, text="📊 Выгрузить в Excel", command=self._export_to_excel).pack(fill="x", pady=(0, 4))
        ttk.Button(btn_frame, text="📈 Отчёт по заполненности", command=self._export_fill_report).pack(fill="x")

        tbl_frame = tk.LabelFrame(
            self,
            text=" 📋 Список табелей ",
            font=("Segoe UI", 9, "bold"),
            bg=TS_COLORS["panel"],
            fg=TS_COLORS["accent"],
            relief="groove",
            bd=1,
        )
        tbl_frame.pack(fill="both", expand=True, padx=10, pady=(4, 4))

        cols = ("year", "month", "object", "department", "user", "updated_at")
        self.tree = ttk.Treeview(tbl_frame, columns=cols, show="headings", selectmode="browse")

        heads = {
            "year": ("Год", 65, "center"),
            "month": ("Месяц", 95, "center"),
            "object": ("Объект", 310, "w"),
            "department": ("Подразделение", 190, "w"),
            "user": ("Пользователь", 190, "w"),
            "updated_at": ("Обновлён", 140, "center"),
        }
        for col, (text, width, anchor) in heads.items():
            self.tree.heading(col, text=text)
            self.tree.column(col, width=width, anchor=anchor, stretch=(col == "object"))

        vsb = ttk.Scrollbar(tbl_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        self.tree.bind("<Double-1>", self._on_open)
        self.tree.bind("<Return>", self._on_open)

        bottom = tk.Frame(self, bg=TS_COLORS["accent_light"], pady=5)
        bottom.pack(fill="x", padx=10, pady=(0, 8))

        self.lbl_count = tk.Label(
            bottom,
            text="Табелей: 0",
            font=("Segoe UI", 9, "bold"),
            fg=TS_COLORS["accent"],
            bg=TS_COLORS["accent_light"],
        )
        self.lbl_count.pack(side="left", padx=10)

        tk.Label(
            bottom,
            text="Двойной щелчок или Enter — открыть табель",
            font=("Segoe UI", 9, "italic"),
            fg="#555",
            bg=TS_COLORS["accent_light"],
        ).pack(side="right", padx=10)

    def _load_departments(self):
        self._all_departments = []
        try:
            with db_cursor() as (_conn, cur):
                cur.execute(
                    """
                    SELECT DISTINCT department
                    FROM timesheet_headers
                    WHERE department IS NOT NULL
                      AND TRIM(department) <> ''
                    ORDER BY department
                    """
                )
                self._all_departments = [normalize_spaces(r[0]) for r in cur.fetchall() if normalize_spaces(r[0])]
        except Exception:
            logger.exception("Ошибка загрузки списка подразделений реестра")

        values = ["Все"] + self._all_departments
        self._cmb_dep.configure(values=values)
        if not self.var_dep.get() or self.var_dep.get() == "Все":
            self.var_dep.set("Все")

    def _on_text_filter_changed(self, *_):
        if self._filter_job is not None:
            self.after_cancel(self._filter_job)
        self._filter_job = self.after(400, self._load_data)

    def _on_year_changed(self, *_):
        if self._filter_job is not None:
            self.after_cancel(self._filter_job)
        self._filter_job = self.after(600, self._load_data)

    def _reset_filters(self):
        if self._filter_job is not None:
            self.after_cancel(self._filter_job)
            self._filter_job = None

        self.var_year.set(str(datetime.now().year))
        self.var_month.set("Все")
        self.var_dep.set("Все")
        self.var_obj_addr.set("")
        self.var_obj_id.set("")
        self._load_data()

    def _load_data(self):
        if self._filter_job is not None:
            self.after_cancel(self._filter_job)
        self._filter_job = None

        self.tree.delete(*self.tree.get_children())
        self._headers.clear()

        year = None
        try:
            y = int(normalize_spaces(self.var_year.get() or ""))
            if 2000 <= y <= 2100:
                year = y
        except Exception:
            pass

        month = None
        month_name = normalize_spaces(self.var_month.get() or "")
        if month_name and month_name != "Все":
            for i in range(1, 13):
                if month_name_ru(i) == month_name:
                    month = i
                    break

        dep = normalize_spaces(self.var_dep.get() or "")
        if not dep or dep == "Все":
            dep = None

        addr_sub = normalize_spaces(self.var_obj_addr.get() or "") or None
        oid_sub = normalize_spaces(self.var_obj_id.get() or "") or None

        try:
            headers = load_all_timesheet_headers(
                year=year,
                month=month,
                department=dep,
                object_addr_substr=addr_sub,
                object_id_substr=oid_sub,
            )
        except Exception as e:
            messagebox.showerror("Реестр табелей", f"Ошибка загрузки из БД:\n{e}", parent=self)
            return

        self._headers = headers

        for h in headers:
            yr = h["year"]
            mn = h["month"]
            addr = h.get("object_addr") or ""
            oid = h.get("object_id") or ""
            dep_val = h.get("department") or ""
            user_name = h.get("full_name") or h.get("username") or ""
            upd = h.get("updated_at")

            month_ru = month_name_ru(mn) if 1 <= mn <= 12 else str(mn)
            obj_display = f"[{oid}] {addr}" if oid else addr
            upd_str = upd.strftime("%d.%m.%Y %H:%M") if isinstance(upd, datetime) else str(upd or "")

            self.tree.insert(
                "",
                "end",
                iid=str(h["id"]),
                values=(yr, month_ru, obj_display, dep_val, user_name, upd_str),
            )

        self.lbl_count.config(text=f"Табелей: {len(headers)}")

    def _export_to_excel(self):
        if not self._headers:
            messagebox.showinfo("Экспорт", "Нет данных для выгрузки.", parent=self)
            return

        path = filedialog.asksaveasfilename(
            parent=self,
            title="Сохранить реестр табелей в Excel",
            defaultextension=".xlsx",
            initialfile=f"Реестр_табелей_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Все", "*.*")],
        )
        if not path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Реестр табелей"

            header_row = (
                ["Год", "Месяц", "Адрес", "ID объекта", "Подразделение", "Пользователь", "ФИО", "Табельный №"]
                + [str(i) for i in range(1, 32)]
                + ["Итого_дней", "Итого_часов", "В т.ч. ночных", "Переработка_день", "Переработка_ночь"]
            )
            ws.append(header_row)

            widths = [6, 10, 40, 14, 22, 22, 28, 12] + [6] * 31 + [10, 14, 16, 16, 16]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

            total_rows = 0
            for h in self._headers:
                rows = load_timesheet_rows_by_header_id(int(h["id"]))
                user_display = h.get("full_name") or h.get("username") or ""
                for r in rows:
                    ws.append(
                        [
                            h["year"],
                            h["month"],
                            h.get("object_addr", ""),
                            h.get("object_id", ""),
                            h.get("department", ""),
                            user_display,
                            r["fio"],
                            r["tbn"],
                        ]
                        + (r.get("hours_raw") or [None] * 31)
                        + [
                            r.get("total_days"),
                            r.get("total_hours"),
                            r.get("night_hours"),
                            r.get("overtime_day"),
                            r.get("overtime_night"),
                        ]
                    )
                    total_rows += 1

            wb.save(path)
            messagebox.showinfo("Экспорт", f"Готово.\nСтрок: {total_rows}\nФайл: {path}", parent=self)
        except Exception as e:
            logger.exception("Ошибка экспорта реестра табелей")
            messagebox.showerror("Экспорт", f"Ошибка:\n{e}", parent=self)

    def _export_fill_report(self):
        if not self._headers:
            messagebox.showinfo("Отчёт по заполненности", "Нет данных для выгрузки.", parent=self)
            return

        import calendar as _cal

        today = datetime.now().date()
        path = filedialog.asksaveasfilename(
            parent=self,
            title="Сохранить отчёт по заполненности",
            defaultextension=".xlsx",
            initialfile=f"Заполненность_табелей_{today.strftime('%Y%m%d')}.xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Все", "*.*")],
        )
        if not path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Заполненность"

            ws.append(
                [
                    "Объект (адрес)",
                    "ID объекта",
                    "Подразделение",
                    "Пользователь",
                    "Год",
                    "Месяц",
                    "Дата обновления",
                    "Дней в периоде",
                    "Дней заполнено",
                    "Заполненность, %",
                ]
            )

            col_widths = [45, 14, 24, 24, 8, 12, 20, 16, 16, 18]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

            for cell in ws[1]:
                cell.font = Font(bold=True)

            red_fill = PatternFill("solid", fgColor="FFC7CE")
            yellow_fill = PatternFill("solid", fgColor="FFEB9C")
            green_fill = PatternFill("solid", fgColor="C6EFCE")

            row_num = 2
            for h in self._headers:
                yr = int(h["year"])
                mn = int(h["month"])
                addr = h.get("object_addr") or ""
                obj_id = h.get("object_id") or ""
                dep = h.get("department") or ""
                user_disp = h.get("full_name") or h.get("username") or ""
                upd = h.get("updated_at")
                upd_str = upd.strftime("%d.%m.%Y %H:%M") if isinstance(upd, datetime) else ""
                month_ru = month_name_ru(mn) if 1 <= mn <= 12 else str(mn)

                last_day = _cal.monthrange(yr, mn)[1]
                period_end = min(today, date(yr, mn, last_day))
                period_start = date(yr, mn, 1)

                if period_end < period_start:
                    days_in_period = 0
                    days_filled = 0
                else:
                    days_in_period = (period_end - period_start).days + 1
                    rows = load_timesheet_rows_by_header_id(int(h["id"]))
                    days_filled = 0
                    for d_idx in range(days_in_period):
                        for row in rows:
                            hrs = row.get("hours_raw") or []
                            if d_idx < len(hrs) and hrs[d_idx] is not None and str(hrs[d_idx]).strip():
                                days_filled += 1
                                break

                pct = round(days_filled / days_in_period * 100, 1) if days_in_period > 0 else 0.0

                ws.append(
                    [
                        addr,
                        obj_id,
                        dep,
                        user_disp,
                        yr,
                        month_ru,
                        upd_str,
                        days_in_period,
                        days_filled,
                        pct,
                    ]
                )

                cell = ws.cell(row=row_num, column=10)
                try:
                    v = float(cell.value or 0)
                    cell.fill = red_fill if v < 50 else yellow_fill if v < 90 else green_fill
                except Exception:
                    pass

                row_num += 1

            wb.save(path)
            messagebox.showinfo(
                "Отчёт по заполненности",
                f"Готово.\nТабелей: {len(self._headers)}\nФайл: {path}",
                parent=self,
            )
        except Exception as e:
            logger.exception("Ошибка выгрузки отчёта по заполненности")
            messagebox.showerror("Отчёт по заполненности", f"Ошибка:\n{e}", parent=self)

    def _get_selected_header(self) -> Optional[Dict[str, Any]]:
        sel = self.tree.selection()
        if not sel:
            return None
        try:
            hid = int(sel[0])
            return next((h for h in self._headers if int(h["id"]) == hid), None)
        except Exception:
            return None

    def _on_open(self, event=None):
        h = self._get_selected_header()
        if not h:
            return
    
        role = (getattr(self.app_ref, "current_user", None) or {}).get("role") or "specialist"
        read_only = role != "admin"
    
        key = f"timesheet_{int(h['id'])}"
        builder = lambda parent: TimesheetPage(
            parent,
            app_ref=self.app_ref,
            init_header_id=int(h["id"]),
            init_object_id=h.get("object_id"),
            init_object_addr=h.get("object_addr"),
            init_department=h.get("department"),
            init_year=int(h.get("year") or 0),
            init_month=int(h.get("month") or 0),
            read_only=read_only,
            owner_user_id=h.get("user_id"),
        )
    
        if hasattr(self.app_ref, "open_page_in_tab"):
            self.app_ref.open_page_in_tab(key, builder)
        else:
            self.app_ref._show_page(key, builder)
    
        _set_timesheet_tab_title(self.app_ref, key, h)
# ============================================================
# API для main_app
# ============================================================


def create_timesheet_page(parent, app_ref, **kwargs) -> TimesheetPage:
    return TimesheetPage(parent, app_ref=app_ref, **kwargs)


def create_my_timesheets_page(parent, app_ref) -> MyTimesheetsPage:
    return MyTimesheetsPage(parent, app_ref=app_ref)


def create_timesheet_registry_page(parent, app_ref) -> TimesheetRegistryPage:
    return TimesheetRegistryPage(parent, app_ref=app_ref)


__all__ = [
    "set_db_pool",
    "TimesheetPage",
    "MyTimesheetsPage",
    "TimesheetRegistryPage",
    "create_timesheet_page",
    "create_my_timesheets_page",
    "create_timesheet_registry_page",
]
