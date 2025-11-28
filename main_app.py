import os
import re
import sys
import csv
import json
import math
import calendar
import subprocess
import configparser
import urllib.request
import urllib.error
import urllib.parse
import traceback
import threading
import psycopg2
from psycopg2 import pool
from psycopg2.extras import RealDictCursor
from urllib.parse import urlparse, parse_qs
import hashlib
import os as _os
from io import BytesIO
from datetime import datetime, date
from pathlib import Path
from typing import List, Tuple, Optional, Any, Dict
import base64

# --- Импорты сторонних библиотек ---
try:
    import pandas as pd
except ImportError:
    pd = None

try:
    from PIL import Image, ImageTk
except Exception:
    Image = ImageTk = None

import logging

# Простейшее логирование в файл рядом с программой
logging.basicConfig(
    filename="main_app_log.txt",
    level=logging.DEBUG,
    format="%(asctime)s [%(levelname)s] %(message)s",
    encoding="utf-8",
)
logging.debug("=== main_app запущен ===")

# Мягкий импорт модулей
try:
    import BudgetAnalyzer  # должен содержать create_page(parent)
except Exception:
    BudgetAnalyzer = None

try:
    import assets_logo as _assets_logo
    _LOGO_BASE64 = getattr(_assets_logo, "LOGO_BASE64", None)
except Exception:
    _LOGO_BASE64 = None

try:
    import SpecialOrders  # должен содержать create_page/create_planning_page
except Exception:
    SpecialOrders = None

try:
    import timesheet_transformer  # должен содержать open_converter(parent)
except Exception:
    timesheet_transformer = None

# --- логируем импорт модуля питания ---
logging.debug("Пробуем импортировать meals_module...")
try:
    import meals_module  # обновлённый модуль питания (работает с БД)
    logging.debug(f"meals_module импортирован: {meals_module}")
except Exception:
    logging.exception("Ошибка при импорте meals_module")
    meals_module = None

try:
    import objects  # наш новый модуль
except Exception:
    logging.exception("Ошибка при импорте модуля objects")
    objects = None

# --- логируем импорт settings_manager ---
logging.debug("Пробуем импортировать settings_manager...")
try:
    import settings_manager as Settings
    logging.debug("settings_manager импортирован успешно")
except Exception:
    logging.exception("Ошибка при импорте settings_manager")
    Settings = None

import tkinter as tk
from tkinter import ttk, messagebox, simpledialog

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

APP_NAME = "Управление строительством (Главное меню)"

# ------------- КОНФИГ, СХЕМЫ И КОНСТАНТЫ -------------

CONFIG_FILE = "tabel_config.ini"
CONFIG_SECTION_PATHS = "Paths"
CONFIG_SECTION_UI = "UI"
CONFIG_SECTION_INTEGR = "Integrations"

KEY_OUTPUT_DIR = "output_dir"
KEY_EXPORT_PWD = "export_password"
KEY_SELECTED_DEP = "selected_department"

OUTPUT_DIR_DEFAULT = "Объектные_табели"
RAW_LOGO_URL = "https://raw.githubusercontent.com/alekseyvz-dotcom/TimesheetTransformer/main/logo.png"
TINY_PNG_BASE64 = (
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8"
    "/w8AAn8B9w3G2kIAAAAASUVORK5CYII="
)


def exe_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def config_path() -> Path:
    return exe_dir() / CONFIG_FILE


# Если settings_manager есть — используем его
if Settings:
    ensure_config = Settings.ensure_config
    read_config = Settings.read_config
    write_config = Settings.write_config

    get_output_dir_from_config = Settings.get_output_dir_from_config
    get_export_password_from_config = Settings.get_export_password_from_config

    get_selected_department_from_config = Settings.get_selected_department_from_config
    set_selected_department_in_config = Settings.set_selected_department_in_config
else:
    # fallback на ini‑файл
    def ensure_config():
        cp = config_path()
        if cp.exists():
            cfg = configparser.ConfigParser()
            cfg.read(cp, encoding="utf-8")
            changed = False
            if not cfg.has_section(CONFIG_SECTION_PATHS):
                cfg[CONFIG_SECTION_PATHS] = {}
                changed = True
            if KEY_OUTPUT_DIR not in cfg[CONFIG_SECTION_PATHS]:
                cfg[CONFIG_SECTION_PATHS][KEY_OUTPUT_DIR] = str(exe_dir() / OUTPUT_DIR_DEFAULT)
                changed = True
            if not cfg.has_section(CONFIG_SECTION_UI):
                cfg[CONFIG_SECTION_UI] = {}
                changed = True
            if KEY_SELECTED_DEP not in cfg[CONFIG_SECTION_UI]:
                cfg[CONFIG_SECTION_UI][KEY_SELECTED_DEP] = "Все"
                changed = True
            if not cfg.has_section(CONFIG_SECTION_INTEGR):
                cfg[CONFIG_SECTION_INTEGR] = {}
                changed = True
            if KEY_EXPORT_PWD not in cfg[CONFIG_SECTION_INTEGR]:
                cfg[CONFIG_SECTION_INTEGR][KEY_EXPORT_PWD] = "2025"
                changed = True
            if changed:
                with open(cp, "w", encoding="utf-8") as f:
                    cfg.write(f)
            return

        cfg = configparser.ConfigParser()
        cfg[CONFIG_SECTION_PATHS] = {
            KEY_OUTPUT_DIR: str(exe_dir() / OUTPUT_DIR_DEFAULT),
        }
        cfg[CONFIG_SECTION_UI] = {KEY_SELECTED_DEP: "Все"}
        cfg[CONFIG_SECTION_INTEGR] = {KEY_EXPORT_PWD: "2025"}
        with open(cp, "w", encoding="utf-8") as f:
            cfg.write(f)

    def read_config() -> configparser.ConfigParser:
        ensure_config()
        cfg = configparser.ConfigParser()
        cfg.read(config_path(), encoding="utf-8")
        return cfg

    def write_config(cfg: configparser.ConfigParser):
        with open(config_path(), "w", encoding="utf-8") as f:
            cfg.write(f)

    def get_output_dir_from_config() -> Path:
        cfg = read_config()
        raw = cfg.get(CONFIG_SECTION_PATHS, KEY_OUTPUT_DIR, fallback=str(exe_dir() / OUTPUT_DIR_DEFAULT))
        return Path(os.path.expandvars(raw))

    def get_export_password_from_config() -> str:
        cfg = read_config()
        return cfg.get(CONFIG_SECTION_INTEGR, KEY_EXPORT_PWD, fallback="2025")

    def get_selected_department_from_config() -> str:
        cfg = read_config()
        return cfg.get(CONFIG_SECTION_UI, KEY_SELECTED_DEP, fallback="Все")

    def set_selected_department_in_config(dep: str):
        cfg = read_config()
        if not cfg.has_section(CONFIG_SECTION_UI):
            cfg[CONFIG_SECTION_UI] = {}
        cfg[CONFIG_SECTION_UI][KEY_SELECTED_DEP] = dep or "Все"
        write_config(cfg)


def embedded_logo_image(parent, max_w=360, max_h=160):
    """
    Источники по приоритету:
    1) _LOGO_BASE64 из assets_logo.py (если есть)
    2) RAW-скачивание из GitHub
    3) tiny PNG
    """
    b64 = _LOGO_BASE64

    if not b64:
        try:
            data = urllib.request.urlopen(RAW_LOGO_URL, timeout=5).read()
            b64 = base64.b64encode(data).decode("ascii")
        except Exception:
            b64 = TINY_PNG_BASE64

    if Image and ImageTk:
        try:
            raw = base64.b64decode(b64.strip())
            im = Image.open(BytesIO(raw))
            im.thumbnail((max_w, max_h), Image.LANCZOS)
            return ImageTk.PhotoImage(im, master=parent)
        except Exception:
            pass

    try:
        ph = tk.PhotoImage(data=b64.strip(), master=parent)
        w, h = ph.width(), ph.height()
        k = max(w / max_w, h / max_h, 1)
        if k > 1:
            k = max(1, int(k))
            ph = ph.subsample(k, k)
        return ph
    except Exception:
        return None


# ================= БД: подключение и пользователи =================

# ================= БД: подключение и пользователи =================

# Глобальная переменная для хранения пула соединений
db_connection_pool = None

def initialize_db_pool():
    """
    Создает пул соединений с БД. Вызывается один раз при старте приложения.
    """
    global db_connection_pool
    if db_connection_pool:
        logging.info("Пул соединений уже был инициализирован.")
        return

    if not Settings:
        logging.error("settings_manager не доступен, не могу инициализировать пул БД.")
        raise RuntimeError("settings_manager не доступен, не могу прочитать параметры БД")

    try:
        provider = Settings.get_db_provider().strip().lower()
        if provider != "postgres":
            raise RuntimeError(f"Ожидался provider=postgres, а в настройках: {provider!r}")

        db_url = Settings.get_database_url().strip()
        if not db_url:
            raise RuntimeError("В настройках не указана строка подключения (DATABASE_URL)")

        url = urlparse(db_url)
        if url.scheme not in ("postgresql", "postgres"):
            raise RuntimeError(f"Неверная схема в DATABASE_URL: {url.scheme}")

        user = url.username
        password = url.password
        host = url.hostname or "localhost"
        port = url.port or 5432
        dbname = url.path.lstrip("/")
        q = parse_qs(url.query)
        sslmode = (q.get("sslmode", [Settings.get_db_sslmode()])[0] or "require")

        logging.debug("Инициализация пула соединений с БД...")
        # Создаем пул. minconn=1 (минимум 1 соединение держать открытым), maxconn=10 (до 10 одновременных)
        db_connection_pool = pool.SimpleConnectionPool(
            minconn=1,
            maxconn=10,
            host=host,
            port=port,
            dbname=dbname,
            user=user,
            password=password,
            sslmode=sslmode,
        )
        logging.info("Пул соединений с БД успешно инициализирован.")

    except Exception as e:
        logging.exception("Критическая ошибка: не удалось создать пул соединений с БД.")
        # Здесь можно либо показать messagebox и завершить приложение, либо просто выбросить исключение
        db_connection_pool = None
        raise e

def get_db_connection():
    """
    Получает соединение из пула.
    ВАЖНО: соединение нужно обязательно вернуть в пул с помощью putconn().
    """
    if db_connection_pool is None:
        # Попытка ленивой инициализации, если пул еще не создан
        initialize_db_pool()
        if db_connection_pool is None: # Если и после этого не создан, то беда
             raise RuntimeError("Пул соединений с БД не доступен.")
    return db_connection_pool.getconn()

def release_db_connection(conn):
    """
    Возвращает соединение обратно в пул.
    """
    if db_connection_pool:
        db_connection_pool.putconn(conn)

def close_db_pool():
    """
    Закрывает все соединения в пуле. Вызывается при выходе из приложения.
    """
    global db_connection_pool
    if db_connection_pool:
        logging.info("Закрытие пула соединений с БД...")
        db_connection_pool.closeall()
        db_connection_pool = None
        logging.info("Пул соединений закрыт.")

def _hash_password(password: str, salt: Optional[bytes] = None) -> str:
    if salt is None:
        salt = _os.urandom(16)
    iterations = 260000
    dk = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, iterations)
    return f"pbkdf2_sha256${iterations}${salt.hex()}${dk.hex()}"


def _verify_password(password: str, stored_hash: str) -> bool:
    try:
        if stored_hash.startswith("pbkdf2_sha256$"):
            _, it_str, salt_hex, hash_hex = stored_hash.split("$", 3)
            iterations = int(it_str)
            salt = bytes.fromhex(salt_hex)
            dk = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, iterations)
            return dk.hex() == hash_hex
        else:
            return password == stored_hash
    except Exception:
        return False


def authenticate_user(username: str, password: str) -> Optional[Dict[str, Any]]:
    """
    Проверяет логин/пароль в таблице app_users.
    При успехе возвращает dict с данными пользователя (без password_hash), иначе None.
    """
    logging.debug(f"authenticate_user: пытаемся авторизовать {username!r}")
    conn = None
    try:
        conn = get_db_connection() # <-- Получаем соединение из пула
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT id,
                       username,
                       password_hash,
                       is_active,
                       full_name,
                       role
                FROM app_users
                WHERE username = %s
                """,
                (username,),
            )
            row = cur.fetchone()
            if not row:
                return None
            if not row["is_active"]:
                return None
            if not _verify_password(password, row["password_hash"]):
                return None
            row.pop("password_hash", None)
            return dict(row)
    finally:
        if conn:
            release_db_connection(conn) # <-- ОБЯЗАТЕЛЬНО возвращаем 

def find_object_db_id_by_excel_or_address(
    excel_id: Optional[str],
    address: str,
) -> Optional[int]:
    """
    Ищет объект в таблице objects:
      - сначала по excel_id (если задан),
      - если не найдено — по точному адресу.
    Возвращает id объекта или None.
    """
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            if excel_id:
                cur.execute(
                    """
                    SELECT id
                      FROM objects
                     WHERE COALESCE(NULLIF(excel_id, ''), '') = %s
                    """,
                    (excel_id,),
                )
                row = cur.fetchone()
                if row:
                    return row[0]

            # По адресу (если по excel_id не нашли / не задан)
            cur.execute(
                """
                SELECT id
                  FROM objects
                 WHERE address = %s
                """,
                (address,),
            )
            row = cur.fetchone()
            return row[0] if row else None
    finally:
        if conn:
            release_db_connection(conn)

# ---------- Справочники из БД ----------

def load_employees_from_db() -> List[Tuple[str, str, str, str]]:
    """
    Возвращает список сотрудников:
      [(fio, tbn, position, department), ...]
    """
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT e.fio, e.tbn, e.position, d.name AS dep
                  FROM employees e
                  LEFT JOIN departments d ON d.id = e.department_id
                 WHERE COALESCE(e.is_fired, FALSE) = FALSE
              ORDER BY e.fio
                """
            )
            rows = cur.fetchall()
            return [(fio or "", tbn or "", pos or "", dep or "") for fio, tbn, pos, dep in rows]
    finally:
        if conn:
            release_db_connection(conn)


def load_objects_from_db() -> List[Tuple[str, str]]:
    """
    Возвращает список объектов [(code, address)].
    code = excel_id (если есть) или пустая строка.
    """
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT
                    COALESCE(NULLIF(excel_id, ''), '') AS code,
                    address
                  FROM objects
                 ORDER BY address
                """
            )
            rows = cur.fetchall()
            return [(code or "", addr or "") for code, addr in rows]
    finally:
        if conn:
            release_db_connection(conn)

def upsert_timesheet_header(
    object_id: str,
    object_addr: str,
    department: str,
    year: int,
    month: int,
    user_id: int,
) -> int:
    """
    Находит или создаёт заголовок табеля и возвращает его id.
    Привязан к конкретному user_id.
    object_db_id (FK на objects.id) берётся из таблицы objects.
    """
    # Находим объект в БД
    object_db_id = find_object_db_id_by_excel_or_address(
        object_id or None,
        object_addr,
    )
    if object_db_id is None:
        # Можно вернуть None, но лучше сразу понятная ошибка
        raise RuntimeError(
            f"В БД не найден объект (excel_id={object_id!r}, address={object_addr!r}). "
            f"Сначала создайте объект в разделе «Объекты»."
        )

    conn = None
    try:
        conn = get_db_connection()
        with conn, conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO timesheet_headers (
                    object_id,
                    object_addr,
                    department,
                    year,
                    month,
                    user_id,
                    object_db_id
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (object_id, object_addr, department, year, month, user_id)
                DO UPDATE SET
                    updated_at = now(),
                    object_db_id = EXCLUDED.object_db_id
                RETURNING id;
                """,
                (
                    object_id or None,
                    object_addr,
                    department or None,
                    year,
                    month,
                    user_id,
                    object_db_id,
                ),
            )
            ts_id = cur.fetchone()[0]
        return ts_id
    finally:
        if conn:
            release_db_connection(conn)

def replace_timesheet_rows(
    header_id: int,
    rows: List[Dict[str, Any]],
):
    """
    Полностью заменяет строки табеля для заданного header_id.
    rows: [{'fio': str, 'tbn': str, 'hours': [str|None x31]}, ...]
    """
    conn = None
    try:
        conn = get_db_connection()
        with conn, conn.cursor() as cur:
            # Удаляем старые строки
            cur.execute("DELETE FROM timesheet_rows WHERE header_id = %s", (header_id,))

            # Вставляем новые
            for rec in rows:
                fio = rec["fio"]
                tbn = rec.get("tbn") or None
                hours_list = rec.get("hours") or [None] * 31
                if len(hours_list) != 31:
                    hours_list = (hours_list + [None] * 31)[:31]

                # Подсчёты такие же, как в save_all
                total_hours = 0.0
                total_days = 0
                total_ot_day = 0.0
                total_ot_night = 0.0

                for raw in hours_list:
                    if not raw:
                        continue
                    hrs = parse_hours_value(raw)
                    d_ot, n_ot = parse_overtime(raw)
                    if isinstance(hrs, (int, float)) and hrs > 1e-12:
                        total_hours += hrs
                        total_days += 1
                    if isinstance(d_ot, (int, float)):
                        total_ot_day += float(d_ot)
                    if isinstance(n_ot, (int, float)):
                        total_ot_night += float(n_ot)

                cur.execute(
                    """
                    INSERT INTO timesheet_rows (
                        header_id, fio, tbn, hours_raw,
                        total_days, total_hours, overtime_day, overtime_night
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        header_id,
                        fio,
                        tbn,
                        hours_list,
                        total_days if total_days else None,
                        None if abs(total_hours) < 1e-12 else total_hours,
                        None if abs(total_ot_day) < 1e-12 else total_ot_day,
                        None if abs(total_ot_night) < 1e-12 else total_ot_night,
                    ),
                )
    finally:
        if conn:
            release_db_connection(conn)

def load_timesheet_rows_from_db(
    object_id: str,
    object_addr: str,
    department: str,
    year: int,
    month: int,
    user_id: int,
) -> List[Dict[str, Any]]:
    """
    Загружает строки табеля для конкретного пользователя и контекста
    (объект, подразделение, период).
    Возвращает список dict: {'fio', 'tbn', 'hours': [str|None x31]}
    """
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT h.id
                  FROM timesheet_headers h
                 WHERE COALESCE(h.object_id, '') = COALESCE(%s, '')
                   AND h.object_addr = %s
                   AND COALESCE(h.department, '') = COALESCE(%s, '')
                   AND h.year = %s
                   AND h.month = %s
                   AND h.user_id = %s
                """,
                (object_id or None, object_addr, department or None, year, month, user_id),
            )
            row = cur.fetchone()
            if not row:
                return []
            header_id = row[0]

            cur.execute(
                """
                SELECT fio, tbn, hours_raw
                  FROM timesheet_rows
                 WHERE header_id = %s
                 ORDER BY fio, tbn
                """,
                (header_id,),
            )
            result = []
            for fio, tbn, hours_raw in cur.fetchall():
                # hours_raw приходит как list/tuple из psycopg2
                hrs = list(hours_raw) if hours_raw is not None else [None] * 31
                if len(hrs) != 31:
                    hrs = (hrs + [None] * 31)[:31]
                result.append({
                    "fio": fio or "",
                    "tbn": tbn or "",
                    "hours": [h if h is not None else None for h in hrs],
                })
            return result
    finally:
        if conn:
            release_db_connection(conn)

def load_timesheet_rows_by_header_id(header_id: int) -> List[Dict[str, Any]]:
    """
    Загружает строки табеля по ID заголовка (timesheet_headers.id).
    Возвращает: список dict:
      {
        'fio': str,
        'tbn': str,
        'hours_raw': [str|None x31],
        'total_days': int|None,
        'total_hours': float|None,
        'overtime_day': float|None,
        'overtime_night': float|None,
      }
    """
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT fio, tbn, hours_raw,
                       total_days, total_hours,
                       overtime_day, overtime_night
                  FROM timesheet_rows
                 WHERE header_id = %s
                 ORDER BY fio, tbn
                """,
                (header_id,),
            )
            result: List[Dict[str, Any]] = []
            for fio, tbn, hours_raw, total_days, total_hours, ot_day, ot_night in cur.fetchall():
                hrs = list(hours_raw) if hours_raw is not None else [None] * 31
                if len(hrs) != 31:
                    hrs = (hrs + [None] * 31)[:31]
                result.append({
                    "fio": fio or "",
                    "tbn": tbn or "",
                    "hours_raw": [h if h is not None else None for h in hrs],
                    "total_days": total_days,
                    "total_hours": float(total_hours) if total_hours is not None else None,
                    "overtime_day": float(ot_day) if ot_day is not None else None,
                    "overtime_night": float(ot_night) if ot_night is not None else None,
                })
            return result
    finally:
        if conn:
            release_db_connection(conn)

def load_user_timesheet_headers(user_id: int) -> List[Dict[str, Any]]:
    """
    Возвращает список заголовков табелей, созданных пользователем.
    Каждый элемент:
      {
        'id': int,
        'object_id': str|None,
        'object_addr': str,
        'department': str|None,
        'year': int,
        'month': int,
        'created_at': datetime,
        'updated_at': datetime
      }
    """
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT id,
                       object_id,
                       object_addr,
                       department,
                       year,
                       month,
                       created_at,
                       updated_at
                  FROM timesheet_headers
                 WHERE user_id = %s
                 ORDER BY year DESC, month DESC, object_addr, COALESCE(department, '')
                """,
                (user_id,),
            )
            rows = cur.fetchall()
            return [dict(r) for r in rows]
    finally:
        if conn:
            release_db_connection(conn)

def load_all_timesheet_headers(
    year: Optional[int] = None,
    month: Optional[int] = None,
    department: Optional[str] = None,
    object_addr_substr: Optional[str] = None,
    object_id_substr: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """
    Возвращает список заголовков табелей всех пользователей с фильтрами.
    Фильтры:
      - year, month — если заданы, фильтруем по периоду;
      - department — если не пустая строка, фильтруем по точному совпадению;
      - object_addr_substr — подстрока в адресе объекта (ILIKE %...%);
      - object_id_substr — подстрока в ID объекта (ILIKE %...%).

    Каждый результат:
      {
        'id', 'object_id', 'object_addr', 'department',
        'year', 'month', 'user_id', 'username', 'full_name',
        'created_at', 'updated_at'
      }
    """
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            where = ["1=1"]
            params: List[Any] = []

            if year is not None:
                where.append("h.year = %s")
                params.append(year)
            if month is not None:
                where.append("h.month = %s")
                params.append(month)
            if department:
                where.append("COALESCE(h.department, '') = %s")
                params.append(department)

            if object_addr_substr:
                where.append("h.object_addr ILIKE %s")
                params.append(f"%{object_addr_substr}%")
            if object_id_substr:
                where.append("COALESCE(h.object_id, '') ILIKE %s")
                params.append(f"%{object_id_substr}%")

            where_sql = " AND ".join(where)

            cur.execute(
                f"""
                SELECT h.id,
                       h.object_id,
                       h.object_addr,
                       h.department,
                       h.year,
                       h.month,
                       h.user_id,
                       u.username,
                       u.full_name,
                       h.created_at,
                       h.updated_at
                  FROM timesheet_headers h
                  JOIN app_users u ON u.id = h.user_id
                 WHERE {where_sql}
                 ORDER BY h.year DESC, h.month DESC,
                          h.object_addr,
                          COALESCE(h.department, ''),
                          u.full_name
                """,
                params,
            )
            rows = cur.fetchall()
            return [dict(r) for r in rows]
    finally:
        if conn:
            release_db_connection(conn)
# ------------- Утилиты для времени и табеля -------------

def month_days(year: int, month: int) -> int:
    return calendar.monthrange(year, month)[1]


def month_name_ru(month: int) -> str:
    names = [
        "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
        "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"
    ]
    return names[month - 1]


def parse_hours_value(v: Any) -> Optional[float]:
    s = str(v or "").strip()
    if not s:
        return None

    if "(" in s:
        s = s.split("(")[0].strip()

    if "/" in s:
        total = 0.0
        any_part = False
        for part in s.split("/"):
            n = parse_hours_value(part)
            if isinstance(n, (int, float)):
                total += float(n)
                any_part = True
        return total if any_part else None

    if ":" in s:
        p = s.split(":")
        try:
            hh = float(p[0].replace(",", "."))
            mm = float((p[1] if len(p) > 1 else "0").replace(",", "."))
            ss = float((p[2] if len(p) > 2 else "0").replace(",", "."))
            return hh + mm / 60.0 + ss / 3600.0
        except Exception:
            pass

    s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def parse_overtime(v: Any) -> Tuple[Optional[float], Optional[float]]:
    s = str(v or "").strip()
    if "(" not in s or ")" not in s:
        return None, None
    try:
        start = s.index("(")
        end = s.index(")")
        overtime_str = s[start + 1:end].strip()
        if "/" in overtime_str:
            parts = overtime_str.split("/")
            day_ot = float(parts[0].replace(",", ".")) if parts[0].strip() else 0.0
            night_ot = float(parts[1].replace(",", ".")) if len(parts) > 1 and parts[1].strip() else 0.0
            return day_ot, night_ot
        else:
            ot = float(overtime_str.replace(",", "."))
            return ot, 0.0
    except Exception:
        return None, None


def safe_filename(s: str, maxlen: int = 60) -> str:
    if not s:
        return "NOID"
    s = re.sub(r'[<>:"/\\|?*\n\r\t]+', "_", str(s)).strip()
    s = re.sub(r"_+", "_", s)
    return s[:maxlen] if len(s) > maxlen else s


# ------------- Логотип / домашняя страница -------------

def find_logo_path() -> Optional[Path]:
    candidates = [
        exe_dir() / "assets" / "logo.png",
        exe_dir() / "assets" / "logo.gif",
        exe_dir() / "assets" / "logo.jpg",
        exe_dir() / "logo.png",
        exe_dir() / "logo.gif",
        exe_dir() / "logo.jpg",
    ]
    for p in candidates:
        if p.exists():
            return p
    return None


class HomePage(tk.Frame):
    def __init__(self, master):
        super().__init__(master, bg="#f7f7f7")

        outer = tk.Frame(self, bg="#f7f7f7")
        outer.pack(fill="both", expand=True)

        center = tk.Frame(outer, bg="#f7f7f7")
        center.place(relx=0.5, rely=0.5, anchor="center")

        self.logo_img = embedded_logo_image(center, max_w=360, max_h=360)
        if self.logo_img:
            tk.Label(center, image=self.logo_img, bg="#f7f7f7").pack(anchor="center", pady=(0, 12))

        tk.Label(
            center,
            text="Добро пожаловать!",
            font=("Segoe UI", 18, "bold"),
            bg="#f7f7f7",
        ).pack(anchor="center", pady=(4, 6))
        tk.Label(
            center,
            text="Выберите раздел в верхнем меню.",
            font=("Segoe UI", 10),
            fg="#444",
            bg="#f7f7f7",
            justify="center",
        ).pack(anchor="center")


class LoginPage(tk.Frame):
    def __init__(self, master, app_ref: "MainApp"):
        super().__init__(master, bg="#f7f7f7")
        self.app_ref = app_ref

        outer = tk.Frame(self, bg="#f7f7f7")
        outer.pack(fill="both", expand=True)

        center = tk.Frame(outer, bg="#f7f7f7")
        center.place(relx=0.5, rely=0.5, anchor="center")

        tk.Label(
            center,
            text="Управление строительством",
            font=("Segoe UI", 16, "bold"),
            bg="#f7f7f7",
        ).grid(row=0, column=0, columnspan=2, pady=(0, 10))

        tk.Label(
            center,
            text="Вход в систему",
            font=("Segoe UI", 11),
            fg="#555",
            bg="#f7f7f7",
        ).grid(row=1, column=0, columnspan=2, pady=(0, 15))

        tk.Label(center, text="Логин:", bg="#f7f7f7").grid(row=2, column=0, sticky="e", padx=(0, 6), pady=4)
        tk.Label(center, text="Пароль:", bg="#f7f7f7").grid(row=3, column=0, sticky="e", padx=(0, 6), pady=4)

        self.ent_login = ttk.Entry(center, width=26)
        self.ent_login.grid(row=2, column=1, sticky="w", pady=4)
        self.ent_pass = ttk.Entry(center, width=26, show="*")
        self.ent_pass.grid(row=3, column=1, sticky="w", pady=4)

        btns = tk.Frame(center, bg="#f7f7f7")
        btns.grid(row=4, column=0, columnspan=2, pady=(12, 0), sticky="e")

        ttk.Button(btns, text="Войти", width=12, command=self._on_login).pack(side="left", padx=5)
        ttk.Button(btns, text="Выход", width=10, command=self._on_exit).pack(side="left", padx=5)

        self.ent_login.focus_set()
        self.bind_all("<Return>", self._on_enter)

    def _on_enter(self, event):
        if self.winfo_ismapped():
            self._on_login()

    def _on_login(self):
        username = self.ent_login.get().strip()
        password = self.ent_pass.get().strip()
        if not username or not password:
            messagebox.showwarning("Вход", "Укажите логин и пароль.", parent=self)
            return
        try:
            logging.debug(f"LoginPage: пробуем авторизовать {username!r}")
            user = authenticate_user(username, password)
        except Exception as e:
            logging.exception("Ошибка при обращении к БД в authenticate_user")
            messagebox.showerror("Вход", f"Ошибка при обращении к БД:\n{e}", parent=self)
            return

        if not user:
            messagebox.showerror("Вход", "Неверный логин или пароль.", parent=self)
            return

        self.app_ref.on_login_success(user)

    def _on_exit(self):
        self.app_ref.destroy()

# ================= ДОПОЛНИТЕЛЬНЫЕ ВИДЖЕТЫ И ДИАЛОГИ ДЛЯ ТАБЕЛЕЙ =================

class CopyFromDialog(simpledialog.Dialog):
    def __init__(self, parent, init_year: int, init_month: int):
        self.init_year = init_year
        self.init_month = init_month
        self.result = None
        super().__init__(parent, title="Копировать сотрудников из месяца")

    def body(self, master):
        tk.Label(master, text="Источник").grid(row=0, column=0, sticky="w", pady=(2, 6), columnspan=4)

        tk.Label(master, text="Месяц:").grid(row=1, column=0, sticky="e")
        self.cmb_month = ttk.Combobox(master, state="readonly", width=18,
                                      values=[month_name_ru(i) for i in range(1, 13)])
        self.cmb_month.grid(row=1, column=1, sticky="w")
        self.cmb_month.current(max(0, min(11, self.init_month - 1)))

        tk.Label(master, text="Год:").grid(row=1, column=2, sticky="e", padx=(10, 4))
        self.spn_year = tk.Spinbox(master, from_=2000, to=2100, width=6)
        self.spn_year.grid(row=1, column=3, sticky="w")
        self.spn_year.delete(0, "end")
        self.spn_year.insert(0, str(self.init_year))

        self.var_copy_hours = tk.BooleanVar(value=False)
        ttk.Checkbutton(master, text="Копировать часы", variable=self.var_copy_hours)\
            .grid(row=2, column=1, sticky="w", pady=(8, 2))

        tk.Label(master, text="Режим:").grid(row=3, column=0, sticky="e", pady=(6, 2))
        self.var_mode = tk.StringVar(value="replace")
        frame_mode = tk.Frame(master)
        frame_mode.grid(row=3, column=1, columnspan=3, sticky="w", pady=(6, 2))
        ttk.Radiobutton(frame_mode, text="Заменить текущий список",
                        value="replace", variable=self.var_mode).pack(anchor="w")
        ttk.Radiobutton(frame_mode, text="Объединить (добавить недостающих)",
                        value="merge", variable=self.var_mode).pack(anchor="w")
        return self.cmb_month

    def validate(self):
        try:
            y = int(self.spn_year.get())
            if not (2000 <= y <= 2100):
                raise ValueError
            return True
        except Exception:
            messagebox.showwarning("Копирование", "Введите корректный год (2000–2100).")
            return False

    def apply(self):
        self.result = {
            "year": int(self.spn_year.get()),
            "month": self.cmb_month.current() + 1,
            "with_hours": bool(self.var_copy_hours.get()),
            "mode": self.var_mode.get(),
        }


class BatchAddDialog(tk.Toplevel):
    def __init__(self, parent, total: int, title: str = "Добавление сотрудников"):
        super().__init__(parent)
        self.parent = parent
        self.total = max(1, int(total))
        self.done = 0
        self.cancelled = False
        self.title(title)
        self.resizable(False, False)
        self.grab_set()

        frm = tk.Frame(self, padx=12, pady=12)
        frm.pack(fill="both", expand=True)

        self.lbl = tk.Label(frm, text=f"Добавлено: 0 из {self.total}")
        self.lbl.pack(fill="x")

        self.pb = ttk.Progressbar(frm, mode="determinate",
                                  maximum=self.total, length=420)
        self.pb.pack(fill="x", pady=(8, 8))

        self.btn_cancel = ttk.Button(frm, text="Отмена", command=self._on_cancel)
        self.btn_cancel.pack(anchor="e", pady=(6, 0))

        try:
            self.update_idletasks()
            px = parent.winfo_rootx()
            py = parent.winfo_rooty()
            pw = parent.winfo_width()
            ph = parent.winfo_height()
            sw = self.winfo_width()
            sh = self.winfo_height()
            self.geometry(f"+{px + (pw - sw)//2}+{py + (ph - sh)//2}")
        except Exception:
            pass

    def step(self, n: int = 1):
        if self.cancelled:
            return
        self.done += n
        if self.done > self.total:
            self.done = self.total
        self.pb["value"] = self.done
        self.lbl.config(text=f"Добавлено: {self.done} из {self.total}")
        self.update_idletasks()

    def _on_cancel(self):
        self.cancelled = True

    def close(self):
        try:
            self.grab_release()
        except Exception:
            pass
        self.destroy()


class HoursFillDialog(simpledialog.Dialog):
    def __init__(self, parent, max_day: int):
        self.max_day = max_day
        self.result = None
        super().__init__(parent, title="Проставить часы всем")

    def body(self, master):
        tk.Label(master, text=f"В текущем месяце дней: {self.max_day}")\
            .grid(row=0, column=0, columnspan=3, sticky="w", pady=(2, 6))
        tk.Label(master, text="День:").grid(row=1, column=0, sticky="e")
        self.spn_day = tk.Spinbox(master, from_=1, to=31, width=4)
        self.spn_day.grid(row=1, column=1, sticky="w")
        self.spn_day.delete(0, "end")
        self.spn_day.insert(0, "1")

        self.var_clear = tk.BooleanVar(value=False)
        ttk.Checkbutton(master, text="Очистить день (пусто)",
                        variable=self.var_clear,
                        command=self._on_toggle_clear)\
            .grid(row=2, column=1, sticky="w", pady=(6, 2))

        tk.Label(master, text="Часы:").grid(row=3, column=0, sticky="e", pady=(6, 0))
        self.ent_hours = ttk.Entry(master, width=12)
        self.ent_hours.grid(row=3, column=1, sticky="w", pady=(6, 0))
        self.ent_hours.insert(0, "8")

        tk.Label(master, text="Форматы: 8 | 8,25 | 8:30 | 1/7")\
            .grid(row=4, column=0, columnspan=3, sticky="w", pady=(6, 2))
        return self.spn_day

    def _on_toggle_clear(self):
        if self.var_clear.get():
            self.ent_hours.configure(state="disabled")
        else:
            self.ent_hours.configure(state="normal")

    def validate(self):
        try:
            d = int(self.spn_day.get())
            if not (1 <= d <= 31):
                raise ValueError
        except Exception:
            messagebox.showwarning(
                "Проставить часы",
                "День должен быть числом от 1 до 31.",
            )
            return False

        if self.var_clear.get():
            self._d = d
            self._h = 0.0
            self._clear = True
            return True

        hv = parse_hours_value(self.ent_hours.get().strip())
        if hv is None or hv < 0:
            messagebox.showwarning(
                "Проставить часы",
                "Введите корректное значение часов (например, 8, 8:30, 1/7).",
            )
            return False
        self._d = d
        self._h = float(hv)
        self._clear = False
        return True

    def apply(self):
        self.result = {
            "day": self._d,
            "hours": self._h,
            "clear": self._clear,
        }


class AutoCompleteCombobox(ttk.Combobox):
    def __init__(self, master=None, **kw):
        super().__init__(master, **kw)
        self._all_values: List[str] = []
        self.bind("<KeyRelease>", self._on_keyrelease)
        self.bind("<Control-BackSpace>", self._clear_all)

    def set_completion_list(self, values: List[str]):
        self._all_values = list(values)
        self["values"] = self._all_values

    def _clear_all(self, _=None):
        self.delete(0, tk.END)
        self["values"] = self._all_values

    def _on_keyrelease(self, event):
        if event.keysym in (
            "Up", "Down", "Left", "Right", "Home",
            "End", "Return", "Escape", "Tab"
        ):
            return
        typed = self.get().strip()
        if not typed:
            self["values"] = self._all_values
            return
        self["values"] = [x for x in self._all_values if typed.lower() in x.lower()]


class RowWidget:
    WEEK_BG_SAT = "#fff8e1"
    WEEK_BG_SUN = "#ffebee"
    ZEBRA_EVEN = "#ffffff"
    ZEBRA_ODD = "#f6f8fa"
    ERR_BG = "#ffccbc"
    DISABLED_BG = "#f0f0f0"

    def __init__(self, table: tk.Frame, row_index: int, fio: str, tbn: str,
                 get_year_month_callable, on_delete_callable):
        self.table = table
        self.row = row_index
        self.get_year_month = get_year_month_callable
        self.on_delete = on_delete_callable
        self._suspend_sync = False

        zebra_bg = self.ZEBRA_EVEN if (row_index % 2 == 0) else self.ZEBRA_ODD
        self.widgets: List[tk.Widget] = []

        # ФИО
        self.lbl_fio = tk.Label(
            self.table,
            text=fio,
            anchor="w",
            bg=zebra_bg,
            width=35,          # подбери по вкусу
        )
        self.lbl_fio.grid(row=self.row, column=0, padx=0, pady=1, sticky="nsew")
        self.widgets.append(self.lbl_fio)

        # Таб.№
        self.lbl_tbn = tk.Label(self.table, text=tbn, anchor="center", bg=zebra_bg)
        self.lbl_tbn.grid(row=self.row, column=1, padx=0, pady=1, sticky="nsew")
        self.widgets.append(self.lbl_tbn)

        # Дни месяца (col 2..32)
        self.day_entries: List[tk.Entry] = []
        for d in range(1, 32):
            e = tk.Entry(self.table, width=4, justify="center", relief="solid", bd=1)
            e.grid(row=self.row, column=1 + d, padx=0, pady=1, sticky="nsew")
            e.bind("<FocusOut>", lambda ev, _d=d: self.update_total())
            e.bind("<Button-2>", lambda ev: "break")
            e.bind("<ButtonRelease-2>", lambda ev: "break")
            self.day_entries.append(e)
            self.widgets.append(e)

        # Итоги
        self.lbl_days = tk.Label(self.table, text="0", anchor="e", bg=zebra_bg)
        self.lbl_days.grid(row=self.row, column=33, padx=(4, 1), pady=1, sticky="nsew")
        self.widgets.append(self.lbl_days)

        self.lbl_total = tk.Label(self.table, text="0", anchor="e", bg=zebra_bg)
        self.lbl_total.grid(row=self.row, column=34, padx=(4, 1), pady=1, sticky="nsew")
        self.widgets.append(self.lbl_total)

        self.lbl_overtime_day = tk.Label(self.table, text="0", anchor="e", bg=zebra_bg)
        self.lbl_overtime_day.grid(row=self.row, column=35, padx=(4, 1), pady=1, sticky="nsew")
        self.widgets.append(self.lbl_overtime_day)

        self.lbl_overtime_night = tk.Label(self.table, text="0", anchor="e", bg=zebra_bg)
        self.lbl_overtime_night.grid(row=self.row, column=36, padx=(4, 1), pady=1, sticky="nsew")
        self.widgets.append(self.lbl_overtime_night)

        # 5/2
        self.btn_52 = ttk.Button(self.table, text="5/2", width=4, command=self.fill_52)
        self.btn_52.grid(row=self.row, column=37, padx=1, pady=0, sticky="nsew")
        self.widgets.append(self.btn_52)

        # Удалить
        self.btn_del = ttk.Button(self.table, text="Удалить", width=7, command=self.delete_row)
        self.btn_del.grid(row=self.row, column=38, padx=1, pady=0, sticky="nsew")
        self.widgets.append(self.btn_del)

    def set_day_font(self, font_tuple):
        for e in self.day_entries:
            e.configure(font=font_tuple)

    def regrid_to(self, new_row: int):
        self.row = new_row
        self.lbl_fio.grid_configure(row=new_row, column=0)
        self.lbl_tbn.grid_configure(row=new_row, column=1)
        for i, e in enumerate(self.day_entries, start=2):
            e.grid_configure(row=new_row, column=i)
        self.lbl_days.grid_configure(row=new_row, column=33)
        self.lbl_total.grid_configure(row=new_row, column=34)
        self.lbl_overtime_day.grid_configure(row=new_row, column=35)
        self.lbl_overtime_night.grid_configure(row=new_row, column=36)
        self.btn_52.grid_configure(row=new_row, column=37)
        self.btn_del.grid_configure(row=new_row, column=38)

    def fio(self) -> str:
        return self.lbl_fio.cget("text")

    def tbn(self) -> str:
        return self.lbl_tbn.cget("text")

    def set_hours(self, arr: List[Optional[str]]):
        days = len(arr)
        for i in range(31):
            self.day_entries[i].delete(0, "end")
            if i < days and arr[i]:
                self.day_entries[i].insert(0, str(arr[i]))
        self.update_total()

    def get_hours_with_overtime(self) -> List[Tuple[Optional[float], Optional[float], Optional[float]]]:
        result = []
        for e in self.day_entries:
            raw = e.get().strip()
            hours = parse_hours_value(raw) if raw else None
            day_ot, night_ot = parse_overtime(raw) if raw else (None, None)
            result.append((hours, day_ot, night_ot))
        return result

    def _bg_for_day(self, year: int, month: int, day: int) -> str:
        from datetime import datetime as _dt
        wd = _dt(year, month, day).weekday()
        if wd == 5:
            return self.WEEK_BG_SAT
        if wd == 6:
            return self.WEEK_BG_SUN
        return "white"

    def _repaint_day_cell(self, i0: int, year: int, month: int):
        from datetime import datetime as _dt
        day = i0 + 1
        e = self.day_entries[i0]
        days = month_days(year, month)

        if day > days:
            e.configure(state="disabled", disabledbackground=self.DISABLED_BG)
            e.delete(0, "end")
            return

        e.configure(state="normal")
        raw = e.get().strip()

        invalid = False
        if raw:
            val = parse_hours_value(raw)
            if val is None or val < 0 or val > 24:
                invalid = True
            if "(" in raw:
                day_ot, night_ot = parse_overtime(raw)
                if day_ot is None and night_ot is None:
                    invalid = True

        if invalid:
            e.configure(bg=self.ERR_BG)
        else:
            e.configure(bg=self._bg_for_day(year, month, day))

    def update_days_enabled(self, year: int, month: int):
        for i in range(31):
            self._repaint_day_cell(i, year, month)
        self.update_total()

    def update_total(self):
        total_hours = 0.0
        total_days = 0
        total_overtime_day = 0.0
        total_overtime_night = 0.0

        y, m = self.get_year_month()
        days_in_m = month_days(y, m)

        for i, e in enumerate(self.day_entries, start=1):
            raw = e.get().strip()
            self._repaint_day_cell(i - 1, y, m)
            if i <= days_in_m and raw:
                hours = parse_hours_value(raw)
                day_ot, night_ot = parse_overtime(raw)
                if isinstance(hours, (int, float)) and hours > 1e-12:
                    total_hours += float(hours)
                    total_days += 1
                if isinstance(day_ot, (int, float)):
                    total_overtime_day += float(day_ot)
                if isinstance(night_ot, (int, float)):
                    total_overtime_night += float(night_ot)

        self.lbl_days.config(text=str(total_days))
        sh = f"{total_hours:.2f}".rstrip("0").rstrip(".")
        self.lbl_total.config(text=sh)
        sod = f"{total_overtime_day:.2f}".rstrip("0").rstrip(".")
        self.lbl_overtime_day.config(text=sod)
        son = f"{total_overtime_night:.2f}".rstrip("0").rstrip(".")
        self.lbl_overtime_night.config(text=son)

    def fill_52(self):
        y, m = self.get_year_month()
        days = month_days(y, m)
        for d in range(1, days + 1):
            wd = datetime(y, m, d).weekday()
            e = self.day_entries[d - 1]
            e.delete(0, "end")
            if wd < 4:
                e.insert(0, "8,25")
            elif wd == 4:
                e.insert(0, "7")
        for d in range(days + 1, 32):
            self.day_entries[d - 1].delete(0, "end")
        self.update_total()

    def delete_row(self):
        self.on_delete(self)
        
# ================= СТРАНИЦА ТАБЕЛЕЙ (ИСПОЛЬЗУЕТ БАЗУ) =================

class TimesheetPage(tk.Frame):
    COLPX = {"fio": 200, "tbn": 100, "day": 36, "days": 46, "hours": 56, "btn52": 40, "del": 66}
    MIN_FIO_PX = 140
    MAX_FIO_PX = 260
    HEADER_BG = "#d0d0d0"

    def __init__(
        self,
        master,
        app_ref,
        init_object_id: Optional[str] = None,
        init_object_addr: Optional[str] = None,
        init_department: Optional[str] = None,
        init_year: Optional[int] = None,
        init_month: Optional[int] = None,
        read_only: bool = False,
        owner_user_id: Optional[int] = None,
    ):
        super().__init__(master)
        self.app_ref = app_ref  # ссылка на MainApp, чтобы брать current_user
        self.read_only = bool(read_only)
        # ВЛАДЕЛЕЦ ТАБЕЛЯ (для загрузки строк)
        self.owner_user_id: Optional[int] = owner_user_id
        # Параметры инициализации (могут быть None)
        self._init_object_id = init_object_id
        self._init_object_addr = init_object_addr
        self._init_department = init_department
        self._init_year = init_year
        self._init_month = init_month

        self.base_dir = exe_dir()
        self.out_dir = get_output_dir_from_config()
        self.out_dir.mkdir(parents=True, exist_ok=True)

        self.DAY_ENTRY_FONT = ("Segoe UI", 8)
        self._fit_job = None

        self._load_spr_data_from_db()

        self.model_rows: List[Dict[str, Any]] = []
        self.current_page = 1
        self.page_size = tk.IntVar(value=50)
        self._suspend_sync = False

        self._build_ui()
        self._render_page(1)
        self._load_existing_rows()

        self.bind("<Configure>", self._on_window_configure)
        self.after(120, self._auto_fit_columns)

    def _load_spr_data_from_db(self):
        employees = load_employees_from_db()
        objects = load_objects_from_db()

        self.employees = employees
        self.objects = objects

        self.emp_names = [fio for (fio, _, _, _) in self.employees]
        self.emp_info = {fio: (tbn, pos) for (fio, tbn, pos, _) in self.employees}

        deps = sorted({(dep or "").strip() for (_, _, _, dep) in self.employees if (dep or "").strip()})
        self.departments = ["Все"] + deps

        self.addr_to_ids: Dict[str, List[str]] = {}
        for oid, addr in self.objects:
            if not addr:
                continue
            self.addr_to_ids.setdefault(addr, [])
            if oid and oid not in self.addr_to_ids[addr]:
                self.addr_to_ids[addr].append(oid)
        addresses_set = set(self.addr_to_ids.keys()) | {addr for _, addr in self.objects if addr}
        self.address_options = sorted(addresses_set)

    def _build_ui(self):
        # Верхняя панель
        top = tk.Frame(self)
        top.pack(fill="x", padx=8, pady=8)

        tk.Label(top, text="Подразделение:").grid(row=0, column=0, sticky="w")
        deps = self.departments or ["Все"]
        self.cmb_department = ttk.Combobox(top, state="readonly", values=deps, width=48)
        self.cmb_department.grid(row=0, column=1, sticky="w", padx=(4, 12))
        try:
            saved_dep = get_selected_department_from_config()
            self.cmb_department.set(saved_dep if saved_dep in deps else deps[0])
        except Exception:
            self.cmb_department.set(deps[0])
        self.cmb_department.bind("<<ComboboxSelected>>", lambda e: self._on_department_select())

        tk.Label(top, text="Месяц:").grid(row=1, column=0, sticky="w", padx=(0, 4), pady=(8, 0))
        self.cmb_month = ttk.Combobox(top, state="readonly", width=12, values=[month_name_ru(i) for i in range(1, 13)])
        self.cmb_month.grid(row=1, column=1, sticky="w", pady=(8, 0))
        self.cmb_month.current(datetime.now().month - 1)
        self.cmb_month.bind("<<ComboboxSelected>>", lambda e: self._on_period_change())

        tk.Label(top, text="Год:").grid(row=1, column=2, sticky="w", padx=(16, 4), pady=(8, 0))
        self.spn_year = tk.Spinbox(top, from_=2000, to=2100, width=6, command=self._on_period_change)
        self.spn_year.grid(row=1, column=3, sticky="w", pady=(8, 0))
        self.spn_year.delete(0, "end")
        self.spn_year.insert(0, datetime.now().year)
        self.spn_year.bind("<FocusOut>", lambda e: self._on_period_change())

        tk.Label(top, text="Адрес:").grid(row=1, column=4, sticky="w", padx=(20, 4), pady=(8, 0))
        self.cmb_address = AutoCompleteCombobox(top, width=46)
        self.cmb_address.set_completion_list(self.address_options)
        self.cmb_address.grid(row=1, column=5, sticky="w", pady=(8, 0))
        self.cmb_address.bind("<<ComboboxSelected>>", self._on_address_select)
        self.cmb_address.bind("<FocusOut>", self._on_address_select)
        self.cmb_address.bind("<Return>", lambda e: self._on_address_select())
        self.cmb_address.bind("<KeyRelease>", lambda e: self._on_address_change(), add="+")

        tk.Label(top, text="ID объекта:").grid(row=1, column=6, sticky="w", padx=(16, 4), pady=(8, 0))
        self.cmb_object_id = ttk.Combobox(top, state="readonly", values=[], width=18)
        self.cmb_object_id.grid(row=1, column=7, sticky="w", pady=(8, 0))
        self.cmb_object_id.bind("<<ComboboxSelected>>", lambda e: self._load_existing_rows())

        tk.Label(top, text="ФИО:").grid(row=2, column=0, sticky="w", pady=(8, 0))
        self.fio_var = tk.StringVar()
        self.cmb_fio = AutoCompleteCombobox(top, textvariable=self.fio_var, width=30)
        self.cmb_fio.set_completion_list(self.emp_names)
        self.cmb_fio.grid(row=2, column=1, sticky="w", pady=(8, 0))
        self.cmb_fio.bind("<<ComboboxSelected>>", self._on_fio_select)

        tk.Label(top, text="Табельный №:").grid(row=2, column=2, sticky="w", padx=(16, 4), pady=(8, 0))
        self.ent_tbn = ttk.Entry(top, width=14)
        self.ent_tbn.grid(row=2, column=3, sticky="w", pady=(8, 0))

        tk.Label(top, text="Должность:").grid(row=2, column=4, sticky="w", padx=(16, 4), pady=(8, 0))
        self.pos_var = tk.StringVar()
        self.ent_pos = ttk.Entry(top, textvariable=self.pos_var, width=40, state="readonly")
        self.ent_pos.grid(row=2, column=5, sticky="w", pady=(8, 0))

        # Кнопки действий
        btns = tk.Frame(top)
        btns.grid(row=3, column=0, columnspan=8, sticky="w", pady=(8, 0))
        ttk.Button(btns, text="Добавить в табель", command=self.add_row).grid(row=0, column=0, padx=4)
        ttk.Button(btns, text="Добавить подразделение", command=self.add_department_all).grid(row=0, column=1, padx=4)
        ttk.Button(btns, text="5/2 всем", command=self.fill_52_all).grid(row=0, column=2, padx=4)
        ttk.Button(btns, text="Проставить часы", command=self.fill_hours_all).grid(row=0, column=3, padx=4)
        ttk.Button(btns, text="Очистить все строки", command=self.clear_all_rows).grid(row=0, column=4, padx=4)
        ttk.Button(btns, text="Загрузить из Excel", command=self.import_from_excel).grid(row=0, column=5, padx=4)
        ttk.Button(btns, text="Обновить справочник", command=self.reload_spravochnik).grid(row=0, column=6, padx=4)
        ttk.Button(btns, text="Копировать из месяца…", command=self.copy_from_month).grid(row=0, column=7, padx=4)
        ttk.Button(btns, text="Сохранить", command=self.save_all).grid(row=0, column=8, padx=4)


        # Основной контейнер с прокруткой
        main_frame = tk.Frame(self)
        main_frame.pack(fill="both", expand=True, padx=8, pady=(4, 8))

        # Канвас для шапки (фиксирован сверху)
        self.header_canvas = tk.Canvas(main_frame, borderwidth=0, highlightthickness=0, height=28)
        self.header_canvas.grid(row=0, column=0, sticky="ew")

        # Канвас с телом таблицы (вертикально скроллится)
        self.main_canvas = tk.Canvas(main_frame, borderwidth=0, highlightthickness=0)
        self.main_canvas.grid(row=1, column=0, sticky="nsew")

        # Скроллбары
        self.vscroll = ttk.Scrollbar(main_frame, orient="vertical", command=self.main_canvas.yview)
        self.vscroll.grid(row=1, column=1, sticky="ns")
        self.hscroll = ttk.Scrollbar(main_frame, orient="horizontal")
        self.hscroll.grid(row=2, column=0, sticky="ew")

        main_frame.grid_rowconfigure(1, weight=1)
        main_frame.grid_columnconfigure(0, weight=1)

        # Таблицы внутри канвасов
        self.header_table = tk.Frame(self.header_canvas, bg="#ffffff")
        self.header_window = self.header_canvas.create_window((0, 0), window=self.header_table, anchor="nw")

        self.table = tk.Frame(self.main_canvas, bg="#ffffff")
        self.canvas_window = self.main_canvas.create_window((0, 0), window=self.table, anchor="nw")

        # Привязки скролла
        self.main_canvas.configure(yscrollcommand=self.vscroll.set, xscrollcommand=self._on_xscroll_main)
        # Горизонтальный скролл двигает оба канваса
        self.hscroll.configure(command=self._xscroll_both)

        # Обновление области прокрутки
        self.table.bind("<Configure>", self._on_scroll_frame_configure)

        # Создаём шапку в первой строке таблицы
        self._configure_table_columns()
        self._configure_table_columns()   # обновим оба фрейма (см. ниже)
        self._build_header_row(self.header_table)

        # Обработчики колеса мыши
        self.main_canvas.bind("<MouseWheel>", self._on_wheel)
        self.main_canvas.bind("<Shift-MouseWheel>", self._on_shift_wheel)
        self.bind_all("<MouseWheel>", self._on_wheel_anywhere)

        # Коллекция строк
        self.rows: List[RowWidget] = []

        # Нижняя панель
        bottom = tk.Frame(self)
        bottom.pack(fill="x", padx=8, pady=(0, 8))

        self.lbl_object_total = tk.Label(
            bottom, text="Сумма: сотрудников 0 | дней 0 | часов 0",
            font=("Segoe UI", 10, "bold")
        )
        self.lbl_object_total.pack(side="left")

        # Пагинация справа
        pag = tk.Frame(bottom)
        pag.pack(side="right")

        ttk.Label(pag, text="На странице:").pack(side="left", padx=(0, 4))
        self.cmb_page_size = ttk.Combobox(pag, state="readonly", width=6,
                                          values=[25, 50, 100])
        self.cmb_page_size.pack(side="left")
        self.cmb_page_size.set(str(self.page_size.get()))
        self.cmb_page_size.bind(
            "<<ComboboxSelected>>",
            lambda e: self._on_page_size_change()
        )

        ttk.Button(pag, text="⟨", width=3, command=lambda: self._render_page(self.current_page - 1)).pack(side="left", padx=4)
        self.lbl_page = ttk.Label(pag, text="Стр. 1 / 1")
        self.lbl_page.pack(side="left")
        ttk.Button(pag, text="⟩", width=3, command=lambda: self._render_page(self.current_page + 1)).pack(side="left", padx=4)

        # Применяем переданные значения (если открываем существующий табель)
        # Подразделение
        if self._init_department:
            if self._init_department in deps:
                self.cmb_department.set(self._init_department)

        # Период
        if self._init_year:
            self.spn_year.delete(0, "end")
            self.spn_year.insert(0, str(self._init_year))
        if self._init_month:
            if 1 <= self._init_month <= 12:
                self.cmb_month.current(self._init_month - 1)

        # Адрес и ID объекта
        if self._init_object_addr:
            if self._init_object_addr in self.address_options:
                self.cmb_address.set(self._init_object_addr)
        if self._init_object_id:
            # сначала заполним ID для текущего адреса
            self._on_address_change()
            if self._init_object_id in (self.cmb_object_id.cget("values") or []):
                self.cmb_object_id.set(self._init_object_id)

        self._on_department_select()

        # Если страница в режиме "только просмотр" — блокируем редактирование
        if self.read_only:
            # Отключаем верхние кнопки действий
            try:
                for child in btns.winfo_children():
                    child.configure(state="disabled")
            except Exception:
                pass

            # Сохраняем ссылку на панель кнопок, чтобы при надобности ещё обращаться
            self._btns_frame = btns

            # Чуть меняем подсказку внизу
            try:
                self.lbl_object_total.config(
                    text=self.lbl_object_total.cget("text") + " (режим просмотра)"
                )
            except Exception:
                pass

    def _build_header_row(self, parent):
        hb = self.HEADER_BG
        tk.Label(parent, text="ФИО", bg=hb, anchor="w", font=("Segoe UI", 9, "bold")).grid(
            row=0, column=0, padx=0, pady=(0, 2), sticky="nsew")
        tk.Label(parent, text="Таб.№", bg=hb, anchor="center", font=("Segoe UI", 9, "bold")).grid(
            row=0, column=1, padx=0, pady=(0, 2), sticky="nsew")

        for d in range(1, 32):
            tk.Label(parent, text=str(d), bg=hb, anchor="center", font=("Segoe UI", 9, "bold")).grid(
                row=0, column=1 + d, padx=0, pady=(0, 2), sticky="nsew")

        tk.Label(parent, text="Дней", bg=hb, anchor="e", font=("Segoe UI", 9, "bold")).grid(
            row=0, column=33, padx=(4, 1), pady=(0, 2), sticky="nsew")
        tk.Label(parent, text="Часы", bg=hb, anchor="e", font=("Segoe UI", 9, "bold")).grid(
            row=0, column=34, padx=(4, 1), pady=(0, 2), sticky="nsew")

        tk.Label(parent, text="Пер.день", bg=hb, anchor="e", font=("Segoe UI", 9, "bold")).grid(
            row=0, column=35, padx=(4, 1), pady=(0, 2), sticky="nsew")
        tk.Label(parent, text="Пер.ночь", bg=hb, anchor="e", font=("Segoe UI", 9, "bold")).grid(
            row=0, column=36, padx=(4, 1), pady=(0, 2), sticky="nsew")

        tk.Label(parent, text="5/2", bg=hb, anchor="center", font=("Segoe UI", 9, "bold")).grid(
            row=0, column=37, padx=1, pady=(0, 2), sticky="nsew")
        tk.Label(parent, text="Удалить", bg=hb, anchor="center", font=("Segoe UI", 9, "bold")).grid(
            row=0, column=38, padx=1, pady=(0, 2), sticky="nsew")

    def _on_scroll_frame_configure(self, _=None):
        """
        Вызывается при изменении размеров фрейма с телом таблицы.
        Обновляет область прокрутки и синхронизирует ширину шапки с телом.
        """
        # Область прокрутки для тела
        self.main_canvas.configure(scrollregion=self.main_canvas.bbox("all"))
        try:
            content_bbox = self.main_canvas.bbox("all")
            if content_bbox:
                x1, y1, x2, y2 = content_bbox
                # Область прокрутки по X для шапки
                self.header_canvas.configure(scrollregion=(0, 0, x2, 0))
            # ВАЖНО: делаем ширину header_canvas такой же, как у main_canvas,
            # чтобы grid‑колонки шапки и тела физически совпадали по ширине.
            self.header_canvas.configure(width=self.main_canvas.winfo_width())
        except Exception:
            pass

    def _configure_table_columns(self):
        px = self.COLPX
        # для тела
        for frame in (self.table, self.header_table):
            if not frame:
                continue
            frame.grid_columnconfigure(0, minsize=px['fio'], weight=0)
            frame.grid_columnconfigure(1, minsize=px['tbn'], weight=0)
            for col in range(2, 33):
                frame.grid_columnconfigure(col, minsize=px['day'], weight=0)
            frame.grid_columnconfigure(33, minsize=px['days'], weight=0)
            frame.grid_columnconfigure(34, minsize=px['hours'], weight=0)
            frame.grid_columnconfigure(35, minsize=px['hours'], weight=0)
            frame.grid_columnconfigure(36, minsize=px['hours'], weight=0)
            frame.grid_columnconfigure(37, minsize=px['btn52'], weight=0)
            frame.grid_columnconfigure(38, minsize=px['del'], weight=0)

    def _on_wheel(self, event):
        if self.main_canvas.winfo_exists():
            self.main_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        return "break"

    def _on_wheel_anywhere(self, event):
        try:
            widget = event.widget
            while widget:
                if widget == self.main_canvas or widget == self.table:
                    return self._on_wheel(event)
                widget = widget.master
        except:
            pass
        return None

    def _on_shift_wheel(self, event):
        if self.main_canvas.winfo_exists():
            dx = int(-1 * (event.delta / 120))
            self.main_canvas.xview_scroll(dx, "units")
            try:
                self.header_canvas.xview_scroll(dx, "units")
            except Exception:
                pass
        return "break"

    def _xscroll_both(self, *args):
        try:
            self.main_canvas.xview(*args)
            self.header_canvas.xview(*args)
        except Exception:
            pass

    def _on_xscroll_main(self, first, last):
        try:
            self.hscroll.set(first, last)
            # Двигаем шапку вслед за телом
            self.header_canvas.xview_moveto(first)
        except Exception:
            pass

    def _on_period_change(self):
        self._update_rows_days_enabled()
        self._load_existing_rows()

    def _on_address_change(self, *_):
        addr = self.cmb_address.get().strip()
        ids = sorted(self.addr_to_ids.get(addr, []))
        if ids:
            self.cmb_object_id.config(state="readonly", values=ids)
            if self.cmb_object_id.get() not in ids:
                self.cmb_object_id.set(ids[0])
        else:
            self.cmb_object_id.config(state="normal", values=[])
            self.cmb_object_id.set("")

    def _on_address_select(self, *_):
        self._on_address_change()
        # очищаем модель и UI при смене адреса
        self.model_rows.clear()
        for r in list(self.rows):
            r.destroy()
        self.rows.clear()
        self._regrid_rows()
        self._load_existing_rows()

    def get_year_month(self) -> Tuple[int, int]:
        return int(self.spn_year.get()), self.cmb_month.current() + 1

    def _update_rows_days_enabled(self):
        y, m = self.get_year_month()
        CHUNK = 20
        rows_list = list(self.rows)

        def apply_chunk(idx: int = 0):
            end = min(idx + CHUNK, len(rows_list))
            for j in range(idx, end):
                r = rows_list[j]
                r.set_day_font(self.DAY_ENTRY_FONT)
                r.update_days_enabled(y, m)
            if end < len(rows_list):
                self.after(1, lambda: apply_chunk(end))
            else:
                self._recalc_object_total()

        apply_chunk(0)

    def _regrid_rows(self):
        # Перегрид всех строк под заголовком (начиная с 1)
        for idx, r in enumerate(self.rows, start=1):
            r.regrid_to(idx)
            r.set_day_font(self.DAY_ENTRY_FONT)
        self.after(30, self._on_scroll_frame_configure)
        self._recalc_object_total()

    def _on_page_size_change(self):
        try:
            sz = int(self.cmb_page_size.get())
            if sz not in (25, 50, 100):
                sz = 50
            self.page_size.set(sz)
        except Exception:
            self.page_size.set(50)
        # Перед сменой страницы — сохранить правки из видимых строк в модель
        self._sync_visible_to_model()
        self._render_page(1)

    def _page_count(self) -> int:
        sz = max(1, int(self.page_size.get()))
        n = len(self.model_rows)
        return max(1, math.ceil(n / sz))

    def _update_page_label(self):
        self.lbl_page.config(text=f"Стр. {self.current_page} / {self._page_count()}")

    def _sync_visible_to_model(self):
        """Считывает значения из видимых RowWidget в модель."""
        if not self.rows:
            return
        sz = max(1, int(self.page_size.get()))
        start = (self.current_page - 1) * sz
        for i, roww in enumerate(self.rows):
            idx = start + i
            if 0 <= idx < len(self.model_rows):
                # забираем сырые значения строками (с переработкой)
                vals = []
                for e in roww.day_entries:
                    raw = e.get().strip()
                    vals.append(raw if raw else None)
                self.model_rows[idx]["hours"] = vals

    def _render_page(self, page: Optional[int] = None):
        """Рендерит только текущую страницу из модели."""
        # Сохраняем видимые правки, если не в массовом режиме
        if not getattr(self, "_suspend_sync", False):
            self._sync_visible_to_model()

        # Очистка текущих UI-строк
        for r in list(getattr(self, "rows", [])):
            try:
                r.destroy()
            except Exception:
                pass
        self.rows = []

        total_pages = self._page_count()
        if page is None:
            page = self.current_page
        page = max(1, min(total_pages, page))
        self.current_page = page

        sz = max(1, int(self.page_size.get()))
        start = (page - 1) * sz
        end = min(start + sz, len(self.model_rows))

        y, m = self.get_year_month()
        # Создаём виджеты только для среза
        for i in range(start, end):
            rec = self.model_rows[i]
            row_index = len(self.rows) + 1
            w = RowWidget(self.table, row_index, rec["fio"], rec["tbn"], self.get_year_month, self.delete_row)
            w.set_day_font(self.DAY_ENTRY_FONT)

            # применим формат дней только один раз
            w.update_days_enabled(y, m)

            # подставим значения часов
            hours = rec.get("hours") or [None] * 31
            w.set_hours(hours)
            self.rows.append(w)

        self._regrid_rows()
        self._update_page_label()
        self._recalc_object_total()  # итоги по всей модели, не только по странице

    def _recalc_object_total(self):
        tot_h = 0.0
        tot_d = 0
        tot_ot_day = 0.0
        tot_ot_night = 0.0

        for rec in self.model_rows:
            hours = rec.get("hours") or [None] * 31
            for raw in hours:
                if not raw:
                    continue
                hv = parse_hours_value(raw)
                d_ot, n_ot = parse_overtime(raw)
                if isinstance(hv, (int, float)) and hv > 1e-12:
                    tot_h += float(hv)
                    tot_d += 1
                if isinstance(d_ot, (int, float)):
                    tot_ot_day += float(d_ot)
                if isinstance(n_ot, (int, float)):
                    tot_ot_night += float(n_ot)

        sh = f"{tot_h:.2f}".rstrip("0").rstrip(".")
        sod = f"{tot_ot_day:.2f}".rstrip("0").rstrip(".")
        son = f"{tot_ot_night:.2f}".rstrip("0").rstrip(".")
        cnt = len(self.model_rows)

        self.lbl_object_total.config(
            text=f"Сумма: сотрудников {cnt} | дней {tot_d} | часов {sh} | пер.день {sod} | пер.ночь {son}"
        )

    def add_row(self):
        if self.read_only:
            return
        fio = self.fio_var.get().strip()
        tbn = self.ent_tbn.get().strip()
        if not fio:
            messagebox.showwarning("Объектный табель", "Выберите ФИО.")
            return

        key = (fio.strip().lower(), tbn.strip())
        existing = {(r["fio"].strip().lower(), r["tbn"].strip()) for r in self.model_rows}
        if key in existing:
            if not messagebox.askyesno("Дублирование",
                                       f"Сотрудник уже есть в реестре:\n{fio} (Таб.№ {tbn}).\nДобавить ещё одну строку?"):
                return

        self.model_rows.append({"fio": fio, "tbn": tbn, "hours": [None] * 31})
        self._render_page(self.current_page)

    def add_department_all(self):
        if self.read_only:
            return
        dep_sel = (self.cmb_department.get() or "Все").strip()

        # Подбор списка сотрудников по подразделению
        if dep_sel == "Все":
            candidates = self.employees[:]  # все сотрудники
            if not candidates:
                messagebox.showinfo("Объектный табель", "Справочник сотрудников пуст.")
                return
            if not messagebox.askyesno("Добавить всех", f"Добавить в реестр всех сотрудников ({len(candidates)})?"):
                return
        else:
            candidates = [e for e in self.employees if len(e) > 3 and (e[3] or "").strip() == dep_sel]
            if not candidates:
                messagebox.showinfo("Объектный табель", f"В подразделении «{dep_sel}» нет сотрудников.")
                return

        # Уникальность по (fio.lower, tbn)
        existing = {(r["fio"].strip().lower(), r["tbn"].strip()) for r in self.model_rows}

        # Диалог прогресса и пакетная обработка
        dlg = BatchAddDialog(self, total=len(candidates), title="Добавление сотрудников")

        CHUNK = 50  # крупнее пакет, т.к. мы не создаем виджеты
        added = 0

        def add_chunk(start_idx: int = 0):
            nonlocal added, existing
            if dlg.cancelled:
                finalize()
                return

            end_idx = min(start_idx + CHUNK, len(candidates))
            for i in range(start_idx, end_idx):
                fio, tbn, pos, dep = candidates[i]
                key = (fio.strip().lower(), (tbn or "").strip())
                if key in existing:
                    dlg.step(1)
                    continue
                self.model_rows.append({"fio": fio, "tbn": tbn, "hours": [None] * 31})
                existing.add(key)
                added += 1
                dlg.step(1)

            if end_idx >= len(candidates):
                finalize()
            else:
                self.after(1, lambda: add_chunk(end_idx))

        def finalize():
            try:
                dlg.close()
            except Exception:
                pass
            self._render_page(1)
            messagebox.showinfo("Объектный табель", f"Добавлено сотрудников: {added}")

        add_chunk(0)

    def _on_fio_select(self, *_):
        fio = self.fio_var.get().strip()
        tbn, pos = self.emp_info.get(fio, ("", ""))
        self.ent_tbn.delete(0, "end")
        self.ent_tbn.insert(0, tbn)
        self.pos_var.set(pos)

    def reload_spravochnik(self):
        try:
            cur_dep = (self.cmb_department.get() or "Все").strip()
            cur_addr = (self.cmb_address.get() or "").strip()
            cur_id = (self.cmb_object_id.get() or "").strip()
            cur_fio = (self.fio_var.get() or "").strip()

            self._load_spr_data_from_db()

            self.cmb_department.config(values=self.departments)
            if cur_dep in self.departments:
                self.cmb_department.set(cur_dep)
            else:
                try:
                    saved_dep = get_selected_department_from_config()
                    self.cmb_department.set(saved_dep if saved_dep in self.departments else self.departments[0])
                except Exception:
                    self.cmb_department.set(self.departments[0] if self.departments else "Все")

            self.cmb_address.set_completion_list(self.address_options)
            if cur_addr in self.address_options:
                self.cmb_address.set(cur_addr)
            else:
                self.cmb_address.set("")
            self._on_address_change()
            if cur_id and cur_id in (self.cmb_object_id.cget("values") or []):
                self.cmb_object_id.set(cur_id)

            self._on_department_select()
            dep_sel = (self.cmb_department.get() or "Все").strip()
            if dep_sel == "Все":
                allowed = [e[0] for e in self.employees]
            else:
                allowed = [e[0] for e in self.employees if len(e) > 3 and (e[3] or "").strip() == dep_sel]
            seen = set()
            allowed = [n for n in allowed if (n not in seen and not seen.add(n))]
            if cur_fio and cur_fio in allowed:
                self.fio_var.set(cur_fio)
                self._on_fio_select()
            else:
                self.fio_var.set("")
                self.ent_tbn.delete(0, "end")
                self.pos_var.set("")

            messagebox.showinfo("Справочник", "Справочник обновлён.")
        except Exception as e:
            messagebox.showerror("Справочник", f"Ошибка перечтения справочника:\n{e}")

    def fill_52_all(self):
        if self.read_only:
            return
        if not self.model_rows:
            messagebox.showinfo("5/2 всем", "Список сотрудников пуст.")
            return

        y, m = self.get_year_month()
        days = month_days(y, m)

        for rec in self.model_rows:
            hrs = [None] * 31
            for d in range(1, days + 1):
                wd = datetime(y, m, d).weekday()
                if wd < 4:
                    hrs[d - 1] = "8,25"
                elif wd == 4:
                    hrs[d - 1] = "7"
                else:
                    hrs[d - 1] = None
            rec["hours"] = hrs

        # ВАЖНО: перерисовываем без синхронизации видимых значений
        self._suspend_sync = True
        try:
            self._render_page(self.current_page)
        finally:
            self._suspend_sync = False

        messagebox.showinfo("5/2 всем", "Режим 5/2 установлен всем сотрудникам текущего реестра.")

    def fill_hours_all(self):
        if self.read_only:
            return
        if not self.model_rows:
            messagebox.showinfo("Проставить часы", "Список сотрудников пуст.")
            return
        y, m = self.get_year_month()
        max_day = month_days(y, m)
        dlg = HoursFillDialog(self, max_day)
        if not getattr(dlg, "result", None):
            return
        day = dlg.result["day"]
        clear = bool(dlg.result.get("clear", False))
        if day > max_day:
            messagebox.showwarning("Проставить часы", f"В {month_name_ru(m)} {y} только {max_day} дней.")
            return

        if clear:
            for rec in self.model_rows:
                hrs = rec.get("hours") or [None] * 31
                hrs[day - 1] = None
                rec["hours"] = hrs
            self._suspend_sync = True
            try:
                self._render_page(self.current_page)
            finally:
                self._suspend_sync = False
            messagebox.showinfo("Проставить часы", f"День {day} очищен у {len(self.model_rows)} сотрудников.")
            return

        hours_val = float(dlg.result["hours"])
        s = f"{hours_val:.2f}".rstrip("0").rstrip(".").replace(".", ",")
        for rec in self.model_rows:
            hrs = rec.get("hours") or [None] * 31
            hrs[day - 1] = s if hours_val > 1e-12 else None
            rec["hours"] = hrs

        self._suspend_sync = True
        try:
            self._render_page(self.current_page)
        finally:
            self._suspend_sync = False

        messagebox.showinfo("Проставить часы", f"Проставлено {s} ч в день {day} для {len(self.model_rows)} сотрудников.")


    def delete_row(self, roww: RowWidget):
        if self.read_only:
            return
        # Синхронизируем видимые правки
        self._sync_visible_to_model()
        try:
            # Определяем глобальный индекс
            sz = max(1, int(self.page_size.get()))
            start = (self.current_page - 1) * sz
            local_idx = self.rows.index(roww)
            global_idx = start + local_idx
        except Exception:
            global_idx = None

        try:
            roww.destroy()
        except Exception:
            pass
        try:
            self.rows.remove(roww)
        except Exception:
            pass

        if global_idx is not None and 0 <= global_idx < len(self.model_rows):
            del self.model_rows[global_idx]

        # Перерендерим текущую страницу (возможно, перелистнём назад, если страница опустела)
        if self.current_page > self._page_count():
            self.current_page = self._page_count()
        self._render_page(self.current_page)

    def clear_all_rows(self):
        if self.read_only:
            return
        if not self.model_rows:
            return
        if not messagebox.askyesno("Объектный табель", "Очистить все строки?"):
            return
        self.model_rows.clear()
        self._render_page(1)

    def _current_file_path(self) -> Optional[Path]:
        """Генерирует путь к файлу с учетом подразделения"""
        addr = self.cmb_address.get().strip()
        oid = self.cmb_object_id.get().strip()
        dep = self.cmb_department.get().strip()
    
        if not addr and not oid:
            return None
    
        y, m = self.get_year_month()
        id_part = oid if oid else safe_filename(addr)
    
        # Добавляем подразделение в имя файла
        dep_part = safe_filename(dep) if dep and dep != "Все" else "ВсеПодразделения"
    
        return self.out_dir / f"Объектный_табель_{id_part}_{dep_part}_{y}_{m:02d}.xlsx"

    def _file_path_for(self, year: int, month: int, addr: Optional[str] = None, 
                   oid: Optional[str] = None, department: Optional[str] = None) -> Optional[Path]:
        """Генерирует путь к файлу для заданных параметров"""
        addr = (addr if addr is not None else self.cmb_address.get().strip())
        oid = (oid if oid is not None else self.cmb_object_id.get().strip())
        dep = (department if department is not None else self.cmb_department.get().strip())
    
        if not addr and not oid:
            return None
    
        id_part = oid if oid else safe_filename(addr)
        dep_part = safe_filename(dep) if dep and dep != "Все" else "ВсеПодразделения"
    
        return self.out_dir / f"Объектный_табель_{id_part}_{dep_part}_{year}_{month:02d}.xlsx"

    def _ensure_sheet(self, wb) -> Any:
        """Проверяет наличие листа 'Табель' с правильной структурой и создает его при необходимости"""
        if "Табель" in wb.sheetnames:
            ws = wb["Табель"]
            hdr_first = str(ws.cell(1, 1).value or "")
            # Проверяем наличие новых столбцов (включая Подразделение)
            if hdr_first == "ID объекта" and ws.max_column >= (7 + 31 + 4):  # +1 для подразделения, +4 для итогов и переработок
                return ws
            # Если структура не совпадает, переименовываем старый лист
            base = "Табель_OLD"
            new_name = base
            i = 1
            while new_name in wb.sheetnames:
                i += 1
                new_name = f"{base}{i}"
            ws.title = new_name

        # Создаем новый лист с правильной структурой
        ws2 = wb.create_sheet("Табель")
        hdr = ["ID объекта", "Адрес", "Месяц", "Год", "ФИО", "Табельный №", "Подразделение"] + \
              [str(i) for i in range(1, 32)] + \
              ["Итого дней", "Итого часов по табелю", "Переработка день", "Переработка ночь"]
        ws2.append(hdr)

        # Настройка ширины столбцов
        ws2.column_dimensions["A"].width = 14  # ID объекта
        ws2.column_dimensions["B"].width = 40  # Адрес
        ws2.column_dimensions["C"].width = 10  # Месяц
        ws2.column_dimensions["D"].width = 8   # Год
        ws2.column_dimensions["E"].width = 28  # ФИО
        ws2.column_dimensions["F"].width = 14  # Табельный №
        ws2.column_dimensions["G"].width = 20  # Подразделение

        # Дни месяца (1-31) - столбцы 8-38
        for i in range(8, 8 + 31):
            ws2.column_dimensions[get_column_letter(i)].width = 6

        # Итоговые столбцы
        ws2.column_dimensions[get_column_letter(39)].width = 10  # Итого дней
        ws2.column_dimensions[get_column_letter(40)].width = 18  # Итого часов по табелю
        ws2.column_dimensions[get_column_letter(41)].width = 14  # Переработка день
        ws2.column_dimensions[get_column_letter(42)].width = 14  # Переработка ночь

        ws2.freeze_panes = "A2"
        return ws2

    def _load_existing_rows(self):
        """
        Загружает существующие строки табеля из БД для текущих:
        - адреса / ID объекта
        - месяца / года
        - подразделения
        - пользователя (current_user.id)
        """
        # Очистим модель
        self.model_rows.clear()

        addr = self.cmb_address.get().strip()
        oid = self.cmb_object_id.get().strip()
        y, m = self.get_year_month()
        current_dep = self.cmb_department.get().strip()
        if current_dep == "Все":
            # В БД храним конкретное подразделение, а "Все" — это
            # объединение, которого пока не будет. Можно здесь
            # просто не грузить ничего или придумать отдельную логику.
            self._render_page(1)
            return

        # user_id владельца табеля:
        # если передан явно (например, из реестра) — используем его,
        # иначе берём текущего пользователя
        if self.owner_user_id is not None:
            user_id = self.owner_user_id
        else:
            user = getattr(self.app_ref, "current_user", None) if hasattr(self, "app_ref") else None
            user_id = (user or {}).get("id")

        if not user_id:
            self._render_page(1)
            return

        try:
            rows = load_timesheet_rows_from_db(
                object_id=oid or None,
                object_addr=addr,
                department=current_dep,
                year=y,
                month=m,
                user_id=user_id,
            )

            self.model_rows.extend(rows)
            self._render_page(1)
        except Exception as e:
            logging.exception("Ошибка загрузки табеля из БД")
            messagebox.showerror("Загрузка", f"Не удалось загрузить табель из БД:\n{e}")
            self._render_page(1)

    def save_all(self):
        if self.read_only:
            messagebox.showinfo(
                "Объектный табель",
                "Табель открыт в режиме только просмотра. Сохранение недоступно."
            )
            return
        # Сохраним правки с текущей страницы в модель
        self._sync_visible_to_model()

        addr = self.cmb_address.get().strip()
        oid = self.cmb_object_id.get().strip()
        y, m = self.get_year_month()
        current_dep = self.cmb_department.get().strip()

        if not addr and not oid:
            messagebox.showwarning("Сохранение", "Укажите адрес и/или ID объекта, а также период.")
            return

        if current_dep == "Все":
            messagebox.showwarning(
                "Сохранение",
                "Для сохранения в БД выберите конкретное подразделение (не «Все»).",
            )
            return

        # Текущий пользователь из MainApp
        user = getattr(self.app_ref, "current_user", None) if hasattr(self, "app_ref") else None
        user_id = (user or {}).get("id")
        if not user_id:
            messagebox.showerror(
                "Сохранение",
                "Не удалось определить текущего пользователя. Повторите вход в систему.",
            )
            return

        # Сохранение в БД
        try:
            header_id = upsert_timesheet_header(
                object_id=oid or None,
                object_addr=addr,
                department=current_dep,
                year=y,
                month=m,
                user_id=user_id,
            )
            replace_timesheet_rows(header_id, self.model_rows)
        except Exception as e:
            logging.exception("Ошибка сохранения табеля в БД")
            messagebox.showerror("Сохранение", f"Ошибка сохранения в БД:\n{e}")
            return

        # Если хочешь оставить Excel как резервную копию — можно вызвать старый код.
        # Если нет необходимости — блок ниже можно удалить.
        try:
            fpath = self._current_file_path()
            if fpath:
                addr_local = addr
                oid_local = oid
                y_local, m_local = y, m
                current_dep_local = current_dep

                if fpath.exists():
                    wb = load_workbook(fpath)
                else:
                    fpath.parent.mkdir(parents=True, exist_ok=True)
                    wb = Workbook()
                    if wb.active:
                        wb.remove(wb.active)

                ws = self._ensure_sheet(wb)

                # Удаляем старые записи ТЕКУЩЕГО подразделения
                to_del = []
                for r in range(2, ws.max_row + 1):
                    row_oid = (ws.cell(r, 1).value or "")
                    row_addr = (ws.cell(r, 2).value or "")
                    row_m = int(ws.cell(r, 3).value or 0)
                    row_y = int(ws.cell(r, 4).value or 0)
                    row_dep = (ws.cell(r, 7).value or "")

                    match_obj = (oid_local and row_oid == oid_local) or (not oid_local and row_addr == addr_local)
                    match_period = (row_m == m_local and row_y == y_local)
                    match_dep = (current_dep_local == "Все" or row_dep == current_dep_local)

                    if match_obj and match_period and match_dep:
                        to_del.append(r)

                for r in reversed(to_del):
                    ws.delete_rows(r, 1)

                # Записываем модель (логика как раньше)
                for rec in self.model_rows:
                    fio = rec["fio"]
                    tbn = rec["tbn"]
                    hours_list = rec.get("hours") or [None] * 31

                    department = current_dep_local if current_dep_local != "Все" else ""
                    for emp_fio, emp_tbn, emp_pos, emp_dep in self.employees:
                        if emp_fio == fio:
                            if emp_dep:
                                department = emp_dep
                            break

                    total_hours = 0.0
                    total_days = 0
                    total_ot_day = 0.0
                    total_ot_night = 0.0

                    day_values = []
                    for raw in hours_list:
                        if not raw:
                            day_values.append(None)
                            continue
                        hrs = parse_hours_value(raw)
                        d_ot, n_ot = parse_overtime(raw)

                        if isinstance(hrs, (int, float)) and hrs > 1e-12:
                            total_hours += hrs
                            total_days += 1

                        cell_str = None
                        try:
                            base = f"{hrs:.2f}".rstrip("0").rstrip(".") if hrs is not None else None
                            if base:
                                if d_ot or n_ot:
                                    d_ot_val = d_ot if d_ot else 0
                                    n_ot_val = n_ot if n_ot else 0
                                    cell_str = f"{base}({d_ot_val:.0f}/{n_ot_val:.0f})"
                                    total_ot_day += d_ot_val
                                    total_ot_night += n_ot_val
                                else:
                                    cell_str = base
                        except Exception:
                            cell_str = str(raw)

                        day_values.append(cell_str)

                    row_values = [oid_local, addr_local, m_local, y_local, fio, tbn, department] + day_values + [
                        total_days if total_days else None,
                        None if abs(total_hours) < 1e-12 else total_hours,
                        None if abs(total_ot_day) < 1e-12 else total_ot_day,
                        None if abs(total_ot_night) < 1e-12 else total_ot_night
                    ]
                    ws.append(row_values)

                wb.save(fpath)
                messagebox.showinfo(
                    "Сохранение",
                    f"Сохранено в БД (user_id={user_id}) и в файл:\n{fpath}",
                )
            else:
                messagebox.showinfo(
                    "Сохранение",
                    f"Сохранено в БД (user_id={user_id}). Локальный файл не создан (нет адреса/ID).",
                )
        except Exception as e:
            logging.exception("Ошибка резервного сохранения в Excel")
            messagebox.showwarning(
                "Сохранение",
                f"В БД табель сохранён, но резервное сохранение в Excel завершилось ошибкой:\n{e}",
            )

    def _on_department_select(self):
        """Обработчик смены подразделения"""
        dep_sel = (self.cmb_department.get() or "Все").strip()
        set_selected_department_in_config(dep_sel)
    
        # КЛЮЧЕВОЕ ИСПРАВЛЕНИЕ: очищаем реестр и загружаем данные для нового подразделения
        for r in list(self.rows):
            r.destroy()
        self.rows.clear()
    
        # Фильтруем список сотрудников
        if dep_sel == "Все":
            names = [e[0] for e in self.employees]
        else:
            names = [e[0] for e in self.employees if len(e) > 3 and (e[3] or "").strip() == dep_sel]
    
        seen = set()
        filtered = []
        for n in names:
            if n not in seen:
                seen.add(n)
                filtered.append(n)
    
        self.cmb_fio.set_completion_list(filtered)
    
        cur = self.fio_var.get().strip()
        if cur and cur not in filtered:
            self.fio_var.set("")
            self.ent_tbn.delete(0, "end")
            self.pos_var.set("")
    
        # Загружаем сохраненные данные для выбранного подразделения
        self._load_existing_rows()

    def import_from_excel(self):
        """
        Загрузка сотрудников и часов из старого Excel-файла табеля в текущий реестр (в модель/память).
        Файл должен иметь лист 'Табель' со структурой, как в _ensure_sheet.
        Фильтрация по:
          - текущему объекту (ID/Адрес),
          - месяцу/году,
          - подразделению (если не 'Все').
        """
        if self.read_only:
            messagebox.showinfo("Импорт из Excel", "Табель открыт в режиме только просмотра.")
            return

        from tkinter import filedialog

        addr = self.cmb_address.get().strip()
        oid = self.cmb_object_id.get().strip()
        current_dep = self.cmb_department.get().strip()
        y, m = self.get_year_month()

        if not addr and not oid:
            messagebox.showwarning(
                "Импорт из Excel",
                "Укажите адрес и/или ID объекта, а также период (месяц и год) перед импортом.",
            )
            return

        # выбор файла
        path = filedialog.askopenfilename(
            parent=self,
            title="Выберите Excel-файл табеля",
            filetypes=[
                ("Excel файлы", "*.xlsx *.xlsm *.xltx *.xltm"),
                ("Все файлы", "*.*"),
            ],
        )
        if not path:
            return

        try:
            wb = load_workbook(path, data_only=True)
        except Exception as e:
            messagebox.showerror("Импорт из Excel", f"Не удалось открыть файл:\n{e}", parent=self)
            return

        try:
            ws = self._ensure_sheet(wb)
        except Exception as e:
            messagebox.showerror("Импорт из Excel", f"Не удалось подготовить лист 'Табель':\n{e}", parent=self)
            return

        imported: List[Dict[str, Any]] = []

        try:
            for r in range(2, ws.max_row + 1):
                row_oid = (ws.cell(r, 1).value or "")
                row_addr = (ws.cell(r, 2).value or "")
                row_m = int(ws.cell(r, 3).value or 0)
                row_y = int(ws.cell(r, 4).value or 0)
                fio = str(ws.cell(r, 5).value or "").strip()
                tbn = str(ws.cell(r, 6).value or "").strip()
                row_dep = str(ws.cell(r, 7).value or "").strip()

                # фильтр по периоду
                if row_m != m or row_y != y:
                    continue

                # фильтр по объекту
                if oid:
                    if row_oid != oid:
                        continue
                else:
                    if row_addr != addr:
                        continue

                # фильтр по подразделению
                if current_dep != "Все" and row_dep != current_dep:
                    continue

                # читаем часы по дням (столбцы 8..8+31-1)
                hours_raw: List[Optional[str]] = []
                for c in range(8, 8 + 31):
                    v = ws.cell(r, c).value
                    if v is None or str(v).strip() == "":
                        hours_raw.append(None)
                    else:
                        hours_raw.append(str(v).strip())

                if fio:
                    imported.append({
                        "fio": fio,
                        "tbn": tbn,
                        "hours": hours_raw,
                    })

        except Exception as e:
            messagebox.showerror("Импорт из Excel", f"Ошибка чтения данных с листа 'Табель':\n{e}", parent=self)
            return

        if not imported:
            messagebox.showinfo(
                "Импорт из Excel",
                "В выбранном файле не найдено подходящих строк для текущего объекта/периода/подразделения.",
                parent=self,
            )
            return

        # убираем дубликаты внутри файла
        uniq: Dict[tuple, Dict[str, Any]] = {}
        for rec in imported:
            key = (rec["fio"].strip().lower(), rec["tbn"].strip())
            if key not in uniq:
                uniq[key] = rec
        imported = list(uniq.values())

        # спросим режим: заменить или объединить
        if self.model_rows:
            mode = messagebox.askyesno(
                "Импорт из Excel",
                "Заменить текущий список сотрудников на импортированный?\n"
                "Да — заменить полностью\n"
                "Нет — объединить (добавить недостающих)",
                parent=self,
            )
            replace_mode = mode  # True = заменить, False = merge
        else:
            replace_mode = True

        # сохраняем текущую страницу в модель
        self._sync_visible_to_model()

        if replace_mode:
            self.model_rows.clear()

        existing = {(r["fio"].strip().lower(), r["tbn"].strip()) for r in self.model_rows}
        added = 0
        for rec in imported:
            key = (rec["fio"].strip().lower(), rec["tbn"].strip())
            if not replace_mode and key in existing:
                continue
            self.model_rows.append({
                "fio": rec["fio"],
                "tbn": rec["tbn"],
                "hours": rec.get("hours") or [None] * 31,
            })
            existing.add(key)
            added += 1

        # перерисовываем без лишней синхронизации
        self._suspend_sync = True
        try:
            self._render_page(1)
        finally:
            self._suspend_sync = False

        messagebox.showinfo(
            "Импорт из Excel",
            f"Импортировано сотрудников: {added}\n\nТеперь можно нажать «Сохранить», чтобы записать табель в БД.",
            parent=self,
        )

    def copy_from_month(self):
        if self.read_only:
            return
        """Копирование с учетом подразделения (в модель с пагинацией)"""
        addr = self.cmb_address.get().strip()
        oid = self.cmb_object_id.get().strip()
        current_dep = self.cmb_department.get().strip()

        if not addr and not oid:
            messagebox.showwarning("Копирование", "Укажите адрес и/или ID объекта для назначения.")
            return

        cy, cm = self.get_year_month()
        src_y, src_m = cy, cm - 1
        if src_m < 1:
            src_m = 12
            src_y -= 1

        dlg = CopyFromDialog(self, init_year=src_y, init_month=src_m)
        if not getattr(dlg, "result", None):
            return

        src_y = dlg.result["year"]
        src_m = dlg.result["month"]
        with_hours = dlg.result["with_hours"]
        mode = dlg.result["mode"]

        # Путь к исходному файлу С УЧЕТОМ подразделения
        src_path = self._file_path_for(src_y, src_m, addr=addr, oid=oid, department=current_dep)
        if not src_path or not src_path.exists():
            messagebox.showwarning("Копирование",
                f"Не найден файл источника для подразделения «{current_dep}»:\n{src_path.name if src_path else 'N/A'}")
            return

        try:
            wb = load_workbook(src_path, data_only=True)
            ws = self._ensure_sheet(wb)

            found = []
            for r in range(2, ws.max_row + 1):
                row_oid = (ws.cell(r, 1).value or "")
                row_addr = (ws.cell(r, 2).value or "")
                row_m = int(ws.cell(r, 3).value or 0)
                row_y = int(ws.cell(r, 4).value or 0)
                fio = str(ws.cell(r, 5).value or "").strip()
                tbn = str(ws.cell(r, 6).value or "").strip()
                row_dep = str(ws.cell(r, 7).value or "").strip()

                if row_m != src_m or row_y != src_y:
                    continue
                if oid:
                    if row_oid != oid:
                        continue
                else:
                    if row_addr != addr:
                        continue
                if current_dep != "Все" and row_dep != current_dep:
                    continue

                hrs = None
                if with_hours:
                    hrs = []
                    for c in range(8, 8 + 31):
                        v = ws.cell(r, c).value
                        hrs.append(str(v) if v else None)

                if fio:
                    found.append((fio, tbn, hrs))

            if not found:
                messagebox.showinfo("Копирование",
                    f"В источнике нет сотрудников подразделения «{current_dep}» для выбранного объекта и периода.")
                return

            # Убираем дубликаты
            uniq = {}
            for fio, tbn, hrs in found:
                key = (fio.strip().lower(), tbn.strip())
                if key not in uniq:
                    uniq[key] = (fio, tbn, hrs)
            found = list(uniq.values())

            # Сохраним правки видимой страницы
            self._sync_visible_to_model()

            added = 0
            if mode == "replace":
                self.model_rows.clear()

            existing = {(r["fio"].strip().lower(), r["tbn"].strip()) for r in self.model_rows}
            for fio, tbn, hrs in found:
                key = (fio.strip().lower(), tbn.strip())
                if mode == "merge" and key in existing:
                    continue
                self.model_rows.append({
                    "fio": fio,
                    "tbn": tbn,
                    "hours": hrs if hrs is not None else [None] * 31
                })
                existing.add(key)
                added += 1

            # Перерисовываем без синхронизации, чтобы не затирать модель пустыми Entry
            self._suspend_sync = True
            try:
                self._render_page(1 if mode == "replace" else self.current_page)
            finally:
                self._suspend_sync = False

            messagebox.showinfo("Копирование", f"Добавлено сотрудников: {added}")

        except Exception as e:
            messagebox.showerror("Копирование", f"Ошибка копирования:\n{e}")

    def _content_total_width(self, fio_px: Optional[int] = None) -> int:
        """
        Полная ширина содержимого таблицы в пикселях, с учетом всех колонок:
        ФИО, Таб.№, 31 день, Дней, Часы, Пер.день, Пер.ночь, 5/2, Удалить.
        """
        px = self.COLPX.copy()
        if fio_px is not None:
            px["fio"] = fio_px

        # fio + tbn + 31 * day + days + hours + overtime_day + overtime_night + btn52 + del
        return (
            px["fio"] +
            px["tbn"] +
            31 * px["day"] +
            px["days"] +
            px["hours"] +  # "Часы"
            px["hours"] +  # "Пер.день"
            px["hours"] +  # "Пер.ночь"
            px["btn52"] +
            px["del"]
        )

    def _auto_fit_columns(self):
        """
        Автоматически подгоняет ширину колонки ФИО под текущую ширину окна.
        ВАЖНО: при сужении окна ниже минимальной ширины таблицы
        ширина колонки ФИО больше НЕ уменьшается — включается горизонтальный скролл.
        """
        try:
            viewport = self.main_canvas.winfo_width()
        except Exception:
            viewport = 0

        # Окно еще не отрисовано – повторим позже
        if viewport <= 1:
            self.after(120, self._auto_fit_columns)
            return

        total = self._content_total_width()
        new_fio = self.COLPX["fio"]

        if total < viewport:
            # Есть запас по ширине — можно немного расширить ФИО
            surplus = viewport - total
            new_fio = min(self.MAX_FIO_PX, self.COLPX["fio"] + surplus)
        else:
            # Контент уже не помещается — НЕ уменьшаем ФИО,
            # чтобы не ломать выравнивание, просто оставляем горизонтальный скролл
            new_fio = self.COLPX["fio"]

        if int(new_fio) != int(self.COLPX["fio"]):
            self.COLPX["fio"] = int(new_fio)
            self._configure_table_columns()
            # перегрид строк, чтобы они подстроились под новую ширину
            self._regrid_rows()
            self._on_scroll_frame_configure()
        else:
            # Даже если ширина не изменилась, синхронизируем ширину шапки
            try:
                self.header_canvas.configure(width=self.main_canvas.winfo_width())
            except Exception:
                pass

    def _on_window_configure(self, _evt):
        try:
            self.after_cancel(self._fit_job)
        except Exception:
            pass
        self._fit_job = self.after(150, self._auto_fit_columns)

class MyTimesheetsPage(tk.Frame):
    """
    Реестр табелей текущего пользователя.
    """
    def __init__(self, master, app_ref: "MainApp"):
        super().__init__(master)
        self.app_ref = app_ref

        self.tree = None
        self._headers: List[Dict[str, Any]] = []

        self._build_ui()
        self._load_data()

    def _build_ui(self):
        top = tk.Frame(self)
        top.pack(fill="x", padx=8, pady=(8, 4))

        tk.Label(
            top,
            text="Мои табели",
            font=("Segoe UI", 12, "bold"),
        ).pack(side="left")

        ttk.Button(
            top,
            text="Обновить",
            command=self._load_data,
        ).pack(side="right", padx=4)

        # Таблица
        frame = tk.Frame(self)
        frame.pack(fill="both", expand=True, padx=8, pady=(4, 8))

        cols = ("year", "month", "object", "department", "updated_at")
        self.tree = ttk.Treeview(
            frame,
            columns=cols,
            show="headings",
            selectmode="browse",
        )

        self.tree.heading("year", text="Год")
        self.tree.heading("month", text="Месяц")
        self.tree.heading("object", text="Объект")
        self.tree.heading("department", text="Подразделение")
        self.tree.heading("updated_at", text="Обновлён")

        self.tree.column("year", width=60, anchor="center")
        self.tree.column("month", width=80, anchor="center")
        self.tree.column("object", width=260, anchor="w")
        self.tree.column("department", width=180, anchor="w")
        self.tree.column("updated_at", width=140, anchor="center")

        vsb = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)

        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        # Двойной клик — открыть табель
        self.tree.bind("<Double-1>", self._on_open)
        # Enter — тоже открыть
        self.tree.bind("<Return>", self._on_open)

        # Нижняя панель с подсказкой
        bottom = tk.Frame(self)
        bottom.pack(fill="x", padx=8, pady=(0, 8))
        tk.Label(
            bottom,
            text="Двойной щелчок или Enter по строке — открыть табель для редактирования.",
            font=("Segoe UI", 9),
            fg="#555",
        ).pack(side="left")

    def _load_data(self):
        self.tree.delete(*self.tree.get_children())
        self._headers.clear()

        user = getattr(self.app_ref, "current_user", None) if hasattr(self, "app_ref") else None
        user_id = (user or {}).get("id")
        if not user_id:
            messagebox.showwarning("Мои табели", "Не определён текущий пользователь.")
            return

        try:
            headers = load_user_timesheet_headers(user_id)
        except Exception as e:
            logging.exception("Ошибка загрузки списка табелей пользователя")
            messagebox.showerror("Мои табели", f"Ошибка загрузки списка табелей из БД:\n{e}")
            return

        self._headers = headers

        for h in headers:
            year = h["year"]
            month = h["month"]
            addr = h["object_addr"] or ""
            obj_id = h.get("object_id") or ""
            dep = h.get("department") or ""
            upd = h.get("updated_at")

            month_ru = month_name_ru(month) if 1 <= month <= 12 else str(month)
            obj_display = addr
            if obj_id:
                obj_display = f"[{obj_id}] {addr}"

            if isinstance(upd, datetime):
                upd_str = upd.strftime("%d.%m.%Y %H:%M")
            else:
                upd_str = str(upd or "")

            iid = str(h["id"])
            self.tree.insert(
                "",
                "end",
                iid=iid,
                values=(year, month_ru, obj_display, dep, upd_str),
            )

    def _get_selected_header(self) -> Optional[Dict[str, Any]]:
        sel = self.tree.selection()
        if not sel:
            return None
        iid = sel[0]
        try:
            hid = int(iid)
        except Exception:
            return None
        for h in self._headers:
            if int(h["id"]) == hid:
                return h
        return None

    def _on_open(self, event=None):
        h = self._get_selected_header()
        if not h:
            return

        object_id = h.get("object_id") or None
        object_addr = h.get("object_addr") or ""
        department = h.get("department") or ""
        year = int(h.get("year") or 0)
        month = int(h.get("month") or 0)

        # Табель из "Моих табелей" всегда редактируемый для владельца
        self.app_ref._show_page(
            "timesheet",
            lambda parent: TimesheetPage(
                parent,
                app_ref=self.app_ref,
                init_object_id=object_id,
                init_object_addr=object_addr,
                init_department=department,
                init_year=year,
                init_month=month,
                read_only=False,
                owner_user_id=None,
            ),
        )

class TimesheetRegistryPage(tk.Frame):
    """
    Реестр табелей всех пользователей (для руководителей/админов).
    """
    def __init__(self, master, app_ref: "MainApp"):
        super().__init__(master)
        self.app_ref = app_ref

        self.tree = None
        self._headers: List[Dict[str, Any]] = []

        self.var_year = tk.StringVar()
        self.var_month = tk.StringVar()
        self.var_dep = tk.StringVar()
        self.var_obj_addr = tk.StringVar()
        self.var_obj_id = tk.StringVar()

        self._build_ui()
        self._load_data()

    def _build_ui(self):
        top = tk.Frame(self)
        top.pack(fill="x", padx=8, pady=(8, 4))

        tk.Label(top, text="Реестр табелей", font=("Segoe UI", 12, "bold")).grid(
            row=0, column=0, columnspan=6, sticky="w", pady=(0, 4)
        )

        # Фильтры
        row_f = 1

        tk.Label(top, text="Год:").grid(row=row_f, column=0, sticky="e", padx=(0, 4))
        spn_year = tk.Spinbox(top, from_=2000, to=2100, width=6, textvariable=self.var_year)
        spn_year.grid(row=row_f, column=1, sticky="w")
        # по умолчанию текущий год
        self.var_year.set(str(datetime.now().year))

        tk.Label(top, text="Месяц:").grid(row=row_f, column=2, sticky="e", padx=(12, 4))
        cmb_month = ttk.Combobox(
            top,
            state="readonly",
            width=12,
            textvariable=self.var_month,
            values=["Все"] + [month_name_ru(i) for i in range(1, 12 + 1)],
        )
        cmb_month.grid(row=row_f, column=3, sticky="w")
        self.var_month.set("Все")

        tk.Label(top, text="Подразделение:").grid(row=row_f, column=4, sticky="e", padx=(12, 4))
        # список подразделений возьмём из TimesheetPage: employees уже загружены там,
        # но здесь мы не хотим зависеть; поэтому просто будем вводить текстом.
        ent_dep = ttk.Entry(top, width=24, textvariable=self.var_dep)
        ent_dep.grid(row=row_f, column=5, sticky="w")

        row_f += 1

        tk.Label(top, text="Объект (адрес):").grid(row=row_f, column=0, sticky="e", padx=(0, 4), pady=(4, 0))
        ent_addr = ttk.Entry(top, width=34, textvariable=self.var_obj_addr)
        ent_addr.grid(row=row_f, column=1, columnspan=2, sticky="w", pady=(4, 0))

        tk.Label(top, text="ID объекта:").grid(row=row_f, column=3, sticky="e", padx=(12, 4), pady=(4, 0))
        ent_oid = ttk.Entry(top, width=18, textvariable=self.var_obj_id)
        ent_oid.grid(row=row_f, column=4, sticky="w", pady=(4, 0))

        btns = tk.Frame(top)
        btns.grid(row=row_f, column=5, sticky="e", padx=(8, 0), pady=(4, 0))
        ttk.Button(btns, text="Применить фильтр", command=self._load_data).pack(side="left", padx=2)
        ttk.Button(btns, text="Сброс", command=self._reset_filters).pack(side="left", padx=2)
        ttk.Button(btns, text="Выгрузить в Excel", command=self._export_to_excel).pack(side="left", padx=2)

        # Таблица
        frame = tk.Frame(self)
        frame.pack(fill="both", expand=True, padx=8, pady=(4, 8))

        cols = ("year", "month", "object", "department", "user", "updated_at")
        self.tree = ttk.Treeview(
            frame,
            columns=cols,
            show="headings",
            selectmode="browse",
        )

        self.tree.heading("year", text="Год")
        self.tree.heading("month", text="Месяц")
        self.tree.heading("object", text="Объект")
        self.tree.heading("department", text="Подразделение")
        self.tree.heading("user", text="Пользователь")
        self.tree.heading("updated_at", text="Обновлён")

        self.tree.column("year", width=60, anchor="center")
        self.tree.column("month", width=80, anchor="center")
        self.tree.column("object", width=280, anchor="w")
        self.tree.column("department", width=160, anchor="w")
        self.tree.column("user", width=180, anchor="w")
        self.tree.column("updated_at", width=140, anchor="center")

        vsb = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)

        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        # Открытие табеля
        self.tree.bind("<Double-1>", self._on_open)
        self.tree.bind("<Return>", self._on_open)

        # Нижняя панель
        bottom = tk.Frame(self)
        bottom.pack(fill="x", padx=8, pady=(0, 8))
        tk.Label(
            bottom,
            text="Двойной щелчок или Enter по строке — открыть табель для просмотра/редактирования.",
            font=("Segoe UI", 9),
            fg="#555",
        ).pack(side="left")

    def _reset_filters(self):
        self.var_year.set(str(datetime.now().year))
        self.var_month.set("Все")
        self.var_dep.set("")
        self.var_obj_addr.set("")
        self.var_obj_id.set("")
        self._load_data()

    def _export_to_excel(self):
        """
        Выгружает все табели, показанные в реестре (с учётом фильтров),
        в один Excel-файл.
        Формат строк:
          Год, Месяц, Адрес, ID объекта, Подразделение, Пользователь,
          ФИО, Таб.№, D1..D31, Итого_дней, Итого_часов, Переработка_день, Переработка_ночь
        """
        if not self._headers:
            messagebox.showinfo("Экспорт в Excel", "Нет данных для выгрузки.")
            return

        from tkinter import filedialog

        # Выбор файла для сохранения
        default_name = f"Реестр_табелей_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        path = filedialog.asksaveasfilename(
            parent=self,
            title="Сохранить реестр табелей в Excel",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel файлы", "*.xlsx"), ("Все файлы", "*.*")],
        )
        if not path:
            return

        try:
            # Создаём новую рабочую книгу
            wb = Workbook()
            ws = wb.active
            ws.title = "Реестр табелей"

            # Заголовок
            header_row = [
                "Год",
                "Месяц",
                "Адрес",
                "ID объекта",
                "Подразделение",
                "Пользователь",
                "ФИО",
                "Табельный №",
            ] + [f"{i}" for i in range(1, 32)] + [
                "Итого_дней",
                "Итого_часов",
                "Переработка_день",
                "Переработка_ночь",
            ]
            ws.append(header_row)

            # Немного ширин столбцов
            ws.column_dimensions["A"].width = 6   # Год
            ws.column_dimensions["B"].width = 10  # Месяц
            ws.column_dimensions["C"].width = 40  # Адрес
            ws.column_dimensions["D"].width = 14  # ID объекта
            ws.column_dimensions["E"].width = 22  # Подразделение
            ws.column_dimensions["F"].width = 22  # Пользователь
            ws.column_dimensions["G"].width = 28  # ФИО
            ws.column_dimensions["H"].width = 12  # Таб.№
            for col_idx in range(9, 9 + 31):      # дни
                ws.column_dimensions[get_column_letter(col_idx)].width = 6
            # Итоги
            base = 9 + 31
            ws.column_dimensions[get_column_letter(base)].width = 10   # Итого_дней
            ws.column_dimensions[get_column_letter(base + 1)].width = 14  # Итого_часов
            ws.column_dimensions[get_column_letter(base + 2)].width = 16  # Переработка_день
            ws.column_dimensions[get_column_letter(base + 3)].width = 16  # Переработка_ночь

            # Заполняем данные
            total_rows = 0
            for h in self._headers:
                header_id = int(h["id"])
                year = int(h["year"])
                month = int(h["month"])
                addr = h.get("object_addr") or ""
                obj_id = h.get("object_id") or ""
                dep = h.get("department") or ""
                user_display = h.get("full_name") or h.get("username") or ""

                rows = load_timesheet_rows_by_header_id(header_id)

                for row in rows:
                    fio = row["fio"]
                    tbn = row["tbn"]
                    hours_raw = row.get("hours_raw") or [None] * 31
                    total_days = row.get("total_days")
                    total_hours = row.get("total_hours")
                    ot_day = row.get("overtime_day")
                    ot_night = row.get("overtime_night")

                    excel_row = [
                        year,
                        month,
                        addr,
                        obj_id,
                        dep,
                        user_display,
                        fio,
                        tbn,
                    ]

                    # 1..31 дни (как в БД/табеле — строковые значения)
                    for v in hours_raw:
                        excel_row.append(v if v is not None else None)

                    excel_row.append(total_days if total_days is not None else None)
                    excel_row.append(total_hours if total_hours is not None else None)
                    excel_row.append(ot_day if ot_day is not None else None)
                    excel_row.append(ot_night if ot_night is not None else None)

                    ws.append(excel_row)
                    total_rows += 1

            wb.save(path)
            messagebox.showinfo(
                "Экспорт в Excel",
                f"Выгрузка завершена.\nФайл: {path}\nСтрок табеля: {total_rows}",
                parent=self,
            )
        except Exception as e:
            logging.exception("Ошибка экспорта реестра табелей в Excel")
            messagebox.showerror("Экспорт в Excel", f"Ошибка при выгрузке:\n{e}", parent=self)

    def _load_data(self):
        self.tree.delete(*self.tree.get_children())
        self._headers.clear()

        # Определяем фильтры
        year = None
        try:
            y = int(self.var_year.get().strip())
            if 2000 <= y <= 2100:
                year = y
        except Exception:
            pass

        month = None
        m_name = (self.var_month.get() or "").strip()
        if m_name and m_name != "Все":
            # преобразуем русское имя в номер
            for i in range(1, 13):
                if month_name_ru(i) == m_name:
                    month = i
                    break

        dep = self.var_dep.get().strip()
        if not dep:
            dep = None

        addr_sub = self.var_obj_addr.get().strip() or None
        oid_sub = self.var_obj_id.get().strip() or None

        try:
            headers = load_all_timesheet_headers(
                year=year,
                month=month,
                department=dep,
                object_addr_substr=addr_sub,
                object_id_substr=oid_sub,
            )
        except Exception as e:
            logging.exception("Ошибка загрузки реестра табелей")
            messagebox.showerror("Реестр табелей", f"Ошибка загрузки реестра из БД:\n{e}")
            return

        self._headers = headers

        for h in headers:
            year = h["year"]
            month = h["month"]
            addr = h["object_addr"] or ""
            obj_id = h.get("object_id") or ""
            dep = h.get("department") or ""
            upd = h.get("updated_at")
            full_name = h.get("full_name") or h.get("username") or ""

            month_ru = month_name_ru(month) if 1 <= month <= 12 else str(month)
            obj_display = addr
            if obj_id:
                obj_display = f"[{obj_id}] {addr}"

            if isinstance(upd, datetime):
                upd_str = upd.strftime("%d.%m.%Y %H:%M")
            else:
                upd_str = str(upd or "")

            iid = str(h["id"])
            self.tree.insert(
                "",
                "end",
                iid=iid,
                values=(year, month_ru, obj_display, dep, full_name, upd_str),
            )

    def _get_selected_header(self) -> Optional[Dict[str, Any]]:
        sel = self.tree.selection()
        if not sel:
            return None
        iid = sel[0]
        try:
            hid = int(iid)
        except Exception:
            return None
        for h in self._headers:
            if int(h["id"]) == hid:
                return h
        return None

    def _on_open(self, event=None):
        h = self._get_selected_header()
        if not h:
            return

        object_id = h.get("object_id") or None
        object_addr = h.get("object_addr") or ""
        department = h.get("department") or ""
        year = int(h.get("year") or 0)
        month = int(h.get("month") or 0)

        owner_user_id = h.get("user_id")  # владелец табеля в БД

        # роль текущего пользователя
        role = (self.app_ref.current_user or {}).get("role") or "specialist"
        # только admin может редактировать, остальные — только просмотр
        read_only = (role != "admin")

        self.app_ref._show_page(
            "timesheet",
            lambda parent: TimesheetPage(
                parent,
                app_ref=self.app_ref,
                init_object_id=object_id,
                init_object_addr=object_addr,
                init_department=department,
                init_year=year,
                init_month=month,
                read_only=read_only,
                owner_user_id=owner_user_id,   # <-- передаём id автора
            ),
        )

class MainApp(tk.Tk):
    def __init__(self, current_user: Optional[Dict[str, Any]] = None):
        super().__init__()
        if meals_module:
            meals_module.set_db_pool(db_connection_pool)
        if SpecialOrders:
            SpecialOrders.set_db_pool(db_connection_pool)
        if objects:
            objects.set_db_pool(db_connection_pool)

        self._menu_timesheets = None
        self._menu_timesheets_registry_index = None
        self.current_user: Dict[str, Any] = current_user or {}
        self.is_authenticated: bool = bool(current_user)
        self.title(APP_NAME)
        self.geometry("1024x720")
        self.minsize(980, 640)
        self.resizable(True, True)

        ensure_config()
        self._pages: Dict[str, tk.Widget] = {}
        self._menubar = None
        self._menu_meals = None
        self._menu_meals_planning_index = None
        self._menu_meals_settings_index = None
        self._menu_transport = None
        self._menu_transport_planning_index = None
        self._menu_transport_registry_index = None
        self._menu_objects = None
        self._menu_objects_create_index = None
        self._menu_settings_index = None

        menubar = tk.Menu(self)

        menubar.add_command(label="Главная", command=self.show_home)

        m_ts = tk.Menu(menubar, tearoff=0)
        m_ts.add_command(
            label="Создать",
            command=lambda: self._show_page(
                "timesheet",
                lambda parent: TimesheetPage(parent, app_ref=self),
            ),
        )
        m_ts.add_command(
            label="Мои табели",
            command=lambda: self._show_page(
                "my_timesheets",
                lambda parent: MyTimesheetsPage(parent, app_ref=self),
            ),
        )
        m_ts.add_command(
            label="Реестр табелей",
            command=lambda: self._show_page(
                "timesheet_registry",
                lambda parent: TimesheetRegistryPage(parent, app_ref=self),
            ),
        )

        menubar.add_cascade(label="Объектный табель", menu=m_ts)
        self._menu_timesheets = m_ts
        self._menu_timesheets_registry_index = 2  # 0 - Создать, 1 - Мои табели, 2 - Реестр табелей

        # Автотранспорт
        m_transport = tk.Menu(menubar, tearoff=0)
        self._menu_transport = m_transport
        if SpecialOrders and hasattr(SpecialOrders, "create_page"):
            m_transport.add_command(
                label="📝 Создать заявку",
                command=lambda: self._show_page("transport", lambda parent: SpecialOrders.create_page(parent)),
            )
        else:
            m_transport.add_command(label="📝 Создать заявку", command=self.run_special_orders_exe)

        self._menu_transport_planning_index = None
        if SpecialOrders and hasattr(SpecialOrders, "create_planning_page"):
            self._menu_transport_planning_index = 1
            m_transport.add_command(
                label="🚛Планирование транспорта",
                command=lambda: self._show_page(
                    "planning", lambda parent: SpecialOrders.create_planning_page(parent)
                ),
            )

        self._menu_transport_registry_index = None
        if SpecialOrders and hasattr(SpecialOrders, "create_transport_registry_page"):
            idx = m_transport.index("end")
            self._menu_transport_registry_index = idx + 1 if idx is not None else 0
            m_transport.add_command(
                label="🚘Реестр транспорта",
                command=lambda: self._show_page(
                    "transport_registry",
                    lambda parent: SpecialOrders.create_transport_registry_page(parent),
                ),
            )
        menubar.add_cascade(label="Автотранспорт", menu=m_transport)

        # Питание
        logging.debug(f"Строим меню Питание. meals_module={meals_module}")
        m_meals = tk.Menu(menubar, tearoff=0)
        self._menu_meals = m_meals

        if meals_module and hasattr(meals_module, "create_meals_order_page"):
            m_meals.add_command(
                label="📝 Создать заявку",
                command=lambda: self._show_page(
                    "meals_order",
                    lambda parent: meals_module.create_meals_order_page(parent, app_ref=self),
                ),
            )
        else:
            m_meals.add_command(label="📝 Создать заявку", command=self.run_meals_exe)

        if meals_module and hasattr(meals_module, "create_my_meals_orders_page"):
            m_meals.add_command(
                label="📄 Мои заявки",
                command=lambda: self._show_page(
                    "my_meals_orders",
                    lambda parent: meals_module.create_my_meals_orders_page(parent, app_ref=self),
                ),
            )
        else:
            m_meals.add_command(label="📝 Создать заявку", command=self.run_meals_exe)

        self._menu_meals_planning_index = None
        if meals_module and hasattr(meals_module, "create_meals_planning_page"):
            self._menu_meals_planning_index = 1
            m_meals.add_command(
                label="🍽️Планирование питания",
                command=lambda: self._show_page(
                    "meals_planning",
                    lambda parent: meals_module.create_meals_planning_page(parent),
                ),
            )

        # Настройки питания (видны только admin — управляем в _apply_role_visibility)
        self._menu_meals_settings_index = None
        if meals_module and hasattr(meals_module, "create_meals_settings_page"):
            idx = m_meals.index("end")
            self._menu_meals_settings_index = idx + 1 if idx is not None else 0
            m_meals.add_command(
                label="⚙ Настройки питания",
                command=lambda: self._show_page(
                    "meals_settings",
                    lambda parent: meals_module.create_meals_settings_page(
                        parent, (self.current_user or {}).get("role") or "specialist"
                    ),
                ),
            )

        m_meals.add_separator()
        m_meals.add_command(label="📂 Открыть папку заявок", command=self.open_meals_folder)
        menubar.add_cascade(label="Питание", menu=m_meals)

        # Объекты
        m_objects = tk.Menu(menubar, tearoff=0)
        self._menu_objects = m_objects

        if objects and hasattr(objects, "ObjectCreatePage"):
            m_objects.add_command(
                label="Создать",
                command=lambda: self._show_page(
                    "object_create",
                    lambda parent: objects.ObjectCreatePage(parent, app_ref=self),
                ),
            )
        else:
            m_objects.add_command(
                label="Создать",
                command=lambda: messagebox.showwarning(
                    "Объекты", "Модуль objects.py не найден или некорректен."
                ),
            )
        self._menu_objects_create_index = 0  # первый пункт

        if objects and hasattr(objects, "ObjectsRegistryPage"):
            m_objects.add_command(
                label="Реестр",
                command=lambda: self._show_page(
                    "objects_registry",
                    lambda parent: objects.ObjectsRegistryPage(parent, app_ref=self),
                ),
            )
        else:
            m_objects.add_command(
                label="Реестр",
                command=lambda: messagebox.showwarning(
                    "Объекты", "Модуль objects.py не найден или некорректен."
                ),
            )

        menubar.add_cascade(label="Объекты", menu=m_objects)

        # Аналитика
        m_analytics = tk.Menu(menubar, tearoff=0)
        m_analytics.add_command(label="Экспорт свода (XLSX/CSV)", command=self.summary_export)
        menubar.add_cascade(label="Аналитика", menu=m_analytics)

        # Инструменты
        m_tools = tk.Menu(menubar, tearoff=0)
        if timesheet_transformer and hasattr(timesheet_transformer, "open_converter"):
            m_tools.add_command(
                label="Конвертер табеля (1С)",
                command=lambda: timesheet_transformer.open_converter(self),
            )
        else:
            m_tools.add_command(label="Конвертер табеля (1С)", command=self.run_converter_exe)
        if BudgetAnalyzer and hasattr(BudgetAnalyzer, "create_page"):
            m_tools.add_command(
                label="Анализ смет",
                command=lambda: self._show_page(
                    "budget", lambda parent: BudgetAnalyzer.create_page(parent)
                ),
            )
        else:
            m_tools.add_command(
                label="Анализ смет",
                command=lambda: messagebox.showwarning(
                    "Анализ смет", "Модуль BudgetAnalyzer.py не найден."
                ),
            )
        menubar.add_cascade(label="Инструменты", menu=m_tools)

        # Настройки (общие)
        m_settings = tk.Menu(menubar, tearoff=0)
        m_settings.add_command(
            label="Открыть настройки",
            command=lambda: Settings.open_settings_window(self)
            if Settings
            else messagebox.showwarning(
                "Настройки", "Модуль settings_manager не найден."
            ),
        )
        menubar.add_cascade(label="Настройки", menu=m_settings)
        self._menu_settings_index = menubar.index("end")

        self.config(menu=menubar)
        self._menubar = menubar

        self._set_user(None)

        self.header = tk.Frame(self)
        self.header.pack(fill="x", padx=12, pady=(10, 4))

        self.lbl_header_title = tk.Label(
            self.header,
            text="Управление строительством",
            font=("Segoe UI", 16, "bold"),
        )
        self.lbl_header_title.pack(side="left")

        self.lbl_header_hint = tk.Label(
            self.header,
            text="Выберите раздел в верхнем меню",
            font=("Segoe UI", 10),
            fg="#555",
        )
        self.lbl_header_hint.pack(side="right")

        self.content = tk.Frame(self, bg="#f7f7f7")
        self.content.pack(fill="both", expand=True)

        footer = tk.Frame(self)
        footer.pack(fill="x", padx=12, pady=(0, 10))
        tk.Label(
            footer,
            text="Разработал Алексей Зезюкин, 2025",
            font=("Segoe UI", 8),
            fg="#666",
        ).pack(side="right")

        self.show_login()

    # --- управление пользователем / страницами ---

    def _set_user(self, user: Optional[Dict[str, Any]]):
        self.current_user = user or {}
        self.is_authenticated = bool(user)
        caption = ""
        if user:
            fn = user.get("full_name") or ""
            un = user.get("username") or ""
            caption = f" — {fn or un}"
        self.title(APP_NAME + caption)
        self._apply_role_visibility()

    def _set_header(self, title: str, hint: str = ""):
        """Обновляет заголовок над содержимым."""
        if hasattr(self, "lbl_header_title"):
            self.lbl_header_title.config(text=title)
        if hasattr(self, "lbl_header_hint"):
            self.lbl_header_hint.config(text=hint or "")

    def show_login(self):
        self._show_page("login", lambda parent: LoginPage(parent, app_ref=self))

    def on_login_success(self, user: Dict[str, Any]):
        logging.debug(f"MainApp.on_login_success: {user!r}")
        self._set_user(user)
        self.show_home()

    def _show_page(self, key: str, builder):
        if not self.is_authenticated and key not in ("login",):
            messagebox.showwarning(
                "Доступ ограничен",
                "Для доступа к разделу необходимо войти в систему.",
                parent=self,
            )
            self.show_login()
            return

        # --- устанавливаем заголовок в зависимости от раздела ---
        if key == "home":
            self._set_header("Управление строительством", "Выберите раздел в верхнем меню")
        elif key == "timesheet":
            self._set_header("Объектный табель", "")
        elif key == "my_timesheets":
            self._set_header("Мои табели", "")
        elif key == "timesheet_registry":
            self._set_header("Реестр табелей", "")
        elif key == "transport":
            self._set_header("Заявка на спецтехнику", "")
        elif key == "planning":
            self._set_header("Планирование транспорта", "")
        elif key == "transport_registry":
            self._set_header("Реестр транспорта", "")
        elif key == "meals_order":
            self._set_header("Заказ питания", "")
        elif key == "my_meals_orders":
            self._set_header("Мои заявки на питание", "")
        elif key == "meals_planning":
            self._set_header("Планирование питания", "")
        elif key == "meals_settings":
            self._set_header("Настройки питания", "")
        elif key == "object_create":
            self._set_header("Объекты — создание/редактирование", "")
        elif key == "objects_registry":
            self._set_header("Объекты — реестр", "")
        elif key == "budget":
            self._set_header("Анализ смет", "")
        elif key == "login":
            self._set_header("Управление строительством", "Вход в систему")
        else:
            # по умолчанию — просто ключ
            self._set_header(key, "")

        for w in self.content.winfo_children():
            try:
                w.destroy()
            except Exception:
                pass
        try:
            page = builder(self.content)
        except Exception as e:
            traceback.print_exc()
            messagebox.showerror("Ошибка", f"Не удалось открыть страницу:\n{e}")
            if self.is_authenticated:
                self.show_home()
            else:
                self.show_login()
            return

        if isinstance(page, tk.Widget) and page.master is self.content:
            try:
                page.pack_forget()
            except Exception:
                pass
        try:
            page.pack(fill="both", expand=True)
        except Exception:
            pass
        self._pages[key] = page

    def show_home(self):
        # Главная страница — основной заголовок
        self._set_header("Управление строительством", "Выберите раздел в верхнем меню")
        self._show_page("home", lambda parent: HomePage(parent))


    def _apply_role_visibility(self):
        role = (self.current_user or {}).get("role") or "specialist"

        # Реестр табелей — только для руководителей и админов
        if self._menu_timesheets is not None and self._menu_timesheets_registry_index is not None:
            try:
                # разрешаем для ролей admin, head, planner (можно сузить список)
                st = "normal" if role in ("admin", "manager") else "disabled"
                self._menu_timesheets.entryconfig(self._menu_timesheets_registry_index, state=st)
            except Exception:
                pass

        # Питание
        if self._menu_meals is not None:
            try:
                self._menu_meals.entryconfig(0, state="normal")  # Создать
                if self._menu_meals_planning_index is not None:
                    st = "normal" if role in ("admin", "planner", "manager") else "disabled"
                    self._menu_meals.entryconfig(self._menu_meals_planning_index, state=st)
                if self._menu_meals_settings_index is not None:
                    st = "normal" if role == "admin" else "disabled"
                    self._menu_meals.entryconfig(self._menu_meals_settings_index, state=st)
            except Exception:
                pass

        # Автотранспорт
        if self._menu_transport is not None:
            try:
                self._menu_transport.entryconfig(0, state="normal")
                if self._menu_transport_planning_index is not None:
                    st = "normal" if role in ("admin", "planner", "manager") else "disabled"
                    self._menu_transport.entryconfig(self._menu_transport_planning_index, state=st)
                if self._menu_transport_registry_index is not None:
                    st = "normal" if role in ("admin", "planner", "manager") else "disabled"
                    self._menu_transport.entryconfig(self._menu_transport_registry_index, state=st)
            except Exception:
                pass

        # Объекты: пункт "Создать" 
        if self._menu_objects is not None and self._menu_objects_create_index is not None:
            try:
                st = "normal" if role in ("admin", "manager") else "disabled"
                self._menu_objects.entryconfig(self._menu_objects_create_index, state=st)
            except Exception:
                pass


        # Верхнее "Настройки" только для admin (отключаем для остальных)
        if self._menubar is not None and self._menu_settings_index is not None:
            try:
                state = "normal" if role == "admin" else "disabled"
                self._menubar.entryconfig(self._menu_settings_index, state=state)
            except Exception:
                pass

    # --- Папки / внешние EXE ---

    def open_meals_folder(self):
        try:
            meals_dir = exe_dir() / "Заявки_питание"
            meals_dir.mkdir(parents=True, exist_ok=True)
            os.startfile(meals_dir)
        except Exception as e:
            messagebox.showerror("Папка заявок", f"Не удалось открыть папку:\n{e}")

    def run_meals_exe(self):
        try:
            p = exe_dir() / "meals_module.exe"
            if not p.exists():
                messagebox.showwarning("Заказ питания", "Не найден meals_module.exe рядом с программой.")
                return
            subprocess.Popen([str(p)], shell=False)
        except Exception as e:
            messagebox.showerror("Заказ питания", f"Не удалось запустить модуль:\n{e}")

    def open_orders_folder(self):
        try:
            orders_dir = exe_dir() / "Заявки_спецтехники"
            orders_dir.mkdir(parents=True, exist_ok=True)
            os.startfile(orders_dir)
        except Exception as e:
            messagebox.showerror("Папка заявок", f"Не удалось открыть папку:\n{e}")

    def summary_export(self):
        pwd = simpledialog.askstring("Сводный экспорт", "Введите пароль:", show="*", parent=self)
        if pwd is None:
            return
        if pwd != get_export_password_from_config():
            messagebox.showerror("Сводный экспорт", "Неверный пароль.")
            return

    def run_special_orders_exe(self):
        try:
            p = exe_dir() / "SpecialOrders.exe"
            if not p.exists():
                messagebox.showwarning("Заказ спецтехники", "Не найден SpecialOrders.exe рядом с программой.")
                return
            subprocess.Popen([str(p)], shell=False)
        except Exception as e:
            messagebox.showerror("Заказ спецтехники", f"Не удалось запустить модуль:\n{e}")

    def run_converter_exe(self):
        try:
            p = exe_dir() / "TabelConverter.exe"
            if not p.exists():
                messagebox.showwarning("Конвертер", "Не найден TabelConverter.exe рядом с программой.")
                return
            subprocess.Popen([str(p)], shell=False)
        except Exception as e:
            messagebox.showerror("Конвертер", f"Не удалось запустить конвертер:\n{e}")

    def destroy(self):
        """Переопределяем стандартный метод destroy для корректного завершения работы."""
        logging.info("Приложение закрывается. Закрываем пул соединений.")
        close_db_pool()  # Закрываем все соединения в пуле
        super().destroy() # Вызываем оригинальный метод destroy

logging.debug("Модуль main_app импортирован, готов к запуску.")

if __name__ == "__main__":
    logging.debug("Старт приложения...")

    # Инициализируем конфиг и пул ДО создания окна
    try:
        # 1. Убеждаемся, что файл конфигурации существует.
        # Эта функция уже есть в вашем коде.
        ensure_config()
        initialize_db_pool()

    except Exception as e:
        logging.critical("Приложение не может быть запущено из-за ошибки инициализации.", exc_info=True)
        root = tk.Tk()
        root.withdraw() # Прячем пустое окно
        messagebox.showerror(
            "Критическая ошибка",
            f"Не удалось инициализировать приложение.\n\n"
            f"Ошибка: {e}\n\n"
            "Проверьте настройки в файле settings.json (DATABASE_URL) и доступность базы данных.",
            parent=root
        )
        root.destroy()
        sys.exit(1)

    # 3. Только если всё прошло успешно, создаём и запускаем главное окно приложения.
    logging.debug("Инициализация успешна. Запускаем главный цикл приложения.")
    app = MainApp()
    app.mainloop()
