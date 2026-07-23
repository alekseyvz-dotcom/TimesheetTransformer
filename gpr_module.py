# gpr_module.py  — профессиональный модуль ГПР v3 (bugfix + perf)
from __future__ import annotations

import sys
import logging
import calendar
from datetime import datetime, date, timedelta
from typing import Any, Dict, List, Optional, Tuple, Set
from pathlib import Path
from gpr_planning_module import GprPlanningPanel

import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog

import psycopg2
from psycopg2.extras import RealDictCursor, execute_values

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logger = logging.getLogger(__name__)

# ═══════════════════════════════════════════════════════════════
#  COLORS / THEME
# ═══════════════════════════════════════════════════════════════
C = {
    "bg": "#f0f2f5",
    "panel": "#ffffff",
    "accent": "#1565c0",
    "accent_light": "#e3f2fd",
    "success": "#2e7d32",
    "warning": "#ed6c02",
    "error": "#d32f2f",
    "border": "#dde1e7",
    "text": "#1a1a2e",
    "text2": "#555",
    "text3": "#999",
    "btn_bg": "#1565c0",
    "btn_fg": "#ffffff",
}

STATUS_COLORS = {
    "planned": ("#90caf9", "#1565c0", "Запланировано"),
    "in_progress": ("#ffcc80", "#e65100", "В работе"),
    "done": ("#a5d6a7", "#1b5e20", "Выполнено"),
    "paused": ("#fff176", "#f9a825", "Приостановлено"),
    "canceled": ("#ef9a9a", "#b71c1c", "Отменено"),
}

STATUS_LIST = ["planned", "in_progress", "done", "paused", "canceled"]
STATUS_LABELS = {k: v[2] for k, v in STATUS_COLORS.items()}

# Обратное отображение: label → code
_STATUS_LABEL_TO_CODE = {v[2]: k for k, v in STATUS_COLORS.items()}


# ═══════════════════════════════════════════════════════════════
#  DB POOL  (с context-manager для безопасности)
# ═══════════════════════════════════════════════════════════════
db_connection_pool = None


def set_db_pool(pool):
    global db_connection_pool
    db_connection_pool = pool


class _DBConn:
    """Context-manager: гарантированный возврат соединения в пул."""

    def __init__(self):
        self.conn = None

    def __enter__(self):
        if not db_connection_pool:
            raise RuntimeError("DB pool not set (gpr_module.set_db_pool)")
        self.conn = db_connection_pool.getconn()
        return self.conn

    def __exit__(self, exc_type, exc_val, exc_tb):
        if db_connection_pool and self.conn:
            if exc_type is not None:
                try:
                    self.conn.rollback()
                except Exception:
                    pass
            db_connection_pool.putconn(self.conn)
            self.conn = None
        return False  # не подавляем исключения

# ═══════════════════════════════════════════════════════════════
#  ОБРАТНАЯ СОВМЕСТИМОСТЬ: _conn / _release
#  Используются в gpr_task_dialog.py, gpr_dictionaries.py и др.
# ═══════════════════════════════════════════════════════════════
def _conn():
    """Получить соединение из пула (legacy API).
    
    ВНИМАНИЕ: вызывающий код ОБЯЗАН вызвать _release(conn)
    в finally-блоке. Для нового кода используйте _DBConn().
    """
    if not db_connection_pool:
        raise RuntimeError("DB pool not set (gpr_module.set_db_pool)")
    return db_connection_pool.getconn()


def _release(conn):
    """Вернуть соединение в пул (legacy API)."""
    if db_connection_pool and conn:
        try:
            db_connection_pool.putconn(conn)
        except Exception:
            logger.exception("Error releasing DB connection")

# ═══════════════════════════════════════════════════════════════
#  UTILITIES
# ═══════════════════════════════════════════════════════════════
def _parse_date(s: str) -> date:
    """Парсит дату из строки дд.мм.гггг"""
    s = s.strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Неверный формат даты: '{s}' (ожидается дд.мм.гггг)")


def _to_date(d) -> Optional[date]:
    """Безопасное приведение к date."""
    if isinstance(d, date) and not isinstance(d, datetime):
        return d
    if isinstance(d, datetime):
        return d.date()
    if isinstance(d, str) and d.strip():
        try:
            return _parse_date(d)
        except ValueError:
            try:
                return datetime.fromisoformat(d).date()
            except Exception:
                return None
    return None


def _fmt_date(d) -> str:
    dt = _to_date(d) if not isinstance(d, date) else d
    if isinstance(dt, date):
        return dt.strftime("%d.%m.%Y")
    return str(d or "")


def _today() -> date:
    return datetime.now().date()


def _quarter_range() -> Tuple[date, date]:
    t = _today()
    q_start_month = ((t.month - 1) // 3) * 3 + 1
    d0 = date(t.year, q_start_month, 1)
    end_month = q_start_month + 2
    d1 = date(t.year, end_month, calendar.monthrange(t.year, end_month)[1])
    return d0, d1


def _safe_float(v) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(str(v).replace(",", ".").strip())
    except (ValueError, TypeError):
        return None


def _fmt_qty(v) -> str:
    f = _safe_float(v)
    if f is None:
        return ""
    return f"{f:.3f}".rstrip("0").rstrip(".")

def _overlap_days(a0: date, a1: date, b0: date, b1: date) -> int:
    """
    Количество календарных дней пересечения двух периодов включительно.
    """
    if not a0 or not a1 or not b0 or not b1:
        return 0

    s = max(a0, b0)
    f = min(a1, b1)

    if f < s:
        return 0

    return (f - s).days + 1


def _calc_plan_qty_for_period(
    plan_qty: Any,
    plan_start: Any,
    plan_finish: Any,
    period_from: date,
    period_to: date,
) -> Optional[float]:
    """
    Расчёт планового объёма на период.

    Логика:
    - если у работы есть общий плановый объём;
    - и работа пересекается с выбранным периодом;
    - то объём распределяется равномерно по календарным дням работы.
    """
    qty = _safe_float(plan_qty)
    if qty is None:
        return None

    ds = _to_date(plan_start)
    df = _to_date(plan_finish)

    if not ds or not df or df < ds:
        return None

    total_days = (df - ds).days + 1
    if total_days <= 0:
        return None

    days_in_period = _overlap_days(ds, df, period_from, period_to)
    if days_in_period <= 0:
        return 0.0

    return qty * days_in_period / total_days

def _mouse_delta(event) -> int:
    """Кроссплатформенный расчёт направления колёсика мыши."""
    if event.delta:
        # Windows/macOS
        return -1 if event.delta > 0 else 1
    # Linux: event.num == 4 (up) / 5 (down)
    if hasattr(event, 'num'):
        return -1 if event.num == 4 else 1
    return 0

class GprExcelImportService:
    """
    Импорт задач ГПР из Excel-листа 'ГПР'.

    Поддерживает строки:
    - task   : обычная работа
    - group  : если в колонке 'Тип работ' указано 'ГРУППА'
    - title  : если в колонке 'Тип работ' указано 'ТИТУЛ'

    Импорт не пишет в БД — только формирует список задач
    для подстановки в редактор.
    """

    REQUIRED_HEADERS = {
        "type": "Тип работ",
        "name": "Вид работ",
        "uom": "Ед. изм.",
        "qty": "Объём план",
        "start": "Начало",
        "finish": "Окончание",
        "status": "Статус",
    }

    @staticmethod
    def _norm(s: Any) -> str:
        return " ".join(str(s or "").strip().lower().split())

    @staticmethod
    def _cell_to_date(value) -> Optional[date]:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str) and value.strip():
            return _to_date(value)
        return None

    @staticmethod
    def _build_work_type_map(
        work_types: List[Dict[str, Any]]
    ) -> Dict[str, Dict[str, Any]]:
        out: Dict[str, Dict[str, Any]] = {}
        for w in work_types:
            name = GprExcelImportService._norm(w.get("name"))
            code = GprExcelImportService._norm(w.get("code"))
            if name:
                out[name] = w
            if code:
                out[code] = w
        return out

    @staticmethod
    def _build_uom_map(
        uoms: List[Dict[str, Any]]
    ) -> Dict[str, Dict[str, Any]]:
        out: Dict[str, Dict[str, Any]] = {}
        for u in uoms:
            code = GprExcelImportService._norm(u.get("code"))
            name = GprExcelImportService._norm(u.get("name"))
            pair1 = GprExcelImportService._norm(
                f"{u.get('code', '')} — {u.get('name', '')}"
            )
            pair2 = GprExcelImportService._norm(
                f"{u.get('code', '')}-{u.get('name', '')}"
            )

            if code:
                out[code] = u
            if name:
                out[name] = u
            if pair1:
                out[pair1] = u
            if pair2:
                out[pair2] = u
        return out

    @staticmethod
    def _build_status_map() -> Dict[str, str]:
        out: Dict[str, str] = {}
        for code in STATUS_LIST:
            out[GprExcelImportService._norm(code)] = code
            out[GprExcelImportService._norm(STATUS_LABELS.get(code, code))] = code
        return out

    @staticmethod
    def _find_sheet(wb):
        preferred = ["ГПР", "гпр", "GPR", "Гпр"]
        for name in preferred:
            if name in wb.sheetnames:
                return wb[name]
        if wb.sheetnames:
            return wb[wb.sheetnames[0]]
        raise ValueError("В книге Excel нет листов")

    @staticmethod
    def _find_header_row(ws) -> Tuple[int, Dict[str, int]]:
        """
        Ищет строку заголовков в первых 30 строках.
        Возвращает:
            (номер_строки_заголовка, маппинг_ключ->номер_колонки)
        """
        max_scan_rows = min(ws.max_row, 30)

        for row_idx in range(1, max_scan_rows + 1):
            row_values = [
                ws.cell(row=row_idx, column=col_idx).value
                for col_idx in range(1, ws.max_column + 1)
            ]

            normalized: Dict[str, int] = {}
            for col_idx, val in enumerate(row_values, start=1):
                key = GprExcelImportService._norm(val)
                if key:
                    normalized[key] = col_idx

            found: Dict[str, int] = {}
            ok = True
            for key, title in GprExcelImportService.REQUIRED_HEADERS.items():
                col_idx = normalized.get(GprExcelImportService._norm(title))
                if not col_idx:
                    ok = False
                    break
                found[key] = col_idx

            if ok:
                return row_idx, found

        raise ValueError(
            "Не найдена строка заголовков. "
            "Ожидаются колонки: "
            + ", ".join(GprExcelImportService.REQUIRED_HEADERS.values())
        )

    @staticmethod
    def import_tasks_from_excel(
        path: str,
        work_types: List[Dict[str, Any]],
        uoms: List[Dict[str, Any]],
    ) -> Dict[str, Any]:
        if not HAS_OPENPYXL:
            raise RuntimeError(
                "Для импорта необходима библиотека openpyxl"
            )

        wb = load_workbook(path, data_only=True)
        ws = GprExcelImportService._find_sheet(wb)

        header_row, cols = GprExcelImportService._find_header_row(ws)

        if not work_types:
            raise ValueError("Справочник типов работ пуст")

        wt_map = GprExcelImportService._build_work_type_map(work_types)
        uom_map = GprExcelImportService._build_uom_map(uoms)
        status_map = GprExcelImportService._build_status_map()

        default_work_type_id = int(work_types[0]["id"])

        tasks: List[Dict[str, Any]] = []
        errors: List[str] = []
        skipped_empty = 0

        for row_idx in range(header_row + 1, ws.max_row + 1):
            raw_type = ws.cell(row=row_idx, column=cols["type"]).value
            raw_name = ws.cell(row=row_idx, column=cols["name"]).value
            raw_uom = ws.cell(row=row_idx, column=cols["uom"]).value
            raw_qty = ws.cell(row=row_idx, column=cols["qty"]).value
            raw_start = ws.cell(row=row_idx, column=cols["start"]).value
            raw_finish = ws.cell(row=row_idx, column=cols["finish"]).value
            raw_status = ws.cell(row=row_idx, column=cols["status"]).value

            row_values = [
                raw_type, raw_name, raw_uom,
                raw_qty, raw_start, raw_finish, raw_status
            ]
            if all(not str(v or "").strip() for v in row_values):
                skipped_empty += 1
                continue

            type_text = str(raw_type or "").strip()
            name_text = str(raw_name or "").strip()

            if not name_text:
                errors.append(
                    f"Строка {row_idx}: пустое поле 'Вид работ'"
                )
                continue

            type_norm = GprExcelImportService._norm(type_text)

            # ── GROUP ─────────────────────────────────────────
            if type_norm in ("группа", "group"):
                tasks.append(
                    {
                        "id": None,
                        "parent_id": None,
                        "row_kind": "group",
                        "work_type_id": default_work_type_id,
                        "work_type_name": "",
                        "work_item_id": None,
                        "labor_norm_id": None,
                        "labor_hours_per_unit": None,
                        "productivity_factor": None,
                        "name": name_text,
                        "uom_code": None,
                        "plan_qty": None,
                        "plan_start": _today(),
                        "plan_finish": _today(),
                        "status": "planned",
                        "is_milestone": False,
                        "sort_order": len(tasks) * 10,
                    }
                )
                continue

            # ── TITLE ─────────────────────────────────────────
            if type_norm in ("титул", "title"):
                tasks.append(
                    {
                        "id": None,
                        "parent_id": None,
                        "row_kind": "title",
                        "work_type_id": default_work_type_id,
                        "work_type_name": "",
                        "work_item_id": None,
                        "labor_norm_id": None,
                        "labor_hours_per_unit": None,
                        "productivity_factor": None,
                        "name": name_text,
                        "uom_code": None,
                        "plan_qty": None,
                        "plan_start": _today(),
                        "plan_finish": _today(),
                        "status": "planned",
                        "is_milestone": False,
                        "sort_order": len(tasks) * 10,
                    }
                )
                continue

            # ── TASK ──────────────────────────────────────────
            wt = wt_map.get(type_norm)
            if not wt:
                errors.append(
                    f"Строка {row_idx}: тип работ '{type_text}' "
                    f"не найден в справочнике"
                )
                continue

            uom_code = None
            if str(raw_uom or "").strip():
                uom = uom_map.get(GprExcelImportService._norm(raw_uom))
                if not uom:
                    errors.append(
                        f"Строка {row_idx}: единица измерения '{raw_uom}' "
                        f"не найдена в справочнике"
                    )
                    continue
                uom_code = uom["code"]

            qty = None
            if raw_qty not in (None, ""):
                qty = _safe_float(raw_qty)
                if qty is None:
                    errors.append(
                        f"Строка {row_idx}: неверное значение объёма "
                        f"'{raw_qty}'"
                    )
                    continue

            ds = GprExcelImportService._cell_to_date(raw_start)
            df = GprExcelImportService._cell_to_date(raw_finish)

            if not ds:
                errors.append(
                    f"Строка {row_idx}: неверная дата начала "
                    f"'{raw_start}'"
                )
                continue

            if not df:
                errors.append(
                    f"Строка {row_idx}: неверная дата окончания "
                    f"'{raw_finish}'"
                )
                continue

            if df < ds:
                errors.append(
                    f"Строка {row_idx}: окончание раньше начала"
                )
                continue

            status = "planned"
            if str(raw_status or "").strip():
                status = status_map.get(
                    GprExcelImportService._norm(raw_status)
                )
                if not status:
                    errors.append(
                        f"Строка {row_idx}: неизвестный статус "
                        f"'{raw_status}'"
                    )
                    continue

            tasks.append(
                {
                    "id": None,
                    "parent_id": None,
                    "row_kind": "task",
                    "work_type_id": int(wt["id"]),
                    "work_type_name": wt["name"],
                    "name": name_text,
                    "uom_code": uom_code,
                    "plan_qty": qty,
                    "plan_start": ds,
                    "plan_finish": df,
                    "status": status,
                    "is_milestone": False,
                    "sort_order": len(tasks) * 10,
                }
            )

        return {
            "sheet_name": ws.title,
            "header_row": header_row,
            "tasks": tasks,
            "errors": errors,
            "count": len(tasks),
            "skipped_empty": skipped_empty,
        }

# ═══════════════════════════════════════════════════════════════
#  SERVICE LAYER
# ═══════════════════════════════════════════════════════════════
class GprService:

    # ── objects ──
    @staticmethod
    def load_objects_short() -> List[Dict[str, Any]]:
        with _DBConn() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("""
                    SELECT id,
                           COALESCE(short_name,'') AS short_name,
                           address,
                           COALESCE(excel_id,'') AS excel_id,
                           COALESCE(status,'') AS status
                    FROM public.objects
                    ORDER BY address, short_name
                """)
                return [dict(r) for r in cur.fetchall()]

    @staticmethod
    def load_task_fact_upto(task_ids: List[int], cutoff: date) -> Dict[int, float]:
        """
        Накопительный факт по задачам на дату cutoff (включительно).
        Возвращает: { task_id: fact_qty_upto }
        """
        if not task_ids:
            return {}
        with _DBConn() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    """
                    SELECT task_id, COALESCE(SUM(fact_qty), 0) AS fact_qty_upto
                    FROM public.gpr_task_facts
                    WHERE task_id = ANY(%s)
                      AND fact_date <= %s
                    GROUP BY task_id
                    """,
                    (task_ids, cutoff),
                )
                out: Dict[int, float] = {}
                for r in cur.fetchall():
                    out[int(r["task_id"])] = float(r.get("fact_qty_upto") or 0.0)
                return out

    # ── dictionaries ──
    @staticmethod
    def load_work_types() -> List[Dict[str, Any]]:
        with _DBConn() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("""
                    SELECT id, COALESCE(code,'') AS code, name
                    FROM public.gpr_work_types WHERE is_active=true
                    ORDER BY sort_order, name
                """)
                return [dict(r) for r in cur.fetchall()]

    @staticmethod
    def load_uoms() -> List[Dict[str, Any]]:
        with _DBConn() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("SELECT code, name FROM public.gpr_uom ORDER BY code")
                return [dict(r) for r in cur.fetchall()]

    @staticmethod
    def load_statuses() -> List[Dict[str, Any]]:
        with _DBConn() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    "SELECT code, name FROM public.gpr_statuses ORDER BY code"
                )
                return [dict(r) for r in cur.fetchall()]

    @staticmethod
    def load_gpr_registry() -> List[Dict[str, Any]]:
        with _DBConn() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    """
                    SELECT
                        o.id AS object_db_id,
                        COALESCE(o.excel_id, '') AS excel_id,
                        COALESCE(o.short_name, '') AS short_name,
                        o.address,
                        COALESCE(o.status, '') AS object_status,
    
                        p.id AS plan_id,
                        p.version_no,
                        p.updated_at,
                        COALESCE(u.full_name, '') AS creator_name,
    
                        COALESCE(s.task_count, 0) AS task_count,
                        COALESCE(s.done_count, 0) AS done_count,
                        COALESCE(s.in_progress_count, 0) AS in_progress_count,
                        COALESCE(s.overdue_count, 0) AS overdue_count
    
                    FROM public.objects o
                    LEFT JOIN public.gpr_plans p
                           ON p.object_db_id = o.id
                          AND p.is_current = true
                    LEFT JOIN public.app_users u
                           ON u.id = p.created_by
                    LEFT JOIN (
                        SELECT
                            t.plan_id,
                            COUNT(*) FILTER (
                                WHERE COALESCE(t.is_deleted, false) = false
                                  AND COALESCE(t.row_kind, 'task') = 'task'
                            ) AS task_count,
                            COUNT(*) FILTER (
                                WHERE COALESCE(t.is_deleted, false) = false
                                  AND COALESCE(t.row_kind, 'task') = 'task'
                                  AND t.status = 'done'
                            ) AS done_count,
                            COUNT(*) FILTER (
                                WHERE COALESCE(t.is_deleted, false) = false
                                  AND COALESCE(t.row_kind, 'task') = 'task'
                                  AND t.status = 'in_progress'
                            ) AS in_progress_count,
                            COUNT(*) FILTER (
                                WHERE COALESCE(t.is_deleted, false) = false
                                  AND COALESCE(t.row_kind, 'task') = 'task'
                                  AND t.status NOT IN ('done', 'canceled')
                                  AND t.plan_finish < CURRENT_DATE
                            ) AS overdue_count
                        FROM public.gpr_tasks t
                        GROUP BY t.plan_id
                    ) s ON s.plan_id = p.id
                    ORDER BY o.address, o.short_name
                    """
                )
                return [dict(r) for r in cur.fetchall()]

    @staticmethod
    def load_task_fact_info(task_ids: List[int]) -> Dict[int, Dict[str, Any]]:
        if not task_ids:
            return {}
    
        with _DBConn() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    """
                    WITH qty AS (
                        SELECT
                            task_id,
                            COALESCE(SUM(fact_qty), 0) AS fact_qty_total
                        FROM public.gpr_task_facts
                        WHERE task_id = ANY(%s)
                        GROUP BY task_id
                    ),
                    last_fact AS (
                        SELECT DISTINCT ON (task_id)
                            task_id,
                            workers_count,
                            fact_date,
                            id
                        FROM public.gpr_task_facts
                        WHERE task_id = ANY(%s)
                        ORDER BY task_id, fact_date DESC, id DESC
                    ),
                    agg AS (
                        SELECT
                            task_id,
                            MAX(workers_count) AS workers_max,
                            COALESCE(SUM(workers_count), 0) AS workers_sum
                        FROM public.gpr_task_facts
                        WHERE task_id = ANY(%s)
                        GROUP BY task_id
                    )
                    SELECT
                        q.task_id,
                        q.fact_qty_total,
                        lf.workers_count AS workers_last,
                        a.workers_max,
                        a.workers_sum
                    FROM qty q
                    LEFT JOIN last_fact lf ON lf.task_id = q.task_id
                    LEFT JOIN agg a ON a.task_id = q.task_id
                    """
                    ,
                    (task_ids, task_ids, task_ids),
                )
    
                out = {}
                for r in cur.fetchall():
                    d = dict(r)
                    out[int(d["task_id"])] = {
                        "fact_qty_total": float(d.get("fact_qty_total") or 0),
                        "workers_last": int(d["workers_last"]) if d.get("workers_last") is not None else None,
                        "workers_max": int(d["workers_max"]) if d.get("workers_max") is not None else None,
                        "workers_sum": int(d["workers_sum"]) if d.get("workers_sum") is not None else 0,
                    }
                return out

    @staticmethod
    def load_task_fact_period_info(
        task_ids: List[int],
        period_from: date,
        period_to: date,
    ) -> Dict[int, Dict[str, Any]]:
        """
        Загружает факт только за выбранный период.

        Возвращает:
        {
            task_id: {
                "fact_qty_period": float,
                "workers_last_period": int | None,
                "workers_max_period": int | None,
                "workers_sum_period": int,
            }
        }
        """
        if not task_ids:
            return {}

        with _DBConn() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    """
                    WITH qty AS (
                        SELECT
                            task_id,
                            COALESCE(SUM(fact_qty), 0) AS fact_qty_period
                        FROM public.gpr_task_facts
                        WHERE task_id = ANY(%s)
                          AND fact_date BETWEEN %s AND %s
                        GROUP BY task_id
                    ),
                    last_fact AS (
                        SELECT DISTINCT ON (task_id)
                            task_id,
                            workers_count,
                            fact_date,
                            id
                        FROM public.gpr_task_facts
                        WHERE task_id = ANY(%s)
                          AND fact_date BETWEEN %s AND %s
                        ORDER BY task_id, fact_date DESC, id DESC
                    ),
                    agg AS (
                        SELECT
                            task_id,
                            MAX(workers_count) AS workers_max_period,
                            COALESCE(SUM(workers_count), 0) AS workers_sum_period
                        FROM public.gpr_task_facts
                        WHERE task_id = ANY(%s)
                          AND fact_date BETWEEN %s AND %s
                        GROUP BY task_id
                    )
                    SELECT
                        q.task_id,
                        q.fact_qty_period,
                        lf.workers_count AS workers_last_period,
                        a.workers_max_period,
                        a.workers_sum_period
                    FROM qty q
                    LEFT JOIN last_fact lf ON lf.task_id = q.task_id
                    LEFT JOIN agg a ON a.task_id = q.task_id
                    """,
                    (
                        task_ids,
                        period_from,
                        period_to,
                        task_ids,
                        period_from,
                        period_to,
                        task_ids,
                        period_from,
                        period_to,
                    ),
                )

                out: Dict[int, Dict[str, Any]] = {}

                for r in cur.fetchall():
                    d = dict(r)
                    tid = int(d["task_id"])

                    out[tid] = {
                        "fact_qty_period": float(d.get("fact_qty_period") or 0),
                        "workers_last_period": (
                            int(d["workers_last_period"])
                            if d.get("workers_last_period") is not None
                            else None
                        ),
                        "workers_max_period": (
                            int(d["workers_max_period"])
                            if d.get("workers_max_period") is not None
                            else None
                        ),
                        "workers_sum_period": (
                            int(d["workers_sum_period"])
                            if d.get("workers_sum_period") is not None
                            else 0
                        ),
                    }

                return out

    # ── plans ──
    @staticmethod
    def get_or_create_current_plan(
        object_db_id: int, user_id: Optional[int]
    ) -> Dict[str, Any]:
        with _DBConn() as conn:
            with conn:  # autocommit-block
                with conn.cursor(cursor_factory=RealDictCursor) as cur:
                    cur.execute(
                        """
                        SELECT p.*, u.full_name AS creator_name
                        FROM public.gpr_plans p
                        LEFT JOIN public.app_users u ON u.id = p.created_by
                        WHERE p.object_db_id=%s AND p.is_current=true
                        LIMIT 1
                    """,
                        (object_db_id,),
                    )
                    row = cur.fetchone()
                    if row:
                        return dict(row)

                    cur.execute(
                        """
                        INSERT INTO public.gpr_plans
                            (object_db_id, version_no, is_current,
                             is_baseline, created_by)
                        VALUES (%s, 1, true, false, %s)
                        RETURNING id
                    """,
                        (object_db_id, user_id),
                    )
                    pid = cur.fetchone()["id"]

                    cur.execute(
                        """
                        SELECT p.*, u.full_name AS creator_name
                        FROM public.gpr_plans p
                        LEFT JOIN public.app_users u ON u.id = p.created_by
                        WHERE p.id=%s
                    """,
                        (pid,),
                    )
                    return dict(cur.fetchone())

    @staticmethod
    def load_plan_tasks(plan_id: int) -> List[Dict[str, Any]]:
        with _DBConn() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    """
                    SELECT
                        t.id,
                        t.parent_id,
                        t.work_type_id,
                        wt.name AS work_type_name,

                        t.work_item_id,
                        t.labor_norm_id,
                        t.labor_hours_per_unit,
                        t.productivity_factor,

                        t.name,
                        t.uom_code,
                        t.plan_qty,
                        t.plan_start,
                        t.plan_finish,
                        t.status,
                        t.sort_order,
                        t.is_milestone,
                        t.row_kind,
                        t.created_by,
                        t.created_at,
                        t.updated_at
                    FROM public.gpr_tasks t
                    JOIN public.gpr_work_types wt
                        ON wt.id = t.work_type_id
                    WHERE t.plan_id = %s
                      AND COALESCE(t.is_deleted, false) = false
                    ORDER BY
                        t.sort_order,
                        wt.sort_order,
                        wt.name,
                        t.name,
                        t.plan_start,
                        t.id
                    """,
                    (plan_id,),
                )

                rows = []

                for r in cur.fetchall():
                    d = dict(r)
                    d["row_kind"] = (d.get("row_kind") or "task").strip()
                    d["plan_start"] = _to_date(d.get("plan_start"))
                    d["plan_finish"] = _to_date(d.get("plan_finish"))

                    if d.get("productivity_factor") is None:
                        d["productivity_factor"] = 1.0

                    rows.append(d)

                return rows

    @staticmethod
    def load_task_facts_cumulative(task_ids: List[int]) -> Dict[int, float]:
        info = GprService.load_task_fact_info(task_ids)
        return {
            task_id: float(v.get("fact_qty_total") or 0)
            for task_id, v in info.items()
        }

    @staticmethod
    def replace_plan_tasks(
        plan_id: int,
        user_id: Optional[int],
        tasks: List[Dict[str, Any]],
    ) -> None:
        """
        Полностью синхронизирует строки ГПР с БД.

        Новые строки вставляются.
        Существующие обновляются.
        Удалённые из редактора строки мягко архивируются.

        Для задач типа task дополнительно хранится связь с конкретной
        работой и снимок нормы ЗТР на дату создания задачи.
        """
        with _DBConn() as conn:
            with conn:
                with conn.cursor(cursor_factory=RealDictCursor) as cur:
                    cur.execute(
                        """
                        SELECT
                            id,
                            COALESCE(is_deleted, false) AS is_deleted
                        FROM public.gpr_tasks
                        WHERE plan_id = %s
                        """,
                        (plan_id,),
                    )

                    existing_rows = cur.fetchall()

                    existing_ids: Set[int] = {
                        int(row["id"])
                        for row in existing_rows
                    }

                    existing_active_ids: Set[int] = {
                        int(row["id"])
                        for row in existing_rows
                        if not row["is_deleted"]
                    }

                    seen_ids: Set[int] = set()
                    inserts: List[Tuple[Any, ...]] = []

                    for i, t in enumerate(tasks):
                        row_kind = (t.get("row_kind") or "task").strip()

                        if row_kind not in ("task", "group", "title"):
                            row_kind = "task"

                        name = (t.get("name") or "").strip()
                        if not name:
                            raise ValueError(
                                f"Строка {i + 1}: пустое название"
                            )

                        work_type_id = t.get("work_type_id")
                        if work_type_id is None:
                            raise ValueError(
                                f"Строка {i + 1} «{name}»: "
                                "не указан тип работ"
                            )

                        try:
                            work_type_id = int(work_type_id)
                        except (ValueError, TypeError) as exc:
                            raise ValueError(
                                f"Строка {i + 1} «{name}»: "
                                "неверный идентификатор типа работ"
                            ) from exc

                        parent_id = t.get("parent_id")

                        if parent_id in ("", 0):
                            parent_id = None
                        elif parent_id is not None:
                            try:
                                parent_id = int(parent_id)
                            except (ValueError, TypeError):
                                parent_id = None

                        plan_start = _to_date(t.get("plan_start"))
                        plan_finish = _to_date(t.get("plan_finish"))

                        if row_kind == "task":
                            if not plan_start or not plan_finish:
                                raise ValueError(
                                    f"Задача «{name}»: "
                                    "не указаны даты начала или окончания"
                                )

                            if plan_finish < plan_start:
                                raise ValueError(
                                    f"Задача «{name}»: "
                                    "окончание раньше начала"
                                )
                        else:
                            # Технические даты для групп и титулов.
                            if not plan_start:
                                plan_start = _today()

                            if not plan_finish:
                                plan_finish = plan_start

                        status = (t.get("status") or "planned").strip()
                        if status not in STATUS_LIST:
                            status = "planned"

                        try:
                            sort_order = int(
                                t.get("sort_order")
                                if t.get("sort_order") is not None
                                else i * 10
                            )
                        except (ValueError, TypeError):
                            sort_order = i * 10

                        is_milestone = bool(t.get("is_milestone") or False)
                        uom_code = t.get("uom_code") or None
                        plan_qty = _safe_float(t.get("plan_qty"))

                        # ─────────────────────────────────────
                        # Связь с профессиональным справочником.
                        # Для групп и титулов всё должно быть NULL.
                        # ─────────────────────────────────────
                        work_item_id = None
                        labor_norm_id = None
                        labor_hours_per_unit = None
                        productivity_factor = None

                        if row_kind == "task":
                            raw_work_item_id = t.get("work_item_id")
                            raw_labor_norm_id = t.get("labor_norm_id")

                            if raw_work_item_id is not None:
                                try:
                                    work_item_id = int(raw_work_item_id)
                                except (TypeError, ValueError):
                                    work_item_id = None

                            if raw_labor_norm_id is not None:
                                try:
                                    labor_norm_id = int(raw_labor_norm_id)
                                except (TypeError, ValueError):
                                    labor_norm_id = None

                            labor_hours_per_unit = _safe_float(
                                t.get("labor_hours_per_unit")
                            )

                            productivity_factor = _safe_float(
                                t.get("productivity_factor")
                            )

                            if (
                                productivity_factor is None
                                or productivity_factor <= 0
                            ):
                                productivity_factor = 1.0

                        task_id = t.get("id")

                        if task_id is not None:
                            try:
                                task_id = int(task_id)
                            except (ValueError, TypeError):
                                task_id = None

                        # ─────────────────────────────────────
                        # Обновление существующей строки.
                        # ─────────────────────────────────────
                        if task_id and task_id in existing_ids:
                            cur.execute(
                                """
                                UPDATE public.gpr_tasks
                                SET
                                    parent_id = %s,
                                    work_type_id = %s,

                                    work_item_id = %s,
                                    labor_norm_id = %s,
                                    labor_hours_per_unit = %s,
                                    productivity_factor = %s,

                                    name = %s,
                                    uom_code = %s,
                                    plan_qty = %s,
                                    plan_start = %s,
                                    plan_finish = %s,
                                    status = %s,
                                    sort_order = %s,
                                    is_milestone = %s,
                                    row_kind = %s,

                                    is_deleted = false,
                                    deleted_at = NULL,
                                    deleted_by = NULL,
                                    updated_at = now()

                                WHERE id = %s
                                  AND plan_id = %s
                                """,
                                (
                                    parent_id,
                                    work_type_id,

                                    work_item_id,
                                    labor_norm_id,
                                    labor_hours_per_unit,
                                    productivity_factor,

                                    name,
                                    uom_code,
                                    plan_qty,
                                    plan_start,
                                    plan_finish,
                                    status,
                                    sort_order,
                                    is_milestone,
                                    row_kind,

                                    task_id,
                                    plan_id,
                                ),
                            )

                            seen_ids.add(task_id)

                        # ─────────────────────────────────────
                        # Вставка новой строки.
                        # ─────────────────────────────────────
                        else:
                            inserts.append(
                                (
                                    plan_id,
                                    work_type_id,
                                    parent_id,

                                    work_item_id,
                                    labor_norm_id,
                                    labor_hours_per_unit,
                                    productivity_factor,

                                    name,
                                    uom_code,
                                    plan_qty,
                                    plan_start,
                                    plan_finish,
                                    status,
                                    is_milestone,
                                    sort_order,
                                    user_id,
                                    user_id,
                                    row_kind,
                                )
                            )

                    if inserts:
                        execute_values(
                            cur,
                            """
                            INSERT INTO public.gpr_tasks (
                                plan_id,
                                work_type_id,
                                parent_id,

                                work_item_id,
                                labor_norm_id,
                                labor_hours_per_unit,
                                productivity_factor,

                                name,
                                uom_code,
                                plan_qty,
                                plan_start,
                                plan_finish,
                                status,
                                is_milestone,
                                sort_order,
                                created_by,
                                updated_by,
                                row_kind
                            )
                            VALUES %s
                            """,
                            inserts,
                        )

                    # Архивируем активные строки, которых больше нет
                    # в текущем наборе редактора.
                    ids_to_soft_delete = list(
                        existing_active_ids - seen_ids
                    )

                    if ids_to_soft_delete:
                        cur.execute(
                            """
                            UPDATE public.gpr_tasks
                            SET
                                is_deleted = true,
                                deleted_at = now(),
                                deleted_by = %s,
                                status = CASE
                                    WHEN status IN (
                                        'planned',
                                        'in_progress',
                                        'paused'
                                    )
                                    THEN 'canceled'
                                    ELSE status
                                END,
                                updated_at = now()

                            WHERE id = ANY(%s)
                              AND plan_id = %s
                            """,
                            (
                                user_id,
                                ids_to_soft_delete,
                                plan_id,
                            ),
                        )

                    cur.execute(
                        """
                        UPDATE public.gpr_plans
                        SET updated_at = now()
                        WHERE id = %s
                        """,
                        (plan_id,),
                    )
    @staticmethod
    def update_task_status(task_id: int, new_status: str) -> None:
        if new_status not in STATUS_LIST:
            raise ValueError(f"Неизвестный статус: {new_status}")
        with _DBConn() as conn:
            with conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        UPDATE public.gpr_tasks
                        SET status=%s, updated_at=now()
                        WHERE id=%s
                          AND COALESCE(is_deleted, false)=false
                        """,
                        (new_status, task_id),
                    )

    # ── templates ──
    @staticmethod
    def load_templates() -> List[Dict[str, Any]]:
        with _DBConn() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    """
                    SELECT id, name FROM public.gpr_templates
                    WHERE is_active=true ORDER BY name
                """
                )
                return [dict(r) for r in cur.fetchall()]

    @staticmethod
    def load_template_tasks(template_id: int) -> List[Dict[str, Any]]:
        with _DBConn() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    """
                        SELECT id, parent_id, work_type_id, name, uom_code,
                               default_qty, is_milestone, sort_order, row_kind
                        FROM public.gpr_template_tasks
                        WHERE template_id=%s ORDER BY sort_order, id
                    """,
                    (template_id,),
                )
                return [dict(r) for r in cur.fetchall()]


# ═══════════════════════════════════════════════════════════════
#  AUTOCOMPLETE COMBOBOX (исправлен)
# ═══════════════════════════════════════════════════════════════
class _AutoCombo(ttk.Combobox):
    """Combobox с автодополнением. Хранит полный список и
    отображение label→index для корректного получения выбранного
    элемента даже после фильтрации."""

    # Клавиши, которые НЕ должны вызывать фильтрацию
    _IGNORE_KEYS = {
        "Return", "Escape", "Tab", "Up", "Down",
        "Left", "Right", "Home", "End",
        "Shift_L", "Shift_R", "Control_L", "Control_R",
        "Alt_L", "Alt_R", "Caps_Lock",
    }

    def __init__(self, master=None, **kw):
        super().__init__(master, **kw)
        self._all: List[str] = []
        self._label_to_idx: Dict[str, int] = {}
        self.bind("<KeyRelease>", self._on_key)

    def set_values(self, vals: List[str]):
        self._all = list(vals or [])
        self._label_to_idx = {v: i for i, v in enumerate(self._all)}
        self.config(values=self._all)

    def get_original_index(self) -> int:
        """Возвращает индекс в ОРИГИНАЛЬНОМ списке, -1 если не найден."""
        text = self.get().strip()
        return self._label_to_idx.get(text, -1)

    def _on_key(self, event):
        if event.keysym in self._IGNORE_KEYS:
            return
        q = self.get().strip().lower()
        if not q:
            self.config(values=self._all)
            return
        filtered = [v for v in self._all if q in v.lower()]
        self.config(values=filtered)


# ═══════════════════════════════════════════════════════════════
#  DIALOGS
# ═══════════════════════════════════════════════════════════════
class DateRangeDialog(simpledialog.Dialog):
    def __init__(self, parent, d0: date, d1: date):
        self._d0, self._d1 = d0, d1
        self.result: Optional[Tuple[date, date]] = None
        super().__init__(parent, title="Диапазон дат отображения")

    def body(self, m):
        tk.Label(m, text="С (дд.мм.гггг):").grid(
            row=0, column=0, sticky="e", padx=(0, 6), pady=4
        )
        self.e0 = ttk.Entry(m, width=14)
        self.e0.grid(row=0, column=1, pady=4)
        self.e0.insert(0, _fmt_date(self._d0))

        tk.Label(m, text="По (дд.мм.гггг):").grid(
            row=1, column=0, sticky="e", padx=(0, 6), pady=4
        )
        self.e1 = ttk.Entry(m, width=14)
        self.e1.grid(row=1, column=1, pady=4)
        self.e1.insert(0, _fmt_date(self._d1))

        ttk.Button(m, text="Текущий квартал", command=self._set_quarter).grid(
            row=2, column=0, columnspan=2, pady=(8, 0)
        )
        return self.e0

    def _set_quarter(self):
        d0, d1 = _quarter_range()
        self.e0.delete(0, "end")
        self.e0.insert(0, _fmt_date(d0))
        self.e1.delete(0, "end")
        self.e1.insert(0, _fmt_date(d1))

    def validate(self):
        try:
            a = _parse_date(self.e0.get())
            b = _parse_date(self.e1.get())
            if b < a:
                raise ValueError("Дата окончания раньше даты начала")
            self._a, self._b = a, b
            return True
        except Exception as e:
            messagebox.showwarning("Даты", str(e), parent=self)
            return False

    def apply(self):
        self.result = (self._a, self._b)


class TaskEditDialog(simpledialog.Dialog):
    """Встроенный диалог редактирования задачи (fallback если нет
    внешнего gpr_task_dialog)."""

    def __init__(self, parent, wt, uoms, statuses_db=None, init=None):
        self.wt = wt
        self.uoms = uoms
        self.init = init or {}
        self._statuses_db = statuses_db or []
        self.result: Optional[Dict[str, Any]] = None
        super().__init__(parent, title="Работа ГПР")

    def body(self, m):
        wt_v = [w["name"] for w in self.wt]
        uom_v = [f"{u['code']} — {u['name']}" for u in self.uoms]
        st_v = [STATUS_LABELS.get(s, s) for s in STATUS_LIST]

        r = 0
        tk.Label(m, text="Тип работ *:").grid(
            row=r, column=0, sticky="e", padx=(0, 6), pady=3
        )
        self.cmb_wt = ttk.Combobox(
            m, state="readonly", width=42, values=wt_v
        )
        self.cmb_wt.grid(row=r, column=1, pady=3)
        r += 1

        tk.Label(m, text="Вид работ *:").grid(
            row=r, column=0, sticky="e", padx=(0, 6), pady=3
        )
        self.ent_name = ttk.Entry(m, width=46)
        self.ent_name.grid(row=r, column=1, pady=3)
        r += 1

        tk.Label(m, text="Ед. изм.:").grid(
            row=r, column=0, sticky="e", padx=(0, 6), pady=3
        )
        self.cmb_uom = ttk.Combobox(
            m, state="readonly", width=42, values=["—"] + uom_v
        )
        self.cmb_uom.grid(row=r, column=1, pady=3)
        r += 1

        tk.Label(m, text="Объём план:").grid(
            row=r, column=0, sticky="e", padx=(0, 6), pady=3
        )
        self.ent_qty = ttk.Entry(m, width=18)
        self.ent_qty.grid(row=r, column=1, sticky="w", pady=3)
        r += 1

        tk.Label(m, text="Начало *:").grid(
            row=r, column=0, sticky="e", padx=(0, 6), pady=3
        )
        self.ent_s = ttk.Entry(m, width=14)
        self.ent_s.grid(row=r, column=1, sticky="w", pady=3)
        r += 1

        tk.Label(m, text="Окончание *:").grid(
            row=r, column=0, sticky="e", padx=(0, 6), pady=3
        )
        self.ent_f = ttk.Entry(m, width=14)
        self.ent_f.grid(row=r, column=1, sticky="w", pady=3)
        r += 1

        tk.Label(m, text="Статус:").grid(
            row=r, column=0, sticky="e", padx=(0, 6), pady=3
        )
        self.cmb_st = ttk.Combobox(
            m, state="readonly", width=20, values=st_v
        )
        self.cmb_st.grid(row=r, column=1, sticky="w", pady=3)
        r += 1

        self.var_ms = tk.BooleanVar(
            value=bool(self.init.get("is_milestone"))
        )
        ttk.Checkbutton(
            m, text="Веха (milestone)", variable=self.var_ms
        ).grid(row=r, column=1, sticky="w", pady=3)

        # ── заполняем начальные значения ──
        iw = self.init.get("work_type_id")
        if iw:
            for i, w in enumerate(self.wt):
                if int(w["id"]) == int(iw):
                    self.cmb_wt.current(i)
                    break
            else:
                if self.wt:
                    self.cmb_wt.current(0)
        elif self.wt:
            self.cmb_wt.current(0)

        self.ent_name.insert(0, self.init.get("name", ""))

        iu = self.init.get("uom_code")
        if iu:
            for i, u in enumerate(self.uoms):
                if u["code"] == iu:
                    self.cmb_uom.current(i + 1)  # +1 из-за "—"
                    break
            else:
                self.cmb_uom.current(0)
        else:
            self.cmb_uom.current(0)

        if self.init.get("plan_qty") is not None:
            self.ent_qty.insert(0, _fmt_qty(self.init["plan_qty"]))

        d0 = _to_date(self.init.get("plan_start")) or _today()
        d1 = _to_date(self.init.get("plan_finish")) or _today()
        self.ent_s.insert(0, _fmt_date(d0))
        self.ent_f.insert(0, _fmt_date(d1))

        ist = self.init.get("status", "planned")
        try:
            self.cmb_st.current(STATUS_LIST.index(ist))
        except ValueError:
            self.cmb_st.current(0)

        return self.ent_name

    def validate(self):
        try:
            wi = self.cmb_wt.current()
            if wi < 0:
                raise ValueError("Выберите тип работ")
            wt_id = int(self.wt[wi]["id"])
            nm = (self.ent_name.get() or "").strip()
            if not nm:
                raise ValueError("Введите вид работ")

            uom = None
            ui = self.cmb_uom.current()
            if ui > 0:
                uom = self.uoms[ui - 1]["code"]

            qty = _safe_float(self.ent_qty.get())
            ds = _parse_date(self.ent_s.get())
            df = _parse_date(self.ent_f.get())
            if df < ds:
                raise ValueError("Окончание раньше начала")

            si = self.cmb_st.current()
            st = STATUS_LIST[si] if 0 <= si < len(STATUS_LIST) else "planned"

            self._out = dict(
                work_type_id=wt_id,
                name=nm,
                uom_code=uom,
                plan_qty=qty,
                plan_start=ds,
                plan_finish=df,
                status=st,
                is_milestone=bool(self.var_ms.get()),
            )
            return True
        except Exception as e:
            messagebox.showwarning("Работа", str(e), parent=self)
            return False

    def apply(self):
        self.result = dict(self._out)


class TemplateSelectDialog(simpledialog.Dialog):
    def __init__(self, parent, templates):
        self.templates = templates
        self.result: Optional[int] = None
        super().__init__(parent, title="Выбор шаблона ГПР")

    def body(self, m):
        tk.Label(m, text="Выберите шаблон:").pack(anchor="w", pady=(0, 6))
        self.lb = tk.Listbox(
            m, width=50, height=min(15, max(4, len(self.templates)))
        )
        for t in self.templates:
            self.lb.insert("end", t["name"])
        self.lb.pack(fill="both", expand=True)
        if self.templates:
            self.lb.selection_set(0)
        return self.lb

    def validate(self):
        sel = self.lb.curselection()
        if not sel:
            messagebox.showwarning(
                "Шаблон", "Выберите шаблон.", parent=self
            )
            return False
        self._idx = sel[0]
        return True

    def apply(self):
        self.result = int(self.templates[self._idx]["id"])


# ═══════════════════════════════════════════════════════════════
#  GANTT CANVAS (optimized scroll performance)
# ═══════════════════════════════════════════════════════════════

class GanttCanvas(tk.Frame):
    """Гант, синхронизированный с Treeview по позициям строк."""

    MONTH_H = 20
    DAY_H = 22
    HEADER_H = MONTH_H + DAY_H

    def __init__(self, master, *, day_px=20, linked_tree=None):
        super().__init__(master, bg=C["panel"])
        self.day_px = day_px
        self._tree = linked_tree
    
        self.hdr = tk.Canvas(
            self, height=self.HEADER_H, bg="#e8eaed", highlightthickness=0
        )
        self.body = tk.Canvas(self, bg="#ffffff", highlightthickness=0)
        self.hsb = ttk.Scrollbar(
            self, orient="horizontal", command=self._xview
        )
    
        self.hdr.grid(row=0, column=0, sticky="ew")
        self.body.grid(row=1, column=0, sticky="nsew")
        self.hsb.grid(row=2, column=0, sticky="ew")
        self.grid_rowconfigure(1, weight=1)
        self.grid_columnconfigure(0, weight=1)
    
        self.body.configure(xscrollcommand=self._on_xscroll)
    
        self._range: Tuple[date, date] = _quarter_range()
        self._rows: List[Dict[str, Any]] = []
        self._facts: Dict[int, float] = {}
        self._fact_info: Dict[int, Dict[str, Any]] = {}
    
        self._heavy_redraw_pending: Optional[str] = None
        self._is_mapped = False
        self._header_cache_key = None
    
        self.body.bind("<Configure>", self._on_configure)
        self.body.bind("<MouseWheel>", self._wheel)
        self.body.bind("<Button-4>", self._wheel)
        self.body.bind("<Button-5>", self._wheel)
        self.body.bind("<Shift-MouseWheel>", self._hwheel)
        self.bind("<Map>", self._on_map)

    def _on_map(self, _e=None):
        self._is_mapped = True
        self._schedule_heavy_redraw()

    def _on_configure(self, _e=None):
        self._schedule_heavy_redraw()

    def set_tree(self, tree):
        self._tree = tree

    def set_range(self, d0, d1):
        self._range = (d0, d1)
        self._header_cache_key = None
        self._schedule_heavy_redraw()

    def set_data(self, rows, facts=None, fact_info=None):
        self._rows = rows or []
        self._facts = facts or {}
        self._fact_info = fact_info or {}
        self._schedule_heavy_redraw()

    def _schedule_heavy_redraw(self):
        if self._heavy_redraw_pending:
            self.after_cancel(self._heavy_redraw_pending)
        self._heavy_redraw_pending = self.after(40, self._full_redraw)

    def _full_redraw(self):
        self._heavy_redraw_pending = None
        if not self._is_mapped:
            return
        try:
            self._draw_header()
            self._draw_bars()
        except Exception:
            logger.exception("GanttCanvas._full_redraw error")

    def redraw_bars_only(self):
        if not self._is_mapped:
            return
        try:
            self._draw_bars()
        except Exception:
            logger.exception("GanttCanvas.redraw_bars_only error")

    def _xview(self, *a):
        self.body.xview(*a)
        self.hdr.xview(*a)

    def _on_xscroll(self, f0, f1):
        self.hsb.set(f0, f1)
        self.hdr.xview_moveto(float(f0))

    def _wheel(self, e):
        if self._tree:
            d = _mouse_delta(e)
            if d:
                self._tree.yview_scroll(d, "units")
                self.after_idle(self.redraw_bars_only)
        return "break"

    def _hwheel(self, e):
        d = _mouse_delta(e)
        if d:
            self.body.xview_scroll(d, "units")
            self.hdr.xview_scroll(d, "units")
        return "break"

    def _get_tree_row_positions(self) -> List[Optional[Tuple[int, int]]]:
        """Позиции строк Treeview, нормализованные относительно
        первой видимой строки. Это устраняет потерю первой строки
        и выравнивает бары по строкам."""
        if not self._tree:
            return []
    
        items = self._tree.get_children()
        if not items:
            return []
    
        raw_positions: List[Optional[Tuple[int, int]]] = []
        first_visible_y: Optional[int] = None
    
        for iid in items:
            try:
                bbox = self._tree.bbox(iid)
                if bbox:
                    _x, y, _w, h = bbox
                    raw_positions.append((y, y + h))
                    if first_visible_y is None:
                        first_visible_y = y
                else:
                    raw_positions.append(None)
            except (tk.TclError, Exception):
                raw_positions.append(None)
    
        if first_visible_y is None:
            return raw_positions
    
        norm_positions: List[Optional[Tuple[int, int]]] = []
        for pos in raw_positions:
            if pos is None:
                norm_positions.append(None)
            else:
                y0, y1 = pos
                norm_positions.append((y0 - first_visible_y, y1 - first_visible_y))
    
        return norm_positions

    def _draw_header(self):
        d0, d1 = self._range
        if d1 < d0:
            return

        days = (d1 - d0).days + 1
        tw = max(1, days * self.day_px)

        cache_key = (d0, d1, self.day_px)
        if self._header_cache_key == cache_key:
            return
        self._header_cache_key = cache_key

        self.hdr.delete("all")
        self.hdr.configure(scrollregion=(0, 0, tw, self.HEADER_H))

        cur = date(d0.year, d0.month, 1)
        while cur <= d1:
            mr = calendar.monthrange(cur.year, cur.month)[1]
            ms = max(cur, d0)
            me = min(date(cur.year, cur.month, mr), d1)
            x0 = (ms - d0).days * self.day_px
            x1 = ((me - d0).days + 1) * self.day_px
            self.hdr.create_rectangle(
                x0, 0, x1, self.MONTH_H, fill="#d6dbe0", outline="#bbb"
            )
            if (x1 - x0) > 40:
                self.hdr.create_text(
                    (x0 + x1) / 2,
                    self.MONTH_H / 2,
                    text=cur.strftime("%b %Y"),
                    font=("Segoe UI", 8, "bold"),
                    fill="#333",
                )
            if cur.month == 12:
                cur = date(cur.year + 1, 1, 1)
            else:
                cur = date(cur.year, cur.month + 1, 1)

        for i in range(days):
            x0 = i * self.day_px
            x1 = x0 + self.day_px
            d = d0 + timedelta(days=i)
            fill = "#ffecec" if d.weekday() >= 5 else "#f3f4f6"
            self.hdr.create_rectangle(
                x0, self.MONTH_H, x1, self.HEADER_H,
                fill=fill, outline="#d0d0d0"
            )
            if self.day_px >= 14:
                self.hdr.create_text(
                    (x0 + x1) / 2,
                    self.MONTH_H + self.DAY_H / 2,
                    text=str(d.day),
                    font=("Segoe UI", 7),
                    fill="#555",
                )

        td = _today()
        if d0 <= td <= d1:
            tx = (td - d0).days * self.day_px + self.day_px // 2
            self.hdr.create_line(
                tx, 0, tx, self.HEADER_H, fill=C["error"], width=2
            )

    def _draw_bars(self):
        d0, d1 = self._range
        if d1 < d0:
            return
    
        days = (d1 - d0).days + 1
        tw = max(1, days * self.day_px)
        body_h = self.body.winfo_height()
        if body_h < 10:
            body_h = 600
    
        self.body.delete("all")
        self.body.configure(scrollregion=(0, 0, tw, body_h))
    
        td = _today()
        if d0 <= td <= d1:
            tx = (td - d0).days * self.day_px + self.day_px // 2
            self.body.create_line(
                tx, 0, tx, body_h,
                fill=C["error"], width=1, dash=(4, 2)
            )
    
        step = 7 if self.day_px >= 10 else 14
        for i in range(0, days, step):
            x = i * self.day_px
            self.body.create_line(x, 0, x, body_h, fill="#eeeeee")
    
        positions = self._get_tree_row_positions()
        if not positions:
            return
    
        for row_idx, t in enumerate(self._rows):
            if row_idx >= len(positions) or positions[row_idx] is None:
                continue
    
            y0, y1 = positions[row_idx]
    
            if y1 < -5 or y0 > body_h + 5:
                continue
    
            row_kind = (t.get("row_kind") or "task").strip()
    
            if row_kind == "group":
                bg = "#eef5ff"
            elif row_kind == "title":
                bg = "#dff1ff"
            else:
                bg = "#ffffff" if row_idx % 2 == 0 else "#f8f9fa"
    
            self.body.create_rectangle(0, y0, tw, y1, fill=bg, outline="")
    
            if row_kind in ("group", "title"):
                fg = "#1a3d7c" if row_kind == "group" else "#0b5394"
                prefix = "📁 " if row_kind == "group" else "🟦 "
                self.body.create_text(
                    6,
                    (y0 + y1) / 2,
                    text=prefix + (t.get("name") or ""),
                    anchor="w",
                    font=("Segoe UI", 8, "bold"),
                    fill=fg,
                )
                continue
    
            ts = _to_date(t.get("plan_start"))
            tf = _to_date(t.get("plan_finish"))
            if not ts or not tf:
                continue
            if tf < d0 or ts > d1:
                continue
    
            s2 = max(ts, d0)
            f2 = min(tf, d1)
            bx0 = (s2 - d0).days * self.day_px
            bx1 = ((f2 - d0).days + 1) * self.day_px
    
            st = (t.get("status") or "planned").strip()
            col, _, _ = STATUS_COLORS.get(st, ("#90caf9", "#555", ""))
    
            by0 = y0 + 4
            by1 = y1 - 4
            if by1 - by0 < 4:
                by0 = y0 + 2
                by1 = y1 - 2
    
            self.body.create_rectangle(
                bx0 + 1, by0, bx1 - 1, by1,
                fill=col, outline="#5f6368"
            )
    
            tid = t.get("id")
            pq = _safe_float(t.get("plan_qty"))
            fq = self._facts.get(tid, 0) if tid else 0
            fact_info = self._fact_info.get(tid, {}) if tid else {}
            workers_last = fact_info.get("workers_last")
    
            if pq and pq > 0 and fq > 0:
                pct = min(1.0, fq / pq)
                fw = max(2, int((bx1 - bx0 - 2) * pct))
                self.body.create_rectangle(
                    bx0 + 1, by0, bx0 + 1 + fw, by1,
                    fill="#388e3c", outline=""
                )
    
            if t.get("is_milestone"):
                cx = bx0 + 6
                cy = (y0 + y1) / 2
                self.body.create_polygon(
                    cx, cy, cx + 7, cy - 5,
                    cx + 14, cy, cx + 7, cy + 5,
                    fill="#1a73e8", outline=""
                )
    
            bar_w = bx1 - bx0
    
            info_parts = []
            if fq and pq and pq > 0:
                info_parts.append(f"{_fmt_qty(fq)}/{_fmt_qty(pq)}")
            elif fq:
                info_parts.append(_fmt_qty(fq))
    
            if workers_last:
                info_parts.append(f"{workers_last}чел")
    
            info_text = " · ".join(info_parts)
    
            if info_text and bar_w > 90:
                self.body.create_text(
                    bx1 - 4,
                    (y0 + y1) / 2,
                    text=info_text,
                    anchor="e",
                    font=("Segoe UI", 7),
                    fill="#222"
                )
    
            if bar_w > 60:
                nm = (t.get("name") or "")[:30]
                self.body.create_text(
                    bx0 + 4, (y0 + y1) / 2,
                    text=nm,
                    anchor="w",
                    font=("Segoe UI", 7),
                    fill="#333"
                )
            elif workers_last and bar_w > 26:
                self.body.create_text(
                    (bx0 + bx1) / 2,
                    (y0 + y1) / 2,
                    text=f"{workers_last}",
                    anchor="center",
                    font=("Segoe UI", 7, "bold"),
                    fill="#222"
                )

# ═══════════════════════════════════════════════════════════════
#  MAIN PAGE
# ═══════════════════════════════════════════════════════════════

class GprPage(tk.Frame):

    def __init__(self, master, app_ref):
        super().__init__(master, bg=C["bg"])
        self.app_ref = app_ref
    
        self.objects: List[Dict[str, Any]] = []
        self.work_types: List[Dict[str, Any]] = []
        self.uoms: List[Dict[str, Any]] = []
    
        self.object_db_id: Optional[int] = None
        self.plan_info: Optional[Dict[str, Any]] = None
        self.plan_id: Optional[int] = None
    
        self.tasks: List[Dict[str, Any]] = []
        self.tasks_filtered: List[Dict[str, Any]] = []
        self.facts: Dict[int, float] = {}
        self.fact_info: Dict[int, Dict[str, Any]] = {}
    
        self.registry_rows: List[Dict[str, Any]] = []
        self.registry_filtered: List[Dict[str, Any]] = []
    
        self._new_task_counter = 0
    
        q = _quarter_range()
        self.range_from: date = q[0]
        self.range_to: date = q[1]
    
        self._build_ui()
        self._load_refs()
        self._refresh_registry()
        self._update_range_label()

    def _build_ui(self):
        hdr = tk.Frame(self, bg=C["accent"], pady=6)
        hdr.pack(fill="x")

        tk.Label(
            hdr,
            text="📊  ГПР — График производства работ",
            font=("Segoe UI", 12, "bold"),
            bg=C["accent"],
            fg="white",
            padx=12,
        ).pack(side="left")

        self.lbl_plan_info = tk.Label(
            hdr,
            text="",
            font=("Segoe UI", 8),
            bg=C["accent"],
            fg="#bbdefb",
            padx=12,
        )
        self.lbl_plan_info.pack(side="right")

        self.nb_main = ttk.Notebook(self)
        self.nb_main.pack(fill="both", expand=True, padx=0, pady=0)

        self.tab_registry = tk.Frame(self.nb_main, bg=C["bg"])
        self.tab_editor = tk.Frame(self.nb_main, bg=C["bg"])
        self.tab_planning = tk.Frame(self.nb_main, bg=C["bg"])

        self.nb_main.add(self.tab_registry, text="  📚 Реестр ГПР  ")
        self.nb_main.add(self.tab_editor, text="  🛠 Редактор ГПР  ")
        self.nb_main.add(
            self.tab_planning,
            text="  📅 Планирование  ",
        )

        self.planning_panel = GprPlanningPanel(
            self.tab_planning,
            get_tasks_callback=self._get_tasks_for_planning,
            on_saved_callback=self._after_planning_saved,
        )
        self.planning_panel.pack(fill="both", expand=True)
        self._build_registry_tab(self.tab_registry)
        self._build_editor_tab(self.tab_editor)

    # ══════════════════════════════════════════════════════
    #  REGISTRY TAB
    # ══════════════════════════════════════════════════════
    def _build_registry_tab(self, parent):
        top = tk.LabelFrame(
            parent,
            text=" 🔎 Поиск и фильтрация ",
            font=("Segoe UI", 9, "bold"),
            bg=C["panel"],
            fg=C["accent"],
            relief="groove",
            bd=1,
            padx=10,
            pady=8,
        )
        top.pack(fill="x", padx=10, pady=(8, 4))

        row1 = tk.Frame(top, bg=C["panel"])
        row1.pack(fill="x")

        tk.Label(
            row1, text="Поиск:", bg=C["panel"], font=("Segoe UI", 9)
        ).pack(side="left")
        self.var_registry_search = tk.StringVar()
        ent = ttk.Entry(row1, textvariable=self.var_registry_search, width=40)
        ent.pack(side="left", padx=(6, 12))
        ent.bind("<KeyRelease>", lambda _e: self._apply_registry_filter())

        tk.Label(
            row1, text="Фильтр:", bg=C["panel"], font=("Segoe UI", 9)
        ).pack(side="left")
        self.cmb_registry_filter = ttk.Combobox(
            row1,
            state="readonly",
            width=20,
            values=[
                "Все объекты",
                "Только с ГПР",
                "Без ГПР",
                "С просрочкой",
            ],
        )
        self.cmb_registry_filter.pack(side="left", padx=(6, 12))
        self.cmb_registry_filter.current(0)
        self.cmb_registry_filter.bind(
            "<<ComboboxSelected>>",
            lambda _e: self._apply_registry_filter(),
        )

        ttk.Button(
            row1, text="Обновить", command=self._refresh_registry
        ).pack(side="right", padx=2)

        ttk.Button(
            row1, text="Открыть выбранный", command=self._open_selected_registry
        ).pack(side="right", padx=2)

        self.lbl_registry_summary = tk.Label(
            parent,
            text="",
            bg=C["bg"],
            fg=C["text2"],
            font=("Segoe UI", 8),
            anchor="w",
        )
        self.lbl_registry_summary.pack(fill="x", padx=14, pady=(2, 0))

        wrap = tk.Frame(parent, bg=C["panel"])
        wrap.pack(fill="both", expand=True, padx=10, pady=(4, 8))

        cols = (
            "excel_id",
            "short_name",
            "address",
            "object_status",
            "version_no",
            "updated_at",
            "creator_name",
            "task_count",
            "done_count",
            "overdue_count",
        )
        self.registry_tree = ttk.Treeview(
            wrap, columns=cols, show="headings", selectmode="browse"
        )

        heads = {
            "excel_id": ("Код", 80),
            "short_name": ("Краткое имя", 180),
            "address": ("Адрес", 420),
            "object_status": ("Статус объекта", 110),
            "version_no": ("Версия", 60),
            "updated_at": ("Обновлён", 120),
            "creator_name": ("Создал", 140),
            "task_count": ("Работ", 70),
            "done_count": ("Выполнено", 85),
            "overdue_count": ("Просрочено", 90),
        }

        for c, (t, w) in heads.items():
            self.registry_tree.heading(c, text=t)
            anc = "center" if c in (
                "excel_id", "version_no", "task_count", "done_count", "overdue_count"
            ) else "w"
            self.registry_tree.column(c, width=w, anchor=anc)

        vsb = ttk.Scrollbar(
            wrap, orient="vertical", command=self.registry_tree.yview
        )
        hsb = ttk.Scrollbar(
            parent.winfo_toplevel() if False else wrap,
            orient="horizontal",
            command=self.registry_tree.xview,
        )
        self.registry_tree.configure(
            yscrollcommand=vsb.set,
            xscrollcommand=hsb.set,
        )

        self.registry_tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")

        self.registry_tree.bind(
            "<Double-1>", lambda _e: self._open_selected_registry()
        )
        self.registry_tree.bind(
            "<Return>", lambda _e: self._open_selected_registry()
        )

        self.registry_tree.tag_configure("no_plan", foreground="#777")
        self.registry_tree.tag_configure("overdue", background="#fff3f3")

    def _refresh_registry(self):
        try:
            self.registry_rows = GprService.load_gpr_registry()
        except Exception as e:
            logger.exception("GPR registry load error")
            messagebox.showerror(
                "ГПР", f"Ошибка загрузки реестра ГПР:\n{e}", parent=self
            )
            return

        self._apply_registry_filter()

    def _get_tasks_for_planning(self):
        """
        Возвращает список задач открытого ГПР.
    
        ВАЖНО:
        замените self.tasks на фактическое имя списка,
        в котором ваш GprPage хранит задачи текущего графика.
        """
        return list(getattr(self, "tasks", []) or [])
    
    
    def _after_planning_saved(self):
        """
        Вызывается после сохранения недельного плана.
    
        Недельное планирование не изменяет даты и общий объём задачи,
        поэтому редактор ГПР можно не перезагружать из БД.
        Но обновляем текущую отрисовку и статус.
        """
        try:
            self._apply_filter()
            self._update_summary()
    
            self.lbl_bottom.config(
                text=(
                    f"Объект: {self.object_db_id or '—'}  |  "
                    "Недельный план сохранён."
                )
            )
        except Exception:
            logger.exception(
                "Не удалось обновить интерфейс после сохранения недельного плана"
            )
    
    def _apply_registry_filter(self):
        q = (self.var_registry_search.get() or "").strip().lower()
        mode = self.cmb_registry_filter.get().strip()

        res = []
        for r in self.registry_rows:
            has_plan = bool(r.get("plan_id"))
            overdue = int(r.get("overdue_count") or 0)

            if mode == "Только с ГПР" and not has_plan:
                continue
            if mode == "Без ГПР" and has_plan:
                continue
            if mode == "С просрочкой" and overdue <= 0:
                continue

            if q:
                hay = " ".join([
                    str(r.get("excel_id") or ""),
                    str(r.get("short_name") or ""),
                    str(r.get("address") or ""),
                    str(r.get("creator_name") or ""),
                    str(r.get("object_status") or ""),
                ]).lower()
                if q not in hay:
                    continue

            res.append(r)

        self.registry_filtered = res
        self._render_registry()

    def _render_registry(self):
        self.registry_tree.delete(*self.registry_tree.get_children())

        total = len(self.registry_filtered)
        with_plan = 0
        overdue_cnt = 0

        for r in self.registry_filtered:
            oid = int(r["object_db_id"])
            iid = f"obj_{oid}"

            plan_id = r.get("plan_id")
            if plan_id:
                with_plan += 1

            overdue = int(r.get("overdue_count") or 0)
            if overdue > 0:
                overdue_cnt += 1

            upd = r.get("updated_at")
            if isinstance(upd, datetime):
                upd_s = upd.strftime("%d.%m.%Y %H:%M")
            else:
                upd_s = ""

            tags = []
            if not plan_id:
                tags.append("no_plan")
            if overdue > 0:
                tags.append("overdue")

            self.registry_tree.insert(
                "",
                "end",
                iid=iid,
                values=(
                    r.get("excel_id") or "",
                    r.get("short_name") or "",
                    r.get("address") or "",
                    r.get("object_status") or "",
                    r.get("version_no") or "",
                    upd_s,
                    r.get("creator_name") or "",
                    r.get("task_count") or 0,
                    r.get("done_count") or 0,
                    overdue,
                ),
                tags=tuple(tags),
            )

        self.lbl_registry_summary.config(
            text=(
                f"Объектов: {total}  |  "
                f"С ГПР: {with_plan}  |  "
                f"С просрочкой: {overdue_cnt}"
            )
        )

    def _open_selected_registry(self):
        sel = self.registry_tree.selection()
        if not sel:
            messagebox.showinfo(
                "ГПР", "Выберите объект в реестре.", parent=self
            )
            return

        iid = sel[0]
        if not iid.startswith("obj_"):
            return

        try:
            oid = int(iid[4:])
        except ValueError:
            return

        self._open_object_by_id(oid)

    def _open_fact_batch(self):
        if not self.plan_id:
            messagebox.showinfo("ГПР", "Сначала откройте объект.", parent=self)
            return
    
        task_rows = [
            t for t in self.tasks
            if (t.get("row_kind") or "task") == "task" and t.get("id")
        ]
        if not task_rows:
            messagebox.showinfo(
                "ГПР",
                "Нет сохранённых работ для массового ввода факта.",
                parent=self,
            )
            return
    
        uid = (self.app_ref.current_user or {}).get("id")
    
        try:
            from gpr_task_dialog import open_task_fact_batch_dialog
        except ImportError as e:
            logger.exception("Cannot import batch fact dialog")
            messagebox.showerror(
                "ГПР",
                f"Не удалось открыть диалог массового ввода факта:\n{e}",
                parent=self,
            )
            return
    
        result = open_task_fact_batch_dialog(
            self,
            tasks=self.tasks,
            user_id=uid,
            fact_date=_today(),
        )
    
        if not result or not result.get("saved"):
            return
    
        try:
            changed_ids = [
                int(x) for x in (result.get("changed_task_ids") or [])
                if x is not None
            ]
    
            if changed_ids:
                updated = GprService.load_task_fact_info(changed_ids)
                for tid in changed_ids:
                    if tid in updated:
                        self.fact_info[tid] = updated[tid]
                        self.facts[tid] = float(
                            updated[tid].get("fact_qty_total") or 0
                        )
                    else:
                        self.fact_info.pop(tid, None)
                        self.facts.pop(tid, None)
            else:
                tids = [
                    t["id"] for t in self.tasks
                    if t.get("id") and (t.get("row_kind") or "task") == "task"
                ]
                self.fact_info = GprService.load_task_fact_info(tids)
                self.facts = {
                    task_id: float(v.get("fact_qty_total") or 0)
                    for task_id, v in self.fact_info.items()
                }
    
            self._apply_filter()
            self._update_summary()
    
            messagebox.showinfo(
                "ГПР",
                f"Сохранено записей факта: {result.get('count', 0)}",
                parent=self,
            )
        except Exception as e:
            logger.exception("Refresh facts after batch save error")
            messagebox.showwarning(
                "ГПР",
                f"Факт сохранён, но не удалось обновить отображение:\n{e}",
                parent=self,
            )

    # ══════════════════════════════════════════════════════
    #  EDITOR TAB
    # ══════════════════════════════════════════════════════
    def _build_editor_tab(self, parent):
        top = tk.LabelFrame(
            parent,
            text=" 📍 Объект и диапазон ",
            font=("Segoe UI", 9, "bold"),
            bg=C["panel"],
            fg=C["accent"],
            relief="groove",
            bd=1,
            padx=10,
            pady=8,
        )
        top.pack(fill="x", padx=10, pady=(8, 4))
        top.grid_columnconfigure(1, weight=1)
    
        tk.Label(
            top, text="Объект:", bg=C["panel"], font=("Segoe UI", 9)
        ).grid(row=0, column=0, sticky="e", padx=(0, 6))
    
        self.cmb_obj = _AutoCombo(top, width=60, font=("Segoe UI", 9))
        self.cmb_obj.grid(row=0, column=1, sticky="ew", pady=3)
    
        btn_f = tk.Frame(top, bg=C["panel"])
        btn_f.grid(row=0, column=2, padx=(8, 0))
        self._accent_btn(btn_f, "▶  Открыть", self._open_object)
    
        tk.Label(
            top, text="Диапазон:", bg=C["panel"], font=("Segoe UI", 9)
        ).grid(row=1, column=0, sticky="e", padx=(0, 6))
    
        range_f = tk.Frame(top, bg=C["panel"])
        range_f.grid(row=1, column=1, sticky="w", pady=3)
    
        self.lbl_range = tk.Label(
            range_f,
            text="",
            bg=C["panel"],
            fg=C["text2"],
            font=("Segoe UI", 9),
        )
        self.lbl_range.pack(side="left")
    
        ttk.Button(
            range_f, text="Изменить…", command=self._change_range
        ).pack(side="left", padx=(12, 0))
        ttk.Button(
            range_f, text="По работам", command=self._fit_range
        ).pack(side="left", padx=(6, 0))
    
        style = ttk.Style(self)
        style.configure(
            "GPR.Toolbar.TButton",
            font=("Segoe UI", 8),
            padding=(5, 2),
        )

        bar = tk.Frame(parent, bg=C["accent_light"], pady=2)
        bar.pack(fill="x", padx=10)

        tools_left = tk.Frame(bar, bg=C["accent_light"])
        tools_left.pack(side="left", fill="x", expand=True)

        tools_save = tk.Frame(bar, bg=C["accent_light"])
        tools_save.pack(side="right", fill="y", padx=(8, 4))

        row1 = tk.Frame(tools_left, bg=C["accent_light"])
        row1.pack(anchor="w", fill="x", pady=(0, 1))

        row2 = tk.Frame(tools_left, bg=C["accent_light"])
        row2.pack(anchor="w", fill="x", pady=(1, 0))

        def _grid_btn(row_frame, col, text, cmd):
            btn = self._tb_btn(
                row_frame,
                text,
                cmd,
                pack=False,
                style="GPR.Toolbar.TButton",
            )
            btn.grid(row=0, column=col, sticky="w", padx=(0, 4), pady=0)
            return btn

        def _grid_sep(row_frame, col):
            sep = tk.Frame(row_frame, bg=C["border"], width=1, height=22)
            sep.grid(row=0, column=col, sticky="ns", padx=(3, 7), pady=1)
            return sep

        # 1-я строка: основные операции редактирования.
        c = 0
        self.btn_add = _grid_btn(row1, c, "➕ Работа", self._add_task)
        c += 1

        self.btn_group = _grid_btn(row1, c, "📁 Группа", self._add_group)
        c += 1

        self.btn_title = _grid_btn(row1, c, "🟦 Титул", self._add_title)
        c += 1

        _grid_sep(row1, c)
        c += 1

        self.btn_edit = _grid_btn(row1, c, "✏️ Правка", self._edit_selected)
        c += 1

        self.btn_delete = _grid_btn(row1, c, "🗑 Удалить", self._delete_selected)
        c += 1

        _grid_sep(row1, c)
        c += 1

        self.btn_up = _grid_btn(row1, c, "⬆ Вверх", self._move_selected_up)
        c += 1

        self.btn_down = _grid_btn(row1, c, "⬇ Вниз", self._move_selected_down)
        c += 1

        _grid_sep(row1, c)
        c += 1

        _grid_btn(row1, c, "🔍−", lambda: self._zoom(-2))
        c += 1

        _grid_btn(row1, c, "🔍+", lambda: self._zoom(2))
        c += 1

        # 2-я строка: шаблоны, факт, отчёты и Excel.
        c = 0
        self.btn_template = _grid_btn(row2, c, "📋 Шаблон", self._apply_template)
        c += 1

        self.btn_fact_batch = _grid_btn(row2, c, "📈 Факт", self._open_fact_batch)
        c += 1

        self.btn_period_slice = _grid_btn(row2, c, "📊 Срез", self._export_period_slice)
        c += 1

        _grid_sep(row2, c)
        c += 1

        self.btn_import = _grid_btn(row2, c, "📤 Импорт", self._import_excel)
        c += 1

        self.btn_export = _grid_btn(row2, c, "📥 Экспорт", self._export_excel)
        c += 1

        self.btn_save = self._accent_btn(
            tools_save,
            "💾  СОХРАНИТЬ",
            self._save,
            pack=False,
            compact=True,
        )
        self.btn_save.pack(side="right", fill="y", ipady=5)

        fbar = tk.Frame(parent, bg=C["bg"], pady=4)
        fbar.pack(fill="x", padx=10)
    
        tk.Label(
            fbar, text="Фильтр тип:", bg=C["bg"], font=("Segoe UI", 8)
        ).pack(side="left")
        self.cmb_filt_wt = ttk.Combobox(
            fbar, state="readonly", width=20, values=["Все"]
        )
        self.cmb_filt_wt.pack(side="left", padx=(4, 12))
        self.cmb_filt_wt.current(0)
        self.cmb_filt_wt.bind(
            "<<ComboboxSelected>>", lambda _e: self._apply_filter()
        )
    
        tk.Label(
            fbar, text="Статус:", bg=C["bg"], font=("Segoe UI", 8)
        ).pack(side="left")
        self.cmb_filt_st = ttk.Combobox(
            fbar,
            state="readonly",
            width=16,
            values=["Все"] + [STATUS_LABELS[s] for s in STATUS_LIST],
        )
        self.cmb_filt_st.pack(side="left", padx=(4, 12))
        self.cmb_filt_st.current(0)
        self.cmb_filt_st.bind(
            "<<ComboboxSelected>>", lambda _e: self._apply_filter()
        )
    
        tk.Label(
            fbar, text="Поиск:", bg=C["bg"], font=("Segoe UI", 8)
        ).pack(side="left")
        self.var_search = tk.StringVar()
        ent_s = ttk.Entry(fbar, textvariable=self.var_search, width=24)
        ent_s.pack(side="left", padx=(4, 0))
        ent_s.bind("<KeyRelease>", lambda _e: self._apply_filter())
    
        self.lbl_summary = tk.Label(
            parent,
            text="",
            bg=C["bg"],
            font=("Segoe UI", 8),
            fg=C["text2"],
            anchor="w",
        )
        self.lbl_summary.pack(fill="x", padx=14, pady=(2, 0))
    
        leg = tk.Frame(parent, bg=C["bg"])
        leg.pack(fill="x", padx=14, pady=(0, 2))
        for code in STATUS_LIST:
            col, _, label = STATUS_COLORS[code]
            f = tk.Frame(leg, bg=C["bg"])
            f.pack(side="left", padx=(0, 12))
            tk.Canvas(
                f,
                width=12,
                height=12,
                bg=col,
                highlightthickness=1,
                highlightbackground="#999",
            ).pack(side="left", padx=(0, 3))
            tk.Label(
                f,
                text=label,
                bg=C["bg"],
                font=("Segoe UI", 7),
                fg=C["text2"],
            ).pack(side="left")
    
        pw = tk.PanedWindow(
            parent, orient="horizontal", sashrelief="raised", bg=C["bg"]
        )
        pw.pack(fill="both", expand=True, padx=10, pady=(4, 4))
    
        left = tk.Frame(pw, bg=C["panel"])
        right = tk.Frame(pw, bg=C["panel"])
        pw.add(left, minsize=560)
        pw.add(right, minsize=400)
    
        self.tree_top_spacer = tk.Frame(
            left,
            bg="#d6dbe0",
            height=GanttCanvas.MONTH_H,
            highlightthickness=0,
            bd=0,
        )
        self.tree_top_spacer.pack(side="top", fill="x")
        self.tree_top_spacer.pack_propagate(False)
    
        tree_wrap = tk.Frame(left, bg=C["panel"])
        tree_wrap.pack(side="top", fill="both", expand=True)
    
        style = ttk.Style(self)
        style.configure("GPR.Treeview", rowheight=24)
        style.configure("GPR.Treeview.Heading", font=("Segoe UI", 9, "bold"))
    
        cols = ("type", "name", "start", "finish", "uom", "qty", "labor", "workers", "status")
        self.tree = ttk.Treeview(
            tree_wrap,
            columns=cols,
            show="headings",
            selectmode="browse",
            style="GPR.Treeview",
        )
    
        heads = {
            "type": ("Тип работ", 130),
            "name": ("Вид работ", 260),
            "start": ("Начало", 85),
            "finish": ("Конец", 85),
            "uom": ("Ед.", 50),
            "qty": ("Объём", 75),
            "labor": ("ЗТР", 70),
            "workers": ("Людей", 70),
            "status": ("Статус", 100),
        }
    
        for c, (t, w) in heads.items():
            self.tree.heading(c, text=t)
            anc = (
                "center"
                if c in ("start", "finish", "uom", "workers", "status")
                else ("e" if c in ("qty", "labor") else "w")
            )
            self.tree.column(c, width=w, anchor=anc)
    
        vsb = ttk.Scrollbar(
            tree_wrap, orient="vertical", command=self._on_tree_scroll
        )
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
    
        self.tree.bind("<Double-1>", lambda _e: self._edit_selected())
        self.tree.bind("<Return>", lambda _e: self._edit_selected())
        self.tree.bind("<MouseWheel>", self._on_tree_wheel)
        self.tree.bind("<Button-4>", self._on_tree_wheel)
        self.tree.bind("<Button-5>", self._on_tree_wheel)
        self.tree.bind("<Control-Up>", lambda _e: self._move_selected_up())
        self.tree.bind("<Control-Down>", lambda _e: self._move_selected_down())
    
        self.gantt = GanttCanvas(right, day_px=20, linked_tree=self.tree)
        self.gantt.pack(fill="both", expand=True)
    
        self.after_idle(self._sync_tree_header_spacer)
    
        bot = tk.Frame(parent, bg=C["border"], height=1)
        bot.pack(fill="x", padx=10)
    
        self.lbl_bottom = tk.Label(
            parent,
            text="Выберите объект в реестре или откройте его вручную",
            bg=C["bg"],
            fg=C["text3"],
            font=("Segoe UI", 8),
            anchor="w",
            padx=14,
            pady=2,
        )
        self.lbl_bottom.pack(fill="x", padx=0, pady=(0, 6))
    # ══════════════════════════════════════════════════════
    #  COMMON
    # ══════════════════════════════════════════════════════
    def _sync_tree_header_spacer(self):
        try:
            self.update_idletasks()

            items = self.tree.get_children()
            if not items:
                self.tree_top_spacer.config(height=GanttCanvas.MONTH_H)
                return

            first_bbox = None
            for iid in items:
                try:
                    bb = self.tree.bbox(iid)
                    if bb:
                        first_bbox = bb
                        break
                except tk.TclError:
                    continue

            if not first_bbox:
                self.tree_top_spacer.config(height=GanttCanvas.MONTH_H)
                return

            tree_header_h = max(0, int(first_bbox[1]))
            spacer_h = max(0, int(self.gantt.HEADER_H) - tree_header_h)
            self.tree_top_spacer.config(height=spacer_h)

        except Exception:
            logger.exception("Error syncing tree/gantt header heights")
            self.tree_top_spacer.config(height=GanttCanvas.MONTH_H)

    def _accent_btn(self, parent, text, cmd, *, pack=True, compact=False):
        b = tk.Button(
            parent,
            text=text,
            font=("Segoe UI", 8 if compact else 9, "bold"),
            bg=C["btn_bg"],
            fg=C["btn_fg"],
            activebackground="#0d47a1",
            activeforeground="white",
            relief="flat",
            cursor="hand2",
            padx=8 if compact else 10,
            pady=2 if compact else 3,
            command=cmd,
        )

        if pack:
            b.pack(side="left", padx=2)

        b.bind("<Enter>", lambda _e: b.config(bg="#0d47a1"))
        b.bind("<Leave>", lambda _e: b.config(bg=C["btn_bg"]))

        return b

    def _tb_btn(self, parent, text, cmd, *, pack=True, style=None):
        kw = {
            "text": text,
            "command": cmd,
        }

        if style:
            kw["style"] = style

        b = ttk.Button(parent, **kw)

        if pack:
            b.pack(side="left", padx=2)

        return b

    def _on_tree_scroll(self, *args):
        self.tree.yview(*args)
        self.gantt.after_idle(self.gantt.redraw_bars_only)

    def _on_tree_wheel(self, event):
        d = _mouse_delta(event)
        if d:
            self.tree.yview_scroll(d, "units")
            self.gantt.after_idle(self.gantt.redraw_bars_only)
        return "break"

    def _load_refs(self):
        try:
            self.objects = GprService.load_objects_short()
            self.work_types = GprService.load_work_types()
            self.uoms = GprService.load_uoms()
        except Exception as e:
            logger.exception("GPR refs error")
            messagebox.showerror(
                "ГПР", f"Ошибка загрузки справочников:\n{e}", parent=self
            )
            return

        vals = []
        for o in self.objects:
            sn = (o.get("short_name") or "").strip()
            addr = (o.get("address") or "").strip()
            eid = str(o.get("excel_id") or "").strip()
            db_id = str(o.get("id") or "")

            tag = f"[{eid}]" if eid else f"[id:{db_id}]"
            lbl = f"{sn} — {addr} — {tag}" if sn else f"{addr} — {tag}"
            vals.append(lbl)

        self.cmb_obj.set_values(vals)

        wt_names = ["Все"] + [w["name"] for w in self.work_types]
        self.cmb_filt_wt.config(values=wt_names)

    def _update_range_label(self):
        self.lbl_range.config(
            text=f"{_fmt_date(self.range_from)} — {_fmt_date(self.range_to)}"
        )
        self.gantt.set_range(self.range_from, self.range_to)

    def _update_plan_info(self):
        p = self.plan_info
        if not p:
            self.lbl_plan_info.config(text="")
            return

        cr = p.get("creator_name") or "—"
        upd = p.get("updated_at")
        if isinstance(upd, datetime):
            upd_s = upd.strftime("%d.%m.%Y %H:%M")
        else:
            upd_s = str(upd or "")

        v = p.get("version_no", 1)
        self.lbl_plan_info.config(
            text=f"Версия: {v}  |  Создал: {cr}  |  Обновлён: {upd_s}"
        )

    def _update_summary(self):
        total = len([t for t in self.tasks if (t.get("row_kind") or "task") == "task"])
        by_st: Dict[str, int] = {}
        overdue = 0
        td = _today()

        for t in self.tasks:
            if (t.get("row_kind") or "task") != "task":
                continue
            st = t.get("status", "planned")
            by_st[st] = by_st.get(st, 0) + 1
            if st not in ("done", "canceled"):
                pf = _to_date(t.get("plan_finish"))
                if pf and pf < td:
                    overdue += 1

        groups = sum(1 for t in self.tasks if (t.get("row_kind") or "task") == "group")
        titles = sum(1 for t in self.tasks if (t.get("row_kind") or "task") == "title")

        parts = [f"Работ: {total}"]
        if groups:
            parts.append(f"Групп: {groups}")
        if titles:
            parts.append(f"Титулов: {titles}")

        for s in STATUS_LIST:
            cnt = by_st.get(s, 0)
            if cnt > 0:
                parts.append(f"{STATUS_LABELS[s]}: {cnt}")
        if overdue > 0:
            parts.append(f"⚠ Просрочено: {overdue}")

        self.lbl_summary.config(text="  |  ".join(parts))

    def _sel_obj_id(self) -> Optional[int]:
        idx = self.cmb_obj.get_original_index()
        if idx < 0 or idx >= len(self.objects):
            return None
        return int(self.objects[idx]["id"])

    def _recalc_sort_order(self):
        for i, t in enumerate(self.tasks):
            t["sort_order"] = i * 10

    def _preserve_selection_by_task(self, task_ref):
        if task_ref is None:
            return
        for iid in self.tree.get_children():
            try:
                idx = self.tree.index(iid)
            except tk.TclError:
                continue
            if 0 <= idx < len(self.tasks_filtered) and self.tasks_filtered[idx] is task_ref:
                self.tree.selection_set(iid)
                self.tree.focus(iid)
                self.tree.see(iid)
                break

    # ══════════════════════════════════════════════════════
    #  OPEN OBJECT
    # ══════════════════════════════════════════════════════
    def _open_object(self):
        oid = self._sel_obj_id()
        if not oid:
            messagebox.showwarning(
                "ГПР", "Выберите объект из списка.", parent=self
            )
            return
        self._open_object_by_id(oid)

    def _open_object_by_id(self, oid: int):
        self.object_db_id = oid
        uid = (self.app_ref.current_user or {}).get("id")
    
        try:
            self.plan_info = GprService.get_or_create_current_plan(oid, uid)
            self.plan_id = int(self.plan_info["id"])
            self.tasks = GprService.load_plan_tasks(self.plan_id)
    
            for t in self.tasks:
                t["row_kind"] = (t.get("row_kind") or "task").strip()
    
            tids = [
                t["id"] for t in self.tasks
                if t.get("id") and (t.get("row_kind") or "task") == "task"
            ]
    
            self.fact_info = GprService.load_task_fact_info(tids)
            self.facts = {
                task_id: float(v.get("fact_qty_total") or 0)
                for task_id, v in self.fact_info.items()
            }
    
        except Exception as e:
            logger.exception("GPR open error")
            messagebox.showerror(
                "ГПР", f"Не удалось открыть ГПР:\n{e}", parent=self
            )
            return
    
        obj = next((o for o in self.objects if int(o["id"]) == oid), None)
        if obj:
            sn = (obj.get("short_name") or "").strip()
            addr = (obj.get("address") or "").strip()
            name = sn if sn else addr
    
            label = None
            for i, o in enumerate(self.objects):
                if int(o["id"]) == oid:
                    eid = str(o.get("excel_id") or "").strip()
                    tag = f"[{eid}]" if eid else f"[id:{oid}]"
                    label = f"{sn} — {addr} — {tag}" if sn else f"{addr} — {tag}"
                    if label:
                        self.cmb_obj.set(label)
                    break
        else:
            name = str(oid)
    
        self._update_plan_info()
        self._apply_filter()
        self._update_summary()
        
        if hasattr(self, "planning_panel"):
            try:
                self.planning_panel.reload()
            except Exception:
                logger.exception(
                    "Не удалось обновить вкладку планирования после открытия объекта"
                )
        
        self.lbl_bottom.config(
            text=f"Объект: {name}  |  Строк: {len(self.tasks)}"
        )
        
        self.nb_main.select(self.tab_editor)

    # ══════════════════════════════════════════════════════
    #  FILTER / RENDER
    # ══════════════════════════════════════════════════════
    def _apply_filter(self):
        wt_idx = self.cmb_filt_wt.current()
        wt_name = None
        if wt_idx > 0 and wt_idx <= len(self.work_types):
            wt_name = self.work_types[wt_idx - 1]["name"]

        st_idx = self.cmb_filt_st.current()
        st_code = None
        if st_idx > 0 and st_idx <= len(STATUS_LIST):
            st_code = STATUS_LIST[st_idx - 1]

        q = (self.var_search.get() or "").strip().lower()

        res = []
        for t in self.tasks:
            row_kind = (t.get("row_kind") or "task").strip()

            if row_kind in ("group", "title"):
                res.append(t)
                continue

            if wt_name and (t.get("work_type_name") or "") != wt_name:
                continue
            if st_code and (t.get("status") or "") != st_code:
                continue
            if q:
                nm = (t.get("name") or "").lower()
                wtn = (t.get("work_type_name") or "").lower()
                if q not in nm and q not in wtn:
                    continue
            res.append(t)

        self.tasks_filtered = res
        self._render()

    def _gen_iid(self, task: Dict[str, Any]) -> str:
        tid = task.get("id")
        if tid is not None:
            return f"db_{tid}"
        self._new_task_counter += 1
        return f"new_{self._new_task_counter}"

    def _render(self):
        self.tree.delete(*self.tree.get_children())
    
        for t in self.tasks_filtered:
            iid = self._gen_iid(t)
            row_kind = (t.get("row_kind") or "task").strip()
    
            if row_kind == "group":
                values = ("", f"📁 {t.get('name', '')}", "", "", "", "", "", "", "")
                self.tree.insert("", "end", iid=iid, values=values, tags=("group",))
            elif row_kind == "title":
                values = ("", f"🟦 {t.get('name', '')}", "", "", "", "", "", "", "")
                self.tree.insert("", "end", iid=iid, values=values, tags=("title",))
            else:
                st_label = STATUS_LABELS.get(
                    t.get("status", ""), t.get("status", "")
                )
    
                tid = t.get("id")
                workers_last = ""
                if tid is not None:
                    info = self.fact_info.get(tid) or {}
                    if info.get("workers_last") is not None:
                        workers_last = str(info["workers_last"])
    
                self.tree.insert(
                    "",
                    "end",
                    iid=iid,
                    values=(
                        t.get("work_type_name", ""),
                        t.get("name", ""),
                        _fmt_date(t.get("plan_start")),
                        _fmt_date(t.get("plan_finish")),
                        t.get("uom_code") or "",
                        _fmt_qty(t.get("plan_qty")),
                        _fmt_qty(t.get("labor_hours_per_unit")),
                        workers_last,
                        st_label,
                    ),
                    tags=("task",),
                )
    
        self.tree.tag_configure("group", font=("Segoe UI", 9, "bold"))
        self.tree.tag_configure(
            "title",
            font=("Segoe UI", 9, "bold"),
            background="#e3f2fd",
        )
    
        self.gantt.set_data(self.tasks_filtered, self.facts, self.fact_info)
        self.after_idle(self._sync_tree_header_spacer)
        self.after_idle(self.gantt.redraw_bars_only)

    # ══════════════════════════════════════════════════════
    #  RANGE / ZOOM
    # ══════════════════════════════════════════════════════
    def _change_range(self):
        dlg = DateRangeDialog(self, self.range_from, self.range_to)
        if dlg.result:
            self.range_from, self.range_to = dlg.result
            self._update_range_label()
            self.gantt.set_data(self.tasks_filtered, self.facts, self.fact_info)

    def _fit_range(self):
        task_rows = [t for t in self.tasks if (t.get("row_kind") or "task") == "task"]
        if not task_rows:
            messagebox.showinfo(
                "ГПР", "Нет работ для определения диапазона.", parent=self
            )
            return
    
        starts = [
            _to_date(t["plan_start"])
            for t in task_rows
            if _to_date(t.get("plan_start"))
        ]
        finishes = [
            _to_date(t["plan_finish"])
            for t in task_rows
            if _to_date(t.get("plan_finish"))
        ]
    
        if not starts or not finishes:
            messagebox.showinfo(
                "ГПР", "Нет работ с валидными датами.", parent=self
            )
            return
    
        d0 = min(starts)
        d1 = max(finishes)
        self.range_from = d0 - timedelta(days=7)
        self.range_to = d1 + timedelta(days=7)
        self._update_range_label()
        self.gantt.set_data(self.tasks_filtered, self.facts, self.fact_info)

    def _zoom(self, delta):
        self.gantt.day_px = max(6, min(50, self.gantt.day_px + delta))
        self.gantt._header_cache_key = None
        self.gantt._schedule_heavy_redraw()

    # ══════════════════════════════════════════════════════
    #  CRUD
    # ══════════════════════════════════════════════════════
    def _find_task_idx(self) -> Optional[int]:
        sel = self.tree.selection()
        if not sel:
            return None
        iid = sel[0]

        if iid.startswith("db_"):
            try:
                tid = int(iid[3:])
                for i, t in enumerate(self.tasks):
                    if t.get("id") is not None and int(t["id"]) == tid:
                        return i
            except (ValueError, TypeError):
                pass

        try:
            tree_idx = self.tree.index(iid)
            if 0 <= tree_idx < len(self.tasks_filtered):
                task_ref = self.tasks_filtered[tree_idx]
                for i, t in enumerate(self.tasks):
                    if t is task_ref:
                        return i
        except (tk.TclError, ValueError):
            pass

        return None

    def _get_insert_index_after_selection(self) -> int:
        """
        Возвращает индекс, куда нужно вставить новую строку:
        - если строка выделена — сразу после неё;
        - если ничего не выделено — в конец списка.
        """
        idx = self._find_task_idx()
        if idx is None:
            return len(self.tasks)
        return idx + 1

    def _open_task_dialog(self, init=None):
        uid = (self.app_ref.current_user or {}).get("id")

        if not hasattr(self, "_ext_dialog_func"):
            try:
                from gpr_task_dialog import open_task_dialog as _func
                self._ext_dialog_func = _func
                logger.info("gpr_task_dialog loaded successfully")
            except ImportError as e:
                logger.warning("gpr_task_dialog not available: %s", e)
                self._ext_dialog_func = None

        if self._ext_dialog_func is not None:
            try:
                return self._ext_dialog_func(
                    self, self.work_types, self.uoms,
                    init=init, user_id=uid
                )
            except Exception:
                logger.exception(
                    "External task dialog error, falling back to built-in"
                )

        logger.warning("Using built-in TaskEditDialog (no gpr_task_dialog)")
        dlg = TaskEditDialog(self, self.work_types, self.uoms, init=init)
        return dlg.result

    def _prepare_new_task_row(
        self,
        data: Dict[str, Any],
        *,
        fallback_start: Optional[date] = None,
        fallback_finish: Optional[date] = None,
    ) -> Dict[str, Any]:
        """
        Готовит новую строку задачи для self.tasks.

        Используется и для одной работы, и для массового выбора
        работ из профессионального справочника.
        """
        t = dict(data)

        work_type_id = t.get("work_type_id")
        if work_type_id is None:
            raise ValueError("Не указан тип работ")

        work_type_id = int(work_type_id)

        t["id"] = None
        t["parent_id"] = t.get("parent_id")
        t["row_kind"] = "task"
        t["work_type_id"] = work_type_id

        t["work_type_name"] = next(
            (
                w["name"]
                for w in self.work_types
                if int(w["id"]) == work_type_id
            ),
            "",
        )

        t["plan_start"] = (
            _to_date(t.get("plan_start"))
            or fallback_start
            or _today()
        )

        t["plan_finish"] = (
            _to_date(t.get("plan_finish"))
            or fallback_finish
            or t["plan_start"]
        )

        if t["plan_finish"] < t["plan_start"]:
            raise ValueError(
                f"Для работы «{t.get('name') or '—'}» "
                "дата окончания раньше даты начала."
            )

        t["status"] = (
            t.get("status")
            if t.get("status") in STATUS_LIST
            else "planned"
        )

        t["is_milestone"] = bool(t.get("is_milestone"))

        t["plan_qty"] = _safe_float(t.get("plan_qty"))

        # Данные профессионального справочника.
        if t.get("work_item_id") is not None:
            try:
                t["work_item_id"] = int(t["work_item_id"])
            except (TypeError, ValueError):
                t["work_item_id"] = None
        else:
            t["work_item_id"] = None

        if t.get("labor_norm_id") is not None:
            try:
                t["labor_norm_id"] = int(t["labor_norm_id"])
            except (TypeError, ValueError):
                t["labor_norm_id"] = None
        else:
            t["labor_norm_id"] = None

        t["labor_hours_per_unit"] = _safe_float(
            t.get("labor_hours_per_unit")
        )

        factor = _safe_float(t.get("productivity_factor"))
        t["productivity_factor"] = (
            factor
            if factor is not None and factor > 0
            else 1.0
        )

        return t
        
    def _add_task(self):
        if not self.plan_id:
            messagebox.showinfo(
                "ГПР",
                "Сначала откройте объект.",
                parent=self,
            )
            return

        selected_idx = self._find_task_idx()

        base_start = self.range_from
        base_finish = self.range_from

        if selected_idx is not None and 0 <= selected_idx < len(self.tasks):
            selected_task = self.tasks[selected_idx]

            base_start = (
                _to_date(selected_task.get("plan_start"))
                or self.range_from
            )

            base_finish = (
                _to_date(selected_task.get("plan_finish"))
                or base_start
            )

        result = self._open_task_dialog(
            init={
                "plan_start": base_start,
                "plan_finish": base_finish,
                "row_kind": "task",
            }
        )

        if not result:
            return

        insert_at = (
            selected_idx + 1
            if selected_idx is not None
            else len(self.tasks)
        )

        # ──────────────────────────────────────────────────
        # Множественный выбор работ из gpr_task_dialog.py.
        # ──────────────────────────────────────────────────
        bulk_tasks = result.get("_bulk_tasks")

        if bulk_tasks:
            prepared_tasks: List[Dict[str, Any]] = []

            try:
                for task_data in bulk_tasks:
                    prepared_tasks.append(
                        self._prepare_new_task_row(
                            task_data,
                            fallback_start=base_start,
                            fallback_finish=base_finish,
                        )
                    )
            except Exception as exc:
                messagebox.showwarning(
                    "ГПР",
                    f"Не удалось подготовить список работ:\n{exc}",
                    parent=self,
                )
                return

            if not prepared_tasks:
                messagebox.showinfo(
                    "ГПР",
                    "Не выбрано ни одной работы.",
                    parent=self,
                )
                return

            self.tasks[insert_at:insert_at] = prepared_tasks

            self._recalc_sort_order()
            self._apply_filter()
            self._update_summary()

            self._preserve_selection_by_task(prepared_tasks[0])

            self.lbl_bottom.config(
                text=(
                    f"Добавлено работ из справочника: "
                    f"{len(prepared_tasks)}. "
                    "Нажмите «СОХРАНИТЬ» для записи в БД."
                )
            )

            return

        # ──────────────────────────────────────────────────
        # Обычное добавление одной работы.
        # ──────────────────────────────────────────────────
        try:
            task = self._prepare_new_task_row(
                result,
                fallback_start=base_start,
                fallback_finish=base_finish,
            )
        except Exception as exc:
            messagebox.showwarning(
                "ГПР",
                f"Не удалось добавить работу:\n{exc}",
                parent=self,
            )
            return

        self.tasks.insert(insert_at, task)

        self._recalc_sort_order()
        self._apply_filter()
        self._update_summary()
        self._preserve_selection_by_task(task)
        
    def _add_group(self):
        if not self.plan_id:
            messagebox.showinfo("ГПР", "Сначала откройте объект.", parent=self)
            return
    
        name = simpledialog.askstring(
            "Группа",
            "Введите название группы:",
            parent=self,
        )
        if not name:
            return
    
        wt_id = int(self.work_types[0]["id"]) if self.work_types else 1
    
        t = {
            "id": None,
            "row_kind": "group",
            "work_type_id": wt_id,
            "work_type_name": "",
            "name": name.strip(),
            "uom_code": None,
            "plan_qty": None,
            "plan_start": self.range_from,
            "plan_finish": self.range_from,
            "status": "planned",
            "is_milestone": False,
            "sort_order": len(self.tasks) * 10,
        }
    
        insert_at = self._get_insert_index_after_selection()
        self.tasks.insert(insert_at, t)
        
        self._recalc_sort_order()
        self._apply_filter()
        self._update_summary()
        self._preserve_selection_by_task(t)

    def _add_title(self):
        if not self.plan_id:
            messagebox.showinfo("ГПР", "Сначала откройте объект.", parent=self)
            return
    
        name = simpledialog.askstring(
            "Титульная строка",
            "Введите текст титульной строки:",
            parent=self,
        )
        if not name:
            return
    
        wt_id = int(self.work_types[0]["id"]) if self.work_types else 1
    
        t = {
            "id": None,
            "row_kind": "title",
            "work_type_id": wt_id,
            "work_type_name": "",
            "name": name.strip(),
            "uom_code": None,
            "plan_qty": None,
            "plan_start": self.range_from,
            "plan_finish": self.range_from,
            "status": "planned",
            "is_milestone": False,
            "sort_order": len(self.tasks) * 10,
        }
    
        insert_at = self._get_insert_index_after_selection()
        self.tasks.insert(insert_at, t)
        
        self._recalc_sort_order()
        self._apply_filter()
        self._update_summary()
        self._preserve_selection_by_task(t)

    def _edit_selected(self):
        idx = self._find_task_idx()
        if idx is None:
            messagebox.showinfo(
                "ГПР", "Выберите работу для редактирования.", parent=self
            )
            return
    
        t0 = self.tasks[idx]
        row_kind = (t0.get("row_kind") or "task").strip()
    
        if row_kind in ("group", "title"):
            title = "Группа" if row_kind == "group" else "Титульная строка"
            name = simpledialog.askstring(
                title,
                "Введите новый текст:",
                initialvalue=t0.get("name", ""),
                parent=self,
            )
            if not name:
                return
            t0["name"] = name.strip()
            self._apply_filter()
            self._update_summary()
            return
    
        result = self._open_task_dialog(init=t0)
        if not result:
            return

        if result.get("_bulk_tasks"):
            messagebox.showwarning(
                "ГПР",
                "При редактировании одной существующей задачи "
                "нельзя добавить несколько работ.\n\n"
                "Для массового добавления используйте кнопку «➕ Работа».",
                parent=self,
            )
            return
    
        upd = dict(result)
        upd["id"] = t0.get("id")
        upd["row_kind"] = "task"
        upd["sort_order"] = t0.get("sort_order", idx * 10)
        upd["work_type_name"] = next(
            (
                w["name"]
                for w in self.work_types
                if int(w["id"]) == int(upd["work_type_id"])
            ),
            "",
        )
        upd["plan_start"] = _to_date(upd.get("plan_start")) or _today()
        upd["plan_finish"] = _to_date(upd.get("plan_finish")) or _today()
    
        task_id = t0.get("id")
        assignments = upd.pop("_assignments", None)
        facts_payload = upd.pop("_facts", None)
        facts_changed = bool(upd.pop("_facts_changed", False))
    
        uid = (self.app_ref.current_user or {}).get("id")
    
        if task_id and assignments is not None:
            try:
                from gpr_task_dialog import _EmployeeService
                _EmployeeService.save_task_assignments(task_id, assignments, uid)
            except ImportError:
                logger.warning(
                    "gpr_task_dialog not available — assignments not saved"
                )
            except Exception as e:
                logger.exception("Save assignments error")
                messagebox.showwarning(
                    "ГПР",
                    f"Ошибка сохранения назначений:\n{e}",
                    parent=self,
                )
    
        if task_id and facts_changed:
            try:
                from gpr_task_dialog import _TaskFactService
                _TaskFactService.save_task_facts(task_id, facts_payload or [], uid)
    
                fact_info_map = GprService.load_task_fact_info([task_id])
    
                if task_id in fact_info_map:
                    self.fact_info[task_id] = fact_info_map[task_id]
                    self.facts[task_id] = float(
                        fact_info_map[task_id].get("fact_qty_total") or 0
                    )
                else:
                    self.fact_info.pop(task_id, None)
                    self.facts.pop(task_id, None)
    
            except ImportError:
                logger.warning("gpr_task_dialog not available — facts not saved")
            except Exception as e:
                logger.exception("Save facts error")
                messagebox.showwarning(
                    "ГПР",
                    f"Ошибка сохранения факта:\n{e}",
                    parent=self,
                )
    
        self.tasks[idx] = upd
        self._apply_filter()
        self._update_summary()

    def _delete_selected(self):
        idx = self._find_task_idx()
        if idx is None:
            messagebox.showinfo(
                "ГПР", "Выберите строку для удаления.", parent=self
            )
            return
        row_name = self.tasks[idx].get("name", "")
        if not messagebox.askyesno(
            "ГПР",
            f"Удалить строку «{row_name}»?",
            parent=self,
        ):
            return
        self.tasks.pop(idx)
        self._recalc_sort_order()
        self._apply_filter()
        self._update_summary()

    def _move_selected_up(self):
        idx = self._find_task_idx()
        if idx is None:
            messagebox.showinfo("ГПР", "Выберите строку.", parent=self)
            return
        if idx <= 0:
            return

        task_ref = self.tasks[idx]
        self.tasks[idx - 1], self.tasks[idx] = self.tasks[idx], self.tasks[idx - 1]
        self._recalc_sort_order()
        self._apply_filter()
        self._update_summary()
        self._preserve_selection_by_task(task_ref)

    def _move_selected_down(self):
        idx = self._find_task_idx()
        if idx is None:
            messagebox.showinfo("ГПР", "Выберите строку.", parent=self)
            return
        if idx >= len(self.tasks) - 1:
            return

        task_ref = self.tasks[idx]
        self.tasks[idx], self.tasks[idx + 1] = self.tasks[idx + 1], self.tasks[idx]
        self._recalc_sort_order()
        self._apply_filter()
        self._update_summary()
        self._preserve_selection_by_task(task_ref)

    def _apply_template(self):
        if not self.plan_id:
            messagebox.showinfo(
                "ГПР", "Сначала откройте объект.", parent=self
            )
            return
        try:
            tpls = GprService.load_templates()
        except Exception as e:
            logger.exception("Load templates error")
            messagebox.showerror(
                "ГПР", f"Ошибка загрузки шаблонов:\n{e}", parent=self
            )
            return
        if not tpls:
            messagebox.showinfo("ГПР", "Шаблонов нет.", parent=self)
            return

        dlg = TemplateSelectDialog(self, tpls)
        if not dlg.result:
            return

        try:
            tt = GprService.load_template_tasks(dlg.result)
        except Exception as e:
            logger.exception("Load template tasks error")
            messagebox.showerror(
                "ГПР", f"Ошибка загрузки задач шаблона:\n{e}", parent=self
            )
            return
        if not tt:
            messagebox.showinfo(
                "ГПР", "В шаблоне нет задач.", parent=self
            )
            return
        if self.tasks and not messagebox.askyesno(
            "ГПР", "Заменить текущие работы шаблоном?", parent=self
        ):
            return

        base = self.range_from
        out = []
        for i, x in enumerate(tt):
            row_kind = (x.get("row_kind") or "task").strip()
            wid = int(x["work_type_id"])
            wn = next(
                (w["name"] for w in self.work_types if int(w["id"]) == wid),
                "",
            )

            out.append(
                dict(
                    id=None,
                    row_kind=row_kind,
                    work_type_id=wid,
                    work_type_name="" if row_kind != "task" else wn,

                    work_item_id=None,
                    labor_norm_id=None,
                    labor_hours_per_unit=None,
                    productivity_factor=1.0,

                    name=x["name"],
                    uom_code=x.get("uom_code"),
                    plan_qty=x.get("default_qty"),
                    plan_start=base,
                    plan_finish=base,
                    status="planned",
                    is_milestone=bool(x.get("is_milestone")),
                    sort_order=int(
                        x.get("sort_order")
                        if x.get("sort_order") is not None
                        else i * 10
                    ),
                )
            )
        self.tasks = out
        self._recalc_sort_order()
        self._apply_filter()
        self._update_summary()

    def _save(self):
        if not self.plan_id:
            messagebox.showinfo(
                "ГПР", "Сначала откройте объект.", parent=self
            )
            return
    
        errors = []
        for i, t in enumerate(self.tasks):
            row_kind = (t.get("row_kind") or "task").strip()
            name = (t.get("name") or "").strip()
    
            if not name:
                errors.append(f"Строка {i + 1}: нет названия")
                continue
    
            if row_kind != "task":
                continue
    
            ds = _to_date(t.get("plan_start"))
            df = _to_date(t.get("plan_finish"))
            if not ds or not df:
                errors.append(f"«{name}»: невалидные даты")
            elif df < ds:
                errors.append(f"«{name}»: окончание раньше начала")
    
        if errors:
            msg = "Ошибки валидации:\n\n" + "\n".join(errors[:10])
            if len(errors) > 10:
                msg += f"\n\n...и ещё {len(errors) - 10} ошибок"
            messagebox.showwarning("ГПР", msg, parent=self)
            return
    
        uid = (self.app_ref.current_user or {}).get("id")
        try:
            GprService.replace_plan_tasks(self.plan_id, uid, self.tasks)
            self.tasks = GprService.load_plan_tasks(self.plan_id)
            for t in self.tasks:
                t["row_kind"] = (t.get("row_kind") or "task").strip()
    
            tids = [
                t["id"] for t in self.tasks
                if t.get("id") and (t.get("row_kind") or "task") == "task"
            ]
    
            self.fact_info = GprService.load_task_fact_info(tids)
            self.facts = {
                task_id: float(v.get("fact_qty_total") or 0)
                for task_id, v in self.fact_info.items()
            }
    
            self.plan_info = GprService.get_or_create_current_plan(
                self.object_db_id, uid
            )
            self._update_plan_info()
            self._apply_filter()
            self._update_summary()
            if hasattr(self, "planning_panel"):
                try:
                    self.planning_panel.reload()
                except Exception:
                    logger.exception(
                        "Не удалось обновить планирование после сохранения ГПР"
                    )
            self._refresh_registry()
            messagebox.showinfo("ГПР", "Сохранено успешно.", parent=self)
        except Exception as e:
            logger.exception("GPR save error")
            messagebox.showerror(
                "ГПР", f"Ошибка сохранения:\n{e}", parent=self
            )

    @staticmethod
    def _xl_fill(color: str) -> PatternFill:
        """Excel fill из hex-цвета."""
        return PatternFill(
            "solid",
            fgColor=(color or "FFFFFF").replace("#", "").upper()
        )

    def _import_excel(self):
        if not self.plan_id:
            messagebox.showinfo(
                "ГПР",
                "Сначала откройте объект.",
                parent=self,
            )
            return
    
        if not HAS_OPENPYXL:
            messagebox.showwarning(
                "ГПР",
                "Для импорта необходима библиотека openpyxl.\n"
                "Установите: pip install openpyxl",
                parent=self,
            )
            return
    
        path = filedialog.askopenfilename(
            parent=self,
            title="Выбрать Excel-файл для импорта ГПР",
            filetypes=[("Excel", "*.xlsx"), ("Все файлы", "*.*")],
        )
        if not path:
            return
    
        try:
            result = GprExcelImportService.import_tasks_from_excel(
                path=path,
                work_types=self.work_types,
                uoms=self.uoms,
            )
        except Exception as e:
            logger.exception("GPR excel import error")
            messagebox.showerror(
                "ГПР",
                f"Ошибка чтения Excel:\n{e}",
                parent=self,
            )
            return
    
        errors = result.get("errors") or []
        tasks = result.get("tasks") or []
    
        if errors:
            preview = "\n".join(errors[:20])
            if len(errors) > 20:
                preview += f"\n... и ещё {len(errors) - 20} ошибок"
    
            messagebox.showwarning(
                "ГПР",
                "Импорт не выполнен, потому что в файле есть ошибки:\n\n"
                f"{preview}",
                parent=self,
            )
            return
    
        if not tasks:
            messagebox.showinfo(
                "ГПР",
                "В выбранном файле не найдено строк для импорта.",
                parent=self,
            )
            return
    
        msg = (
            f"Найдено строк для загрузки: {len(tasks)}\n"
            f"Лист: {result.get('sheet_name')}\n"
        )
        skipped_empty = int(result.get("skipped_empty") or 0)
        if skipped_empty:
            msg += f"Пустых строк пропущено: {skipped_empty}\n"
        msg += "\nЗаменить текущий список работ данными из Excel?"
    
        if self.tasks:
            ok = messagebox.askyesno("ГПР", msg, parent=self)
            if not ok:
                return
    
        self.tasks = tasks
        self._recalc_sort_order()
        self._apply_filter()
        self._update_summary()
    
        self.lbl_bottom.config(
            text=(
                f"Импортировано строк: {len(tasks)}  |  "
                "Проверьте данные и нажмите 'СОХРАНИТЬ'"
            )
        )
    
        messagebox.showinfo(
            "ГПР",
            f"Импорт выполнен успешно.\n\n"
            f"Загружено строк: {len(tasks)}\n"
            f"Из них:\n"
            f"  • работ: {sum(1 for t in tasks if (t.get('row_kind') or 'task') == 'task')}\n"
            f"  • групп: {sum(1 for t in tasks if (t.get('row_kind') or 'task') == 'group')}\n"
            f"  • титулов: {sum(1 for t in tasks if (t.get('row_kind') or 'task') == 'title')}\n\n"
            "Данные пока загружены только в редактор.\n"
            "Для записи в базу нажмите 'СОХРАНИТЬ'.",
            parent=self,
        )

    def _build_period_slice_rows(
        self,
        period_from: date,
        period_to: date,
    ) -> List[Dict[str, Any]]:
        """
        Формирует строки среза:
        - план на период / факт на период / отклонение за период
        - план к дате (на конец периода) / факт к дате / общее отклонение
        """
        task_ids = [
            int(t["id"])
            for t in self.tasks
            if t.get("id") and (t.get("row_kind") or "task") == "task"
        ]

        period_fact_info = GprService.load_task_fact_period_info(
            task_ids,
            period_from,
            period_to,
        )

        # Накопительный факт на дату (включая period_to)
        fact_upto_map = GprService.load_task_fact_upto(task_ids, period_to)  # NEW

        rows: List[Dict[str, Any]] = []

        for t in self.tasks:
            row_kind = (t.get("row_kind") or "task").strip()

            if row_kind in ("group", "title"):
                rows.append(
                    {"row_kind": row_kind, "name": t.get("name", "")}
                )
                continue

            ds = _to_date(t.get("plan_start"))
            df = _to_date(t.get("plan_finish"))

            tid = t.get("id")
            tid_int = int(tid) if tid else None

            fact_period_info = period_fact_info.get(tid_int, {}) if tid_int else {}
            fact_qty_period = float(fact_period_info.get("fact_qty_period") or 0)

            has_plan_overlap = False
            if ds and df:
                has_plan_overlap = _overlap_days(ds, df, period_from, period_to) > 0

            if not has_plan_overlap and fact_qty_period <= 0:
                # В срезе оставляем только реально попадающие работы
                continue

            plan_qty_total = _safe_float(t.get("plan_qty"))

            # План на период
            plan_qty_period = _calc_plan_qty_for_period(
                plan_qty_total, ds, df, period_from, period_to
            )

            # Факт накоп. за всё время (как и раньше)
            fact_qty_total = float(self.facts.get(tid_int, 0) or 0) if tid_int else 0.0

            # Отклонение за период
            deviation_period = None
            period_pct = None
            if plan_qty_period is not None:
                deviation_period = fact_qty_period - plan_qty_period
                if plan_qty_period > 0:
                    period_pct = fact_qty_period / plan_qty_period * 100

            # План к дате (на конец периода) — равномерное распределение
            plan_qty_upto = None
            if plan_qty_total is not None and ds and df:
                cutoff = min(df, period_to)
                if cutoff < ds:
                    plan_qty_upto = 0.0
                else:
                    plan_qty_upto = _calc_plan_qty_for_period(
                        plan_qty_total, ds, df, ds, cutoff
                    )

            # Факт к дате (на конец периода)
            fact_qty_upto = float(fact_upto_map.get(tid_int, 0.0)) if tid_int else 0.0

            # Общее отклонение (к концу периода)
            deviation_total = None
            total_pct_to_date = None
            if plan_qty_upto is not None:
                deviation_total = fact_qty_upto - plan_qty_upto
                if plan_qty_upto > 0:
                    total_pct_to_date = fact_qty_upto / plan_qty_upto * 100

            # Старый общий % к общему плану (оставляем как было)
            total_pct = None
            if plan_qty_total and plan_qty_total > 0:
                total_pct = fact_qty_total / plan_qty_total * 100

            rows.append(
                {
                    "row_kind": "task",
                    "task_id": tid_int,
                    "work_type_name": t.get("work_type_name", ""),
                    "name": t.get("name", ""),
                    "uom_code": t.get("uom_code") or "",
                    "plan_start": ds,
                    "plan_finish": df,
                    "status": t.get("status", "planned"),

                    "plan_qty_total": plan_qty_total,
                    "plan_qty_period": plan_qty_period,

                    "fact_qty_period": fact_qty_period,
                    "deviation": deviation_period,
                    "period_pct": period_pct,

                    "plan_qty_upto": plan_qty_upto,           # NEW
                    "fact_qty_upto": fact_qty_upto,           # NEW
                    "deviation_total": deviation_total,       # NEW
                    "total_pct_to_date": total_pct_to_date,   # NEW

                    "fact_qty_total": fact_qty_total,
                    "total_pct": total_pct,

                    "workers_last_period": fact_period_info.get("workers_last_period"),
                    "workers_max_period": fact_period_info.get("workers_max_period"),
                    "workers_sum_period": fact_period_info.get("workers_sum_period", 0),
                }
            )

        return rows

    def _export_period_slice(self):
        """
        Экспортирует срез за выбранный период:
        - план на период / факт на период / отклонение за период
        - план к дате / факт к дате / отклонение общее (на конец периода)
        """
        if not self.plan_id:
            messagebox.showinfo(
                "ГПР",
                "Сначала откройте объект.",
                parent=self,
            )
            return
    
        if not self.tasks:
            messagebox.showinfo(
                "ГПР",
                "Нет работ для формирования среза.",
                parent=self,
            )
            return
    
        if not HAS_OPENPYXL:
            messagebox.showwarning(
                "ГПР",
                "Для экспорта необходима библиотека openpyxl.\n"
                "Установите: pip install openpyxl",
                parent=self,
            )
            return
    
        dlg = DateRangeDialog(self, self.range_from, self.range_to)
        if not dlg.result:
            return
    
        period_from, period_to = dlg.result
    
        try:
            rows = self._build_period_slice_rows(period_from, period_to)
        except Exception as e:
            logger.exception("GPR period slice build error")
            messagebox.showerror(
                "ГПР",
                f"Не удалось сформировать срез:\n{e}",
                parent=self,
            )
            return
    
        task_rows_count = sum(1 for r in rows if r.get("row_kind") == "task")
        if task_rows_count <= 0:
            messagebox.showinfo(
                "ГПР",
                "В выбранном периоде нет плановых работ и факта.",
                parent=self,
            )
            return
    
        obj = next((o for o in self.objects if int(o["id"]) == self.object_db_id), None)
        if obj:
            obj_name = obj.get("short_name") or obj.get("address") or "объект"
            addr = obj.get("address") or ""
        else:
            obj_name = "объект"
            addr = ""
    
        default_name = (
            f"ГПР_срез_{obj_name}_"
            f"{period_from.strftime('%Y%m%d')}_"
            f"{period_to.strftime('%Y%m%d')}.xlsx"
        )
        default_name = "".join(c if c.isalnum() or c in "._- ()" else "_" for c in default_name)
    
        path = filedialog.asksaveasfilename(
            parent=self,
            title="Сохранить срез ГПР в Excel",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel", "*.xlsx"), ("Все файлы", "*.*")],
        )
        if not path:
            return
    
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Срез периода"
    
            thin_side = Side(style="thin", color="D0D0D0")
            thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
            header_fill = PatternFill("solid", fgColor="D6DCE4")
            group_fill = PatternFill("solid", fgColor="EEF5FF")
            title_fill = PatternFill("solid", fgColor="DFF1FF")
            negative_fill = PatternFill("solid", fgColor="F8D7DA")
            positive_fill = PatternFill("solid", fgColor="D4EDDA")
            neutral_fill = PatternFill("solid", fgColor="FFF3CD")
    
            status_fill = {
                "planned": PatternFill("solid", fgColor="D6EAFF"),
                "in_progress": PatternFill("solid", fgColor="FFF3CD"),
                "done": PatternFill("solid", fgColor="D4EDDA"),
                "paused": PatternFill("solid", fgColor="FFF9C4"),
                "canceled": PatternFill("solid", fgColor="F8D7DA"),
            }
    
            headers = [
                "№",
                "Тип работ",
                "Вид работ",
                "Ед.",
                "Начало",
                "Окончание",
                "Статус",
                "План всего",
                "План на период",
                "Факт на период",
                "Отклонение",
                "% периода",
                "Отклонение общее",  # NEW
                "Факт накоп.",
                "% общий",
                "Людей посл.",
                "Людей сумма",
            ]
    
            widths = [
                6, 22, 40, 8, 12, 12, 16, 14, 16, 16, 14, 12,
                16,  # ширина для "Отклонение общее"
                14, 12, 12, 12,
            ]
    
            header_row = 4
            data_row = header_row + 1
    
            # Заголовки отчета (динамическое слияние по ширине)
            last_col_letter = get_column_letter(len(headers))
            ws.merge_cells(f"A1:{last_col_letter}1")
            title = f"Срез ГПР за период {_fmt_date(period_from)} — {_fmt_date(period_to)}"
            ws.cell(1, 1, title).font = Font(bold=True, size=13)
            ws.cell(1, 1).alignment = Alignment(horizontal="left")
    
            ws.merge_cells(f"A2:{last_col_letter}2")
            obj_title = f"Объект: {obj_name}"
            if addr:
                obj_title += f" — {addr}"
            ws.cell(2, 1, obj_title).font = Font(size=10, italic=True)
    
            # Заголовки таблицы
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(header_row, col_idx, header)
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
    
            for col_idx, width in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width
    
            excel_no = 0
            row_num = data_row
    
            total_plan_period = 0.0
            total_fact_period = 0.0
            total_plan_all = 0.0
            total_fact_all = 0.0
            total_dev_total = 0.0  # NEW: сумма общего отклонения
    
            for r in rows:
                row_kind = (r.get("row_kind") or "task").strip()
    
                if row_kind == "group":
                    ws.cell(row_num, 2, "ГРУППА").font = Font(bold=True)
                    ws.cell(row_num, 3, r.get("name", "")).font = Font(bold=True)
                    for col in range(1, len(headers) + 1):
                        ws.cell(row_num, col).fill = group_fill
                        ws.cell(row_num, col).border = thin_border
                    row_num += 1
                    continue
    
                if row_kind == "title":
                    ws.cell(row_num, 2, "ТИТУЛ").font = Font(bold=True)
                    ws.cell(row_num, 3, r.get("name", "")).font = Font(bold=True)
                    for col in range(1, len(headers) + 1):
                        ws.cell(row_num, col).fill = title_fill
                        ws.cell(row_num, col).border = thin_border
                    row_num += 1
                    continue
    
                excel_no += 1
    
                st_code = r.get("status", "planned")
                st_label = STATUS_LABELS.get(st_code, st_code)
    
                plan_total = r.get("plan_qty_total")
                plan_period = r.get("plan_qty_period")
                fact_period = float(r.get("fact_qty_period") or 0)
                deviation = r.get("deviation")
                period_pct = r.get("period_pct")
    
                deviation_total = r.get("deviation_total")  # NEW
                fact_total = float(r.get("fact_qty_total") or 0)
                total_pct = r.get("total_pct")
    
                if plan_period is not None:
                    total_plan_period += float(plan_period or 0)
                total_fact_period += fact_period
                if plan_total is not None:
                    total_plan_all += float(plan_total or 0)
                total_fact_all += fact_total
                if deviation_total is not None:
                    total_dev_total += float(deviation_total or 0)
    
                values = [
                    excel_no,                              # 1
                    r.get("work_type_name", ""),           # 2
                    r.get("name", ""),                     # 3
                    r.get("uom_code", ""),                 # 4
                    _fmt_date(r.get("plan_start")),        # 5
                    _fmt_date(r.get("plan_finish")),       # 6
                    st_label,                              # 7
                    _fmt_qty(plan_total),                  # 8
                    _fmt_qty(plan_period),                 # 9
                    _fmt_qty(fact_period),                 # 10
                    _fmt_qty(deviation),                   # 11
                    f"{period_pct:.1f}%" if period_pct is not None else "",  # 12
                    _fmt_qty(deviation_total),             # 13 NEW
                    _fmt_qty(fact_total),                  # 14
                    f"{total_pct:.1f}%" if total_pct is not None else "",    # 15
                    r.get("workers_last_period") or "",    # 16
                    r.get("workers_sum_period") or "",     # 17
                ]
    
                for col_idx, value in enumerate(values, start=1):
                    cell = ws.cell(row_num, col_idx, value)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    if col_idx in (2, 3):
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
                # Подсветка статуса
                status_cell = ws.cell(row_num, 7)
                if st_code in status_fill:
                    status_cell.fill = status_fill[st_code]
    
                # Подсветка отклонений (период и общее)
                deviation_cell = ws.cell(row_num, 11)
                if deviation is not None:
                    if deviation < 0:
                        deviation_cell.fill = negative_fill
                    elif deviation > 0:
                        deviation_cell.fill = positive_fill
                    else:
                        deviation_cell.fill = neutral_fill
    
                deviation_total_cell = ws.cell(row_num, 13)  # NEW
                if deviation_total is not None:
                    if deviation_total < 0:
                        deviation_total_cell.fill = negative_fill
                    elif deviation_total > 0:
                        deviation_total_cell.fill = positive_fill
                    else:
                        deviation_total_cell.fill = neutral_fill
    
                row_num += 1
    
            # Итоги
            total_row = row_num + 1
            ws.cell(total_row, 2, "ИТОГО").font = Font(bold=True)
            ws.cell(total_row, 8, _fmt_qty(total_plan_all)).font = Font(bold=True)
            ws.cell(total_row, 9, _fmt_qty(total_plan_period)).font = Font(bold=True)
            ws.cell(total_row, 10, _fmt_qty(total_fact_period)).font = Font(bold=True)
    
            total_dev = total_fact_period - total_plan_period
            ws.cell(total_row, 11, _fmt_qty(total_dev)).font = Font(bold=True)
    
            if total_plan_period > 0:
                ws.cell(total_row, 12, f"{total_fact_period / total_plan_period * 100:.1f}%").font = Font(bold=True)
    
            # NEW: итого по «Отклонение общее»
            ws.cell(total_row, 13, _fmt_qty(total_dev_total)).font = Font(bold=True)
    
            ws.cell(total_row, 14, _fmt_qty(total_fact_all)).font = Font(bold=True)
    
            if total_plan_all > 0:
                ws.cell(total_row, 15, f"{total_fact_all / total_plan_all * 100:.1f}%").font = Font(bold=True)
    
            for col in range(1, len(headers) + 1):
                ws.cell(total_row, col).border = thin_border
                ws.cell(total_row, col).fill = PatternFill("solid", fgColor="E2F0D9")
    
            ws.cell(total_row + 2, 2, f"Сформировано: {_today().strftime('%d.%m.%Y')}").font = Font(
                italic=True, size=8, color="888888"
            )
    
            ws.freeze_panes = ws.cell(data_row, 1)
            ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{max(row_num - 1, header_row)}"
    
            # Лист Ганта по срезу
            self._export_period_slice_gantt_sheet(
                wb=wb,
                rows=rows,
                obj=obj,
                obj_name=obj_name,
                period_from=period_from,
                period_to=period_to,
            )
    
            wb.save(path)
    
            messagebox.showinfo(
                "ГПР",
                f"Срез периода сохранён:\n{path}\n\n"
                f"Листы: 'Срез периода' и 'Гант среза'",
                parent=self,
            )
    
        except PermissionError:
            messagebox.showerror(
                "ГПР",
                f"Нет доступа к файлу:\n{path}\n\n"
                "Возможно файл открыт в другой программе.",
                parent=self,
            )
        except Exception as e:
            logger.exception("GPR period slice export error")
            messagebox.showerror(
                "ГПР",
                f"Ошибка экспорта среза:\n{e}",
                parent=self,
            )

    def _export_period_slice_gantt_sheet(
        self,
        wb,
        rows: List[Dict[str, Any]],
        obj,
        obj_name: str,
        period_from: date,
        period_to: date,
    ) -> None:
        """
        Создаёт лист Excel с диаграммой Ганта только по строкам среза.

        Диапазон диаграммы = выбранный период среза.
        Прогресс считается как:
            факт на период / план на период
        """
        d0 = _to_date(period_from) or _today()
        d1 = _to_date(period_to) or d0

        if d1 < d0:
            d0, d1 = d1, d0

        days = (d1 - d0).days + 1

        ws = wb.create_sheet("Гант среза")

        fixed_cols = [
            ("№", 6),
            ("Тип работ", 18),
            ("Вид работ", 36),
            ("Начало", 12),
            ("Окончание", 12),
            ("Статус", 16),
            ("План период", 14),
            ("Факт период", 14),
            ("% период", 12),
            ("Людей", 10),
        ]

        gantt_col_start = len(fixed_cols) + 1
        total_cols = len(fixed_cols) + days

        addr = (obj or {}).get("address", "") if obj else ""

        title = (
            f"Диаграмма Ганта по срезу: "
            f"{_fmt_date(d0)} — {_fmt_date(d1)}"
        )
        if obj_name:
            title += f" | {obj_name}"
        if addr:
            title += f" — {addr}"

        thin_side = Side(style="thin", color="D0D0D0")
        thin_border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side,
        )

        header_fill = PatternFill("solid", fgColor="D6DCE4")
        month_fill = PatternFill("solid", fgColor="D6DBE0")
        weekday_fill = PatternFill("solid", fgColor="F3F4F6")
        weekend_fill = PatternFill("solid", fgColor="FFECEC")
        group_fill = PatternFill("solid", fgColor="EEF5FF")
        title_fill = PatternFill("solid", fgColor="DFF1FF")
        progress_fill = PatternFill("solid", fgColor="388E3C")

        status_fill = {
            "planned": self._xl_fill("#90caf9"),
            "in_progress": self._xl_fill("#ffcc80"),
            "done": self._xl_fill("#a5d6a7"),
            "paused": self._xl_fill("#fff176"),
            "canceled": self._xl_fill("#ef9a9a"),
        }

        month_names = {
            1: "Январь",
            2: "Февраль",
            3: "Март",
            4: "Апрель",
            5: "Май",
            6: "Июнь",
            7: "Июль",
            8: "Август",
            9: "Сентябрь",
            10: "Октябрь",
            11: "Ноябрь",
            12: "Декабрь",
        }

        weekday_names = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]

        ws.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=total_cols,
        )
        c = ws.cell(1, 1, title)
        c.font = Font(bold=True, size=12)
        c.alignment = Alignment(horizontal="left", vertical="center")

        month_row = 2
        day_row = 3
        week_row = 4
        data_row = 5

        # ── фиксированные заголовки ───────────────────────
        for col_idx, (caption, width) in enumerate(fixed_cols, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

            ws.merge_cells(
                start_row=month_row,
                start_column=col_idx,
                end_row=week_row,
                end_column=col_idx,
            )

            cell = ws.cell(month_row, col_idx, caption)
            cell.font = Font(bold=True, size=10)
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = thin_border

        # ── месяцы ────────────────────────────────────────
        cur = date(d0.year, d0.month, 1)

        while cur <= d1:
            month_last_day = calendar.monthrange(cur.year, cur.month)[1]
            ms = max(cur, d0)
            me = min(date(cur.year, cur.month, month_last_day), d1)

            c0 = gantt_col_start + (ms - d0).days
            c1 = gantt_col_start + (me - d0).days

            if c0 != c1:
                ws.merge_cells(
                    start_row=month_row,
                    start_column=c0,
                    end_row=month_row,
                    end_column=c1,
                )

            mcell = ws.cell(
                month_row,
                c0,
                f"{month_names[cur.month]} {cur.year}",
            )
            mcell.font = Font(bold=True, size=10)
            mcell.fill = month_fill
            mcell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )
            mcell.border = thin_border

            if cur.month == 12:
                cur = date(cur.year + 1, 1, 1)
            else:
                cur = date(cur.year, cur.month + 1, 1)

        # ── дни ───────────────────────────────────────────
        for i in range(days):
            col = gantt_col_start + i
            dt = d0 + timedelta(days=i)

            ws.column_dimensions[get_column_letter(col)].width = 3.2

            fill = weekend_fill if dt.weekday() >= 5 else weekday_fill

            day_cell = ws.cell(day_row, col, dt.day)
            day_cell.font = Font(bold=True, size=8)
            day_cell.fill = fill
            day_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )
            day_cell.border = thin_border

            wd_cell = ws.cell(week_row, col, weekday_names[dt.weekday()])
            wd_cell.font = Font(size=8)
            wd_cell.fill = fill
            wd_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )
            wd_cell.border = thin_border

        excel_no = 0
        last_row = data_row - 1

        # ── строки диаграммы ──────────────────────────────
        for row_num, r in enumerate(rows, start=data_row):
            last_row = row_num
            ws.row_dimensions[row_num].height = 20

            row_kind = (r.get("row_kind") or "task").strip()

            if row_kind == "group":
                for col in range(1, total_cols + 1):
                    ws.cell(row_num, col).fill = group_fill
                    ws.cell(row_num, col).border = thin_border

                ws.cell(row_num, 2, "ГРУППА").font = Font(bold=True)
                ws.cell(row_num, 3, r.get("name", "")).font = Font(bold=True)
                ws.cell(row_num, 3).alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                )
                continue

            if row_kind == "title":
                for col in range(1, total_cols + 1):
                    ws.cell(row_num, col).fill = title_fill
                    ws.cell(row_num, col).border = thin_border

                ws.cell(row_num, 2, "ТИТУЛ").font = Font(bold=True)
                ws.cell(row_num, 3, r.get("name", "")).font = Font(bold=True)
                ws.cell(row_num, 3).alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                )
                continue

            excel_no += 1

            ds = _to_date(r.get("plan_start"))
            df = _to_date(r.get("plan_finish"))

            st_code = (r.get("status") or "planned").strip()
            st_label = STATUS_LABELS.get(st_code, st_code)

            plan_period = _safe_float(r.get("plan_qty_period"))
            fact_period = _safe_float(r.get("fact_qty_period")) or 0.0

            period_pct = None
            if plan_period and plan_period > 0:
                period_pct = fact_period / plan_period * 100

            workers_last = r.get("workers_last_period") or ""

            left_values = [
                excel_no,
                r.get("work_type_name", ""),
                r.get("name", ""),
                _fmt_date(ds) if ds else "",
                _fmt_date(df) if df else "",
                st_label,
                _fmt_qty(plan_period),
                _fmt_qty(fact_period),
                f"{period_pct:.1f}%" if period_pct is not None else "",
                workers_last,
            ]

            for col_idx, val in enumerate(left_values, start=1):
                cell = ws.cell(row_num, col_idx, val)
                cell.border = thin_border

                if col_idx in (1, 4, 5, 6, 7, 8, 9, 10):
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                        wrap_text=True,
                    )
                else:
                    cell.alignment = Alignment(
                        horizontal="left",
                        vertical="center",
                        wrap_text=True,
                    )

            st_fill = status_fill.get(st_code)
            if st_fill:
                ws.cell(row_num, 6).fill = st_fill

            if not ds or not df:
                continue

            # Если работа не пересекается с периодом, бар не рисуем.
            if df < d0 or ds > d1:
                continue

            clip_start = max(ds, d0)
            clip_finish = min(df, d1)

            col_start = gantt_col_start + (clip_start - d0).days
            col_finish = gantt_col_start + (clip_finish - d0).days

            bar_fill = status_fill.get(st_code, self._xl_fill("#90caf9"))

            # Плановый бар в пределах выбранного периода.
            for col in range(col_start, col_finish + 1):
                cell = ws.cell(row_num, col)
                cell.fill = bar_fill
                cell.border = thin_border

            # Прогресс по периоду: факт периода / план периода.
            if plan_period and plan_period > 0 and fact_period > 0:
                pct = max(0.0, min(1.0, fact_period / plan_period))

                clipped_days = max(1, (clip_finish - clip_start).days + 1)
                progress_days = max(
                    1,
                    min(clipped_days, int(round(clipped_days * pct))),
                )

                progress_finish = clip_start + timedelta(days=progress_days - 1)

                pcol0 = gantt_col_start + (clip_start - d0).days
                pcol1 = gantt_col_start + (progress_finish - d0).days

                for col in range(pcol0, pcol1 + 1):
                    cell = ws.cell(row_num, col)
                    cell.fill = progress_fill
                    cell.border = thin_border

        # ── линия сегодняшнего дня ────────────────────────
        td = _today()

        if d0 <= td <= d1:
            today_col = gantt_col_start + (td - d0).days
            red_side = Side(style="medium", color="FF0000")

            for rr in range(day_row, max(last_row, week_row) + 1):
                cell = ws.cell(rr, today_col)
                old_border = cell.border

                cell.border = Border(
                    left=red_side,
                    right=red_side,
                    top=old_border.top,
                    bottom=old_border.bottom,
                )

        # ── легенда ───────────────────────────────────────
        legend_row = last_row + 2

        ws.cell(legend_row, 1, "Легенда:").font = Font(bold=True)

        lc = 2
        for code in STATUS_LIST:
            ws.cell(legend_row, lc, " ").fill = status_fill[code]
            ws.cell(legend_row, lc).border = thin_border
            ws.cell(legend_row, lc + 1, STATUS_LABELS[code])
            lc += 2

        ws.cell(legend_row + 1, 2, " ").fill = progress_fill
        ws.cell(legend_row + 1, 2).border = thin_border
        ws.cell(legend_row + 1, 3, "Фактическое выполнение за период")

        ws.freeze_panes = ws.cell(data_row, gantt_col_start)
    
    def _export_excel_gantt_sheet(self, wb, obj, obj_name: str) -> None:
        """
        Создаёт второй лист Excel с диаграммой Ганта.
        Использует текущий диапазон self.range_from / self.range_to.
        """
        rows = self.tasks
    
        ws = wb.create_sheet("Диаграмма Ганта")
    
        d0 = _to_date(self.range_from) or _today()
        d1 = _to_date(self.range_to) or d0
        if d1 < d0:
            d0, d1 = d1, d0
    
        days = (d1 - d0).days + 1
    
        fixed_cols = [
            ("№", 6),
            ("Тип работ", 18),
            ("Вид работ", 36),
            ("Начало", 12),
            ("Окончание", 12),
            ("Статус", 16),
            ("Людей", 10),
        ]
        gantt_col_start = len(fixed_cols) + 1
        total_cols = len(fixed_cols) + days
    
        addr = (obj or {}).get("address", "") if obj else ""
        title = f"Диаграмма Ганта: {obj_name}"
        if addr:
            title += f" — {addr}"
    
        thin_side = Side(style="thin", color="D0D0D0")
        thin_border = Border(
            left=thin_side, right=thin_side,
            top=thin_side, bottom=thin_side
        )
    
        header_fill = PatternFill("solid", fgColor="D6DCE4")
        month_fill = PatternFill("solid", fgColor="D6DBE0")
        weekday_fill = PatternFill("solid", fgColor="F3F4F6")
        weekend_fill = PatternFill("solid", fgColor="FFECEC")
        group_fill = PatternFill("solid", fgColor="EEF5FF")
        title_fill = PatternFill("solid", fgColor="DFF1FF")
        progress_fill = PatternFill("solid", fgColor="388E3C")
    
        status_fill = {
            "planned": self._xl_fill("#90caf9"),
            "in_progress": self._xl_fill("#ffcc80"),
            "done": self._xl_fill("#a5d6a7"),
            "paused": self._xl_fill("#fff176"),
            "canceled": self._xl_fill("#ef9a9a"),
        }
    
        month_names = {
            1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
            5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
            9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь",
        }
        weekday_names = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        c = ws.cell(1, 1, title)
        c.font = Font(bold=True, size=12)
        c.alignment = Alignment(horizontal="left", vertical="center")
    
        month_row = 2
        day_row = 3
        week_row = 4
        data_row = 5
    
        for col_idx, (caption, width) in enumerate(fixed_cols, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
            ws.merge_cells(
                start_row=month_row, start_column=col_idx,
                end_row=week_row, end_column=col_idx
            )
            cell = ws.cell(month_row, col_idx, caption)
            cell.font = Font(bold=True, size=10)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
    
        cur = date(d0.year, d0.month, 1)
        while cur <= d1:
            month_last_day = calendar.monthrange(cur.year, cur.month)[1]
            ms = max(cur, d0)
            me = min(date(cur.year, cur.month, month_last_day), d1)
    
            c0 = gantt_col_start + (ms - d0).days
            c1 = gantt_col_start + (me - d0).days
    
            if c0 != c1:
                ws.merge_cells(
                    start_row=month_row, start_column=c0,
                    end_row=month_row, end_column=c1
                )
    
            mcell = ws.cell(month_row, c0, f"{month_names[cur.month]} {cur.year}")
            mcell.font = Font(bold=True, size=10)
            mcell.fill = month_fill
            mcell.alignment = Alignment(horizontal="center", vertical="center")
    
            if cur.month == 12:
                cur = date(cur.year + 1, 1, 1)
            else:
                cur = date(cur.year, cur.month + 1, 1)
    
        for i in range(days):
            col = gantt_col_start + i
            dt = d0 + timedelta(days=i)
    
            ws.column_dimensions[get_column_letter(col)].width = 3.2
            fill = weekend_fill if dt.weekday() >= 5 else weekday_fill
    
            day_cell = ws.cell(day_row, col, dt.day)
            day_cell.font = Font(bold=True, size=8)
            day_cell.fill = fill
            day_cell.alignment = Alignment(horizontal="center", vertical="center")
            day_cell.border = thin_border
    
            wd_cell = ws.cell(week_row, col, weekday_names[dt.weekday()])
            wd_cell.font = Font(size=8)
            wd_cell.fill = fill
            wd_cell.alignment = Alignment(horizontal="center", vertical="center")
            wd_cell.border = thin_border
    
        excel_no = 0
        last_row = data_row - 1
    
        for row_num, t in enumerate(rows, start=data_row):
            last_row = row_num
            ws.row_dimensions[row_num].height = 20
    
            row_kind = (t.get("row_kind") or "task").strip()
    
            if row_kind == "group":
                for col in range(1, total_cols + 1):
                    ws.cell(row_num, col).fill = group_fill
                ws.cell(row_num, 2, "ГРУППА").font = Font(bold=True)
                ws.cell(row_num, 3, t.get("name", "")).font = Font(bold=True)
                ws.cell(row_num, 3).alignment = Alignment(horizontal="left", vertical="center")
                continue
    
            if row_kind == "title":
                for col in range(1, total_cols + 1):
                    ws.cell(row_num, col).fill = title_fill
                ws.cell(row_num, 2, "ТИТУЛ").font = Font(bold=True)
                ws.cell(row_num, 3, t.get("name", "")).font = Font(bold=True)
                ws.cell(row_num, 3).alignment = Alignment(horizontal="left", vertical="center")
                continue
    
            excel_no += 1
    
            ds = _to_date(t.get("plan_start"))
            df = _to_date(t.get("plan_finish"))
            st_code = (t.get("status") or "planned").strip()
            st_label = STATUS_LABELS.get(st_code, st_code)
    
            tid = t.get("id")
            workers_last = ""
            if tid is not None:
                info = self.fact_info.get(tid) or {}
                if info.get("workers_last") is not None:
                    workers_last = info["workers_last"]
    
            left_values = [
                excel_no,
                t.get("work_type_name", ""),
                t.get("name", ""),
                _fmt_date(ds) if ds else "",
                _fmt_date(df) if df else "",
                st_label,
                workers_last,
            ]
    
            for col_idx, val in enumerate(left_values, start=1):
                cell = ws.cell(row_num, col_idx, val)
                cell.border = thin_border
                if col_idx in (1, 4, 5, 6, 7):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
    
            st_fill = status_fill.get(st_code)
            if st_fill:
                ws.cell(row_num, 6).fill = st_fill
    
            if not ds or not df:
                continue
            if df < d0 or ds > d1:
                continue
    
            clip_start = max(ds, d0)
            clip_finish = min(df, d1)
    
            col_start = gantt_col_start + (clip_start - d0).days
            col_finish = gantt_col_start + (clip_finish - d0).days
    
            bar_fill = status_fill.get(st_code, self._xl_fill("#90caf9"))
    
            for col in range(col_start, col_finish + 1):
                cell = ws.cell(row_num, col)
                cell.fill = bar_fill
                cell.border = thin_border
    
            pq = _safe_float(t.get("plan_qty"))
            fq = self.facts.get(tid, 0) if tid else 0.0
    
            if pq and pq > 0 and fq > 0:
                pct = max(0.0, min(1.0, fq / pq))
                total_task_days = max(1, (df - ds).days + 1)
                progress_days = max(1, min(total_task_days, int(round(total_task_days * pct))))
                progress_finish = ds + timedelta(days=progress_days - 1)
    
                p0 = max(ds, d0)
                p1 = min(progress_finish, d1)
    
                if p1 >= p0:
                    pcol0 = gantt_col_start + (p0 - d0).days
                    pcol1 = gantt_col_start + (p1 - d0).days
                    for col in range(pcol0, pcol1 + 1):
                        cell = ws.cell(row_num, col)
                        cell.fill = progress_fill
                        cell.border = thin_border
    
            if t.get("is_milestone"):
                milestone_date = min(max(ds, d0), d1)
                mcol = gantt_col_start + (milestone_date - d0).days
                mcell = ws.cell(row_num, mcol, "◆")
                mcell.font = Font(bold=True, color="1A73E8")
                mcell.alignment = Alignment(horizontal="center", vertical="center")
    
        td = _today()
        if d0 <= td <= d1:
            today_col = gantt_col_start + (td - d0).days
            red_side = Side(style="medium", color="FF0000")
            for rr in range(day_row, max(last_row, week_row) + 1):
                cell = ws.cell(rr, today_col)
                old_border = cell.border
                cell.border = Border(
                    left=red_side,
                    right=red_side,
                    top=old_border.top,
                    bottom=old_border.bottom
                )
    
        legend_row = last_row + 2
        ws.cell(legend_row, 1, "Легенда:").font = Font(bold=True)
    
        lc = 2
        for code in STATUS_LIST:
            ws.cell(legend_row, lc, " ").fill = status_fill[code]
            ws.cell(legend_row, lc).border = thin_border
            ws.cell(legend_row, lc + 1, STATUS_LABELS[code])
            lc += 2
    
        ws.cell(legend_row + 1, 2, " ").fill = progress_fill
        ws.cell(legend_row + 1, 2).border = thin_border
        ws.cell(legend_row + 1, 3, "Фактическое выполнение")
    
        ws.freeze_panes = ws.cell(data_row, gantt_col_start)
    
    def _export_excel(self):
        if not self.tasks:
            messagebox.showinfo(
                "ГПР", "Нет данных для выгрузки.", parent=self
            )
            return
    
        if not HAS_OPENPYXL:
            messagebox.showwarning(
                "ГПР",
                "Для экспорта необходима библиотека openpyxl.\n"
                "Установите: pip install openpyxl",
                parent=self,
            )
            return
    
        obj = next(
            (o for o in self.objects if int(o["id"]) == self.object_db_id),
            None,
        )
        if obj:
            obj_name = obj.get("short_name") or obj.get("address") or "объект"
        else:
            obj_name = "объект"
    
        default_name = f"ГПР_{obj_name}_{_today().strftime('%Y%m%d')}.xlsx"
        default_name = "".join(
            c if c.isalnum() or c in "._- ()" else "_"
            for c in default_name
        )
    
        path = filedialog.asksaveasfilename(
            parent=self,
            title="Сохранить ГПР в Excel",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel", "*.xlsx"), ("Все файлы", "*.*")],
        )
        if not path:
            return
    
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "ГПР"
    
            if obj:
                addr = obj.get("address", "")
                ws.merge_cells("A1:L1")
                title_cell = ws.cell(1, 1, f"ГПР: {obj_name} — {addr}")
                title_cell.font = Font(bold=True, size=12)
                title_cell.alignment = Alignment(horizontal="left")
                data_start_row = 3
            else:
                data_start_row = 1
    
            headers = [
                "№",
                "Тип работ",
                "Вид работ",
                "Ед. изм.",
                "Объём план",
                "Начало",
                "Окончание",
                "Длительность (дн.)",
                "Статус",
                "Факт (накоп.)",
                "% выполнения",
                "Людей",
            ]
            widths = [6, 22, 36, 8, 14, 14, 14, 14, 16, 14, 14, 10]
    
            hdr_row = data_start_row
            for i, h in enumerate(headers, 1):
                cell = ws.cell(hdr_row, i, h)
                cell.font = Font(bold=True, size=10)
                cell.fill = PatternFill("solid", fgColor="D6DCE4")
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
    
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
    
            ws.freeze_panes = f"A{hdr_row + 1}"
    
            status_fill = {
                "planned": PatternFill("solid", fgColor="D6EAFF"),
                "in_progress": PatternFill("solid", fgColor="FFF3CD"),
                "done": PatternFill("solid", fgColor="D4EDDA"),
                "paused": PatternFill("solid", fgColor="FFF9C4"),
                "canceled": PatternFill("solid", fgColor="F8D7DA"),
            }
    
            excel_no = 0
            for row_num, t in enumerate(self.tasks, start=hdr_row + 1):
                row_kind = (t.get("row_kind") or "task").strip()
    
                if row_kind == "group":
                    ws.cell(row_num, 2, "ГРУППА").font = Font(bold=True)
                    ws.cell(row_num, 3, t.get("name", "")).font = Font(bold=True)
                    for c in range(1, 13):
                        ws.cell(row_num, c).fill = PatternFill("solid", fgColor="EEF5FF")
                    continue
    
                if row_kind == "title":
                    ws.cell(row_num, 2, "ТИТУЛ").font = Font(bold=True)
                    ws.cell(row_num, 3, t.get("name", "")).font = Font(bold=True)
                    for c in range(1, 13):
                        ws.cell(row_num, c).fill = PatternFill("solid", fgColor="DFF1FF")
                    continue
    
                excel_no += 1
                ds = _to_date(t.get("plan_start"))
                df = _to_date(t.get("plan_finish"))
                dur = (df - ds).days + 1 if ds and df else ""
    
                pq = _safe_float(t.get("plan_qty"))
                tid = t.get("id")
                fq = self.facts.get(tid, 0) if tid else 0
                pct = ""
                if pq and pq > 0:
                    pct = f"{min(100.0, fq / pq * 100):.1f}%"
    
                workers_last = ""
                if tid is not None:
                    info = self.fact_info.get(tid) or {}
                    if info.get("workers_last") is not None:
                        workers_last = info["workers_last"]
    
                st_code = t.get("status", "planned")
                st_label = STATUS_LABELS.get(st_code, st_code)
    
                values = [
                    excel_no,
                    t.get("work_type_name", ""),
                    t.get("name", ""),
                    t.get("uom_code") or "",
                    _fmt_qty(pq) if pq else "",
                    _fmt_date(ds) if ds else "",
                    _fmt_date(df) if df else "",
                    dur,
                    st_label,
                    _fmt_qty(fq) if fq else "",
                    pct,
                    workers_last,
                ]
    
                for col, val in enumerate(values, 1):
                    cell = ws.cell(row_num, col, val)
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                    if col == 9:
                        fill = status_fill.get(st_code)
                        if fill:
                            cell.fill = fill
    
            last_row = hdr_row + len(self.tasks) + 1
            task_count = sum(
                1 for t in self.tasks if (t.get("row_kind") or "task") == "task"
            )
            done_cnt = sum(
                1 for t in self.tasks
                if (t.get("row_kind") or "task") == "task" and t.get("status") == "done"
            )
    
            ws.cell(last_row, 2, f"Итого работ: {task_count}").font = Font(bold=True)
            ws.cell(last_row, 9, f"Выполнено: {done_cnt}").font = Font(bold=True)
            ws.cell(
                last_row + 1,
                2,
                f"Выгружено: {_today().strftime('%d.%m.%Y')}",
            ).font = Font(italic=True, size=8, color="888888")
    
            self._export_excel_gantt_sheet(wb, obj, obj_name)
    
            wb.save(path)
            messagebox.showinfo(
                "ГПР",
                f"Файл сохранён:\n{path}\n\n"
                f"Листы: 'ГПР' и 'Диаграмма Ганта'",
                parent=self
            )
    
        except PermissionError:
            messagebox.showerror(
                "ГПР",
                f"Нет доступа к файлу:\n{path}\n\n"
                "Возможно файл открыт в другой программе.",
                parent=self,
            )
        except Exception as e:
            logger.exception("GPR excel export error")
            messagebox.showerror(
                "ГПР", f"Ошибка экспорта:\n{e}", parent=self
            )

# ═══════════════════════════════════════════════════════════════
#  API for main_app
# ═══════════════════════════════════════════════════════════════
def create_gpr_page(parent, app_ref) -> GprPage:
    """Фабричная функция — вызывается из main_app._show_page."""
    return GprPage(parent, app_ref=app_ref)
