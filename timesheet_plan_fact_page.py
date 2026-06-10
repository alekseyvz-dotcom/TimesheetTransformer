import calendar
import logging
import re
import threading
from datetime import datetime, timedelta, date
from typing import Optional, List, Dict, Any, Tuple

import tkinter as tk
from tkinter import ttk, messagebox, filedialog

import pandas as pd

from psycopg2 import pool
from psycopg2.extras import RealDictCursor
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

try:
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
except ImportError:
    ILLEGAL_CHARACTERS_RE = re.compile(r"[\000-\010]|[\013-\014]|[\016-\037]")

db_connection_pool: Optional[pool.SimpleConnectionPool] = None


def set_db_pool(db_pool: pool.SimpleConnectionPool):
    global db_connection_pool
    db_connection_pool = db_pool
    logging.info("Timesheet Plan/Fact Module: DB pool set.")


PALETTE = {
    "primary": "#1565C0",
    "success": "#2E7D32",
    "warning": "#E65100",
    "accent": "#6A1B9A",
    "negative": "#C62828",
    "text_muted": "#78909C",
    "offday": "#90A4AE",
    "noschedule": "#8E24AA",
}

# Названия русских месяцев для подписей
RU_MONTHS = [
    "", "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
]


def _norm_text(v: Any) -> str:
    return str(v or "").strip()


def _norm_fio(v: Any) -> str:
    return " ".join(_norm_text(v).lower().split())


def _norm_tbn(v: Any) -> str:
    return " ".join(_norm_text(v).lower().split())


def _person_key(fio: Any, tbn: Any) -> str:
    tbn_val = _norm_tbn(tbn)
    if tbn_val:
        return f"tbn:{tbn_val}"
    return f"fio:{_norm_fio(fio)}"


def _extract_hours(raw: Any) -> float:
    if raw is None:
        return 0.0

    if isinstance(raw, (int, float)):
        try:
            val = float(raw)
            return val if val > 0 else 0.0
        except Exception:
            return 0.0

    s = str(raw).strip().replace(",", ".").lower()
    if not s:
        return 0.0

    non_work_codes = {
        "null", "[null]",
        "в", "вых", "выходной",
        "б", "больничный",
        "о", "от", "отп", "отпуск",
        "к", "ком", "командировка",
        "п", "пр", "прогул",
        "н", "нн",
        "д", "до",
    }
    if s in non_work_codes:
        return 0.0

    m = re.search(r"[-+]?\d+(?:\.\d+)?", s)
    if not m:
        return 0.0

    try:
        val = float(m.group(0))
        return val if val > 0 else 0.0
    except Exception:
        return 0.0


class TimesheetPlanFactData:
    """
    mode:
        "day"   -> аналитика за конкретный день selected_date
        "month" -> аналитика за весь месяц selected_date (год + месяц)
    """

    def __init__(self, selected_date: date, object_type_filter: str = "", mode: str = "day"):
        self.selected_date = selected_date
        self.object_type_filter = object_type_filter or ""
        self.mode = mode if mode in ("day", "month") else "day"

        self._detail_cache: Optional[pd.DataFrame] = None
        self._object_cache: Optional[pd.DataFrame] = None
        self._position_cache: Optional[pd.DataFrame] = None
        self._kpi_cache: Optional[Dict[str, Any]] = None
        self._trend_cache: Optional[pd.DataFrame] = None

    # ---------- Вспомогательные свойства периода ----------

    @property
    def is_month_mode(self) -> bool:
        return self.mode == "month"

    @property
    def period_label(self) -> str:
        if self.is_month_mode:
            return f"{RU_MONTHS[self.selected_date.month]} {self.selected_date.year}"
        return self.selected_date.strftime("%d.%m.%Y")

    def _month_days(self) -> List[date]:
        year = self.selected_date.year
        month = self.selected_date.month
        days_in_month = calendar.monthrange(year, month)[1]
        return [date(year, month, d) for d in range(1, days_in_month + 1)]

    # ---------- БД ----------

    def _execute_query(self, query: str, params: tuple = ()) -> List[Dict[str, Any]]:
        if not db_connection_pool:
            raise ConnectionError("Пул соединений с БД не инициализирован.")

        conn = None
        try:
            conn = db_connection_pool.getconn()
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("SET LOCAL statement_timeout = 120000")
                cur.execute(query, params)
                return [dict(r) for r in cur.fetchall()]
        finally:
            if conn:
                db_connection_pool.putconn(conn)

    def get_object_types(self) -> List[str]:
        rows = self._execute_query(
            """
            SELECT DISTINCT short_name
            FROM objects
            WHERE COALESCE(short_name, '') <> ''
            ORDER BY short_name
            """
        )
        return [r["short_name"] for r in rows]

    def _load_timesheet_rows_for_month_array(self, year: int, month: int) -> List[Dict[str, Any]]:
        """
        Загружаем строки табеля за месяц, забирая ВЕСЬ массив hours_raw.
        Дальше в Python раскладываем по дням.
        """
        params: List[Any] = [year, month]
        object_filter_sql = ""

        if self.object_type_filter:
            object_filter_sql = " AND COALESCE(o.short_name, '') = %s "
            params.append(self.object_type_filter)

        sql = f"""
            SELECT
                th.id AS header_id,
                th.object_id,
                th.object_db_id,
                COALESCE(o.address, th.object_addr, '—') AS object_name,
                COALESCE(o.short_name, '') AS object_type,
                th.year,
                th.month,
                COALESCE(dep_hdr.name, th.department, '—') AS header_department,
                tr.id AS row_id,
                tr.fio,
                tr.tbn,
                tr.hours_raw AS hours_raw
            FROM timesheet_headers th
            JOIN timesheet_rows tr
              ON tr.header_id = th.id
            LEFT JOIN objects o
              ON o.id = th.object_db_id
            LEFT JOIN departments dep_hdr
              ON dep_hdr.id = th.department_id
            WHERE th.year = %s
              AND th.month = %s
              {object_filter_sql}
            ORDER BY object_name, tr.fio
        """
        return self._execute_query(sql, tuple(params))

    def _load_employees(self) -> Tuple[Dict[str, Dict[str, Any]], Dict[str, Dict[str, Any]]]:
        rows = self._execute_query(
            """
            SELECT
                e.id,
                e.fio,
                e.tbn,
                e.position,
                e.work_schedule,
                COALESCE(d.name, '—') AS department_name
            FROM employees e
            LEFT JOIN departments d
              ON d.id = e.department_id
            """
        )

        by_tbn: Dict[str, Dict[str, Any]] = {}
        by_fio: Dict[str, Dict[str, Any]] = {}

        for r in rows:
            item = {
                "employee_id": r.get("id"),
                "position_name": r.get("position") or "—",
                "department_name": r.get("department_name") or "—",
                "work_schedule_name": _norm_text(r.get("work_schedule")),
            }

            tbn_norm = _norm_tbn(r.get("tbn"))
            fio_norm = _norm_fio(r.get("fio"))

            if tbn_norm and tbn_norm not in by_tbn:
                by_tbn[tbn_norm] = item
            if fio_norm and fio_norm not in by_fio:
                by_fio[fio_norm] = item

        return by_tbn, by_fio

    def _load_schedule_map_for_month(self, year: int, month: int) -> Dict[Tuple[str, date], Dict[str, Any]]:
        start_dt = date(year, month, 1)
        end_dt = date(year, month, calendar.monthrange(year, month)[1])

        rows = self._execute_query(
            """
            SELECT
                LOWER(BTRIM(ws.schedule_name)) AS schedule_name_norm,
                wsd.work_date,
                wsd.is_workday,
                wsd.planned_hours
            FROM work_schedules ws
            JOIN work_schedule_days wsd
              ON wsd.schedule_id = ws.id
            WHERE wsd.work_date BETWEEN %s AND %s
            """,
            (start_dt, end_dt)
        )

        result: Dict[Tuple[str, date], Dict[str, Any]] = {}
        for r in rows:
            result[(r["schedule_name_norm"], r["work_date"])] = {
                "is_workday": bool(r["is_workday"]),
                "planned_hours": float(r["planned_hours"] or 0),
            }
        return result

    def _match_employee(
        self,
        fio: Any,
        tbn: Any,
        employees_by_tbn: Dict[str, Dict[str, Any]],
        employees_by_fio: Dict[str, Dict[str, Any]],
    ) -> Optional[Dict[str, Any]]:
        tbn_norm = _norm_tbn(tbn)
        if tbn_norm and tbn_norm in employees_by_tbn:
            return employees_by_tbn[tbn_norm]

        fio_norm = _norm_fio(fio)
        if fio_norm and fio_norm in employees_by_fio:
            return employees_by_fio[fio_norm]

        return None

    @staticmethod
    def _get_day_value_from_array(hours_raw: Any, day_num: int) -> Any:
        """
        hours_raw приходит как list (PostgreSQL array). День day_num -> индекс day_num-1.
        """
        if hours_raw is None:
            return None
        try:
            if isinstance(hours_raw, (list, tuple)):
                if 1 <= day_num <= len(hours_raw):
                    return hours_raw[day_num - 1]
                return None
        except Exception:
            return None
        return None

    def _build_one_day_row(
        self,
        row: Dict[str, Any],
        current_date: date,
        raw_val: Any,
        emp: Optional[Dict[str, Any]],
        schedule_map: Dict[Tuple[str, date], Dict[str, Any]],
    ) -> Dict[str, Any]:
        hours = _extract_hours(raw_val)
        worked_flag = 1 if hours > 0 else 0

        if emp:
            employee_id = emp.get("employee_id")
            department_name = emp.get("department_name") or row.get("header_department") or "—"
            position_name = emp.get("position_name") or "—"
            work_schedule_name = emp.get("work_schedule_name") or ""
        else:
            employee_id = None
            department_name = row.get("header_department") or "—"
            position_name = "—"
            work_schedule_name = ""

        schedule_info = None
        if work_schedule_name:
            schedule_info = schedule_map.get((work_schedule_name.lower(), current_date))

        no_schedule_flag = 0
        plan_flag = 0
        off_flag = 0
        fact_flag = 0
        absent_flag = 0
        worked_on_off_flag = 0
        is_workday = None
        planned_hours = 0.0

        if schedule_info is None:
            no_schedule_flag = 1
            if worked_flag:
                fact_flag = 1
        else:
            is_workday = bool(schedule_info["is_workday"])
            planned_hours = float(schedule_info.get("planned_hours") or 0)

            if is_workday:
                plan_flag = 1
                if worked_flag:
                    fact_flag = 1
                else:
                    absent_flag = 1
            else:
                off_flag = 1
                if worked_flag:
                    worked_on_off_flag = 1
                    fact_flag = 1

        return {
            "work_date": pd.Timestamp(current_date),
            "object_id": row.get("object_id") or row.get("object_db_id"),
            "object_db_id": row.get("object_db_id"),
            "object_name": row.get("object_name") or "—",
            "object_type": row.get("object_type") or "",
            "department_name": department_name,
            "position_name": position_name,
            "fio": row.get("fio") or "",
            "tbn": row.get("tbn") or "",
            "person_key": _person_key(row.get("fio"), row.get("tbn")),
            "employee_id": employee_id,
            "work_schedule_name": work_schedule_name,
            "day_raw": raw_val,
            "hours": hours,
            "worked_flag": worked_flag,
            "is_workday": is_workday,
            "planned_hours": planned_hours,
            "plan_flag": plan_flag,
            "off_flag": off_flag,
            "fact_flag": fact_flag,
            "absent_flag": absent_flag,
            "worked_on_off_flag": worked_on_off_flag,
            "no_schedule_flag": no_schedule_flag,
        }

    @staticmethod
    def _empty_detail_df() -> pd.DataFrame:
        return pd.DataFrame(columns=[
            "work_date", "object_id", "object_db_id", "object_name", "object_type",
            "department_name", "position_name", "fio", "tbn", "person_key",
            "employee_id", "work_schedule_name", "day_raw", "hours",
            "worked_flag", "is_workday", "planned_hours", "plan_flag",
            "off_flag", "fact_flag", "absent_flag", "worked_on_off_flag",
            "no_schedule_flag",
        ])

    def _build_detail(self) -> pd.DataFrame:
        """
        Универсальная сборка детализации.
        В режиме day -> один день.
        В режиме month -> все дни месяца (каждая запись = человек+день).
        """
        year = self.selected_date.year
        month = self.selected_date.month

        timesheet_rows = self._load_timesheet_rows_for_month_array(year, month)
        employees_by_tbn, employees_by_fio = self._load_employees()
        schedule_map = self._load_schedule_map_for_month(year, month)

        if self.is_month_mode:
            target_days = self._month_days()
        else:
            target_days = [self.selected_date]

        result_rows: List[Dict[str, Any]] = []

        for row in timesheet_rows:
            hours_raw = row.get("hours_raw")
            emp = self._match_employee(
                row.get("fio"),
                row.get("tbn"),
                employees_by_tbn,
                employees_by_fio,
            )

            for current_date in target_days:
                raw_val = self._get_day_value_from_array(hours_raw, current_date.day)
                result_rows.append(
                    self._build_one_day_row(row, current_date, raw_val, emp, schedule_map)
                )

        df = pd.DataFrame(result_rows)

        if df.empty:
            return self._empty_detail_df()

        for col in [
            "worked_flag", "plan_flag", "off_flag", "fact_flag",
            "absent_flag", "worked_on_off_flag", "no_schedule_flag",
        ]:
            df[col] = df[col].fillna(0).astype(int)

        df["hours"] = pd.to_numeric(df["hours"], errors="coerce").fillna(0.0)

        return df

    def get_day_detail(self) -> pd.DataFrame:
        """
        Историческое имя метода. Возвращает детализацию за период
        (день или все дни месяца — зависит от mode).
        """
        if self._detail_cache is None:
            self._detail_cache = self._build_detail()
        return self._detail_cache.copy()

    def get_kpi(self) -> Dict[str, Any]:
        if self._kpi_cache is not None:
            return dict(self._kpi_cache)

        df = self.get_day_detail()

        if df.empty:
            self._kpi_cache = {
                "plan_total": 0,
                "off_total": 0,
                "fact_total": 0,
                "absent_total": 0,
                "worked_on_off_total": 0,
                "no_schedule_total": 0,
                "attendance_pct": 0.0,
                "employees_count": 0,
                "objects_count": 0,
            }
            return dict(self._kpi_cache)

        plan_total = int(df["plan_flag"].sum())
        off_total = int(df["off_flag"].sum())
        fact_total = int(df["fact_flag"].sum())
        absent_total = int(df["absent_flag"].sum())
        worked_on_off_total = int(df["worked_on_off_flag"].sum())
        no_schedule_total = int(df["no_schedule_flag"].sum())

        self._kpi_cache = {
            "plan_total": plan_total,
            "off_total": off_total,
            "fact_total": fact_total,
            "absent_total": absent_total,
            "worked_on_off_total": worked_on_off_total,
            "no_schedule_total": no_schedule_total,
            "attendance_pct": round(fact_total / plan_total * 100.0, 1) if plan_total > 0 else 0.0,
            "employees_count": int(df["person_key"].nunique()),
            "objects_count": int(df["object_name"].nunique()),
        }
        return dict(self._kpi_cache)

    def get_by_object(self) -> pd.DataFrame:
        if self._object_cache is not None:
            return self._object_cache.copy()

        df = self.get_day_detail()
        if df.empty:
            self._object_cache = pd.DataFrame(columns=[
                "object_name", "plan_count", "off_count", "fact_count",
                "absent_count", "worked_on_off_count", "no_schedule_count", "attendance_pct"
            ])
            return self._object_cache.copy()

        result = (
            df.groupby("object_name", as_index=False)
            .agg(
                plan_count=("plan_flag", "sum"),
                off_count=("off_flag", "sum"),
                fact_count=("fact_flag", "sum"),
                absent_count=("absent_flag", "sum"),
                worked_on_off_count=("worked_on_off_flag", "sum"),
                no_schedule_count=("no_schedule_flag", "sum"),
            )
            .sort_values(["plan_count", "fact_count", "object_name"], ascending=[False, False, True])
        )

        result["attendance_pct"] = result.apply(
            lambda r: round(r["fact_count"] / r["plan_count"] * 100.0, 1) if r["plan_count"] > 0 else 0.0,
            axis=1
        )

        self._object_cache = result.copy()
        return self._object_cache.copy()

    def get_by_position(self) -> pd.DataFrame:
        if self._position_cache is not None:
            return self._position_cache.copy()

        df = self.get_day_detail()
        if df.empty:
            self._position_cache = pd.DataFrame(columns=[
                "department_name", "position_name", "plan_count", "off_count",
                "fact_count", "absent_count", "worked_on_off_count",
                "no_schedule_count", "attendance_pct"
            ])
            return self._position_cache.copy()

        result = (
            df.groupby(["department_name", "position_name"], as_index=False)
            .agg(
                plan_count=("plan_flag", "sum"),
                off_count=("off_flag", "sum"),
                fact_count=("fact_flag", "sum"),
                absent_count=("absent_flag", "sum"),
                worked_on_off_count=("worked_on_off_flag", "sum"),
                no_schedule_count=("no_schedule_flag", "sum"),
            )
            .sort_values(
                ["plan_count", "fact_count", "department_name", "position_name"],
                ascending=[False, False, True, True]
            )
        )

        result["attendance_pct"] = result.apply(
            lambda r: round(r["fact_count"] / r["plan_count"] * 100.0, 1) if r["plan_count"] > 0 else 0.0,
            axis=1
        )

        self._position_cache = result.copy()
        return self._position_cache.copy()

    def get_trend(self, days_back: int = 7) -> pd.DataFrame:
        """
        В режиме day -> динамика за последние N дней.
        В режиме month -> динамика по дням внутри выбранного месяца.
        """
        if self._trend_cache is not None:
            return self._trend_cache.copy()

        if self.is_month_mode:
            df = self.get_day_detail()
            if df.empty:
                self._trend_cache = pd.DataFrame(columns=[
                    "work_date", "plan_count", "off_count", "fact_count",
                    "absent_count", "worked_on_off_count", "no_schedule_count",
                    "attendance_pct"
                ])
                return self._trend_cache.copy()

            grouped = (
                df.groupby("work_date", as_index=False)
                .agg(
                    plan_count=("plan_flag", "sum"),
                    off_count=("off_flag", "sum"),
                    fact_count=("fact_flag", "sum"),
                    absent_count=("absent_flag", "sum"),
                    worked_on_off_count=("worked_on_off_flag", "sum"),
                    no_schedule_count=("no_schedule_flag", "sum"),
                )
                .sort_values("work_date")
            )
            grouped["attendance_pct"] = grouped.apply(
                lambda r: round(r["fact_count"] / r["plan_count"] * 100.0, 1) if r["plan_count"] > 0 else 0.0,
                axis=1
            )
            self._trend_cache = grouped.copy()
            return self._trend_cache.copy()

        # day mode
        records = []
        for i in range(days_back - 1, -1, -1):
            day = self.selected_date - timedelta(days=i)
            day_provider = TimesheetPlanFactData(day, self.object_type_filter, mode="day")
            kpi = day_provider.get_kpi()
            records.append({
                "work_date": pd.Timestamp(day),
                "plan_count": kpi["plan_total"],
                "off_count": kpi["off_total"],
                "fact_count": kpi["fact_total"],
                "absent_count": kpi["absent_total"],
                "worked_on_off_count": kpi["worked_on_off_total"],
                "no_schedule_count": kpi["no_schedule_total"],
                "attendance_pct": kpi["attendance_pct"],
            })

        self._trend_cache = pd.DataFrame(records)
        return self._trend_cache.copy()

    def get_export_summary(self) -> pd.DataFrame:
        df = self.get_day_detail()

        if df.empty:
            return pd.DataFrame(columns=[
                "Объект",
                "Подразделение",
                "Должность",
                "План",
                "Факт",
            ])

        result = (
            df.groupby(["object_name", "department_name", "position_name"], as_index=False)
            .agg(
                plan_count=("plan_flag", "sum"),
                fact_count=("fact_flag", "sum"),
            )
            .sort_values(["object_name", "department_name", "position_name"], ascending=[True, True, True])
            .rename(columns={
                "object_name": "Объект",
                "department_name": "Подразделение",
                "position_name": "Должность",
                "plan_count": "План",
                "fact_count": "Факт",
            })
        )

        return result

    def get_people_summary(self) -> pd.DataFrame:
        """
        Сводка по людям за период.
        В режиме day по сути по одной строке на человека.
        В режиме month агрегирует все дни месяца по каждому человеку.
        """
        df = self.get_day_detail()

        empty_cols = [
            "fio", "tbn", "object_name", "department_name", "position_name",
            "work_schedule_name", "plan_count", "fact_count", "absent_count",
            "off_count", "worked_on_off_count", "no_schedule_count",
            "total_hours", "attendance_pct",
        ]

        if df.empty:
            return pd.DataFrame(columns=empty_cols)

        # Берём «представительные» текстовые поля по человеку (первое непустое)
        def _first_non_empty(series: pd.Series) -> str:
            for v in series:
                if _norm_text(v):
                    return v
            return "—"

        grouped = (
            df.groupby("person_key", as_index=False)
            .agg(
                fio=("fio", _first_non_empty),
                tbn=("tbn", _first_non_empty),
                object_name=("object_name", _first_non_empty),
                department_name=("department_name", _first_non_empty),
                position_name=("position_name", _first_non_empty),
                work_schedule_name=("work_schedule_name", _first_non_empty),
                plan_count=("plan_flag", "sum"),
                fact_count=("fact_flag", "sum"),
                absent_count=("absent_flag", "sum"),
                off_count=("off_flag", "sum"),
                worked_on_off_count=("worked_on_off_flag", "sum"),
                no_schedule_count=("no_schedule_flag", "sum"),
                total_hours=("hours", "sum"),
            )
        )

        grouped["attendance_pct"] = grouped.apply(
            lambda r: round(r["fact_count"] / r["plan_count"] * 100.0, 1) if r["plan_count"] > 0 else 0.0,
            axis=1
        )

        grouped = grouped.sort_values(
            ["object_name", "department_name", "position_name", "fio"],
            ascending=[True, True, True, True]
        )

        return grouped


class TimesheetPlanFactPage(ttk.Frame):
    def __init__(self, master, app_ref=None):
        super().__init__(master)
        self.app_ref = app_ref
        self.data_provider: Optional[TimesheetPlanFactData] = None
        self._loading = False

        self._build_header()
        self._build_body()
        self.load_filters()

        self.after(100, self.set_today_and_refresh)

    def _build_header(self):
        hdr = ttk.Frame(self, padding="8 6 8 6")
        hdr.pack(fill="x", side="top")

        # Переключатель режима день/месяц
        ttk.Label(hdr, text="Период:").pack(side="left", padx=(0, 4))
        self.mode_var = tk.StringVar(value="day")

        self.rb_day = ttk.Radiobutton(
            hdr, text="День", value="day",
            variable=self.mode_var, command=self._on_mode_change
        )
        self.rb_day.pack(side="left", padx=2)

        self.rb_month = ttk.Radiobutton(
            hdr, text="Месяц", value="month",
            variable=self.mode_var, command=self._on_mode_change
        )
        self.rb_month.pack(side="left", padx=2)

        # --- Блок выбора дня ---
        self.day_frame = ttk.Frame(hdr)
        self.day_frame.pack(side="left", padx=(12, 0))

        ttk.Label(self.day_frame, text="Дата:").pack(side="left", padx=(0, 4))

        self.date_var = tk.StringVar(value=datetime.today().strftime("%d.%m.%Y"))
        self.date_entry = ttk.Entry(self.day_frame, textvariable=self.date_var, width=12)
        self.date_entry.pack(side="left", padx=4)

        self.btn_today = ttk.Button(self.day_frame, text="Сегодня", command=self.set_today_and_refresh)
        self.btn_today.pack(side="left", padx=2)

        self.btn_yesterday = ttk.Button(self.day_frame, text="Вчера", command=self.set_yesterday_and_refresh)
        self.btn_yesterday.pack(side="left", padx=2)

        # --- Блок выбора месяца ---
        self.month_frame = ttk.Frame(hdr)
        # пока не упаковываем, покажем при выборе режима "month"

        ttk.Label(self.month_frame, text="Месяц:").pack(side="left", padx=(0, 4))
        self.month_var = tk.StringVar(value=RU_MONTHS[datetime.today().month])
        self.month_combo = ttk.Combobox(
            self.month_frame,
            textvariable=self.month_var,
            state="readonly",
            width=12,
            values=RU_MONTHS[1:],
        )
        self.month_combo.pack(side="left", padx=4)

        ttk.Label(self.month_frame, text="Год:").pack(side="left", padx=(8, 4))
        current_year = datetime.today().year
        self.year_var = tk.StringVar(value=str(current_year))
        self.year_combo = ttk.Combobox(
            self.month_frame,
            textvariable=self.year_var,
            state="readonly",
            width=6,
            values=[str(y) for y in range(current_year - 5, current_year + 2)],
        )
        self.year_combo.pack(side="left", padx=4)

        # Общие элементы
        ttk.Label(hdr, text="Тип объекта:").pack(side="left", padx=(12, 4))
        self.object_type_var = tk.StringVar(value="Все типы")
        self.object_type_combo = ttk.Combobox(
            hdr,
            textvariable=self.object_type_var,
            state="readonly",
            width=28,
        )
        self.object_type_combo.pack(side="left", padx=4)

        self.btn_refresh = ttk.Button(hdr, text="⟳ Обновить", command=self.refresh_data)
        self.btn_refresh.pack(side="left", padx=10)

        self.btn_export = ttk.Button(hdr, text="📥 Экспорт в Excel", command=self.export_to_excel)
        self.btn_export.pack(side="left", padx=4)

        self.last_update_var = tk.StringVar(value="")
        ttk.Label(
            hdr,
            textvariable=self.last_update_var,
            font=("Segoe UI", 8),
            foreground=PALETTE["text_muted"],
        ).pack(side="right", padx=8)
    
    def _on_mode_change(self):
        mode = self.mode_var.get()
    
        if mode == "month":
            self.day_frame.pack_forget()
            self.month_frame.pack(side="left", padx=(12, 0), before=self.object_type_combo)
        else:
            self.month_frame.pack_forget()
            self.day_frame.pack(side="left", padx=(12, 0), before=self.object_type_combo)

    def _build_body(self):
        self.canvas = tk.Canvas(self, highlightthickness=0, bg="#F5F7FA")
        self.v_scroll = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.inner = ttk.Frame(self.canvas)

        self.inner.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        self.canvas_window = self.canvas.create_window((0, 0), window=self.inner, anchor="nw")
        self.canvas.configure(yscrollcommand=self.v_scroll.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        self.v_scroll.pack(side="right", fill="y")

        self.canvas.bind(
            "<Configure>",
            lambda e: self.canvas.itemconfig(self.canvas_window, width=e.width)
        )

    def load_filters(self):
        try:
            types = TimesheetPlanFactData(datetime.now().date(), "", mode="day").get_object_types()
            self.object_type_combo["values"] = ["Все типы"] + types
            self.object_type_combo.current(0)
        except Exception:
            logging.exception("Не удалось загрузить типы объектов")
            self.object_type_combo["values"] = ["Все типы"]
            self.object_type_combo.current(0)

    def _parse_selected_date(self) -> date:
        try:
            return datetime.strptime(self.date_var.get().strip(), "%d.%m.%Y").date()
        except ValueError:
            raise ValueError("Дата должна быть в формате ДД.ММ.ГГГГ")

    def _parse_selected_month(self) -> date:
        try:
            month_name = self.month_var.get().strip()
            month_num = RU_MONTHS.index(month_name)
        except ValueError:
            raise ValueError("Не удалось определить месяц")

        try:
            year = int(self.year_var.get().strip())
        except ValueError:
            raise ValueError("Год должен быть числом")

        return date(year, month_num, 1)

    def set_today_and_refresh(self):
        self.mode_var.set("day")
        self._on_mode_change()
        self.date_var.set(datetime.today().strftime("%d.%m.%Y"))
        self.refresh_data()

    def set_yesterday_and_refresh(self):
        self.mode_var.set("day")
        self._on_mode_change()
        self.date_var.set((datetime.today() - timedelta(days=1)).strftime("%d.%m.%Y"))
        self.refresh_data()

    def _set_loading_state(self, loading: bool, status_text: str = ""):
        self._loading = loading
        state = "disabled" if loading else "normal"
        combo_state = "disabled" if loading else "readonly"

        self.rb_day.configure(state=state)
        self.rb_month.configure(state=state)
        self.date_entry.configure(state=state)
        self.btn_today.configure(state=state)
        self.btn_yesterday.configure(state=state)
        self.month_combo.configure(state=combo_state)
        self.year_combo.configure(state=combo_state)
        self.object_type_combo.configure(state=combo_state)
        self.btn_refresh.configure(state=state)
        self.btn_export.configure(state=("normal" if (not loading and self.data_provider) else "disabled"))

        if status_text:
            self.last_update_var.set(status_text)

    def _clear_inner(self):
        for w in self.inner.winfo_children():
            w.destroy()

    def _show_loading_placeholder(self, text: str):
        self._clear_inner()
        wrap = ttk.Frame(self.inner)
        wrap.pack(fill="both", expand=True, padx=20, pady=30)
        ttk.Label(wrap, text=text, font=("Segoe UI", 11)).pack(pady=20)

    def refresh_data(self):
        if self._loading:
            return

        mode = self.mode_var.get()

        try:
            if mode == "month":
                selected_date = self._parse_selected_month()
            else:
                selected_date = self._parse_selected_date()
        except Exception as e:
            messagebox.showerror("План / факт", str(e))
            return

        obj_filter = self.object_type_var.get()
        if obj_filter == "Все типы":
            obj_filter = ""

        if mode == "month":
            period_text = f"{RU_MONTHS[selected_date.month]} {selected_date.year}"
        else:
            period_text = selected_date.strftime("%d.%m.%Y")

        self._set_loading_state(True, f"Загрузка... {period_text}")
        self._show_loading_placeholder("Загрузка аналитики по выбранному периоду...")

        def worker():
            try:
                dp = TimesheetPlanFactData(selected_date, obj_filter, mode=mode)
                dp.get_kpi()
                dp.get_by_object()
                dp.get_by_position()
                dp.get_day_detail()
                dp.get_trend(7)
                self.after(0, lambda: self._on_data_loaded(dp))
            except Exception as e:
                logging.exception("Ошибка загрузки аналитики план/факт")
                self.after(0, lambda: self._on_data_error(e))

        threading.Thread(target=worker, daemon=True).start()

    def _on_data_loaded(self, dp: TimesheetPlanFactData):
        self.data_provider = dp
        self._render()
        self._set_loading_state(
            False,
            f"Обновлено: {datetime.now().strftime('%H:%M:%S')} | Период: {dp.period_label}"
        )

    def _on_data_error(self, error: Exception):
        self._set_loading_state(False, "Ошибка загрузки данных")
        self._show_loading_placeholder("Не удалось загрузить данные.")
        messagebox.showerror("План / факт", f"Не удалось загрузить данные:\n{error}")

    def _create_card(self, parent, title: str, value: str, unit: str, color: str, note: str = ""):
        card = tk.Frame(
            parent,
            bg="white",
            highlightbackground="#DDE3EA",
            highlightthickness=1,
        )

        tk.Frame(card, bg=color, height=4).pack(fill="x", side="top")

        inner = tk.Frame(card, bg="white", padx=12, pady=10)
        inner.pack(fill="both", expand=True)

        tk.Label(
            inner,
            text=title,
            font=("Segoe UI", 9, "bold"),
            fg="#455A64",
            bg="white",
            anchor="w",
        ).pack(fill="x")

        tk.Label(
            inner,
            text=value,
            font=("Segoe UI", 22, "bold"),
            fg=color,
            bg="white",
            anchor="w",
        ).pack(fill="x", pady=(5, 0))

        bottom_text = unit if not note else f"{unit} · {note}"

        tk.Label(
            inner,
            text=bottom_text,
            font=("Segoe UI", 8),
            fg=PALETTE["text_muted"],
            bg="white",
            anchor="w",
        ).pack(fill="x", pady=(2, 0))

        return card

    def _create_treeview(self, parent, columns: List[Tuple[str, str]], height: int = 10) -> ttk.Treeview:
        tree = ttk.Treeview(parent, columns=[c[0] for c in columns], show="headings", height=height)
        for col_id, col_text in columns:
            tree.heading(col_id, text=col_text)
            tree.column(col_id, anchor="w", width=120)

        tree.tag_configure("odd", background="#F5F7FA")
        tree.tag_configure("even", background="white")

        vsb = ttk.Scrollbar(parent, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(parent, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        parent.grid_rowconfigure(0, weight=1)
        parent.grid_columnconfigure(0, weight=1)
        return tree

    def _insert_rows(self, tree: ttk.Treeview, rows: list):
        for i, vals in enumerate(rows):
            tree.insert("", "end", values=vals, tags=("odd" if i % 2 == 0 else "even",))

    def _render(self):
        self._clear_inner()

        dp = self.data_provider
        if not dp:
            return

        kpi = dp.get_kpi()

        period_text = dp.period_label
        object_filter_text = dp.object_type_filter if dp.object_type_filter else "Все типы объектов"
        is_month = dp.is_month_mode

        # Подписи единиц зависят от режима
        unit_label = "чел.-дн." if is_month else "чел."

        # Верхний информационный блок
        hero = tk.Frame(
            self.inner,
            bg="white",
            highlightbackground="#DDE3EA",
            highlightthickness=1,
        )
        hero.pack(fill="x", padx=10, pady=10)

        hero_inner = tk.Frame(hero, bg="white", padx=14, pady=12)
        hero_inner.pack(fill="x")

        mode_title = "Аналитика выходов сотрудников за месяц" if is_month \
            else "Аналитика выходов сотрудников за день"

        tk.Label(
            hero_inner,
            text=mode_title,
            font=("Segoe UI", 15, "bold"),
            fg="#263238",
            bg="white",
            anchor="w",
        ).pack(fill="x")

        period_prefix = "Месяц" if is_month else "Дата"

        tk.Label(
            hero_inner,
            text=(
                f"{period_prefix}: {period_text}  ·  "
                f"Фильтр: {object_filter_text}  ·  "
                f"Сотрудников в выборке: {kpi.get('employees_count', 0)}  ·  "
                f"Объектов: {kpi.get('objects_count', 0)}"
            ),
            font=("Segoe UI", 9),
            fg=PALETTE["text_muted"],
            bg="white",
            anchor="w",
        ).pack(fill="x", pady=(4, 0))

        # KPI
        kpi_frame = tk.Frame(self.inner, bg="#F5F7FA")
        kpi_frame.pack(fill="x", padx=10, pady=(0, 10))

        cards = [
            ("План", str(kpi.get("plan_total", 0)), unit_label, PALETTE["primary"], "по графику"),
            ("Факт", str(kpi.get("fact_total", 0)), unit_label, PALETTE["success"], "есть часы в табеле"),
            ("Выходной", str(kpi.get("off_total", 0)), unit_label, PALETTE["offday"], "по графику"),
            ("Вышли в выходной", str(kpi.get("worked_on_off_total", 0)), unit_label, PALETTE["warning"], "есть часы"),
            ("Без графика", str(kpi.get("no_schedule_total", 0)), unit_label, PALETTE["noschedule"], "не найден график"),
        ]

        for i, (title, value, unit, color, note) in enumerate(cards):
            c = self._create_card(kpi_frame, title, value, unit, color, note)
            c.grid(row=0, column=i, padx=5, pady=4, sticky="nsew")
            kpi_frame.grid_columnconfigure(i, weight=1)

        # По объектам
        df_obj = dp.get_by_object()

        obj_wrap = self._create_section(
            self.inner,
            "Сводка по объектам",
            "План, факт и отклонения по каждому объекту за выбранный период.",
        )

        tree_obj = self._create_treeview(
            obj_wrap,
            [
                ("object", "Объект"),
                ("plan", "План"),
                ("fact", "Факт"),
                ("absent", "Не вышли"),
                ("off", "Выходной"),
                ("workoff", "Вышли в вых."),
                ("noschedule", "Без графика"),
                ("pct", "Явка, %"),
            ],
            height=10,
        )

        tree_obj.column("object", width=360)
        for col in ["plan", "fact", "absent", "off", "workoff", "noschedule", "pct"]:
            tree_obj.column(col, width=105, anchor="e")

        self._insert_rows(tree_obj, [
            (
                r["object_name"],
                int(r["plan_count"]),
                int(r["fact_count"]),
                int(r["absent_count"]),
                int(r["off_count"]),
                int(r["worked_on_off_count"]),
                int(r["no_schedule_count"]),
                f"{float(r['attendance_pct']):.1f}",
            )
            for _, r in df_obj.iterrows()
        ])

        # По подразделениям и должностям
        df_pos = dp.get_by_position()

        pos_wrap = self._create_section(
            self.inner,
            "Сводка по подразделениям и должностям",
            "Помогает понять, по каким подразделениям и должностям есть недовыходы или выходы вне графика.",
        )

        tree_pos = self._create_treeview(
            pos_wrap,
            [
                ("department", "Подразделение"),
                ("position", "Должность"),
                ("plan", "План"),
                ("fact", "Факт"),
                ("absent", "Не вышли"),
                ("off", "Выходной"),
                ("workoff", "Вышли в вых."),
                ("noschedule", "Без графика"),
                ("pct", "Явка, %"),
            ],
            height=10,
        )

        tree_pos.column("department", width=220)
        tree_pos.column("position", width=220)
        for col in ["plan", "fact", "absent", "off", "workoff", "noschedule", "pct"]:
            tree_pos.column(col, width=100, anchor="e")

        self._insert_rows(tree_pos, [
            (
                r["department_name"],
                r["position_name"],
                int(r["plan_count"]),
                int(r["fact_count"]),
                int(r["absent_count"]),
                int(r["off_count"]),
                int(r["worked_on_off_count"]),
                int(r["no_schedule_count"]),
                f"{float(r['attendance_pct']):.1f}",
            )
            for _, r in df_pos.iterrows()
        ])

        # Динамика
        df_trend = dp.get_trend(7)

        if dp.is_month_mode:
            trend_title = "Динамика по дням месяца"
            trend_subtitle = "План-факт показатели по каждому дню выбранного месяца."
        else:
            trend_title = "Динамика за последние 7 дней"
            trend_subtitle = "Краткая история план-факт показателей за неделю."

        trend_wrap = self._create_section(
            self.inner,
            trend_title,
            trend_subtitle,
        )

        # Высота таблицы динамики: для месяца побольше
        trend_height = 15 if dp.is_month_mode else 7

        tree_trend = self._create_treeview(
            trend_wrap,
            [
                ("date", "Дата"),
                ("plan", "План"),
                ("fact", "Факт"),
                ("absent", "Не вышли"),
                ("off", "Выходной"),
                ("workoff", "Вышли в вых."),
                ("noschedule", "Без графика"),
                ("pct", "Явка, %"),
            ],
            height=trend_height,
        )

        tree_trend.column("date", width=120)
        for col in ["plan", "fact", "absent", "off", "workoff", "noschedule", "pct"]:
            tree_trend.column(col, width=105, anchor="e")

        self._insert_rows(tree_trend, [
            (
                pd.to_datetime(r["work_date"]).strftime("%d.%m.%Y"),
                int(r["plan_count"]),
                int(r["fact_count"]),
                int(r["absent_count"]),
                int(r["off_count"]),
                int(r["worked_on_off_count"]),
                int(r["no_schedule_count"]),
                f"{float(r['attendance_pct']):.1f}",
            )
            for _, r in df_trend.iterrows()
        ])

        # Детализация по людям
        if dp.is_month_mode:
            self._render_people_month(dp)
        else:
            self._render_people_day(dp)

    def _render_people_day(self, dp: TimesheetPlanFactData):
        """Детализация по людям за день (как было раньше)."""
        detail_df = dp.get_day_detail().copy()

        people_wrap = self._create_section(
            self.inner,
            "Сотрудники на выбранную дату",
            "Детальная расшифровка по людям: объект, должность, график, часы и итоговый статус.",
        )

        tree_det = self._create_treeview(
            people_wrap,
            [
                ("fio", "ФИО"),
                ("tbn", "Таб. №"),
                ("object", "Объект"),
                ("dept", "Подразделение"),
                ("pos", "Должность"),
                ("schedule", "График"),
                ("hours", "Часы"),
                ("status", "Статус"),
            ],
            height=15,
        )

        tree_det.column("fio", width=240)
        tree_det.column("tbn", width=90)
        tree_det.column("object", width=260)
        tree_det.column("dept", width=180)
        tree_det.column("pos", width=180)
        tree_det.column("schedule", width=120)
        tree_det.column("hours", width=80, anchor="e")
        tree_det.column("status", width=160)

        if not detail_df.empty:
            detail_df["status_text"] = detail_df.apply(self._get_status_text, axis=1)

        if detail_df.empty:
            return

        self._insert_rows(tree_det, [
            (
                r["fio"],
                r["tbn"],
                r["object_name"],
                r["department_name"],
                r["position_name"],
                r["work_schedule_name"],
                f"{float(r['hours']):.2f}",
                r["status_text"],
            )
            for _, r in detail_df.sort_values(
                ["object_name", "department_name", "position_name", "fio"]
            ).iterrows()
        ])

    def _render_people_month(self, dp: TimesheetPlanFactData):
        """Сводка по людям за месяц (агрегаты человеко-дней)."""
        people_df = dp.get_people_summary().copy()

        people_wrap = self._create_section(
            self.inner,
            "Сотрудники за месяц (сводка)",
            "По каждому сотруднику: сколько дней по плану, фактических выходов, невыходов и суммарные часы за месяц.",
        )

        tree_det = self._create_treeview(
            people_wrap,
            [
                ("fio", "ФИО"),
                ("tbn", "Таб. №"),
                ("object", "Объект"),
                ("dept", "Подразделение"),
                ("pos", "Должность"),
                ("plan", "План, дн."),
                ("fact", "Факт, дн."),
                ("absent", "Не вышел, дн."),
                ("workoff", "В вых., дн."),
                ("hours", "Часы за мес."),
                ("pct", "Явка, %"),
            ],
            height=18,
        )

        tree_det.column("fio", width=220)
        tree_det.column("tbn", width=80)
        tree_det.column("object", width=240)
        tree_det.column("dept", width=160)
        tree_det.column("pos", width=160)
        for col in ["plan", "fact", "absent", "workoff", "hours", "pct"]:
            tree_det.column(col, width=95, anchor="e")

        if people_df.empty:
            return

        self._insert_rows(tree_det, [
            (
                r["fio"],
                r["tbn"],
                r["object_name"],
                r["department_name"],
                r["position_name"],
                int(r["plan_count"]),
                int(r["fact_count"]),
                int(r["absent_count"]),
                int(r["worked_on_off_count"]),
                f"{float(r['total_hours']):.2f}",
                f"{float(r['attendance_pct']):.1f}",
            )
            for _, r in people_df.iterrows()
        ])

    @staticmethod
    def _strip_tz(df: pd.DataFrame) -> pd.DataFrame:
        if df is None or df.empty:
            return df
        df = df.copy()
        for col in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                try:
                    if hasattr(df[col].dt, "tz") and df[col].dt.tz is not None:
                        df[col] = df[col].dt.tz_convert("UTC").dt.tz_localize(None)
                except Exception:
                    pass
        return df

    def export_to_excel(self):
        if not self.data_provider:
            messagebox.showwarning("Экспорт", "Сначала загрузите данные.")
            return

        dp = self.data_provider

        if dp.is_month_mode:
            period_part = f"{dp.selected_date.year}{dp.selected_date.month:02d}"
        else:
            period_part = dp.selected_date.strftime("%Y%m%d")

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel файл", "*.xlsx")],
            title="Сохранить отчёт",
            initialfile=f"план_факт_выходов_{period_part}_{datetime.now().strftime('%H%M')}.xlsx",
        )

        if not path:
            return

        try:
            frames = self._build_excel_frames(dp)

            table_names = {
                "KPI": "tbl_kpi",
                "По объектам": "tbl_by_object",
                "По должностям": "tbl_by_position",
                "Динамика по дням": "tbl_trend",
                "Сотрудники": "tbl_people",
                "Свод объект-подразделение": "tbl_summary",
            }

            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                for sheet_name, df in frames.items():
                    df = self._sanitize_excel_df(df)
                    df = self._strip_tz(df)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

                    ws = writer.sheets[sheet_name]
                    self._format_excel_sheet(
                        ws,
                        table_names.get(sheet_name, f"tbl_{len(writer.sheets)}"),
                    )

                wb = writer.book
                wb.properties.title = "План-факт выходов сотрудников"
                wb.properties.subject = f"Период отчёта: {dp.period_label}"
                wb.properties.creator = "Модуль аналитики табелей"

            messagebox.showinfo("Экспорт завершён", f"Файл успешно сохранён:\n{path}")

        except Exception as e:
            logging.exception("Ошибка экспорта")
            messagebox.showerror("Ошибка экспорта", f"Не удалось сохранить файл:\n{e}")

    def _get_status_text(self, r) -> str:
        if int(r.get("worked_on_off_flag", 0)) == 1:
            return "Вышел в выходной"

        if int(r.get("fact_flag", 0)) == 1 and int(r.get("plan_flag", 0)) == 1:
            return "Вышел по графику"

        if int(r.get("fact_flag", 0)) == 1 and int(r.get("no_schedule_flag", 0)) == 1:
            return "Вышел без графика"

        if int(r.get("absent_flag", 0)) == 1:
            return "Не вышел"

        if int(r.get("off_flag", 0)) == 1:
            return "Выходной"

        if int(r.get("no_schedule_flag", 0)) == 1:
            return "Нет графика"

        return "Не определено"

    def _create_section(self, parent, title: str, subtitle: str = ""):
        section = tk.Frame(
            parent,
            bg="white",
            highlightbackground="#DDE3EA",
            highlightthickness=1,
        )
        section.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        header = tk.Frame(section, bg="white", padx=10, pady=8)
        header.pack(fill="x")

        tk.Label(
            header,
            text=title,
            font=("Segoe UI", 11, "bold"),
            fg="#263238",
            bg="white",
            anchor="w",
        ).pack(fill="x")

        if subtitle:
            tk.Label(
                header,
                text=subtitle,
                font=("Segoe UI", 8),
                fg=PALETTE["text_muted"],
                bg="white",
                anchor="w",
            ).pack(fill="x", pady=(2, 0))

        body = ttk.Frame(section)
        body.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        return body

    def _build_excel_frames(self, dp: TimesheetPlanFactData) -> Dict[str, pd.DataFrame]:
        kpi = dp.get_kpi()
        is_month = dp.is_month_mode
        unit = "чел.-дн." if is_month else "чел."

        kpi_df = pd.DataFrame([
            {"Показатель": "Период", "Значение": dp.period_label, "Ед. изм.": ""},
            {"Показатель": "План по графику", "Значение": kpi.get("plan_total", 0), "Ед. изм.": unit},
            {"Показатель": "Факт выходов", "Значение": kpi.get("fact_total", 0), "Ед. изм.": unit},
            {"Показатель": "Выходной по графику", "Значение": kpi.get("off_total", 0), "Ед. изм.": unit},
            {"Показатель": "Вышли в выходной", "Значение": kpi.get("worked_on_off_total", 0), "Ед. изм.": unit},
            {"Показатель": "Без графика", "Значение": kpi.get("no_schedule_total", 0), "Ед. изм.": unit},
            {"Показатель": "Сотрудников в выборке", "Значение": kpi.get("employees_count", 0), "Ед. изм.": "чел."},
            {"Показатель": "Объектов в выборке", "Значение": kpi.get("objects_count", 0), "Ед. изм.": "объектов"},
        ])

        by_object_df = dp.get_by_object().rename(columns={
            "object_name": "Объект",
            "plan_count": "План",
            "fact_count": "Факт",
            "absent_count": "Не вышли",
            "off_count": "Выходной",
            "worked_on_off_count": "Вышли в выходной",
            "no_schedule_count": "Без графика",
            "attendance_pct": "Явка, %",
        })

        by_object_df = by_object_df[
            ["Объект", "План", "Факт", "Не вышли", "Выходной",
             "Вышли в выходной", "Без графика", "Явка, %"]
        ]

        by_position_df = dp.get_by_position().rename(columns={
            "department_name": "Подразделение",
            "position_name": "Должность",
            "plan_count": "План",
            "fact_count": "Факт",
            "absent_count": "Не вышли",
            "off_count": "Выходной",
            "worked_on_off_count": "Вышли в выходной",
            "no_schedule_count": "Без графика",
            "attendance_pct": "Явка, %",
        })

        by_position_df = by_position_df[
            ["Подразделение", "Должность", "План", "Факт", "Не вышли",
             "Выходной", "Вышли в выходной", "Без графика", "Явка, %"]
        ]

        trend_df = dp.get_trend(7).copy()

        if not trend_df.empty:
            trend_df["work_date"] = pd.to_datetime(trend_df["work_date"]).dt.date

        trend_df = trend_df.rename(columns={
            "work_date": "Дата",
            "plan_count": "План",
            "fact_count": "Факт",
            "absent_count": "Не вышли",
            "off_count": "Выходной",
            "worked_on_off_count": "Вышли в выходной",
            "no_schedule_count": "Без графика",
            "attendance_pct": "Явка, %",
        })

        if not trend_df.empty:
            trend_df = trend_df[
                ["Дата", "План", "Факт", "Не вышли", "Выходной",
                 "Вышли в выходной", "Без графика", "Явка, %"]
            ]
        else:
            trend_df = pd.DataFrame(columns=[
                "Дата", "План", "Факт", "Не вышли", "Выходной",
                "Вышли в выходной", "Без графика", "Явка, %"
            ])

        # Лист "Сотрудники" формируем по-разному в зависимости от режима
        if is_month:
            people_df = self._build_people_month_frame(dp)
        else:
            people_df = self._build_people_day_frame(dp)

        summary_df = dp.get_export_summary().copy()

        return {
            "KPI": kpi_df,
            "По объектам": by_object_df,
            "По должностям": by_position_df,
            "Динамика по дням": trend_df,
            "Сотрудники": people_df,
            "Свод объект-подразделение": summary_df,
        }

    def _build_people_day_frame(self, dp: TimesheetPlanFactData) -> pd.DataFrame:
        detail_df = dp.get_day_detail().copy()

        cols = [
            "Дата", "Объект", "Тип объекта", "Подразделение", "Должность",
            "ФИО", "Табельный номер", "График", "Рабочий день по графику",
            "Плановые часы", "Значение в табеле", "Часы", "Статус",
        ]

        if detail_df.empty:
            return pd.DataFrame(columns=cols)

        detail_df["Статус"] = detail_df.apply(self._get_status_text, axis=1)
        detail_df["work_date"] = pd.to_datetime(detail_df["work_date"]).dt.date

        detail_df["Рабочий день по графику"] = detail_df["is_workday"].map({
            True: "Да",
            False: "Нет",
        }).fillna("—")

        detail_df = detail_df.rename(columns={
            "work_date": "Дата",
            "object_name": "Объект",
            "object_type": "Тип объекта",
            "department_name": "Подразделение",
            "position_name": "Должность",
            "fio": "ФИО",
            "tbn": "Табельный номер",
            "work_schedule_name": "График",
            "day_raw": "Значение в табеле",
            "hours": "Часы",
            "planned_hours": "Плановые часы",
        })

        return detail_df[cols]

    def _build_people_month_frame(self, dp: TimesheetPlanFactData) -> pd.DataFrame:
        people_df = dp.get_people_summary().copy()

        cols = [
            "Объект", "Подразделение", "Должность", "ФИО", "Табельный номер",
            "График", "План, дн.", "Факт, дн.", "Не вышел, дн.",
            "Выходной, дн.", "В вых., дн.", "Без графика, дн.",
            "Часы за месяц", "Явка, %",
        ]

        if people_df.empty:
            return pd.DataFrame(columns=cols)

        people_df = people_df.rename(columns={
            "object_name": "Объект",
            "department_name": "Подразделение",
            "position_name": "Должность",
            "fio": "ФИО",
            "tbn": "Табельный номер",
            "work_schedule_name": "График",
            "plan_count": "План, дн.",
            "fact_count": "Факт, дн.",
            "absent_count": "Не вышел, дн.",
            "off_count": "Выходной, дн.",
            "worked_on_off_count": "В вых., дн.",
            "no_schedule_count": "Без графика, дн.",
            "total_hours": "Часы за месяц",
            "attendance_pct": "Явка, %",
        })

        return people_df[cols]

    def _format_excel_sheet(self, ws, table_name: str):
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"

        header_fill = PatternFill("solid", fgColor="1565C0")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        thin_side = Side(style="thin", color="D9E2EC")
        border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side,
        )

        even_fill = PatternFill("solid", fgColor="F5F7FA")

        max_row = ws.max_row
        max_col = ws.max_column

        if max_row < 1 or max_col < 1:
            return

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        for row_idx in range(2, max_row + 1):
            for cell in ws[row_idx]:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

                if row_idx % 2 == 0:
                    cell.fill = even_fill

        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0

            for row_idx in range(1, max_row + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is None:
                    continue

                value_len = len(str(value))
                if value_len > max_len:
                    max_len = value_len

            width = min(max(max_len + 2, 12), 55)
            ws.column_dimensions[col_letter].width = width

        ws.row_dimensions[1].height = 28

        for row_idx in range(2, max_row + 1):
            ws.row_dimensions[row_idx].height = 22

        if max_row >= 2:
            ref = f"A1:{get_column_letter(max_col)}{max_row}"

            tab = Table(displayName=table_name, ref=ref)
            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            tab.tableStyleInfo = style
            ws.add_table(tab)

    def _clean_excel_value(self, value):
        """
        Убирает значения и символы, которые могут ломать xlsx.
        """
        if value is None:
            return ""

        try:
            if pd.isna(value):
                return ""
        except Exception:
            pass

        if isinstance(value, float):
            if value == float("inf") or value == float("-inf"):
                return ""

        if isinstance(value, str):
            value = ILLEGAL_CHARACTERS_RE.sub("", value)
            return value

        return value

    def _sanitize_excel_df(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Подготавливает DataFrame для безопасной записи в Excel.
        """
        if df is None:
            return pd.DataFrame()

        df = df.copy()

        new_columns = []
        seen = {}

        for col in df.columns:
            col_name = str(self._clean_excel_value(col)).strip()
            if not col_name:
                col_name = "Колонка"

            if col_name in seen:
                seen[col_name] += 1
                col_name = f"{col_name}_{seen[col_name]}"
            else:
                seen[col_name] = 1

            new_columns.append(col_name)

        df.columns = new_columns

        df = df.replace([float("inf"), float("-inf")], "")

        for col in df.columns:
            df[col] = df[col].map(self._clean_excel_value)

        return df
