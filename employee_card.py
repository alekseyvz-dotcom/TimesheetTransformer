# employee_card.py
from __future__ import annotations

import threading
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime

from psycopg2 import pool
from psycopg2.extras import RealDictCursor

db_connection_pool: Optional[pool.SimpleConnectionPool] = None


def set_db_pool(db_pool: pool.SimpleConnectionPool):
    global db_connection_pool
    db_connection_pool = db_pool


def _fetch_all(query: str, params: Tuple[Any, ...] = ()) -> List[Dict[str, Any]]:
    if not db_connection_pool:
        raise RuntimeError("DB pool is not set (employee_card.set_db_pool was not called).")
    conn = None
    try:
        conn = db_connection_pool.getconn()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(query, params)
            return [dict(r) for r in cur.fetchall()]
    finally:
        if conn:
            db_connection_pool.putconn(conn)


def _fetch_one(query: str, params: Tuple[Any, ...] = ()) -> Optional[Dict[str, Any]]:
    rows = _fetch_all(query, params)
    return rows[0] if rows else None


# =========================
# Data queries
# =========================

def search_employees(needle: str, limit: int = 200) -> List[Dict[str, Any]]:
    needle = (needle or "").strip()
    if not needle:
        # по умолчанию показываем первых активных
        return _fetch_all(
            """
            SELECT e.id, e.fio, e.tbn, e.position,
                   COALESCE(d.name,'') AS department,
                   COALESCE(e.is_fired,false) AS is_fired
            FROM employees e
            LEFT JOIN departments d ON d.id = e.department_id
            ORDER BY COALESCE(e.is_fired,false), e.fio
            LIMIT %s
            """,
            (limit,),
        )

    like = f"%{needle}%"
    return _fetch_all(
        """
        SELECT e.id, e.fio, e.tbn, e.position,
               COALESCE(d.name,'') AS department,
               COALESCE(e.is_fired,false) AS is_fired
        FROM employees e
        LEFT JOIN departments d ON d.id = e.department_id
        WHERE e.tbn ILIKE %s OR e.fio ILIKE %s
        ORDER BY COALESCE(e.is_fired,false), e.fio
        LIMIT %s
        """,
        (like, like, limit),
    )


def load_employee_profile(employee_id: int) -> Optional[Dict[str, Any]]:
    return _fetch_one(
        """
        SELECT e.id, e.fio, e.tbn, e.position,
               COALESCE(d.name,'') AS department,
               COALESCE(e.is_fired,false) AS is_fired
        FROM employees e
        LEFT JOIN departments d ON d.id = e.department_id
        WHERE e.id = %s
        """,
        (employee_id,),
    )


def load_timesheet_summary_by_month(tbn: str) -> List[Dict[str, Any]]:
    return _fetch_all(
        """
        SELECT th.year, th.month,
               COALESCE(SUM(tr.total_days),0)::int AS days,
               COALESCE(SUM(tr.total_hours),0)::numeric AS hours,
               COALESCE(SUM(tr.night_hours),0)::numeric AS night_hours,
               COALESCE(SUM(tr.overtime_day),0)::numeric AS ot_day,
               COALESCE(SUM(tr.overtime_night),0)::numeric AS ot_night
        FROM timesheet_rows tr
        JOIN timesheet_headers th ON th.id = tr.header_id
        WHERE tr.tbn = %s
        GROUP BY th.year, th.month
        ORDER BY th.year DESC, th.month DESC
        """,
        (tbn,),
    )


def load_timesheet_by_object(tbn: str, limit: int = 30) -> List[Dict[str, Any]]:
    return _fetch_all(
        """
        SELECT o.address,
               COALESCE(SUM(tr.total_hours),0)::numeric AS hours,
               COALESCE(SUM(tr.total_days),0)::int AS days,
               COUNT(DISTINCT (th.year*100+th.month))::int AS months_cnt
        FROM timesheet_rows tr
        JOIN timesheet_headers th ON th.id = tr.header_id
        JOIN objects o ON o.id = th.object_db_id
        WHERE tr.tbn = %s
        GROUP BY o.address
        ORDER BY hours DESC
        LIMIT %s
        """,
        (tbn, limit),
    )


def load_meals_kpi(tbn: str) -> Dict[str, Any]:
    row = _fetch_one(
        """
        SELECT
          COUNT(*)::int AS rows_cnt,
          COUNT(DISTINCT mo.date)::int AS days_cnt,
          COALESCE(SUM(moi.quantity),0)::numeric AS qty_sum
        FROM meal_order_items moi
        JOIN meal_orders mo ON mo.id = moi.order_id
        WHERE moi.tbn_text = %s
        """,
        (tbn,),
    ) or {}
    return {
        "rows_cnt": int(row.get("rows_cnt", 0) or 0),
        "days_cnt": int(row.get("days_cnt", 0) or 0),
        "qty_sum": float(row.get("qty_sum", 0) or 0),
    }


def load_meals_by_object(tbn: str, limit: int = 30) -> List[Dict[str, Any]]:
    return _fetch_all(
        """
        SELECT
          COALESCE(mo.fact_address, o.address, '—') AS address,
          COUNT(DISTINCT mo.date)::int AS days_cnt,
          COUNT(*)::int AS rows_cnt,
          COALESCE(SUM(moi.quantity),0)::numeric AS qty_sum
        FROM meal_order_items moi
        JOIN meal_orders mo ON mo.id = moi.order_id
        LEFT JOIN objects o ON o.id = mo.object_id
        WHERE moi.tbn_text = %s
        GROUP BY COALESCE(mo.fact_address, o.address, '—')
        ORDER BY days_cnt DESC, qty_sum DESC
        LIMIT %s
        """,
        (tbn, limit),
    )


def load_meals_history(tbn: str, limit: int = 1000) -> List[Dict[str, Any]]:
    # limit нужен, чтобы не тащить бесконечно (в UI можно сделать "показать ещё")
    return _fetch_all(
        """
        SELECT
          mo.date,
          COALESCE(mo.fact_address, o.address, '—') AS address,
          COALESCE(d.name,'') AS department,
          COALESCE(mo.team_name,'') AS team_name,
          COALESCE(mt.name, moi.meal_type_text, '') AS meal_type,
          COALESCE(moi.quantity,1)::numeric AS qty
        FROM meal_order_items moi
        JOIN meal_orders mo ON mo.id = moi.order_id
        LEFT JOIN objects o ON o.id = mo.object_id
        LEFT JOIN departments d ON d.id = mo.department_id
        LEFT JOIN meal_types mt ON mt.id = moi.meal_type_id
        WHERE moi.tbn_text = %s
        ORDER BY mo.date DESC, address, meal_type
        LIMIT %s
        """,
        (tbn, limit),
    )


def load_lodging_current(employee_id: int) -> Optional[Dict[str, Any]]:
    return _fetch_one(
        """
        SELECT
          s.check_in, s.check_out, s.status,
          d.name AS dorm_name,
          r.room_no, r.capacity
        FROM dorm_stays s
        JOIN dorms d ON d.id = s.dorm_id
        JOIN dorm_rooms r ON r.id = s.room_id
        WHERE s.employee_id = %s
          AND s.status='active'
          AND s.check_out IS NULL
        ORDER BY s.check_in DESC
        LIMIT 1
        """,
        (employee_id,),
    )


def load_lodging_history(employee_id: int) -> List[Dict[str, Any]]:
    return _fetch_all(
        """
        SELECT
          s.id AS stay_id,
          s.check_in,
          s.check_out,
          s.status,
          d.name AS dorm_name,
          r.room_no
        FROM dorm_stays s
        JOIN dorms d ON d.id = s.dorm_id
        JOIN dorm_rooms r ON r.id = s.room_id
        WHERE s.employee_id = %s
        ORDER BY s.check_in DESC
        """,
        (employee_id,),
    )


def load_lodging_bed_days_total(employee_id: int) -> int:
    row = _fetch_one(
        """
        SELECT
          COALESCE(SUM((COALESCE(s.check_out, CURRENT_DATE) - s.check_in)),0)::int AS bed_days_total
        FROM dorm_stays s
        WHERE s.employee_id = %s
          AND s.status IN ('active','closed')
        """,
        (employee_id,),
    ) or {}
    return int(row.get("bed_days_total", 0) or 0)


def load_lodging_charges(employee_id: int) -> List[Dict[str, Any]]:
    return _fetch_all(
        """
        SELECT
          dc.year, dc.month, dc.days,
          dc.amount, dc.avg_price_per_day,
          dc.rate_source,
          dc.stay_id
        FROM dorm_charges dc
        JOIN dorm_stays s ON s.id = dc.stay_id
        WHERE s.employee_id = %s
        ORDER BY dc.year DESC, dc.month DESC
        """,
        (employee_id,),
    )


# =========================
# UI
# =========================

class EmployeeCardPage(tk.Frame):
    def __init__(self, master, app_ref=None):
        super().__init__(master, bg="#f7f7f7")
        self.app_ref = app_ref

        self._selected_employee: Optional[Dict[str, Any]] = None

        # background worker guard
        self._load_token = 0

        self._build_ui()
        self._search()  # initial load

    def _build_ui(self):
        root = tk.Frame(self, bg="#f7f7f7")
        root.pack(fill="both", expand=True, padx=10, pady=10)

        root.grid_columnconfigure(0, weight=1)
        root.grid_columnconfigure(1, weight=3)
        root.grid_rowconfigure(0, weight=1)

        # left: search + list
        left = tk.LabelFrame(root, text="Поиск сотрудника", bg="#f7f7f7")
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 10))

        top = tk.Frame(left, bg="#f7f7f7")
        top.pack(fill="x", padx=8, pady=8)

        tk.Label(top, text="ФИО / Таб№:", bg="#f7f7f7").pack(side="left")
        self.var_q = tk.StringVar()
        ent = ttk.Entry(top, textvariable=self.var_q, width=30)
        ent.pack(side="left", padx=(6, 6))
        ent.bind("<Return>", lambda e: self._search())

        ttk.Button(top, text="Найти", command=self._search).pack(side="left")

        self.tree = ttk.Treeview(
            left,
            columns=("fio", "tbn", "dep", "fired"),
            show="headings",
            height=20,
            selectmode="browse",
        )
        self.tree.heading("fio", text="ФИО")
        self.tree.heading("tbn", text="Таб№")
        self.tree.heading("dep", text="Подразделение")
        self.tree.heading("fired", text="Уволен")

        self.tree.column("fio", width=240)
        self.tree.column("tbn", width=90, anchor="center")
        self.tree.column("dep", width=150)
        self.tree.column("fired", width=60, anchor="center")

        vsb = ttk.Scrollbar(left, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)

        self.tree.pack(side="left", fill="both", expand=True, padx=(8, 0), pady=(0, 8))
        vsb.pack(side="right", fill="y", padx=(0, 8), pady=(0, 8))

        self.tree.bind("<<TreeviewSelect>>", lambda e: self._on_select_employee())

        # right: card notebook
        right = tk.Frame(root, bg="#f7f7f7")
        right.grid(row=0, column=1, sticky="nsew")

        self.lbl_title = tk.Label(
            right,
            text="Карточка сотрудника",
            font=("Segoe UI", 14, "bold"),
            bg="#f7f7f7",
        )
        self.lbl_title.pack(anchor="w", pady=(0, 8))

        self.status = tk.Label(right, text="", fg="#555", bg="#f7f7f7")
        self.status.pack(anchor="w", pady=(0, 6))

        btns = tk.Frame(right, bg="#f7f7f7")
        btns.pack(fill="x", pady=(0, 6))
        
        ttk.Button(btns, text="Экспорт досье…", command=self._export_dossier).pack(side="left")

        self.nb = ttk.Notebook(right)
        self.nb.pack(fill="both", expand=True)

        self.tab_profile = ttk.Frame(self.nb)
        self.tab_work = ttk.Frame(self.nb)
        self.tab_meals = ttk.Frame(self.nb)
        self.tab_lodging = ttk.Frame(self.nb)

        self.nb.add(self.tab_profile, text="Профиль")
        self.nb.add(self.tab_work, text="Работа (табели)")
        self.nb.add(self.tab_meals, text="Питание")
        self.nb.add(self.tab_lodging, text="Проживание")

        # build tab skeletons
        self._build_profile_tab()
        self._build_work_tab()
        self._build_meals_tab()
        self._build_lodging_tab()

        # lazy loading per tab
        self.nb.bind("<<NotebookTabChanged>>", lambda e: self._load_current_tab_data())

    def _set_status(self, text: str):
        try:
            self.status.config(text=text or "")
        except Exception:
            pass

    def _clear_tree(self, tree: ttk.Treeview):
        tree.delete(*tree.get_children())

    def _search(self):
        q = (self.var_q.get() or "").strip()
        self._set_status("Поиск...")
        self.tree.delete(*self.tree.get_children())

        try:
            rows = search_employees(q, limit=200)
        except Exception as e:
            messagebox.showerror("Сотрудники", f"Ошибка поиска:\n{e}", parent=self)
            self._set_status("")
            return

        for r in rows:
            iid = str(r["id"])
            self.tree.insert(
                "",
                "end",
                iid=iid,
                values=(
                    r.get("fio", ""),
                    r.get("tbn", ""),
                    r.get("department", ""),
                    "Да" if r.get("is_fired") else "Нет",
                ),
            )

        self._set_status(f"Найдено: {len(rows)}")
        if rows:
            self.tree.selection_set(str(rows[0]["id"]))
            self._on_select_employee()

    def _on_select_employee(self):
        sel = self.tree.selection()
        if not sel:
            return
        try:
            emp_id = int(sel[0])
        except Exception:
            return

        try:
            prof = load_employee_profile(emp_id)
        except Exception as e:
            messagebox.showerror("Сотрудники", f"Ошибка загрузки профиля:\n{e}", parent=self)
            return

        if not prof:
            return

        self._selected_employee = prof
        fio = prof.get("fio") or ""
        tbn = prof.get("tbn") or ""
        self.lbl_title.config(text=f"{fio}  (Таб№ {tbn})")

        # fill profile immediately
        self._render_profile(prof)

        # invalidate previous background loads
        self._load_token += 1
        self._set_status("")

        # load data for active tab
        self._load_current_tab_data()

    # ---------- tabs: build ----------

    def _build_profile_tab(self):
        frm = ttk.Frame(self.tab_profile, padding=10)
        frm.pack(fill="both", expand=True)

        self._profile_labels: Dict[str, ttk.Label] = {}
        rows = [
            ("fio", "ФИО:"),
            ("tbn", "Табельный №:"),
            ("position", "Должность:"),
            ("department", "Подразделение:"),
            ("is_fired", "Уволен:"),
        ]
        for i, (key, title) in enumerate(rows):
            ttk.Label(frm, text=title).grid(row=i, column=0, sticky="e", padx=(0, 8), pady=4)
            lbl = ttk.Label(frm, text="")
            lbl.grid(row=i, column=1, sticky="w", pady=4)
            self._profile_labels[key] = lbl

        frm.grid_columnconfigure(1, weight=1)

    def _build_work_tab(self):
        root = ttk.Frame(self.tab_work, padding=10)
        root.pack(fill="both", expand=True)

        top = ttk.LabelFrame(root, text="Итоги по месяцам")
        top.pack(fill="both", expand=True, pady=(0, 8))

        cols = ("period", "days", "hours", "night", "ot_day", "ot_night")
        self.work_months = ttk.Treeview(top, columns=cols, show="headings", height=10)
        for c, t, w, a in [
            ("period", "Период", 100, "center"),
            ("days", "Дней", 70, "e"),
            ("hours", "Часы", 90, "e"),
            ("night", "Ночные", 90, "e"),
            ("ot_day", "Пер.день", 90, "e"),
            ("ot_night", "Пер.ночь", 90, "e"),
        ]:
            self.work_months.heading(c, text=t)
            self.work_months.column(c, width=w, anchor=a)
        vsb = ttk.Scrollbar(top, orient="vertical", command=self.work_months.yview)
        self.work_months.configure(yscrollcommand=vsb.set)
        self.work_months.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        bot = ttk.LabelFrame(root, text="ТОП объектов (по часам)")
        bot.pack(fill="both", expand=True)

        cols2 = ("address", "hours", "days", "months")
        self.work_objects = ttk.Treeview(bot, columns=cols2, show="headings", height=12)
        for c, t, w, a in [
            ("address", "Объект", 360, "w"),
            ("hours", "Часы", 90, "e"),
            ("days", "Дней", 70, "e"),
            ("months", "Месяцев", 80, "e"),
        ]:
            self.work_objects.heading(c, text=t)
            self.work_objects.column(c, width=w, anchor=a)
        vsb2 = ttk.Scrollbar(bot, orient="vertical", command=self.work_objects.yview)
        self.work_objects.configure(yscrollcommand=vsb2.set)
        self.work_objects.pack(side="left", fill="both", expand=True)
        vsb2.pack(side="right", fill="y")

    def _build_meals_tab(self):
        root = ttk.Frame(self.tab_meals, padding=10)
        root.pack(fill="both", expand=True)

        kpi = ttk.LabelFrame(root, text="Итоги")
        kpi.pack(fill="x", pady=(0, 8))
        self.lbl_meals_kpi = ttk.Label(kpi, text="—")
        self.lbl_meals_kpi.pack(anchor="w", padx=8, pady=6)

        top = ttk.LabelFrame(root, text="ТОП объектов питания")
        top.pack(fill="both", expand=True, pady=(0, 8))

        cols = ("address", "days", "rows", "qty")
        self.meals_objects = ttk.Treeview(top, columns=cols, show="headings", height=10)
        for c, t, w, a in [
            ("address", "Объект", 360, "w"),
            ("days", "Дней", 70, "e"),
            ("rows", "Записей", 80, "e"),
            ("qty", "Порций", 90, "e"),
        ]:
            self.meals_objects.heading(c, text=t)
            self.meals_objects.column(c, width=w, anchor=a)
        vsb = ttk.Scrollbar(top, orient="vertical", command=self.meals_objects.yview)
        self.meals_objects.configure(yscrollcommand=vsb.set)
        self.meals_objects.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        bot = ttk.LabelFrame(root, text="История (последние записи)")
        bot.pack(fill="both", expand=True)

        cols2 = ("date", "address", "meal_type", "qty", "team", "dep")
        self.meals_history = ttk.Treeview(bot, columns=cols2, show="headings", height=12)
        for c, t, w, a in [
            ("date", "Дата", 95, "center"),
            ("address", "Объект", 260, "w"),
            ("meal_type", "Тип", 140, "w"),
            ("qty", "Кол-во", 70, "e"),
            ("team", "Бригада", 160, "w"),
            ("dep", "Подразделение", 160, "w"),
        ]:
            self.meals_history.heading(c, text=t)
            self.meals_history.column(c, width=w, anchor=a)
        vsb2 = ttk.Scrollbar(bot, orient="vertical", command=self.meals_history.yview)
        self.meals_history.configure(yscrollcommand=vsb2.set)
        self.meals_history.pack(side="left", fill="both", expand=True)
        vsb2.pack(side="right", fill="y")

    def _build_lodging_tab(self):
        root = ttk.Frame(self.tab_lodging, padding=10)
        root.pack(fill="both", expand=True)

        cur = ttk.LabelFrame(root, text="Текущее проживание")
        cur.pack(fill="x", pady=(0, 8))
        self.lbl_lodging_current = ttk.Label(cur, text="—")
        self.lbl_lodging_current.pack(anchor="w", padx=8, pady=6)

        self.lbl_lodging_totals = ttk.Label(root, text="")
        self.lbl_lodging_totals.pack(anchor="w", pady=(0, 8))

        hist = ttk.LabelFrame(root, text="История проживаний")
        hist.pack(fill="both", expand=True, pady=(0, 8))

        cols = ("check_in", "check_out", "status", "dorm", "room")
        self.lodging_history = ttk.Treeview(hist, columns=cols, show="headings", height=10)
        for c, t, w, a in [
            ("check_in", "Заезд", 95, "center"),
            ("check_out", "Выезд", 95, "center"),
            ("status", "Статус", 90, "center"),
            ("dorm", "Общежитие", 240, "w"),
            ("room", "Комната", 90, "center"),
        ]:
            self.lodging_history.heading(c, text=t)
            self.lodging_history.column(c, width=w, anchor=a)
        vsb = ttk.Scrollbar(hist, orient="vertical", command=self.lodging_history.yview)
        self.lodging_history.configure(yscrollcommand=vsb.set)
        self.lodging_history.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        charges = ttk.LabelFrame(root, text="Начисления (dorm_charges)")
        charges.pack(fill="both", expand=True)

        cols2 = ("period", "days", "amount", "avg", "source", "stay_id")
        self.lodging_charges = ttk.Treeview(charges, columns=cols2, show="headings", height=10)
        for c, t, w, a in [
            ("period", "Период", 90, "center"),
            ("days", "Дней", 60, "e"),
            ("amount", "Сумма", 90, "e"),
            ("avg", "Средняя", 90, "e"),
            ("source", "Источник", 110, "center"),
            ("stay_id", "stay_id", 80, "e"),
        ]:
            self.lodging_charges.heading(c, text=t)
            self.lodging_charges.column(c, width=w, anchor=a)
        vsb2 = ttk.Scrollbar(charges, orient="vertical", command=self.lodging_charges.yview)
        self.lodging_charges.configure(yscrollcommand=vsb2.set)
        self.lodging_charges.pack(side="left", fill="both", expand=True)
        vsb2.pack(side="right", fill="y")

    # ---------- render helpers ----------

    def _render_profile(self, prof: Dict[str, Any]):
        self._profile_labels["fio"].config(text=str(prof.get("fio") or ""))
        self._profile_labels["tbn"].config(text=str(prof.get("tbn") or ""))
        self._profile_labels["position"].config(text=str(prof.get("position") or ""))
        self._profile_labels["department"].config(text=str(prof.get("department") or ""))
        self._profile_labels["is_fired"].config(text=("Да" if prof.get("is_fired") else "Нет"))

    # ---------- background loading ----------

    def _load_current_tab_data(self):
        emp = self._selected_employee
        if not emp:
            return

        tab = self.nb.select()
        tab_name = self.nb.tab(tab, "text")

        if tab_name.startswith("Работа"):
            self._load_work(emp)
        elif tab_name.startswith("Питание"):
            self._load_meals(emp)
        elif tab_name.startswith("Проживание"):
            self._load_lodging(emp)

    def _run_bg(self, label: str, func, on_ok):
        token = self._load_token

        def worker():
            try:
                data = func()
            except Exception as e:
                data = e
            def deliver():
                if token != self._load_token:
                    return
                if isinstance(data, Exception):
                    messagebox.showerror("Сотрудники", f"{label}:\n{data}", parent=self)
                    self._set_status("")
                    return
                on_ok(data)
                self._set_status("")
            self.after(0, deliver)

        self._set_status(label)
        threading.Thread(target=worker, daemon=True).start()

    def _load_work(self, emp: Dict[str, Any]):
        tbn = (emp.get("tbn") or "").strip()
        if not tbn:
            return

        def load():
            months = load_timesheet_summary_by_month(tbn)
            objs = load_timesheet_by_object(tbn, limit=30)
            return {"months": months, "objs": objs}

        def render(payload):
            self._clear_tree(self.work_months)
            for r in payload["months"]:
                period = f"{int(r['year']):04d}-{int(r['month']):02d}"
                self.work_months.insert(
                    "", "end",
                    values=(
                        period,
                        int(r.get("days", 0) or 0),
                        f"{float(r.get('hours', 0) or 0):.2f}",
                        f"{float(r.get('night_hours', 0) or 0):.2f}",
                        f"{float(r.get('ot_day', 0) or 0):.2f}",
                        f"{float(r.get('ot_night', 0) or 0):.2f}",
                    )
                )

            self._clear_tree(self.work_objects)
            for r in payload["objs"]:
                self.work_objects.insert(
                    "", "end",
                    values=(
                        r.get("address") or "—",
                        f"{float(r.get('hours', 0) or 0):.2f}",
                        int(r.get("days", 0) or 0),
                        int(r.get("months_cnt", 0) or 0),
                    )
                )

        self._run_bg("Загрузка табелей…", load, render)

    def _load_meals(self, emp: Dict[str, Any]):
        tbn = (emp.get("tbn") or "").strip()
        if not tbn:
            return

        def load():
            kpi = load_meals_kpi(tbn)
            top = load_meals_by_object(tbn, limit=30)
            hist = load_meals_history(tbn, limit=1000)
            return {"kpi": kpi, "top": top, "hist": hist}

        def render(payload):
            k = payload["kpi"]
            self.lbl_meals_kpi.config(
                text=f"Дней питания: {k['days_cnt']} | Записей: {k['rows_cnt']} | Порций (SUM quantity): {k['qty_sum']:.2f}"
            )

            self._clear_tree(self.meals_objects)
            for r in payload["top"]:
                self.meals_objects.insert(
                    "", "end",
                    values=(
                        r.get("address") or "—",
                        int(r.get("days_cnt", 0) or 0),
                        int(r.get("rows_cnt", 0) or 0),
                        f"{float(r.get('qty_sum', 0) or 0):.2f}",
                    )
                )

            self._clear_tree(self.meals_history)
            for r in payload["hist"]:
                dt = r.get("date")
                dt_s = dt.strftime("%Y-%m-%d") if hasattr(dt, "strftime") else str(dt or "")
                self.meals_history.insert(
                    "", "end",
                    values=(
                        dt_s,
                        r.get("address") or "—",
                        r.get("meal_type") or "",
                        f"{float(r.get('qty', 0) or 0):.2f}",
                        r.get("team_name") or "",
                        r.get("department") or "",
                    )
                )

        self._run_bg("Загрузка питания…", load, render)

    def _load_lodging(self, emp: Dict[str, Any]):
        emp_id = int(emp["id"])

        def load():
            cur = load_lodging_current(emp_id)
            hist = load_lodging_history(emp_id)
            bed_days = load_lodging_bed_days_total(emp_id)
            charges = load_lodging_charges(emp_id)
            return {"cur": cur, "hist": hist, "bed_days": bed_days, "charges": charges}

        def render(payload):
            cur = payload["cur"]
            if not cur:
                self.lbl_lodging_current.config(text="Нет активного проживания.")
            else:
                ci = cur.get("check_in")
                co = cur.get("check_out")
                ci_s = ci.strftime("%Y-%m-%d") if hasattr(ci, "strftime") else str(ci or "")
                co_s = co.strftime("%Y-%m-%d") if hasattr(co, "strftime") else str(co or "")
                self.lbl_lodging_current.config(
                    text=f"{cur.get('dorm_name','')} | комната {cur.get('room_no','')} | "
                         f"заезд: {ci_s} | выезд: {co_s or '—'}"
                )

            self.lbl_lodging_totals.config(text=f"Койко-дней всего (оценка): {int(payload['bed_days'])}")

            self._clear_tree(self.lodging_history)
            for r in payload["hist"]:
                ci = r.get("check_in")
                co = r.get("check_out")
                ci_s = ci.strftime("%Y-%m-%d") if hasattr(ci, "strftime") else str(ci or "")
                co_s = co.strftime("%Y-%m-%d") if hasattr(co, "strftime") else str(co or "")
                self.lodging_history.insert(
                    "", "end",
                    values=(
                        ci_s,
                        co_s,
                        r.get("status") or "",
                        r.get("dorm_name") or "",
                        r.get("room_no") or "",
                    )
                )

            self._clear_tree(self.lodging_charges)
            for r in payload["charges"]:
                period = f"{int(r.get('year') or 0):04d}-{int(r.get('month') or 0):02d}"
                self.lodging_charges.insert(
                    "", "end",
                    values=(
                        period,
                        int(r.get("days", 0) or 0),
                        f"{float(r.get('amount', 0) or 0):.2f}",
                        f"{float(r.get('avg_price_per_day', 0) or 0):.2f}" if r.get("avg_price_per_day") is not None else "",
                        r.get("rate_source") or "",
                        int(r.get("stay_id", 0) or 0),
                    )
                )

        self._run_bg("Загрузка проживания…", load, render)

    def _export_dossier(self):
        emp = self._selected_employee
        if not emp:
            messagebox.showinfo("Экспорт", "Сначала выберите сотрудника.", parent=self)
            return

        fio = (emp.get("fio") or "").strip() or "Сотрудник"
        tbn = (emp.get("tbn") or "").strip() or "—"
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"Досье_{tbn}_{fio}_{ts}.xlsx".replace("/", "_").replace("\\", "_")

        path = filedialog.asksaveasfilename(
            parent=self,
            title="Сохранить досье сотрудника",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel", "*.xlsx"), ("Все файлы", "*.*")],
        )
        if not path:
            return

        # грузим все данные одним пакетом (сейчас “всё показываем всегда”)
        self._run_bg(
            "Формирование досье…",
            func=lambda: self._collect_dossier_data(emp),
            on_ok=lambda data: self._export_dossier_write_xlsx(path, data),
        )

    def _collect_dossier_data(self, emp: Dict[str, Any]) -> Dict[str, Any]:
        """
        Собирает полный набор данных для досье.
        Вызывается в фоне.
        """
        emp_id = int(emp["id"])
        tbn = (emp.get("tbn") or "").strip()

        profile = load_employee_profile(emp_id) or emp

        work_months = load_timesheet_summary_by_month(tbn)
        work_objects = load_timesheet_by_object(tbn, limit=9999)

        meals_kpi = load_meals_kpi(tbn)
        meals_objects = load_meals_by_object(tbn, limit=9999)
        meals_history = load_meals_history(tbn, limit=20000)

        lodging_current = load_lodging_current(emp_id)
        lodging_history = load_lodging_history(emp_id)
        lodging_bed_days = load_lodging_bed_days_total(emp_id)
        lodging_charges = load_lodging_charges(emp_id)

        return {
            "profile": profile,
            "work_months": work_months,
            "work_objects": work_objects,
            "meals_kpi": meals_kpi,
            "meals_objects": meals_objects,
            "meals_history": meals_history,
            "lodging_current": lodging_current,
            "lodging_history": lodging_history,
            "lodging_bed_days_total": lodging_bed_days,
            "lodging_charges": lodging_charges,
        }

    def _export_dossier_write_xlsx(self, path: str, data: Dict[str, Any]):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except Exception as e:
            messagebox.showerror("Экспорт", f"Не удалось импортировать openpyxl:\n{e}", parent=self)
            return
    
        prof = data.get("profile") or {}
        fio = (prof.get("fio") or "").strip()
        tbn = (prof.get("tbn") or "").strip()
    
        wb = Workbook()
        ws = wb.active
        ws.title = "Досье"
    
        # --- styles
        title_font = Font(size=16, bold=True)
        section_font = Font(size=12, bold=True)
        header_font = Font(bold=True)
        gray_fill = PatternFill("solid", fgColor="F2F2F2")
        blue_fill = PatternFill("solid", fgColor="DDEBFF")
        thin = Side(style="thin", color="BFBFBF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
        def write_section_title(row: int, text: str, col_from: int = 1, col_to: int = 8) -> int:
            ws.cell(row=row, column=col_from, value=text).font = section_font
            ws.cell(row=row, column=col_from).fill = blue_fill
            ws.merge_cells(start_row=row, start_column=col_from, end_row=row, end_column=col_to)
            return row + 1
    
        def write_kv(row: int, key: str, value: Any, key_col: int = 1, val_col: int = 2) -> int:
            ws.cell(row=row, column=key_col, value=key).font = header_font
            ws.cell(row=row, column=val_col, value=value if value is not None else "")
            return row + 1
    
        def write_table(row: int, headers: List[str], rows: List[List[Any]], col_from: int = 1) -> int:
            # header
            for j, h in enumerate(headers):
                c = ws.cell(row=row, column=col_from + j, value=h)
                c.font = header_font
                c.fill = gray_fill
                c.border = border
            row += 1
    
            # body
            for r in rows:
                for j, v in enumerate(r):
                    c = ws.cell(row=row, column=col_from + j, value=v)
                    c.border = border
                row += 1
    
            return row
    
        def autosize(max_col: int, min_w: int = 10, max_w: int = 60):
            for col in range(1, max_col + 1):
                max_len = 0
                for r in range(1, ws.max_row + 1):
                    v = ws.cell(r, col).value
                    if v is None:
                        continue
                    s = str(v)
                    if len(s) > max_len:
                        max_len = len(s)
                ws.column_dimensions[get_column_letter(col)].width = min(max_w, max(min_w, max_len + 2))
    
        # ========== TOP TITLE ==========
        ws["A1"] = "Досье сотрудника"
        ws["A1"].font = title_font
        ws["A2"] = f"Сформировано: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
        row = 4
    
        # ========== PROFILE ==========
        row = write_section_title(row, "Профиль", 1, 8)
        row = write_kv(row, "ФИО", fio)
        row = write_kv(row, "Табельный №", tbn)
        row = write_kv(row, "Должность", prof.get("position") or "")
        row = write_kv(row, "Подразделение", prof.get("department") or "")
        row = write_kv(row, "Уволен", "Да" if prof.get("is_fired") else "Нет")
        row += 1
    
        # ========== WORK (timesheets) ==========
        row = write_section_title(row, "Работа (табели) — итоги по месяцам", 1, 8)
        months_rows: List[List[Any]] = []
        for r in (data.get("work_months") or []):
            period = f"{int(r['year']):04d}-{int(r['month']):02d}"
            months_rows.append([
                period,
                int(r.get("days", 0) or 0),
                float(r.get("hours", 0) or 0),
                float(r.get("night_hours", 0) or 0),
                float(r.get("ot_day", 0) or 0),
                float(r.get("ot_night", 0) or 0),
            ])
        row = write_table(
            row,
            headers=["Период", "Дней", "Часы", "Ночные", "Пер.день", "Пер.ночь"],
            rows=months_rows,
            col_from=1,
        )
        row += 2
    
        row = write_section_title(row, "Работа (табели) — ТОП объектов", 1, 8)
        obj_rows: List[List[Any]] = []
        for r in (data.get("work_objects") or []):
            obj_rows.append([
                r.get("address") or "—",
                float(r.get("hours", 0) or 0),
                int(r.get("days", 0) or 0),
                int(r.get("months_cnt", 0) or 0),
            ])
        row = write_table(
            row,
            headers=["Объект", "Часы", "Дней", "Месяцев"],
            rows=obj_rows,
            col_from=1,
        )
        row += 2
    
        # ========== MEALS ==========
        row = write_section_title(row, "Питание — итоги", 1, 8)
        k = data.get("meals_kpi") or {}
        row = write_kv(row, "Дней питания (distinct date)", int(k.get("days_cnt", 0) or 0))
        row = write_kv(row, "Записей (rows)", int(k.get("rows_cnt", 0) or 0))
        row = write_kv(row, "Порций (SUM quantity)", float(k.get("qty_sum", 0) or 0))
        row += 1
    
        row = write_section_title(row, "Питание — ТОП объектов", 1, 8)
        meals_obj_rows: List[List[Any]] = []
        for r in (data.get("meals_objects") or []):
            meals_obj_rows.append([
                r.get("address") or "—",
                int(r.get("days_cnt", 0) or 0),
                int(r.get("rows_cnt", 0) or 0),
                float(r.get("qty_sum", 0) or 0),
            ])
        row = write_table(
            row,
            headers=["Объект", "Дней", "Записей", "Порций"],
            rows=meals_obj_rows,
            col_from=1,
        )
        row += 2
    
        # История питания — лучше ограничить, иначе “досье” будет огромным
        meals_hist = data.get("meals_history") or []
        max_hist = 1500  # можно настроить
        meals_hist = meals_hist[:max_hist]
    
        row = write_section_title(row, f"Питание — история (последние {len(meals_hist)})", 1, 8)
        hist_rows: List[List[Any]] = []
        for r in meals_hist:
            dt = r.get("date")
            dt_s = dt.strftime("%Y-%m-%d") if hasattr(dt, "strftime") else str(dt or "")
            hist_rows.append([
                dt_s,
                r.get("address") or "—",
                r.get("meal_type") or "",
                float(r.get("qty", 0) or 0),
                r.get("team_name") or "",
                r.get("department") or "",
            ])
        row = write_table(
            row,
            headers=["Дата", "Объект", "Тип питания", "Кол-во", "Бригада", "Подразделение"],
            rows=hist_rows,
            col_from=1,
        )
        row += 2
    
        # ========== LODGING ==========
        row = write_section_title(row, "Проживание — текущее", 1, 8)
        cur = data.get("lodging_current")
        if not cur:
            row = write_kv(row, "Статус", "Нет активного проживания")
        else:
            ci = cur.get("check_in")
            co = cur.get("check_out")
            ci_s = ci.strftime("%Y-%m-%d") if hasattr(ci, "strftime") else str(ci or "")
            co_s = co.strftime("%Y-%m-%d") if hasattr(co, "strftime") else str(co or "")
            row = write_kv(row, "Общежитие", cur.get("dorm_name") or "")
            row = write_kv(row, "Комната", cur.get("room_no") or "")
            row = write_kv(row, "Заезд", ci_s)
            row = write_kv(row, "Выезд", co_s or "—")
        row += 1
    
        row = write_kv(row, "Койко-дней всего (оценка)", int(data.get("lodging_bed_days_total") or 0))
        row += 1
    
        row = write_section_title(row, "Проживание — история", 1, 8)
        lod_hist_rows: List[List[Any]] = []
        for r in (data.get("lodging_history") or []):
            ci = r.get("check_in")
            co = r.get("check_out")
            ci_s = ci.strftime("%Y-%m-%d") if hasattr(ci, "strftime") else str(ci or "")
            co_s = co.strftime("%Y-%m-%d") if hasattr(co, "strftime") else str(co or "")
            lod_hist_rows.append([ci_s, co_s, r.get("status") or "", r.get("dorm_name") or "", r.get("room_no") or ""])
        row = write_table(
            row,
            headers=["Заезд", "Выезд", "Статус", "Общежитие", "Комната"],
            rows=lod_hist_rows,
            col_from=1,
        )
        row += 2
    
        row = write_section_title(row, "Проживание — начисления (dorm_charges)", 1, 8)
        ch_rows: List[List[Any]] = []
        for r in (data.get("lodging_charges") or []):
            period = f"{int(r.get('year') or 0):04d}-{int(r.get('month') or 0):02d}"
            ch_rows.append([
                period,
                int(r.get("days", 0) or 0),
                float(r.get("amount", 0) or 0),
                float(r.get("avg_price_per_day", 0) or 0) if r.get("avg_price_per_day") is not None else "",
                r.get("rate_source") or "",
                int(r.get("stay_id", 0) or 0),
            ])
        row = write_table(
            row,
            headers=["Период", "Дней", "Сумма", "Средняя", "Источник", "stay_id"],
            rows=ch_rows,
            col_from=1,
        )
    
        # оформление
        ws.freeze_panes = "A4"
        autosize(max_col=8)
    
        try:
            wb.save(path)
        except Exception as e:
            messagebox.showerror("Экспорт", f"Не удалось сохранить файл:\n{e}", parent=self)
            return
    
        messagebox.showinfo("Экспорт", f"Досье сформировано (1 лист):\n{path}", parent=self)

def create_employee_card_page(parent, app_ref=None) -> tk.Frame:
    try:
        return EmployeeCardPage(parent, app_ref=app_ref)
    except Exception as e:
        messagebox.showerror("Сотрудники", f"Не удалось открыть карточку сотрудника:\n{e}", parent=parent)
        return tk.Frame(parent)
