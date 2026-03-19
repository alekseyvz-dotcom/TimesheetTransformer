import os
import re
import sys
import json
import urllib.request
import urllib.error
import urllib.parse
from pathlib import Path
from typing import List, Tuple, Optional, Dict, Any
import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, date, timedelta
import psycopg2
from psycopg2 import pool
from psycopg2.extras import RealDictCursor
from urllib.parse import urlparse, parse_qs

db_connection_pool = None

def set_db_pool(pool):
    global db_connection_pool
    db_connection_pool = pool

def release_db_connection(conn):
    if db_connection_pool:
        db_connection_pool.putconn(conn)

# ========================= БАЗОВЫЕ КОНСТАНТЫ =========================

APP_TITLE = "Заказ питания"

CONFIG_FILE = "tabel_config.ini"
CONFIG_SECTION_UI = "UI"
CONFIG_SECTION_INTEGR = "Integrations"

KEY_SELECTED_DEP = "selected_department"
KEY_MEALS_PLANNING_ENABLED = "meals_planning_enabled"

SPRAVOCHNIK_FILE = "Справочник.xlsx"  # оставлен для совместимости, больше не используется
ORDERS_DIR = "Заявки_питание"

def exe_dir() -> Path:
    """Каталог, откуда запущена программа/скрипт."""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def config_path() -> Path:
    """Путь к ini‑конфигу для fallback‑режима (без settings_manager)."""
    return exe_dir() / CONFIG_FILE

try:
    import settings_manager as Settings
except Exception:
    Settings = None

if Settings:
    ensure_config = Settings.ensure_config
    read_config = Settings.read_config
    write_config = Settings.write_config

    def get_saved_dep() -> str:
        return Settings.get_selected_department_from_config()

    def set_saved_dep(dep: str):
        return Settings.set_selected_department_from_config(dep)

else:
    def ensure_config():
        cp = config_path()
        if cp.exists():
            cfg = configparser.ConfigParser()
            cfg.read(cp, encoding="utf-8")
            changed = False

            if not cfg.has_section(CONFIG_SECTION_UI):
                cfg[CONFIG_SECTION_UI] = {}
                changed = True
            if KEY_SELECTED_DEP not in cfg[CONFIG_SECTION_UI]:
                cfg[CONFIG_SECTION_UI][KEY_SELECTED_DEP] = "Все"
                changed = True

            if not cfg.has_section(CONFIG_SECTION_INTEGR):
                cfg[CONFIG_SECTION_INTEGR] = {}
                changed = True
            if KEY_MEALS_PLANNING_ENABLED not in cfg[CONFIG_SECTION_INTEGR]:
                cfg[CONFIG_SECTION_INTEGR][KEY_MEALS_PLANNING_ENABLED] = "true"
                changed = True

            if changed:
                with open(cp, "w", encoding="utf-8") as f:
                    cfg.write(f)
            return

        cfg = configparser.ConfigParser()
        cfg[CONFIG_SECTION_UI] = {KEY_SELECTED_DEP: "Все"}
        cfg[CONFIG_SECTION_INTEGR] = {
            KEY_MEALS_PLANNING_ENABLED: "true",
        }
        with open(cp, "w", encoding="utf-8") as f:
            cfg.write(f)

    import configparser

    def read_config() -> configparser.ConfigParser:
        ensure_config()
        cfg = configparser.ConfigParser()
        cfg.read(config_path(), encoding="utf-8")
        return cfg

    def write_config(cfg: configparser.ConfigParser):
        with open(config_path(), "w", encoding="utf-8") as f:
            cfg.write(f)

    def get_saved_dep() -> str:
        cfg = read_config()
        return cfg.get(CONFIG_SECTION_UI, KEY_SELECTED_DEP, fallback="Все")

    def set_saved_dep(dep: str):
        cfg = read_config()
        if not cfg.has_section(CONFIG_SECTION_UI):
            cfg[CONFIG_SECTION_UI] = {}
        cfg[CONFIG_SECTION_UI][KEY_SELECTED_DEP] = dep or "Все"
        write_config(cfg)

def get_meals_planning_enabled() -> bool:
    if Settings and hasattr(Settings, "get_meals_planning_enabled_from_config"):
        return Settings.get_meals_planning_enabled_from_config()
    cfg = read_config()
    v = cfg.get(CONFIG_SECTION_INTEGR, KEY_MEALS_PLANNING_ENABLED, fallback="true").strip().lower()
    return v in ("1", "true", "yes", "on")

def get_db_connection():
    """Получает соединение из установленного пула."""
    if db_connection_pool is None:
         raise RuntimeError("Пул соединений не был установлен из главного приложения.")
    return db_connection_pool.getconn()

def get_or_create_department(cur, name: str):
    if not name:
        return None
    cur.execute("SELECT id FROM departments WHERE name = %s", (name,))
    row = cur.fetchone()
    if row:
        return row[0]
    cur.execute("INSERT INTO departments (name) VALUES (%s) RETURNING id", (name,))
    return cur.fetchone()[0]

def get_setting(key: str, default: str = "") -> str:
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("SELECT value FROM app_settings WHERE key = %s", (key,))
            row = cur.fetchone()
            if row:
                return row[0]
            return default
    finally:
        if conn:
            release_db_connection(conn)

def set_setting(key: str, value: str):
    conn = None
    try:
        conn = get_db_connection()
        with conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO app_settings (key, value)
                    VALUES (%s, %s)
                    ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value
                    """,
                    (key, value),
                )
    finally:
        if conn:
            release_db_connection(conn)

def get_meals_limit_next_day_only() -> bool:

    v = (get_setting("meals_limit_next_day_only", "1") or "").strip().lower()
    return v in ("1", "true", "yes", "on")

def get_meals_next_day_deadline() -> Optional[datetime.time]:

    s = (get_setting("meals_next_day_deadline", "").strip())
    if not s:
        return None
    try:

        for fmt in ("%H:%M", "%H:%M:%S"):
            try:
                return datetime.strptime(s, fmt).time()
            except Exception:
                pass
    except Exception:
        pass
    return None

def get_object_id(cur, excel_id: str, address: str) -> Optional[int]:
    excel_id = (excel_id or "").strip()
    address = (address or "").strip()

    if excel_id and address:
        cur.execute(
            """
            SELECT id
            FROM objects
            WHERE excel_id = %s
              AND address = %s
            LIMIT 1
            """,
            (excel_id, address),
        )
        row = cur.fetchone()
        if row:
            return row[0]
        return None

    if excel_id:
        cur.execute(
            "SELECT id FROM objects WHERE excel_id = %s LIMIT 1",
            (excel_id,),
        )
        row = cur.fetchone()
        if row:
            return row[0]

    if address:
        cur.execute(
            "SELECT id FROM objects WHERE address = %s LIMIT 1",
            (address,),
        )
        row = cur.fetchone()
        if row:
            return row[0]

    return None

def get_or_create_object_by_excel_id(cur, excel_id: str, address: str) -> int:
    raise RuntimeError("Создание объектов из модуля питания запрещено.")

def get_or_create_meal_type(cur, name: str):
    name = (name or "").strip()
    if not name:
        return None
    cur.execute("SELECT id FROM meal_types WHERE name = %s", (name,))
    row = cur.fetchone()
    if row:
        return row[0]
    cur.execute("INSERT INTO meal_types (name, price) VALUES (%s, 0) RETURNING id", (name,))
    return cur.fetchone()[0]

def find_employee(cur, fio: str, tbn: str = None):
    fio = (fio or "").strip()
    tbn = (tbn or "").strip()
    if tbn:
        cur.execute("SELECT id FROM employees WHERE tbn = %s", (tbn,))
        row = cur.fetchone()
        if row:
            return row[0]
    if fio:
        cur.execute("SELECT id FROM employees WHERE fio = %s", (fio,))
        row = cur.fetchone()
        if row:
            return row[0]
    return None

# ---------- Загрузка справочников из БД ----------

def load_employees_from_db() -> List[Dict[str, Any]]:
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT e.fio, e.tbn, e.position, d.name AS dep
                  FROM employees e
                  LEFT JOIN departments d ON d.id = e.department_id
                 WHERE COALESCE(e.is_fired, FALSE) = FALSE
              ORDER BY e.fio
                """
            )
            res = []
            for fio, tbn, pos, dep in cur.fetchall():
                res.append({
                    "fio": fio or "",
                    "tbn": tbn or "",
                    "pos": pos or "",
                    "dep": dep or "",
                })
            return res
    finally:
        if conn:
            release_db_connection(conn)

def load_objects_from_db() -> List[Tuple[str, str]]:
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT
                    COALESCE(NULLIF(excel_id, ''), '') AS code,
                    address
                  FROM objects
                 ORDER BY address
                """
            )
            res = []
            for code, addr in cur.fetchall():
                res.append((code or "", addr or ""))
            return res
    finally:
        if conn:
            release_db_connection(conn)

def load_meal_types_from_db() -> List[Dict[str, Any]]:
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT id, name, COALESCE(price, 0) AS price
                  FROM meal_types
              ORDER BY id
                """
            )
            rows = cur.fetchall()
            if not rows:
                # создаём три стандартных типа, если таблица пустая
                defaults = [("Одноразовое", 0), ("Двухразовое", 0), ("Трехразовое", 0)]
                for name, price in defaults:
                    cur.execute(
                        "INSERT INTO meal_types (name, price) VALUES (%s, %s) RETURNING id, name, price",
                        (name, price),
                    )
                    rows.append(cur.fetchone())
                conn.commit()
            return rows
    finally:
        if conn:
            release_db_connection(conn)

def get_meal_type_price_map() -> Dict[str, float]:
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute(
                "SELECT name, COALESCE(price, 0) FROM meal_types"
            )
            return {
                (name or "").strip(): float(price or 0)
                for name, price in cur.fetchall()
            }
    finally:
        if conn:
            release_db_connection(conn)

# ---------------- Сохранение заказов, реестры, проверки ----------------

def get_current_user_id() -> Optional[int]:
    return None

def save_order_to_db(data: dict) -> int:
    conn = None
    try:
        conn = get_db_connection()
        with conn:
            with conn.cursor() as cur:
                dept_name = (data.get("department") or "").strip()
                dept_id = get_or_create_department(cur, dept_name) if dept_name else None

                obj = data.get("object") or {}
                obj_excel_id = (obj.get("id") or "").strip()
                obj_address = (obj.get("address") or "").strip()

                object_id = get_object_id(cur, obj_excel_id, obj_address)
                if object_id is None:
                    raise ValueError(
                        "Объект не найден в реестре. "
                        "Выберите объект из списка, а не вводите адрес вручную."
                    )

                created_at = datetime.strptime(data["created_at"], "%Y-%m-%dT%H:%M:%S")
                order_date = datetime.strptime(data["date"], "%Y-%m-%d").date()
                team_name = (data.get("team_name") or "").strip()
                fact_address = (data.get("fact_address") or "").strip()
                user_id = data.get("user_id")

                cur.execute(
                    """
                    INSERT INTO meal_orders (created_at, date, department_id, team_name, object_id, user_id, fact_address)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                    """,
                    (created_at, order_date, dept_id, team_name, object_id, user_id, fact_address),
                )
                order_id = cur.fetchone()[0]

                for emp in data.get("employees", []):
                    fio = (emp.get("fio") or "").strip()
                    tbn = (emp.get("tbn") or "").strip()
                    position = (emp.get("position") or "").strip()
                    meal_type_name = (emp.get("meal_type") or "").strip()
                    comment = (emp.get("comment") or "").strip()
                    quantity = float(emp.get("quantity") or 1)

                    meal_type_id = get_or_create_meal_type(cur, meal_type_name)
                    employee_id = find_employee(cur, fio, tbn)

                    # перезапись по этому же объекту/дате/сотруднику
                    if tbn:
                        cur.execute(
                            """
                            DELETE FROM meal_order_items moi
                            WHERE EXISTS (
                                SELECT 1
                                FROM meal_orders mo
                                LEFT JOIN employees e ON e.id = moi.employee_id
                                WHERE moi.order_id = mo.id
                                  AND mo.date = %s
                                  AND mo.object_id = %s
                                  AND (
                                       moi.tbn_text = %s
                                       OR (e.tbn IS NOT NULL AND e.tbn = %s)
                                  )
                            )
                            """,
                            (order_date, object_id, tbn, tbn),
                        )
                    else:
                        cur.execute(
                            """
                            DELETE FROM meal_order_items moi
                            WHERE EXISTS (
                                SELECT 1
                                FROM meal_orders mo
                                LEFT JOIN employees e ON e.id = moi.employee_id
                                WHERE moi.order_id = mo.id
                                  AND mo.date = %s
                                  AND mo.object_id = %s
                                  AND (
                                       moi.fio_text = %s
                                       OR (e.fio IS NOT NULL AND e.fio = %s)
                                  )
                            )
                            """,
                            (order_date, object_id, fio, fio),
                        )

                    cur.execute(
                        """
                        INSERT INTO meal_order_items
                        (order_id, employee_id, fio_text, tbn_text, position_text,
                         meal_type_id, meal_type_text, comment, quantity)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """,
                        (
                            order_id,
                            employee_id,
                            fio,
                            tbn,
                            position,
                            meal_type_id,
                            meal_type_name,
                            comment,
                            quantity,
                        ),
                    )

        return order_id
    finally:
        if conn:
            release_db_connection(conn)

def get_registry_from_db(
    filter_date: Optional[str] = None,
    filter_address: Optional[str] = None,
    filter_department: Optional[str] = None,
) -> List[Dict[str, Any]]:
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            params = []
            where_clauses = []

            if filter_date:
                where_clauses.append("mo.date = %s")
                params.append(filter_date)

            if filter_address:
                where_clauses.append("COALESCE(mo.fact_address, o.address) ILIKE %s")
                params.append(f"%{filter_address}%")

            if filter_department and filter_department.lower() != "все":
                where_clauses.append("d.name = %s")
                params.append(filter_department)

            where_sql = ""
            if where_clauses:
                where_sql = "WHERE " + " AND ".join(where_clauses)

            sql = f"""
                SELECT
                    mo.id                    AS order_id,
                    mo.date::text            AS date,
                    COALESCE(mo.fact_address, o.address, '')  AS address,
                    COALESCE(d.name, '')     AS department,
                    COALESCE(mo.team_name, '') AS team_name,
                    COALESCE(mti.name, moi.meal_type_text, '') AS meal_type
                FROM meal_orders mo
                JOIN meal_order_items moi ON moi.order_id = mo.id
                LEFT JOIN objects o       ON o.id = mo.object_id
                LEFT JOIN departments d   ON d.id = mo.department_id
                LEFT JOIN meal_types mti  ON mti.id = moi.meal_type_id
                {where_sql}
            """
            cur.execute(sql, params)
            rows = cur.fetchall()

        result: Dict[Tuple[str, str, str, str], Dict[str, Any]] = {}

        for order_id, date_str, address, dept, team_name, meal_type in rows:
            dep_name = dept or "Без подразделения"
            team_name = team_name or ""
            key = (date_str, address, dep_name, team_name)

            rec = result.setdefault(
                key,
                {
                    "date": date_str,
                    "address": address,
                    "department": dep_name,
                    "team_name": team_name,
                    "total_count": 0,
                    "by_meal_type": {},
                    "order_ids": set(),
                },
            )
            rec["total_count"] += 1
            rec["order_ids"].add(order_id)

            mt = meal_type or "Не указан"
            by_mt = rec["by_meal_type"]
            by_mt[mt] = by_mt.get(mt, 0) + 1

        return [
            {
                **rec,
                "order_ids": list(rec["order_ids"]),
            }
            for rec in result.values()
        ]

    finally:
        if conn:
            release_db_connection(conn)

def load_all_meal_orders(
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    department: Optional[str] = None,
    address_substr: Optional[str] = None,
) -> List[Dict[str, Any]]:
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            where_clauses = []
            params: List[Any] = []

            if date_from:
                where_clauses.append("mo.date >= %s")
                params.append(date_from)
            if date_to:
                where_clauses.append("mo.date <= %s")
                params.append(date_to)

            if department and department.strip() and department.strip().lower() != "все":
                where_clauses.append("d.name = %s")
                params.append(department.strip())

            if address_substr and address_substr.strip():
                where_clauses.append("COALESCE(mo.fact_address, o.address) ILIKE %s")
                params.append(f"%{address_substr.strip()}%")

            where_sql = ""
            if where_clauses:
                where_sql = "WHERE " + " AND ".join(where_clauses)

            sql = f"""
                SELECT
                    mo.id,
                    mo.date,
                    mo.created_at,
                    COALESCE(d.name, '')        AS department,
                    COALESCE(mo.team_name, '')  AS team_name,
                    COALESCE(o.excel_id, '')    AS object_id,
                    COALESCE(mo.fact_address, o.address, '')     AS object_address,
                    COUNT(moi.id)               AS employees_count,
                    COALESCE(au.full_name,
                             au.username,
                             '')                 AS user_name
                FROM meal_orders mo
                JOIN meal_order_items moi ON moi.order_id = mo.id
                LEFT JOIN departments d    ON d.id = mo.department_id
                LEFT JOIN objects o        ON o.id = mo.object_id
                LEFT JOIN app_users au     ON au.id = mo.user_id
                {where_sql}
                GROUP BY
                    mo.id,
                    mo.date,
                    mo.created_at,
                    d.name,
                    mo.team_name,
                    o.excel_id,
                    o.address,
                    au.full_name,
                    au.username
                ORDER BY mo.date DESC, mo.id DESC
            """
            cur.execute(sql, params)
            rows = cur.fetchall()
            return [dict(r) for r in rows]
    finally:
        if conn:
            release_db_connection(conn)

def load_user_meal_orders(user_id: int) -> List[Dict[str, Any]]:
    if not user_id:
        return []

    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT
                    mo.id,
                    mo.date,
                    mo.created_at,
                    COALESCE(d.name, '')      AS department,
                    COALESCE(mo.team_name, '') AS team_name,
                    COALESCE(o.excel_id, '')  AS object_id,
                    COALESCE(mo.fact_address, o.address, '')   AS object_address,
                    COUNT(moi.id)             AS employees_count
                FROM meal_orders mo
                JOIN meal_order_items moi ON moi.order_id = mo.id
                LEFT JOIN departments d    ON d.id = mo.department_id
                LEFT JOIN objects o        ON o.id = mo.object_id
                WHERE mo.user_id = %s
                GROUP BY
                    mo.id,
                    mo.date,
                    mo.created_at,
                    d.name,
                    mo.team_name,
                    o.excel_id,
                    o.address
                ORDER BY mo.date DESC, mo.id DESC
                """,
                (user_id,),
            )
            rows = cur.fetchall()
            return [dict(r) for r in rows]
    finally:
        if conn:
            release_db_connection(conn)

def get_details_from_db(
    filter_date: Optional[str] = None,
    filter_address: Optional[str] = None,
    filter_department: Optional[str] = None,
) -> List[Dict[str, Any]]:
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            params = []
            where_clauses = []

            if filter_date:
                where_clauses.append("mo.date = %s")
                params.append(filter_date)

            if filter_address:
                where_clauses.append("COALESCE(mo.fact_address, o.address) ILIKE %s")
                params.append(f"%{filter_address}%")

            if filter_department and filter_department.lower() != "все":
                where_clauses.append("d.name = %s")
                params.append(filter_department)

            where_sql = ""
            if where_clauses:
                where_sql = "WHERE " + " AND ".join(where_clauses)

            sql = f"""
                SELECT
                    mo.date::text        AS date,
                    COALESCE(mo.fact_address, o.address, '')       AS address,
                    COALESCE(o.excel_id, '')        AS object_excel_id,
            
                    -- Подразделение из заявки
                    COALESCE(d_order.name, '')     AS department,
            
                    -- Подразделение сотрудника из справочника (1С)
                    COALESCE(d_emp.name, '')       AS employee_department,
            
                    COALESCE(mo.team_name, '')    AS team_name,
                    COALESCE(moi.fio_text, '')    AS fio,
                    COALESCE(moi.tbn_text, '')    AS tbn,
                    COALESCE(moi.position_text, '') AS position,
                    COALESCE(mti.name, moi.meal_type_text, '') AS meal_type,
                    COALESCE(moi.comment, '')     AS comment,
                    COALESCE(moi.quantity, 1)     AS quantity
                FROM meal_orders mo
                JOIN meal_order_items moi ON moi.order_id = mo.id
                LEFT JOIN objects o       ON o.id = mo.object_id
                LEFT JOIN departments d_order ON d_order.id = mo.department_id
                LEFT JOIN employees e        ON e.id = moi.employee_id
                LEFT JOIN departments d_emp  ON d_emp.id = e.department_id
                LEFT JOIN meal_types mti  ON mti.id = moi.meal_type_id
                {where_sql}
                ORDER BY mo.date, o.address, d_order.name, mo.team_name, moi.fio_text
            """
            cur.execute(sql, params)
            rows = cur.fetchall()

        result = []
        for r in rows:
            (
                date_str,
                address,
                object_excel_id,
                department,
                employee_department,
                team_name,
                fio,
                tbn,
                position,
                meal_type,
                comment,
                quantity,
            ) = r
            result.append(
                {
                    "date": date_str,
                    "address": address,
                    "object_id": object_excel_id,
                    "department": department,
                    "employee_department": employee_department,  # из employees (1С)
                    "team_name": team_name,
                    "fio": fio,
                    "tbn": tbn,
                    "position": position,
                    "meal_type": meal_type,
                    "comment": comment,
                    "quantity": float(quantity or 1),
                }
            )
        return result

    finally:
        if conn:
            release_db_connection(conn)

def find_conflicting_meal_orders_same_date_other_object(data: dict) -> List[Dict[str, Any]]:
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            order_date = datetime.strptime(data["date"], "%Y-%m-%d").date()
            obj = data.get("object") or {}
            obj_excel_id = (obj.get("id") or "").strip()
            obj_address = (obj.get("address") or "").strip()

            # только ищем объект, не создаём
            current_object_id = get_object_id(cur, obj_excel_id, obj_address)
            if current_object_id is None:
                return []

            conflicts: List[Dict[str, Any]] = []

            for emp in data.get("employees", []):
                fio = (emp.get("fio") or "").strip()
                tbn = (emp.get("tbn") or "").strip()
                if not fio and not tbn:
                    continue

                params = [order_date, current_object_id]
                where_emp = []

                if tbn:
                    where_emp.append("(moi.tbn_text = %s OR e.tbn = %s)")
                    params.extend([tbn, tbn])
                else:
                    where_emp.append("(moi.fio_text = %s OR e.fio = %s)")
                    params.extend([fio, fio])

                where_emp_sql = " AND ".join(where_emp)

                sql = f"""
                    SELECT
                        mo.date::text,
                        COALESCE(o.address, '')       AS address,
                        COALESCE(mo.team_name, '')    AS team_name,
                        COALESCE(d.name, '')          AS department
                    FROM meal_orders mo
                    JOIN meal_order_items moi ON moi.order_id = mo.id
                    LEFT JOIN employees e    ON e.id = moi.employee_id
                    LEFT JOIN objects o      ON o.id = mo.object_id
                    LEFT JOIN departments d  ON d.id = mo.department_id
                    WHERE mo.date = %s
                      AND mo.object_id <> %s
                      AND {where_emp_sql}
                    LIMIT 5
                """
                cur.execute(sql, params)
                rows = cur.fetchall()
                for r in rows:
                    date_str, addr, team_name, dep = r
                    conflicts.append(
                        {
                            "fio": fio,
                            "tbn": tbn,
                            "date": date_str,
                            "address": addr,
                            "team_name": team_name,
                            "department": dep,
                        }
                    )

            return conflicts
    finally:
        if conn:
            release_db_connection(conn)

def parse_date_any(s: str) -> Optional[date]:
    s = (s or "").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None

def post_json(url: str, payload: dict, token: str = "") -> Tuple[bool, str]:
    try:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        if token:
            sep = "&" if ("?" in url) else "?"
            url = f"{url}{sep}token={urllib.parse.quote(token)}"
        req = urllib.request.Request(
            url,
            data=body,
            headers={"Content-Type": "application/json; charset=utf-8"},
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=25) as resp:
            code = resp.getcode()
            text = resp.read().decode("utf-8", errors="replace")
            return (200 <= code < 300, f"{code}: {text}")
    except urllib.error.HTTPError as e:
        try:
            txt = e.read().decode("utf-8", errors="replace")
        except Exception:
            txt = str(e)
        return (False, f"HTTPError {e.code}: {txt}")
    except Exception as e:
        return (False, f"Error: {e}")

def safe_filename(s: str, maxlen: int = 60) -> str:
    if not s:
        return "NOID"
    s = re.sub(r'[<>:"/\\|?*\n\r\t]+', "_", str(s)).strip()
    s = re.sub(r"_+", "_", s)
    return s[:maxlen] if len(s) > maxlen else s

def disable_mousewheel(widget: tk.Widget):
    def _block(event):
        return "break"
    widget.bind("<MouseWheel>", _block)

class AutoCompleteCombobox(ttk.Combobox):
    def __init__(self, master=None, **kw):
        super().__init__(master, **kw)
        self._all_values: List[str] = []
        self.bind("<KeyRelease>", self._on_keyrelease)
        self.bind("<Control-BackSpace>", self._clear_all)

    def set_completion_list(self, values: List[str]):
        self._all_values = list(values)
        self["values"] = self._all_values

    def _clear_all(self, _=None):
        self.delete(0, tk.END)
        self["values"] = self._all_values

    def _on_keyrelease(self, event):
        if event.keysym in (
            "Up",
            "Down",
            "Left",
            "Right",
            "Home",
            "End",
            "Return",
            "Escape",
            "Tab",
        ):
            return
        typed = self.get().strip()
        if not typed:
            self["values"] = self._all_values
            return
        self["values"] = [x for x in self._all_values if typed.lower() in x.lower()]

EMP_COL_WIDTHS = {
    0: 320,
    1: 90,
    2: 230,
    3: 140,
    4: 200,
    5: 60,
    6: 80,
}

class SelectEmployeesDialog(tk.Toplevel):

    def __init__(self, parent, employees: List[Dict[str, Any]], current_dep: str):
        super().__init__(parent)
        self.parent = parent
        self.employees = employees
        self.current_dep = (current_dep or "").strip()
        self.result: Optional[List[Dict[str, Any]]] = None

        self.title("Выбор сотрудников")
        self.resizable(True, True)
        self.grab_set()

        self.var_only_dep = tk.BooleanVar(
            value=bool(self.current_dep and self.current_dep != "Все")
        )
        self.var_search = tk.StringVar()

        main = tk.Frame(self, padx=10, pady=10)
        main.pack(fill="both", expand=True)

        top = tk.Frame(main)
        top.pack(fill="x")

        tk.Label(
            top,
            text=f"Подразделение: {self.current_dep or 'Все'}",
            font=("Segoe UI", 10, "bold"),
        ).grid(row=0, column=0, columnspan=2, sticky="w")

        ttk.Checkbutton(
            top,
            text="Показывать только сотрудников этого подразделения",
            variable=self.var_only_dep,
            command=self._refilter,
        ).grid(row=1, column=0, columnspan=2, sticky="w", pady=(4, 4))

        tk.Label(top, text="Поиск (ФИО / таб.№):").grid(
            row=2, column=0, sticky="w", pady=(4, 2)
        )
        ent_search = ttk.Entry(top, textvariable=self.var_search, width=40)
        ent_search.grid(row=2, column=1, sticky="w", pady=(4, 2))
        ent_search.bind("<KeyRelease>", lambda e: self._refilter())

        tbl_frame = tk.Frame(main)
        tbl_frame.pack(fill="both", expand=True, pady=(8, 4))

        columns = ("fio", "tbn", "pos", "dep")
        self.tree = ttk.Treeview(
            tbl_frame,
            columns=columns,
            show="headings",
            selectmode="none",
        )

        self.tree.heading("fio", text="ФИО")
        self.tree.heading("tbn", text="Таб. №")
        self.tree.heading("pos", text="Должность")
        self.tree.heading("dep", text="Подразделение")

        self.tree.column("fio", width=260, anchor="w")
        self.tree.column("tbn", width=80, anchor="center", stretch=False)
        self.tree.column("pos", width=200, anchor="w")
        self.tree.column("dep", width=160, anchor="w")

        normal_font = ("Segoe UI", 9)
        bold_font = ("Segoe UI", 9, "bold")
        self.tree.tag_configure("checked", font=bold_font)
        self.tree.tag_configure("unchecked", font=normal_font)

        vsb = ttk.Scrollbar(tbl_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)

        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        self.tree.bind("<Button-1>", self._on_tree_click)

        self._filtered_indices: List[int] = []
        self._selected_indices: set[int] = set()

        self._refilter()
        self._update_selected_count()

        sel_frame = tk.Frame(main)
        sel_frame.pack(fill="x")
        ttk.Button(sel_frame, text="Отметить всех", command=self._select_all).pack(
            side="left", padx=(0, 4)
        )
        ttk.Button(sel_frame, text="Снять все", command=self._clear_all).pack(
            side="left", padx=4
        )
        self.lbl_selected = tk.Label(
            sel_frame,
            text="Выбрано: 0",
            bg=sel_frame["bg"],
        )
        self.lbl_selected.pack(side="right")

        btns = tk.Frame(main)
        btns.pack(fill="x", pady=(8, 0))
        ttk.Button(btns, text="OK", command=self._on_ok).pack(
            side="right", padx=(4, 0)
        )
        ttk.Button(btns, text="Отмена", command=self._on_cancel).pack(side="right")

        main.rowconfigure(2, weight=1)
        main.columnconfigure(0, weight=1)

        try:
            self.update_idletasks()
            px = parent.winfo_rootx()
            py = parent.winfo_rooty()
            pw = parent.winfo_width()
            ph = parent.winfo_height()
            sw = self.winfo_width()
            sh = self.winfo_height()
            self.geometry(f"+{px + (pw - sw)//2}+{py + (ph - sh)//2}")
        except Exception:
            pass

    def _update_selected_count(self):
        try:
            self.lbl_selected.config(text=f"Выбрано: {len(self._selected_indices)}")
        except Exception:
            pass

    def _refilter(self):
        search = self.var_search.get().strip().lower()
        only_dep = self.var_only_dep.get()
        dep_sel = self.current_dep

        self.tree.delete(*self.tree.get_children())
        self._filtered_indices.clear()

        for idx, emp in enumerate(self.employees):
            fio = emp.get("fio", "")
            tbn = emp.get("tbn", "")
            pos = emp.get("pos", "")
            dep = emp.get("dep", "")

            if only_dep and dep_sel and dep_sel != "Все":
                if (dep or "").strip() != dep_sel:
                    continue

            if search:
                if search not in fio.lower() and search not in tbn.lower():
                    continue

            checked = (idx in self._selected_indices)
            display_fio = f"[{'x' if checked else ' '}] {fio}"

            iid = f"emp_{idx}"
            self.tree.insert(
                "",
                "end",
                iid=iid,
                values=(display_fio, tbn, pos, dep),
                tags=("checked" if checked else "unchecked",),
            )
            self._filtered_indices.append(idx)

    def _toggle_index(self, idx: int):
        if idx in self._selected_indices:
            self._selected_indices.remove(idx)
        else:
            self._selected_indices.add(idx)
        self._update_selected_count()

    def _on_tree_click(self, event):
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return

        row_id = self.tree.identify_row(event.y)
        if not row_id:
            return

        try:
            list_index = self.tree.index(row_id)
            emp_index = self._filtered_indices[list_index]
        except Exception:
            return

        self._toggle_index(emp_index)

        emp = self.employees[emp_index]
        fio = emp.get("fio", "")
        tbn = emp.get("tbn", "")
        pos = emp.get("pos", "")
        dep = emp.get("dep", "")
        checked = (emp_index in self._selected_indices)
        display_fio = f"[{'x' if checked else ' '}] {fio}"
        self.tree.item(
            row_id,
            values=(display_fio, tbn, pos, dep),
            tags=("checked" if checked else "unchecked",),
        )
    def _select_all(self):
        for idx in self._filtered_indices:
            self._selected_indices.add(idx)
        self._refilter()
        self._update_selected_count()

    def _clear_all(self):
        self._selected_indices.clear()
        self._refilter()
        self._update_selected_count()

    def _on_ok(self):
        if not self._selected_indices:
            if not messagebox.askyesno(
                "Выбор сотрудников",
                "Не выбрано ни одного сотрудника.\nЗакрыть окно?",
                parent=self,
            ):
                return
            self.result = []
        else:
            self.result = [self.employees[i] for i in sorted(self._selected_indices)]
        self.destroy()

    def _on_cancel(self):
        self.result = None
        self.destroy()

# ============================================================
# УЛУЧШЕННАЯ ВЕРСИЯ: MealOrderPage + EmployeeRow
# Вставить вместо существующих классов в meals_module.py
# ============================================================

# Цветовая схема
COLORS = {
    "bg":           "#f0f2f5",   # фон страницы
    "panel":        "#ffffff",   # фон панелей
    "accent":       "#1565c0",   # акцентный синий
    "accent_light": "#e3f2fd",   # светло-синий фон заголовков
    "success":      "#2e7d32",   # зелёный (OK)
    "warning":      "#b00020",   # красный (ошибка)
    "border":       "#dde1e7",   # цвет рамок
    "row_even":     "#ffffff",
    "row_odd":      "#f8f9fb",
    "row_hover":    "#e8f4fd",
    "err":          "#ffccbc",
    "btn_save_bg":  "#1565c0",
    "btn_save_fg":  "#ffffff",
    "lbl_req":      "#b00020",   # цвет звёздочки обязательного поля
}

EMP_COL_WIDTHS = {
    0: 320,   # ФИО
    1: 90,    # Таб. №
    2: 200,   # Должность
    3: 150,   # Тип питания
    4: 180,   # Комментарий
    5: 60,    # Кол-во
    6: 32,    # Удалить
}


class EmployeeRow:
    """Строка сотрудника в таблице заявки — улучшенная версия."""

    ERR_BG   = COLORS["err"]
    ZEBRA_EVEN = COLORS["row_even"]
    ZEBRA_ODD  = COLORS["row_odd"]

    def __init__(self, parent, idx: int, emp_names: List[str],
                 meal_types: List[str], on_delete, on_change=None):
        self.parent    = parent
        self.idx       = idx
        self.on_delete = on_delete
        self.on_change = on_change   # колбэк при любом изменении строки
        self.emp_names = emp_names
        self.meal_types = meal_types

        self.frame = tk.Frame(parent, bg=self.ZEBRA_EVEN,
                              relief="flat", bd=0)

        # ── Номер строки ──────────────────────────────────────
        self.lbl_num = tk.Label(
            self.frame, text=f"{idx}.", width=3, anchor="e",
            bg=self.ZEBRA_EVEN, fg="#888", font=("Segoe UI", 8)
        )
        self.lbl_num.grid(row=0, column=0, padx=(4, 0), pady=2, sticky="e")

        # ── ФИО (автодополнение) ──────────────────────────────
        self.fio_var = tk.StringVar()
        self.cmb_fio = AutoCompleteCombobox(
            self.frame, textvariable=self.fio_var, width=36,
            font=("Segoe UI", 9)
        )
        self.cmb_fio.set_completion_list(emp_names)
        self.cmb_fio.grid(row=0, column=1, padx=(2, 1), pady=2, sticky="ew")
        self.cmb_fio.bind("<MouseWheel>", lambda e: "break")

        # ── Таб. № + Должность (только чтение, авто-заполнение) ─
        self.lbl_info = tk.Label(
            self.frame, text="", width=22, anchor="w",
            bg=self.ZEBRA_EVEN, fg="#444", font=("Segoe UI", 8),
            cursor="arrow"
        )
        self.lbl_info.grid(row=0, column=2, padx=(2, 1), pady=2, sticky="w")

        # ── Тип питания ───────────────────────────────────────
        self.cmb_meal_type = ttk.Combobox(
            self.frame, values=meal_types, state="readonly", width=16
        )
        if meal_types:
            self.cmb_meal_type.set(meal_types[0])
        self.cmb_meal_type.grid(row=0, column=3, padx=(2, 1), pady=2, sticky="ew")
        disable_mousewheel(self.cmb_meal_type)
        if self.on_change:
            self.cmb_meal_type.bind("<<ComboboxSelected>>",
                                    lambda e: self.on_change())

        # ── Комментарий ───────────────────────────────────────
        self.ent_comment = ttk.Entry(self.frame, width=22,
                                     font=("Segoe UI", 9))
        self.ent_comment.grid(row=0, column=4, padx=(2, 1), pady=2, sticky="ew")

        # ── Кол-во ────────────────────────────────────────────
        self.ent_quantity = ttk.Entry(self.frame, width=5,
                                      font=("Segoe UI", 9),
                                      justify="center")
        self.ent_quantity.insert(0, "1")
        self.ent_quantity.grid(row=0, column=5, padx=(2, 1), pady=2, sticky="ew")

        # ── Кнопка удаления (маленькая «×») ──────────────────
        self.btn_del = tk.Label(
            self.frame, text="×", width=2,
            bg="#ffebee", fg="#c62828",
            font=("Segoe UI", 11, "bold"),
            cursor="hand2", relief="flat"
        )
        self.btn_del.grid(row=0, column=6, padx=(2, 4), pady=2)
        self.btn_del.bind("<Button-1>", lambda e: self._delete())
        self.btn_del.bind("<Enter>",
                          lambda e: self.btn_del.config(bg="#ef9a9a"))
        self.btn_del.bind("<Leave>",
                          lambda e: self.btn_del.config(bg="#ffebee"))

        # Растягиваем колонки
        self.frame.grid_columnconfigure(1, weight=3)
        self.frame.grid_columnconfigure(2, weight=2)
        self.frame.grid_columnconfigure(3, weight=1)
        self.frame.grid_columnconfigure(4, weight=2)

    # ── Публичный API ─────────────────────────────────────────────

    def grid(self, row: int):
        self.frame.grid(row=row, column=0, sticky="ew", pady=0)
        self.lbl_num.config(text=f"{row + 1}.")

    def destroy(self):
        self.frame.destroy()

    def apply_zebra(self, row_idx: int):
        bg = self.ZEBRA_ODD if (row_idx % 2 == 1) else self.ZEBRA_EVEN
        self.frame.config(bg=bg)
        for w in (self.lbl_num, self.lbl_info):
            w.config(bg=bg)
        # ttk-виджеты фон через style не меняем — они и так нормально смотрятся

    def set_info(self, tbn: str, pos: str):
        """Отображает таб. № и должность одной строкой."""
        parts = []
        if tbn:
            parts.append(tbn)
        if pos:
            # Обрезаем длинную должность
            parts.append(pos[:25] + "…" if len(pos) > 25 else pos)
        self.lbl_info.config(text=" | ".join(parts))

        # Подсказка (tooltip) при наведении — полный текст
        full = f"Таб. №: {tbn}\nДолжность: {pos}" if (tbn or pos) else ""
        self._set_tooltip(self.lbl_info, full)

    def set_meal_type(self, meal_type: str):
        if meal_type in self.meal_types:
            self.cmb_meal_type.set(meal_type)

    def validate(self) -> bool:
        ok = True
        fio = (self.cmb_fio.get() or "").strip()
        if not fio:
            self._mark_err(self.cmb_fio)
            ok = False
        else:
            self._clear_err(self.cmb_fio)

        meal_type = (self.cmb_meal_type.get() or "").strip()
        if not meal_type:
            self._mark_err(self.cmb_meal_type)
            ok = False
        else:
            self._clear_err(self.cmb_meal_type)

        qty_str = (self.ent_quantity.get() or "").replace(",", ".").strip()
        try:
            qty = float(qty_str) if qty_str else 1.0
            if qty <= 0:
                raise ValueError
            self._clear_err(self.ent_quantity)
        except Exception:
            self._mark_err(self.ent_quantity)
            ok = False

        return ok

    def get_dict(self) -> Dict:
        qty_str = (self.ent_quantity.get() or "").replace(",", ".").strip()
        try:
            qty = float(qty_str) if qty_str else 1.0
        except Exception:
            qty = 1.0
        # Извлекаем tabn и pos из lbl_info обратно через атрибуты
        return {
            "fio":      (self.cmb_fio.get() or "").strip(),
            "tbn":      getattr(self, "_tbn", ""),
            "position": getattr(self, "_pos", ""),
            "meal_type": (self.cmb_meal_type.get() or "").strip(),
            "comment":  (self.ent_comment.get() or "").strip(),
            "quantity": qty,
        }

    # ── Хранение tabn/pos ─────────────────────────────────────────

    def store_emp_info(self, tbn: str, pos: str):
        """Сохраняет tabn и pos как атрибуты для последующего get_dict()."""
        self._tbn = tbn
        self._pos = pos
        self.set_info(tbn, pos)

    # ── Внутренние методы ─────────────────────────────────────────

    def _delete(self):
        self.on_delete(self)

    def _mark_err(self, widget):
        try:
            widget.configure(background=self.ERR_BG)
        except Exception:
            pass

    def _clear_err(self, widget):
        try:
            widget.configure(background="white")
        except Exception:
            pass

    def _set_tooltip(self, widget, text: str):
        """Простой tooltip на основе bind."""
        def _show(event):
            if not text:
                return
            tip = tk.Toplevel(widget)
            tip.wm_overrideredirect(True)
            tip.wm_geometry(f"+{event.x_root + 12}+{event.y_root + 6}")
            lbl = tk.Label(
                tip, text=text, justify="left",
                bg="#ffffe0", relief="solid", bd=1,
                font=("Segoe UI", 8), padx=4, pady=2
            )
            lbl.pack()
            widget._tip = tip

        def _hide(event):
            tip = getattr(widget, "_tip", None)
            if tip:
                try:
                    tip.destroy()
                except Exception:
                    pass
                widget._tip = None

        widget.bind("<Enter>", _show, add="+")
        widget.bind("<Leave>", _hide, add="+")


class MealOrderPage(tk.Frame):
    """Улучшенная страница создания/редактирования заявки на питание."""

    def __init__(self, master, existing_data: dict = None,
                 order_id: int = None, on_saved=None):
        super().__init__(master, bg=COLORS["bg"])
        ensure_config()
        self.base_dir   = exe_dir()
        self.orders_dir = self.base_dir / ORDERS_DIR
        self.orders_dir.mkdir(parents=True, exist_ok=True)

        self.edit_order_id = order_id
        self.on_saved      = on_saved
        self.emp_rows: List[EmployeeRow] = []

        self._load_refs_from_db()
        self._build_ui()

        if existing_data:
            self._fill_from_existing(existing_data)

    # ════════════════════════════════════════════════════════════
    #  Загрузка справочников
    # ════════════════════════════════════════════════════════════

    def _load_refs_from_db(self):
        emps = load_employees_from_db()
        self.emps       = emps
        self.emp_by_fio = {r["fio"]: r for r in emps}

        self.objects         = load_objects_from_db()
        self.meal_types_full = load_meal_types_from_db()
        self.meal_types      = [mt["name"] for mt in self.meal_types_full] or [
            "Одноразовое", "Двухразовое", "Трехразовое"
        ]

        self.deps = ["Все"] + sorted(
            {(r["dep"] or "").strip() for r in self.emps
             if (r["dep"] or "").strip()}
        )
        self.emp_names_all = [r["fio"] for r in self.emps]

        self.addr_to_ids: Dict[str, List[str]] = {}
        for oid, addr in self.objects:
            if not addr:
                continue
            self.addr_to_ids.setdefault(addr, [])
            if oid and oid not in self.addr_to_ids[addr]:
                self.addr_to_ids[addr].append(oid)
        addresses_set = set(self.addr_to_ids.keys())
        addresses_set.update(addr for _, addr in self.objects if addr)
        self.addresses = sorted(addresses_set)

    # ════════════════════════════════════════════════════════════
    #  Построение UI
    # ════════════════════════════════════════════════════════════

    def _build_ui(self):
        # Заголовок страницы
        self._build_header()
        # Верхняя форма (2 панели рядом)
        self._build_top_form()
        # Секция сотрудников
        self._build_employees_section()
        # Нижняя панель с кнопками
        self._build_bottom_bar()

    # ── Заголовок ─────────────────────────────────────────────

    def _build_header(self):
        hdr = tk.Frame(self, bg=COLORS["accent"], pady=6)
        hdr.pack(fill="x")

        title_text = (
            "✏️  Редактирование заявки" if self.edit_order_id
            else "📋  Новая заявка на питание"
        )
        tk.Label(
            hdr, text=title_text,
            font=("Segoe UI", 12, "bold"),
            bg=COLORS["accent"], fg="white",
            padx=12
        ).pack(side="left")

        # Индикатор дедлайна (подгружается после)
        self.lbl_deadline_info = tk.Label(
            hdr, text="", font=("Segoe UI", 9),
            bg=COLORS["accent"], fg="#bbdefb", padx=8
        )
        self.lbl_deadline_info.pack(side="right")
        self._show_deadline_info()

    def _show_deadline_info(self):
        """Показывает информацию о дедлайне приёма заявок в заголовке."""
        try:
            deadline = get_meals_next_day_deadline()
            limit    = get_meals_limit_next_day_only()
            if limit and deadline:
                self.lbl_deadline_info.config(
                    text=f"⏰ Приём заявок до {deadline.strftime('%H:%M')} сегодня"
                )
            elif limit:
                self.lbl_deadline_info.config(
                    text="⏰ Заявки только на завтра"
                )
        except Exception:
            pass

    # ── Верхняя форма ─────────────────────────────────────────

    def _build_top_form(self):
        """Две колонки: левая — детали заявки, правая — объект."""
        form_outer = tk.Frame(self, bg=COLORS["bg"])
        form_outer.pack(fill="x", padx=10, pady=(8, 4))
        form_outer.grid_columnconfigure(0, weight=1)
        form_outer.grid_columnconfigure(1, weight=2)

        self._build_order_panel(form_outer)   # левая панель
        self._build_object_panel(form_outer)  # правая панель

    def _build_order_panel(self, parent):
        """Левая панель: дата, подразделение, бригада."""
        pnl = tk.LabelFrame(
            parent, text=" 📄 Параметры заявки ",
            font=("Segoe UI", 9, "bold"),
            bg=COLORS["panel"], fg=COLORS["accent"],
            relief="groove", bd=1,
            padx=10, pady=8
        )
        pnl.grid(row=0, column=0, sticky="nsew", padx=(0, 4), pady=2)
        pnl.grid_columnconfigure(1, weight=1)

        row = 0

        # Дата
        self._lbl(pnl, "Дата заявки", row, required=True)
        date_frame = tk.Frame(pnl, bg=COLORS["panel"])
        date_frame.grid(row=row, column=1, sticky="ew", padx=(0, 0), pady=3)

        self.ent_date = ttk.Entry(date_frame, width=12,
                                  font=("Segoe UI", 10))
        self.ent_date.pack(side="left")
        tomorrow_str = (date.today() + timedelta(days=1)).strftime("%d.%m.%Y")
        self.ent_date.insert(0, tomorrow_str)
        self.ent_date.bind("<KeyRelease>",  lambda e: self._update_date_hint())
        self.ent_date.bind("<FocusOut>",    lambda e: self._update_date_hint())

        self.lbl_date_hint = tk.Label(
            date_frame, text="", fg=COLORS["success"],
            bg=COLORS["panel"], font=("Segoe UI", 8)
        )
        self.lbl_date_hint.pack(side="left", padx=(6, 0))
        row += 1

        # Подразделение
        self._lbl(pnl, "Подразделение", row, required=True)
        self.cmb_dep = ttk.Combobox(
            pnl, state="readonly", values=self.deps, width=32
        )
        saved_dep = get_saved_dep()
        self.cmb_dep.set(saved_dep if saved_dep in self.deps else self.deps[0])
        self.cmb_dep.grid(row=row, column=1, sticky="ew", pady=3)
        disable_mousewheel(self.cmb_dep)
        self.cmb_dep.bind("<<ComboboxSelected>>", self._on_dep_changed)
        row += 1

        # Наименование бригады
        self._lbl(pnl, "Бригада", row, required=True)
        self.ent_team = ttk.Entry(pnl, width=32, font=("Segoe UI", 9))
        self.ent_team.grid(row=row, column=1, sticky="ew", pady=3)
        row += 1

    def _build_object_panel(self, parent):
        """Правая панель: адрес объекта, ID объекта, фактический адрес."""
        pnl = tk.LabelFrame(
            parent, text=" 📍 Объект (место работы) ",
            font=("Segoe UI", 9, "bold"),
            bg=COLORS["panel"], fg=COLORS["accent"],
            relief="groove", bd=1,
            padx=10, pady=8
        )
        pnl.grid(row=0, column=1, sticky="nsew", padx=(4, 0), pady=2)
        pnl.grid_columnconfigure(1, weight=1)

        row = 0

        # Адрес объекта
        self._lbl(pnl, "Адрес объекта", row, required=True)
        addr_frame = tk.Frame(pnl, bg=COLORS["panel"])
        addr_frame.grid(row=row, column=1, sticky="ew", pady=3)
        addr_frame.grid_columnconfigure(0, weight=1)

        self.cmb_address = AutoCompleteCombobox(
            addr_frame, width=38, font=("Segoe UI", 9)
        )
        self.cmb_address.set_completion_list(self.addresses)
        self.cmb_address.grid(row=0, column=0, sticky="ew")
        disable_mousewheel(self.cmb_address)
        self.cmb_address.bind("<<ComboboxSelected>>", self._on_address_selected)
        self.cmb_address.bind("<FocusOut>",
            lambda e: self._sync_ids_by_address(self.cmb_address.get()))
        self.cmb_address.bind("<Return>",
            lambda e: self._sync_ids_by_address(self.cmb_address.get()))
        row += 1

        # ID объекта (автоматически по адресу)
        self._lbl(pnl, "ID объекта", row)
        id_frame = tk.Frame(pnl, bg=COLORS["panel"])
        id_frame.grid(row=row, column=1, sticky="ew", pady=3)
    
        self.cmb_object_id = ttk.Combobox(
            id_frame, state="readonly", values=[], width=20
        )
        self.cmb_object_id.pack(side="left")
        disable_mousewheel(self.cmb_object_id)
    
        tk.Label(
            id_frame,
            text="← выбирается автоматически по адресу",
            font=("Segoe UI", 7), fg="#888", bg=COLORS["panel"]
        ).pack(side="left", padx=6)
        row += 1

        # Фактический адрес
        self._lbl(pnl, "Факт. адрес", row)
        self.ent_fact_address = ttk.Entry(pnl, width=38, font=("Segoe UI", 9))
        self.ent_fact_address.grid(row=row, column=1, sticky="ew", pady=3)
        tk.Label(
            pnl, text="(если отличается от объекта)",
            font=("Segoe UI", 7), fg="#888", bg=COLORS["panel"]
        ).grid(row=row + 1, column=1, sticky="w")
        row += 2

    # ── Секция сотрудников ────────────────────────────────────

    def _build_employees_section(self):
        emp_frame = tk.LabelFrame(
            self, text=" 👥 Список сотрудников ",
            font=("Segoe UI", 9, "bold"),
            bg=COLORS["panel"], fg=COLORS["accent"],
            relief="groove", bd=1,
        )
        emp_frame.pack(fill="both", expand=True, padx=10, pady=(4, 4))

        # ── Тулбар (кнопки + массовый тип питания) ──────────
        toolbar = tk.Frame(emp_frame, bg=COLORS["accent_light"], pady=5)
        toolbar.pack(fill="x")

        # Кнопки добавления
        btn_add = ttk.Button(
            toolbar, text="➕ Добавить сотрудника",
            command=self.add_employee
        )
        btn_add.pack(side="left", padx=(8, 4))

        btn_dep = ttk.Button(
            toolbar, text="👥 Добавить всё подразделение",
            command=self.add_department
        )
        btn_dep.pack(side="left", padx=4)

        btn_sel = ttk.Button(
            toolbar, text="☑ Выбрать из подразделения…",
            command=self.add_department_partial
        )
        btn_sel.pack(side="left", padx=4)

        # Разделитель
        tk.Frame(toolbar, bg=COLORS["border"], width=1).pack(
            side="left", fill="y", padx=8
        )

        # Массовое изменение типа питания
        tk.Label(
            toolbar, text="Тип питания для всех:",
            font=("Segoe UI", 9), bg=COLORS["accent_light"]
        ).pack(side="left")

        self.cmb_mass_meal = ttk.Combobox(
            toolbar, values=self.meal_types,
            state="readonly", width=14
        )
        if self.meal_types:
            self.cmb_mass_meal.set(self.meal_types[0])
        self.cmb_mass_meal.pack(side="left", padx=4)
        disable_mousewheel(self.cmb_mass_meal)

        ttk.Button(
            toolbar, text="Применить",
            command=self._apply_mass_meal_type
        ).pack(side="left", padx=(0, 8))

        # Счётчик справа
        self.lbl_emp_count = tk.Label(
            toolbar, text="Сотрудников: 0",
            font=("Segoe UI", 9, "bold"),
            fg=COLORS["accent"], bg=COLORS["accent_light"]
        )
        self.lbl_emp_count.pack(side="right", padx=12)

        # ── Заголовок таблицы ────────────────────────────────
        self._build_table_header(emp_frame)

        # ── Скроллируемая область строк ──────────────────────
        wrap = tk.Frame(emp_frame, bg=COLORS["panel"])
        wrap.pack(fill="both", expand=True)

        self.cv = tk.Canvas(
            wrap, bg=COLORS["panel"],
            borderwidth=0, highlightthickness=0
        )
        self.rows_holder = tk.Frame(self.cv, bg=COLORS["panel"])
        self.cv.create_window((0, 0), window=self.rows_holder, anchor="nw")

        self.vscroll = ttk.Scrollbar(wrap, orient="vertical",
                                     command=self.cv.yview)
        self.cv.configure(yscrollcommand=self.vscroll.set)

        self.cv.pack(side="left", fill="both", expand=True)
        self.vscroll.pack(side="right", fill="y")

        self.rows_holder.bind(
            "<Configure>",
            lambda e: self.cv.configure(scrollregion=self.cv.bbox("all"))
        )
        self.cv.bind(
            "<MouseWheel>",
            lambda e: (
                self.cv.yview_scroll(int(-1 * (e.delta / 120)), "units"),
                "break"
            )
        )

        # Сообщение "список пуст"
        self.lbl_empty = tk.Label(
            self.rows_holder,
            text="Список сотрудников пуст.\n"
                 "Нажмите «Добавить сотрудника» или выберите из подразделения.",
            font=("Segoe UI", 9), fg="#999",
            bg=COLORS["panel"], pady=20
        )
        self.lbl_empty.grid(row=0, column=0, sticky="ew", padx=20)
        self.rows_holder.grid_columnconfigure(0, weight=1)

    def _build_table_header(self, parent):
        hdr = tk.Frame(parent, bg=COLORS["accent_light"], pady=4)
        hdr.pack(fill="x")

        # Порядок и ширины колонок должны совпадать с EmployeeRow
        columns = [
            ("№",              3,   "e"),
            ("ФИО сотрудника *", 36, "w"),
            ("Таб. № / Должность", 22, "w"),
            ("Тип питания *",  16, "w"),
            ("Комментарий",    22, "w"),
            ("Кол-во",          5,  "c"),
            ("",                2,  "c"),   # кнопка удаления
        ]
        for i, (text, width, anchor) in enumerate(columns):
            lbl = tk.Label(
                hdr, text=text,
                font=("Segoe UI", 8, "bold"),
                fg=COLORS["accent"], bg=COLORS["accent_light"],
                anchor=anchor
            )
            lbl.grid(row=0, column=i, padx=3, sticky="ew")
            hdr.grid_columnconfigure(i, weight=(3 if i == 1 else 1))

    # ── Нижняя панель ─────────────────────────────────────────

    def _build_bottom_bar(self):
        bar = tk.Frame(self, bg=COLORS["bg"], pady=6)
        bar.pack(fill="x", padx=10)

        # Кнопка "Сохранить" — крупная и акцентная
        self.btn_save = tk.Button(
            bar,
            text="💾  СОХРАНИТЬ ЗАЯВКУ",
            font=("Segoe UI", 10, "bold"),
            bg=COLORS["btn_save_bg"],
            fg=COLORS["btn_save_fg"],
            activebackground="#0d47a1",
            activeforeground="white",
            relief="flat", cursor="hand2",
            padx=18, pady=6,
            command=self.save_order
        )
        self.btn_save.pack(side="left", padx=(0, 8))
        self.btn_save.bind("<Enter>",
                           lambda e: self.btn_save.config(bg="#0d47a1"))
        self.btn_save.bind("<Leave>",
                           lambda e: self.btn_save.config(bg=COLORS["btn_save_bg"]))

        ttk.Button(
            bar, text="🗑  Очистить форму",
            command=self.clear_form
        ).pack(side="left", padx=4)

        ttk.Button(
            bar, text="📁  Открыть папку заявок",
            command=self.open_orders_dir
        ).pack(side="left", padx=4)

        # Инициализируем
        self._update_emp_list()
        self._update_date_hint()

    # ── Вспомогательный метод для меток ──────────────────────

    def _lbl(self, parent, text: str, row: int, required: bool = False):
        """Создаёт метку поля формы."""
        display = f"{text}{'  *' if required else ''}:"
        fg = COLORS["warning"] if required else "#333"
        tk.Label(
            parent, text=display,
            font=("Segoe UI", 9),
            bg=COLORS["panel"], fg=fg,
            anchor="e"
        ).grid(row=row, column=0, sticky="e", padx=(0, 8), pady=3)

    # ════════════════════════════════════════════════════════════
    #  Логика работы с сотрудниками
    # ════════════════════════════════════════════════════════════

    def add_employee(self):
        """Добавляет пустую строку сотрудника."""
        self._hide_empty_label()
        names = self._get_emp_names_for_dep()
        row = EmployeeRow(
            self.rows_holder,
            len(self.emp_rows) + 1,
            names,
            self.meal_types,
            self.delete_employee,
            on_change=self._update_emp_count,
        )
        row.grid(len(self.emp_rows))
        row.apply_zebra(len(self.emp_rows))

        row.cmb_fio.bind("<<ComboboxSelected>>",
                         lambda e, r=row: self._fill_emp_info(r))
        row.cmb_fio.bind("<FocusOut>",
                         lambda e, r=row: self._fill_emp_info(r))

        self.emp_rows.append(row)
        self._update_emp_list()
        self._update_emp_count()

        # Скролл вниз к новой строке
        self.cv.update_idletasks()
        self.cv.yview_moveto(1.0)

    def delete_employee(self, emp_row: EmployeeRow):
        try:
            self.emp_rows.remove(emp_row)
        except Exception:
            pass
        emp_row.destroy()
        for i, r in enumerate(self.emp_rows):
            r.grid(i)
            r.apply_zebra(i)
        self._show_empty_label_if_needed()
        self._update_emp_count()

    def add_department(self):
        dep = (self.cmb_dep.get() or "Все").strip()
        candidates = (
            self.emps[:] if dep == "Все"
            else [e for e in self.emps if (e["dep"] or "").strip() == dep]
        )
        if not candidates:
            messagebox.showinfo("Питание",
                                f"В подразделении «{dep}» нет сотрудников.")
            return

        existing = {r.cmb_fio.get().strip() for r in self.emp_rows}
        added = 0
        for e in candidates:
            fio = e["fio"]
            if fio in existing:
                continue
            self._add_emp_from_dict(e)
            existing.add(fio)
            added += 1

        self._update_emp_list()
        self._update_emp_count()
        messagebox.showinfo("Питание", f"Добавлено сотрудников: {added}")

    def add_department_partial(self):
        dep = (self.cmb_dep.get() or "Все").strip()
        if not self.emps:
            messagebox.showinfo("Питание", "Справочник сотрудников пуст.")
            return

        dlg = SelectEmployeesDialog(self, self.emps, dep)
        self.wait_window(dlg)
        if dlg.result is None:
            return

        existing = {r.cmb_fio.get().strip() for r in self.emp_rows}
        added = 0
        for emp in (dlg.result or []):
            fio = (emp.get("fio") or "").strip()
            if not fio or fio in existing:
                continue
            self._add_emp_from_dict(emp)
            existing.add(fio)
            added += 1

        self._update_emp_list()
        self._update_emp_count()
        if added:
            messagebox.showinfo("Питание", f"Добавлено сотрудников: {added}")
        else:
            messagebox.showinfo("Питание",
                                "Все выбранные сотрудники уже есть в заявке.")

    def _add_emp_from_dict(self, emp: dict, meal_type: str = ""):
        """Внутренний метод: добавляет строку по словарю сотрудника."""
        self._hide_empty_label()
        row = EmployeeRow(
            self.rows_holder,
            len(self.emp_rows) + 1,
            [],
            self.meal_types,
            self.delete_employee,
            on_change=self._update_emp_count,
        )
        row.grid(len(self.emp_rows))
        row.apply_zebra(len(self.emp_rows))
        row.fio_var.set(emp.get("fio", ""))
        row.store_emp_info(
            tbn=emp.get("tbn", ""),
            pos=emp.get("pos", "") or emp.get("position", "")
        )
        if meal_type and meal_type in self.meal_types:
            row.set_meal_type(meal_type)
        self.emp_rows.append(row)

        row.cmb_fio.bind("<<ComboboxSelected>>",
                         lambda e, r=row: self._fill_emp_info(r))
        row.cmb_fio.bind("<FocusOut>",
                         lambda e, r=row: self._fill_emp_info(r))

    def _apply_mass_meal_type(self):
        """Устанавливает один тип питания для всех строк."""
        mt = (self.cmb_mass_meal.get() or "").strip()
        if not mt:
            return
        for row in self.emp_rows:
            row.set_meal_type(mt)

    # ── Вспомогательные методы ────────────────────────────────

    def _hide_empty_label(self):
        try:
            self.lbl_empty.grid_remove()
        except Exception:
            pass

    def _show_empty_label_if_needed(self):
        if not self.emp_rows:
            try:
                self.lbl_empty.grid(row=0, column=0, sticky="ew", padx=20)
            except Exception:
                pass

    def _get_emp_names_for_dep(self) -> List[str]:
        dep = (self.cmb_dep.get() or "Все").strip()
        if dep == "Все":
            names = [r["fio"] for r in self.emps]
        else:
            names = [r["fio"] for r in self.emps
                     if (r["dep"] or "").strip() == dep]
        seen, filtered = set(), []
        for n in names:
            if n not in seen:
                seen.add(n)
                filtered.append(n)
        if not filtered and dep != "Все":
            filtered = [r["fio"] for r in self.emps]
        return filtered

    def _update_emp_list(self):
        names = self._get_emp_names_for_dep()
        for row in self.emp_rows:
            row.cmb_fio.set_completion_list(names)

    def _update_emp_count(self):
        cnt = len(self.emp_rows)
        try:
            color = COLORS["success"] if cnt > 0 else "#888"
            self.lbl_emp_count.config(
                text=f"Сотрудников: {cnt}",
                fg=color
            )
        except Exception:
            pass

    def _fill_emp_info(self, row: EmployeeRow):
        fio  = row.fio_var.get().strip()
        info = self.emp_by_fio.get(fio)
        if info:
            row.store_emp_info(
                tbn=info.get("tbn", ""),
                pos=info.get("pos", "")
            )
        else:
            row.store_emp_info("", "")

    def _on_dep_changed(self, event=None):
        set_saved_dep(self.cmb_dep.get())
        self._update_emp_list()

    def _update_date_hint(self):
        try:
            # Принимаем оба формата: YYYY-MM-DD и DD.MM.YYYY
            raw = self.ent_date.get().strip()
            req  = parse_date_any(raw)
            today    = date.today()
            tomorrow = today + timedelta(days=1)

            if req is None:
                self.lbl_date_hint.config(
                    text="⚠ дата некорректна", fg=COLORS["warning"]
                )
            elif req < tomorrow:
                self.lbl_date_hint.config(
                    text=f"⚠ минимум {tomorrow.strftime('%d.%m.%Y')}",
                    fg=COLORS["warning"]
                )
            else:
                day_name = ["пн", "вт", "ср", "чт", "пт", "сб", "вс"][req.weekday()]
                self.lbl_date_hint.config(
                    text=f"✓ {day_name}", fg=COLORS["success"]
                )
        except Exception:
            self.lbl_date_hint.config(text="", fg="#555")

    def _sync_ids_by_address(self, address_value=None):
        addr = (address_value if address_value is not None
                else self.cmb_address.get() or "").strip()
        if address_value is not None:
            try:
                self.cmb_address.set(addr)
            except Exception:
                pass

        ids = sorted(self.addr_to_ids.get(addr, []))
        if ids:
            self.cmb_object_id.config(state="readonly", values=ids)
            if self.cmb_object_id.get() not in ids:
                self.cmb_object_id.set(ids[0])
        else:
            self.cmb_object_id.config(state="readonly", values=[])
            self.cmb_object_id.set("")

        # Авто-заполнение фактического адреса, если пустой
        fact = (self.ent_fact_address.get() or "").strip()
        if not fact and addr:
            try:
                self.ent_fact_address.delete(0, tk.END)
                self.ent_fact_address.insert(0, addr)
            except Exception:
                pass

    def _on_address_selected(self, event=None):
        full_addr = (self.cmb_address.get() or "").strip()
        if not full_addr:
            return
        if full_addr not in self.addresses:
            typed = full_addr.lower()
            for a in self.addresses:
                if a.lower().startswith(typed):
                    full_addr = a
                    break
        self.cmb_address.set(full_addr)
        self._sync_ids_by_address(full_addr)

    # ════════════════════════════════════════════════════════════
    #  Заполнение из существующей заявки
    # ════════════════════════════════════════════════════════════

    def _fill_from_existing(self, data: dict):
        # Дата
        raw_date = data.get("date", "")
        # Конвертируем в DD.MM.YYYY для отображения
        try:
            d = datetime.strptime(raw_date, "%Y-%m-%d").date()
            display_date = d.strftime("%d.%m.%Y")
        except Exception:
            display_date = raw_date
        self.ent_date.delete(0, "end")
        self.ent_date.insert(0, display_date)

        # Подразделение
        dep = data.get("department", "") or "Все"
        if dep not in self.deps:
            self.deps.append(dep)
            self.cmb_dep["values"] = self.deps
        self.cmb_dep.set(dep)

        # Объект
        obj  = data.get("object") or {}
        addr = obj.get("address", "") or ""
        oid  = obj.get("id", "") or ""
        if addr and addr not in self.addresses:
            self.addresses.append(addr)
            self.addresses.sort()
            self.cmb_address.set_completion_list(self.addresses)
        self.cmb_address.set(addr)
        self._sync_ids_by_address(addr)
        if oid:
            ids = list(self.cmb_object_id["values"])
            if oid not in ids:
                ids.append(oid)
                self.cmb_object_id["values"] = ids
            self.cmb_object_id.set(oid)

        # Фактический адрес
        fact_addr = (data.get("fact_address") or "").strip()
        self.ent_fact_address.delete(0, tk.END)
        self.ent_fact_address.insert(0, fact_addr or addr)

        # Бригада
        self.ent_team.delete(0, "end")
        self.ent_team.insert(0, data.get("team_name", ""))

        # Сотрудники
        for r in self.emp_rows:
            r.destroy()
        self.emp_rows.clear()

        for emp in data.get("employees", []):
            self._add_emp_from_dict(
                {
                    "fio":      emp.get("fio", ""),
                    "tbn":      emp.get("tbn", ""),
                    "position": emp.get("position", ""),
                },
                meal_type=emp.get("meal_type", "")
            )
            row = self.emp_rows[-1]
            mt = (emp.get("meal_type") or "").strip()
            if mt and mt not in self.meal_types:
                self.meal_types.append(mt)
                for r in self.emp_rows:
                    r.cmb_meal_type["values"] = self.meal_types
            row.cmb_meal_type.set(mt or (self.meal_types[0] if self.meal_types else ""))
            row.ent_comment.delete(0, "end")
            row.ent_comment.insert(0, emp.get("comment", ""))
            qty = emp.get("quantity") or 1
            row.ent_quantity.delete(0, tk.END)
            row.ent_quantity.insert(0, str(qty))

        self._show_empty_label_if_needed()
        self._update_date_hint()
        self._update_emp_count()

    # ════════════════════════════════════════════════════════════
    #  Сборка данных заявки
    # ════════════════════════════════════════════════════════════

    def _build_order_dict_core(self) -> Dict:
        req_date = parse_date_any(self.ent_date.get()) or date.today()
        addr     = (self.cmb_address.get() or "").strip()
        oid      = (self.cmb_object_id.get() or "").strip()
        fact_address = (self.ent_fact_address.get() or "").strip()
        employees    = [r.get_dict() for r in self.emp_rows]

        user_id  = None
        app_ref  = getattr(self, "app_ref", None)
        if app_ref is not None and hasattr(app_ref, "current_user"):
            try:
                user_id = int((app_ref.current_user or {}).get("id") or 0) or None
            except Exception:
                pass

        core = {
            "date":       req_date.strftime("%Y-%m-%d"),
            "department": (self.cmb_dep.get() or "").strip(),
            "team_name":  (self.ent_team.get() or "").strip(),
            "object":     {"id": oid, "address": addr},
            "fact_address": fact_address,
            "employees":  employees,
        }
        if user_id is not None:
            core["user_id"] = user_id
        return core

    def _build_order_dict(self) -> Dict:
        core = self._build_order_dict_core()
        core["created_at"] = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
        return core

    def _validate_form(self) -> bool:
        req = parse_date_any(self.ent_date.get())
        today = date.today()
        tomorrow = today + timedelta(days=1)
    
        if req is None:
            messagebox.showwarning(
                "Заявка",
                "Укажите корректную дату заявки."
            )
            self.ent_date.focus_set()
            return False
    
        limit_next_day_only = get_meals_limit_next_day_only()
        deadline_time = get_meals_next_day_deadline()
        now = datetime.now()
    
        if limit_next_day_only:
            if req != tomorrow:
                messagebox.showwarning(
                    "Заявка",
                    f"Сейчас можно подать заявку только на "
                    f"{tomorrow.strftime('%d.%m.%Y')}."
                )
                return False
            if deadline_time is not None:
                if now > datetime.combine(today, deadline_time):
                    messagebox.showwarning(
                        "Заявка",
                        f"Приём заявок на завтра закрыт после "
                        f"{deadline_time.strftime('%H:%M')}."
                    )
                    return False
        else:
            if req < today:
                messagebox.showwarning(
                    "Заявка",
                    f"Дата не может быть раньше сегодня "
                    f"({today.strftime('%d.%m.%Y')})."
                )
                return False
    
        if not (self.cmb_dep.get() or "").strip():
            messagebox.showwarning("Заявка", "Выберите Подразделение.")
            return False
    
        addr = (self.cmb_address.get() or "").strip()
        if not addr:
            messagebox.showwarning("Заявка", "Укажите Адрес объекта.")
            self.cmb_address.focus_set()
            return False
    
        oid = (self.cmb_object_id.get() or "").strip()
        if not oid:
            messagebox.showwarning(
                "Заявка",
                "Не выбран ID объекта.\n"
                "Выберите адрес из списка — ID подставится автоматически."
            )
            return False
    
        # Жёсткая проверка: адрес должен существовать в реестре
        if addr not in self.addr_to_ids:
            messagebox.showwarning(
                "Заявка",
                "Адрес объекта не найден в реестре.\n"
                "Выберите адрес из выпадающего списка, а не вводите его вручную."
            )
            self.cmb_address.focus_set()
            return False
    
        # Жёсткая проверка: ID должен принадлежать выбранному адресу
        valid_ids = self.addr_to_ids.get(addr, [])
        if oid not in valid_ids:
            messagebox.showwarning(
                "Заявка",
                "ID объекта не соответствует выбранному адресу.\n"
                "Повторно выберите адрес из списка."
            )
            self.cmb_address.focus_set()
            return False
    
        if not (self.ent_team.get() or "").strip():
            messagebox.showwarning("Заявка", "Укажите Наименование бригады.")
            self.ent_team.focus_set()
            return False
    
        if not self.emp_rows:
            messagebox.showwarning(
                "Заявка",
                "Добавьте хотя бы одного сотрудника."
            )
            return False
    
        all_ok = all(r.validate() for r in self.emp_rows)
        if not all_ok:
            messagebox.showwarning(
                "Заявка",
                "Исправьте подсвеченные поля.\n"
                "(ФИО и Тип питания обязательны для каждого сотрудника)"
            )
            return False
    
        return True
    # ════════════════════════════════════════════════════════════
    #  Сохранение, очистка, вспомогательные действия
    # (логика save_order НЕ изменяется — только UI)
    # ════════════════════════════════════════════════════════════

    def save_order(self):
        if not self._validate_form():
            return

        try:
            self.btn_save.configure(state="disabled")
        except Exception:
            pass

        # ── Весь код save_order начиная отсюда — БЕЗ ИЗМЕНЕНИЙ ──
        try:
            data = self._build_order_dict()
            total_items = len(data.get("employees", []))

            try:
                conflicts = find_conflicting_meal_orders_same_date_other_object(data)
            except Exception as e:
                if not messagebox.askokcancel(
                    "Проверка пересечений",
                    f"Не удалось проверить пересечения по БД:\n{e}\n\n"
                    "Продолжить сохранение?",
                ):
                    return
                conflicts = []

            if conflicts:
                lines = []
                for c in conflicts:
                    fio  = c.get("fio") or "?"
                    tbn  = c.get("tbn") or ""
                    who  = f"{fio} ({tbn})" if tbn else fio
                    addr = c.get("address") or "неизвестный адрес"
                    date_str = c.get("date") or ""
                    team = c.get("team_name") or ""
                    dep  = c.get("department") or ""
                    extra = f", бригада: {team}" if team else ""
                    if dep:
                        extra += f", подразделение: {dep}"
                    lines.append(f"- {who}: {date_str}, объект: {addr}{extra}")

                text = (
                    "Обнаружены сотрудники, на которых в ЭТУ ЖЕ дату уже заказано питание\n"
                    "на ДРУГОМ объекте:\n\n"
                    + "\n".join(lines[:20])
                    + ("\n\n(Показаны первые 20)" if len(lines) > 20 else "")
                    + "\n\nВсё равно сохранить?"
                )
                if not messagebox.askokcancel(
                    "Пересечение заявок", text
                ):
                    return

            order_db_id = None
            conn = None
            try:
                conn = get_db_connection()
                with conn.cursor() as cur:
                    with conn:
                        dept_name = (data.get("department") or "").strip()
                        dept_id = get_or_create_department(cur, dept_name) if dept_name else None

                        obj = data.get("object") or {}
                        obj_excel_id = (obj.get("id") or "").strip()
                        obj_address  = (obj.get("address") or "").strip()

                        object_id = get_object_id(cur, obj_excel_id, obj_address)
                        if object_id is None:
                            raise ValueError(
                                "Не найден объект в реестре.\n"
                                "Выберите адрес из списка, а не вводите его вручную."
                            )

                        order_date   = datetime.strptime(data["date"], "%Y-%m-%d").date()
                        team_name    = (data.get("team_name") or "").strip()
                        user_id      = data.get("user_id")
                        fact_address = (data.get("fact_address") or "").strip()

                        if self.edit_order_id:
                            existing_id = self.edit_order_id
                        else:
                            existing_id = None
                            cur.execute(
                                """
                                SELECT id FROM meal_orders
                                 WHERE date = %s AND object_id = %s
                                   AND COALESCE(department_id, 0) = COALESCE(%s, 0)
                                   AND COALESCE(team_name, '') = COALESCE(%s, '')
                                   AND COALESCE(user_id, 0) = COALESCE(%s, 0)
                                 ORDER BY id DESC LIMIT 1
                                """,
                                (order_date, object_id, dept_id, team_name, user_id),
                            )
                            row_db = cur.fetchone()
                            if row_db:
                                existing_id = row_db[0]

                        if existing_id:
                            if not messagebox.askyesno(
                                "Обновление заявки",
                                "Заявка с такими параметрами уже существует.\n\n"
                                "Перезаписать?",
                            ):
                                return

                            cur.execute(
                                "DELETE FROM meal_order_items WHERE order_id = %s",
                                (existing_id,),
                            )
                            cur.execute(
                                """
                                UPDATE meal_orders
                                   SET date = %s,
                                       department_id = %s,
                                       team_name = %s,
                                       object_id = %s,
                                       fact_address = %s
                                 WHERE id = %s
                                """,
                                (order_date, dept_id, team_name, object_id,
                                 fact_address, existing_id),
                            )
                            order_db_id = existing_id
                        else:
                            created_at = datetime.strptime(
                                data["created_at"], "%Y-%m-%dT%H:%M:%S"
                            )
                            cur.execute(
                                """
                                INSERT INTO meal_orders
                                    (created_at, date, department_id, team_name,
                                     object_id, user_id, fact_address)
                                VALUES (%s, %s, %s, %s, %s, %s, %s)
                                RETURNING id
                                """,
                                (created_at, order_date, dept_id, team_name,
                                 object_id, user_id, fact_address),
                            )
                            order_db_id = cur.fetchone()[0]

                        for emp in data.get("employees", []):
                            fio           = (emp.get("fio") or "").strip()
                            tbn           = (emp.get("tbn") or "").strip()
                            position      = (emp.get("position") or "").strip()
                            meal_type_name = (emp.get("meal_type") or "").strip()
                            comment       = (emp.get("comment") or "").strip()
                            quantity      = float(emp.get("quantity") or 1)

                            meal_type_id = get_or_create_meal_type(cur, meal_type_name)
                            employee_id  = find_employee(cur, fio, tbn)

                            cur.execute(
                                """
                                INSERT INTO meal_order_items
                                (order_id, employee_id, fio_text, tbn_text,
                                 position_text, meal_type_id, meal_type_text,
                                 comment, quantity)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                                """,
                                (
                                    order_db_id,
                                    employee_id,
                                    fio,
                                    tbn,
                                    position,
                                    meal_type_id,
                                    meal_type_name,
                                    comment,
                                    quantity,
                                ),
                            )

            except Exception as e:
                messagebox.showerror(
                    "Сохранение в БД",
                    f"Не удалось сохранить заявку:\n{e}"
                )
                return
            finally:
                if conn:
                    release_db_connection(conn)

            # ── Сохранение XLSX ───────────────────────────────────────
            ts      = datetime.now().strftime("%H%M%S")
            id_part = data["object"]["id"] or safe_filename(
                data["object"]["address"]
            )
            fname = (
                f"Заявка_питание_{data['date']}_{ts}"
                f"_{id_part or 'NOID'}.xlsx"
            )
            fpath = self.orders_dir / fname

            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Заявка"
                ws.append(["Создано",              data["created_at"]])
                ws.append(["Дата",                 data["date"]])
                ws.append(["Подразделение",         data["department"]])
                ws.append(["Наименование бригады",  data.get("team_name", "")])
                ws.append(["ID объекта",            data["object"]["id"]])
                ws.append(["Адрес (объект)",        data["object"]["address"]])
                ws.append(["Фактический адрес",     data.get("fact_address", "")])
                ws.append([])
                ws.append(["#", "ФИО", "Тип питания", "Комментарий"])
                for i, emp in enumerate(data["employees"], start=1):
                    ws.append([i, emp["fio"], emp["meal_type"], emp["comment"]])
                for col, w in enumerate([4, 40, 20, 40], start=1):
                    ws.column_dimensions[get_column_letter(col)].width = w
                ws.freeze_panes = "A10"
                wb.save(fpath)
            except Exception as e:
                messagebox.showerror(
                    "Сохранение",
                    f"Заявка сохранена в БД, но не удалось создать XLSX:\n{e}",
                )
                return

            messagebox.showinfo(
                "Готово ✓",
                f"Заявка успешно сохранена!\n\n"
                f"Сотрудников: {total_items}\n"
                f"Файл: {fpath}",
            )

            if not self.edit_order_id:
                self.clear_form()

            if self.on_saved:
                self.on_saved()

        finally:
            try:
                self.btn_save.configure(state="normal")
            except Exception:
                pass

    def clear_form(self):
        tomorrow_str = (date.today() + timedelta(days=1)).strftime("%d.%m.%Y")
        self.ent_date.delete(0, "end")
        self.ent_date.insert(0, tomorrow_str)

        self.cmb_address.set("")
        self.cmb_object_id.config(values=[])
        self.cmb_object_id.set("")

        self.ent_team.delete(0, "end")

        try:
            self.ent_fact_address.delete(0, tk.END)
        except Exception:
            pass

        for r in self.emp_rows:
            r.destroy()
        self.emp_rows.clear()

        self._show_empty_label_if_needed()
        self._update_emp_list()
        self._update_date_hint()
        self._update_emp_count()

    def open_orders_dir(self):
        try:
            os.startfile(self.orders_dir)
        except Exception as e:
            messagebox.showerror("Папка", f"Не удалось открыть папку:\n{e}")

class MyMealsOrdersPage(tk.Frame):
    def __init__(self, master, app_ref=None):
        super().__init__(master, bg="#f7f7f7")
        self.app_ref = app_ref
        self.tree = None
        self._orders: List[Dict[str, Any]] = []
        self._build_ui()
        self._load_data()

    def _get_current_user_id(self) -> Optional[int]:
        if self.app_ref is not None and hasattr(self.app_ref, "current_user"):
            try:
                return int((self.app_ref.current_user or {}).get("id") or 0) or None
            except Exception:
                return None
        return None

    def _build_ui(self):
        # ── Верхняя панель ──
        top = tk.Frame(self, bg="#f7f7f7")
        top.pack(fill="x", padx=8, pady=(8, 4))

        tk.Label(
            top,
            text="Мои заявки на питание",
            font=("Segoe UI", 12, "bold"),
            bg="#f7f7f7",
        ).pack(side="left")

        ttk.Button(
            top,
            text="Обновить",
            command=self._load_data,
        ).pack(side="right", padx=4)

        # ── Таблица ──
        frame = tk.Frame(self, bg="#f7f7f7")
        frame.pack(fill="both", expand=True, padx=8, pady=(4, 8))

        cols = ("date", "object", "department", "team", "count", "created_at")
        self.tree = ttk.Treeview(
            frame,
            columns=cols,
            show="headings",
            selectmode="browse",
        )

        self.tree.heading("date", text="Дата")
        self.tree.heading("object", text="Объект")
        self.tree.heading("department", text="Подразделение")
        self.tree.heading("team", text="Бригада")
        self.tree.heading("count", text="Сотр., чел.")
        self.tree.heading("created_at", text="Создана")

        self.tree.column("date", width=90, anchor="center")
        self.tree.column("object", width=280, anchor="w")
        self.tree.column("department", width=180, anchor="w")
        self.tree.column("team", width=220, anchor="w")
        self.tree.column("count", width=90, anchor="center")
        self.tree.column("created_at", width=140, anchor="center")

        vsb = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)

        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        # ══════════════════════════════════════════════════════════════
        # УБРАН двойной клик и Enter — вместо них кнопки ниже
        # self.tree.bind("<Double-1>", self._on_open)   # УДАЛЕНО
        # self.tree.bind("<Return>", self._on_open)      # УДАЛЕНО
        # ══════════════════════════════════════════════════════════════

        # ── Панель кнопок действий ──
        btn_frame = tk.Frame(self, bg="#f7f7f7")
        btn_frame.pack(fill="x", padx=8, pady=(0, 4))

        self.btn_edit = ttk.Button(
            btn_frame,
            text="✏️  Редактировать выбранную заявку",
            command=self._on_edit_order,
        )
        self.btn_edit.pack(side="left", padx=(0, 12), ipady=4)

        self.btn_copy = ttk.Button(
            btn_frame,
            text="📋  Создать копию на следующий день",
            command=self._on_copy_order,
        )
        self.btn_copy.pack(side="left", padx=(0, 12), ipady=4)

        # ── Подсказка внизу ──
        bottom = tk.Frame(self, bg="#f7f7f7")
        bottom.pack(fill="x", padx=8, pady=(0, 8))
        tk.Label(
            bottom,
            text="Выберите заявку в таблице, затем нажмите нужную кнопку.",
            font=("Segoe UI", 9),
            fg="#555",
            bg="#f7f7f7",
        ).pack(side="left")

    # ── Загрузка данных (без изменений) ──
    def _load_data(self):
        self.tree.delete(*self.tree.get_children())
        self._orders.clear()

        user_id = self._get_current_user_id()
        if not user_id:
            messagebox.showwarning(
                "Мои заявки",
                "Не определён текущий пользователь.",
                parent=self,
            )
            return

        try:
            orders = load_user_meal_orders(user_id)
        except Exception as e:
            messagebox.showerror(
                "Мои заявки",
                f"Ошибка загрузки списка заявок из БД:\n{e}",
                parent=self,
            )
            return
        self._orders = orders

        for o in orders:
            oid = o["id"]
            date_val = o.get("date")
            if isinstance(date_val, datetime):
                date_str = date_val.date().strftime("%Y-%m-%d")
            else:
                date_str = str(date_val or "")

            addr = o.get("object_address") or ""
            obj_code = o.get("object_id") or ""
            obj_display = f"[{obj_code}] {addr}" if obj_code else addr

            dep = o.get("department") or ""
            team = o.get("team_name") or ""
            cnt = o.get("employees_count") or 0
            created_at = o.get("created_at")
            if isinstance(created_at, datetime):
                created_str = created_at.strftime("%d.%m.%Y %H:%M")
            else:
                created_str = str(created_at or "")
            iid = str(oid)
            self.tree.insert(
                "",
                "end",
                iid=iid,
                values=(date_str, obj_display, dep, team, cnt, created_str),
            )

    def _get_selected_order(self) -> Optional[Dict[str, Any]]:
        sel = self.tree.selection()
        if not sel:
            return None
        iid = sel[0]
        try:
            oid = int(iid)
        except Exception:
            return None
        for o in self._orders:
            if int(o["id"]) == oid:
                return o
        return None

    # ══════════════════════════════════════════════════════════════════
    #  Загрузка данных заявки из БД (общий хелпер)
    # ══════════════════════════════════════════════════════════════════
    def _load_order_data(self) -> Optional[tuple]:
        """Возвращает (order_id, order_data) или None, если ничего не выбрано."""
        info = self._get_selected_order()
        if not info:
            messagebox.showinfo(
                "Мои заявки",
                "Сначала выберите заявку в таблице.",
                parent=self,
            )
            return None

        order_id = int(info["id"])
        try:
            order_data = get_order_with_items_from_db(order_id)
        except Exception as e:
            messagebox.showerror(
                "Мои заявки",
                f"Не удалось загрузить заявку id={order_id}:\n{e}",
                parent=self,
            )
            return None

        return order_id, order_data

    # ══════════════════════════════════════════════════════════════════
    #  Кнопка «Редактировать»
    # ══════════════════════════════════════════════════════════════════
    def _on_edit_order(self):
        result = self._load_order_data()
        if result is None:
            return
        order_id, order_data = result

        # Подтверждение, чтобы исключить случайное нажатие
        ok = messagebox.askokcancel(
            "Редактирование заявки",
            f"Вы собираетесь РЕДАКТИРОВАТЬ заявку #{order_id} "
            f"на дату {order_data.get('date', '?')}.\n\n"
            "Все изменения перезапишут текущую заявку.\n"
            "Продолжить?",
            parent=self,
        )
        if not ok:
            return

        self._open_order_window(
            order_data=order_data,
            edit_id=order_id,
            title=f"Редактирование заявки #{order_id}",
        )

    # ══════════════════════════════════════════════════════════════════
    #  Кнопка «Создать копию на следующий день»
    # ══════════════════════════════════════════════════════════════════
    def _on_copy_order(self):
        result = self._load_order_data()
        if result is None:
            return
        order_id, order_data = result

        try:
            old_date = datetime.strptime(order_data["date"], "%Y-%m-%d").date()
            new_date = old_date + timedelta(days=1)
            order_data["date"] = new_date.strftime("%Y-%m-%d")
        except Exception:
            pass

        self._open_order_window(
            order_data=order_data,
            edit_id=None,                          # ← None = новая заявка
            title=f"Новая заявка (копия #{order_id}) на {order_data.get('date', '?')}",
        )

    # ══════════════════════════════════════════════════════════════════
    #  Общий метод открытия окна заявки
    # ══════════════════════════════════════════════════════════════════
    def _open_order_window(self, order_data: dict, edit_id: Optional[int], title: str):
        win = tk.Toplevel(self)
        win.title(title)
        win.geometry("1300x720")

        if hasattr(self, "app_ref"):
            setattr(win, "app_ref", self.app_ref)

        def on_saved_callback():
            self._load_data()

        page = MealOrderPage(
            win,
            existing_data=order_data,
            order_id=edit_id,
            on_saved=on_saved_callback,
        )
        page.app_ref = self.app_ref
        page.pack(fill="both", expand=True)


class AllMealsOrdersPage(tk.Frame):
    def __init__(self, master, app_ref=None):
        super().__init__(master, bg="#f7f7f7")
        self.app_ref = app_ref
        self.tree = None
        self._orders: List[Dict[str, Any]] = []
        self._build_ui()
        self._load_data()  # стартовая загрузка без фильтра (или за сегодня/месяц при желании)

    def _build_ui(self):
        top = tk.Frame(self, bg="#f7f7f7")
        top.pack(fill="x", padx=8, pady=(8, 4))

        tk.Label(
            top,
            text="Реестр заявок (все пользователи)",
            font=("Segoe UI", 12, "bold"),
            bg="#f7f7f7",
        ).grid(row=0, column=0, columnspan=8, sticky="w")

        tk.Label(top, text="Дата с:", bg="#f7f7f7").grid(row=1, column=0, sticky="w", pady=(4, 0))
        self.ent_date_from = ttk.Entry(top, width=12)
        self.ent_date_from.grid(row=1, column=1, sticky="w", padx=(4, 8), pady=(4, 0))

        tk.Label(top, text="по:", bg="#f7f7f7").grid(row=1, column=2, sticky="w", pady=(4, 0))
        self.ent_date_to = ttk.Entry(top, width=12)
        self.ent_date_to.grid(row=1, column=3, sticky="w", padx=(4, 8), pady=(4, 0))
        
        tk.Label(top, text="Подразделение:", bg="#f7f7f7").grid(row=1, column=4, sticky="w", pady=(4, 0))
        self.cmb_dep_filter = ttk.Combobox(top, state="readonly", width=40)
        self.cmb_dep_filter.grid(row=1, column=5, sticky="w", padx=(4, 8), pady=(4, 0))

        tk.Label(top, text="Адрес (часть):", bg="#f7f7f7").grid(row=1, column=6, sticky="w", pady=(4, 0))
        self.ent_address_filter = ttk.Entry(top, width=24)
        self.ent_address_filter.grid(row=1, column=7, sticky="w", padx=(4, 8), pady=(4, 0))

        btn_frame = tk.Frame(top, bg="#f7f7f7")
        btn_frame.grid(row=0, column=9, rowspan=2, sticky="e")

        ttk.Button(
            btn_frame,
            text="Выгрузить в Excel",
            command=self._export_to_excel,
        ).pack(side="right", padx=4)

        ttk.Button(
            btn_frame,
            text="Применить фильтр",
            command=self._load_data,
        ).pack(side="right", padx=4)

        ttk.Button(
            btn_frame,
            text="Обновить",
            command=self._load_data,
        ).pack(side="right", padx=4)

        self._init_dep_filter()

        frame = tk.Frame(self, bg="#f7f7f7")
        frame.pack(fill="both", expand=True, padx=8, pady=(4, 8))

        cols = ("date", "object", "department", "team", "count", "user", "created_at")
        self.tree = ttk.Treeview(
            frame,
            columns=cols,
            show="headings",
            selectmode="browse",
        )

        self.tree.heading("date", text="Дата")
        self.tree.heading("object", text="Объект")
        self.tree.heading("department", text="Подразделение")
        self.tree.heading("team", text="Бригада")
        self.tree.heading("count", text="Сотр., чел.")
        self.tree.heading("user", text="Пользователь")
        self.tree.heading("created_at", text="Создана")

        self.tree.column("date", width=90, anchor="center")
        self.tree.column("object", width=260, anchor="w")
        self.tree.column("department", width=150, anchor="w")
        self.tree.column("team", width=200, anchor="w")
        self.tree.column("count", width=80, anchor="center")
        self.tree.column("user", width=160, anchor="w")
        self.tree.column("created_at", width=140, anchor="center")

        vsb = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)

        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        self.tree.bind("<Double-1>", self._on_open)
        self.tree.bind("<Return>", self._on_open)

        bottom = tk.Frame(self, bg="#f7f7f7")
        bottom.pack(fill="x", padx=8, pady=(0, 8))

        tk.Label(
            bottom,
            text="Двойной щелчок или Enter по строке — открыть заявку для редактирования или дублирования.",
            font=("Segoe UI", 9),
            fg="#555",
            bg="#f7f7f7",
        ).pack(side="left")

        self.btn_delete = ttk.Button(
            bottom,
            text="Удалить заявку",
            command=self._on_delete_clicked,
        )
        self.btn_delete.pack(side="right")
        self._update_delete_button_state()

    def _get_current_role(self) -> str:
        if self.app_ref is not None and hasattr(self.app_ref, "current_user"):
            try:
                return str((self.app_ref.current_user or {}).get("role") or "").lower()
            except Exception:
                return ""
        return ""

    def _update_delete_button_state(self):
        role = self._get_current_role()
        is_admin = (role == "admin")
        state = "normal" if is_admin else "disabled"
        try:
            self.btn_delete.configure(state=state)
        except Exception:
            pass

    def _parse_period(self) -> Tuple[Optional[date], Optional[date]]:
        txt_from = (self.ent_date_from.get() or "").strip()
        txt_to = (self.ent_date_to.get() or "").strip()

        d_from = parse_date_any(txt_from) if txt_from else None
        d_to = parse_date_any(txt_to) if txt_to else None

        return d_from, d_to

    def _init_dep_filter(self):
        conn = None
        try:
            conn = get_db_connection()
            with conn.cursor() as cur:
                cur.execute("SELECT name FROM departments ORDER BY name;")
                rows = cur.fetchall()
                deps = ["Все"]
                for (name,) in rows:
                    if name:
                        deps.append(name)
                self.cmb_dep_filter["values"] = deps
                self.cmb_dep_filter.set("Все")
        except Exception:
            self.cmb_dep_filter["values"] = ["Все"]
            self.cmb_dep_filter.set("Все")
        finally:
            if conn:
                release_db_connection(conn)

    def _load_data(self):
        self.tree.delete(*self.tree.get_children())
        self._orders.clear()

        date_from, date_to = self._parse_period()

        if date_from and date_to and date_from > date_to:
            messagebox.showwarning(
                "Реестр заявок",
                "Дата 'с' больше даты 'по'. Исправьте период.",
                parent=self,
            )
            return

        dep_filter = (self.cmb_dep_filter.get() or "").strip() if hasattr(self, "cmb_dep_filter") else ""
        addr_filter = (self.ent_address_filter.get() or "").strip() if hasattr(self, "ent_address_filter") else ""

        try:
            orders = load_all_meal_orders(
                date_from=date_from,
                date_to=date_to,
                department=dep_filter or None,
                address_substr=addr_filter or None,
            )
        except Exception as e:
            messagebox.showerror(
                "Реестр заявок",
                f"Ошибка загрузки списка заявок из БД:\n{e}",
                parent=self,
            )
            return

        self._orders = orders

        for o in orders:
            oid = o["id"]
            date_val = o.get("date")
            if isinstance(date_val, datetime):
                date_str = date_val.date().strftime("%Y-%m-%d")
            else:
                date_str = str(date_val or "")

            addr = o.get("object_address") or ""
            obj_code = o.get("object_id") or ""
            obj_display = f"[{obj_code}] {addr}" if obj_code else addr

            dep = o.get("department") or ""
            team = o.get("team_name") or ""
            cnt = o.get("employees_count") or 0
            user_name = o.get("user_name") or ""
            created_at = o.get("created_at")
            if isinstance(created_at, datetime):
                created_str = created_at.strftime("%d.%m.%Y %H:%M")
            else:
                created_str = str(created_at or "")

            iid = str(oid)
            self.tree.insert(
                "",
                "end",
                iid=iid,
                values=(date_str, obj_display, dep, team, cnt, user_name, created_str),
            )

    def _export_to_excel(self):
        try:
            date_from, date_to = self._parse_period()

            if date_from and date_to and date_from > date_to:
                messagebox.showwarning(
                    "Экспорт в Excel",
                    "Дата 'с' больше даты 'по'. Исправьте период.",
                    parent=self,
                )
                return

            dep_filter = (self.cmb_dep_filter.get() or "").strip() if hasattr(self, "cmb_dep_filter") else ""
            addr_filter = (self.ent_address_filter.get() or "").strip() if hasattr(self, "ent_address_filter") else ""

            filter_date_str: Optional[str] = None
            if date_from and date_to and date_from == date_to:
                filter_date_str = date_from.strftime("%Y-%m-%d")
            elif date_from and not date_to:
                filter_date_str = date_from.strftime("%Y-%m-%d")
            elif date_to and not date_from:
                filter_date_str = date_to.strftime("%Y-%m-%d")
            else:
                filter_date_str = None

            try:
                raw_orders = get_details_from_db(
                    filter_date=filter_date_str,
                    filter_address=addr_filter or None,
                    filter_department=dep_filter or None,
                )
            except Exception as e:
                messagebox.showerror(
                    "Экспорт в Excel",
                    f"Ошибка загрузки детализированных данных из БД:\n{e}",
                    parent=self,
                )
                return

            if not raw_orders:
                messagebox.showinfo(
                    "Экспорт в Excel",
                    "Нет данных для выгрузки (по заданным фильтрам).",
                    parent=self,
                )
                return
                
            orders: List[Dict[str, Any]] = []
            for o in raw_orders:
                d_str = o.get("date") or ""
                try:
                    d = datetime.strptime(d_str, "%Y-%m-%d").date()
                except Exception:
                    d = None

                if date_from and d and d < date_from:
                    continue
                if date_to and d and d > date_to:
                    continue
                orders.append(o)

            if not orders:
                messagebox.showinfo(
                    "Экспорт в Excel",
                    "Нет данных для выгрузки (по периодам дат).",
                    parent=self,
                )
                return

            price_map = get_meal_type_price_map()
            
            freq: Dict[tuple, int] = {}
            for o in orders:
                fio = (o.get("fio") or "").strip()
                tbn = (o.get("tbn") or "").strip()
                key = (fio.lower(), tbn.lower())
                if fio or tbn:
                    freq[key] = freq.get(key, 0) + 1

            duplicates_mark: List[str] = []
            for o in orders:
                fio = (o.get("fio") or "").strip()
                tbn = (o.get("tbn") or "").strip()
                key = (fio.lower(), tbn.lower())
                mark = "дубль" if (fio or tbn) and freq.get(key, 0) > 1 else ""
                duplicates_mark.append(mark)

            wb = Workbook()
            ws = wb.active
            ws.title = "Реестр питания"

            summary: Dict[str, Dict[str, Dict[str, float]]] = {}

            for o in orders:
                addr = o.get("address", "") or ""
                mt = (o.get("meal_type", "") or "").strip()
                if not addr or not mt:
                    continue
                qty = float(o.get("quantity") or 1)
                price = price_map.get(mt, 0.0)

                addr_dict = summary.setdefault(addr, {})
                mt_dict = addr_dict.setdefault(mt, {"count": 0.0, "amount": 0.0})
                mt_dict["count"] += qty
                mt_dict["amount"] += price * qty

            headers = [
                "Дата",
                "Адрес",
                "ID объекта",
                "Подразделение (из заявки)",
                "Подразделение сотрудника (из 1С)",
                "Наименование бригады",
                "ФИО",
                "Табельный №",
                "Должность",
                "Тип питания",
                "Цена, руб.",
                "Сумма, руб.",
                "Комментарий",
                "Дубликаты",
            ]

            end_col = len(headers)
            # Шапка
            period_str = ""
            if date_from:
                period_str += f"с {date_from.strftime('%Y-%m-%d')} "
            if date_to:
                period_str += f"по {date_to.strftime('%Y-%m-%d')}"
            period_str = period_str.strip() or "все даты"

            dep_str = dep_filter or "Все"
            addr_str = addr_filter or "(любой адрес)"

            ws.append([f"Реестр питания ({period_str})"])
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)

            ws.append([f"Подразделение: {dep_str}"])
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)

            ws.append([f"Адрес содержит: {addr_str}"])
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=end_col)

            ws.append([])

            ws.append(["Свод по объектам, типам питания и стоимости"])
            ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=4)
            ws.append(["Адрес", "Тип питания", "Кол-во человек", "Сумма, руб."])

            for addr, by_type in summary.items():
                for mt, agg in by_type.items():
                    ws.append([
                        addr,
                        mt,
                        agg["count"],
                        agg["amount"],
                    ])

            ws.append([])

            ws.append(headers)

            for order, mark in zip(orders, duplicates_mark):
                mt = (order.get("meal_type") or "").strip()
                qty = float(order.get("quantity") or 1)
                price = float(price_map.get(mt, 0.0))
                amount = price * qty

                ws.append([
                    order.get("date", ""),
                    order.get("address", ""),
                    order.get("object_id", ""),
                    order.get("department", ""),
                    order.get("employee_department", ""),
                    order.get("team_name", ""),
                    order.get("fio", ""),
                    order.get("tbn", ""),
                    order.get("position", ""),
                    mt,
                    price,
                    amount,
                    order.get("comment", ""),
                    mark,
                ])

            widths = [
                12,  # Дата
                40,  # Адрес
                15,  # ID объекта
                25,  # Подразделение
                28,  # Подразделение сотрудника (из 1С)
                25,  # Наименование бригады
                30,  # ФИО
                15,  # Табельный №
                25,  # Должность
                18,  # Тип питания
                12,  # Цена, руб.
                14,  # Сумма, руб.
                40,  # Комментарий
                12,  # Дубликаты
            ]
            for col, width in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(col)].width = width

            ws.freeze_panes = "A7"

            ts = datetime.now().strftime("%Y%m%d_%H%M%S")

            period_part = ""
            if date_from:
                period_part += date_from.strftime("%Y%m%d")
            if date_to:
                period_part += f"-{date_to.strftime('%Y%m%d')}"
            if not period_part:
                period_part = "все"

            fname = f"Реестр_питания_детально_{period_part}_{ts}.xlsx"
            fpath = exe_dir() / "Заявки_питание" / fname
            fpath.parent.mkdir(parents=True, exist_ok=True)

            wb.save(fpath)

            messagebox.showinfo(
                "Экспорт в Excel",
                f"Файл успешно сформирован:\n{fpath}\n\nСтрок (сотрудников): {len(orders)}",
                parent=self,
            )

            try:
                os.startfile(fpath)
            except Exception:
                pass

        except Exception as e:
            messagebox.showerror(
                "Экспорт в Excel",
                f"Ошибка при формировании файла:\n{e}",
                parent=self,
            )

    def _get_selected_order(self) -> Optional[Dict[str, Any]]:
        sel = self.tree.selection()
        if not sel:
            return None
        iid = sel[0]
        try:
            oid = int(iid)
        except Exception:
            return None
        for o in self._orders:
            if int(o["id"]) == oid:
                return o
        return None

    def _on_delete_clicked(self):
        role = self._get_current_role()
        if role != "admin":
            messagebox.showwarning(
                "Удаление заявки",
                "Удалять заявки может только администратор.",
                parent=self,
            )
            return

        info = self._get_selected_order()
        if not info:
            messagebox.showinfo(
                "Удаление заявки",
                "Не выбрана ни одна заявка.",
                parent=self,
            )
            return

        order_id = int(info["id"])
        date_val = info.get("date")
        if isinstance(date_val, datetime):
            date_str = date_val.date().strftime("%Y-%m-%d")
        else:
            date_str = str(date_val or "")

        addr = info.get("object_address") or ""
        dep = info.get("department") or ""
        team = info.get("team_name") or ""

        text = (
            f"Вы действительно хотите ПОЛНОСТЬЮ удалить заявку #{order_id}?\n\n"
            f"Дата: {date_str}\n"
            f"Объект: {addr}\n"
            f"Подразделение: {dep}\n"
            f"Бригада: {team}\n\n"
            f"Будут удалены все сотрудники и сам заголовок заявки.\n"
            f"Отменить это действие будет невозможно."
        )

        if not messagebox.askyesno(
            "Подтверждение удаления заявки",
            text,
            parent=self,
        ):
            return

        try:
            delete_meal_order_from_db(order_id)
        except Exception as e:
            messagebox.showerror(
                "Удаление заявки",
                f"Не удалось удалить заявку #{order_id}:\n{e}",
                parent=self,
            )
            return

        messagebox.showinfo(
            "Удаление заявки",
            f"Заявка #{order_id} успешно удалена.",
            parent=self,
        )

        self._load_data()

    def _on_open(self, event=None):
        info = self._get_selected_order()
        if not info:
            return

        order_id = int(info["id"])

        try:
            order_data = get_order_with_items_from_db(order_id)
        except Exception as e:
            messagebox.showerror(
                "Реестр заявок",
                f"Не удалось загрузить заявку id={order_id}:\n{e}",
                parent=self,
            )
            return

        choice = messagebox.askyesnocancel(
            "Открыть заявку",
            "Нажмите «Да», чтобы ОТКРЫТЬ заявку для РЕДАКТИРОВАНИЯ.\n"
            "Нажмите «Нет», чтобы СОЗДАТЬ КОПИЮ заявки (на другой день).\n"
            "Отмена — закрыть.",
            parent=self,
        )
        if choice is None:
            return

        if choice is False:
            try:
                old_date = datetime.strptime(order_data["date"], "%Y-%m-%d").date()
                new_date = old_date + timedelta(days=1)
                order_data["date"] = new_date.strftime("%Y-%m-%d")
            except Exception:
                pass
            edit_id = None
            title = f"Новая заявка (копия #{order_id})"
        else:
            edit_id = order_id
            title = f"Редактирование заявки #{order_id}"

        win = tk.Toplevel(self)
        win.title(title)
        win.geometry("1300x720")

        if hasattr(self, "app_ref"):
            setattr(win, "app_ref", self.app_ref)

        def on_saved_callback():
            self._load_data()

        page = MealOrderPage(
            win,
            existing_data=order_data,
            order_id=edit_id,
            on_saved=on_saved_callback,
        )
        page.app_ref = self.app_ref
        page.pack(fill="both", expand=True)

class MealPlanningPage(tk.Frame):
    def __init__(self, master):
        super().__init__(master, bg="#f7f7f7")
        self.row_meta: Dict[str, Dict[str, Any]] = {}
        self._build_ui()

    def _build_ui(self):
        top = tk.Frame(self, bg="#f7f7f7")
        top.pack(fill="x", padx=10, pady=8)

        tk.Label(top, text="Дата:", bg="#f7f7f7").grid(row=0, column=0, sticky="w")
        self.ent_filter_date = ttk.Entry(top, width=12)
        self.ent_filter_date.grid(row=0, column=1, padx=4)
        self.ent_filter_date.insert(0, date.today().strftime("%Y-%m-%d"))

        tk.Label(top, text="Подразделение:", bg="#f7f7f7").grid(
            row=0, column=2, sticky="w", padx=(12, 0)
        )

        self.cmb_filter_dep = ttk.Combobox(top, state="readonly", values=["Все"], width=20)
        self.cmb_filter_dep.grid(row=0, column=3, padx=4)
        self.cmb_filter_dep.set("Все")

        tk.Label(top, text="Адрес:", bg="#f7f7f7").grid(
            row=0, column=4, sticky="w", padx=(12, 0)
        )
        self.ent_filter_address = ttk.Entry(top, width=30)
        self.ent_filter_address.grid(row=0, column=5, padx=4)

        ttk.Button(top, text="🔄 Загрузить реестр", command=self.load_registry).grid(
            row=0, column=6, padx=12
        )
        ttk.Button(top, text="📊 Сформировать Excel", command=self.export_to_excel).grid(
            row=0, column=7, padx=4
        )

        ttk.Button(top, text="Заявка поставщику", command=self.export_supplier_order).grid(
            row=0, column=8, padx=4
        )

        table_frame = tk.LabelFrame(self, text="Реестр заказа питания по объектам")
        table_frame.pack(fill="both", expand=True, padx=10, pady=8)

        columns = ("date", "address", "department", "team_name", "total_count", "details")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=15)

        headers = {
            "date": "Дата",
            "address": "Адрес объекта",
            "department": "Подразделение",
            "team_name": "Бригада",
            "total_count": "Всего чел.",
            "details": "По типам питания",
        }
        widths = {
            "date": 90,
            "address": 260,
            "department": 180,
            "team_name": 180,
            "total_count": 90,
            "details": 260,
        }

        for col in columns:
            self.tree.heading(col, text=headers.get(col, col))
            self.tree.column(col, width=widths.get(col, 100))

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        self.tree.bind("<Double-1>", self.on_row_double_click)

    def export_supplier_order(self):
        try:
            filter_date = self.ent_filter_date.get().strip()
            if not filter_date:
                messagebox.showwarning("Заявка поставщика", "Укажите дату фильтра.")
                return

            orders = get_details_from_db(
                filter_date=filter_date or None,
                filter_address=None,
                filter_department=None,
            )

            if not orders:
                messagebox.showinfo(
                    "Заявка поставщика",
                    "Нет данных по указанной дате.",
                )
                return

            total_by_type: Dict[str, int] = {}
            per_object_team_type: Dict[tuple, int] = {}

            for o in orders:
                mt = (o.get("meal_type") or "").strip() or "Не указан"
                addr = (o.get("address") or "").strip()
                team = (o.get("team_name") or "").strip()
                qty = float(o.get("quantity") or 1)

                total_by_type[mt] = total_by_type.get(mt, 0) + qty

                key_row = (addr, team, mt)
                per_object_team_type[key_row] = per_object_team_type.get(key_row, 0) + qty

            wb = Workbook()
            ws = wb.active
            ws.title = "Заявка поставщика"

            ws.append([f"Заявка питания на {filter_date}"])
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

            ws.append([])  # пустая строка

            ws.append(["Итоги по видам питания"])
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=10)
            ws.append(["Тип питания", "Кол-во человек"])
            for mt in sorted(total_by_type.keys()):
                ws.append([mt, total_by_type[mt]])

            ws.append([])

            all_meal_types = sorted({mt for _, _, mt in per_object_team_type.keys()})

            header = ["Объект (адрес)", "Бригада"] + all_meal_types
            ws.append(header)

            row_data: Dict[tuple, Dict[str, int]] = {}
            for (addr, team, mt), cnt in per_object_team_type.items():
                key = (addr, team)
                mt_map = row_data.setdefault(key, {})
                mt_map[mt] = mt_map.get(mt, 0) + cnt

            for (addr, team) in sorted(row_data.keys(), key=lambda x: (x[0], x[1])):
                mt_map = row_data[(addr, team)]
                row = [addr, team]
                for mt in all_meal_types:
                    row.append(mt_map.get(mt, 0))
                ws.append(row)

            ws.column_dimensions[get_column_letter(1)].width = 40  # Объект
            ws.column_dimensions[get_column_letter(2)].width = 30  # Бригада
            for col in range(3, 3 + len(all_meal_types)):
                ws.column_dimensions[get_column_letter(col)].width = 15

            ws.freeze_panes = "A8"

            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            fname = f"Заявка_поставщика_{filter_date}_{ts}.xlsx"
            fpath = exe_dir() / ORDERS_DIR / fname
            fpath.parent.mkdir(parents=True, exist_ok=True)
            wb.save(fpath)

            messagebox.showinfo(
                "Заявка поставщика",
                f"Файл сформирован:\n{fpath}",
            )

            try:
                os.startfile(fpath)
            except Exception:
                pass

        except Exception as e:
            messagebox.showerror(
                "Заявка поставщика",
                f"Ошибка формирования Excel:\n{e}",
            )

    def load_registry(self):
        try:
            filter_date = self.ent_filter_date.get().strip()
            filter_address = self.ent_filter_address.get().strip()
            filter_dep = self.cmb_filter_dep.get().strip()

            registry = get_registry_from_db(
                filter_date=filter_date or None,
                filter_address=filter_address or None,
                filter_department=filter_dep or None,
            )

            self._populate_tree(registry)
            messagebox.showinfo("Загрузка", f"Загружено объектов: {len(registry)}")
        except Exception as e:
            messagebox.showerror(
                "Ошибка", f"Не удалось загрузить реестр из БД:\n{e}"
            )

    def _populate_tree(self, registry: List[Dict]):
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.row_meta = {}

        for entry in registry:
            req_date = entry.get("date", "")
            address = entry.get("address", "")
            department = entry.get("department", "")
            team_name = entry.get("team_name", "")
            total = entry.get("total_count", 0)
            details_text = self._format_details(entry.get("by_meal_type", {}))

            item_id = self.tree.insert(
                "",
                "end",
                values=(req_date, address, department, team_name, total, details_text),
            )
            self.row_meta[item_id] = entry

    def _format_details(self, by_meal_type: Dict[str, int]) -> str:
        if not by_meal_type:
            return "Нет данных"
        parts = [f"{mt}: {cnt}" for mt, cnt in sorted(by_meal_type.items())]
        # ограничим первые 3–4 для компактности
        if len(parts) > 4:
            return " | ".join(parts[:4]) + " ..."
        return " | ".join(parts)

    def on_row_double_click(self, event):
        selection = self.tree.selection()
        if not selection:
            return
        item_id = selection[0]
        entry = self.row_meta.get(item_id)
        if not entry:
            return

        order_ids = entry.get("order_ids") or []
        if not order_ids:
            messagebox.showinfo("Заявка", "Для этого объекта нет связанных заявок.")
            return

        order_id = order_ids[0]

        try:
            order_data = get_order_with_items_from_db(order_id)
        except Exception as e:
            messagebox.showerror(
                "Загрузка заявки",
                f"Не удалось загрузить заявку id={order_id}:\n{e}",
                parent=self,
            )
            return

        win = tk.Toplevel(self)
        win.title(f"Редактирование заявки #{order_id}")
        win.geometry("1300x720")

        def on_saved_callback():
            self.load_registry()

        page = MealOrderPage(win, existing_data=order_data, order_id=order_id, on_saved=on_saved_callback)
        page.pack(fill="both", expand=True)

    def _show_details_dialog(self, entry: Dict):
        dialog = tk.Toplevel(self)
        dialog.title("Детальная информация")
        dialog.geometry("800x600")
        dialog.resizable(True, True)
        dialog.transient(self)
        dialog.grab_set()

        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (800 // 2)
        y = (dialog.winfo_screenheight() // 2) - (600 // 2)
        dialog.geometry(f"800x600+{x}+{y}")

        header = tk.Frame(dialog, bg="#e3f2fd", relief="solid", borderwidth=1)
        header.pack(fill="x", padx=0, pady=0)
        tk.Label(
            header,
            text=f"📅 Дата: {entry.get('date', '')} | 📍 {entry.get('address', '')}",
            font=("Arial", 12, "bold"),
            bg="#e3f2fd",
            fg="#0066cc",
            padx=15,
            pady=12,
        ).pack(anchor="w")

        info_frame = tk.Frame(dialog, bg="#f7f7f7")
        info_frame.pack(fill="x", padx=15, pady=10)
        tk.Label(
            info_frame,
            text=f"Всего заявок: {entry.get('total_count', 0)} человек",
            font=("Arial", 11, "bold"),
            bg="#f7f7f7",
        ).pack(anchor="w")

        table_frame = tk.LabelFrame(
            dialog,
            text="Детализация по подразделениям и типам питания",
            padx=10,
            pady=10,
        )
        table_frame.pack(fill="both", expand=True, padx=15, pady=(0, 10))

        columns = ("department", "meal_type", "count")
        tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=15)
        tree.heading("department", text="Подразделение")
        tree.heading("meal_type", text="Тип питания")
        tree.heading("count", text="Количество")
        tree.column("department", width=300)
        tree.column("meal_type", width=200)
        tree.column("count", width=100)

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        by_dept = entry.get("by_department", {})
        totals_by_type = {}

        for dept, data in sorted(by_dept.items()):
            by_type = data.get("by_meal_type", {})
            for meal_type, count in sorted(by_type.items()):
                tree.insert("", "end", values=(dept, meal_type, count))
                totals_by_type[meal_type] = totals_by_type.get(meal_type, 0) + count

        if totals_by_type:
            tree.insert("", "end", values=("", "", ""), tags=("separator",))
            tree.tag_configure("separator", background="#e0e0e0")
            for meal_type, total in sorted(totals_by_type.items()):
                tree.insert(
                    "", "end", values=("ИТОГО", meal_type, total), tags=("total",)
                )
            tree.tag_configure(
                "total", background="#fff3cd", font=("Arial", 9, "bold")
            )

        ttk.Button(dialog, text="Закрыть", command=dialog.destroy, width=20).pack(
            pady=15
        )

    def export_to_excel(self):
        try:
            filter_date = self.ent_filter_date.get().strip()
            filter_address = self.ent_filter_address.get().strip()
            filter_dep = self.cmb_filter_dep.get().strip()

            orders = get_details_from_db(
                filter_date=filter_date or None,
                filter_address=filter_address or None,
                filter_department=filter_dep or None,
            )

            if not orders:
                messagebox.showinfo(
                    "Экспорт",
                    "Нет данных для экспорта (по заданным фильтрам)",
                )
                return

            price_map = get_meal_type_price_map()

            freq: Dict[tuple, int] = {}
            for o in orders:
                fio = (o.get("fio") or "").strip()
                tbn = (o.get("tbn") or "").strip()
                key = (fio.lower(), tbn.lower())
                if fio or tbn:
                    freq[key] = freq.get(key, 0) + 1

            duplicates_mark: List[str] = []
            for o in orders:
                fio = (o.get("fio") or "").strip()
                tbn = (o.get("tbn") or "").strip()
                key = (fio.lower(), tbn.lower())
                mark = "дубль" if (fio or tbn) and freq.get(key, 0) > 1 else ""
                duplicates_mark.append(mark)

            wb = Workbook()
            ws = wb.active
            ws.title = "Реестр питания"

            summary: Dict[str, Dict[str, Dict[str, float]]] = {}

            for o in orders:
                addr = o.get("address", "") or ""
                mt = (o.get("meal_type", "") or "").strip()
                if not addr or not mt:
                    continue
                qty = float(o.get("quantity") or 1)
                price = price_map.get(mt, 0.0)

                addr_dict = summary.setdefault(addr, {})
                mt_dict = addr_dict.setdefault(mt, {"count": 0.0, "amount": 0.0})
                mt_dict["count"] += qty
                mt_dict["amount"] += price * qty

            ws.append(["Свод по объектам, типам питания и стоимости"])
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
            ws.append(["Адрес", "Тип питания", "Кол-во человек", "Сумма, руб."])

            for addr, by_type in summary.items():
                for mt, agg in by_type.items():
                    ws.append([
                        addr,
                        mt,
                        agg["count"],
                        agg["amount"],
                    ])

            ws.append([])

            headers = [
                "Дата",
                "Адрес",
                "ID объекта",
                "Подразделение",
                "Наименование бригады",
                "ФИО",
                "Табельный №",
                "Должность",
                "Тип питания",
                "Цена, руб.",
                "Сумма, руб.",
                "Комментарий",
                "Дубликаты",
            ]
            ws.append(headers)

            for order, mark in zip(orders, duplicates_mark):
                mt = (order.get("meal_type") or "").strip()
                qty = float(order.get("quantity") or 1)
                price = float(price_map.get(mt, 0.0))
                amount = price * qty

                ws.append([
                    order.get("date", ""),
                    order.get("address", ""),
                    order.get("object_id", ""),
                    order.get("department", ""),
                    order.get("team_name", ""),
                    order.get("fio", ""),
                    order.get("tbn", ""),
                    order.get("position", ""),
                    mt,
                    price,
                    amount,
                    order.get("comment", ""),
                    mark,
                ])

            widths = [
                12,  # Дата
                40,  # Адрес
                15,  # ID объекта
                25,  # Подразделение
                25,  # Наименование бригады
                30,  # ФИО
                15,  # Табельный №
                25,  # Должность
                18,  # Тип питания
                12,  # Цена, руб.
                14,  # Сумма, руб.
                40,  # Комментарий
                12,  # Дубликаты
            ]
            for col, width in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(col)].width = width

            ws.freeze_panes = "A5"

            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            fname = f"Реестр_питания_{filter_date or 'все'}_{ts}.xlsx"
            fpath = exe_dir() / ORDERS_DIR / fname
            fpath.parent.mkdir(parents=True, exist_ok=True)

            wb.save(fpath)
            messagebox.showinfo(
                "Экспорт",
                f"Реестр успешно сформирован:\n{fpath}\n\nЗаписей: {len(orders)}",
            )

            try:
                os.startfile(fpath)
            except Exception:
                pass

        except Exception as e:
            messagebox.showerror(
                "Ошибка",
                f"Не удалось сформировать реестр из БД:\n{e}",
            )

class MealsSettingsPage(tk.Frame):
    def __init__(self, master):
        super().__init__(master, bg="#f7f7f7")
        self.name_vars: List[tk.StringVar] = []
        self.price_vars: List[tk.StringVar] = []

        # --- новые переменные для настроек приёма заявок ---
        self.var_limit_next_day_only = tk.BooleanVar(value=True)
        self.var_deadline_time = tk.StringVar(value="14:00")

        self._build_ui()
        self.load_meal_types()
        self.load_meals_order_rules()  # <<< новая загрузка настроек

    def _build_ui(self):
        top = tk.Frame(self, bg="#f7f7f7")
        top.pack(fill="x", padx=10, pady=10)

        tk.Label(
            top,
            text="Настройки типов питания",
            font=("Arial", 12, "bold"),
            bg="#f7f7f7",
        ).grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 8))

        hdr = tk.Frame(self, bg="#f7f7f7")
        hdr.pack(fill="x", padx=10)
        tk.Label(hdr, text="Название типа питания", bg="#f7f7f7").grid(
            row=0, column=0, sticky="w", padx=4
        )
        tk.Label(hdr, text="Цена, руб.", bg="#f7f7f7").grid(
            row=0, column=1, sticky="w", padx=4
        )

        self.rows_frame = tk.Frame(self, bg="#f7f7f7")
        self.rows_frame.pack(fill="x", padx=10, pady=4)

        for i in range(3):
            nv = tk.StringVar()
            pv = tk.StringVar()
            self.name_vars.append(nv)
            self.price_vars.append(pv)
            ttk.Entry(self.rows_frame, textvariable=nv, width=30).grid(
                row=i, column=0, padx=4, pady=2, sticky="w"
            )
            ttk.Entry(self.rows_frame, textvariable=pv, width=10).grid(
                row=i, column=1, padx=4, pady=2, sticky="w"
            )

        btns = tk.Frame(self, bg="#f7f7f7")
        btns.pack(fill="x", padx=10, pady=(8, 10))
        ttk.Button(
            btns, text="Сохранить типы питания", command=self.save_meal_types
        ).pack(side="left", padx=4)
        rules_frame = tk.LabelFrame(
            self,
            text="Правила приёма заявок",
            bg="#f7f7f7",
            padx=10,
            pady=8,
        )
        rules_frame.pack(fill="x", padx=10, pady=(0, 10))

        ttk.Checkbutton(
            rules_frame,
            text="Разрешать заявки только на следующий день",
            variable=self.var_limit_next_day_only,
        ).grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 4))

        tk.Label(
            rules_frame,
            text="Время окончания приёма заявок на завтра (HH:MM):",
            bg="#f7f7f7",
        ).grid(row=1, column=0, sticky="w", pady=(4, 0))

        ttk.Entry(
            rules_frame,
            textvariable=self.var_deadline_time,
            width=10,
        ).grid(row=1, column=1, sticky="w", padx=(4, 0), pady=(4, 0))

        ttk.Label(
            rules_frame,
            text="(пусто — без ограничения по времени)",
            background="#f7f7f7",
        ).grid(row=1, column=2, sticky="w", padx=(6, 0), pady=(4, 0))

        btns2 = tk.Frame(self, bg="#f7f7f7")
        btns2.pack(fill="x", padx=10, pady=(0, 10))
        ttk.Button(
            btns2,
            text="Сохранить правила приёма заявок",
            command=self.save_meals_order_rules,
        ).pack(side="left", padx=4)

    def load_meals_order_rules(self):
        try:
            limit = get_meals_limit_next_day_only()
            deadline = get_meals_next_day_deadline()

            self.var_limit_next_day_only.set(bool(limit))
            if deadline is not None:
                self.var_deadline_time.set(deadline.strftime("%H:%M"))
            else:
                self.var_deadline_time.set("")
        except Exception as e:
            messagebox.showerror(
                "Настройки питания",
                f"Ошибка загрузки правил приёма заявок:\n{e}",
                parent=self,
            )

    def save_meals_order_rules(self):
        t_str = (self.var_deadline_time.get() or "").strip()
        if t_str:
            try:
                # допускаем только HH:MM
                dt = datetime.strptime(t_str, "%H:%M")
                t_str_norm = dt.strftime("%H:%M")
            except Exception:
                messagebox.showwarning(
                    "Настройки питания",
                    "Время окончания приёма заявок должно быть в формате HH:MM, "
                    "например '14:00', или оставьте поле пустым.",
                    parent=self,
                )
                return
        else:
            t_str_norm = ""

        try:
            set_setting("meals_limit_next_day_only",
                        "1" if self.var_limit_next_day_only.get() else "0")
            set_setting("meals_next_day_deadline", t_str_norm)

            messagebox.showinfo(
                "Настройки питания",
                "Правила приёма заявок сохранены.",
                parent=self,
            )
        except Exception as e:
            messagebox.showerror(
                "Настройки питания",
                f"Ошибка сохранения правил приёма заявок:\n{e}",
                parent=self,
            )

    def load_meal_types(self):
        try:
            mts = load_meal_types_from_db()
        except Exception as e:
            messagebox.showerror(
                "Настройки питания",
                f"Ошибка загрузки типов питания:\n{e}",
                parent=self,
            )
            return

        for i in range(3):
            if i < len(mts):
                self.name_vars[i].set(mts[i]["name"] or "")
                self.price_vars[i].set(f'{mts[i]["price"]:.2f}')
            else:
                self.name_vars[i].set("")
                self.price_vars[i].set("0.00")

    def save_meal_types(self):
        data: List[Tuple[str, float]] = []
        for nv, pv in zip(self.name_vars, self.price_vars):
            name = (nv.get() or "").strip()
            if not name:
                continue
            p_str = (pv.get() or "0").replace(",", ".")
            try:
                price = float(p_str)
            except Exception:
                messagebox.showwarning(
                    "Настройки питания",
                    f"Цена должна быть числом: '{p_str}'",
                    parent=self,
                )
                return
            data.append((name, price))

        if not data:
            messagebox.showwarning(
                "Настройки питания",
                "Укажите хотя бы один тип питания.",
                parent=self,
            )
            return

        conn = None
        try:
            conn = get_db_connection()
            with conn:
                with conn.cursor(cursor_factory=RealDictCursor) as cur:
                    # читаем текущие типы
                    cur.execute("SELECT id, name, price FROM meal_types;")
                    existing = list(cur.fetchall())

                    by_name: Dict[str, Dict[str, Any]] = {
                        (row["name"] or "").strip().lower(): row for row in existing
                    }

                    for name, price in data:
                        key = name.strip().lower()
                        row = by_name.get(key)
                        if row:
                            cur.execute(
                                "UPDATE meal_types SET price = %s WHERE id = %s",
                                (price, row["id"]),
                            )
                        else:
                            cur.execute(
                                "INSERT INTO meal_types (name, price) VALUES (%s, %s)",
                                (name, price),
                            )

            messagebox.showinfo(
                "Настройки питания",
                "Типы питания и цены сохранены.",
                parent=self,
            )
            self.load_meal_types()

        except Exception as e:
            messagebox.showerror(
                "Настройки питания",
                f"Ошибка сохранения:\n{e}",
                parent=self,
            )
        finally:
            if conn:
                release_db_connection(conn)

class MealsApp(tk.Tk):

    def __init__(self, current_user_role: str = "user"):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1100x720")
        self.resizable(True, True)

        notebook = ttk.Notebook(self)
        notebook.pack(fill="both", expand=True)

        order_page = MealOrderPage(notebook)
        notebook.add(order_page, text="Создать заявку")

        if get_meals_planning_enabled():
            planning_page = MealPlanningPage(notebook)
            notebook.add(planning_page, text="Планирование питания")

        if current_user_role == "admin":
            settings_page = MealsSettingsPage(notebook)
            notebook.add(settings_page, text="Настройки")

    def destroy(self):
        """Переопределяем для закрытия локального пула."""
        global db_connection_pool, USING_SHARED_POOL
        if not USING_SHARED_POOL and db_connection_pool:
            print("Closing local DB connection pool for meals_module...")
            db_connection_pool.closeall()
            db_connection_pool = None
        super().destroy()

def get_order_with_items_from_db(order_id: int) -> Dict[str, Any]:
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT
                    mo.id,
                    mo.created_at,
                    mo.date::text,
                    COALESCE(d.name, '') AS department,
                    COALESCE(mo.team_name, '') AS team_name,
                    COALESCE(o.excel_id, '') AS object_excel_id,
                    COALESCE(o.address, '') AS object_address,
                    COALESCE(mo.fact_address, '') AS fact_address
                FROM meal_orders mo
                LEFT JOIN departments d ON d.id = mo.department_id
                LEFT JOIN objects o     ON o.id = mo.object_id
                WHERE mo.id = %s
                """,
                (order_id,),
            )
            row = cur.fetchone()
            if not row:
                raise ValueError(f"Заявка id={order_id} не найдена")

            (oid, created_at, date_str,
             department, team_name, obj_excel_id, obj_address, fact_address) = row

            cur.execute(
                """
                SELECT
                    COALESCE(moi.fio_text, '')      AS fio,
                    COALESCE(moi.tbn_text, '')      AS tbn,
                    COALESCE(moi.position_text, '') AS position,
                    COALESCE(mti.name, moi.meal_type_text, '') AS meal_type,
                    COALESCE(moi.comment, '')       AS comment,
                    COALESCE(moi.quantity, 1)       AS quantity
                FROM meal_order_items moi
                LEFT JOIN meal_types mti ON mti.id = moi.meal_type_id
                WHERE moi.order_id = %s
                ORDER BY moi.fio_text
                """,
                (order_id,),
            )
            emps = []
            for fio, tbn, position, meal_type, comment, quantity in cur.fetchall():
                emps.append(
                    {
                        "fio": fio,
                        "tbn": tbn,
                        "position": position,
                        "meal_type": meal_type,
                        "comment": comment,
                        "quantity": float(quantity or 1),
                    }
                )

        return {
            "id": oid,
            "created_at": created_at.strftime("%Y-%m-%dT%H:%M:%S"),
            "date": date_str,
            "department": department,
            "team_name": team_name,
            "object": {"id": obj_excel_id, "address": obj_address},
            "fact_address": fact_address,
            "employees": emps,
        }
    finally:
        if conn:
            release_db_connection(conn)

def create_meals_order_page(parent, app_ref=None) -> tk.Frame:
    ensure_config()
    try:
        page = MealOrderPage(parent)
        page.app_ref = app_ref
        return page
    except Exception:
        import traceback
        messagebox.showerror(
            "Питание — ошибка", traceback.format_exc(), parent=parent
        )
        return tk.Frame(parent)

def create_all_meals_orders_page(parent, app_ref=None) -> tk.Frame:
    ensure_config()
    try:
        page = AllMealsOrdersPage(parent, app_ref=app_ref)
        return page
    except Exception:
        import traceback
        messagebox.showerror(
            "Питание — реестр заявок", traceback.format_exc(), parent=parent
        )
        return tk.Frame(parent)

def create_my_meals_orders_page(parent, app_ref=None) -> tk.Frame:
    ensure_config()
    try:
        page = MyMealsOrdersPage(parent, app_ref=app_ref)
        return page
    except Exception:
        import traceback
        messagebox.showerror(
            "Питание — мои заявки", traceback.format_exc(), parent=parent
        )
        return tk.Frame(parent)

def create_meals_planning_page(parent, app_ref=None) -> tk.Frame:
    ensure_config()
    try:
        page = MealPlanningPage(parent)
        page.app_ref = app_ref
        return page
    except Exception:
        import traceback
        messagebox.showerror(
            "Планирование питания — ошибка", traceback.format_exc(), parent=parent
        )
        return tk.Frame(parent)

def delete_order_items_from_db(order_id: int):
    conn = None
    try:
        conn = get_db_connection()
        with conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM meal_order_items WHERE order_id = %s", (order_id,))
    finally:
        if conn:
            release_db_connection(conn)

def delete_meal_order_from_db(order_id: int):
    conn = None
    try:
        conn = get_db_connection()
        with conn:
            with conn.cursor() as cur:
                cur.execute(
                    "DELETE FROM meal_order_items WHERE order_id = %s",
                    (order_id,),
                )
                cur.execute(
                    "DELETE FROM meal_orders WHERE id = %s",
                    (order_id,),
                )
    finally:
        if conn:
            release_db_connection(conn)

def create_meals_settings_page(parent, current_user_role: str) -> Optional[tk.Frame]:
    if current_user_role != "admin":
        return None
    ensure_config()
    try:
        return MealsSettingsPage(parent)
    except Exception:
        import traceback
        messagebox.showerror(
            "Настройки питания — ошибка", traceback.format_exc(), parent=parent
        )
        return tk.Frame(parent)

def open_meals_module(parent=None, current_user_role: str = "user"):
    if parent is None:
        app = MealsApp(current_user_role=current_user_role)
        app.mainloop()
        return app

    win = tk.Toplevel(parent)
    win.title(APP_TITLE)
    win.geometry("1100x720")

    notebook = ttk.Notebook(win)
    notebook.pack(fill="both", expand=True)

    order_page = MealOrderPage(notebook)
    notebook.add(order_page, text="Создать заявку")

    if get_meals_planning_enabled():
        planning_page = MealPlanningPage(notebook)
        notebook.add(planning_page, text="Планирование питания")

    if current_user_role == "admin":
        settings_page = MealsSettingsPage(notebook)
        notebook.add(settings_page, text="Настройки")

    return win

if __name__ == "__main__":
    ensure_config()
    try:
        conn = get_db_connection()
        release_db_connection(conn)
    except Exception as e:
        messagebox.showerror(
            "Критическая ошибка",
            f"Не удалось подключиться к базе данных:\n{e}"
        )
        sys.exit(1)

    app = MealsApp(current_user_role="admin")
    app.mainloop()
