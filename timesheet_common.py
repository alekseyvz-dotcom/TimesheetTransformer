from __future__ import annotations

import calendar
import difflib
import logging
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

MAX_HOURS_PER_DAY = 24

TS_COLORS = {
    "bg": "#f0f2f5",
    "panel": "#ffffff",
    "accent": "#1565c0",
    "accent_light": "#e3f2fd",
    "success": "#2e7d32",
    "warning": "#b00020",
    "border": "#dde1e7",
    "btn_save_bg": "#1565c0",
    "btn_save_fg": "#ffffff",
    "suspicious": "#FF6B6B",
    "suspicious_fg": "#FFFFFF",
}

# Единый справочник поддерживаемых кодов.
# Если позже потребуется поменять бизнес-логику по конкретному коду,
# достаточно будет исправить этот словарь и/или parse_timesheet_cell().
SPECIAL_CODES = {
    "Б": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Больничный"},
    "Т": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Больничный неоплачиваемый"},
    "ВМ": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Вахта"},
    "ВЧ": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Вечерние часы"},
    "ПВ": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Время вынужденного прогула"},
    "РП": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Время простоя по вине работодателя"},
    "Г": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Выполнение государственных обязанностей"},
    "В": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Выходной"},
    "ДВ": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Дни в пути (вахта)"},
    "ДБ": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Доп. отпуск без сохранения заработной платы"},
    "НВ": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Доп. выходной без оплаты"},
    "ЕО": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Дополнительные выходные дни (неоплачиваемые)"},
    "ОВ": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Отгул донорский"},
    "ОД": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Дополнительный отпуск"},
    "ЗБ": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Забастовка"},
    "К": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Командировка"},
    "МО": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Междувахтовый отдых"},
    "НН": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Неявка"},
    "Н": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Ночные часы"},
    "ОТ": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Отпуск"},
    "УД": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Отпуск дополнительный (неоплачиваемый учебный)"},
    "У": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Отпуск дополнительный (оплачиваемый учебный)"},
    "ОЗ": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Отпуск неоплачиваемый в соответствии с законом"},
    "ДО": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Отпуск без сохранения"},
    "Р": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Отпуск по беременности и родам"},
    "ОЖ": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Отпуск по уходу за ребенком"},
    "НБ": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Отстранение от работы без оплаты"},
    "НО": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Отстранение от работы с оплатой"},
    "КР": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Перерывы для кормления ребенка"},
    "ПК": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Повышение квалификации"},
    "ПМ": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Повышение квалификации в другой местности"},
    "РВ": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Праздники"},
    "РВВ": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Праздники (вечернее время)"},
    "РВН": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Праздники (ночное время)"},
    "ПН": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Праздники без повышенной оплаты"},
    "ПНВ": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Праздники без повышенной оплаты (вечернее время)"},
    "ПНН": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Праздники без повышенной оплаты (ночное время)"},
    "НЗ": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Приостановка работы в случае задержки выплаты з/п"},
    "ПТД": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Приостановление трудового договора"},
    "ПР": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Прогул"},
    "ВП": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Простой по вине работника"},
    "НП": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Простой, не зависящий от работодателя и работника"},
    "НРВ": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Работа в выходные и праздники (ночное время)"},
    "НС": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Работа в режиме неполного рабочего времени"},
    "С": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Сверхурочно"},
    "СВВ": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Сверхурочно (вечернее время)"},
    "СВН": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Сверхурочно (ночное время)"},
    "СН": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Сверхурочные без повышенной оплаты"},
    "СНВ": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Сверхурочные без повышенной оплаты (вечер. время)"},
    "СНН": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Сверхурочные без повышенной оплаты (ночное время)"},
    "УВ": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Уволен"},
    "ЛЧ": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Сокращенное рабочее время в соответствии с законом"},
    "Я": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Явка"},

    # твои кастомные коды
    "О": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Отсутствие"},
    "П": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Простой / прочее"},
    "КВ": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Командировка выходного дня"},
    "снег": {"hours": 0.0, "night_hours": 0.0, "counts_day": False, "description": "Снег"},
}
    # Для РВ логичнее учитывать часы как фактические часы работы.
    # При необходимости можно перенастроить.
    "РВ 8": {"hours": 8.0, "night_hours": 0.0, "counts_day": True, "description": "Работа в выходной 8 ч"},
    "РВ 11": {"hours": 11.0, "night_hours": 0.0, "counts_day": True, "description": "Работа в выходной 11 ч"},
}


@dataclass(frozen=True)
class ParsedTimesheetCell:
    raw: str
    normalized: str
    is_empty: bool
    is_code: bool
    code: Optional[str]
    total_hours: Optional[float]
    night_hours: float
    overtime_day: float
    overtime_night: float
    counts_day: bool
    suspicious: bool


def exe_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def month_days(year: int, month: int) -> int:
    return calendar.monthrange(year, month)[1]


def month_name_ru(month: int) -> str:
    names = [
        "Январь",
        "Февраль",
        "Март",
        "Апрель",
        "Май",
        "Июнь",
        "Июль",
        "Август",
        "Сентябрь",
        "Октябрь",
        "Ноябрь",
        "Декабрь",
    ]
    if 1 <= month <= 12:
        return names[month - 1]
    return str(month)


def safe_filename(value: str, maxlen: int = 60) -> str:
    s = re.sub(r'[<>:"/\\|?*\n\r\t]+', "_", str(value or "")).strip()
    s = re.sub(r"_+", "_", s)
    return s[:maxlen] if maxlen > 0 else s


def normalize_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def normalize_tbn(value: Any) -> str:
    return normalize_spaces(str(value or ""))


def normalize_code(value: str) -> str:
    s = normalize_spaces(str(value or "").upper())
    s = s.replace("РВ8", "РВ 8").replace("РВ11", "РВ 11")
    s = re.sub(r"^РВ\s+0*8$", "РВ 8", s)
    s = re.sub(r"^РВ\s+0*11$", "РВ 11", s)
    return s


def is_allowed_timesheet_code(value: str) -> bool:
    s = normalize_code(value)
    if s in SPECIAL_CODES:
        return True

    # Разрешаем общий формат "РВ <число>", если потребуется в будущем.
    # Если хочешь оставить строго только РВ 8 / РВ 11, этот блок можно удалить.
    if s.startswith("РВ "):
        tail = s[3:].strip()
        try:
            number = float(tail.replace(",", "."))
            return number > 0
        except Exception:
            return False

    return False


def format_hours_for_cell(value: float | int | None) -> Optional[str]:
    if value is None:
        return None
    v = float(value)
    if abs(v) < 1e-12:
        return "0"
    return f"{v:.2f}".rstrip("0").rstrip(".").replace(".", ",")


def _to_float_number(value: str) -> Optional[float]:
    try:
        return float(value.replace(",", "."))
    except Exception:
        return None


def _parse_time_token_to_hours(token: str) -> Optional[float]:
    token = normalize_spaces(token)
    if not token:
        return None

    if ":" in token:
        parts = token.split(":")
        if len(parts) > 2:
            return None
        hh = _to_float_number(parts[0].strip())
        mm_raw = parts[1].strip() if len(parts) == 2 else "0"
        mm = _to_float_number(mm_raw)
        if hh is None or mm is None:
            return None
        if mm < 0 or mm >= 60:
            return None
        return hh + mm / 60.0

    return _to_float_number(token)


def parse_hours_value(value: Any) -> Optional[float]:
    """
    Парсит только числовой формат часов:
      8
      8,25
      8:30
      1/7
      8/2
      8/2/1

    ВАЖНО:
    - буквенные коды здесь не поддерживаются;
    - часть в скобках (...) игнорируется.
    """
    s = normalize_spaces(str(value or ""))
    if not s:
        return None

    if "(" in s:
        s = s.split("(", 1)[0].strip()
    if not s:
        return None

    if "/" in s:
        total = 0.0
        found_any = False
        for part in s.split("/"):
            hours = _parse_time_token_to_hours(part.strip())
            if hours is None:
                return None
            total += float(hours)
            found_any = True
        return total if found_any else None

    return _parse_time_token_to_hours(s)


def parse_overtime(value: Any) -> Tuple[Optional[float], Optional[float]]:
    """
    Парсинг переработки из скобок:
      8/2(1/1) -> (1.0, 1.0)
      8(2)     -> (2.0, 0.0)
    """
    s = normalize_spaces(str(value or ""))
    if "(" not in s or ")" not in s:
        return None, None

    try:
        overtime_str = s[s.index("(") + 1 : s.index(")")].strip()
    except ValueError:
        return None, None

    if not overtime_str:
        return None, None

    try:
        if "/" in overtime_str:
            parts = [p.strip() for p in overtime_str.split("/")]
            day_ot = _to_float_number(parts[0]) if len(parts) >= 1 and parts[0] else 0.0
            night_ot = _to_float_number(parts[1]) if len(parts) >= 2 and parts[1] else 0.0
            if day_ot is None or night_ot is None:
                return None, None
            return float(day_ot), float(night_ot)

        single = _to_float_number(overtime_str)
        if single is None:
            return None, None
        return float(single), 0.0
    except Exception:
        return None, None


def parse_hours_and_night(value: Any) -> Tuple[Optional[float], Optional[float]]:
    """
    Старый совместимый интерфейс:
      "8/2" -> total=10, night=2
      "8"   -> total=8, night=0
      "РВ 8" -> total=8, night=0
      "ОТ"   -> total=0, night=0
    """
    parsed = parse_timesheet_cell(value)
    if parsed.is_empty:
        return None, None
    if parsed.total_hours is None:
        return None, None
    return parsed.total_hours, parsed.night_hours


def parse_timesheet_cell(value: Any) -> ParsedTimesheetCell:
    raw = str(value or "")
    normalized = normalize_spaces(raw)

    if not normalized:
        return ParsedTimesheetCell(
            raw=raw,
            normalized="",
            is_empty=True,
            is_code=False,
            code=None,
            total_hours=None,
            night_hours=0.0,
            overtime_day=0.0,
            overtime_night=0.0,
            counts_day=False,
            suspicious=False,
        )

    code = normalize_code(normalized)
    if code in SPECIAL_CODES:
        info = SPECIAL_CODES[code]
        total_hours = float(info.get("hours", 0.0))
        night_hours = float(info.get("night_hours", 0.0))
        counts_day = bool(info.get("counts_day", total_hours > 1e-12))
        return ParsedTimesheetCell(
            raw=raw,
            normalized=code,
            is_empty=False,
            is_code=True,
            code=code,
            total_hours=total_hours,
            night_hours=night_hours,
            overtime_day=0.0,
            overtime_night=0.0,
            counts_day=counts_day,
            suspicious=False,
        )

    # Поддержка общего формата "РВ <число>"
    if code.startswith("РВ "):
        tail = code[3:].strip()
        generic_rv = _to_float_number(tail)
        if generic_rv is not None and generic_rv > 0:
            return ParsedTimesheetCell(
                raw=raw,
                normalized=code,
                is_empty=False,
                is_code=True,
                code=code,
                total_hours=float(generic_rv),
                night_hours=0.0,
                overtime_day=0.0,
                overtime_night=0.0,
                counts_day=True,
                suspicious=False,
            )

    base_part = normalized
    if "(" in base_part:
        base_part = base_part.split("(", 1)[0].strip()

    total_hours: Optional[float] = None
    night_hours = 0.0

    if "/" in base_part:
        parts = [p.strip() for p in base_part.split("/") if p.strip()]
        if parts:
            base = _parse_time_token_to_hours(parts[0])
            if base is not None:
                total = float(base)
                for night_part in parts[1:]:
                    nv = _parse_time_token_to_hours(night_part)
                    if nv is None:
                        total = None
                        night_hours = 0.0
                        break
                    total += float(nv)
                    night_hours += float(nv)
                total_hours = total
    else:
        total_hours = parse_hours_value(base_part)
        night_hours = 0.0

    ot_day, ot_night = parse_overtime(normalized)
    overtime_day = float(ot_day) if isinstance(ot_day, (int, float)) else 0.0
    overtime_night = float(ot_night) if isinstance(ot_night, (int, float)) else 0.0

    suspicious = False
    counts_day = False
    if isinstance(total_hours, (int, float)):
        suspicious = float(total_hours) > MAX_HOURS_PER_DAY
        counts_day = float(total_hours) > 1e-12

    return ParsedTimesheetCell(
        raw=raw,
        normalized=normalized,
        is_empty=False,
        is_code=False,
        code=None,
        total_hours=float(total_hours) if isinstance(total_hours, (int, float)) else None,
        night_hours=float(night_hours),
        overtime_day=overtime_day,
        overtime_night=overtime_night,
        counts_day=counts_day,
        suspicious=suspicious,
    )


def is_suspicious_hours(raw_value: Any) -> bool:
    parsed = parse_timesheet_cell(raw_value)
    return bool(parsed.suspicious)


def normalize_hours_list(
    hours_list: Sequence[Any] | None,
    year: int,
    month: int,
) -> list[Optional[str]]:
    """
    Нормализует массив часов:
    - всегда длина 31;
    - все значения str|None;
    - дни после конца месяца зануляются.
    """
    out: list[Optional[str]] = []
    for item in list(hours_list or [])[:31]:
        if item in (None, ""):
            out.append(None)
        else:
            text = normalize_spaces(str(item))
            out.append(text or None)

    if len(out) < 31:
        out.extend([None] * (31 - len(out)))

    days_in_m = month_days(year, month)
    for idx in range(days_in_m, 31):
        out[idx] = None

    return out


def calc_row_totals(hours_list: Sequence[Any] | None, year: int, month: int) -> Dict[str, Any]:
    normalized = normalize_hours_list(hours_list, year, month)
    days_in_m = month_days(year, month)

    total_hours = 0.0
    total_days = 0
    total_night = 0.0
    total_ot_day = 0.0
    total_ot_night = 0.0

    for i in range(days_in_m):
        raw = normalized[i]
        if not raw:
            continue

        parsed = parse_timesheet_cell(raw)

        if isinstance(parsed.total_hours, (int, float)):
            total_hours += float(parsed.total_hours)
        total_night += float(parsed.night_hours)
        total_ot_day += float(parsed.overtime_day)
        total_ot_night += float(parsed.overtime_night)

        if parsed.counts_day:
            total_days += 1

    return {
        "days": total_days,
        "hours": float(f"{total_hours:.2f}"),
        "night_hours": float(f"{total_night:.2f}"),
        "ot_day": float(f"{total_ot_day:.2f}"),
        "ot_night": float(f"{total_ot_night:.2f}"),
    }


def calc_rows_summary(rows: Sequence[Mapping[str, Any]], year: int, month: int) -> Dict[str, Any]:
    total_hours = 0.0
    total_days = 0
    total_night = 0.0
    total_ot_day = 0.0
    total_ot_night = 0.0

    for rec in rows:
        totals = rec.get("_totals")
        if not isinstance(totals, dict):
            totals = calc_row_totals(rec.get("hours"), year, month)

        total_days += int(totals.get("days") or 0)
        total_hours += float(totals.get("hours") or 0.0)
        total_night += float(totals.get("night_hours") or 0.0)
        total_ot_day += float(totals.get("ot_day") or 0.0)
        total_ot_night += float(totals.get("ot_night") or 0.0)

    return {
        "employees": len(rows),
        "days": total_days,
        "hours": float(f"{total_hours:.2f}"),
        "night_hours": float(f"{total_night:.2f}"),
        "ot_day": float(f"{total_ot_day:.2f}"),
        "ot_night": float(f"{total_ot_night:.2f}"),
    }


def format_summary_value(value: float | int) -> str:
    return f"{float(value):.2f}".rstrip("0").rstrip(".")


def make_row_key(fio: str, tbn: str) -> tuple[str, str]:
    return normalize_spaces(fio).lower(), normalize_tbn(tbn)


def normalize_row_record(
    record: Mapping[str, Any],
    year: int,
    month: int,
) -> Dict[str, Any]:
    fio = normalize_spaces(str(record.get("fio") or ""))
    tbn = normalize_tbn(record.get("tbn"))
    hours_source = record.get("hours")
    if hours_source is None:
        hours_source = record.get("hours_raw")

    hours = normalize_hours_list(hours_source, year, month)
    totals = calc_row_totals(hours, year, month)

    return {
        "fio": fio,
        "tbn": tbn,
        "hours": hours,
        "_totals": totals,
    }


def deduplicate_timesheet_rows(
    rows: Sequence[Mapping[str, Any]],
    year: int,
    month: int,
) -> List[Dict[str, Any]]:
    """
    Удаляет дубли по ключу (fio.lower(), tbn).
    Последняя запись побеждает.
    """
    uniq: dict[tuple[str, str], Dict[str, Any]] = {}
    for rec in rows:
        normalized = normalize_row_record(rec, year, month)
        key = make_row_key(normalized["fio"], normalized["tbn"])
        if not key[0] and not key[1]:
            continue
        uniq[key] = normalized
    return list(uniq.values())


def find_suspicious_cells(
    rows: Sequence[Mapping[str, Any]],
    year: int,
    month: int,
) -> List[Dict[str, Any]]:
    days_in_m = month_days(year, month)
    suspicious: List[Dict[str, Any]] = []

    for row_idx, rec in enumerate(rows):
        fio = str(rec.get("fio") or "")
        tbn = str(rec.get("tbn") or "")
        hours_list = normalize_hours_list(rec.get("hours"), year, month)

        for day_idx in range(days_in_m):
            raw = hours_list[day_idx]
            if not raw:
                continue

            parsed = parse_timesheet_cell(raw)
            if parsed.suspicious:
                suspicious.append(
                    {
                        "row_idx": row_idx,
                        "day": day_idx + 1,
                        "fio": fio,
                        "tbn": tbn,
                        "raw": str(raw),
                        "parsed": parsed.total_hours,
                    }
                )

    return suspicious


def validate_row_record(
    record: Mapping[str, Any],
    year: int,
    month: int,
) -> List[str]:
    errors: List[str] = []
    fio = normalize_spaces(str(record.get("fio") or ""))
    tbn = normalize_tbn(record.get("tbn"))
    hours = normalize_hours_list(record.get("hours"), year, month)
    days_in_m = month_days(year, month)

    if not fio and not tbn:
        errors.append("Строка сотрудника не содержит ни ФИО, ни табельного номера.")

    for idx in range(days_in_m):
        raw = hours[idx]
        if not raw:
            continue

        parsed = parse_timesheet_cell(raw)
        if parsed.is_empty:
            continue

        if parsed.is_code:
            continue

        numeric_ok = parsed.total_hours is not None
        has_overtime = parsed.overtime_day > 0 or parsed.overtime_night > 0
        if not numeric_ok and not has_overtime:
            errors.append(
                f"Некорректное значение в дне {idx + 1}: '{raw}'"
            )

    return errors


def validate_rows_before_save(
    rows: Sequence[Mapping[str, Any]],
    year: int,
    month: int,
) -> List[str]:
    errors: List[str] = []
    seen: set[tuple[str, str]] = set()

    for index, rec in enumerate(rows, start=1):
        normalized = normalize_row_record(rec, year, month)
        row_key = make_row_key(normalized["fio"], normalized["tbn"])

        if row_key in seen and (row_key[0] or row_key[1]):
            errors.append(
                f"Дублирующаяся строка в текущем табеле: "
                f"{normalized['fio']} (таб.№ {normalized['tbn']})"
            )
        seen.add(row_key)

        row_errors = validate_row_record(normalized, year, month)
        for msg in row_errors:
            errors.append(f"Строка {index}: {msg}")

    return errors


def _norm_fio(value: str) -> str:
    s = normalize_spaces(value).lower()
    s = s.replace("ё", "е")
    s = re.sub(r"[.\t\r\n]+", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def best_fio_match_with_score(skud_fio: str, candidates: Sequence[str]) -> Tuple[Optional[str], float]:
    nf = _norm_fio(skud_fio)
    if not nf:
        return None, 0.0

    best_name = None
    best_score = 0.0
    for candidate in candidates:
        nc = _norm_fio(candidate)
        if not nc:
            continue
        score = difflib.SequenceMatcher(None, nf, nc).ratio()
        if score > best_score:
            best_score = score
            best_name = candidate

    return best_name, float(best_score)


def round_hours_nearest(duration_minutes: int) -> int:
    if duration_minutes <= 0:
        return 0
    return int((duration_minutes + 30) // 60)


def _parse_skud_datetime(value: Any) -> Optional[datetime]:
    if isinstance(value, datetime):
        return value

    s = normalize_spaces(str(value or ""))
    if not s:
        return None

    for fmt in ("%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            continue

    return None


def read_skud_events_from_xlsx(path: str) -> List[Dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    header_row: Optional[int] = None
    header_map: Dict[str, int] = {}
    max_scan = min(ws.max_row or 0, 60)
    max_cols = ws.max_column or 0

    for r in range(1, max_scan + 1):
        row_vals = [normalize_spaces(str(ws.cell(r, c).value or "")) for c in range(1, max_cols + 1)]
        if "Время" in row_vals and "Событие" in row_vals and "ФИО сотрудника" in row_vals:
            header_row = r
            for c, name in enumerate(row_vals, start=1):
                if name:
                    header_map[name] = c
            break

    if not header_row:
        raise RuntimeError("Не найден заголовок отчёта СКУД (колонки 'Время', 'Событие', 'ФИО сотрудника').")

    c_time = header_map.get("Время")
    c_fio = header_map.get("ФИО сотрудника")
    c_event = header_map.get("Событие")

    if not (c_time and c_fio and c_event):
        raise RuntimeError("В отчёте СКУД не найдены обязательные колонки.")

    events: List[Dict[str, Any]] = []
    for r in range(header_row + 1, (ws.max_row or header_row) + 1):
        fio = normalize_spaces(str(ws.cell(r, c_fio).value or ""))
        event_raw = normalize_spaces(str(ws.cell(r, c_event).value or ""))
        dt = _parse_skud_datetime(ws.cell(r, c_time).value)

        if not fio or not dt:
            continue

        if event_raw == "Вход":
            event_type = "in"
        elif event_raw == "Выход":
            event_type = "out"
        else:
            continue

        events.append({"dt": dt, "fio": fio, "event": event_type})

    return events


def _apply_default_skud_break(duration_minutes: int) -> int:
    """
    Сохраняем старую бизнес-логику:
    если суммарное присутствие > 4 часов, вычитаем 1 час на обед.
    """
    if duration_minutes > 4 * 60:
        return max(0, duration_minutes - 60)
    return max(0, duration_minutes)


def compute_day_summary_from_events(
    events: Sequence[Mapping[str, Any]],
    target_date: date,
) -> Tuple[Dict[str, Dict[str, Any]], List[Dict[str, Any]]]:
    """
    Более корректный расчёт по СКУД:
    - суммируем пары Вход -> Выход,
    - фиксируем проблемы, если есть висячие входы/выходы.

    Возвращает:
      summary_by_fio, problems
    где problems совместимы с текущим UI.
    """
    by_fio: Dict[str, List[Dict[str, Any]]] = {}
    for e in events:
        dt = e.get("dt")
        if not isinstance(dt, datetime):
            continue
        if dt.date() != target_date:
            continue

        fio = normalize_spaces(str(e.get("fio") or ""))
        ev = str(e.get("event") or "").strip()
        if not fio or ev not in ("in", "out"):
            continue

        by_fio.setdefault(fio, []).append({"dt": dt, "fio": fio, "event": ev})

    summary: Dict[str, Dict[str, Any]] = {}
    problems: List[Dict[str, Any]] = []

    for fio, items in by_fio.items():
        items.sort(key=lambda x: x["dt"])

        count_in = sum(1 for x in items if x["event"] == "in")
        count_out = sum(1 for x in items if x["event"] == "out")
        first_in = next((x["dt"] for x in items if x["event"] == "in"), None)
        last_out = next((x["dt"] for x in reversed(items) if x["event"] == "out"), None)

        total_minutes = 0
        open_in: Optional[datetime] = None
        anomaly = False

        for item in items:
            dt = item["dt"]
            ev = item["event"]

            if ev == "in":
                if open_in is None:
                    open_in = dt
                else:
                    # Повторный вход без выхода.
                    anomaly = True
                    # Сохраняем самый ранний открытый вход, новый игнорируем.
            else:  # out
                if open_in is None:
                    # Выход без входа.
                    anomaly = True
                    continue

                delta_minutes = int((dt - open_in).total_seconds() // 60)
                if delta_minutes < 0:
                    anomaly = True
                    open_in = None
                    continue

                total_minutes += delta_minutes
                open_in = None

        if open_in is not None:
            anomaly = True

        total_minutes = _apply_default_skud_break(total_minutes)
        hours_rounded = round_hours_nearest(total_minutes)

        if total_minutes > 0:
            summary[fio] = {
                "first_in": first_in,
                "last_out": last_out,
                "minutes": total_minutes,
                "hours_rounded": hours_rounded,
                "count_in": count_in,
                "count_out": count_out,
            }

        if anomaly or count_in == 0 or count_out == 0:
            problems.append(
                {
                    "skud_fio": fio,
                    "has_in": bool(count_in),
                    "has_out": bool(count_out),
                    "first_in": first_in,
                    "last_out": last_out,
                    "count_in": count_in,
                    "count_out": count_out,
                }
            )

    return summary, problems


def ensure_current_month_date(selected_date: date, year: int, month: int) -> bool:
    return selected_date.year == year and selected_date.month == month


def rows_have_unsaved_content(rows: Sequence[Mapping[str, Any]]) -> bool:
    for rec in rows:
        if normalize_spaces(str(rec.get("fio") or "")):
            return True
        if normalize_tbn(rec.get("tbn")):
            return True
        for item in rec.get("hours") or []:
            if normalize_spaces(str(item or "")):
                return True
    return False


__all__ = [
    "MAX_HOURS_PER_DAY",
    "TS_COLORS",
    "SPECIAL_CODES",
    "ParsedTimesheetCell",
    "exe_dir",
    "month_days",
    "month_name_ru",
    "safe_filename",
    "normalize_spaces",
    "normalize_tbn",
    "normalize_code",
    "is_allowed_timesheet_code",
    "format_hours_for_cell",
    "parse_hours_value",
    "parse_overtime",
    "parse_hours_and_night",
    "parse_timesheet_cell",
    "normalize_hours_list",
    "calc_row_totals",
    "calc_rows_summary",
    "format_summary_value",
    "make_row_key",
    "normalize_row_record",
    "deduplicate_timesheet_rows",
    "find_suspicious_cells",
    "is_suspicious_hours",
    "validate_row_record",
    "validate_rows_before_save",
    "best_fio_match_with_score",
    "round_hours_nearest",
    "read_skud_events_from_xlsx",
    "compute_day_summary_from_events",
    "ensure_current_month_date",
    "rows_have_unsaved_content",
]
