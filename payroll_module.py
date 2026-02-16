"""
payroll_module.py — Модуль «Затраты (ФОТ)»
Загрузка Excel-файла с начислениями ЗП, распределение по объектам
пропорционально часам из объектного табеля, аналитика.
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
from typing import Optional, List, Dict, Any, Tuple
from psycopg2 import pool
from psycopg2.extras import RealDictCursor
from decimal import Decimal, ROUND_HALF_UP
import logging
import re
import os

import pandas as pd

# Для автоширины колонок Excel
try:
    from openpyxl.utils import get_column_letter
except ImportError:
    get_column_letter = None

# ============================================================
#  DB pool — устанавливается из main_app при старте
# ============================================================

db_connection_pool: Optional[pool.SimpleConnectionPool] = None


def set_db_pool(db_pool: pool.SimpleConnectionPool):
    global db_connection_pool
    db_connection_pool = db_pool
    logging.info("Payroll Module: DB pool set.")


# ============================================================
#  Нормализация ТБН
# ============================================================

def normalize_tbn(raw: Any) -> str:
    """
    Приводит табельный номер к единому формату:
    - убирает пробелы по краям
    - если число — приводит к целому без .0
    - если строка вида 'СТЗК-31896' — извлекает '31896'
      (только если в БД tbn хранится без префикса)
    """
    if raw is None:
        return ""
    if isinstance(raw, float):
        # 31896.0 -> "31896"
        if raw == int(raw):
            return str(int(raw))
        return str(raw)
    if isinstance(raw, int):
        return str(raw)
    s = str(raw).strip()
    # Убираем .0 на конце (Excel иногда даёт "31896.0")
    if s.endswith('.0') and s[:-2].isdigit():
        s = s[:-2]
    return s


# ============================================================
#  EXCEL PARSER
# ============================================================

class PayrollExcelParser:
    """
    Парсер Excel с начислениями ЗП.
    """

    MONTH_MAP = {
        'январь': 1, 'февраль': 2, 'март': 3, 'апрель': 4,
        'май': 5, 'июнь': 6, 'июль': 7, 'август': 8,
        'сентябрь': 9, 'октябрь': 10, 'ноябрь': 11, 'декабрь': 12,
    }

    @staticmethod
    def parse(file_path: str) -> Dict[str, Any]:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        all_rows = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append(list(row))

        if len(all_rows) < 9:
            wb.close()
            raise ValueError("Файл слишком короткий — ожидается минимум 9 строк.")

        # --- Период и организация ---
        organization = ""
        period_label = ""
        year = None
        month = None

        for idx in range(min(7, len(all_rows))):
            for cell_val in all_rows[idx]:
                if cell_val and isinstance(cell_val, str):
                    cell_lower = cell_val.strip().lower()
                    # Ищем месяц
                    if not month:
                        for m_name, m_num in PayrollExcelParser.MONTH_MAP.items():
                            if m_name in cell_lower:
                                month = m_num
                                year_match = re.search(r'(\d{4})', cell_val)
                                if year_match:
                                    year = int(year_match.group(1))
                                period_label = cell_val.strip()
                                break
                    # Ищем организацию
                    if not organization and (
                        'организация' in cell_lower or
                        'ано ' in cell_lower or
                        'ооо ' in cell_lower or
                        cell_lower.startswith('ано ') or
                        cell_lower.startswith('ооо ')
                    ):
                        organization = cell_val.strip()

        now = datetime.now()
        year = year or now.year
        month = month or now.month
        if not period_label:
            period_label = f"{month:02d}.{year}"

        # --- Определяем колонку «Всего начислено» ---
        header_row_idx = 6
        headers = all_rows[header_row_idx] if len(all_rows) > header_row_idx else []
        total_col_idx = None

        for ci, hv in enumerate(headers):
            if hv and isinstance(hv, str) and 'всего' in hv.lower():
                total_col_idx = ci  # берём последнее «Всего»

        if total_col_idx is None:
            for ci in range(len(headers) - 1, -1, -1):
                if headers[ci] is not None:
                    total_col_idx = ci
                    break
            if total_col_idx is None:
                total_col_idx = len(headers) - 1

        # --- Основные индексы колонок ---
        COL_TBN = 0
        COL_FIO = 2
        COL_DEPT = 4
        COL_POS = 6
        COL_DAYS = 10
        COL_HOURS = 11

        # --- Вспомогательные функции ---
        def safe_float(v):
            if v is None:
                return None
            if isinstance(v, (int, float)):
                return float(v)
            try:
                return float(str(v).replace(',', '.').replace(' ', '').strip())
            except (ValueError, TypeError):
                return None

        def safe_int(v):
            f = safe_float(v)
            return int(f) if f is not None else None

        def cell(row_data, idx):
            return row_data[idx] if idx < len(row_data) else None

        # --- Парсим данные ---
        parsed_rows = []
        data_start_idx = 8
        skip_words = frozenset(('итого', 'всего', 'итого:', 'всего:', 'none', ''))

        for ri in range(data_start_idx, len(all_rows)):
            rd = all_rows[ri]
            if not rd or len(rd) < 3:
                continue

            tbn_str = normalize_tbn(cell(rd, COL_TBN))
            fio_raw = cell(rd, COL_FIO)
            fio_str = str(fio_raw).strip() if fio_raw else ""

            # Пропускаем пустые и итоговые строки
            if tbn_str.lower() in skip_words and fio_str.lower() in skip_words:
                continue
            if not tbn_str and not fio_str:
                continue
            if tbn_str.lower() in skip_words:
                continue

            total_accrued = safe_float(cell(rd, total_col_idx))

            parsed_rows.append({
                "tbn": tbn_str,
                "fio": fio_str,
                "department_raw": str(cell(rd, COL_DEPT) or "").strip(),
                "position_raw": str(cell(rd, COL_POS) or "").strip(),
                "worked_days": safe_int(cell(rd, COL_DAYS)),
                "worked_hours": safe_float(cell(rd, COL_HOURS)),
                "total_accrued": total_accrued,
            })

        wb.close()

        logging.info(
            f"PayrollExcelParser: parsed {len(parsed_rows)} rows, "
            f"period={month:02d}.{year}, org='{organization}'"
        )

        return {
            "organization": organization,
            "period_label": period_label,
            "year": year,
            "month": month,
            "rows": parsed_rows,
        }


# ============================================================
#  DATA LAYER
# ============================================================

class PayrollDataManager:
    """Работа с БД: сохранение загрузки, распределение, выборки."""

    @staticmethod
    def _get_conn():
        if not db_connection_pool:
            raise ConnectionError("Пул соединений не инициализирован.")
        return db_connection_pool.getconn()

    @staticmethod
    def _put_conn(conn):
        if conn and db_connection_pool:
            db_connection_pool.putconn(conn)

    @staticmethod
    def _query(sql: str, params: tuple = None) -> List[Dict]:
        conn = PayrollDataManager._get_conn()
        try:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(sql, params)
                return [dict(r) for r in cur.fetchall()]
        finally:
            PayrollDataManager._put_conn(conn)

    # ---- Загрузки ----

    @staticmethod
    def get_uploads() -> List[Dict]:
        return PayrollDataManager._query("""
            SELECT pu.id, pu.organization, pu.period_label,
                   pu.year, pu.month, pu.file_name,
                   pu.uploaded_at, pu.note,
                   au.full_name AS uploaded_by_name,
                   (SELECT COUNT(*) FROM payroll_rows pr
                    WHERE pr.upload_id = pu.id) AS row_count,
                   (SELECT COALESCE(SUM(pr.total_accrued),0)
                    FROM payroll_rows pr
                    WHERE pr.upload_id = pu.id) AS total_sum,
                   (SELECT COUNT(*) FROM payroll_distribution pd
                    JOIN payroll_rows pr2 ON pd.payroll_row_id = pr2.id
                    WHERE pr2.upload_id = pu.id) AS dist_count
            FROM payroll_uploads pu
            LEFT JOIN app_users au ON pu.uploaded_by = au.id
            ORDER BY pu.year DESC, pu.month DESC, pu.uploaded_at DESC
        """)

    @staticmethod
    def check_duplicate(year: int, month: int, file_name: str) -> Optional[Dict]:
        """Проверяет, есть ли уже загрузка за этот период с таким файлом."""
        rows = PayrollDataManager._query("""
            SELECT id, period_label, uploaded_at
            FROM payroll_uploads
            WHERE year = %s AND month = %s AND file_name = %s
            LIMIT 1
        """, (year, month, file_name))
        return rows[0] if rows else None

    @staticmethod
    def save_upload(parsed: Dict, file_name: str, user_id: int) -> int:
        """Сохраняет загрузку + строки, возвращает upload_id."""
        conn = PayrollDataManager._get_conn()
        try:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO payroll_uploads
                        (organization, period_label, year, month, file_name, uploaded_by)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    RETURNING id
                """, (
                    parsed["organization"],
                    parsed["period_label"],
                    parsed["year"],
                    parsed["month"],
                    file_name,
                    user_id,
                ))
                upload_id = cur.fetchone()[0]

                # Batch-поиск всех employee_id по tbn
                all_tbns = [r["tbn"] for r in parsed["rows"] if r["tbn"]]
                tbn_to_emp: Dict[str, int] = {}
                if all_tbns:
                    cur.execute(
                        "SELECT id, tbn FROM employees WHERE tbn = ANY(%s)",
                        (all_tbns,))
                    for row in cur.fetchall():
                        tbn_to_emp[row[1]] = row[0]

                for r in parsed["rows"]:
                    employee_id = tbn_to_emp.get(r["tbn"]) if r["tbn"] else None

                    cur.execute("""
                        INSERT INTO payroll_rows
                            (upload_id, tbn, fio, department_raw, position_raw,
                             worked_days, worked_hours, total_accrued, employee_id)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """, (
                        upload_id,
                        r["tbn"] or None,
                        r["fio"] or None,
                        r["department_raw"] or None,
                        r["position_raw"] or None,
                        r["worked_days"],
                        r["worked_hours"],
                        r["total_accrued"],
                        employee_id,
                    ))
            conn.commit()
            return upload_id
        except Exception:
            conn.rollback()
            raise
        finally:
            PayrollDataManager._put_conn(conn)

    @staticmethod
    def delete_upload(upload_id: int):
        """Удаляет загрузку каскадно (rows + distribution)."""
        conn = PayrollDataManager._get_conn()
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "DELETE FROM payroll_uploads WHERE id = %s", (upload_id,))
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            PayrollDataManager._put_conn(conn)

    # ---- Распределение ----

    @staticmethod
    def distribute(upload_id: int) -> Dict[str, int]:
        """
        Распределяет ФОТ по объектам для загрузки upload_id.
        Возвращает { "distributed": N, "not_found": M, "zero_accrued": K }
        """
        conn = PayrollDataManager._get_conn()
        try:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    "SELECT year, month FROM payroll_uploads WHERE id = %s",
                    (upload_id,))
                upl = cur.fetchone()
                if not upl:
                    raise ValueError(f"Загрузка {upload_id} не найдена")
                y, m = upl["year"], upl["month"]

                # Удаляем старое распределение
                cur.execute("""
                    DELETE FROM payroll_distribution
                    WHERE payroll_row_id IN (
                        SELECT id FROM payroll_rows WHERE upload_id = %s
                    )
                """, (upload_id,))

                # Получаем строки с ненулевым начислением и непустым tbn
                cur.execute("""
                    SELECT id, tbn, total_accrued
                    FROM payroll_rows
                    WHERE upload_id = %s
                      AND COALESCE(tbn, '') <> ''
                      AND COALESCE(total_accrued, 0) > 0
                """, (upload_id,))
                rows = cur.fetchall()

                # Предзагрузка всех табельных данных за период —
                # один запрос вместо N
                all_tbns = [r["tbn"] for r in rows]
                ts_map: Dict[str, List[Dict]] = {}

                if all_tbns:
                    cur.execute("""
                        SELECT
                            tr.tbn,
                            th.object_db_id AS object_id,
                            th.id AS header_id,
                            COALESCE(tr.total_hours, 0) AS hours
                        FROM timesheet_rows tr
                        JOIN timesheet_headers th ON th.id = tr.header_id
                        WHERE tr.tbn = ANY(%s)
                          AND th.year = %s
                          AND th.month = %s
                          AND COALESCE(tr.total_hours, 0) > 0
                    """, (all_tbns, y, m))
                    for tsr in cur.fetchall():
                        ts_map.setdefault(tsr["tbn"], []).append(tsr)

                stats = {"distributed": 0, "not_found": 0, "zero_accrued": 0}
                inserts = []

                for pr in rows:
                    pr_id = pr["id"]
                    tbn = pr["tbn"]
                    total_accrued = float(pr["total_accrued"])

                    ts_rows = ts_map.get(tbn)
                    if not ts_rows:
                        stats["not_found"] += 1
                        continue

                    total_ts_hours = sum(float(r["hours"]) for r in ts_rows)
                    if total_ts_hours <= 0:
                        stats["not_found"] += 1
                        continue

                    # Группируем по объекту
                    obj_hours: Dict[int, Tuple[float, int]] = {}
                    for tsr in ts_rows:
                        oid = tsr["object_id"]
                        h = float(tsr["hours"])
                        hid = tsr["header_id"]
                        if oid in obj_hours:
                            old_h, old_hid = obj_hours[oid]
                            obj_hours[oid] = (old_h + h, old_hid)
                        else:
                            obj_hours[oid] = (h, hid)

                    # Распределяем с точным остатком
                    distributed_sum = Decimal("0")
                    items = list(obj_hours.items())

                    for i, (oid, (h_on_obj, hdr_id)) in enumerate(items):
                        fraction = Decimal(str(h_on_obj)) / Decimal(
                            str(total_ts_hours))

                        if i == len(items) - 1:
                            amount = Decimal(str(total_accrued)) - distributed_sum
                        else:
                            amount = (
                                Decimal(str(total_accrued)) * fraction
                            ).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                            distributed_sum += amount

                        inserts.append((
                            pr_id, oid, hdr_id,
                            round(h_on_obj, 2),
                            round(total_ts_hours, 2),
                            round(float(fraction), 6),
                            float(amount),
                        ))

                    stats["distributed"] += 1

                # Batch insert
                if inserts:
                    from psycopg2.extras import execute_values
                    execute_values(
                        cur,
                        """INSERT INTO payroll_distribution
                               (payroll_row_id, object_id, timesheet_header_id,
                                hours_on_object, total_hours_all_objects,
                                fraction, amount)
                           VALUES %s""",
                        inserts,
                        template="(%s, %s, %s, %s, %s, %s, %s)",
                    )

            conn.commit()
            logging.info(
                f"Payroll distribute upload_id={upload_id}: {stats}")
            return stats
        except Exception:
            conn.rollback()
            raise
        finally:
            PayrollDataManager._put_conn(conn)

    # ---- Аналитические выборки ----

    @staticmethod
    def get_distribution_by_object(upload_id: int) -> pd.DataFrame:
        data = PayrollDataManager._query("""
            SELECT
                o.address AS object_name,
                o.short_name AS object_type,
                COUNT(DISTINCT pr.tbn) AS people_cnt,
                SUM(pd.hours_on_object) AS total_hours,
                SUM(pd.amount) AS total_amount
            FROM payroll_distribution pd
            JOIN payroll_rows pr ON pr.id = pd.payroll_row_id
            JOIN objects o ON o.id = pd.object_id
            WHERE pr.upload_id = %s
            GROUP BY o.id, o.address, o.short_name
            ORDER BY total_amount DESC
        """, (upload_id,))
        df = pd.DataFrame(data)
        if not df.empty:
            df["total_hours"] = df["total_hours"].astype(float)
            df["total_amount"] = df["total_amount"].astype(float)
            df["people_cnt"] = df["people_cnt"].astype(int)
        return df

    @staticmethod
    def get_distribution_by_department(upload_id: int) -> pd.DataFrame:
        data = PayrollDataManager._query("""
            SELECT
                COALESCE(NULLIF(pr.department_raw, ''), '—') AS department_name,
                COUNT(DISTINCT pr.tbn) AS people_cnt,
                SUM(pr.total_accrued) AS total_accrued,
                COALESCE(SUM(pd_sum.distributed), 0) AS total_distributed
            FROM payroll_rows pr
            LEFT JOIN (
                SELECT payroll_row_id, SUM(amount) AS distributed
                FROM payroll_distribution
                GROUP BY payroll_row_id
            ) pd_sum ON pd_sum.payroll_row_id = pr.id
            WHERE pr.upload_id = %s
            GROUP BY COALESCE(NULLIF(pr.department_raw, ''), '—')
            ORDER BY total_accrued DESC
        """, (upload_id,))
        df = pd.DataFrame(data)
        if not df.empty:
            df["total_accrued"] = df["total_accrued"].fillna(0).astype(float)
            df["total_distributed"] = df["total_distributed"].fillna(0).astype(float)
            df["people_cnt"] = df["people_cnt"].astype(int)
        return df

    @staticmethod
    def get_undistributed_rows(upload_id: int) -> pd.DataFrame:
        """Сотрудники, которых не удалось распределить."""
        data = PayrollDataManager._query("""
            SELECT
                pr.tbn, pr.fio, pr.department_raw,
                pr.position_raw, pr.total_accrued
            FROM payroll_rows pr
            WHERE pr.upload_id = %s
              AND pr.id NOT IN (
                  SELECT DISTINCT payroll_row_id FROM payroll_distribution
              )
              AND COALESCE(pr.total_accrued, 0) > 0
            ORDER BY pr.total_accrued DESC
        """, (upload_id,))
        return pd.DataFrame(data)

    @staticmethod
    def get_upload_summary(upload_id: int) -> Dict[str, Any]:
        """Сводка по загрузке — один оптимизированный запрос."""
        rows = PayrollDataManager._query("""
            WITH pr AS (
                SELECT id, total_accrued
                FROM payroll_rows
                WHERE upload_id = %s
            ),
            dist AS (
                SELECT pd.payroll_row_id, pd.object_id, pd.amount
                FROM payroll_distribution pd
                JOIN pr ON pr.id = pd.payroll_row_id
            )
            SELECT
                (SELECT COUNT(*) FROM pr) AS total_rows,
                (SELECT COALESCE(SUM(total_accrued), 0) FROM pr) AS total_accrued,
                (SELECT COALESCE(SUM(amount), 0) FROM dist) AS total_distributed,
                (SELECT COUNT(DISTINCT payroll_row_id) FROM dist) AS rows_distributed,
                (SELECT COUNT(*)
                 FROM pr
                 WHERE COALESCE(total_accrued, 0) > 0
                   AND id NOT IN (SELECT DISTINCT payroll_row_id FROM dist)
                ) AS rows_not_distributed,
                (SELECT COUNT(DISTINCT object_id) FROM dist) AS objects_count
        """, (upload_id,))
        r = rows[0] if rows else {}
        for k in ("total_accrued", "total_distributed"):
            r[k] = float(r.get(k, 0) or 0)
        for k in ("total_rows", "rows_distributed",
                   "rows_not_distributed", "objects_count"):
            r[k] = int(r.get(k, 0) or 0)
        r["undistributed_amount"] = r["total_accrued"] - r["total_distributed"]
        return r

    @staticmethod
    def get_detail_by_employee(upload_id: int) -> pd.DataFrame:
        """Детализация: каждый сотрудник → объект(ы) → сумма."""
        data = PayrollDataManager._query("""
            SELECT
                pr.tbn,
                pr.fio,
                pr.department_raw,
                pr.total_accrued,
                o.address AS object_name,
                pd.hours_on_object,
                pd.total_hours_all_objects,
                pd.fraction,
                pd.amount
            FROM payroll_rows pr
            JOIN payroll_distribution pd ON pd.payroll_row_id = pr.id
            JOIN objects o ON o.id = pd.object_id
            WHERE pr.upload_id = %s
            ORDER BY pr.fio, o.address
        """, (upload_id,))
        return pd.DataFrame(data)


# ============================================================
#  UI: Главная страница модуля
# ============================================================

class PayrollPage(ttk.Frame):
    """Страница «Затраты (ФОТ)»."""

    def __init__(self, master, app_ref):
        super().__init__(master)
        self.app_ref = app_ref

        # ---- Верхняя панель кнопок ----
        toolbar = ttk.Frame(self, padding="8")
        toolbar.pack(fill="x", side="top")

        ttk.Button(
            toolbar, text="📂 Загрузить Excel",
            command=self._on_upload).pack(side="left", padx=4)
        ttk.Button(
            toolbar, text="🔄 Распределить ФОТ",
            command=self._on_distribute).pack(side="left", padx=4)
        ttk.Button(
            toolbar, text="🗑 Удалить загрузку",
            command=self._on_delete).pack(side="left", padx=4)

        ttk.Separator(toolbar, orient="vertical").pack(
            side="left", fill="y", padx=8, pady=2)

        ttk.Button(
            toolbar, text="📥 Полный отчёт в Excel",
            command=self._on_export_full).pack(side="left", padx=4)
        ttk.Button(
            toolbar, text="↻ Обновить",
            command=self._refresh).pack(side="left", padx=4)

        # ---- Основная область: PanedWindow ----
        pw = ttk.PanedWindow(self, orient="horizontal")
        pw.pack(fill="both", expand=True, padx=5, pady=5)

        # Левая панель — список загрузок
        left = ttk.LabelFrame(pw, text="Загрузки")
        pw.add(left, weight=1)

        self.tree_uploads = ttk.Treeview(left, columns=(
            "id", "period", "file", "rows", "sum", "dist", "date"
        ), show="headings", height=12, selectmode="browse")

        cols_cfg = [
            ("id", "ID", 40),
            ("period", "Период", 110),
            ("file", "Файл", 180),
            ("rows", "Строк", 55),
            ("sum", "Сумма ФОТ", 110),
            ("dist", "Распр.", 55),
            ("date", "Загружено", 130),
        ]
        for cid, text, w in cols_cfg:
            self.tree_uploads.heading(cid, text=text)
            self.tree_uploads.column(
                cid, width=w,
                anchor="e" if cid in ("rows", "sum", "dist") else "w")

        vsb = ttk.Scrollbar(
            left, orient="vertical", command=self.tree_uploads.yview)
        self.tree_uploads.configure(yscrollcommand=vsb.set)
        self.tree_uploads.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        self.tree_uploads.bind("<<TreeviewSelect>>", self._on_upload_selected)

        # Правая панель — детали выбранной загрузки
        right = ttk.Frame(pw)
        pw.add(right, weight=3)

        self.detail_notebook = ttk.Notebook(right)
        self.detail_notebook.pack(fill="both", expand=True)

        self.tab_summary = ttk.Frame(self.detail_notebook)
        self.tab_by_object = ttk.Frame(self.detail_notebook)
        self.tab_by_dept = ttk.Frame(self.detail_notebook)
        self.tab_detail = ttk.Frame(self.detail_notebook)
        self.tab_unmatched = ttk.Frame(self.detail_notebook)

        self.detail_notebook.add(self.tab_summary, text="  Сводка  ")
        self.detail_notebook.add(self.tab_by_object, text="  По объектам  ")
        self.detail_notebook.add(self.tab_by_dept, text="  По подразделениям  ")
        self.detail_notebook.add(self.tab_detail, text="  Детализация  ")
        self.detail_notebook.add(self.tab_unmatched, text="  Не распределено  ")

        self._selected_upload_id: Optional[int] = None

        self._refresh()

    # ---- Actions ----

    def _on_upload(self):
        file_path = filedialog.askopenfilename(
            title="Выберите Excel-файл с начислениями",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
        )
        if not file_path:
            return

        try:
            parsed = PayrollExcelParser.parse(file_path)
        except Exception as e:
            logging.exception("Ошибка парсинга Excel")
            messagebox.showerror("Ошибка парсинга",
                                 f"Не удалось прочитать файл:\n{e}")
            return

        file_name = os.path.basename(file_path)
        row_count = len(parsed["rows"])
        total = sum(r["total_accrued"] or 0 for r in parsed["rows"])

        # Проверка дубликата
        dup = PayrollDataManager.check_duplicate(
            parsed["year"], parsed["month"], file_name)
        if dup:
            dt = dup.get("uploaded_at")
            dt_str = dt.strftime("%d.%m.%Y %H:%M") if dt else "?"
            if not messagebox.askyesno(
                "Дубликат",
                f"Файл «{file_name}» за период {parsed['period_label']} "
                f"уже загружен (#{dup['id']}, {dt_str}).\n\n"
                f"Загрузить повторно?"
            ):
                return

        msg = (
            f"Файл: {file_name}\n"
            f"Период: {parsed['period_label']}\n"
            f"Организация: {parsed['organization']}\n"
            f"Строк данных: {row_count}\n"
            f"Сумма «Всего начислено»: {total:,.2f} ₽\n\n"
            f"Загрузить?"
        )
        if not messagebox.askyesno("Подтверждение загрузки", msg):
            return

        try:
            user_id = self.app_ref.current_user.get("id")
            upload_id = PayrollDataManager.save_upload(
                parsed, file_name, user_id)
            messagebox.showinfo(
                "Успех",
                f"Загрузка #{upload_id} сохранена.\n"
                f"{row_count} строк.\n\n"
                f"Теперь нажмите «Распределить ФОТ».")
            self._selected_upload_id = upload_id
            self._refresh()
        except Exception as e:
            logging.exception("Ошибка сохранения загрузки")
            messagebox.showerror("Ошибка", f"Не удалось сохранить:\n{e}")

    def _on_distribute(self):
        if not self._selected_upload_id:
            messagebox.showwarning(
                "Внимание", "Выберите загрузку в списке слева.")
            return
        uid = self._selected_upload_id

        if not messagebox.askyesno(
                "Распределение",
                f"Распределить ФОТ загрузки #{uid} по объектам?\n"
                f"(старое распределение будет пересчитано)"):
            return

        try:
            stats = PayrollDataManager.distribute(uid)
            messagebox.showinfo(
                "Результат распределения",
                f"Распределено сотрудников: {stats['distributed']}\n"
                f"Не найдено в табелях: {stats['not_found']}\n"
                f"Нулевое начисление: {stats['zero_accrued']}")
            self._refresh()
            self._show_upload_details(uid)
        except Exception as e:
            logging.exception("Ошибка распределения")
            messagebox.showerror("Ошибка", f"Не удалось распределить:\n{e}")

    def _on_delete(self):
        if not self._selected_upload_id:
            messagebox.showwarning(
                "Внимание", "Выберите загрузку в списке слева.")
            return
        uid = self._selected_upload_id
        if not messagebox.askyesno(
                "Удаление",
                f"Удалить загрузку #{uid} и все связанные данные?\n"
                f"Это действие нельзя отменить."):
            return
        try:
            PayrollDataManager.delete_upload(uid)
            self._selected_upload_id = None
            self._refresh()
            self._clear_details()
            messagebox.showinfo("Готово", f"Загрузка #{uid} удалена.")
        except Exception as e:
            logging.exception("Ошибка удаления загрузки")
            messagebox.showerror("Ошибка", f"Не удалось удалить:\n{e}")

    def _on_export_full(self):
        """Экспорт всех вкладок в один Excel-файл (разные листы)."""
        if not self._selected_upload_id:
            messagebox.showwarning(
                "Внимание", "Выберите загрузку в списке слева.")
            return
        uid = self._selected_upload_id

        path = self._ask_save_path(f"ФОТ_полный_отчет_{uid}.xlsx")
        if not path:
            return

        try:
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                # Лист 1: По объектам
                df_obj = PayrollDataManager.get_distribution_by_object(uid)
                if not df_obj.empty:
                    grand_total = df_obj["total_amount"].sum()
                    df_e = df_obj.copy()
                    df_e["share_pct"] = df_e["total_amount"].apply(
                        lambda x: round(x / grand_total * 100, 1)
                        if grand_total > 0 else 0)
                    df_e = df_e.rename(columns={
                        "object_name": "Объект",
                        "object_type": "Тип объекта",
                        "people_cnt": "Кол-во сотрудников",
                        "total_hours": "Часов на объекте",
                        "total_amount": "Сумма ФОТ, ₽",
                        "share_pct": "Доля, %",
                    })
                    totals = pd.DataFrame([{
                        "Объект": "ИТОГО",
                        "Тип объекта": "",
                        "Кол-во сотрудников": int(
                            df_e["Кол-во сотрудников"].sum()),
                        "Часов на объекте": round(
                            df_e["Часов на объекте"].sum(), 1),
                        "Сумма ФОТ, ₽": round(
                            df_e["Сумма ФОТ, ₽"].sum(), 2),
                        "Доля, %": 100.0,
                    }])
                    df_e = pd.concat([df_e, totals], ignore_index=True)
                    df_e.to_excel(
                        writer, index=False, sheet_name="По объектам")
                    self._autofit_columns(writer, "По объектам", df_e)

                # Лист 2: По подразделениям
                df_dept = PayrollDataManager.get_distribution_by_department(uid)
                if not df_dept.empty:
                    df_d = df_dept.copy()
                    df_d["diff"] = (
                        df_d["total_accrued"] - df_d["total_distributed"])
                    df_d["pct"] = df_d.apply(
                        lambda r: round(
                            r["total_distributed"] / r["total_accrued"] * 100,
                            1)
                        if r["total_accrued"] > 0 else 0, axis=1)
                    df_d = df_d.rename(columns={
                        "department_name": "Подразделение",
                        "people_cnt": "Кол-во сотрудников",
                        "total_accrued": "Начислено, ₽",
                        "total_distributed": "Распределено, ₽",
                        "diff": "Остаток, ₽",
                        "pct": "Распределено, %",
                    })
                    df_d.to_excel(
                        writer, index=False, sheet_name="По подразделениям")
                    self._autofit_columns(writer, "По подразделениям", df_d)

                # Лист 3: Детализация
                df_det = PayrollDataManager.get_detail_by_employee(uid)
                if not df_det.empty:
                    df_det2 = df_det.rename(columns={
                        "tbn": "Таб. номер",
                        "fio": "ФИО",
                        "department_raw": "Подразделение",
                        "total_accrued": "Всего начислено",
                        "object_name": "Объект",
                        "hours_on_object": "Часы на объекте",
                        "total_hours_all_objects": "Всего часов",
                        "fraction": "Доля",
                        "amount": "Сумма на объект, ₽",
                    })
                    df_det2.to_excel(
                        writer, index=False, sheet_name="Детализация")
                    self._autofit_columns(writer, "Детализация", df_det2)

                # Лист 4: Не распределено
                df_un = PayrollDataManager.get_undistributed_rows(uid)
                if not df_un.empty:
                    df_un2 = df_un.rename(columns={
                        "tbn": "Таб. номер",
                        "fio": "ФИО",
                        "department_raw": "Подразделение",
                        "position_raw": "Должность",
                        "total_accrued": "Начислено, ₽",
                    })
                    df_un2.to_excel(
                        writer, index=False, sheet_name="Не распределено")
                    self._autofit_columns(writer, "Не распределено", df_un2)

            messagebox.showinfo("Экспорт", f"Полный отчёт сохранён:\n{path}")
        except Exception as e:
            logging.exception("Ошибка полного экспорта")
            messagebox.showerror("Ошибка", f"Не удалось сохранить:\n{e}")

    # ---- Refresh / Select ----

    def _refresh(self):
        for item in self.tree_uploads.get_children():
            self.tree_uploads.delete(item)
        try:
            uploads = PayrollDataManager.get_uploads()
        except Exception as e:
            logging.exception("Ошибка загрузки списка payroll_uploads")
            return
        for u in uploads:
            dt = u.get("uploaded_at")
            dt_str = dt.strftime("%d.%m.%Y %H:%M") if dt else ""
            total_sum = float(u.get("total_sum", 0) or 0)
            self.tree_uploads.insert("", "end", iid=str(u["id"]), values=(
                u["id"],
                u.get("period_label") or f"{u['month']:02d}.{u['year']}",
                u.get("file_name") or "",
                u.get("row_count", 0),
                f"{total_sum:,.2f}".replace(",", " "),
                u.get("dist_count", 0),
                dt_str,
            ))
        # Восстанавливаем выделение
        if self._selected_upload_id:
            iid = str(self._selected_upload_id)
            if self.tree_uploads.exists(iid):
                self.tree_uploads.selection_set(iid)
                self.tree_uploads.focus(iid)
                self._show_upload_details(self._selected_upload_id)

    def _on_upload_selected(self, event=None):
        sel = self.tree_uploads.selection()
        if not sel:
            return
        uid = int(sel[0])
        self._selected_upload_id = uid
        self._show_upload_details(uid)

    # ---- Details ----

    def _clear_tab(self, tab):
        for w in tab.winfo_children():
            w.destroy()

    def _clear_details(self):
        self._clear_tab(self.tab_summary)
        self._clear_tab(self.tab_by_object)
        self._clear_tab(self.tab_by_dept)
        self._clear_tab(self.tab_detail)
        self._clear_tab(self.tab_unmatched)

    def _show_upload_details(self, upload_id: int):
        self._clear_details()
        try:
            self._build_summary_tab(upload_id)
            self._build_by_object_tab(upload_id)
            self._build_by_dept_tab(upload_id)
            self._build_detail_tab(upload_id)
            self._build_unmatched_tab(upload_id)
        except Exception as e:
            logging.exception("Ошибка построения деталей загрузки")
            ttk.Label(self.tab_summary,
                      text=f"Ошибка: {e}").pack(padx=10, pady=10)

    # ---- Tab: Сводка ----

    def _create_kpi_card(self, parent, title, value, unit):
        card = ttk.Frame(parent, borderwidth=2, relief="groove", padding=10)
        ttk.Label(card, text=title,
                  font=("Segoe UI", 9, "bold")).pack()
        ttk.Label(card, text=f"{value}",
                  font=("Segoe UI", 16, "bold"),
                  foreground="#0078D7").pack(pady=(4, 0))
        ttk.Label(card, text=unit, font=("Segoe UI", 8)).pack()
        return card

    def _build_summary_tab(self, upload_id: int):
        tab = self.tab_summary
        s = PayrollDataManager.get_upload_summary(upload_id)

        kpi_frame = ttk.Frame(tab)
        kpi_frame.pack(fill="x", pady=10, padx=10)

        cards = [
            ("Всего строк", s["total_rows"], "чел."),
            ("Сумма ФОТ",
             f"{s['total_accrued']:,.0f}".replace(",", " "), "₽"),
            ("Распределено",
             f"{s['total_distributed']:,.0f}".replace(",", " "), "₽"),
            ("Не распределено",
             f"{s['undistributed_amount']:,.0f}".replace(",", " "), "₽"),
            ("Сотр. распред.", s["rows_distributed"], "чел."),
            ("Сотр. без объекта", s["rows_not_distributed"], "чел."),
            ("Объектов", s["objects_count"], "шт."),
        ]
        for i, (title, value, unit) in enumerate(cards):
            card = self._create_kpi_card(kpi_frame, title, value, unit)
            card.grid(row=0, column=i, padx=4, sticky="ew")
            kpi_frame.grid_columnconfigure(i, weight=1)

        # Процент распределения
        pct = 0.0
        if s["total_accrued"] > 0:
            pct = s["total_distributed"] / s["total_accrued"] * 100
        pct_frame = ttk.Frame(tab)
        pct_frame.pack(fill="x", padx=10, pady=(0, 10))

        bar_bg = ttk.Frame(pct_frame, relief="sunken", borderwidth=1)
        bar_bg.pack(fill="x", pady=4)
        bar_fill = tk.Frame(bar_bg, bg="#0078D7", height=20)
        bar_fill.pack(side="left", fill="y")

        def _update_bar(event=None):
            total_w = bar_bg.winfo_width()
            fill_w = max(1, int(total_w * pct / 100))
            bar_fill.configure(width=fill_w)
        bar_bg.bind("<Configure>", _update_bar)

        ttk.Label(
            pct_frame,
            text=f"Распределено {pct:.1f}% от общей суммы ФОТ",
            font=("Segoe UI", 9)).pack(anchor="w")

        if s["rows_not_distributed"] > 0:
            warn_frame = ttk.Frame(tab)
            warn_frame.pack(fill="x", padx=10, pady=5)
            ttk.Label(
                warn_frame,
                text=(
                    f"⚠ {s['rows_not_distributed']} сотрудник(ов) "
                    f"не найдены в табелях за этот период. "
                    f"Их ФОТ ({s['undistributed_amount']:,.0f} ₽) "
                    f"не распределён. См. вкладку «Не распределено»."),
                foreground="#B00020",
                wraplength=700,
                justify="left",
            ).pack(anchor="w")

    # ---- Tab: По объектам ----

    def _build_by_object_tab(self, upload_id: int):
        tab = self.tab_by_object
        df = PayrollDataManager.get_distribution_by_object(upload_id)

        if df.empty:
            ttk.Label(
                tab, text="Нет данных. Нажмите «Распределить ФОТ».",
                font=("Segoe UI", 10)).pack(padx=20, pady=20)
            return

        btn_frame = ttk.Frame(tab)
        btn_frame.pack(fill="x", padx=5, pady=(5, 0))
        ttk.Button(
            btn_frame, text="📥 Выгрузить в Excel",
            command=lambda: self._export_by_object(upload_id)
        ).pack(side="right", padx=5)

        table_frame = ttk.Frame(tab)
        table_frame.pack(fill="both", expand=True, padx=5, pady=5)

        tree = ttk.Treeview(table_frame, columns=(
            "num", "object", "type", "people", "hours", "amount", "share"
        ), show="headings", height=22)

        cols = [
            ("num", "№", 40, "center"),
            ("object", "Объект", 400, "w"),
            ("type", "Тип", 100, "w"),
            ("people", "Людей", 65, "e"),
            ("hours", "Часов", 90, "e"),
            ("amount", "Сумма, ₽", 130, "e"),
            ("share", "Доля %", 70, "e"),
        ]
        for cid, text, w, anchor in cols:
            tree.heading(cid, text=text)
            tree.column(cid, width=w, anchor=anchor, minwidth=40)

        vsb = ttk.Scrollbar(
            table_frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(
            table_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        grand_total = df["total_amount"].sum()
        for idx, (_, row) in enumerate(df.iterrows(), 1):
            share = (
                row["total_amount"] / grand_total * 100
            ) if grand_total > 0 else 0
            tree.insert("", "end", values=(
                idx,
                row.get("object_name", "—"),
                row.get("object_type", ""),
                int(row["people_cnt"]),
                f"{row['total_hours']:,.1f}".replace(",", " "),
                f"{row['total_amount']:,.2f}".replace(",", " "),
                f"{share:.1f}",
            ))

        tree.insert("", "end", values=(
            "", "ИТОГО", "",
            int(df["people_cnt"].sum()),
            f"{df['total_hours'].sum():,.1f}".replace(",", " "),
            f"{grand_total:,.2f}".replace(",", " "),
            "100.0",
        ), tags=("total",))
        tree.tag_configure("total", font=("Segoe UI", 9, "bold"))

    # ---- Tab: По подразделениям ----

    def _build_by_dept_tab(self, upload_id: int):
        tab = self.tab_by_dept
        df = PayrollDataManager.get_distribution_by_department(upload_id)

        if df.empty:
            ttk.Label(tab, text="Нет данных.").pack(padx=20, pady=20)
            return

        btn_frame = ttk.Frame(tab)
        btn_frame.pack(fill="x", padx=5, pady=(5, 0))
        ttk.Button(
            btn_frame, text="📥 Выгрузить в Excel",
            command=lambda: self._export_by_dept(upload_id)
        ).pack(side="right", padx=5)

        table_frame = ttk.Frame(tab)
        table_frame.pack(fill="both", expand=True, padx=5, pady=5)

        tree = ttk.Treeview(table_frame, columns=(
            "num", "dept", "people", "accrued", "distributed", "diff", "pct"
        ), show="headings", height=22)

        for cid, text, w, anc in [
            ("num", "№", 40, "center"),
            ("dept", "Подразделение", 300, "w"),
            ("people", "Людей", 65, "e"),
            ("accrued", "Начислено, ₽", 130, "e"),
            ("distributed", "Распределено, ₽", 130, "e"),
            ("diff", "Остаток, ₽", 120, "e"),
            ("pct", "Распр. %", 75, "e"),
        ]:
            tree.heading(cid, text=text)
            tree.column(cid, width=w, anchor=anc, minwidth=40)

        vsb = ttk.Scrollbar(
            table_frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(
            table_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        for idx, (_, row) in enumerate(df.iterrows(), 1):
            diff = row["total_accrued"] - row["total_distributed"]
            pct = (
                row["total_distributed"] / row["total_accrued"] * 100
            ) if row["total_accrued"] > 0 else 0
            tree.insert("", "end", values=(
                idx,
                row["department_name"],
                int(row["people_cnt"]),
                f"{row['total_accrued']:,.2f}".replace(",", " "),
                f"{row['total_distributed']:,.2f}".replace(",", " "),
                f"{diff:,.2f}".replace(",", " "),
                f"{pct:.1f}",
            ))

        total_accrued = df["total_accrued"].sum()
        total_distributed = df["total_distributed"].sum()
        total_diff = total_accrued - total_distributed
        total_pct = (
            total_distributed / total_accrued * 100
        ) if total_accrued > 0 else 0
        tree.insert("", "end", values=(
            "", "ИТОГО",
            int(df["people_cnt"].sum()),
            f"{total_accrued:,.2f}".replace(",", " "),
            f"{total_distributed:,.2f}".replace(",", " "),
            f"{total_diff:,.2f}".replace(",", " "),
            f"{total_pct:.1f}",
        ), tags=("total",))
        tree.tag_configure("total", font=("Segoe UI", 9, "bold"))

    # ---- Tab: Детализация ----

    def _build_detail_tab(self, upload_id: int):
        tab = self.tab_detail
        df = PayrollDataManager.get_detail_by_employee(upload_id)

        if df.empty:
            ttk.Label(
                tab,
                text="Нет данных. Нажмите «Распределить ФОТ».",
                font=("Segoe UI", 10)).pack(padx=20, pady=20)
            return

        btn_frame = ttk.Frame(tab)
        btn_frame.pack(fill="x", padx=5, pady=(5, 0))
        ttk.Button(
            btn_frame, text="📥 Выгрузить в Excel",
            command=lambda: self._export_detail(upload_id)
        ).pack(side="right", padx=5)

        table_frame = ttk.Frame(tab)
        table_frame.pack(fill="both", expand=True, padx=5, pady=5)

        tree = ttk.Treeview(table_frame, columns=(
            "num", "tbn", "fio", "dept", "accrued",
            "object", "hours_obj", "hours_total", "fraction", "amount"
        ), show="headings", height=22)

        for cid, text, w, anc in [
            ("num", "№", 35, "center"),
            ("tbn", "ТБН", 80, "w"),
            ("fio", "ФИО", 180, "w"),
            ("dept", "Подразделение", 140, "w"),
            ("accrued", "Начислено", 100, "e"),
            ("object", "Объект", 250, "w"),
            ("hours_obj", "Часы объект", 80, "e"),
            ("hours_total", "Часы всего", 80, "e"),
            ("fraction", "Доля", 55, "e"),
            ("amount", "Сумма, ₽", 100, "e"),
        ]:
            tree.heading(cid, text=text)
            tree.column(cid, width=w, anchor=anc, minwidth=30)

        vsb = ttk.Scrollbar(
            table_frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(
            table_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        for idx, (_, row) in enumerate(df.iterrows(), 1):
            frac = float(row.get("fraction", 0) or 0)
            tree.insert("", "end", values=(
                idx,
                row.get("tbn", ""),
                row.get("fio", ""),
                row.get("department_raw", ""),
                f"{float(row.get('total_accrued', 0) or 0):,.2f}".replace(
                    ",", " "),
                row.get("object_name", ""),
                f"{float(row.get('hours_on_object', 0) or 0):,.1f}",
                f"{float(row.get('total_hours_all_objects', 0) or 0):,.1f}",
                f"{frac:.4f}",
                f"{float(row.get('amount', 0) or 0):,.2f}".replace(",", " "),
            ))

        ttk.Label(
            tab,
            text=f"Всего строк: {len(df)}",
            font=("Segoe UI", 8), foreground="#888"
        ).pack(anchor="w", padx=10, pady=(0, 5))

    # ---- Tab: Не распределено ----

    def _build_unmatched_tab(self, upload_id: int):
        tab = self.tab_unmatched
        df = PayrollDataManager.get_undistributed_rows(upload_id)

        if df.empty:
            ttk.Label(
                tab,
                text="✅ Все сотрудники успешно распределены по объектам!",
                font=("Segoe UI", 11),
                foreground="#16A34A").pack(padx=20, pady=30)
            return

        total_lost = df["total_accrued"].fillna(0).astype(float).sum()

        info_frame = ttk.Frame(tab)
        info_frame.pack(fill="x", padx=10, pady=8)

        ttk.Label(
            info_frame,
            text=(
                f"⚠ {len(df)} сотрудник(ов) не найдены в объектном табеле "
                f"за данный месяц.\n"
                f"Нераспределённая сумма: {total_lost:,.2f} ₽\n\n"
                f"Возможные причины:\n"
                f"  • Табельный номер в Excel не совпадает с tbn в табеле\n"
                f"  • Сотрудник не внесён в объектный табель за этот месяц\n"
                f"  • Административный/офисный персонал без объекта"),
            foreground="#B00020",
            wraplength=700,
            justify="left",
        ).pack(side="left", anchor="w")

        ttk.Button(
            info_frame, text="📥 Выгрузить в Excel",
            command=lambda: self._export_unmatched(upload_id),
        ).pack(side="right", padx=5)

        tree_frame = ttk.Frame(tab)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=5)

        tree = ttk.Treeview(tree_frame, columns=(
            "num", "tbn", "fio", "dept", "pos", "accrued"
        ), show="headings", height=20)

        for cid, text, w, anc in [
            ("num", "№", 40, "center"),
            ("tbn", "Таб. номер", 100, "w"),
            ("fio", "ФИО", 250, "w"),
            ("dept", "Подразделение", 200, "w"),
            ("pos", "Должность", 200, "w"),
            ("accrued", "Начислено, ₽", 120, "e"),
        ]:
            tree.heading(cid, text=text)
            tree.column(cid, width=w, anchor=anc, minwidth=40)

        vsb = ttk.Scrollbar(
            tree_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        for idx, (_, row) in enumerate(df.iterrows(), 1):
            accrued = float(row.get("total_accrued", 0) or 0)
            tree.insert("", "end", values=(
                idx,
                row.get("tbn", ""),
                row.get("fio", ""),
                row.get("department_raw", ""),
                row.get("position_raw", ""),
                f"{accrued:,.2f}".replace(",", " "),
            ))

        tree.insert("", "end", values=(
            "", "", "ИТОГО", "", "",
            f"{total_lost:,.2f}".replace(",", " "),
        ), tags=("total",))
        tree.tag_configure("total", font=("Segoe UI", 9, "bold"))

    # ============================================================
    #  ЭКСПОРТ В EXCEL
    # ============================================================

    def _ask_save_path(self, default_name: str) -> Optional[str]:
        path = filedialog.asksaveasfilename(
            title="Сохранить как",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        return path if path else None

    def _autofit_columns(self, writer, sheet_name: str, df: pd.DataFrame):
        """Автоподбор ширины колонок в Excel."""
        try:
            ws = writer.sheets[sheet_name]
            for i, col in enumerate(df.columns):
                max_len = max(
                    len(str(col)),
                    df[col].astype(str).str.len().max() if len(df) > 0 else 0
                )
                col_letter = (
                    get_column_letter(i + 1) if get_column_letter
                    else chr(65 + i) if i < 26
                    else chr(64 + i // 26) + chr(65 + i % 26)
                )
                ws.column_dimensions[col_letter].width = min(max_len + 3, 55)
        except Exception:
            pass

    def _export_by_object(self, upload_id: int):
        df = PayrollDataManager.get_distribution_by_object(upload_id)
        if df.empty:
            messagebox.showinfo("Экспорт", "Нет данных для выгрузки.")
            return

        path = self._ask_save_path(f"ФОТ_по_объектам_{upload_id}.xlsx")
        if not path:
            return

        try:
            grand_total = df["total_amount"].sum()
            df_export = df.copy()
            df_export["share_pct"] = df_export["total_amount"].apply(
                lambda x: round(x / grand_total * 100, 1)
                if grand_total > 0 else 0)
            df_export = df_export.rename(columns={
                "object_name": "Объект",
                "object_type": "Тип объекта",
                "people_cnt": "Кол-во сотрудников",
                "total_hours": "Часов на объекте",
                "total_amount": "Сумма ФОТ, ₽",
                "share_pct": "Доля, %",
            })
            totals = pd.DataFrame([{
                "Объект": "ИТОГО",
                "Тип объекта": "",
                "Кол-во сотрудников": int(
                    df_export["Кол-во сотрудников"].sum()),
                "Часов на объекте": round(
                    df_export["Часов на объекте"].sum(), 1),
                "Сумма ФОТ, ₽": round(df_export["Сумма ФОТ, ₽"].sum(), 2),
                "Доля, %": 100.0,
            }])
            df_export = pd.concat([df_export, totals], ignore_index=True)

            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                df_export.to_excel(
                    writer, index=False, sheet_name="По объектам")
                self._autofit_columns(writer, "По объектам", df_export)

            messagebox.showinfo("Экспорт", f"Файл сохранён:\n{path}")
        except Exception as e:
            logging.exception("Ошибка экспорта по объектам")
            messagebox.showerror("Ошибка",
                                 f"Не удалось сохранить файл:\n{e}")

    def _export_by_dept(self, upload_id: int):
        df = PayrollDataManager.get_distribution_by_department(upload_id)
        if df.empty:
            messagebox.showinfo("Экспорт", "Нет данных для выгрузки.")
            return

        path = self._ask_save_path(
            f"ФОТ_по_подразделениям_{upload_id}.xlsx")
        if not path:
            return

        try:
            df_export = df.copy()
            df_export["diff"] = (
                df_export["total_accrued"] - df_export["total_distributed"])
            df_export["pct"] = df_export.apply(
                lambda r: round(
                    r["total_distributed"] / r["total_accrued"] * 100, 1)
                if r["total_accrued"] > 0 else 0, axis=1)
            df_export = df_export.rename(columns={
                "department_name": "Подразделение",
                "people_cnt": "Кол-во сотрудников",
                "total_accrued": "Начислено, ₽",
                "total_distributed": "Распределено, ₽",
                "diff": "Остаток, ₽",
                "pct": "Распределено, %",
            })
            totals = pd.DataFrame([{
                "Подразделение": "ИТОГО",
                "Кол-во сотрудников": int(
                    df_export["Кол-во сотрудников"].sum()),
                "Начислено, ₽": round(
                    df_export["Начислено, ₽"].sum(), 2),
                "Распределено, ₽": round(
                    df_export["Распределено, ₽"].sum(), 2),
                "Остаток, ₽": round(df_export["Остаток, ₽"].sum(), 2),
                "Распределено, %": "",
            }])
            df_export = pd.concat([df_export, totals], ignore_index=True)

            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                df_export.to_excel(
                    writer, index=False, sheet_name="По подразделениям")
                self._autofit_columns(
                    writer, "По подразделениям", df_export)

            messagebox.showinfo("Экспорт", f"Файл сохранён:\n{path}")
        except Exception as e:
            logging.exception("Ошибка экспорта по подразделениям")
            messagebox.showerror("Ошибка",
                                 f"Не удалось сохранить файл:\n{e}")

    def _export_detail(self, upload_id: int):
        df = PayrollDataManager.get_detail_by_employee(upload_id)
        if df.empty:
            messagebox.showinfo("Экспорт", "Нет данных для выгрузки.")
            return

        path = self._ask_save_path(f"ФОТ_детализация_{upload_id}.xlsx")
        if not path:
            return

        try:
            df_export = df.rename(columns={
                "tbn": "Таб. номер",
                "fio": "ФИО",
                "department_raw": "Подразделение",
                "total_accrued": "Всего начислено",
                "object_name": "Объект",
                "hours_on_object": "Часы на объекте",
                "total_hours_all_objects": "Всего часов",
                "fraction": "Доля",
                "amount": "Сумма на объект, ₽",
            })

            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                df_export.to_excel(
                    writer, index=False, sheet_name="Детализация")
                self._autofit_columns(writer, "Детализация", df_export)

            messagebox.showinfo("Экспорт", f"Файл сохранён:\n{path}")
        except Exception as e:
            logging.exception("Ошибка экспорта детализации")
            messagebox.showerror("Ошибка",
                                 f"Не удалось сохранить файл:\n{e}")

    def _export_unmatched(self, upload_id: int):
        df = PayrollDataManager.get_undistributed_rows(upload_id)
        if df.empty:
            messagebox.showinfo(
                "Экспорт", "Нет нераспределённых сотрудников.")
            return

        path = self._ask_save_path(
            f"ФОТ_нераспределено_{upload_id}.xlsx")
        if not path:
            return

        try:
            df_export = df.rename(columns={
                "tbn": "Таб. номер",
                "fio": "ФИО",
                "department_raw": "Подразделение",
                "position_raw": "Должность",
                "total_accrued": "Начислено, ₽",
            })
            total_lost = (
                df_export["Начислено, ₽"].fillna(0).astype(float).sum())
            totals = pd.DataFrame([{
                "Таб. номер": "",
                "ФИО": "ИТОГО",
                "Подразделение": "",
                "Должность": "",
                "Начислено, ₽": round(total_lost, 2),
            }])
            df_export = pd.concat([df_export, totals], ignore_index=True)

            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                df_export.to_excel(
                    writer, index=False, sheet_name="Не распределено")
                self._autofit_columns(
                    writer, "Не распределено", df_export)

            messagebox.showinfo("Экспорт", f"Файл сохранён:\n{path}")
        except Exception as e:
            logging.exception("Ошибка экспорта нераспределённых")
            messagebox.showerror("Ошибка",
                                 f"Не удалось сохранить файл:\n{e}")

# ============================================================
#  Функция-фабрика для main_app
# ============================================================

def create_payroll_page(parent, app_ref) -> PayrollPage:
    """Фабрика для вызова из main_app._show_page."""
    return PayrollPage(parent, app_ref)
