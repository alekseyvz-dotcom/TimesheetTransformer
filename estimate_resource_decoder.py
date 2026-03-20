import csv
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import tkinter as tk
from tkinter import ttk, messagebox

from openpyxl import Workbook, load_workbook


class EstimateResourceDecoderPage(tk.Frame):
    RESOURCE_TYPE_MAP = {
        1: "ЗП",
        2: "ЭМ",
        3: "МР",
    }

    DELTA_TOLERANCE = 0.05
    DELTA_WARN = 1.0
    DELTA_BAD = 10.0

    def __init__(self, master):
        super().__init__(master, bg="#f7f7f7")

        self.file_path: Optional[Path] = None
        self.workbook = None

        self.local_sheet_name: Optional[str] = None
        self.source_sheet_name: Optional[str] = None
        self.etalon_sheet_name: Optional[str] = None
        self.smtres_sheet_name: Optional[str] = None

        self.local_works: List[Dict[str, Any]] = []
        self.rate_resources: Dict[int, List[Dict[str, Any]]] = {}
        self.code_to_rate_ids: Dict[str, List[int]] = {}

        self.decoded_rows: List[Dict[str, Any]] = []
        self.reconciliation_rows: List[Dict[str, Any]] = []

        self.last_report: str = "Отчет пока не сформирован."

        self._build_ui()

    # ========================= UI =========================

    def _build_ui(self):
        header = tk.Frame(self, bg="#f7f7f7")
        header.pack(fill="x", padx=12, pady=(10, 6))

        tk.Label(
            header,
            text="Раскрытие ресурсов расценок",
            font=("Segoe UI", 16, "bold"),
            bg="#f7f7f7"
        ).pack(side="left")

        ctrl = tk.Frame(self, bg="#f7f7f7")
        ctrl.pack(fill="x", padx=12, pady=(0, 8))

        ttk.Button(ctrl, text="Открыть книгу сметы", command=self._open_file).pack(side="left")
        self.btn_decode = ttk.Button(ctrl, text="Раскрыть ресурсы", command=self._decode_all, state="disabled")
        self.btn_decode.pack(side="left", padx=(8, 0))

        self.btn_export = ttk.Button(ctrl, text="Сохранить результат", command=self._export_result, state="disabled")
        self.btn_export.pack(side="left", padx=(8, 0))

        self.btn_report = ttk.Button(ctrl, text="Показать отчет", command=self._show_report, state="disabled")
        self.btn_report.pack(side="left", padx=(8, 0))

        self.lbl_file = tk.Label(self, text="Файл не выбран", fg="#555", bg="#f7f7f7")
        self.lbl_file.pack(anchor="w", padx=12, pady=(0, 2))

        self.lbl_sheets = tk.Label(self, text="", fg="#777", bg="#f7f7f7", justify="left")
        self.lbl_sheets.pack(anchor="w", padx=12, pady=(0, 8))

        notebook = ttk.Notebook(self)
        notebook.pack(fill="both", expand=True, padx=12, pady=(0, 12))

        self.tab_resources = tk.Frame(notebook, bg="#f7f7f7")
        self.tab_recon = tk.Frame(notebook, bg="#f7f7f7")

        notebook.add(self.tab_resources, text="Ресурсы")
        notebook.add(self.tab_recon, text="Сверка")

        self._build_resources_tab()
        self._build_recon_tab()

    def _build_resources_tab(self):
        filter_frame = tk.Frame(self.tab_resources, bg="#f7f7f7")
        filter_frame.pack(fill="x", padx=0, pady=(0, 6))

        tk.Label(filter_frame, text="Фильтр по позиции:", bg="#f7f7f7").pack(side="left")
        self.var_pos_filter = tk.StringVar()
        ent = ttk.Entry(filter_frame, textvariable=self.var_pos_filter, width=12)
        ent.pack(side="left", padx=(6, 8))
        ent.bind("<KeyRelease>", lambda e: self._fill_resource_tree())

        tk.Label(filter_frame, text="Фильтр по шифру:", bg="#f7f7f7").pack(side="left")
        self.var_code_filter = tk.StringVar()
        ent2 = ttk.Entry(filter_frame, textvariable=self.var_code_filter, width=18)
        ent2.pack(side="left", padx=(6, 8))
        ent2.bind("<KeyRelease>", lambda e: self._fill_resource_tree())

        tk.Label(filter_frame, text="Тип ресурса:", bg="#f7f7f7").pack(side="left")
        self.cmb_type = ttk.Combobox(filter_frame, state="readonly", width=12, values=["Все", "ЗП", "ЭМ", "МР"])
        self.cmb_type.current(0)
        self.cmb_type.pack(side="left", padx=(6, 8))
        self.cmb_type.bind("<<ComboboxSelected>>", lambda e: self._fill_resource_tree())

        wrap = tk.Frame(self.tab_resources)
        wrap.pack(fill="both", expand=True)

        cols = (
            "work_num", "work_code", "rate_id", "res_type",
            "res_code", "res_name", "unit", "norm_qty",
            "work_qty", "res_qty", "price", "base_cost"
        )

        self.tree_resources = ttk.Treeview(wrap, columns=cols, show="headings")
        self.tree_resources.pack(side="left", fill="both", expand=True)
        self.tree_resources.bind("<Double-1>", self._on_resource_double_click)

        headers = {
            "work_num": "Поз.",
            "work_code": "Шифр",
            "rate_id": "Rate ID",
            "res_type": "Тип",
            "res_code": "Код ресурса",
            "res_name": "Наименование ресурса",
            "unit": "Ед.",
            "norm_qty": "Норма",
            "work_qty": "Кол-во поз.",
            "res_qty": "Расход",
            "price": "Цена",
            "base_cost": "Стоимость"
        }

        widths = {
            "work_num": 60,
            "work_code": 120,
            "rate_id": 70,
            "res_type": 60,
            "res_code": 110,
            "res_name": 420,
            "unit": 70,
            "norm_qty": 90,
            "work_qty": 90,
            "res_qty": 90,
            "price": 90,
            "base_cost": 110
        }

        for c in cols:
            self.tree_resources.heading(c, text=headers[c])
            self.tree_resources.column(
                c,
                width=widths[c],
                anchor="w" if c not in ("norm_qty", "work_qty", "res_qty", "price", "base_cost") else "e"
            )

        yscroll = ttk.Scrollbar(wrap, orient="vertical", command=self.tree_resources.yview)
        self.tree_resources.configure(yscrollcommand=yscroll.set)
        yscroll.pack(side="right", fill="y")

    def _build_recon_tab(self):
        top = tk.Frame(self.tab_recon, bg="#f7f7f7")
        top.pack(fill="x", pady=(0, 6))

        tk.Label(top, text="Фильтр по позиции:", bg="#f7f7f7").pack(side="left")
        self.var_recon_pos = tk.StringVar()
        ent = ttk.Entry(top, textvariable=self.var_recon_pos, width=12)
        ent.pack(side="left", padx=(6, 8))
        ent.bind("<KeyRelease>", lambda e: self._fill_recon_tree())

        tk.Label(top, text="Фильтр по шифру:", bg="#f7f7f7").pack(side="left")
        self.var_recon_code = tk.StringVar()
        ent2 = ttk.Entry(top, textvariable=self.var_recon_code, width=18)
        ent2.pack(side="left", padx=(6, 8))
        ent2.bind("<KeyRelease>", lambda e: self._fill_recon_tree())

        tk.Label(top, text="Статус:", bg="#f7f7f7").pack(side="left")
        self.cmb_status = ttk.Combobox(top, state="readonly", width=12, values=["Все", "OK", "Δ"])
        self.cmb_status.current(0)
        self.cmb_status.pack(side="left", padx=(6, 8))
        self.cmb_status.bind("<<ComboboxSelected>>", lambda e: self._fill_recon_tree())

        hint = tk.Label(
            top,
            text="Двойной клик по строке — карточка позиции и фильтр ресурсов",
            bg="#f7f7f7",
            fg="#666"
        )
        hint.pack(side="right")

        wrap = tk.Frame(self.tab_recon)
        wrap.pack(fill="both", expand=True)

        cols = (
            "pos_num", "work_code", "rate_id",
            "zp_local", "zp_res", "zp_delta",
            "em_local", "em_res", "em_delta",
            "mr_local", "mr_res", "mr_delta",
            "zpm_local", "status"
        )

        self.tree_recon = ttk.Treeview(wrap, columns=cols, show="headings")
        self.tree_recon.pack(side="left", fill="both", expand=True)
        self.tree_recon.bind("<Double-1>", self._on_recon_double_click)

        headers = {
            "pos_num": "Поз.",
            "work_code": "Шифр",
            "rate_id": "Rate ID",
            "zp_local": "ЗП лок.",
            "zp_res": "ЗП раскр.",
            "zp_delta": "Δ ЗП",
            "em_local": "ЭМ лок.",
            "em_res": "ЭМ раскр.",
            "em_delta": "Δ ЭМ",
            "mr_local": "МР лок.",
            "mr_res": "МР раскр.",
            "mr_delta": "Δ МР",
            "zpm_local": "ЗПМ лок.",
            "status": "Статус"
        }

        widths = {
            "pos_num": 60,
            "work_code": 120,
            "rate_id": 70,
            "zp_local": 90,
            "zp_res": 90,
            "zp_delta": 90,
            "em_local": 90,
            "em_res": 90,
            "em_delta": 90,
            "mr_local": 90,
            "mr_res": 90,
            "mr_delta": 90,
            "zpm_local": 90,
            "status": 80
        }

        for c in cols:
            self.tree_recon.heading(c, text=headers[c])
            self.tree_recon.column(c, width=widths[c], anchor="e" if c not in ("pos_num", "work_code", "rate_id", "status") else "w")

        self.tree_recon.tag_configure("ok", background="#edf7ed")
        self.tree_recon.tag_configure("warn", background="#fff8e1")
        self.tree_recon.tag_configure("bad", background="#fdecea")

        yscroll = ttk.Scrollbar(wrap, orient="vertical", command=self.tree_recon.yview)
        self.tree_recon.configure(yscrollcommand=yscroll.set)
        yscroll.pack(side="right", fill="y")

    # ========================= FILE =========================

    def _open_file(self):
        try:
            from tkinter import filedialog as fd
        except Exception:
            messagebox.showerror("Файл", "Не удалось открыть диалог выбора файла.")
            return

        fname = fd.askopenfilename(
            title="Выберите Excel-книгу со сметой",
            filetypes=[("Excel", "*.xlsx;*.xlsm"), ("Все файлы", "*.*")]
        )
        if not fname:
            return

        self.file_path = Path(fname)
        self.lbl_file.config(text=f"Файл: {self.file_path}")

        try:
            self._load_workbook()
            self.btn_decode.config(state="normal")
            self.btn_report.config(state="normal")
            self.btn_export.config(state="disabled")
            self.decoded_rows = []
            self.reconciliation_rows = []
            self._fill_resource_tree()
            self._fill_recon_tree()
            messagebox.showinfo("Готово", "Книга загружена. Теперь можно выполнить раскрытие ресурсов.")
        except Exception as e:
            self.btn_decode.config(state="disabled")
            self.btn_export.config(state="disabled")
            self.btn_report.config(state="disabled")
            messagebox.showerror("Ошибка", f"Не удалось загрузить книгу:\n{e}")

    def _load_workbook(self):
        if not self.file_path:
            raise RuntimeError("Файл не выбран")

        self.workbook = load_workbook(self.file_path, data_only=True)

        sheetnames = self.workbook.sheetnames
        self.local_sheet_name = self._detect_local_sheet(sheetnames)
        self.source_sheet_name = self._detect_source_sheet(sheetnames)
        self.etalon_sheet_name = self._detect_etalon_sheet(sheetnames)
        self.smtres_sheet_name = self._detect_smtres_sheet(sheetnames)

        info = [
            f"Локальная смета: {self.local_sheet_name or 'не найдено'}",
            f"Source: {self.source_sheet_name or 'не найдено'}",
            f"EtalonRes: {self.etalon_sheet_name or 'не найдено'}",
            f"SmtRes: {self.smtres_sheet_name or 'не найдено'}",
        ]
        self.lbl_sheets.config(text="\n".join(info))

    # ========================= DETECT / HELPERS =========================

    def _normalize_rate_code(self, raw: Any) -> str:
        s = self._s(raw)
        if not s:
            return ""
        s = s.replace("\r", "\n")
        first = s.split("\n")[0].strip()
        m = re.search(r"\d+\.\d+(?:-\d+)+", first)
        return m.group(0) if m else first

    def _detect_local_sheet(self, names: List[str]) -> Optional[str]:
        for name in names:
            ws = self.workbook[name]
            if self._sheet_has_local_smeta_marker(ws):
                return name
        return names[0] if names else None

    def _detect_source_sheet(self, names: List[str]) -> Optional[str]:
        for name in names:
            low = name.lower()
            if "sourse" in low or "source" in low:
                return name
        return None

    def _detect_etalon_sheet(self, names: List[str]) -> Optional[str]:
        for name in names:
            low = name.lower()
            if "etalon" in low:
                return name
        return None

    def _detect_smtres_sheet(self, names: List[str]) -> Optional[str]:
        for name in names:
            low = name.lower()
            if "smtres" in low:
                return name
        return None

    @staticmethod
    def _sheet_has_local_smeta_marker(ws) -> bool:
        try:
            for row in ws.iter_rows(min_row=1, max_row=30, values_only=True):
                for c in row:
                    if isinstance(c, str) and "локальная смета" in c.lower():
                        return True
        except Exception:
            pass
        return False

    @staticmethod
    def _s(v: Any) -> str:
        return str(v or "").strip()

    @staticmethod
    def _f(v: Any) -> Optional[float]:
        if v is None or v == "":
            return None
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip().replace(" ", "").replace("\u00A0", "").replace(",", ".")
        s = re.sub(r"[^0-9.\-]", "", s)
        if not s:
            return None
        try:
            return float(s)
        except Exception:
            return None

    @staticmethod
    def _try_parse_int(v: Any) -> Optional[int]:
        if v is None:
            return None
        if isinstance(v, int):
            return v
        if isinstance(v, float) and float(v).is_integer():
            return int(v)
        s = str(v).strip()
        if re.fullmatch(r"\d+", s):
            return int(s)
        return None

    @staticmethod
    def _is_main_pos(v: Any) -> bool:
        if isinstance(v, int):
            return True
        if isinstance(v, float) and float(v).is_integer():
            return True
        s = str(v or "").strip()
        return bool(re.fullmatch(r"\d+", s))

    @staticmethod
    def _looks_like_rate_code(s: str) -> bool:
        s = (s or "").strip()
        return bool(re.fullmatch(r"\d+\.\d+(?:-\d+)+", s))

    @staticmethod
    def _sort_pos(v: Any):
        s = str(v or "")
        try:
            return (0, float(s.replace(",", ".")))
        except Exception:
            return (1, s)

    @staticmethod
    def _fmt_num(v: Any) -> str:
        if v is None:
            return ""
        try:
            return f"{float(v):,.2f}".replace(",", " ").replace(".", ",")
        except Exception:
            return str(v)

    def _extract_unit(self, vals: List[Any]) -> str:
        for i in [14, 15, 13, 16]:
            if i < len(vals):
                s = self._s(vals[i])
                if s and len(s) <= 20:
                    return s
        return ""

    def _extract_norm_qty(self, vals: List[Any]) -> Optional[float]:
        for i in [23, 24, 22, 34]:
            if i < len(vals):
                f = self._f(vals[i])
                if f is not None:
                    return f
        return None

    def _extract_price(self, vals: List[Any]) -> Optional[float]:
        for i in [25, 26, 27]:
            if i < len(vals):
                f = self._f(vals[i])
                if f is not None and f >= 0:
                    return f
        return None

    # ========================= CORE =========================

    def _decode_all(self):
        if not self.workbook:
            return

        try:
            self.local_works = self._parse_local_works()
            self.rate_resources = self._parse_rate_resources()
            self.code_to_rate_ids = self._build_code_to_rate_map()
            self.decoded_rows = self._decode_rows()
            self.reconciliation_rows = self._build_reconciliation()

            self._fill_resource_tree()
            self._fill_recon_tree()
            self.btn_export.config(state="normal")

            messagebox.showinfo(
                "Готово",
                f"Раскрытие выполнено.\n"
                f"Позиции: {len(self.local_works)}\n"
                f"Ресурсных строк: {len(self.decoded_rows)}\n"
                f"Сверок: {len(self.reconciliation_rows)}"
            )
        except Exception as e:
            messagebox.showerror("Ошибка раскрытия", str(e))

    def _parse_local_works(self) -> List[Dict[str, Any]]:
        if not self.local_sheet_name:
            raise RuntimeError("Не найден лист локальной сметы.")

        ws = self.workbook[self.local_sheet_name]
        works: List[Dict[str, Any]] = []
        current_section = ""
        current_work: Optional[Dict[str, Any]] = None
        header_found = False

        for row in ws.iter_rows(values_only=True):
            row = list(row)

            if not any(v is not None and str(v).strip() for v in row):
                continue

            first = self._s(row[0]) if len(row) > 0 else ""
            third = self._s(row[2]) if len(row) > 2 else ""

            if first == "№ п/п" or third == "Наименование работ и затрат":
                header_found = True
                continue

            if not header_found:
                continue

            if first.startswith("Раздел:"):
                current_section = first.replace("Раздел:", "").strip()
                continue

            if self._is_main_pos(row[0] if len(row) > 0 else None):
                pos_num = self._s(row[0])
                code = self._normalize_rate_code(row[1] if len(row) > 1 else "")
                name = self._s(row[2] if len(row) > 2 else "")
                unit = self._s(row[3] if len(row) > 3 else "")
                qty = self._f(row[4] if len(row) > 4 else None)

                # строка нумерации колонок
                if code == "2" and name == "3" and self._s(row[3] if len(row) > 3 else "") == "4":
                    continue

                if not name:
                    continue
                if name.lower().startswith("итого"):
                    continue
                if name.lower().startswith("вес "):
                    continue

                current_work = {
                    "pos_num": pos_num,
                    "work_code": code,
                    "name": name,
                    "unit": unit,
                    "qty": qty or 0.0,
                    "section": current_section,
                    "zp_base": None,
                    "em_base": None,
                    "mr_base": None,
                    "zpm_base": None,
                }
                works.append(current_work)
                continue

            if current_work is None:
                continue

            if third == "ЗП":
                current_work["zp_base"] = self._f(row[8] if len(row) > 8 else None)
            elif third == "ЭМ":
                current_work["em_base"] = self._f(row[8] if len(row) > 8 else None)
            elif third == "МР":
                current_work["mr_base"] = self._f(row[8] if len(row) > 8 else None)
            elif third == "в т.ч. ЗПМ":
                current_work["zpm_base"] = self._f(row[8] if len(row) > 8 else None)

        return works

    def _parse_rate_resources(self) -> Dict[int, List[Dict[str, Any]]]:
        sheet_name = self.etalon_sheet_name or self.smtres_sheet_name
        if not sheet_name:
            raise RuntimeError("Не найден лист EtalonRes/SmtRes.")

        ws = self.workbook[sheet_name]
        result: Dict[int, List[Dict[str, Any]]] = {}

        for row in ws.iter_rows(values_only=True):
            vals = list(row)
            if not any(v is not None and str(v).strip() for v in vals):
                continue

            rate_id = self._try_parse_int(vals[0] if len(vals) > 0 else None)
            if rate_id is None:
                continue

            resource_type = self._try_parse_int(vals[7] if len(vals) > 7 else None)
            resource_code = self._s(vals[8] if len(vals) > 8 else "")
            resource_name = self._s(vals[10] if len(vals) > 10 else "")
            unit = self._extract_unit(vals)
            norm_qty = self._extract_norm_qty(vals)
            price = self._extract_price(vals)

            if resource_type not in (1, 2, 3):
                continue
            if not resource_code and not resource_name:
                continue
            if norm_qty is None:
                continue

            result.setdefault(rate_id, []).append({
                "rate_id": rate_id,
                "resource_type_id": resource_type,
                "resource_type": self.RESOURCE_TYPE_MAP.get(resource_type, str(resource_type)),
                "resource_code": resource_code,
                "resource_name": resource_name,
                "unit": unit,
                "norm_qty": norm_qty,
                "price": price,
            })

        return result

    def _build_code_to_rate_map(self) -> Dict[str, List[int]]:
        code_map: Dict[str, List[int]] = {}

        def add_pair(code: str, rate_id: int):
            if not code or rate_id is None:
                return
            code_map.setdefault(code, [])
            if rate_id not in code_map[code]:
                code_map[code].append(rate_id)

        # 1) Source
        if self.source_sheet_name:
            ws = self.workbook[self.source_sheet_name]

            for row in ws.iter_rows(values_only=True):
                vals = list(row)
                if not any(v is not None and str(v).strip() for v in vals):
                    continue

                codes = []
                for v in vals:
                    code = self._normalize_rate_code(v)
                    if self._looks_like_rate_code(code):
                        codes.append(code)
                if not codes:
                    continue

                rate_candidates = []
                for i in range(min(8, len(vals))):
                    rid = self._try_parse_int(vals[i])
                    if rid is not None and rid > 0:
                        rate_candidates.append(rid)

                if not rate_candidates:
                    continue

                # чуть более осторожный выбор
                # сначала ищем кандидатов > 100, потом > 10, потом max
                big = [x for x in rate_candidates if x >= 100]
                mid = [x for x in rate_candidates if x >= 10]

                if big:
                    rate_id = max(big)
                elif mid:
                    rate_id = max(mid)
                else:
                    rate_id = max(rate_candidates)

                for code in codes:
                    add_pair(code, rate_id)

        # 2) fallback
        if not code_map:
            for sheet_name in [self.etalon_sheet_name, self.smtres_sheet_name]:
                if not sheet_name:
                    continue

                ws = self.workbook[sheet_name]

                for row in ws.iter_rows(values_only=True):
                    vals = list(row)
                    if not any(v is not None and str(v).strip() for v in vals):
                        continue

                    rate_id = self._try_parse_int(vals[0] if len(vals) > 0 else None)
                    if rate_id is None or rate_id <= 0:
                        continue

                    codes = []
                    for v in vals:
                        code = self._normalize_rate_code(v)
                        if self._looks_like_rate_code(code):
                            codes.append(code)
                    for code in codes:
                        add_pair(code, rate_id)

        return code_map

    def _decode_rows(self) -> List[Dict[str, Any]]:
        rows: List[Dict[str, Any]] = []
        report_lines = []

        report_lines.append("ОТЧЕТ ПО РАСКРЫТИЮ РЕСУРСОВ")
        report_lines.append("=" * 100)
        report_lines.append(f"Локальный лист: {self.local_sheet_name}")
        report_lines.append(f"Source лист: {self.source_sheet_name}")
        report_lines.append(f"EtalonRes лист: {self.etalon_sheet_name}")
        report_lines.append(f"SmtRes лист: {self.smtres_sheet_name}")
        report_lines.append("")
        report_lines.append("КАРТА code -> rate_ids")
        for code in sorted(self.code_to_rate_ids.keys()):
            report_lines.append(f"  {code} -> {self.code_to_rate_ids[code]}")
        report_lines.append("")

        matched = 0
        unmatched = 0

        for work in self.local_works:
            code = work.get("work_code") or ""
            qty = work.get("qty") or 0.0

            if not code:
                unmatched += 1
                report_lines.append(f"[NO CODE] Поз. {work['pos_num']} | {work['name']}")
                continue

            rate_ids = self.code_to_rate_ids.get(code, [])
            if not rate_ids:
                unmatched += 1
                report_lines.append(f"[NO RATE_ID] Поз. {work['pos_num']} | {code} | {work['name']}")
                continue

            rate_id = self._choose_best_rate_id(work, rate_ids)
            resources = self.rate_resources.get(rate_id, [])

            if not resources:
                unmatched += 1
                report_lines.append(f"[NO RESOURCES] Поз. {work['pos_num']} | {code} | rate_id={rate_id}")
                continue

            matched += 1
            report_lines.append(f"[OK] Поз. {work['pos_num']} | {code} -> rate_id={rate_id} | ресурсов={len(resources)}")

            for res in resources:
                norm_qty = res.get("norm_qty") or 0.0
                price = res.get("price")
                res_qty = norm_qty * qty
                base_cost = (res_qty * price) if isinstance(price, (int, float)) else None

                rows.append({
                    "work_num": work["pos_num"],
                    "work_code": code,
                    "work_name": work["name"],
                    "work_qty": qty,
                    "rate_id": rate_id,
                    "resource_type": res["resource_type"],
                    "resource_type_id": res["resource_type_id"],
                    "resource_code": res["resource_code"],
                    "resource_name": res["resource_name"],
                    "unit": res["unit"],
                    "norm_qty": norm_qty,
                    "resource_qty": res_qty,
                    "price": price,
                    "base_cost": base_cost,
                    "section": work.get("section", ""),
                    "zp_base": work.get("zp_base"),
                    "em_base": work.get("em_base"),
                    "mr_base": work.get("mr_base"),
                    "zpm_base": work.get("zpm_base"),
                })

        report_lines.append("")
        report_lines.append(f"Всего позиций: {len(self.local_works)}")
        report_lines.append(f"Сопоставлено: {matched}")
        report_lines.append(f"Не сопоставлено: {unmatched}")

        self.last_report = "\n".join(report_lines)
        return rows

    def _choose_best_rate_id(self, work: Dict[str, Any], rate_ids: List[int]) -> int:
        if not rate_ids:
            raise RuntimeError("Пустой список rate_id")

        if len(rate_ids) == 1:
            return rate_ids[0]

        local_qty = work.get("qty") or 0.0
        local_zp = work.get("zp_base")
        local_em = work.get("em_base")
        local_mr = work.get("mr_base")

        best_rate_id = rate_ids[0]
        best_score = None

        for rid in rate_ids:
            resources = self.rate_resources.get(rid, [])
            if not resources:
                continue

            zp_sum = 0.0
            em_sum = 0.0
            mr_sum = 0.0

            for res in resources:
                price = res.get("price")
                if price is None:
                    continue

                cost = (res.get("norm_qty") or 0.0) * local_qty * price

                if res["resource_type"] == "ЗП":
                    zp_sum += cost
                elif res["resource_type"] == "ЭМ":
                    em_sum += cost
                elif res["resource_type"] == "МР":
                    mr_sum += cost

            score = 0.0

            if local_em is not None:
                score += abs(local_em - em_sum) * 100.0
            if local_zp is not None:
                score += abs(local_zp - zp_sum) * 10.0
            if local_mr is not None:
                score += abs(local_mr - mr_sum) * 10.0

            score -= min(len(resources), 20) * 0.01

            if best_score is None or score < best_score:
                best_score = score
                best_rate_id = rid

        return best_rate_id

    def _build_reconciliation(self) -> List[Dict[str, Any]]:
        by_pos: Dict[str, Dict[str, Any]] = {}
        works_by_pos = {w["pos_num"]: w for w in self.local_works}

        for row in self.decoded_rows:
            pos = row["work_num"]
            rec = by_pos.setdefault(pos, {
                "pos_num": pos,
                "work_code": row["work_code"],
                "work_name": row["work_name"],
                "rate_id": row["rate_id"],
                "zp_res": 0.0,
                "em_res": 0.0,
                "mr_res": 0.0,
            })

            cost = row.get("base_cost") or 0.0
            if row["resource_type"] == "ЗП":
                rec["zp_res"] += cost
            elif row["resource_type"] == "ЭМ":
                rec["em_res"] += cost
            elif row["resource_type"] == "МР":
                rec["mr_res"] += cost

        result = []
        for pos, work in works_by_pos.items():
            rec = by_pos.get(pos, {
                "pos_num": pos,
                "work_code": work["work_code"],
                "work_name": work["name"],
                "rate_id": None,
                "zp_res": 0.0,
                "em_res": 0.0,
                "mr_res": 0.0,
            })

            zp_local = work.get("zp_base") or 0.0
            em_local = work.get("em_base") or 0.0
            mr_local = work.get("mr_base") or 0.0
            zpm_local = work.get("zpm_base") or 0.0

            zp_delta = zp_local - rec["zp_res"]
            em_delta = em_local - rec["em_res"]
            mr_delta = mr_local - rec["mr_res"]

            max_abs = max(abs(zp_delta), abs(em_delta), abs(mr_delta))

            if max_abs <= self.DELTA_TOLERANCE:
                status = "OK"
                severity = "ok"
            elif max_abs <= self.DELTA_WARN:
                status = "Δ"
                severity = "warn"
            else:
                status = "Δ"
                severity = "bad"

            result.append({
                "pos_num": pos,
                "work_code": work["work_code"],
                "work_name": work["name"],
                "rate_id": rec["rate_id"],
                "zp_local": zp_local,
                "zp_res": rec["zp_res"],
                "zp_delta": zp_delta,
                "em_local": em_local,
                "em_res": rec["em_res"],
                "em_delta": em_delta,
                "mr_local": mr_local,
                "mr_res": rec["mr_res"],
                "mr_delta": mr_delta,
                "zpm_local": zpm_local,
                "status": status,
                "severity": severity,
            })

        result.sort(key=lambda x: self._sort_pos(x["pos_num"]))
        self._append_recon_to_report(result)
        return result

    def _append_recon_to_report(self, rows: List[Dict[str, Any]]):
        lines = [self.last_report, "", "=" * 100, "СВЕРКА ЛОКАЛКА ↔ РАСКРЫТЫЕ РЕСУРСЫ", "=" * 100]

        ok_count = sum(1 for r in rows if r["status"] == "OK")
        delta_count = len(rows) - ok_count

        lines.append(f"Всего строк сверки: {len(rows)}")
        lines.append(f"OK: {ok_count}")
        lines.append(f"С отклонениями: {delta_count}")
        lines.append("")

        bad_rows = [r for r in rows if r["status"] != "OK"]
        for r in bad_rows[:150]:
            lines.append(
                f"Поз. {r['pos_num']} | {r['work_code']} | rate_id={r['rate_id']} | "
                f"ΔЗП={r['zp_delta']:.2f} | ΔЭМ={r['em_delta']:.2f} | ΔМР={r['mr_delta']:.2f}"
            )

        if len(bad_rows) > 150:
            lines.append(f"... и еще {len(bad_rows) - 150} строк с отклонениями")

        self.last_report = "\n".join(lines)

    # ========================= TREES =========================

    def _fill_resource_tree(self):
        for item in self.tree_resources.get_children():
            self.tree_resources.delete(item)

        pos_filter = self.var_pos_filter.get().strip().lower()
        code_filter = self.var_code_filter.get().strip().lower()
        type_filter = self.cmb_type.get().strip()

        for row in self.decoded_rows:
            if pos_filter and pos_filter not in str(row["work_num"]).lower():
                continue
            if code_filter and code_filter not in str(row["work_code"]).lower():
                continue
            if type_filter != "Все" and row["resource_type"] != type_filter:
                continue

            self.tree_resources.insert(
                "",
                "end",
                values=(
                    row["work_num"],
                    row["work_code"],
                    row["rate_id"],
                    row["resource_type"],
                    row["resource_code"],
                    row["resource_name"],
                    row["unit"],
                    self._fmt_num(row["norm_qty"]),
                    self._fmt_num(row["work_qty"]),
                    self._fmt_num(row["resource_qty"]),
                    self._fmt_num(row["price"]),
                    self._fmt_num(row["base_cost"]),
                )
            )

    def _fill_recon_tree(self):
        for item in self.tree_recon.get_children():
            self.tree_recon.delete(item)

        pos_filter = self.var_recon_pos.get().strip().lower()
        code_filter = self.var_recon_code.get().strip().lower()
        status_filter = self.cmb_status.get().strip()

        for row in self.reconciliation_rows:
            if pos_filter and pos_filter not in str(row["pos_num"]).lower():
                continue
            if code_filter and code_filter not in str(row["work_code"]).lower():
                continue
            if status_filter != "Все" and row["status"] != status_filter:
                continue

            tag = row.get("severity", "bad")

            self.tree_recon.insert(
                "",
                "end",
                values=(
                    row["pos_num"],
                    row["work_code"],
                    row["rate_id"] if row["rate_id"] is not None else "",
                    self._fmt_num(row["zp_local"]),
                    self._fmt_num(row["zp_res"]),
                    self._fmt_num(row["zp_delta"]),
                    self._fmt_num(row["em_local"]),
                    self._fmt_num(row["em_res"]),
                    self._fmt_num(row["em_delta"]),
                    self._fmt_num(row["mr_local"]),
                    self._fmt_num(row["mr_res"]),
                    self._fmt_num(row["mr_delta"]),
                    self._fmt_num(row["zpm_local"]),
                    row["status"],
                ),
                tags=(tag,)
            )

    # ========================= DETAILS =========================

    def _on_recon_double_click(self, event):
        item = self.tree_recon.focus()
        if not item:
            return
        vals = self.tree_recon.item(item, "values")
        if not vals:
            return

        pos_num = str(vals[0])
        self.var_pos_filter.set(pos_num)
        self._fill_resource_tree()
        self._open_position_card(pos_num)

    def _on_resource_double_click(self, event):
        item = self.tree_resources.focus()
        if not item:
            return
        vals = self.tree_resources.item(item, "values")
        if not vals:
            return
        pos_num = str(vals[0])
        self._open_position_card(pos_num)

    def _open_position_card(self, pos_num: str):
        work = next((w for w in self.local_works if str(w["pos_num"]) == str(pos_num)), None)
        recon = next((r for r in self.reconciliation_rows if str(r["pos_num"]) == str(pos_num)), None)
        resources = [r for r in self.decoded_rows if str(r["work_num"]) == str(pos_num)]

        if not work:
            return

        win = tk.Toplevel(self)
        win.title(f"Карточка позиции {pos_num}")
        win.geometry("1300x800")

        top = tk.Frame(win, bg="#f7f7f7")
        top.pack(fill="x", padx=10, pady=10)

        title = f"Позиция {work['pos_num']} | {work.get('work_code', '')}"
        tk.Label(top, text=title, font=("Segoe UI", 14, "bold"), bg="#f7f7f7").pack(anchor="w")
        tk.Label(top, text=work.get("name", ""), font=("Segoe UI", 10), bg="#f7f7f7", wraplength=1200, justify="left").pack(anchor="w", pady=(4, 0))
        tk.Label(top, text=f"Ед.: {work.get('unit', '')}    Кол-во: {self._fmt_num(work.get('qty'))}    Раздел: {work.get('section', '')}", bg="#f7f7f7", fg="#555").pack(anchor="w", pady=(4, 0))

        card = tk.Frame(win, bg="#ffffff", bd=1, relief="solid")
        card.pack(fill="x", padx=10, pady=(0, 10))

        grid = tk.Frame(card, bg="#ffffff")
        grid.pack(fill="x", padx=10, pady=10)

        tk.Label(grid, text="Показатель", bg="#ffffff", font=("Segoe UI", 10, "bold")).grid(row=0, column=0, sticky="w", padx=4, pady=4)
        tk.Label(grid, text="Локалка", bg="#ffffff", font=("Segoe UI", 10, "bold")).grid(row=0, column=1, sticky="e", padx=4, pady=4)
        tk.Label(grid, text="Раскрыто", bg="#ffffff", font=("Segoe UI", 10, "bold")).grid(row=0, column=2, sticky="e", padx=4, pady=4)
        tk.Label(grid, text="Δ", bg="#ffffff", font=("Segoe UI", 10, "bold")).grid(row=0, column=3, sticky="e", padx=4, pady=4)

        rows = []
        if recon:
            rows = [
                ("ЗП", recon["zp_local"], recon["zp_res"], recon["zp_delta"]),
                ("ЭМ", recon["em_local"], recon["em_res"], recon["em_delta"]),
                ("МР", recon["mr_local"], recon["mr_res"], recon["mr_delta"]),
            ]
        else:
            rows = [
                ("ЗП", work.get("zp_base") or 0.0, 0.0, work.get("zp_base") or 0.0),
                ("ЭМ", work.get("em_base") or 0.0, 0.0, work.get("em_base") or 0.0),
                ("МР", work.get("mr_base") or 0.0, 0.0, work.get("mr_base") or 0.0),
            ]

        for i, (label, local_v, res_v, delta_v) in enumerate(rows, start=1):
            fg = "#2e7d32" if abs(delta_v) <= self.DELTA_TOLERANCE else ("#ef6c00" if abs(delta_v) <= self.DELTA_WARN else "#c62828")
            tk.Label(grid, text=label, bg="#ffffff").grid(row=i, column=0, sticky="w", padx=4, pady=4)
            tk.Label(grid, text=self._fmt_num(local_v), bg="#ffffff").grid(row=i, column=1, sticky="e", padx=4, pady=4)
            tk.Label(grid, text=self._fmt_num(res_v), bg="#ffffff").grid(row=i, column=2, sticky="e", padx=4, pady=4)
            tk.Label(grid, text=self._fmt_num(delta_v), bg="#ffffff", fg=fg, font=("Segoe UI", 10, "bold")).grid(row=i, column=3, sticky="e", padx=4, pady=4)

        tk.Label(grid, text="ЗПМ (лок., справочно)", bg="#ffffff").grid(row=4, column=0, sticky="w", padx=4, pady=4)
        tk.Label(grid, text=self._fmt_num(work.get("zpm_base") or 0.0), bg="#ffffff").grid(row=4, column=1, sticky="e", padx=4, pady=4)

        wrap = tk.Frame(win)
        wrap.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        cols = ("type", "code", "name", "unit", "norm", "qty", "price", "cost")
        tree = ttk.Treeview(wrap, columns=cols, show="headings")
        tree.pack(side="left", fill="both", expand=True)

        headers = {
            "type": "Тип",
            "code": "Код",
            "name": "Наименование",
            "unit": "Ед.",
            "norm": "Норма",
            "qty": "Расход",
            "price": "Цена",
            "cost": "Стоимость",
        }
        widths = {
            "type": 60, "code": 120, "name": 540, "unit": 70,
            "norm": 90, "qty": 90, "price": 90, "cost": 110,
        }

        for c in cols:
            tree.heading(c, text=headers[c])
            tree.column(c, width=widths[c], anchor="w" if c in ("type", "code", "name", "unit") else "e")

        tree.tag_configure("ЗП", background="#eef7ff")
        tree.tag_configure("ЭМ", background="#fff8e1")
        tree.tag_configure("МР", background="#edf7ed")

        for r in resources:
            tree.insert(
                "",
                "end",
                values=(
                    r["resource_type"],
                    r["resource_code"],
                    r["resource_name"],
                    r["unit"],
                    self._fmt_num(r["norm_qty"]),
                    self._fmt_num(r["resource_qty"]),
                    self._fmt_num(r["price"]),
                    self._fmt_num(r["base_cost"]),
                ),
                tags=(r["resource_type"],)
            )

        yscroll = ttk.Scrollbar(wrap, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=yscroll.set)
        yscroll.pack(side="right", fill="y")

    # ========================= REPORT / EXPORT =========================

    def _show_report(self):
        win = tk.Toplevel(self)
        win.title("Отчет по раскрытию ресурсов")
        win.geometry("1000x700")

        wrap = tk.Frame(win)
        wrap.pack(fill="both", expand=True, padx=10, pady=10)

        txt = tk.Text(wrap, wrap="word", font=("Consolas", 9))
        txt.pack(side="left", fill="both", expand=True)

        scr = ttk.Scrollbar(wrap, orient="vertical", command=txt.yview)
        scr.pack(side="right", fill="y")
        txt.configure(yscrollcommand=scr.set)

        txt.insert("1.0", self.last_report or "")
        txt.configure(state="disabled")

    def _export_result(self):
        if not self.decoded_rows and not self.reconciliation_rows:
            messagebox.showwarning("Экспорт", "Нет данных для сохранения.")
            return

        try:
            from tkinter import filedialog as fd
        except Exception:
            messagebox.showerror("Экспорт", "Не удалось открыть диалог сохранения.")
            return

        fname = fd.asksaveasfilename(
            title="Сохранить раскрытие ресурсов",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv")]
        )
        if not fname:
            return

        out = Path(fname)

        res_headers = [
            "Поз.", "Шифр расценки", "Наименование работы", "Кол-во позиции",
            "Rate ID", "Тип ресурса", "Код ресурса", "Наименование ресурса",
            "Ед.", "Норма", "Расход по позиции", "Цена", "Базовая стоимость"
        ]

        recon_headers = [
            "Поз.", "Шифр", "Наименование", "Rate ID",
            "ЗП лок.", "ЗП раскр.", "Δ ЗП",
            "ЭМ лок.", "ЭМ раскр.", "Δ ЭМ",
            "МР лок.", "МР раскр.", "Δ МР",
            "ЗПМ лок.", "Статус"
        ]

        try:
            if out.suffix.lower() == ".csv":
                base = out.with_suffix("")
                res_csv = base.with_name(base.name + "_resources.csv")
                recon_csv = base.with_name(base.name + "_recon.csv")

                with open(res_csv, "w", encoding="utf-8-sig", newline="") as f:
                    w = csv.writer(f, delimiter=";")
                    w.writerow(res_headers)
                    for r in self.decoded_rows:
                        w.writerow([
                            r["work_num"], r["work_code"], r["work_name"], r["work_qty"],
                            r["rate_id"], r["resource_type"], r["resource_code"], r["resource_name"],
                            r["unit"], r["norm_qty"], r["resource_qty"], r["price"], r["base_cost"]
                        ])

                with open(recon_csv, "w", encoding="utf-8-sig", newline="") as f:
                    w = csv.writer(f, delimiter=";")
                    w.writerow(recon_headers)
                    for r in self.reconciliation_rows:
                        w.writerow([
                            r["pos_num"], r["work_code"], r["work_name"], r["rate_id"],
                            r["zp_local"], r["zp_res"], r["zp_delta"],
                            r["em_local"], r["em_res"], r["em_delta"],
                            r["mr_local"], r["mr_res"], r["mr_delta"],
                            r["zpm_local"], r["status"]
                        ])

                messagebox.showinfo("Экспорт", f"Файлы сохранены:\n{res_csv}\n{recon_csv}")
                return

            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Ресурсы"
            ws1.append(res_headers)

            for r in self.decoded_rows:
                ws1.append([
                    r["work_num"], r["work_code"], r["work_name"], r["work_qty"],
                    r["rate_id"], r["resource_type"], r["resource_code"], r["resource_name"],
                    r["unit"], r["norm_qty"], r["resource_qty"], r["price"], r["base_cost"]
                ])

            ws2 = wb.create_sheet("Сверка")
            ws2.append(recon_headers)

            for r in self.reconciliation_rows:
                ws2.append([
                    r["pos_num"], r["work_code"], r["work_name"], r["rate_id"],
                    r["zp_local"], r["zp_res"], r["zp_delta"],
                    r["em_local"], r["em_res"], r["em_delta"],
                    r["mr_local"], r["mr_res"], r["mr_delta"],
                    r["zpm_local"], r["status"]
                ])

            ws3 = wb.create_sheet("Отчет")
            for line in (self.last_report or "").splitlines():
                ws3.append([line])

            for ws in [ws1, ws2, ws3]:
                widths = {
                    "A": 10, "B": 18, "C": 60, "D": 14, "E": 10, "F": 12,
                    "G": 16, "H": 60, "I": 10, "J": 12, "K": 16, "L": 12,
                    "M": 14, "N": 14, "O": 12
                }
                for col, width in widths.items():
                    ws.column_dimensions[col].width = width

            wb.save(out)
            messagebox.showinfo("Экспорт", f"Файл сохранен:\n{out}")

        except Exception as e:
            messagebox.showerror("Экспорт", f"Ошибка сохранения:\n{e}")


def create_page(parent) -> tk.Frame:
    page = EstimateResourceDecoderPage(parent)
    page.pack(fill="both", expand=True)
    return page


def open_estimate_resource_decoder(parent=None):
    if parent is None:
        root = tk.Tk()
        root.title("Раскрытие ресурсов расценок")
        root.geometry("1500x850")
        EstimateResourceDecoderPage(root).pack(fill="both", expand=True)
        root.mainloop()
        return root

    win = tk.Toplevel(parent)
    win.title("Раскрытие ресурсов расценок")
    win.geometry("1500x850")
    EstimateResourceDecoderPage(win).pack(fill="both", expand=True)
    return win


if __name__ == "__main__":
    open_estimate_resource_decoder()
