import re
import csv
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple

import tkinter as tk
from tkinter import ttk, messagebox, simpledialog

from openpyxl import Workbook, load_workbook

MPL_AVAILABLE = False


class ColumnMappingDialog(simpledialog.Dialog):
    def __init__(self, parent, headers: List[str], cur_map: Dict[str, Optional[int]]):
        self.headers = headers
        self.cur_map = cur_map or {}
        self.result = None
        super().__init__(parent, title="Настройка соответствия колонок")

    def body(self, master):
        tk.Label(
            master,
            text="Укажите, какие колонки соответствуют показателям (для общего режима):",
            font=("Segoe UI", 10, "bold")
        ).grid(row=0, column=0, columnspan=2, sticky="w", pady=(4, 8))

        tk.Label(master, text="Итого (строительные затраты):").grid(row=1, column=0, sticky="e", padx=(0, 6))
        tk.Label(master, text="Материалы:").grid(row=2, column=0, sticky="e", padx=(0, 6))
        tk.Label(master, text="Заработная плата:").grid(row=3, column=0, sticky="e", padx=(0, 6))

        vals = self.headers or ["— нет колонок —"]
        self.cmb_total = ttk.Combobox(master, values=vals, state="readonly", width=46)
        self.cmb_materials = ttk.Combobox(master, values=vals, state="readonly", width=46)
        self.cmb_wages = ttk.Combobox(master, values=vals, state="readonly", width=46)

        def set_by_index(cmb, idx):
            if idx is not None and 0 <= idx < len(vals):
                cmb.current(idx)
            else:
                cmb.set("")

        set_by_index(self.cmb_total, self.cur_map.get("total"))
        set_by_index(self.cmb_materials, self.cur_map.get("materials"))
        set_by_index(self.cmb_wages, self.cur_map.get("wages"))

        self.cmb_total.grid(row=1, column=1, sticky="w")
        self.cmb_materials.grid(row=2, column=1, sticky="w")
        self.cmb_wages.grid(row=3, column=1, sticky="w")
        return self.cmb_total

    def apply(self):
        def idx_of(cmb):
            v = cmb.get().strip()
            try:
                return self.headers.index(v)
            except Exception:
                return None

        self.result = {
            "total": idx_of(self.cmb_total),
            "materials": idx_of(self.cmb_materials),
            "wages": idx_of(self.cmb_wages),
        }


class BudgetAnalysisPage(tk.Frame):
    VAT_RATE = 0.22
    VAT_LABEL = "22%"

    COST_KEYS = ["zp", "em", "mr", "nr", "sp", "nr_sp_zpm"]
    REFERENCE_KEYS = ["zpm_incl"]

    DISPLAY_CATEGORIES = [
        ("zp", "Заработная плата (ЗП)"),
        ("em", "Эксплуатация машин (ЭМ)"),
        ("mr", "Материалы (МР)"),
        ("nr", "Накладные расходы (НР)"),
        ("sp", "Сметная прибыль (СП)"),
        ("nr_sp_zpm", "НР и СП от ЗПМ"),
    ]

    DISPLAY_CATEGORIES_MAP = {
        "zp": "Заработная плата (ЗП)",
        "em_gross": "Эксплуатация машин (ЭМ)",
        "em": "Эксплуатация машин (ЭМ)",
        "mr": "Материалы (МР)",
        "nr": "Накладные расходы (НР)",
        "sp": "Сметная прибыль (СП)",
        "nr_sp_zpm": "НР и СП от ЗПМ",
        "zpm_incl": "в т.ч. ЗПМ",
    }

    def __init__(self, master):
        super().__init__(master, bg="#f7f7f7")
        self.file_path: Optional[Path] = None

        self.headers: List[str] = []
        self.rows: List[List[Any]] = []
        self.mapping: Dict[str, Optional[int]] = {"total": None, "materials": None, "wages": None}

        self.mode: str = "generic"
        self.smeta_sheet_name: Optional[str] = None
        self.smeta_name_col: Optional[int] = None
        self.smeta_total_col: Optional[int] = None
        self.smeta_data_rows: List[List[Any]] = []

        self.stats_base: Dict[str, float] = {
            k: 0.0 for k in self.COST_KEYS + self.REFERENCE_KEYS + ["total", "materials", "wages"]
        }
        self.stats: Dict[str, float] = self.stats_base.copy()
        self.breakdown_rows: List[Dict[str, Any]] = []

        self.last_debug_report: str = "Отчет анализа пока не сформирован."
        self.last_not_classified_rows: List[Dict[str, Any]] = []

        self.vat_enabled = tk.BooleanVar(value=False)

        header = tk.Frame(self, bg="#f7f7f7")
        header.pack(fill="x", padx=12, pady=(10, 6))
        tk.Label(header, text="Анализ смет", font=("Segoe UI", 16, "bold"), bg="#f7f7f7").pack(side="left")

        ctrl = tk.Frame(self, bg="#f7f7f7")
        ctrl.pack(fill="x", padx=12, pady=(0, 8))
        self.btn_open = ttk.Button(ctrl, text="Открыть смету (XLSX/CSV)", command=self._open_file)
        self.btn_open.pack(side="left")
        self.btn_map = ttk.Button(ctrl, text="Настроить соответствие колонок", command=self._open_mapping, state="disabled")
        self.btn_map.pack(side="left", padx=(8, 0))
        self.btn_export = ttk.Button(ctrl, text="Сохранить свод", command=self._export_summary, state="disabled")
        self.btn_export.pack(side="left", padx=(8, 0))
        self.btn_debug = ttk.Button(ctrl, text="Показать отчет анализа", command=self._show_debug_report, state="disabled")
        self.btn_debug.pack(side="left", padx=(8, 0))

        self.chk_vat = ttk.Checkbutton(
            ctrl,
            text=f"Начислить НДС {self.VAT_LABEL}",
            variable=self.vat_enabled,
            command=self._on_vat_toggle
        )
        self.chk_vat.pack(side="left", padx=(16, 0))

        self.lbl_file = tk.Label(self, text="Файл не выбран", fg="#555", bg="#f7f7f7")
        self.lbl_file.pack(anchor="w", padx=12, pady=(0, 2))

        self.lbl_sheet = tk.Label(self, text="", fg="#777", bg="#f7f7f7")
        self.lbl_sheet.pack(anchor="w", padx=12, pady=(0, 6))

        card = tk.Frame(self, bg="#ffffff", bd=1, relief="solid")
        card.pack(fill="x", padx=12, pady=(0, 10))

        grid = tk.Frame(card, bg="#ffffff")
        grid.pack(fill="x", padx=12, pady=12)

        tk.Label(grid, text="Показатель", font=("Segoe UI", 10, "bold"), bg="#ffffff").grid(row=0, column=0, sticky="w")
        tk.Label(grid, text="Сумма (руб.)", font=("Segoe UI", 10, "bold"), bg="#ffffff").grid(row=0, column=1, sticky="e")
        tk.Label(grid, text="Доля", font=("Segoe UI", 10, "bold"), bg="#ffffff").grid(row=0, column=2, sticky="e")

        row_idx = 1
        self._metric_rows = {}

        self._row_total = self._add_metric_row(grid, row_idx, "Строительные затраты (Итого)")
        row_idx += 1

        for key, title in self.DISPLAY_CATEGORIES:
            self._metric_rows[key] = self._add_metric_row(grid, row_idx, title)
            row_idx += 1

        self._row_zpm_ref = self._add_metric_row(grid, row_idx, "в т.ч. ЗПМ (Справочно)")
        self._row_zpm_ref["label"].config(fg="#888888")
        row_idx += 1

        self._row_vat = self._add_metric_row(grid, row_idx, f"НДС {self.VAT_LABEL}")
        self._row_vat["label"].config(text=f"НДС {self.VAT_LABEL}", bg="#ffffff", fg="#d32f2f")
        self._row_vat["label"].grid(row=row_idx, column=0, sticky="w", pady=3)
        self._row_vat["label"].grid_remove()
        self._row_vat["val"].grid_remove()
        self._row_vat["pct"].grid_remove()
        row_idx += 1

        self._row_total_vat = self._add_metric_row(grid, row_idx, "Всего с НДС")
        self._row_total_vat["label"].config(
            text="Всего с НДС", bg="#ffffff", font=("Segoe UI", 10, "bold"), fg="#1976d2"
        )
        self._row_total_vat["label"].grid(row=row_idx, column=0, sticky="w", pady=3)
        self._row_total_vat["label"].grid_remove()
        self._row_total_vat["val"].grid_remove()
        self._row_total_vat["pct"].grid_remove()

        for c in range(3):
            grid.grid_columnconfigure(c, weight=1)

        hint_text = (
            "Smeta.RU: лист «ЛОКАЛЬНАЯ СМЕТА». Расчет ведется по детализированным статьям (ЗП, ЭМ, МР, НР, СП).\n"
            "Эксплуатация машин (ЭМ) автоматически корректируется на сумму 'в т.ч. ЗПМ' для избежания двойного учета.\n"
            f"Поддержка отрицательных значений и дробных номеров позиций. Чекбокс «Начислить НДС {self.VAT_LABEL}» увеличивает все суммы на {self.VAT_LABEL}."
        )
        hint = tk.Label(self, text=hint_text, fg="#666", bg="#f7f7f7", justify="left", wraplength=980)
        hint.pack(fill="x", padx=12, pady=(0, 10))

        main_split = tk.Frame(self, bg="#f7f7f7")
        main_split.pack(fill="both", expand=True, padx=12, pady=(0, 12))

        left = tk.Frame(main_split, bg="#f7f7f7")
        left.pack(side="left", fill="both", expand=True)

        tk.Label(left, text="Расшифровка строк", font=("Segoe UI", 11, "bold"), bg="#f7f7f7").pack(anchor="w", pady=(0, 6))

        flt = tk.Frame(left, bg="#f7f7f7")
        flt.pack(anchor="w", pady=(0, 6))
        self.var_show_mat = tk.BooleanVar(value=True)
        self.var_show_wag = tk.BooleanVar(value=True)
        self.var_show_oth = tk.BooleanVar(value=True)

        ttk.Checkbutton(flt, text="Материалы (МР)", variable=self.var_show_mat, command=self._fill_breakdown_table).pack(side="left", padx=(0, 8))
        ttk.Checkbutton(flt, text="Трудозатраты/Машины (ЗП, ЭМ)", variable=self.var_show_wag, command=self._fill_breakdown_table).pack(side="left", padx=(0, 8))
        ttk.Checkbutton(flt, text="Накладные/Прибыль (НР, СП)", variable=self.var_show_oth, command=self._fill_breakdown_table).pack(side="left")

        tree_wrap = tk.Frame(left)
        tree_wrap.pack(fill="both", expand=True)

        cols = ("pos_num", "rate_code", "category", "name", "amount")
        self.tree = ttk.Treeview(tree_wrap, columns=cols, show="headings", height=12)

        self.tree.heading("pos_num", text="Поз.")
        self.tree.heading("rate_code", text="Шифр расценки")
        self.tree.heading("category", text="Категория")
        self.tree.heading("name", text="Наименование")
        self.tree.heading("amount", text="Сумма, руб.")

        self.tree.column("pos_num", width=60, anchor="w")
        self.tree.column("rate_code", width=140, anchor="w")
        self.tree.column("category", width=180, anchor="w")
        self.tree.column("name", stretch=True, minwidth=250, anchor="w")
        self.tree.column("amount", width=120, anchor="e")

        yscroll = ttk.Scrollbar(tree_wrap, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=yscroll.set)
        self.tree.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")

    def _add_metric_row(self, grid, r, title: str):
        lbl = tk.Label(grid, text=title, bg="#ffffff")
        lbl.grid(row=r, column=0, sticky="w", pady=3)
        val_lbl = tk.Label(grid, text="-", bg="#ffffff", anchor="e")
        val_lbl.grid(row=r, column=1, sticky="e", pady=3)
        pct_lbl = tk.Label(grid, text="-", bg="#ffffff", anchor="e")
        pct_lbl.grid(row=r, column=2, sticky="e", pady=3)
        return {"label": lbl, "val": val_lbl, "pct": pct_lbl}

    def _on_vat_toggle(self):
        if not self.stats_base.get("total"):
            return
        self._apply_vat()
        self._render_stats()
        self._fill_breakdown_table()

    def _apply_vat(self):
        multiplier = (1.0 + self.VAT_RATE) if self.vat_enabled.get() else 1.0
        self.stats = {}

        for key in self.COST_KEYS + self.REFERENCE_KEYS + ["total", "materials", "wages"]:
            self.stats[key] = self.stats_base.get(key, 0.0) * multiplier

        for row in self.breakdown_rows:
            if "amount_base" in row:
                row["amount"] = row["amount_base"] * multiplier

    def _open_file(self):
        try:
            from tkinter import filedialog as fd
        except Exception:
            messagebox.showerror("Файл", "Не удалось открыть диалог выбора файлов.")
            return

        fname = fd.askopenfilename(
            title="Выберите файл сметы (XLSX/CSV)",
            filetypes=[("Excel", "*.xlsx;*.xlsm"), ("CSV", "*.csv"), ("Все файлы", "*.*")]
        )
        if not fname:
            return

        self.file_path = Path(fname)
        self.lbl_file.config(text=f"Файл: {self.file_path}")
        ok = self._load_file(self.file_path)
        self.btn_map.config(state=("normal" if (ok and self.mode == "generic") else "disabled"))
        self.btn_export.config(state=("normal" if ok else "disabled"))
        self.btn_debug.config(state=("normal" if ok and self.mode == "smeta" else "disabled"))

        if not ok:
            messagebox.showwarning("Анализ смет", "Не удалось распознать структуру файла.")

    def _load_file(self, path: Path) -> bool:
        self.mode = "generic"
        self.headers, self.rows = [], []
        self.breakdown_rows = []
        self.smeta_sheet_name = None
        self.smeta_name_col = None
        self.smeta_total_col = None
        self.smeta_data_rows = []
        self.lbl_sheet.config(text="")
        self.last_debug_report = "Отчет анализа пока не сформирован."
        self.last_not_classified_rows = []
        self.btn_debug.config(state="disabled")

        ext = path.suffix.lower()

        self.stats_base = {
            k: 0.0 for k in self.COST_KEYS + self.REFERENCE_KEYS + ["total", "materials", "wages"]
        }

        try:
            if ext in (".xlsx", ".xlsm"):
                if self._parse_xlsx_smeta_ru(path):
                    self.mode = "smeta"
                    self._analyze_smeta()
                    return True
                self._parse_xlsx_generic(path)
                self.mapping = self._detect_mapping(self.headers, self.rows)
                self._analyze_generic()
                return True

            if ext == ".csv":
                self._parse_csv_generic(path)
                self.mapping = self._detect_mapping(self.headers, self.rows)
                self._analyze_generic()
                return True

            try:
                if self._parse_xlsx_smeta_ru(path):
                    self.mode = "smeta"
                    self._analyze_smeta()
                    return True
                self._parse_xlsx_generic(path)
                self.mapping = self._detect_mapping(self.headers, self.rows)
                self._analyze_generic()
                return True
            except Exception:
                self._parse_csv_generic(path)
                self.mapping = self._detect_mapping(self.headers, self.rows)
                self._analyze_generic()
                return True

        except Exception as e:
            messagebox.showerror("Загрузка сметы", f"Ошибка чтения файла:\n{e}")
            return False

    def _parse_xlsx_smeta_ru(self, path: Path) -> bool:
        wb = load_workbook(path, read_only=True, data_only=True)
        target_ws = None

        for ws in wb.worksheets:
            if self._sheet_has_local_smeta_marker(ws):
                target_ws = ws
                break

        if target_ws is None:
            return False

        hdr_row_idx, name_col, total_current_col = self._find_table_header(target_ws)
        if hdr_row_idx is None or name_col is None or total_current_col is None:
            return False

        data_rows: List[List[Any]] = []
        last_local_total: Optional[float] = None
        last_grand_total: Optional[float] = None

        for row in target_ws.iter_rows(min_row=hdr_row_idx + 1, values_only=True):
            cells = list(row)
            if not any(c is not None and str(c).strip() for c in cells):
                continue

            name_cell = self._str(cells[name_col]) if name_col < len(cells) else ""

            if self._is_numbering_row(cells):
                continue

            if self._is_summary_row(cells, name_col):
                val = self._to_number(cells[total_current_col]) if total_current_col < len(cells) else None
                low = name_cell.lower()
                if isinstance(val, float):
                    if "итого по смете" in low or ("итого" in low and "смете" in low):
                        last_grand_total = val
                    if "итого по локальной смете" in low:
                        last_local_total = val
                continue

            data_rows.append(cells)

        self.smeta_sheet_name = target_ws.title
        self.smeta_name_col = name_col
        self.smeta_total_col = total_current_col
        self.smeta_data_rows = data_rows

        total = last_grand_total if isinstance(last_grand_total, float) else last_local_total
        self.stats_base["total"] = float(total or 0.0)

        self.lbl_sheet.config(
            text=f"Лист: {self.smeta_sheet_name} (режим Smeta.RU, колонка текущей цены: {self.smeta_total_col + 1})"
        )
        return True

    @staticmethod
    def _sheet_has_local_smeta_marker(ws) -> bool:
        try:
            for row in ws.iter_rows(min_row=1, max_row=30, values_only=True):
                for c in row:
                    if isinstance(c, str) and "локальная смета" in c.lower():
                        return True
        except Exception:
            pass
        return False

    @staticmethod
    def _normalize_header_text(s: Any) -> str:
        txt = str(s or "").strip()
        txt = txt.replace("\n", " ").replace("\r", " ")
        return re.sub(r"\s+", " ", txt).lower()

    def _find_table_header(self, ws) -> Tuple[Optional[int], Optional[int], Optional[int]]:
        name_col: Optional[int] = None
        hdr_row_idx: Optional[int] = None
        total_current_col: Optional[int] = None

        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            vals_norm = [self._normalize_header_text(v) for v in row]
            if not any(vals_norm):
                continue

            local_name_col = None
            local_total_col = None

            for idx, v in enumerate(vals_norm):
                if ("наименование работ" in v and "затрат" in v) or ("наименование работ и затрат" in v):
                    local_name_col = idx

                if "всего" in v and "затрат" in v and "текущ" in v:
                    local_total_col = idx

            if local_name_col is not None and local_total_col is not None:
                hdr_row_idx = i
                name_col = local_name_col
                total_current_col = local_total_col
                break

        if hdr_row_idx is None:
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                vals_norm = [self._normalize_header_text(v) for v in row]
                if not any(vals_norm):
                    continue

                has_name = any(("наименование работ" in v and "затрат" in v) or ("наименование работ и затрат" in v) for v in vals_norm)
                idx_current = [idx for idx, v in enumerate(vals_norm) if "всего" in v and "текущ" in v]

                if has_name and idx_current:
                    hdr_row_idx = i
                    total_current_col = idx_current[0]
                    try:
                        name_col = vals_norm.index(
                            next(v for v in vals_norm if ("наименование работ" in v and "затрат" in v) or ("наименование работ и затрат" in v))
                        )
                    except StopIteration:
                        name_col = 2
                    break

        return hdr_row_idx, name_col, total_current_col

    def _is_numbering_row(self, cells: List[Any]) -> bool:
        vals = [str(v).strip() for v in cells if v is not None and str(v).strip() != ""]
        if not vals:
            return False
        return self._is_sequential_digits_list(vals)

    @staticmethod
    def _is_sequential_digits_list(vals: List[str]) -> bool:
        try:
            nums = [int(v) for v in vals if v.isdigit()]
        except Exception:
            return False
        if not nums:
            return False
        return nums == list(range(1, len(nums) + 1)) and len(nums) >= 5

    def _is_summary_row(self, cells: List[Any], name_col: int) -> bool:
        col0_empty = (len(cells) < 1) or (self._str(cells[0]) == "")
        col1_empty = (len(cells) < 2) or (self._str(cells[1]) == "")
        name = self._str(cells[name_col]) if name_col < len(cells) else ""
        if not (col0_empty and col1_empty and name):
            return False
        return self._is_summary_name(name)

    @staticmethod
    def _is_summary_name(name: Any) -> bool:
        s = re.sub(r"\s+", " ", str(name or "")).strip().lower()
        if not s:
            return False
        if "по позиции" in s:
            return False

        patterns = [
            "итого по локальной смете",
            "итоги по смете",
            "итоги по разделу",
            "итоги по",
            "итог по",
            "итого прямые затраты",
            "итого прямые",
            "итого по смете",
            "всего по смете",
            "всего по разделу",
            "всего по",
            "справочно",
            "ндс",
            "итого с ндс",
            "всего с ндс"
        ]
        return any(p in s for p in patterns)

    @staticmethod
    def _str(x: Any) -> str:
        return str(x or "").strip()

    @staticmethod
    def _to_number(x: Any) -> Optional[float]:
        if x is None:
            return None

        if isinstance(x, (int, float)):
            return float(x)

        s = str(x).strip()
        if not s:
            return None

        s = s.replace("\u00A0", " ")
        s = re.sub(r"[^0-9,.\-]", "", s)

        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")
        elif "," in s:
            s = s.replace(",", ".")

        try:
            return float(s)
        except Exception:
            return None

    def _get_smeta_current_value(self, row: List[Any]) -> Optional[float]:
        if self.smeta_total_col is None or self.smeta_total_col >= len(row):
            return None
        return self._to_number(row[self.smeta_total_col])

    def _is_labor_or_percent_unit(self, unit: Any) -> bool:
        if unit is None:
            return False
        u = str(unit).strip().lower()
        if not u:
            return False
        if "%" in u:
            return True
        if "чел" in u:
            return True
        return False

    def _is_material_unit(self, unit: Any) -> bool:
        u = self._str(unit).lower()
        if not u:
            return False
        material_units = [
            "м3", "м2", "м", "т", "кг", "шт", "компл", "л",
            "м³", "м²", "комплект", "п.м", "пм", "м.п", "уп", "пач", "пар", "рул"
        ]
        return any(x in u for x in material_units)

    def _has_numeric_position(self, cell: Any) -> bool:
        if cell is None:
            return False

        s = str(cell).strip()
        s = s.replace("\u00A0", "").replace("\xa0", "").replace(" ", "").replace("\t", "")

        if not s:
            return False

        pattern = r"^\d+([.,]\d*)?$"
        return bool(re.match(pattern, s))

    def _diagnose_unclassified_row(self, row: List[Any]) -> str:
        if self.smeta_name_col is None or self.smeta_total_col is None:
            return "Не заданы индексы колонок сметы"

        if self._is_summary_row(row, self.smeta_name_col):
            return "Итоговая/сводная строка"

        val = self._get_smeta_current_value(row)
        if not isinstance(val, float):
            return "Нет числового значения в колонке текущей цены"

        name = self._str(row[self.smeta_name_col]) if self.smeta_name_col < len(row) else ""
        if not name:
            return "Пустое наименование"

        pos_cell = row[0] if len(row) > 0 else None
        pos_str = self._str(pos_cell)
        unit = row[3] if len(row) > 3 else ""
        unit_str = self._str(unit).lower()
        n = re.sub(r"[^а-яa-z0-9]", "", name.lower())

        parts = [
            f"Позиция='{pos_str}'",
            f"Ед.изм='{unit_str}'",
            f"numeric_pos={'да' if self._has_numeric_position(pos_cell) else 'нет'}",
            f"subpos={'да' if bool(pos_str and re.match(r'^\\d+[.,]\\d+$', pos_str)) else 'нет'}",
            f"material_unit={'да' if self._is_material_unit(unit) else 'нет'}",
            f"labor_or_%={'да' if self._is_labor_or_percent_unit(unit) else 'нет'}",
            f"norm='{n[:60]}'"
        ]
        return "Не подошла ни под одно правило классификации; " + ", ".join(parts)

    def _classify_smeta_row(self, row: List[Any]) -> Tuple[Optional[str], Optional[float]]:
        if self.smeta_name_col is None or self.smeta_total_col is None:
            return None, None

        if self._is_summary_row(row, self.smeta_name_col):
            return None, None

        name = self._str(row[self.smeta_name_col]) if self.smeta_name_col < len(row) else ""
        n = re.sub(r"[^а-яa-z0-9]", "", name.lower())

        val = self._get_smeta_current_value(row)
        if not isinstance(val, float):
            return None, None

        for col_idx in [1, 2, 3]:
            if len(row) > col_idx:
                col_val = self._str(row[col_idx]).upper().strip()
                if col_val in ["МР", "МРР"] or col_val.startswith("МР ") or col_val.startswith("МРР "):
                    return "mr", val

        pos_cell = row[0] if len(row) > 0 else None
        pos_str = self._str(pos_cell)
        is_subposition = bool(pos_str and re.match(r"^\d+[.,]\d+$", pos_str))

        unit = row[3] if len(row) > 3 else ""
        is_material_unit = self._is_material_unit(unit)

        if is_subposition and is_material_unit:
            return "mr", val

        if "втчзпм" in n or "втомчислезпм" in n:
            return "zpm_incl", val

        if n == "зп" or n == "зпм" or "оплататруда" in n or "заработн" in n:
            return "zp", val

        if n.startswith("эм") and "зпм" not in n:
            return "em_gross", val
        if n.startswith("эмм") and "зпм" not in n:
            return "em_gross", val
        if "эксплуатациямашин" in n and "зпм" not in n:
            return "em_gross", val

        if "нриспотзпм" in n:
            return "nr_sp_zpm", val

        if n == "нр" or "нротзп" in n or "нротфот" in n or "накладные" in n:
            return "nr", val

        if n == "сп" or "спотзп" in n or "спотфот" in n or "сметнаяприбыль" in n:
            return "sp", val

        not_labor_unit = not self._is_labor_or_percent_unit(unit)

        is_cost_line = True
        if n.startswith("зп") or n.startswith("эм") or n.startswith("нр") or n.startswith("сп"):
            is_cost_line = False
        if n in ["зп", "эм", "нр", "сп", "зпм", "эмм"]:
            is_cost_line = False

        if self._has_numeric_position(pos_cell) and is_material_unit and is_cost_line:
            return "mr", val

        if self._has_numeric_position(pos_cell) and not_labor_unit and is_cost_line and name:
            exclude_words = [
                "машинист", "слесар", "монтаж", "установк", "демонтаж", "разборк",
                "разработка", "устройство", "укладка", "сварка", "бурение", "окраска"
            ]
            name_lower = name.lower()
            is_excluded = any(name_lower.startswith(word) or f" {word}" in name_lower for word in exclude_words)
            if not is_excluded:
                return "mr", val

        return None, None

    def _show_debug_report(self):
        debug_win = tk.Toplevel(self)
        debug_win.title("Отчёт анализа сметы")
        debug_win.geometry("950x650")

        text_frame = tk.Frame(debug_win)
        text_frame.pack(fill="both", expand=True, padx=10, pady=10)

        text_widget = tk.Text(text_frame, wrap="word", font=("Courier New", 9))
        text_widget.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(text_frame, command=text_widget.yview)
        scrollbar.pack(side="right", fill="y")
        text_widget.config(yscrollcommand=scrollbar.set)

        text_widget.insert("1.0", self.last_debug_report or "Отчет отсутствует")
        text_widget.config(state="disabled")

        btn_frame = tk.Frame(debug_win)
        btn_frame.pack(fill="x", padx=10, pady=(0, 10))

        def save_report():
            try:
                with open("smeta_analysis_report.txt", "w", encoding="utf-8") as f:
                    f.write(self.last_debug_report or "")
                messagebox.showinfo("Сохранено", "Отчёт сохранён в файл:\nsmeta_analysis_report.txt")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось сохранить:\n{e}")

        btn_save = ttk.Button(btn_frame, text="Сохранить отчёт в файл", command=save_report)
        btn_save.pack(side="left")

        btn_close = ttk.Button(btn_frame, text="Закрыть", command=debug_win.destroy)
        btn_close.pack(side="right")

    def _analyze_smeta(self):
        if self.smeta_name_col is None or self.smeta_total_col is None:
            raise RuntimeError("Не заданы индексы колонок для сметы.")

        gross_stats: Dict[str, float] = {
            k: 0.0 for k in self.COST_KEYS + self.REFERENCE_KEYS + ["em_gross"]
        }
        self.breakdown_rows = []

        name_col_idx = self.smeta_name_col

        not_classified_rows = []
        classified_count = 0
        total_rows = 0

        for row in self.smeta_data_rows:
            total_rows += 1
            pos_num = self._str(row[0]) if len(row) > 0 else ""
            rate_code = self._str(row[1]) if len(row) > 1 else ""
            name = self._str(row[name_col_idx]) if name_col_idx < len(row) else ""

            cat, val = self._classify_smeta_row(row)

            if cat and isinstance(val, float):
                classified_count += 1
            else:
                test_val = self._get_smeta_current_value(row)
                if isinstance(test_val, float) and abs(test_val) > 0.01:
                    not_classified_rows.append({
                        "pos": pos_num,
                        "code": rate_code,
                        "name": name[:100],
                        "val": test_val,
                        "unit": self._str(row[3]) if len(row) > 3 else "",
                        "is_summary": self._is_summary_row(row, name_col_idx),
                        "reason": self._diagnose_unclassified_row(row)
                    })

            if not cat or not isinstance(val, float):
                continue

            gross_stats[cat] = gross_stats.get(cat, 0.0) + val

            display_cat = self.DISPLAY_CATEGORIES_MAP.get(cat, cat)
            self.breakdown_rows.append({
                "pos_num": pos_num,
                "rate_code": rate_code,
                "category": display_cat,
                "name": name,
                "amount": val,
                "amount_base": val
            })

        em_gross_total = gross_stats.pop("em_gross", 0.0)
        zpm_incl_total = gross_stats.get("zpm_incl", 0.0)
        em_net_total = em_gross_total - zpm_incl_total
        zp_total = gross_stats.get("zp", 0.0) + zpm_incl_total

        final_stats = {
            "zp": zp_total,
            "em": em_net_total,
            "mr": gross_stats.get("mr", 0.0),
            "nr": gross_stats.get("nr", 0.0),
            "sp": gross_stats.get("sp", 0.0),
            "nr_sp_zpm": gross_stats.get("nr_sp_zpm", 0.0),
        }

        total_cost = sum(final_stats.values())

        self.stats_base.update(final_stats)
        self.stats_base["total"] = total_cost
        self.stats_base["zpm_incl"] = zpm_incl_total
        self.stats_base["materials"] = self.stats_base["mr"]
        self.stats_base["wages"] = self.stats_base["zp"]

        debug_lines = []
        debug_lines.append("СТАТИСТИКА АНАЛИЗА:")
        debug_lines.append("=" * 80)
        debug_lines.append(f"Лист: {self.smeta_sheet_name}")
        debug_lines.append(f"Колонка текущей цены: {self.smeta_total_col + 1 if self.smeta_total_col is not None else '-'}")
        debug_lines.append(f"Всего строк данных: {total_rows}")
        debug_lines.append(f"Классифицировано: {classified_count}")
        debug_lines.append(f"Не классифицировано с ценами: {len(not_classified_rows)}")
        debug_lines.append("")
        debug_lines.append("ИТОГИ ПО КАТЕГОРИЯМ:")
        debug_lines.append(f"ЗП: {zp_total:,.2f}")
        debug_lines.append(f"ЭМ (чистая): {em_net_total:,.2f}")
        debug_lines.append(f"МР: {final_stats['mr']:,.2f}")
        debug_lines.append(f"НР: {final_stats['nr']:,.2f}")
        debug_lines.append(f"СП: {final_stats['sp']:,.2f}")
        debug_lines.append(f"НР и СП от ЗПМ: {final_stats['nr_sp_zpm']:,.2f}")
        debug_lines.append(f"в т.ч. ЗПМ (справочно): {zpm_incl_total:,.2f}")
        debug_lines.append("")
        debug_lines.append(f"Вычисленный итог: {total_cost:,.2f} руб.")
        debug_lines.append("=" * 80)
        debug_lines.append("")

        if not_classified_rows:
            total_missing = sum(r["val"] for r in not_classified_rows)
            debug_lines.append("НЕРАСПОЗНАННЫЕ СТРОКИ С ЦЕНАМИ:")
            debug_lines.append(f"Сумма нераспознанных: {total_missing:,.2f} руб.")
            debug_lines.append("")
            debug_lines.append("=" * 80)
            debug_lines.append("")

            for i, r in enumerate(not_classified_rows[:50], 1):
                debug_lines.append(f"{i}. Поз: '{r['pos']}' | Шифр: '{r['code']}'")
                debug_lines.append(f"   Наименование: {r['name']}")
                debug_lines.append(f"   Ед.изм: {r['unit']} | Сумма: {r['val']:,.2f} руб.")
                debug_lines.append(f"   Итоговая строка: {'ДА' if r['is_summary'] else 'НЕТ'}")
                debug_lines.append(f"   Причина: {r['reason']}")
                debug_lines.append("")

            if len(not_classified_rows) > 50:
                debug_lines.append(f"... и ещё {len(not_classified_rows) - 50} строк")
        else:
            debug_lines.append("Все строки с ценами успешно классифицированы!")

        self.last_not_classified_rows = not_classified_rows
        self.last_debug_report = "\n".join(debug_lines)

        self._apply_vat()
        self._render_stats()
        self._fill_breakdown_table()

    def _parse_xlsx_generic(self, path: Path):
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active

        hdr_row_idx = None
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            cells = [self._str(c) for c in row]
            if sum(1 for c in cells if c) >= 2:
                hdr_row_idx = i
                self.headers = [self._norm_header(c) for c in cells]
                break

        if hdr_row_idx is None:
            raise RuntimeError("Не найдена строка заголовков")

        self.rows = [list(row) for row in ws.iter_rows(min_row=hdr_row_idx + 1, values_only=True)]
        self.lbl_sheet.config(text=f"Лист: {ws.title} (общий режим)")

    def _parse_csv_generic(self, path: Path):
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            sample = f.read(4096)
            f.seek(0)
            try:
                sniffer = csv.Sniffer()
                dialect = sniffer.sniff(sample, delimiters=";,")
            except Exception:
                class D:
                    delimiter = ";"
                dialect = D()

            reader = csv.reader(f, dialect=dialect)
            rows = list(reader)

        if not rows:
            raise RuntimeError("CSV пустой")

        hdr_idx = next((i for i, row in enumerate(rows) if any((c or "").strip() for c in row)), None)
        if hdr_idx is None:
            raise RuntimeError("Не найдена строка заголовков")

        self.headers = [self._norm_header(c) for c in rows[hdr_idx]]
        self.rows = rows[hdr_idx + 1:]
        self.lbl_sheet.config(text="CSV (общий режим)")

    @staticmethod
    def _norm_header(s: Any) -> str:
        txt = str(s or "").strip()
        txt = txt.replace("\n", " ").replace("\r", " ")
        return re.sub(r"\s+", " ", txt)

    def _detect_mapping(self, headers: List[str], rows: List[List[Any]]) -> Dict[str, Optional[int]]:
        hlow = [h.lower() for h in headers]

        def find_candidates(patterns: List[str]) -> List[int]:
            return [i for i, h in enumerate(hlow) if any(p in h for p in patterns)]

        def best_index(cands: List[int]) -> Optional[int]:
            best_i, best_sum = None, -1.0
            for idx in cands:
                s = 0.0
                for r in rows:
                    if idx < len(r):
                        v = self._to_number(r[idx])
                        if isinstance(v, float):
                            s += v
                if s > best_sum:
                    best_sum, best_i = s, idx
            return best_i

        return {
            "total": best_index(find_candidates(["итого", "всего", "стоим", "смет", "общая стоимость"])),
            "materials": best_index(find_candidates(["матер", "материа", "мр"])),
            "wages": best_index(find_candidates(["зараб", "оплата труда", "з/п", "зп", "труд"])),
        }

    def _sum_column(self, idx: Optional[int]) -> float:
        if idx is None:
            return 0.0

        s = 0.0
        for r in self.rows:
            if idx < len(r):
                v = self._to_number(r[idx])
                if isinstance(v, float):
                    s += v
        return s

    def _analyze_generic(self):
        total = self._sum_column(self.mapping.get("total"))
        materials = self._sum_column(self.mapping.get("materials"))
        wages = self._sum_column(self.mapping.get("wages"))

        if total <= 0:
            total = materials + wages
        other = max(0.0, total - materials - wages)

        self.stats_base = {
            "total": total,
            "materials": materials,
            "wages": wages,
            "other": other
        }
        self.stats_base.update({k: 0.0 for k in self.COST_KEYS + self.REFERENCE_KEYS})

        self.breakdown_rows = []

        self._apply_vat()
        self._render_stats()
        self._fill_breakdown_table()

    @staticmethod
    def _fmt_money(x: Optional[float]) -> str:
        if x is None:
            return "-"
        try:
            s = f"{float(x):,.2f}"
            s = s.replace(",", " ").replace(".", ",")
            return s
        except Exception:
            return str(x)

    @staticmethod
    def _fmt_pct(x: Optional[float]) -> str:
        if x is None:
            return "-"
        try:
            return f"{x:.1f}%"
        except Exception:
            return "-"

    def _safe_pct(self, part: float) -> Optional[float]:
        t = self.stats.get("total", 0.0)
        return (part / t * 100.0) if t and abs(t) > 1e-12 else None

    def _render_stats(self):
        total = float(self.stats.get("total") or 0.0)

        self._row_total["val"].config(text=self._fmt_money(total))
        self._row_total["pct"].config(text="100%" if abs(total) > 1e-12 else "-")

        if self.mode == "smeta":
            for key, _ in self.DISPLAY_CATEGORIES:
                val = float(self.stats.get(key) or 0.0)
                p = self._safe_pct(val)
                self._metric_rows[key]["label"].grid()
                self._metric_rows[key]["val"].grid()
                self._metric_rows[key]["pct"].grid()
                self._metric_rows[key]["val"].config(text=self._fmt_money(val))
                self._metric_rows[key]["pct"].config(text=self._fmt_pct(p))

            zpm_incl = float(self.stats.get("zpm_incl") or 0.0)
            self._row_zpm_ref["label"].grid()
            self._row_zpm_ref["val"].grid()
            self._row_zpm_ref["pct"].grid()
            self._row_zpm_ref["val"].config(text=self._fmt_money(zpm_incl))
            self._row_zpm_ref["pct"].config(text="-")

        else:
            for key, _ in self.DISPLAY_CATEGORIES:
                self._metric_rows[key]["label"].grid_remove()
                self._metric_rows[key]["val"].grid_remove()
                self._metric_rows[key]["pct"].grid_remove()

            self._row_zpm_ref["label"].grid_remove()
            self._row_zpm_ref["val"].grid_remove()
            self._row_zpm_ref["pct"].grid_remove()

            generic_keys = [("mr", "Материалы"), ("zp", "Заработная плата"), ("nr", "Прочие")]
            generic_vals = {
                "mr": self.stats.get("materials", 0.0),
                "zp": self.stats.get("wages", 0.0),
                "nr": self.stats_base.get("other", 0.0) * ((1.0 + self.VAT_RATE) if self.vat_enabled.get() else 1.0)
            }

            rows_list = list(self._metric_rows.values())
            for i, (key, title) in enumerate(generic_keys):
                val = generic_vals[key]
                p = self._safe_pct(val)
                row_widget = rows_list[i]
                row_widget["label"].config(text=title)
                row_widget["label"].grid()
                row_widget["val"].grid()
                row_widget["pct"].grid()
                row_widget["val"].config(text=self._fmt_money(val))
                row_widget["pct"].config(text=self._fmt_pct(p))

            for i in range(len(generic_keys), len(rows_list)):
                rows_list[i]["label"].grid_remove()
                rows_list[i]["val"].grid_remove()
                rows_list[i]["pct"].grid_remove()

        if self.vat_enabled.get():
            total_base = self.stats_base.get("total", 0.0)
            vat_amount = total_base * self.VAT_RATE
            total_with_vat = total_base * (1.0 + self.VAT_RATE)

            self._row_vat["label"].grid()
            self._row_vat["val"].grid()
            self._row_vat["pct"].grid()

            self._row_total_vat["label"].grid()
            self._row_total_vat["val"].grid()
            self._row_total_vat["pct"].grid()

            self._row_vat["val"].config(text=self._fmt_money(vat_amount))
            self._row_vat["pct"].config(text=self.VAT_LABEL)
            self._row_total_vat["val"].config(text=self._fmt_money(total_with_vat))
            self._row_total_vat["pct"].config(text=f"{int((1.0 + self.VAT_RATE) * 100)}%")
        else:
            self._row_vat["label"].grid_remove()
            self._row_vat["val"].grid_remove()
            self._row_vat["pct"].grid_remove()
            self._row_total_vat["label"].grid_remove()
            self._row_total_vat["val"].grid_remove()
            self._row_total_vat["pct"].grid_remove()

    def _fill_breakdown_table(self):
        for i in self.tree.get_children():
            self.tree.delete(i)

        if not self.breakdown_rows or self.mode != "smeta":
            return

        show_mat = self.var_show_mat.get()
        show_wag = self.var_show_wag.get()
        show_oth = self.var_show_oth.get()

        wage_cats = [
            self.DISPLAY_CATEGORIES_MAP["zp"],
            self.DISPLAY_CATEGORIES_MAP["em"],
            self.DISPLAY_CATEGORIES_MAP["zpm_incl"]
        ]
        other_cats = [
            self.DISPLAY_CATEGORIES_MAP["nr"],
            self.DISPLAY_CATEGORIES_MAP["sp"],
            self.DISPLAY_CATEGORIES_MAP["nr_sp_zpm"]
        ]
        material_cats = [self.DISPLAY_CATEGORIES_MAP["mr"]]

        for row in self.breakdown_rows:
            cat = row["category"]

            is_mat = cat in material_cats
            is_wag = cat in wage_cats
            is_oth = cat in other_cats

            if (is_mat and not show_mat) or (is_wag and not show_wag) or (is_oth and not show_oth):
                continue

            amt = float(row["amount"] or 0.0)

            self.tree.insert(
                "",
                "end",
                values=(
                    row.get("pos_num", ""),
                    row.get("rate_code", ""),
                    cat,
                    str(row["name"]),
                    self._fmt_money(amt)
                )
            )

    def _render_chart(self):
        pass

    def _open_mapping(self):
        if not self.headers or self.mode != "generic":
            return
        dlg = ColumnMappingDialog(self, headers=self.headers, cur_map=self.mapping)
        if getattr(dlg, "result", None):
            self.mapping = dlg.result
            self._analyze_generic()

    def _export_summary(self):
        try:
            from tkinter import filedialog as fd
        except Exception:
            messagebox.showerror("Экспорт", "Не удалось открыть диалог сохранения.")
            return

        if not self.stats:
            return

        fname = fd.asksaveasfilename(
            title="Сохранить свод",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv")]
        )
        if not fname:
            return

        out = Path(fname)

        export_metrics = [
            ("Строительные затраты (Итого)", self.stats.get("total", 0.0), "100%"),
        ]

        if self.mode == "smeta":
            for key, title in self.DISPLAY_CATEGORIES:
                val = self.stats.get(key, 0.0)
                pct = self._fmt_pct(self._safe_pct(val))
                export_metrics.append((title, val, pct))

            ref_zpm = self.stats.get("zpm_incl", 0.0)
            export_metrics.append(("в т.ч. ЗПМ (Справочно)", ref_zpm, "-"))

        else:
            mats = self.stats.get("materials", 0.0)
            wages = self.stats.get("wages", 0.0)
            other = self.stats_base.get("other", 0.0) * ((1.0 + self.VAT_RATE) if self.vat_enabled.get() else 1.0)

            export_metrics.append(("Материалы", mats, self._fmt_pct(self._safe_pct(mats))))
            export_metrics.append(("Заработная плата", wages, self._fmt_pct(self._safe_pct(wages))))
            export_metrics.append(("Прочие", other, self._fmt_pct(self._safe_pct(other))))

        if self.vat_enabled.get():
            total_base = self.stats_base.get("total", 0.0)
            vat_amount = total_base * self.VAT_RATE
            total_with_vat = total_base * (1.0 + self.VAT_RATE)
            export_metrics.append((f"НДС {self.VAT_LABEL}", vat_amount, self.VAT_LABEL))
            export_metrics.append(("Всего с НДС", total_with_vat, f"{int((1.0 + self.VAT_RATE) * 100)}%"))

        try:
            if out.suffix.lower() == ".csv":
                with open(out, "w", encoding="utf-8-sig", newline="") as f:
                    w = csv.writer(f, delimiter=";")
                    w.writerow(["Показатель", "Сумма (руб.)", "Доля"])
                    for title, val, pct in export_metrics:
                        w.writerow([title, self._fmt_money(val), pct])

                    if self.mode == "smeta" and self.breakdown_rows:
                        w.writerow([])
                        w.writerow(["Расшифровка", "", "", "", ""])
                        w.writerow(["Поз.", "Шифр расценки", "Категория", "Наименование", "Сумма, руб."])
                        for row in self.breakdown_rows:
                            w.writerow([
                                row.get("pos_num", ""),
                                row.get("rate_code", ""),
                                row["category"],
                                row["name"],
                                self._fmt_money(row["amount"])
                            ])
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Анализ сметы"

                ws.append(["Показатель", "Сумма (руб.)", "Доля"])
                for title, val, pct in export_metrics:
                    ws.append([title, float(val), pct])

                if self.mode == "smeta" and self.breakdown_rows:
                    ws.append([])
                    ws.append(["Расшифровка"])
                    ws.append(["Поз.", "Шифр расценки", "Категория", "Наименование", "Сумма, руб."])

                    for row in self.breakdown_rows:
                        ws.append([
                            row.get("pos_num", ""),
                            row.get("rate_code", ""),
                            row["category"],
                            row["name"],
                            float(row.get("amount", 0.0) or 0.0)
                        ])

                    ws.column_dimensions["A"].width = 10
                    ws.column_dimensions["B"].width = 20
                    ws.column_dimensions["C"].width = 36
                    ws.column_dimensions["D"].width = 60
                    ws.column_dimensions["E"].width = 18

                wb.save(out)

            messagebox.showinfo("Экспорт", f"Свод сохранён:\n{out}")
        except Exception as e:
            messagebox.showerror("Экспорт", f"Не удалось сохранить свод:\n{e}")


def create_page(parent) -> tk.Frame:
    page = BudgetAnalysisPage(parent)
    page.pack(fill="both", expand=True)
    return page


def open_budget_analyzer(parent=None):
    if parent is None:
        root = tk.Tk()
        root.title("Анализ смет")
        root.geometry("1100x740")
        BudgetAnalysisPage(root).pack(fill="both", expand=True)
        root.mainloop()
        return root

    win = tk.Toplevel(parent)
    win.title("Анализ смет")
    win.geometry("1100x740")
    BudgetAnalysisPage(win).pack(fill="both", expand=True)
    return win


if __name__ == "__main__":
    open_budget_analyzer()
