import os
import re
from datetime import datetime, date
from typing import Optional, Dict, Any, List, Tuple

import tkinter as tk
from tkinter import ttk, messagebox

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from psycopg2.extras import RealDictCursor


# ---------------- DB pool wiring ----------------

db_connection_pool = None

def set_db_pool(pool):
    global db_connection_pool
    db_connection_pool = pool

def get_db_connection():
    if db_connection_pool is None:
        raise RuntimeError("Пул соединений не был установлен из главного приложения.")
    return db_connection_pool.getconn()

def release_db_connection(conn):
    if db_connection_pool and conn:
        db_connection_pool.putconn(conn)


# ---------------- Utils ----------------

def parse_date_any(s: str) -> Optional[date]:
    s = (s or "").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None

def exe_dir() -> str:
    import sys
    from pathlib import Path
    if getattr(sys, "frozen", False):
        return str(Path(sys.executable).resolve().parent)
    return str(Path(__file__).resolve().parent)

def _norm(s: str) -> str:
    return " ".join((s or "").strip().lower().split())

def safe_filename(name: str) -> str:
    name = (name or "").strip()
    name = re.sub(r'[\\/:*?"<>|]+', "_", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name or "file"


# ---------------- Business constants ----------------

COMPLEX_MAP = {
    "Комплекс 1": _norm("Одноразовое"),
    "Комплекс 2": _norm("Двухразовое"),
    "Комплекс 3": _norm("Трехразовое"),
}
COMPLEX_ORDER = ["Комплекс 1", "Комплекс 2", "Комплекс 3"]


# ---------------- Data helpers ----------------

def _load_price_map(conn) -> Dict[str, float]:
    with conn.cursor() as cur:
        cur.execute("SELECT name, COALESCE(price, 0) FROM meal_types")
        res: Dict[str, float] = {}
        for name, price in cur.fetchall():
            res[_norm(name)] = float(price or 0)
        return res

def _complex_by_meal_type_name(meal_type_name: str) -> Optional[str]:
    mt = _norm(meal_type_name)
    if "однораз" in mt:
        return "Комплекс 1"
    if "двухраз" in mt:
        return "Комплекс 2"
    if "трехраз" in mt or "трёхраз" in mt:
        return "Комплекс 3"
    return None

def _fetch_items_for_period(conn, date_from: date, date_to: date) -> List[Dict[str, Any]]:
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute(
            """
            SELECT
                mo.date::date AS service_date,
                COALESCE(mo.fact_address, o.address, '') AS object_address,
                COALESCE(mt.name, moi.meal_type_text, '') AS meal_type_name,
                COALESCE(moi.quantity, 1) AS qty
            FROM meal_orders mo
            JOIN meal_order_items moi ON moi.order_id = mo.id
            LEFT JOIN objects o ON o.id = mo.object_id
            LEFT JOIN meal_types mt ON mt.id = moi.meal_type_id
            WHERE mo.date >= %s AND mo.date <= %s
            """,
            (date_from, date_to),
        )
        return [dict(r) for r in cur.fetchall()]


# ---------------- Report builders ----------------

def build_daily_rows(date_from: date, date_to: date) -> List[Dict[str, Any]]:
    conn = None
    try:
        conn = get_db_connection()
        price_map = _load_price_map(conn)
        items = _fetch_items_for_period(conn, date_from, date_to)

        agg: Dict[Tuple[date, str], Dict[str, float]] = {}
        for it in items:
            d = it["service_date"]
            addr = (it.get("object_address") or "").strip() or "(без адреса)"
            cx = _complex_by_meal_type_name(it.get("meal_type_name") or "")
            if not cx:
                continue
            qty = float(it.get("qty") or 0)
            m = agg.setdefault((d, addr), {c: 0.0 for c in COMPLEX_ORDER})
            m[cx] += qty

        out: List[Dict[str, Any]] = []
        for (d, addr), cx_qty in sorted(agg.items(), key=lambda x: (x[0][0], x[0][1])):
            rec: Dict[str, Any] = {"date": d, "address": addr, "complexes": {}, "total_amount": 0.0}
            total = 0.0
            for cx in COMPLEX_ORDER:
                price = float(price_map.get(COMPLEX_MAP[cx], 0.0))
                q = float(cx_qty.get(cx, 0.0))
                amount = price * q
                rec["complexes"][cx] = {"price": price, "qty": q, "amount": amount}
                total += amount
            rec["total_amount"] = total
            out.append(rec)

        return out
    finally:
        if conn:
            release_db_connection(conn)

def build_monthly_rows(year: int, month: int) -> List[Dict[str, Any]]:
    if month < 1 or month > 12:
        raise ValueError("month must be 1..12")

    date_from = date(year, month, 1)
    date_to_excl = date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)

    conn = None
    try:
        conn = get_db_connection()
        price_map = _load_price_map(conn)

        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT
                    COALESCE(mo.fact_address, o.address, '') AS object_address,
                    COALESCE(mt.name, moi.meal_type_text, '') AS meal_type_name,
                    COALESCE(moi.quantity, 1) AS qty
                FROM meal_orders mo
                JOIN meal_order_items moi ON moi.order_id = mo.id
                LEFT JOIN objects o ON o.id = mo.object_id
                LEFT JOIN meal_types mt ON mt.id = moi.meal_type_id
                WHERE mo.date >= %s AND mo.date < %s
                """,
                (date_from, date_to_excl),
            )
            items = [dict(r) for r in cur.fetchall()]

        agg: Dict[str, Dict[str, float]] = {}
        for it in items:
            addr = (it.get("object_address") or "").strip() or "(без адреса)"
            cx = _complex_by_meal_type_name(it.get("meal_type_name") or "")
            if not cx:
                continue
            qty = float(it.get("qty") or 0)
            m = agg.setdefault(addr, {c: 0.0 for c in COMPLEX_ORDER})
            m[cx] += qty

        out: List[Dict[str, Any]] = []
        for addr, cx_qty in sorted(agg.items(), key=lambda x: x[0]):
            rec: Dict[str, Any] = {"address": addr, "complexes": {}, "total_amount": 0.0}
            total = 0.0
            for cx in COMPLEX_ORDER:
                price = float(price_map.get(COMPLEX_MAP[cx], 0.0))
                q = float(cx_qty.get(cx, 0.0))
                amount = price * q
                rec["complexes"][cx] = {"price": price, "qty": q, "amount": amount}
                total += amount
            rec["total_amount"] = total
            out.append(rec)

        return out
    finally:
        if conn:
            release_db_connection(conn)

def build_dept_employee_rows(department_id: int, date_from: date, date_to: date) -> List[Dict[str, Any]]:
    """
    Вариант B: одна строка на сотрудника.
    В ячейках: список бригад/объектов (уникальные значения) + суммы/кол-ва по комплексам.
    """
    conn = None
    try:
        conn = get_db_connection()
        price_map = _load_price_map(conn)

        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT
                    COALESCE(e.fio, moi.fio_text, '') AS fio,
                    COALESCE(e.tbn, moi.tbn_text, '') AS tbn,
                    COALESCE(e.position, moi.position_text, '') AS position_name,
                    COALESCE(mo.team_name, '') AS team_name,
                    COALESCE(mo.fact_address, o.address, '') AS object_address,
                    COALESCE(mt.name, moi.meal_type_text, '') AS meal_type_name,
                    COALESCE(moi.quantity, 1) AS qty
                FROM meal_orders mo
                JOIN meal_order_items moi ON moi.order_id = mo.id
                LEFT JOIN employees e ON e.id = moi.employee_id
                LEFT JOIN meal_types mt ON mt.id = moi.meal_type_id
                LEFT JOIN objects o ON o.id = mo.object_id
                WHERE mo.date >= %s AND mo.date <= %s
                  AND mo.department_id = %s
                """,
                (date_from, date_to, department_id),
            )
            items = [dict(r) for r in cur.fetchall()]

        agg_qty: Dict[Tuple[str, str, str], Dict[str, float]] = {}
        teams: Dict[Tuple[str, str, str], set[str]] = {}
        objs: Dict[Tuple[str, str, str], set[str]] = {}

        for it in items:
            fio = (it.get("fio") or "").strip()
            tbn = (it.get("tbn") or "").strip()
            pos = (it.get("position_name") or "").strip()
            if not fio:
                continue

            key = (fio, tbn, pos)

            team = (it.get("team_name") or "").strip()
            obj_addr = (it.get("object_address") or "").strip()

            if team:
                teams.setdefault(key, set()).add(team)
            if obj_addr:
                objs.setdefault(key, set()).add(obj_addr)

            cx = _complex_by_meal_type_name(it.get("meal_type_name") or "")
            if not cx:
                continue

            qty = float(it.get("qty") or 0)
            m = agg_qty.setdefault(key, {c: 0.0 for c in COMPLEX_ORDER})
            m[cx] += qty

        out: List[Dict[str, Any]] = []
        for (fio, tbn, pos), cx_qty in sorted(agg_qty.items(), key=lambda x: x[0][0]):
            rec: Dict[str, Any] = {
                "fio": fio,
                "tbn": tbn,
                "position": pos,
                "teams": "; ".join(sorted(teams.get((fio, tbn, pos), set()))),
                "objects": "; ".join(sorted(objs.get((fio, tbn, pos), set()))),
                "qty": {},
                "amount": {},
            }
            for cx in COMPLEX_ORDER:
                price = float(price_map.get(COMPLEX_MAP[cx], 0.0))
                q = float(cx_qty.get(cx, 0.0))
                rec["qty"][cx] = q
                rec["amount"][cx] = price * q
            out.append(rec)

        return out
    finally:
        if conn:
            release_db_connection(conn)


# ---------------- Excel exports ----------------

def export_daily_excel(date_from: date, date_to: date, out_path: str):
    rows = build_daily_rows(date_from, date_to)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ежедневный отчет"

    ws.append(["Ежедневный отчет по питанию"])
    ws.append([f"Период: {date_from:%Y-%m-%d} — {date_to:%Y-%m-%d}"])
    ws.append([])

    header = ["Дата оказания услуги", "Наименование строительного объекта"]
    for cx in COMPLEX_ORDER:
        header += [f"{cx} Цена", f"{cx} Кол-во порций", f"{cx} Стоимость"]
    header += ["ИТОГО стоимость"]
    ws.append(header)

    totals = {cx: 0.0 for cx in COMPLEX_ORDER}
    grand_total = 0.0

    for r in rows:
        row = [r["date"].strftime("%Y-%m-%d"), r["address"]]
        for cx in COMPLEX_ORDER:
            c = r["complexes"][cx]
            row += [c["price"], c["qty"], c["amount"]]
            totals[cx] += float(c["amount"] or 0)
            grand_total += float(c["amount"] or 0)
        row.append(r["total_amount"])
        ws.append(row)

    if not rows:
        ws.append(["Нет данных за выбранный период"])

    if rows:
        total_row = ["ИТОГО", ""]
        for cx in COMPLEX_ORDER:
            total_row += ["", "", totals[cx]]
        total_row += [grand_total]
        ws.append(total_row)

    widths = [14, 45] + [10, 14, 14] * 3 + [14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(out_path)

def export_monthly_excel(year: int, month: int, out_path: str):
    rows = build_monthly_rows(year, month)

    wb = Workbook()
    ws = wb.active
    ws.title = "Месячный отчет"

    ws.append(["Месячный отчет по питанию"])
    ws.append([f"Месяц: {year:04d}-{month:02d}"])
    ws.append([])

    header = ["Наименование строительного объекта"]
    for cx in COMPLEX_ORDER:
        header += [f"{cx} Цена", f"{cx} Кол-во порций", f"{cx} Стоимость"]
    header += ["ИТОГО стоимость"]
    ws.append(header)

    totals = {cx: 0.0 for cx in COMPLEX_ORDER}
    grand_total = 0.0

    for r in rows:
        row = [r["address"]]
        for cx in COMPLEX_ORDER:
            c = r["complexes"][cx]
            row += [c["price"], c["qty"], c["amount"]]
            totals[cx] += float(c["amount"] or 0)
            grand_total += float(c["amount"] or 0)
        row.append(r["total_amount"])
        ws.append(row)

    if not rows:
        ws.append(["Нет данных за выбранный месяц"])

    if rows:
        total_row = ["ИТОГО"]
        for cx in COMPLEX_ORDER:
            total_row += ["", "", totals[cx]]
        total_row += [grand_total]
        ws.append(total_row)

    widths = [55] + [10, 14, 14] * 3 + [14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(out_path)

def export_dept_employee_excel(dept_name: str, department_id: int, date_from: date, date_to: date, out_path: str):
    rows = build_dept_employee_rows(department_id, date_from, date_to)

    conn = None
    try:
        conn = get_db_connection()
        price_map = _load_price_map(conn)
    finally:
        if conn:
            release_db_connection(conn)

    p1 = float(price_map.get(COMPLEX_MAP["Комплекс 1"], 0.0))
    p2 = float(price_map.get(COMPLEX_MAP["Комплекс 2"], 0.0))
    p3 = float(price_map.get(COMPLEX_MAP["Комплекс 3"], 0.0))

    wb = Workbook()
    ws = wb.active
    ws.title = "По подразделению"

    ws.append(["г. Москва"])
    ws.append(["Отчет по расходам на питание (по подразделению)"])
    ws.append([f"Подразделение: {dept_name}"])
    ws.append([f"Период: {date_from:%Y-%m-%d} — {date_to:%Y-%m-%d}"])
    ws.append([])

    header = [
        "№ п/п",
        "Наименование подразделения",
        "Наименование бригад (список)",
        "Адреса объектов (список)",
        "Табельный номер",
        "Фамилия, имя, отчество работника",
        "Наименование профессии, должности",
        "Комплекс 1 (одноразовое питание)",
        "Комплекс 2 (двухразовое питание)",
        "Комплекс 3 (трехразовое питание)",
        f"Сумма расходов на питание Комплекс 1, руб. {p1:.2f} (с НДС)",
        f"Сумма расходов на питание Комплекс 2, руб. {p2:.2f} (с НДС)",
        f"Сумма расходов на питание Комплекс 3, руб. {p3:.2f} (с НДС)",
        "Подпись работника, подтверждающая понесенные расходы",
    ]
    ws.append(header)

    for i, r in enumerate(rows, start=1):
        ws.append([
            i,
            dept_name,
            r.get("teams") or "",
            r.get("objects") or "",
            r.get("tbn") or "",
            r.get("fio") or "",
            r.get("position") or "",
            r["qty"].get("Комплекс 1", 0.0),
            r["qty"].get("Комплекс 2", 0.0),
            r["qty"].get("Комплекс 3", 0.0),
            r["amount"].get("Комплекс 1", 0.0),
            r["amount"].get("Комплекс 2", 0.0),
            r["amount"].get("Комплекс 3", 0.0),
            "",
        ])

    if not rows:
        ws.append(["Нет данных за выбранный период/подразделение"])

    widths = [6, 26, 28, 38, 14, 32, 26, 14, 14, 14, 22, 22, 22, 24]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    wb.save(out_path)


# ---------------- UI Page (no tables) ----------------

class MealsReportsPage(tk.Frame):
    def __init__(self, master, app_ref=None):
        super().__init__(master, bg="#f7f7f7")
        self.app_ref = app_ref

        self.var_report_kind = tk.StringVar(value="daily")  # daily | monthly | dept_employee
        self.var_month = tk.StringVar(value=date.today().strftime("%Y-%m"))

        self._dept_rows: List[Tuple[int, str]] = []

        self._build_ui()
        self._load_departments()
        self._update_mode()

    def _build_ui(self):
        top = tk.Frame(self, bg="#f7f7f7")
        top.pack(fill="x", padx=12, pady=12)

        tk.Label(top, text="Тип отчета:", bg="#f7f7f7").grid(row=0, column=0, sticky="w")

        ttk.Radiobutton(
            top, text="Ежедневный (за период дат)", value="daily",
            variable=self.var_report_kind, command=self._update_mode
        ).grid(row=0, column=1, sticky="w", padx=(8, 10))

        ttk.Radiobutton(
            top, text="Месячный", value="monthly",
            variable=self.var_report_kind, command=self._update_mode
        ).grid(row=0, column=2, sticky="w", padx=(0, 10))

        ttk.Radiobutton(
            top, text="По подразделению (сотрудники)", value="dept_employee",
            variable=self.var_report_kind, command=self._update_mode
        ).grid(row=0, column=3, sticky="w", padx=(0, 10))

        self.frm_daily = tk.Frame(top, bg="#f7f7f7")
        self.frm_daily.grid(row=1, column=0, columnspan=6, sticky="w", pady=(10, 0))

        tk.Label(self.frm_daily, text="Дата с:", bg="#f7f7f7").grid(row=0, column=0, sticky="w")
        self.ent_from = ttk.Entry(self.frm_daily, width=12)
        self.ent_from.grid(row=0, column=1, sticky="w", padx=(6, 10))

        tk.Label(self.frm_daily, text="по:", bg="#f7f7f7").grid(row=0, column=2, sticky="w")
        self.ent_to = ttk.Entry(self.frm_daily, width=12)
        self.ent_to.grid(row=0, column=3, sticky="w", padx=(6, 10))

        today = date.today()
        self.ent_from.insert(0, today.strftime("%Y-%m-%d"))
        self.ent_to.insert(0, today.strftime("%Y-%m-%d"))

        self.frm_monthly = tk.Frame(top, bg="#f7f7f7")
        self.frm_monthly.grid(row=2, column=0, columnspan=6, sticky="w", pady=(10, 0))

        tk.Label(self.frm_monthly, text="Месяц (YYYY-MM):", bg="#f7f7f7").grid(row=0, column=0, sticky="w")
        self.ent_month = ttk.Entry(self.frm_monthly, textvariable=self.var_month, width=12)
        self.ent_month.grid(row=0, column=1, sticky="w", padx=(6, 10))

        self.frm_dept = tk.Frame(top, bg="#f7f7f7")
        self.frm_dept.grid(row=3, column=0, columnspan=6, sticky="w", pady=(10, 0))

        tk.Label(self.frm_dept, text="Подразделение:", bg="#f7f7f7").grid(row=0, column=0, sticky="w")
        self.cmb_dept = ttk.Combobox(self.frm_dept, state="readonly", width=45)
        self.cmb_dept.grid(row=0, column=1, sticky="w", padx=(6, 10))

        btns = tk.Frame(self, bg="#f7f7f7")
        btns.pack(fill="x", padx=12, pady=(0, 12))
        ttk.Button(btns, text="Выгрузить в Excel", command=self._on_export).pack(side="left")

    def _load_departments(self):
        conn = None
        try:
            conn = get_db_connection()
            with conn.cursor() as cur:
                cur.execute("SELECT id, name FROM departments ORDER BY name")
                rows = cur.fetchall()
            self._dept_rows = [(int(r[0]), str(r[1])) for r in rows]
            self.cmb_dept["values"] = [name for _, name in self._dept_rows]
            if self._dept_rows:
                self.cmb_dept.current(0)
        except Exception as e:
            messagebox.showerror("Отчеты", f"Не удалось загрузить подразделения:\n{e}", parent=self)
            self._dept_rows = []
            self.cmb_dept["values"] = []
        finally:
            if conn:
                release_db_connection(conn)

    def _selected_department_id(self) -> Optional[int]:
        name = (self.cmb_dept.get() or "").strip()
        for did, n in self._dept_rows:
            if n == name:
                return did
        return None

    def _update_mode(self):
        kind = self.var_report_kind.get()
        if kind == "daily":
            self.frm_daily.grid()
            self.frm_monthly.grid_remove()
            self.frm_dept.grid_remove()
        elif kind == "monthly":
            self.frm_monthly.grid()
            self.frm_daily.grid_remove()
            self.frm_dept.grid_remove()
        else:
            self.frm_daily.grid()
            self.frm_dept.grid()
            self.frm_monthly.grid_remove()

    def _on_export(self):
        kind = self.var_report_kind.get()

        out_dir = os.path.join(exe_dir(), "Заявки_питание")
        os.makedirs(out_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")

        try:
            if kind == "daily":
                d_from = parse_date_any(self.ent_from.get())
                d_to = parse_date_any(self.ent_to.get())
                if not d_from or not d_to:
                    messagebox.showwarning("Отчеты", "Укажите даты 'с' и 'по'.", parent=self)
                    return
                if d_from > d_to:
                    messagebox.showwarning("Отчеты", "Дата 'с' больше даты 'по'.", parent=self)
                    return

                out_path = os.path.join(out_dir, f"Питание_ежедневный_{d_from:%Y%m%d}-{d_to:%Y%m%d}_{ts}.xlsx")
                export_daily_excel(d_from, d_to, out_path)

            elif kind == "monthly":
                m = (self.var_month.get() or "").strip()
                try:
                    dt = datetime.strptime(m, "%Y-%m")
                    year, month = dt.year, dt.month
                except Exception:
                    messagebox.showwarning("Отчеты", "Месяц должен быть в формате YYYY-MM, например 2026-01.", parent=self)
                    return

                out_path = os.path.join(out_dir, f"Питание_месячный_{year:04d}{month:02d}_{ts}.xlsx")
                export_monthly_excel(year, month, out_path)

            else:
                d_from = parse_date_any(self.ent_from.get())
                d_to = parse_date_any(self.ent_to.get())
                if not d_from or not d_to:
                    messagebox.showwarning("Отчеты", "Укажите даты 'с' и 'по'.", parent=self)
                    return
                if d_from > d_to:
                    messagebox.showwarning("Отчеты", "Дата 'с' больше даты 'по'.", parent=self)
                    return

                dept_id = self._selected_department_id()
                dept_name = (self.cmb_dept.get() or "").strip()
                if not dept_id:
                    messagebox.showwarning("Отчеты", "Выберите подразделение.", parent=self)
                    return

                dept_safe = safe_filename(dept_name)
                out_path = os.path.join(out_dir, f"Питание_подразделение_{dept_safe}_{d_from:%Y%m%d}-{d_to:%Y%m%d}_{ts}.xlsx")
                export_dept_employee_excel(dept_name, dept_id, d_from, d_to, out_path)

            messagebox.showinfo("Отчеты", f"Файл сформирован:\n{out_path}", parent=self)
            try:
                os.startfile(out_path)
            except Exception:
                pass

        except Exception as e:
            messagebox.showerror("Отчеты", f"Не удалось сформировать отчет:\n{e}", parent=self)


def create_meals_reports_page(parent, app_ref=None) -> tk.Frame:
    try:
        return MealsReportsPage(parent, app_ref=app_ref)
    except Exception:
        import traceback
        messagebox.showerror("Питание — отчеты", traceback.format_exc(), parent=parent)
        return tk.Frame(parent)
