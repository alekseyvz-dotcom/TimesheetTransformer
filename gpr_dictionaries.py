# gpr_dictionaries.py — профессиональные справочники ГПР
from __future__ import annotations
import os
import logging
from datetime import datetime, date, timedelta
from typing import Any, Dict, List, Optional, Tuple

import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog

from psycopg2.extras import RealDictCursor


logger = logging.getLogger(__name__)

# ═══════════════════════════════════════════════════════════════
#  DB POOL
# ═══════════════════════════════════════════════════════════════
db_connection_pool = None


def set_db_pool(pool):
    global db_connection_pool
    db_connection_pool = pool


def _conn():
    if not db_connection_pool:
        raise RuntimeError("DB pool not set (gpr_dictionaries)")
    return db_connection_pool.getconn()


def _release(conn):
    if db_connection_pool and conn:
        try:
            db_connection_pool.putconn(conn)
        except Exception:
            logger.exception("Error releasing DB connection in gpr_dictionaries")


# ═══════════════════════════════════════════════════════════════
#  CONST
# ═══════════════════════════════════════════════════════════════
C = {
    "bg": "#f0f2f5",
    "panel": "#ffffff",
    "accent": "#1565c0",
    "accent_light": "#e3f2fd",
    "success": "#2e7d32",
    "warning": "#ed6c02",
    "error": "#d32f2f",
    "border": "#dde1e7",
    "text": "#1a1a2e",
    "text2": "#555",
    "text3": "#999",
    "btn_bg": "#1565c0",
    "btn_fg": "#ffffff",
}

STATUS_LABELS = {
    "planned": "Запланировано",
    "in_progress": "В работе",
    "done": "Выполнено",
    "paused": "Приостановлено",
    "canceled": "Отменено",
}
STATUS_LIST = ["planned", "in_progress", "done", "paused", "canceled"]

ROW_KIND_LABELS = {
    "task": "Работа",
    "group": "Группа",
    "title": "Титул",
}
ROW_KIND_LIST = ["task", "group", "title"]

def _today() -> date:
    return datetime.now().date()


def _safe_float(v):
    if v is None:
        return None
    try:
        s = str(v).strip().replace(",", ".")
        if not s:
            return None
        return float(s)
    except Exception:
        return None


def _fmt_qty(v) -> str:
    f = _safe_float(v)
    if f is None:
        return ""
    return f"{f:.3f}".rstrip("0").rstrip(".")


def _fmt_dt(v) -> str:
    if isinstance(v, datetime):
        return v.strftime("%d.%m.%Y %H:%M")
    return str(v or "—")

def _fmt_date_iso(v) -> str:
    """Дата для ввода в справочнике норм: ГГГГ-ММ-ДД."""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v or "").strip()


def _parse_date_iso(value: str, field_name: str = "Дата") -> date:
    text = (value or "").strip()
    if not text:
        raise ValueError(f"{field_name} обязательна.")

    for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass

    raise ValueError(
        f"{field_name} должна быть указана в формате ГГГГ-ММ-ДД "
        f"или ДД.ММ.ГГГГ."
    )

def _user_id(app_ref) -> Optional[int]:
    try:
        return (getattr(app_ref, "current_user", None) or {}).get("id")
    except Exception:
        return None


# ═══════════════════════════════════════════════════════════════
#  SERVICES
# ═══════════════════════════════════════════════════════════════
class GprTemplateService:
    @staticmethod
    def load_templates_full(search: str = "") -> List[Dict[str, Any]]:
        conn = None
        try:
            conn = _conn()
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                params = []
                where = ""
                if search.strip():
                    where = """
                        WHERE t.name ILIKE %s
                           OR COALESCE(t.category, '') ILIKE %s
                           OR COALESCE(t.description, '') ILIKE %s
                    """
                    q = f"%{search.strip()}%"
                    params.extend([q, q, q])

                cur.execute(f"""
                    SELECT t.id,
                           t.name,
                           t.description,
                           t.category,
                           t.note,
                           t.is_active,
                           t.created_at,
                           t.updated_at,
                           COALESCE(cu.full_name, '') AS creator_name,
                           COALESCE(uu.full_name, '') AS updater_name,
                           COALESCE(cnt.task_count, 0) AS task_count
                    FROM public.gpr_templates t
                    LEFT JOIN public.app_users cu ON cu.id = t.created_by
                    LEFT JOIN public.app_users uu ON uu.id = t.updated_by
                    LEFT JOIN (
                        SELECT template_id, COUNT(*) AS task_count
                        FROM public.gpr_template_tasks
                        GROUP BY template_id
                    ) cnt ON cnt.template_id = t.id
                    {where}
                    ORDER BY t.is_active DESC, t.name
                """, params)
                return [dict(r) for r in cur.fetchall()]
        finally:
            _release(conn)

    @staticmethod
    def load_template_tasks(template_id: int) -> List[Dict[str, Any]]:
        conn = None
        try:
            conn = _conn()
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("""
                    SELECT tt.id,
                           tt.template_id,
                           tt.work_type_id,
                           tt.parent_id,
                           wt.name AS wt_name,
                           tt.name,
                           tt.uom_code,
                           tt.default_qty,
                           tt.is_milestone,
                           tt.sort_order,
                           COALESCE(tt.row_kind, 'task') AS row_kind,
                           tt.created_at,
                           tt.updated_at
                    FROM public.gpr_template_tasks tt
                    JOIN public.gpr_work_types wt ON wt.id = tt.work_type_id
                    WHERE tt.template_id = %s
                    ORDER BY tt.sort_order, tt.id
                """, (template_id,))
                return [dict(r) for r in cur.fetchall()]
        finally:
            _release(conn)

    @staticmethod
    def create_template(
        name: str,
        category: Optional[str] = None,
        description: Optional[str] = None,
        note: Optional[str] = None,
        created_by: Optional[int] = None,
    ) -> int:
        conn = None
        try:
            conn = _conn()
            with conn, conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO public.gpr_templates
                        (name, category, description, note, is_active, created_by, updated_by)
                    VALUES (%s, %s, %s, %s, true, %s, %s)
                    RETURNING id
                """, (
                    name.strip(),
                    (category or "").strip() or None,
                    (description or "").strip() or None,
                    (note or "").strip() or None,
                    created_by,
                    created_by,
                ))
                return int(cur.fetchone()[0])
        finally:
            _release(conn)

    @staticmethod
    def update_template(
        template_id: int,
        name: str,
        category: Optional[str] = None,
        description: Optional[str] = None,
        note: Optional[str] = None,
        updated_by: Optional[int] = None,
    ) -> None:
        conn = None
        try:
            conn = _conn()
            with conn, conn.cursor() as cur:
                cur.execute("""
                    UPDATE public.gpr_templates
                    SET name=%s,
                        category=%s,
                        description=%s,
                        note=%s,
                        updated_by=%s
                    WHERE id=%s
                """, (
                    name.strip(),
                    (category or "").strip() or None,
                    (description or "").strip() or None,
                    (note or "").strip() or None,
                    updated_by,
                    template_id,
                ))
        finally:
            _release(conn)

    @staticmethod
    def toggle_template(template_id: int, updated_by: Optional[int] = None) -> None:
        conn = None
        try:
            conn = _conn()
            with conn, conn.cursor() as cur:
                cur.execute("""
                    UPDATE public.gpr_templates
                    SET is_active = NOT is_active,
                        updated_by=%s
                    WHERE id=%s
                """, (updated_by, template_id))
        finally:
            _release(conn)

    @staticmethod
    def delete_template(template_id: int) -> None:
        conn = None
        try:
            conn = _conn()
            with conn, conn.cursor() as cur:
                cur.execute("DELETE FROM public.gpr_templates WHERE id=%s", (template_id,))
        finally:
            _release(conn)

    @staticmethod
    def duplicate_template(
        template_id: int,
        new_name: str,
        created_by: Optional[int] = None,
    ) -> int:
        conn = None
        try:
            conn = _conn()
            with conn, conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("""
                    SELECT name, category, description, note
                    FROM public.gpr_templates
                    WHERE id=%s
                """, (template_id,))
                src = cur.fetchone()
                if not src:
                    raise ValueError("Шаблон не найден")

                cur.execute("""
                    INSERT INTO public.gpr_templates
                        (name, category, description, note, is_active, created_by, updated_by)
                    VALUES (%s, %s, %s, %s, true, %s, %s)
                    RETURNING id
                """, (
                    new_name.strip(),
                    src.get("category"),
                    src.get("description"),
                    src.get("note"),
                    created_by,
                    created_by,
                ))
                new_tpl_id = int(cur.fetchone()["id"])

                cur.execute("""
                    SELECT id, work_type_id, parent_id, name, uom_code,
                           default_qty, is_milestone, sort_order,
                           COALESCE(row_kind, 'task') AS row_kind
                    FROM public.gpr_template_tasks
                    WHERE template_id=%s
                    ORDER BY sort_order, id
                """, (template_id,))
                old_tasks = [dict(r) for r in cur.fetchall()]

                id_map: Dict[int, int] = {}

                for t in old_tasks:
                    cur.execute("""
                        INSERT INTO public.gpr_template_tasks
                            (template_id, work_type_id, parent_id, name, uom_code,
                             default_qty, is_milestone, sort_order, row_kind)
                        VALUES (%s, %s, NULL, %s, %s, %s, %s, %s, %s)
                        RETURNING id
                    """, (
                        new_tpl_id,
                        t["work_type_id"],
                        t["name"],
                        t.get("uom_code"),
                        t.get("default_qty"),
                        t.get("is_milestone", False),
                        t.get("sort_order", 0),
                        t.get("row_kind", "task"),
                    ))
                    new_id = int(cur.fetchone()["id"])
                    id_map[int(t["id"])] = new_id

                for t in old_tasks:
                    old_parent = t.get("parent_id")
                    if old_parent and int(old_parent) in id_map:
                        cur.execute("""
                            UPDATE public.gpr_template_tasks
                            SET parent_id=%s
                            WHERE id=%s
                        """, (id_map[int(old_parent)], id_map[int(t["id"])]))

                return new_tpl_id
        finally:
            _release(conn)


class GprLaborDictionaryService:
    """Справочник работ и версий нормативов ЗТР."""

    @staticmethod
    def load_work_items(search: str = "") -> List[Dict[str, Any]]:
        conn = None
        try:
            conn = _conn()
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                params: List[Any] = []
                where = ""

                if search.strip():
                    q = f"%{search.strip()}%"
                    where = """
                        WHERE wi.name ILIKE %s
                           OR COALESCE(wi.code, '') ILIKE %s
                           OR wt.name ILIKE %s
                           OR wi.uom_code ILIKE %s
                    """
                    params = [q, q, q, q]

                cur.execute(
                    f"""
                    SELECT
                        wi.id,
                        wi.work_type_id,
                        wt.name AS work_type_name,
                        COALESCE(wi.code, '') AS code,
                        wi.name,
                        wi.uom_code,
                        COALESCE(u.name, '') AS uom_name,
                        wi.sort_order,
                        wi.is_active,
                        COALESCE(wi.note, '') AS note,
                        wi.created_at,
                        wi.updated_at,

                        current_norm.id AS current_norm_id,
                        current_norm.labor_hours_per_unit AS current_labor_hours_per_unit,
                        current_norm.default_productivity_factor
                            AS current_productivity_factor,
                        current_norm.effective_from AS current_norm_from,
                        current_norm.effective_to AS current_norm_to,
                        COALESCE(current_norm.source_name, '') AS current_norm_source

                    FROM public.gpr_work_items wi
                    JOIN public.gpr_work_types wt
                        ON wt.id = wi.work_type_id
                    LEFT JOIN public.gpr_uom u
                        ON u.code = wi.uom_code

                    LEFT JOIN LATERAL (
                        SELECT n.*
                        FROM public.gpr_labor_norms n
                        WHERE n.work_item_id = wi.id
                          AND n.is_active = true
                          AND n.effective_from <= CURRENT_DATE
                          AND (
                              n.effective_to IS NULL
                              OR n.effective_to >= CURRENT_DATE
                          )
                        ORDER BY n.effective_from DESC, n.id DESC
                        LIMIT 1
                    ) current_norm ON true

                    {where}

                    ORDER BY
                        wi.is_active DESC,
                        wt.sort_order,
                        wt.name,
                        wi.sort_order,
                        wi.name
                    """,
                    params,
                )
                return [dict(row) for row in cur.fetchall()]
        finally:
            _release(conn)

    @staticmethod
    def load_norms(work_item_id: int) -> List[Dict[str, Any]]:
        conn = None
        try:
            conn = _conn()
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    """
                    SELECT
                        n.id,
                        n.work_item_id,
                        n.effective_from,
                        n.effective_to,
                        n.labor_hours_per_unit,
                        n.default_productivity_factor,
                        COALESCE(n.source_name, '') AS source_name,
                        COALESCE(n.source_code, '') AS source_code,
                        COALESCE(n.source_version, '') AS source_version,
                        COALESCE(n.note, '') AS note,
                        n.is_active,
                        n.created_at,
                        n.updated_at,
                        COALESCE(cu.full_name, '') AS creator_name,
                        COALESCE(uu.full_name, '') AS updater_name
                    FROM public.gpr_labor_norms n
                    LEFT JOIN public.app_users cu
                        ON cu.id = n.created_by
                    LEFT JOIN public.app_users uu
                        ON uu.id = n.updated_by
                    WHERE n.work_item_id = %s
                    ORDER BY n.effective_from DESC, n.id DESC
                    """,
                    (work_item_id,),
                )
                return [dict(row) for row in cur.fetchall()]
        finally:
            _release(conn)

    @staticmethod
    def create_work_item(
        *,
        work_type_id: int,
        code: Optional[str],
        name: str,
        uom_code: str,
        sort_order: int,
        note: Optional[str],
        user_id: Optional[int],
    ) -> int:
        conn = None
        try:
            conn = _conn()
            with conn, conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO public.gpr_work_items (
                        work_type_id,
                        code,
                        name,
                        uom_code,
                        sort_order,
                        is_active,
                        note,
                        created_by,
                        updated_by
                    )
                    VALUES (%s, %s, %s, %s, %s, true, %s, %s, %s)
                    RETURNING id
                    """,
                    (
                        work_type_id,
                        (code or "").strip() or None,
                        name.strip(),
                        uom_code,
                        sort_order,
                        (note or "").strip() or None,
                        user_id,
                        user_id,
                    ),
                )
                return int(cur.fetchone()[0])
        finally:
            _release(conn)

    @staticmethod
    def update_work_item(
        work_item_id: int,
        *,
        work_type_id: int,
        code: Optional[str],
        name: str,
        uom_code: str,
        sort_order: int,
        note: Optional[str],
        user_id: Optional[int],
    ) -> None:
        conn = None
        try:
            conn = _conn()
            with conn, conn.cursor() as cur:
                cur.execute(
                    """
                    UPDATE public.gpr_work_items
                    SET
                        work_type_id = %s,
                        code = %s,
                        name = %s,
                        uom_code = %s,
                        sort_order = %s,
                        note = %s,
                        updated_by = %s
                    WHERE id = %s
                    """,
                    (
                        work_type_id,
                        (code or "").strip() or None,
                        name.strip(),
                        uom_code,
                        sort_order,
                        (note or "").strip() or None,
                        user_id,
                        work_item_id,
                    ),
                )
        finally:
            _release(conn)

    @staticmethod
    def toggle_work_item(
        work_item_id: int,
        user_id: Optional[int],
    ) -> None:
        conn = None
        try:
            conn = _conn()
            with conn, conn.cursor() as cur:
                cur.execute(
                    """
                    UPDATE public.gpr_work_items
                    SET
                        is_active = NOT is_active,
                        updated_by = %s
                    WHERE id = %s
                    """,
                    (user_id, work_item_id),
                )
        finally:
            _release(conn)

    @staticmethod
    def create_labor_norm(
        *,
        work_item_id: int,
        effective_from: date,
        effective_to: Optional[date],
        labor_hours_per_unit: float,
        default_productivity_factor: float,
        source_name: Optional[str],
        source_code: Optional[str],
        source_version: Optional[str],
        note: Optional[str],
        user_id: Optional[int],
    ) -> int:
        conn = None
        try:
            conn = _conn()
            with conn, conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO public.gpr_labor_norms (
                        work_item_id,
                        effective_from,
                        effective_to,
                        labor_hours_per_unit,
                        default_productivity_factor,
                        source_name,
                        source_code,
                        source_version,
                        note,
                        is_active,
                        created_by,
                        updated_by
                    )
                    VALUES (
                        %s, %s, %s, %s, %s,
                        %s, %s, %s, %s,
                        true, %s, %s
                    )
                    RETURNING id
                    """,
                    (
                        work_item_id,
                        effective_from,
                        effective_to,
                        labor_hours_per_unit,
                        default_productivity_factor,
                        (source_name or "").strip() or None,
                        (source_code or "").strip() or None,
                        (source_version or "").strip() or None,
                        (note or "").strip() or None,
                        user_id,
                        user_id,
                    ),
                )
                return int(cur.fetchone()[0])
        finally:
            _release(conn)

    @staticmethod
    def update_labor_norm(
        norm_id: int,
        *,
        effective_from: date,
        effective_to: Optional[date],
        labor_hours_per_unit: float,
        default_productivity_factor: float,
        source_name: Optional[str],
        source_code: Optional[str],
        source_version: Optional[str],
        note: Optional[str],
        user_id: Optional[int],
    ) -> None:
        conn = None
        try:
            conn = _conn()
            with conn, conn.cursor() as cur:
                cur.execute(
                    """
                    UPDATE public.gpr_labor_norms
                    SET
                        effective_from = %s,
                        effective_to = %s,
                        labor_hours_per_unit = %s,
                        default_productivity_factor = %s,
                        source_name = %s,
                        source_code = %s,
                        source_version = %s,
                        note = %s,
                        updated_by = %s
                    WHERE id = %s
                    """,
                    (
                        effective_from,
                        effective_to,
                        labor_hours_per_unit,
                        default_productivity_factor,
                        (source_name or "").strip() or None,
                        (source_code or "").strip() or None,
                        (source_version or "").strip() or None,
                        (note or "").strip() or None,
                        user_id,
                        norm_id,
                    ),
                )
        finally:
            _release(conn)

    @staticmethod
    def toggle_labor_norm(
        norm_id: int,
        user_id: Optional[int],
    ) -> None:
        conn = None
        try:
            conn = _conn()
            with conn, conn.cursor() as cur:
                cur.execute(
                    """
                    UPDATE public.gpr_labor_norms
                    SET
                        is_active = NOT is_active,
                        updated_by = %s
                    WHERE id = %s
                    """,
                    (user_id, norm_id),
                )
        finally:
            _release(conn)

class GprLaborExcelImportService:
    """
    Импорт работ и норм ЗТР из Excel.

    Ожидаемые колонки:
      - Раздел
      - Вид работы
      - Ед. изм.
      - ЗТР

    Раздел Excel становится типом работ.
    Вид работы становится конкретной работой.
    """

    REQUIRED_HEADERS = {
        "раздел": "section",
        "вид работы": "work_name",
        "ед. изм.": "uom_code",
        "ед изм": "uom_code",
        "единица измерения": "uom_code",
        "зтр": "labor_hours_per_unit",
    }

    @staticmethod
    def _norm_text(value: Any) -> str:
        return " ".join(str(value or "").strip().split())

    @staticmethod
    def _header_key(value: Any) -> str:
        text = GprLaborExcelImportService._norm_text(value).lower()
        text = text.replace("ё", "е")
        text = text.replace(".", "")
        return text

    @staticmethod
    def read_excel_rows(file_path: str) -> Tuple[List[Dict[str, Any]], List[str], str]:
        """
        Читает Excel и возвращает:
          rows       - подготовленные уникальные строки;
          warnings   - предупреждения;
          sheet_name - использованный лист.
        """
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "Не установлена библиотека openpyxl.\n"
                "Установите её командой: pip install openpyxl"
            ) from exc

        wb = load_workbook(
            filename=file_path,
            read_only=True,
            data_only=True,
        )

        try:
            ws = wb.active
            sheet_name = ws.title

            rows_iter = ws.iter_rows(values_only=True)

            try:
                raw_header = next(rows_iter)
            except StopIteration:
                raise ValueError("Файл Excel пустой.")

            column_map: Dict[str, int] = {}

            for index, value in enumerate(raw_header):
                header = GprLaborExcelImportService._header_key(value)
                mapped = GprLaborExcelImportService.REQUIRED_HEADERS.get(header)
                if mapped:
                    column_map[mapped] = index

            required_fields = {
                "section": "Раздел",
                "work_name": "Вид работы",
                "uom_code": "Ед. изм.",
                "labor_hours_per_unit": "ЗТР",
            }

            missing = [
                title
                for key, title in required_fields.items()
                if key not in column_map
            ]

            if missing:
                raise ValueError(
                    "Не найдены обязательные колонки:\n- "
                    + "\n- ".join(missing)
                    + "\n\nОжидаются колонки: Раздел, Вид работы, Ед. изм., ЗТР."
                )

            prepared_rows: List[Dict[str, Any]] = []
            warnings: List[str] = []
            seen: Dict[Tuple[str, str, str], float] = {}

            for excel_row_no, raw_row in enumerate(rows_iter, start=2):
                section = GprLaborExcelImportService._norm_text(
                    raw_row[column_map["section"]]
                    if len(raw_row) > column_map["section"]
                    else None
                )
                work_name = GprLaborExcelImportService._norm_text(
                    raw_row[column_map["work_name"]]
                    if len(raw_row) > column_map["work_name"]
                    else None
                )
                uom_code = GprLaborExcelImportService._norm_text(
                    raw_row[column_map["uom_code"]]
                    if len(raw_row) > column_map["uom_code"]
                    else None
                )

                raw_labor = (
                    raw_row[column_map["labor_hours_per_unit"]]
                    if len(raw_row) > column_map["labor_hours_per_unit"]
                    else None
                )
                labor = _safe_float(raw_labor)

                # Полностью пустую строку не считаем ошибкой.
                if not section and not work_name and not uom_code and labor is None:
                    continue

                if not section:
                    warnings.append(
                        f"Строка {excel_row_no}: не указан раздел — строка пропущена."
                    )
                    continue

                if not work_name:
                    warnings.append(
                        f"Строка {excel_row_no}: не указан вид работы — строка пропущена."
                    )
                    continue

                if not uom_code:
                    warnings.append(
                        f"Строка {excel_row_no}: не указана единица измерения — строка пропущена."
                    )
                    continue

                if labor is None or labor <= 0:
                    warnings.append(
                        f"Строка {excel_row_no}: ЗТР должен быть числом больше 0 — строка пропущена."
                    )
                    continue

                key = (
                    section.casefold(),
                    work_name.casefold(),
                    uom_code.casefold(),
                )

                # Одинаковые строки в Excel разрешены и игнорируются.
                if key in seen:
                    if abs(seen[key] - labor) < 0.0000001:
                        warnings.append(
                            f"Строка {excel_row_no}: дубликат строки проигнорирован."
                        )
                    else:
                        warnings.append(
                            f"Строка {excel_row_no}: для той же работы указано "
                            f"другое значение ЗТР ({_fmt_qty(labor)}). "
                            "Строка пропущена."
                        )
                    continue

                seen[key] = labor

                prepared_rows.append(
                    {
                        "excel_row_no": excel_row_no,
                        "section": section,
                        "work_name": work_name,
                        "uom_code": uom_code,
                        "labor_hours_per_unit": labor,
                    }
                )

            if not prepared_rows:
                raise ValueError(
                    "В Excel не найдено ни одной корректной строки для импорта."
                )

            return prepared_rows, warnings, sheet_name

        finally:
            wb.close()

    @staticmethod
    def import_rows(
        rows: List[Dict[str, Any]],
        *,
        effective_from: date,
        source_name: str,
        user_id: Optional[int],
        import_note: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        Импортирует строки в одной транзакции.

        При повторном импорте:
        - если норма имеет ту же дату effective_from — она обновляется;
        - если дата новая — старая активная норма закрывается предыдущим днем;
        - если есть будущая норма — новая ограничивается днем до будущей нормы.
        """
        conn = None

        result = {
            "work_types_created": 0,
            "uoms_created": 0,
            "work_items_created": 0,
            "norms_created": 0,
            "norms_updated": 0,
            "errors": [],
        }

        try:
            conn = _conn()

            with conn, conn.cursor(cursor_factory=RealDictCursor) as cur:
                # ─────────────────────────────────────────────
                # Кэш типов работ
                # ─────────────────────────────────────────────
                cur.execute("""
                    SELECT id, name
                    FROM public.gpr_work_types
                """)
                work_types_cache = {
                    GprLaborExcelImportService._norm_text(row["name"]).casefold(): int(row["id"])
                    for row in cur.fetchall()
                }

                # ─────────────────────────────────────────────
                # Кэш единиц измерения
                # ─────────────────────────────────────────────
                cur.execute("""
                    SELECT code, name
                    FROM public.gpr_uom
                """)
                uom_cache = {
                    GprLaborExcelImportService._norm_text(row["code"]).casefold(): row["code"]
                    for row in cur.fetchall()
                }

                # ─────────────────────────────────────────────
                # Кэш работ
                # ─────────────────────────────────────────────
                cur.execute("""
                    SELECT id, work_type_id, name, uom_code
                    FROM public.gpr_work_items
                """)

                work_items_cache: Dict[Tuple[int, str, str], int] = {}

                for row in cur.fetchall():
                    key = (
                        int(row["work_type_id"]),
                        GprLaborExcelImportService._norm_text(row["name"]).casefold(),
                        GprLaborExcelImportService._norm_text(row["uom_code"]).casefold(),
                    )
                    work_items_cache[key] = int(row["id"])

                for item in rows:
                    section = item["section"]
                    work_name = item["work_name"]
                    uom_code_from_excel = item["uom_code"]
                    labor_hours = item["labor_hours_per_unit"]

                    # ─────────────────────────────────────────
                    # 1. Тип работ = Раздел Excel
                    # ─────────────────────────────────────────
                    section_key = section.casefold()
                    work_type_id = work_types_cache.get(section_key)

                    if not work_type_id:
                        cur.execute("""
                            INSERT INTO public.gpr_work_types (
                                code,
                                name,
                                sort_order,
                                is_active
                            )
                            VALUES (NULL, %s, 100, true)
                            RETURNING id
                        """, (section,))

                        work_type_id = int(cur.fetchone()["id"])
                        work_types_cache[section_key] = work_type_id
                        result["work_types_created"] += 1

                    # ─────────────────────────────────────────
                    # 2. Единица измерения
                    # ─────────────────────────────────────────
                    uom_key = uom_code_from_excel.casefold()
                    uom_code = uom_cache.get(uom_key)

                    if not uom_code:
                        cur.execute("""
                            INSERT INTO public.gpr_uom (code, name)
                            VALUES (%s, %s)
                        """, (
                            uom_code_from_excel,
                            uom_code_from_excel,
                        ))

                        uom_code = uom_code_from_excel
                        uom_cache[uom_key] = uom_code
                        result["uoms_created"] += 1

                    # ─────────────────────────────────────────
                    # 3. Конкретная работа
                    # ─────────────────────────────────────────
                    work_key = (
                        work_type_id,
                        work_name.casefold(),
                        uom_code.casefold(),
                    )

                    work_item_id = work_items_cache.get(work_key)

                    if not work_item_id:
                        cur.execute("""
                            INSERT INTO public.gpr_work_items (
                                work_type_id,
                                code,
                                name,
                                uom_code,
                                sort_order,
                                is_active,
                                note,
                                created_by,
                                updated_by
                            )
                            VALUES (
                                %s, NULL, %s, %s,
                                100, true, %s, %s, %s
                            )
                            RETURNING id
                        """, (
                            work_type_id,
                            work_name,
                            uom_code,
                            import_note,
                            user_id,
                            user_id,
                        ))

                        work_item_id = int(cur.fetchone()["id"])
                        work_items_cache[work_key] = work_item_id
                        result["work_items_created"] += 1

                    # ─────────────────────────────────────────
                    # 4. Норма ЗТР
                    # ─────────────────────────────────────────
                    # Блокируем все нормы данной работы.
                    cur.execute("""
                        SELECT id,
                               effective_from,
                               effective_to
                        FROM public.gpr_labor_norms
                        WHERE work_item_id = %s
                        ORDER BY effective_from, id
                        FOR UPDATE
                    """, (work_item_id,))

                    existing_norms = [dict(row) for row in cur.fetchall()]

                    same_start_norm = next(
                        (
                            norm
                            for norm in existing_norms
                            if norm["effective_from"] == effective_from
                        ),
                        None,
                    )

                    previous_norm = None
                    next_norm = None

                    for norm in existing_norms:
                        norm_from = norm["effective_from"]

                        if norm_from < effective_from:
                            if (
                                previous_norm is None
                                or norm_from > previous_norm["effective_from"]
                            ):
                                previous_norm = norm

                        elif norm_from > effective_from:
                            if (
                                next_norm is None
                                or norm_from < next_norm["effective_from"]
                            ):
                                next_norm = norm

                    # Если новая редакция нормы начинается позже старой,
                    # закрываем старую норму днем ранее.
                    if previous_norm:
                        prev_end = previous_norm.get("effective_to")

                        if prev_end is None or prev_end >= effective_from:
                            cur.execute("""
                                UPDATE public.gpr_labor_norms
                                SET effective_to = %s,
                                    updated_by = %s
                                WHERE id = %s
                            """, (
                                effective_from - timedelta(days=1),
                                user_id,
                                previous_norm["id"],
                            ))

                    # Если в справочнике уже есть будущая версия,
                    # новая норма действует только до дня перед ней.
                    calculated_effective_to = (
                        next_norm["effective_from"] - timedelta(days=1)
                        if next_norm
                        else None
                    )

                    if same_start_norm:
                        cur.execute("""
                            UPDATE public.gpr_labor_norms
                            SET effective_to = %s,
                                labor_hours_per_unit = %s,
                                default_productivity_factor = 1.000000,
                                source_name = %s,
                                source_code = NULL,
                                source_version = NULL,
                                note = %s,
                                is_active = true,
                                updated_by = %s
                            WHERE id = %s
                        """, (
                            calculated_effective_to,
                            labor_hours,
                            source_name,
                            import_note,
                            user_id,
                            same_start_norm["id"],
                        ))

                        result["norms_updated"] += 1

                    else:
                        cur.execute("""
                            INSERT INTO public.gpr_labor_norms (
                                work_item_id,
                                effective_from,
                                effective_to,
                                labor_hours_per_unit,
                                default_productivity_factor,
                                source_name,
                                source_code,
                                source_version,
                                note,
                                is_active,
                                created_by,
                                updated_by
                            )
                            VALUES (
                                %s, %s, %s, %s,
                                1.000000,
                                %s, NULL, NULL, %s,
                                true, %s, %s
                            )
                        """, (
                            work_item_id,
                            effective_from,
                            calculated_effective_to,
                            labor_hours,
                            source_name,
                            import_note,
                            user_id,
                            user_id,
                        ))

                        result["norms_created"] += 1

            return result

        except Exception:
            logger.exception("Ошибка импорта норм ЗТР из Excel")
            raise

        finally:
            _release(conn)

# ═══════════════════════════════════════════════════════════════
#  DIALOGS
# ═══════════════════════════════════════════════════════════════
class _DictEditDialog(simpledialog.Dialog):
    """Диалог добавления/редактирования одной записи справочника."""

    def __init__(
        self,
        parent,
        title_text: str,
        fields: List[Tuple[str, str, int]],
        init: Optional[Dict[str, str]] = None,
        readonly_keys: Optional[List[str]] = None,
    ):
        self._fields = fields
        self._init = init or {}
        self._readonly_keys = set(readonly_keys or [])
        self.result: Optional[Dict[str, str]] = None
        super().__init__(parent, title=title_text)

    def body(self, m):
        self._vars: Dict[str, tk.StringVar] = {}
        self._entries: Dict[str, ttk.Entry] = {}
        for i, (label, key, width) in enumerate(self._fields):
            tk.Label(m, text=label + ":").grid(
                row=i, column=0, sticky="e", padx=(0, 6), pady=4
            )
            var = tk.StringVar(value=str(self._init.get(key, "")))
            ent = ttk.Entry(m, textvariable=var, width=width)
            ent.grid(row=i, column=1, sticky="w", pady=4)
            if key in self._readonly_keys:
                ent.state(["readonly"])
            self._vars[key] = var
            self._entries[key] = ent

        if self._fields:
            first_key = self._fields[0][1]
            return self._entries[first_key]
        return None

    def validate(self):
        self._out = {k: v.get().strip() for k, v in self._vars.items()}
        return True

    def apply(self):
        self.result = self._out


class _TemplateEditDialog(simpledialog.Dialog):
    def __init__(self, parent, init: Optional[Dict[str, Any]] = None):
        self.init = init or {}
        self.result: Optional[Dict[str, Any]] = None
        super().__init__(parent, title="Карточка шаблона")

    def body(self, m):
        m.grid_columnconfigure(1, weight=1)

        tk.Label(m, text="Наименование *:").grid(
            row=0, column=0, sticky="e", padx=(0, 8), pady=5
        )
        self.var_name = tk.StringVar(value=self.init.get("name", ""))
        self.ent_name = ttk.Entry(m, textvariable=self.var_name, width=52)
        self.ent_name.grid(row=0, column=1, sticky="ew", pady=5)

        tk.Label(m, text="Категория:").grid(
            row=1, column=0, sticky="e", padx=(0, 8), pady=5
        )
        self.var_category = tk.StringVar(value=self.init.get("category", ""))
        self.ent_category = ttk.Entry(m, textvariable=self.var_category, width=52)
        self.ent_category.grid(row=1, column=1, sticky="ew", pady=5)

        tk.Label(m, text="Описание:").grid(
            row=2, column=0, sticky="ne", padx=(0, 8), pady=5
        )
        self.txt_desc = tk.Text(m, width=52, height=4, wrap="word")
        self.txt_desc.grid(row=2, column=1, sticky="ew", pady=5)
        self.txt_desc.insert("1.0", self.init.get("description", "") or "")

        tk.Label(m, text="Примечание:").grid(
            row=3, column=0, sticky="ne", padx=(0, 8), pady=5
        )
        self.txt_note = tk.Text(m, width=52, height=4, wrap="word")
        self.txt_note.grid(row=3, column=1, sticky="ew", pady=5)
        self.txt_note.insert("1.0", self.init.get("note", "") or "")

        info = self.init.get("_info")
        if info:
            tk.Label(
                m,
                text=info,
                fg=C["text3"],
                justify="left",
                anchor="w",
            ).grid(row=4, column=0, columnspan=2, sticky="w", pady=(8, 0))

        return self.ent_name

    def validate(self):
        name = self.var_name.get().strip()
        if not name:
            messagebox.showwarning("Шаблон", "Введите наименование шаблона.", parent=self)
            return False

        self.result = {
            "name": name,
            "category": self.var_category.get().strip(),
            "description": self.txt_desc.get("1.0", "end").strip(),
            "note": self.txt_note.get("1.0", "end").strip(),
        }
        return True

class _TemplateTaskDialog(simpledialog.Dialog):
    """Диалог добавления/редактирования строки шаблона."""

    def __init__(self, parent, work_types, uoms, parents=None, init=None):
        self.work_types = work_types or []
        self.uoms = uoms or []
        self.parents = parents or []
        self.init = init or {}
        self.result = None
        super().__init__(parent, title="Строка шаблона")

    def body(self, master):
        master.grid_columnconfigure(1, weight=1)

        self.var_row_kind = tk.StringVar(value=self.init.get("row_kind", "task"))
        self.var_name = tk.StringVar(value=self.init.get("name", ""))
        self.var_qty = tk.StringVar(value=_fmt_qty(self.init.get("default_qty")))
        self.var_milestone = tk.BooleanVar(value=bool(self.init.get("is_milestone", False)))
        self.var_sort = tk.StringVar(value=str(self.init.get("sort_order", 10)))

        tk.Label(master, text="Тип строки *:").grid(
            row=0, column=0, sticky="e", padx=6, pady=4
        )
        self.cmb_row_kind = ttk.Combobox(
            master,
            state="readonly",
            width=20,
            values=[ROW_KIND_LABELS[k] for k in ROW_KIND_LIST],
        )
        self.cmb_row_kind.grid(row=0, column=1, sticky="w", padx=6, pady=4)

        tk.Label(master, text="Тип работ *:").grid(
            row=1, column=0, sticky="e", padx=6, pady=4
        )
        self.cmb_wt = ttk.Combobox(
            master,
            state="readonly",
            width=42,
            values=[w["name"] for w in self.work_types],
        )
        self.cmb_wt.grid(row=1, column=1, sticky="w", padx=6, pady=4)

        tk.Label(master, text="Наименование *:").grid(
            row=2, column=0, sticky="e", padx=6, pady=4
        )
        self.ent_name = ttk.Entry(master, textvariable=self.var_name, width=44)
        self.ent_name.grid(row=2, column=1, sticky="ew", padx=6, pady=4)

        tk.Label(master, text="Ед. изм.:").grid(
            row=3, column=0, sticky="e", padx=6, pady=4
        )
        self.cmb_uom = ttk.Combobox(
            master,
            state="readonly",
            width=42,
            values=["—"] + [f"{u['code']} — {u['name']}" for u in self.uoms],
        )
        self.cmb_uom.grid(row=3, column=1, sticky="w", padx=6, pady=4)

        tk.Label(master, text="Объём по умолчанию:").grid(
            row=4, column=0, sticky="e", padx=6, pady=4
        )
        self.ent_qty = ttk.Entry(master, textvariable=self.var_qty, width=18)
        self.ent_qty.grid(row=4, column=1, sticky="w", padx=6, pady=4)

        tk.Label(master, text="Родительская строка:").grid(
            row=5, column=0, sticky="e", padx=6, pady=4
        )
        self.cmb_parent = ttk.Combobox(master, state="readonly", width=42)
        parent_values = ["— Нет —"] + [p["name"] for p in self.parents]
        self.cmb_parent["values"] = parent_values
        self.cmb_parent.grid(row=5, column=1, sticky="w", padx=6, pady=4)

        tk.Label(master, text="Порядок:").grid(
            row=6, column=0, sticky="e", padx=6, pady=4
        )
        self.ent_sort = ttk.Entry(master, textvariable=self.var_sort, width=10)
        self.ent_sort.grid(row=6, column=1, sticky="w", padx=6, pady=4)

        self.chk_milestone = ttk.Checkbutton(
            master, text="Веха", variable=self.var_milestone
        )
        self.chk_milestone.grid(row=7, column=1, sticky="w", padx=6, pady=6)

        row_kind = self.init.get("row_kind", "task")
        try:
            self.cmb_row_kind.current(ROW_KIND_LIST.index(row_kind))
        except ValueError:
            self.cmb_row_kind.current(0)

        wt_id = self.init.get("work_type_id")
        if wt_id:
            for i, w in enumerate(self.work_types):
                if int(w["id"]) == int(wt_id):
                    self.cmb_wt.current(i)
                    break
        elif self.work_types:
            self.cmb_wt.current(0)

        uom_code = self.init.get("uom_code")
        if uom_code:
            found = False
            for i, u in enumerate(self.uoms, start=1):
                if u["code"] == uom_code:
                    self.cmb_uom.current(i)
                    found = True
                    break
            if not found:
                self.cmb_uom.current(0)
        else:
            self.cmb_uom.current(0)

        parent_id = self.init.get("parent_id")
        if parent_id:
            found = False
            for i, p in enumerate(self.parents, start=1):
                if int(p["id"]) == int(parent_id):
                    self.cmb_parent.current(i)
                    found = True
                    break
            if not found:
                self.cmb_parent.current(0)
        else:
            self.cmb_parent.current(0)

        self.cmb_row_kind.bind("<<ComboboxSelected>>", lambda _e: self._apply_row_kind_ui())
        self._apply_row_kind_ui()

        return self.ent_name

    def _apply_row_kind_ui(self):
        idx = self.cmb_row_kind.current()
        row_kind = ROW_KIND_LIST[idx] if 0 <= idx < len(ROW_KIND_LIST) else "task"

        is_task = (row_kind == "task")

        wt_state = "readonly"
        uom_state = "readonly" if is_task else "disabled"
        qty_state = "normal" if is_task else "disabled"
        parent_state = "readonly"
        milestone_state = "normal" if is_task else "disabled"

        self.cmb_wt.configure(state=wt_state)
        self.cmb_uom.configure(state=uom_state)
        self.ent_qty.configure(state=qty_state)

        self.cmb_parent.configure(state=parent_state)

        if is_task:
            self.chk_milestone.state(["!disabled"])
        else:
            self.chk_milestone.state(["disabled"])
            self.var_milestone.set(False)
            self.cmb_uom.current(0)
            self.var_qty.set("")

    def validate(self):
        rk_idx = self.cmb_row_kind.current()
        row_kind = ROW_KIND_LIST[rk_idx] if 0 <= rk_idx < len(ROW_KIND_LIST) else "task"

        wi = self.cmb_wt.current()
        if wi < 0:
            messagebox.showwarning("Шаблон", "Выберите тип работ.", parent=self)
            return False

        name = self.var_name.get().strip()
        if not name:
            messagebox.showwarning("Шаблон", "Введите наименование строки.", parent=self)
            return False

        try:
            sort_order = int(self.var_sort.get().strip() or "0")
        except ValueError:
            messagebox.showwarning("Шаблон", "Порядок должен быть целым числом.", parent=self)
            return False

        uom_code = None
        qty = None
        is_milestone = False

        if row_kind == "task":
            qty = _safe_float(self.var_qty.get())
            if self.var_qty.get().strip() and qty is None:
                messagebox.showwarning("Шаблон", "Объём должен быть числом.", parent=self)
                return False

            ui = self.cmb_uom.current()
            if ui > 0:
                uom_code = self.uoms[ui - 1]["code"]

            is_milestone = bool(self.var_milestone.get())

        parent_id = None
        pi = self.cmb_parent.current()
        if pi > 0:
            parent_id = self.parents[pi - 1]["id"]

        self.result = {
            "row_kind": row_kind,
            "work_type_id": self.work_types[wi]["id"],
            "name": name,
            "uom_code": uom_code,
            "default_qty": qty,
            "is_milestone": is_milestone,
            "parent_id": parent_id,
            "sort_order": sort_order,
        }
        return True

class _WorkItemDialog(simpledialog.Dialog):
    """Создание и редактирование конкретной работы."""

    def __init__(
        self,
        parent,
        work_types: List[Dict[str, Any]],
        uoms: List[Dict[str, Any]],
        init: Optional[Dict[str, Any]] = None,
    ):
        self.work_types = work_types or []
        self.uoms = uoms or []
        self.init = init or {}
        self.result: Optional[Dict[str, Any]] = None
        super().__init__(parent, title="Карточка работы")

    def body(self, master):
        master.grid_columnconfigure(1, weight=1)

        self.var_code = tk.StringVar(value=self.init.get("code", "") or "")
        self.var_name = tk.StringVar(value=self.init.get("name", "") or "")
        self.var_sort = tk.StringVar(
            value=str(self.init.get("sort_order", 100))
        )

        tk.Label(master, text="Тип работ *:").grid(
            row=0, column=0, sticky="e", padx=(0, 8), pady=5
        )
        self.cmb_work_type = ttk.Combobox(
            master,
            state="readonly",
            width=48,
            values=[
                f"{w.get('code') or '—'} — {w['name']}"
                for w in self.work_types
            ],
        )
        self.cmb_work_type.grid(
            row=0, column=1, sticky="ew", pady=5
        )

        tk.Label(master, text="Код работы:").grid(
            row=1, column=0, sticky="e", padx=(0, 8), pady=5
        )
        self.ent_code = ttk.Entry(
            master,
            textvariable=self.var_code,
            width=50,
        )
        self.ent_code.grid(row=1, column=1, sticky="ew", pady=5)

        tk.Label(master, text="Наименование *:").grid(
            row=2, column=0, sticky="e", padx=(0, 8), pady=5
        )
        self.ent_name = ttk.Entry(
            master,
            textvariable=self.var_name,
            width=50,
        )
        self.ent_name.grid(row=2, column=1, sticky="ew", pady=5)

        tk.Label(master, text="Ед. измерения *:").grid(
            row=3, column=0, sticky="e", padx=(0, 8), pady=5
        )
        self.cmb_uom = ttk.Combobox(
            master,
            state="readonly",
            width=48,
            values=[
                f"{u['code']} — {u['name']}"
                for u in self.uoms
            ],
        )
        self.cmb_uom.grid(row=3, column=1, sticky="ew", pady=5)

        tk.Label(master, text="Порядок:").grid(
            row=4, column=0, sticky="e", padx=(0, 8), pady=5
        )
        self.ent_sort = ttk.Entry(
            master,
            textvariable=self.var_sort,
            width=12,
        )
        self.ent_sort.grid(row=4, column=1, sticky="w", pady=5)

        tk.Label(master, text="Примечание:").grid(
            row=5, column=0, sticky="ne", padx=(0, 8), pady=5
        )
        self.txt_note = tk.Text(master, height=4, width=50, wrap="word")
        self.txt_note.grid(row=5, column=1, sticky="ew", pady=5)
        self.txt_note.insert("1.0", self.init.get("note", "") or "")

        work_type_id = self.init.get("work_type_id")
        for i, row in enumerate(self.work_types):
            if work_type_id is not None and int(row["id"]) == int(work_type_id):
                self.cmb_work_type.current(i)
                break

        uom_code = self.init.get("uom_code")
        for i, row in enumerate(self.uoms):
            if uom_code and row["code"] == uom_code:
                self.cmb_uom.current(i)
                break

        return self.cmb_work_type

    def validate(self):
        wt_idx = self.cmb_work_type.current()
        if wt_idx < 0:
            messagebox.showwarning(
                "Работа",
                "Выберите тип работ.",
                parent=self,
            )
            return False

        name = self.var_name.get().strip()
        if not name:
            messagebox.showwarning(
                "Работа",
                "Введите наименование работы.",
                parent=self,
            )
            return False

        uom_idx = self.cmb_uom.current()
        if uom_idx < 0:
            messagebox.showwarning(
                "Работа",
                "Выберите единицу измерения.",
                parent=self,
            )
            return False

        try:
            sort_order = int(self.var_sort.get().strip() or "0")
        except ValueError:
            messagebox.showwarning(
                "Работа",
                "Порядок сортировки должен быть целым числом.",
                parent=self,
            )
            return False

        self.result = {
            "work_type_id": int(self.work_types[wt_idx]["id"]),
            "code": self.var_code.get().strip(),
            "name": name,
            "uom_code": self.uoms[uom_idx]["code"],
            "sort_order": sort_order,
            "note": self.txt_note.get("1.0", "end").strip(),
        }
        return True


class _LaborNormDialog(simpledialog.Dialog):
    """Создание и редактирование версии норматива ЗТР."""

    def __init__(
        self,
        parent,
        work_item: Dict[str, Any],
        init: Optional[Dict[str, Any]] = None,
    ):
        self.work_item = work_item
        self.init = init or {}
        self.result: Optional[Dict[str, Any]] = None
        super().__init__(parent, title="Норматив ЗТР")

    def body(self, master):
        master.grid_columnconfigure(1, weight=1)

        header = (
            f"Работа: {self.work_item.get('name')}\n"
            f"Тип: {self.work_item.get('work_type_name') or '—'}\n"
            f"Ед. изм.: {self.work_item.get('uom_code') or '—'}"
        )
        tk.Label(
            master,
            text=header,
            foreground=C["text2"],
            justify="left",
            anchor="w",
        ).grid(
            row=0,
            column=0,
            columnspan=2,
            sticky="ew",
            pady=(0, 10),
        )

        self.var_from = tk.StringVar(
            value=_fmt_date_iso(
                self.init.get("effective_from") or _today()
            )
        )
        self.var_to = tk.StringVar(
            value=_fmt_date_iso(self.init.get("effective_to"))
        )
        self.var_labor = tk.StringVar(
            value=_fmt_qty(self.init.get("labor_hours_per_unit"))
        )
        self.var_factor = tk.StringVar(
            value=_fmt_qty(
                self.init.get("default_productivity_factor", 1)
            )
        )
        self.var_source_name = tk.StringVar(
            value=self.init.get("source_name", "") or ""
        )
        self.var_source_code = tk.StringVar(
            value=self.init.get("source_code", "") or ""
        )
        self.var_source_version = tk.StringVar(
            value=self.init.get("source_version", "") or ""
        )

        fields = [
            ("Действует с *:", self.var_from, 16),
            ("Действует по:", self.var_to, 16),
            ("ЗТР, чел.-ч/ед. *:", self.var_labor, 16),
            ("Коэффициент *:", self.var_factor, 16),
            ("Источник:", self.var_source_name, 42),
            ("Код источника:", self.var_source_code, 42),
            ("Редакция / версия:", self.var_source_version, 42),
        ]

        for i, (label, variable, width) in enumerate(fields, start=1):
            tk.Label(master, text=label).grid(
                row=i,
                column=0,
                sticky="e",
                padx=(0, 8),
                pady=4,
            )
            ttk.Entry(
                master,
                textvariable=variable,
                width=width,
            ).grid(
                row=i,
                column=1,
                sticky="ew",
                pady=4,
            )

        tk.Label(master, text="Примечание:").grid(
            row=8,
            column=0,
            sticky="ne",
            padx=(0, 8),
            pady=4,
        )
        self.txt_note = tk.Text(master, height=4, width=44, wrap="word")
        self.txt_note.grid(row=8, column=1, sticky="ew", pady=4)
        self.txt_note.insert("1.0", self.init.get("note", "") or "")

        tk.Label(
            master,
            text=(
                "Формат дат: ГГГГ-ММ-ДД или ДД.ММ.ГГГГ.\n"
                "Пустая дата «Действует по» означает бессрочную норму."
            ),
            fg=C["text3"],
            justify="left",
            anchor="w",
        ).grid(
            row=9,
            column=0,
            columnspan=2,
            sticky="w",
            pady=(8, 0),
        )

        return master.nametowidget(master.winfo_children()[2])

    def validate(self):
        try:
            effective_from = _parse_date_iso(
                self.var_from.get(),
                "Дата начала действия",
            )

            effective_to = None
            if self.var_to.get().strip():
                effective_to = _parse_date_iso(
                    self.var_to.get(),
                    "Дата окончания действия",
                )

            if effective_to and effective_to < effective_from:
                raise ValueError(
                    "Дата окончания действия не может быть раньше даты начала."
                )

            labor_hours_per_unit = _safe_float(self.var_labor.get())
            if labor_hours_per_unit is None or labor_hours_per_unit <= 0:
                raise ValueError(
                    "Норматив ЗТР должен быть числом больше 0."
                )

            factor = _safe_float(self.var_factor.get())
            if factor is None or factor <= 0:
                raise ValueError(
                    "Коэффициент должен быть числом больше 0."
                )

            self.result = {
                "effective_from": effective_from,
                "effective_to": effective_to,
                "labor_hours_per_unit": labor_hours_per_unit,
                "default_productivity_factor": factor,
                "source_name": self.var_source_name.get().strip(),
                "source_code": self.var_source_code.get().strip(),
                "source_version": self.var_source_version.get().strip(),
                "note": self.txt_note.get("1.0", "end").strip(),
            }
            return True

        except ValueError as exc:
            messagebox.showwarning(
                "Норматив ЗТР",
                str(exc),
                parent=self,
            )
            return False

class _LaborExcelImportDialog(simpledialog.Dialog):
    """Параметры импорта норм ЗТР из Excel."""

    def __init__(self, parent, file_path: str):
        self.file_path = file_path
        self.result: Optional[Dict[str, Any]] = None
        super().__init__(parent, title="Импорт норм ЗТР из Excel")

    def body(self, master):
        master.grid_columnconfigure(1, weight=1)

        file_name = os.path.basename(self.file_path)

        tk.Label(
            master,
            text=(
                f"Файл: {file_name}\n\n"
                "Все строки будут импортированы как версии норм ЗТР.\n"
                "Если повторно загрузить тот же файл с той же датой,\n"
                "существующие нормы будут обновлены без дублей."
            ),
            justify="left",
            anchor="w",
            fg=C["text2"],
        ).grid(
            row=0,
            column=0,
            columnspan=2,
            sticky="ew",
            pady=(0, 12),
        )

        self.var_effective_from = tk.StringVar(value=_today().isoformat())
        self.var_source_name = tk.StringVar(
            value=f"Импорт Excel: {file_name}"
        )

        tk.Label(master, text="Действует с *:").grid(
            row=1,
            column=0,
            sticky="e",
            padx=(0, 8),
            pady=5,
        )

        self.ent_effective_from = ttk.Entry(
            master,
            textvariable=self.var_effective_from,
            width=26,
        )
        self.ent_effective_from.grid(
            row=1,
            column=1,
            sticky="ew",
            pady=5,
        )

        tk.Label(master, text="Источник нормы:").grid(
            row=2,
            column=0,
            sticky="e",
            padx=(0, 8),
            pady=5,
        )

        ttk.Entry(
            master,
            textvariable=self.var_source_name,
            width=50,
        ).grid(
            row=2,
            column=1,
            sticky="ew",
            pady=5,
        )

        tk.Label(
            master,
            text=(
                "Дата вводится в формате ГГГГ-ММ-ДД или ДД.ММ.ГГГГ.\n"
                "Коэффициент производительности при импорте: 1.000000."
            ),
            fg=C["text3"],
            justify="left",
            anchor="w",
        ).grid(
            row=3,
            column=0,
            columnspan=2,
            sticky="w",
            pady=(8, 0),
        )

        return self.ent_effective_from

    def validate(self):
        try:
            effective_from = _parse_date_iso(
                self.var_effective_from.get(),
                "Дата начала действия",
            )
        except ValueError as exc:
            messagebox.showwarning(
                "Импорт Excel",
                str(exc),
                parent=self,
            )
            return False

        self.result = {
            "effective_from": effective_from,
            "source_name": (
                self.var_source_name.get().strip()
                or f"Импорт Excel: {os.path.basename(self.file_path)}"
            ),
        }
        return True

# ═══════════════════════════════════════════════════════════════
#  PAGE
# ═══════════════════════════════════════════════════════════════
class GprDictionariesPage(tk.Frame):
    """
    Управление справочниками ГПР:
      - Типы работ  (gpr_work_types)
      - Ед. измерения (gpr_uom)
      - Шаблоны      (gpr_templates + gpr_template_tasks)
    """

    def __init__(self, master, app_ref=None):
        super().__init__(master, bg=C["bg"])
        self.app_ref = app_ref
        self._build_ui()

    # ══════════════════════════════════════════════════════
    #  UI
    # ══════════════════════════════════════════════════════
    def _build_ui(self):
        hdr = tk.Frame(self, bg=C["accent"], pady=6)
        hdr.pack(fill="x")
        tk.Label(
            hdr,
            text="⚙  Справочники ГПР",
            font=("Segoe UI", 12, "bold"),
            bg=C["accent"],
            fg="white",
            padx=12,
        ).pack(side="left")

        self.nb = ttk.Notebook(self)
        self.nb.pack(fill="both", expand=True, padx=10, pady=10)

        tab_wt = tk.Frame(self.nb, bg=C["panel"])
        self.nb.add(tab_wt, text="  Типы работ  ")
        self._build_wt_tab(tab_wt)

        tab_uom = tk.Frame(self.nb, bg=C["panel"])
        self.nb.add(tab_uom, text="  Единицы измерения  ")
        self._build_uom_tab(tab_uom)

        tab_labor = tk.Frame(self.nb, bg=C["panel"])
        self.nb.add(tab_labor, text="  Работы и нормы ЗТР  ")
        self._build_labor_tab(tab_labor)

        tab_tpl = tk.Frame(self.nb, bg=C["panel"])
        self.nb.add(tab_tpl, text="  Шаблоны  ")
        self._build_tpl_tab(tab_tpl)

    # ══════════════════════════════════════════════════════
    #  ТИПЫ РАБОТ
    # ══════════════════════════════════════════════════════
    def _build_wt_tab(self, parent):
        bar = tk.Frame(parent, bg=C["panel"])
        bar.pack(fill="x", padx=8, pady=6)
        ttk.Button(bar, text="➕ Добавить", command=self._wt_add).pack(side="left", padx=2)
        ttk.Button(bar, text="✏️ Редактировать", command=self._wt_edit).pack(side="left", padx=2)
        ttk.Button(bar, text="🔄 Вкл/Выкл", command=self._wt_toggle).pack(side="left", padx=2)
        ttk.Button(bar, text="🔃 Обновить", command=self._wt_load).pack(side="left", padx=8)

        cols = ("id", "code", "name", "sort_order", "is_active")
        self.wt_tree = ttk.Treeview(
            parent, columns=cols, show="headings",
            selectmode="browse", height=18
        )
        for c, t, w, a in [
            ("id", "ID", 50, "center"),
            ("code", "Код", 80, "w"),
            ("name", "Наименование", 300, "w"),
            ("sort_order", "Сортировка", 80, "center"),
            ("is_active", "Активен", 70, "center"),
        ]:
            self.wt_tree.heading(c, text=t)
            self.wt_tree.column(c, width=w, anchor=a)

        vsb = ttk.Scrollbar(parent, orient="vertical", command=self.wt_tree.yview)
        self.wt_tree.configure(yscrollcommand=vsb.set)
        self.wt_tree.pack(side="left", fill="both", expand=True, padx=(8, 0), pady=(0, 8))
        vsb.pack(side="right", fill="y", padx=(0, 8), pady=(0, 8))

        self.wt_tree.tag_configure("inactive", foreground="#bbb")
        self.wt_tree.bind("<Double-1>", lambda _e: self._wt_edit())

        self._wt_data: List[Dict] = []
        self._wt_load()

    def _wt_load(self):
        self.wt_tree.delete(*self.wt_tree.get_children())
        self._wt_data.clear()
        conn = None
        try:
            conn = _conn()
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("""
                    SELECT id, COALESCE(code,'') AS code, name, sort_order, is_active
                    FROM public.gpr_work_types
                    ORDER BY sort_order, name
                """)
                self._wt_data = [dict(r) for r in cur.fetchall()]
        except Exception as e:
            messagebox.showerror("Справочники", f"Ошибка:\n{e}", parent=self)
            return
        finally:
            _release(conn)

        for w in self._wt_data:
            active = "Да" if w["is_active"] else "Нет"
            tags = () if w["is_active"] else ("inactive",)
            self.wt_tree.insert(
                "", "end",
                values=(w["id"], w["code"], w["name"], w["sort_order"], active),
                tags=tags
            )

    def _wt_sel(self) -> Optional[Dict]:
        sel = self.wt_tree.selection()
        if not sel:
            return None
        idx = self.wt_tree.index(sel[0])
        return self._wt_data[idx] if 0 <= idx < len(self._wt_data) else None

    def _wt_add(self):
        fields = [
            ("Код", "code", 12),
            ("Наименование", "name", 40),
            ("Порядок сортировки", "sort_order", 8),
        ]
        dlg = _DictEditDialog(self, "Новый тип работ", fields, init={"sort_order": "100"})
        if not dlg.result:
            return
        r = dlg.result
        if not r.get("name"):
            messagebox.showwarning("Справочники", "Наименование обязательно.", parent=self)
            return
        conn = None
        try:
            conn = _conn()
            with conn, conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO public.gpr_work_types(code, name, sort_order, is_active)
                    VALUES (%s, %s, %s, true)
                """, (r["code"] or None, r["name"], int(r["sort_order"] or 100)))
        except Exception as e:
            messagebox.showerror("Справочники", f"Ошибка:\n{e}", parent=self)
            return
        finally:
            _release(conn)
        self._wt_load()

    def _wt_edit(self):
        w = self._wt_sel()
        if not w:
            return
        fields = [
            ("Код", "code", 12),
            ("Наименование", "name", 40),
            ("Порядок сортировки", "sort_order", 8),
        ]
        dlg = _DictEditDialog(
            self,
            "Редактирование типа работ",
            fields,
            init={
                "code": w["code"],
                "name": w["name"],
                "sort_order": str(w["sort_order"]),
            }
        )
        if not dlg.result:
            return
        r = dlg.result
        conn = None
        try:
            conn = _conn()
            with conn, conn.cursor() as cur:
                cur.execute("""
                    UPDATE public.gpr_work_types
                    SET code=%s, name=%s, sort_order=%s
                    WHERE id=%s
                """, (r["code"] or None, r["name"], int(r["sort_order"] or 0), w["id"]))
        except Exception as e:
            messagebox.showerror("Справочники", f"Ошибка:\n{e}", parent=self)
            return
        finally:
            _release(conn)
        self._wt_load()

    def _wt_toggle(self):
        w = self._wt_sel()
        if not w:
            return
        conn = None
        try:
            conn = _conn()
            with conn, conn.cursor() as cur:
                cur.execute(
                    "UPDATE public.gpr_work_types SET is_active = NOT is_active WHERE id=%s",
                    (w["id"],)
                )
        except Exception as e:
            messagebox.showerror("Справочники", f"Ошибка:\n{e}", parent=self)
            return
        finally:
            _release(conn)
        self._wt_load()

    # ══════════════════════════════════════════════════════
    #  ЕДИНИЦЫ ИЗМЕРЕНИЯ
    # ══════════════════════════════════════════════════════
    def _build_uom_tab(self, parent):
        bar = tk.Frame(parent, bg=C["panel"])
        bar.pack(fill="x", padx=8, pady=6)
        ttk.Button(bar, text="➕ Добавить", command=self._uom_add).pack(side="left", padx=2)
        ttk.Button(bar, text="✏️ Редактировать", command=self._uom_edit).pack(side="left", padx=2)
        ttk.Button(bar, text="🗑 Удалить", command=self._uom_del).pack(side="left", padx=2)
        ttk.Button(bar, text="🔃 Обновить", command=self._uom_load).pack(side="left", padx=8)

        cols = ("code", "name")
        self.uom_tree = ttk.Treeview(
            parent, columns=cols, show="headings",
            selectmode="browse", height=18
        )
        self.uom_tree.heading("code", text="Код")
        self.uom_tree.heading("name", text="Наименование")
        self.uom_tree.column("code", width=100, anchor="w")
        self.uom_tree.column("name", width=300, anchor="w")

        vsb = ttk.Scrollbar(parent, orient="vertical", command=self.uom_tree.yview)
        self.uom_tree.configure(yscrollcommand=vsb.set)
        self.uom_tree.pack(side="left", fill="both", expand=True, padx=(8, 0), pady=(0, 8))
        vsb.pack(side="right", fill="y", padx=(0, 8), pady=(0, 8))
        self.uom_tree.bind("<Double-1>", lambda _e: self._uom_edit())

        self._uom_data: List[Dict] = []
        self._uom_load()

    def _uom_load(self):
        self.uom_tree.delete(*self.uom_tree.get_children())
        self._uom_data.clear()
        conn = None
        try:
            conn = _conn()
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("SELECT code, name FROM public.gpr_uom ORDER BY code")
                self._uom_data = [dict(r) for r in cur.fetchall()]
        except Exception as e:
            messagebox.showerror("Справочники", f"Ошибка:\n{e}", parent=self)
            return
        finally:
            _release(conn)

        for u in self._uom_data:
            self.uom_tree.insert("", "end", values=(u["code"], u["name"]))

    def _uom_sel(self) -> Optional[Dict]:
        sel = self.uom_tree.selection()
        if not sel:
            return None
        idx = self.uom_tree.index(sel[0])
        return self._uom_data[idx] if 0 <= idx < len(self._uom_data) else None

    def _uom_add(self):
        fields = [("Код", "code", 12), ("Наименование", "name", 30)]
        dlg = _DictEditDialog(self, "Новая единица измерения", fields)
        if not dlg.result:
            return
        r = dlg.result
        if not r["code"]:
            messagebox.showwarning("Справочники", "Код обязателен.", parent=self)
            return
        conn = None
        try:
            conn = _conn()
            with conn, conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO public.gpr_uom(code, name) VALUES(%s,%s)",
                    (r["code"], r["name"])
                )
        except Exception as e:
            messagebox.showerror("Справочники", f"Ошибка:\n{e}", parent=self)
            return
        finally:
            _release(conn)
        self._uom_load()

    def _uom_edit(self):
        u = self._uom_sel()
        if not u:
            return
        fields = [("Код", "code", 12), ("Наименование", "name", 30)]
        dlg = _DictEditDialog(
            self,
            "Редактирование ед. изм.",
            fields,
            init={"code": u["code"], "name": u["name"]},
            readonly_keys=["code"],
        )
        if not dlg.result:
            return
        r = dlg.result
        conn = None
        try:
            conn = _conn()
            with conn, conn.cursor() as cur:
                cur.execute(
                    "UPDATE public.gpr_uom SET name=%s WHERE code=%s",
                    (r["name"], u["code"])
                )
        except Exception as e:
            messagebox.showerror("Справочники", f"Ошибка:\n{e}", parent=self)
            return
        finally:
            _release(conn)
        self._uom_load()

    def _uom_del(self):
        u = self._uom_sel()
        if not u:
            return
        if not messagebox.askyesno(
            "Справочники",
            f"Удалить единицу '{u['code']}'?\n(Не получится, если она используется в задачах)",
            parent=self
        ):
            return
        conn = None
        try:
            conn = _conn()
            with conn, conn.cursor() as cur:
                cur.execute("DELETE FROM public.gpr_uom WHERE code=%s", (u["code"],))
        except Exception as e:
            messagebox.showerror(
                "Справочники",
                f"Ошибка (возможно, используется):\n{e}",
                parent=self
            )
            return
        finally:
            _release(conn)
        self._uom_load()

    # ══════════════════════════════════════════════════════
    #  РАБОТЫ И НОРМЫ ЗТР
    # ══════════════════════════════════════════════════════
    def _build_labor_tab(self, parent):
        self._labor_items_data: List[Dict[str, Any]] = []
        self._labor_norms_data: List[Dict[str, Any]] = []
        self._labor_wt_cache: List[Dict[str, Any]] = []
        self._labor_uom_cache: List[Dict[str, Any]] = []

        search_bar = tk.Frame(parent, bg=C["panel"])
        search_bar.pack(fill="x", padx=8, pady=(8, 4))

        tk.Label(
            search_bar,
            text="Поиск работы:",
            bg=C["panel"],
        ).pack(side="left")

        self.var_labor_search = tk.StringVar()
        ent_search = ttk.Entry(
            search_bar,
            textvariable=self.var_labor_search,
            width=36,
        )
        ent_search.pack(side="left", padx=(6, 8))
        ent_search.bind("<KeyRelease>", lambda _e: self._labor_load_items())

        ttk.Button(
            search_bar,
            text="🔃 Обновить",
            command=self._labor_load_items,
        ).pack(side="left", padx=2)

        self.lbl_labor_summary = tk.Label(
            search_bar,
            text="",
            bg=C["panel"],
            fg=C["text2"],
            font=("Segoe UI", 8),
        )
        self.lbl_labor_summary.pack(side="right", padx=8)

        pw = tk.PanedWindow(
            parent,
            orient="horizontal",
            sashrelief="raised",
            bg=C["bg"],
        )
        pw.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        left = tk.Frame(pw, bg=C["panel"])
        right = tk.Frame(pw, bg=C["panel"])

        pw.add(left, minsize=560)
        pw.add(right, minsize=520)

        # ── Левая панель: работы ──
        work_box = tk.LabelFrame(
            left,
            text=" Конкретные работы ",
            bg=C["panel"],
            padx=8,
            pady=6,
        )
        work_box.pack(fill="both", expand=True)

        bar_work = tk.Frame(work_box, bg=C["panel"])
        bar_work.pack(fill="x")

        ttk.Button(
            bar_work,
            text="➕ Добавить работу",
            command=self._labor_add_item,
        ).pack(side="left", padx=2)

        ttk.Button(
            bar_work,
            text="📥 Импорт Excel",
            command=self._labor_import_excel,
        ).pack(side="left", padx=2)

        ttk.Button(
            bar_work,
            text="✏️ Редактировать",
            command=self._labor_edit_item,
        ).pack(side="left", padx=2)

        ttk.Button(
            bar_work,
            text="🔄 Вкл/Выкл",
            command=self._labor_toggle_item,
        ).pack(side="left", padx=2)

        cols = (
            "type",
            "code",
            "name",
            "uom",
            "norm",
            "from",
            "active",
        )
        self.labor_item_tree = ttk.Treeview(
            work_box,
            columns=cols,
            show="headings",
            selectmode="browse",
            height=20,
        )

        for col, title, width, anchor in [
            ("type", "Тип работ", 160, "w"),
            ("code", "Код", 90, "w"),
            ("name", "Наименование работы", 250, "w"),
            ("uom", "Ед.", 60, "center"),
            ("norm", "ЗТР", 80, "e"),
            ("from", "Действует с", 95, "center"),
            ("active", "Статус", 75, "center"),
        ]:
            self.labor_item_tree.heading(col, text=title)
            self.labor_item_tree.column(col, width=width, anchor=anchor)

        vsb_items = ttk.Scrollbar(
            work_box,
            orient="vertical",
            command=self.labor_item_tree.yview,
        )
        self.labor_item_tree.configure(yscrollcommand=vsb_items.set)

        self.labor_item_tree.pack(
            side="left",
            fill="both",
            expand=True,
            pady=(6, 0),
        )
        vsb_items.pack(side="right", fill="y", pady=(6, 0))

        self.labor_item_tree.tag_configure(
            "inactive",
            foreground="#aaa",
        )
        self.labor_item_tree.tag_configure(
            "without_norm",
            background="#fff3e0",
        )
        self.labor_item_tree.bind(
            "<<TreeviewSelect>>",
            lambda _e: self._labor_load_norms(),
        )
        self.labor_item_tree.bind(
            "<Double-1>",
            lambda _e: self._labor_edit_item(),
        )

        # ── Правая панель: нормы ──
        norm_box = tk.LabelFrame(
            right,
            text=" Нормативы ЗТР выбранной работы ",
            bg=C["panel"],
            padx=8,
            pady=6,
        )
        norm_box.pack(fill="both", expand=True)

        self.lbl_labor_item_info = tk.Label(
            norm_box,
            text="Выберите работу в списке слева.",
            bg=C["panel"],
            fg=C["text2"],
            justify="left",
            anchor="w",
        )
        self.lbl_labor_item_info.pack(fill="x", pady=(0, 6))

        bar_norm = tk.Frame(norm_box, bg=C["panel"])
        bar_norm.pack(fill="x")

        ttk.Button(
            bar_norm,
            text="➕ Новая норма",
            command=self._labor_add_norm,
        ).pack(side="left", padx=2)

        ttk.Button(
            bar_norm,
            text="✏️ Редактировать",
            command=self._labor_edit_norm,
        ).pack(side="left", padx=2)

        ttk.Button(
            bar_norm,
            text="🔄 Вкл/Выкл",
            command=self._labor_toggle_norm,
        ).pack(side="left", padx=2)

        cols_norm = (
            "from",
            "to",
            "labor",
            "factor",
            "source",
            "status",
        )
        self.labor_norm_tree = ttk.Treeview(
            norm_box,
            columns=cols_norm,
            show="headings",
            selectmode="browse",
            height=20,
        )

        for col, title, width, anchor in [
            ("from", "Действует с", 100, "center"),
            ("to", "Действует по", 100, "center"),
            ("labor", "ЗТР, чел.-ч/ед.", 110, "e"),
            ("factor", "Коэфф.", 75, "e"),
            ("source", "Источник", 190, "w"),
            ("status", "Статус", 75, "center"),
        ]:
            self.labor_norm_tree.heading(col, text=title)
            self.labor_norm_tree.column(col, width=width, anchor=anchor)

        vsb_norms = ttk.Scrollbar(
            norm_box,
            orient="vertical",
            command=self.labor_norm_tree.yview,
        )
        self.labor_norm_tree.configure(yscrollcommand=vsb_norms.set)

        self.labor_norm_tree.pack(
            side="left",
            fill="both",
            expand=True,
            pady=(6, 0),
        )
        vsb_norms.pack(side="right", fill="y", pady=(6, 0))

        self.labor_norm_tree.tag_configure(
            "inactive",
            foreground="#aaa",
        )
        self.labor_norm_tree.tag_configure(
            "current",
            background="#e8f5e9",
        )
        self.labor_norm_tree.bind(
            "<Double-1>",
            lambda _e: self._labor_edit_norm(),
        )

        self._labor_load_caches()
        self._labor_load_items()

    def _labor_load_caches(self):
        conn = None
        try:
            conn = _conn()
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    """
                    SELECT id, COALESCE(code, '') AS code, name, sort_order
                    FROM public.gpr_work_types
                    WHERE is_active = true
                    ORDER BY sort_order, name
                    """
                )
                self._labor_wt_cache = [
                    dict(row) for row in cur.fetchall()
                ]

                cur.execute(
                    """
                    SELECT code, name
                    FROM public.gpr_uom
                    ORDER BY code
                    """
                )
                self._labor_uom_cache = [
                    dict(row) for row in cur.fetchall()
                ]
        finally:
            _release(conn)

    def _labor_load_items(self):
        selected_id = None
        current = self._labor_item_sel()
        if current:
            selected_id = int(current["id"])

        try:
            self._labor_items_data = GprLaborDictionaryService.load_work_items(
                self.var_labor_search.get()
                if hasattr(self, "var_labor_search")
                else ""
            )
        except Exception as exc:
            messagebox.showerror(
                "Нормы ЗТР",
                f"Ошибка загрузки работ:\n{exc}",
                parent=self,
            )
            return

        self.labor_item_tree.delete(
            *self.labor_item_tree.get_children()
        )

        selected_iid = None
        no_norm_count = 0

        for item in self._labor_items_data:
            norm = item.get("current_labor_hours_per_unit")
            if norm is None:
                no_norm_count += 1

            active = "Активна" if item.get("is_active") else "Откл."
            tags: Tuple[str, ...] = ()

            if not item.get("is_active"):
                tags = ("inactive",)
            elif norm is None:
                tags = ("without_norm",)

            iid = self.labor_item_tree.insert(
                "",
                "end",
                values=(
                    item.get("work_type_name") or "",
                    item.get("code") or "",
                    item.get("name") or "",
                    item.get("uom_code") or "",
                    _fmt_qty(norm),
                    _fmt_date_iso(item.get("current_norm_from")),
                    active,
                ),
                tags=tags,
            )

            if selected_id and int(item["id"]) == selected_id:
                selected_iid = iid

        self.lbl_labor_summary.config(
            text=(
                f"Работ: {len(self._labor_items_data)}"
                f"  |  Без актуальной нормы: {no_norm_count}"
            )
        )

        if selected_iid:
            self.labor_item_tree.selection_set(selected_iid)
            self.labor_item_tree.focus(selected_iid)
            self.labor_item_tree.see(selected_iid)
            self._labor_load_norms()
        else:
            self._labor_clear_norms()

    def _labor_item_sel(self) -> Optional[Dict[str, Any]]:
        sel = self.labor_item_tree.selection()
        if not sel:
            return None

        idx = self.labor_item_tree.index(sel[0])
        if 0 <= idx < len(self._labor_items_data):
            return self._labor_items_data[idx]

        return None

    def _labor_norm_sel(self) -> Optional[Dict[str, Any]]:
        sel = self.labor_norm_tree.selection()
        if not sel:
            return None

        idx = self.labor_norm_tree.index(sel[0])
        if 0 <= idx < len(self._labor_norms_data):
            return self._labor_norms_data[idx]

        return None

    def _labor_clear_norms(self):
        self.labor_norm_tree.delete(
            *self.labor_norm_tree.get_children()
        )
        self._labor_norms_data = []
        self.lbl_labor_item_info.config(
            text="Выберите работу в списке слева."
        )

    def _labor_load_norms(self):
        item = self._labor_item_sel()
        if not item:
            self._labor_clear_norms()
            return

        try:
            self._labor_norms_data = GprLaborDictionaryService.load_norms(
                int(item["id"])
            )
        except Exception as exc:
            messagebox.showerror(
                "Нормы ЗТР",
                f"Ошибка загрузки норм:\n{exc}",
                parent=self,
            )
            return

        self.labor_norm_tree.delete(
            *self.labor_norm_tree.get_children()
        )

        self.lbl_labor_item_info.config(
            text=(
                f"Работа: {item.get('name') or '—'}\n"
                f"Тип работ: {item.get('work_type_name') or '—'}\n"
                f"Единица измерения: {item.get('uom_code') or '—'}"
            )
        )

        today = _today()

        for norm in self._labor_norms_data:
            is_current = (
                bool(norm.get("is_active"))
                and norm.get("effective_from") <= today
                and (
                    norm.get("effective_to") is None
                    or norm.get("effective_to") >= today
                )
            )

            active = "Активна" if norm.get("is_active") else "Откл."
            source_parts = [
                norm.get("source_name") or "",
                norm.get("source_code") or "",
                norm.get("source_version") or "",
            ]
            source = " | ".join(
                x for x in source_parts if x.strip()
            )

            tags: Tuple[str, ...] = ()
            if not norm.get("is_active"):
                tags = ("inactive",)
            elif is_current:
                tags = ("current",)

            self.labor_norm_tree.insert(
                "",
                "end",
                values=(
                    _fmt_date_iso(norm.get("effective_from")),
                    _fmt_date_iso(norm.get("effective_to")) or "Бессрочно",
                    _fmt_qty(norm.get("labor_hours_per_unit")),
                    _fmt_qty(norm.get("default_productivity_factor")),
                    source,
                    active,
                ),
                tags=tags,
            )

    def _labor_add_item(self):
        if not self._labor_wt_cache or not self._labor_uom_cache:
            self._labor_load_caches()

        if not self._labor_wt_cache:
            messagebox.showwarning(
                "Работы",
                "Нет активных типов работ.",
                parent=self,
            )
            return

        if not self._labor_uom_cache:
            messagebox.showwarning(
                "Работы",
                "Нет единиц измерения.",
                parent=self,
            )
            return

        dlg = _WorkItemDialog(
            self,
            work_types=self._labor_wt_cache,
            uoms=self._labor_uom_cache,
            init={"sort_order": 100},
        )
        if not dlg.result:
            return

        try:
            GprLaborDictionaryService.create_work_item(
                **dlg.result,
                user_id=_user_id(self.app_ref),
            )
        except Exception as exc:
            messagebox.showerror(
                "Работы",
                f"Не удалось создать работу:\n{exc}",
                parent=self,
            )
            return

        self._labor_load_items()

    def _labor_import_excel(self):
        file_path = filedialog.askopenfilename(
            parent=self,
            title="Выберите Excel-файл с нормами ЗТР",
            filetypes=[
                ("Excel files", "*.xlsx *.xlsm"),
                ("All files", "*.*"),
            ],
        )

        if not file_path:
            return

        try:
            rows, warnings, sheet_name = (
                GprLaborExcelImportService.read_excel_rows(file_path)
            )
        except Exception as exc:
            messagebox.showerror(
                "Импорт Excel",
                f"Не удалось прочитать файл:\n\n{exc}",
                parent=self,
            )
            return

        dlg = _LaborExcelImportDialog(self, file_path)
        if not dlg.result:
            return

        preview_text = (
            f"Файл: {os.path.basename(file_path)}\n"
            f"Лист: {sheet_name}\n"
            f"Корректных строк для импорта: {len(rows)}\n"
            f"Предупреждений при чтении: {len(warnings)}\n\n"
            "Продолжить импорт?"
        )

        if not messagebox.askyesno(
            "Импорт Excel",
            preview_text,
            parent=self,
        ):
            return

        import_note = (
            f"Импортировано из файла {os.path.basename(file_path)}, "
            f"лист «{sheet_name}»."
        )

        try:
            result = GprLaborExcelImportService.import_rows(
                rows,
                effective_from=dlg.result["effective_from"],
                source_name=dlg.result["source_name"],
                user_id=_user_id(self.app_ref),
                import_note=import_note,
            )
        except Exception as exc:
            messagebox.showerror(
                "Импорт Excel",
                (
                    "Импорт не выполнен. Все изменения отменены.\n\n"
                    f"Ошибка:\n{exc}"
                ),
                parent=self,
            )
            return

        self._labor_load_caches()
        self._labor_load_items()

        summary = (
            "Импорт завершён успешно.\n\n"
            f"Создано типов работ: {result['work_types_created']}\n"
            f"Создано единиц измерения: {result['uoms_created']}\n"
            f"Создано работ: {result['work_items_created']}\n"
            f"Создано норм: {result['norms_created']}\n"
            f"Обновлено норм: {result['norms_updated']}"
        )

        if warnings:
            preview_warnings = warnings[:15]
            summary += (
                f"\n\nПредупреждений: {len(warnings)}\n"
                + "\n".join(f"• {text}" for text in preview_warnings)
            )

            if len(warnings) > len(preview_warnings):
                summary += (
                    f"\n• ... ещё {len(warnings) - len(preview_warnings)}"
                )

        messagebox.showinfo(
            "Импорт Excel",
            summary,
            parent=self,
        )
    
    def _labor_edit_item(self):
        item = self._labor_item_sel()
        if not item:
            messagebox.showinfo(
                "Работы",
                "Выберите работу.",
                parent=self,
            )
            return

        self._labor_load_caches()

        dlg = _WorkItemDialog(
            self,
            work_types=self._labor_wt_cache,
            uoms=self._labor_uom_cache,
            init=item,
        )
        if not dlg.result:
            return

        try:
            GprLaborDictionaryService.update_work_item(
                int(item["id"]),
                **dlg.result,
                user_id=_user_id(self.app_ref),
            )
        except Exception as exc:
            messagebox.showerror(
                "Работы",
                f"Не удалось изменить работу:\n{exc}",
                parent=self,
            )
            return

        self._labor_load_items()

    def _labor_toggle_item(self):
        item = self._labor_item_sel()
        if not item:
            messagebox.showinfo(
                "Работы",
                "Выберите работу.",
                parent=self,
            )
            return

        action = "отключить" if item.get("is_active") else "включить"
        if not messagebox.askyesno(
            "Работы",
            f"{action.capitalize()} работу «{item.get('name')}»?",
            parent=self,
        ):
            return

        try:
            GprLaborDictionaryService.toggle_work_item(
                int(item["id"]),
                _user_id(self.app_ref),
            )
        except Exception as exc:
            messagebox.showerror(
                "Работы",
                f"Не удалось изменить статус работы:\n{exc}",
                parent=self,
            )
            return

        self._labor_load_items()

    def _labor_add_norm(self):
        item = self._labor_item_sel()
        if not item:
            messagebox.showinfo(
                "Нормы ЗТР",
                "Сначала выберите работу.",
                parent=self,
            )
            return

        dlg = _LaborNormDialog(self, item)
        if not dlg.result:
            return

        try:
            GprLaborDictionaryService.create_labor_norm(
                work_item_id=int(item["id"]),
                **dlg.result,
                user_id=_user_id(self.app_ref),
            )
        except Exception as exc:
            messagebox.showerror(
                "Нормы ЗТР",
                (
                    "Не удалось создать норматив.\n\n"
                    f"{exc}\n\n"
                    "Проверьте, что период действия не пересекается "
                    "с уже существующей нормой."
                ),
                parent=self,
            )
            return

        self._labor_load_items()

    def _labor_edit_norm(self):
        item = self._labor_item_sel()
        norm = self._labor_norm_sel()

        if not item:
            messagebox.showinfo(
                "Нормы ЗТР",
                "Сначала выберите работу.",
                parent=self,
            )
            return

        if not norm:
            messagebox.showinfo(
                "Нормы ЗТР",
                "Выберите норматив.",
                parent=self,
            )
            return

        dlg = _LaborNormDialog(self, item, init=norm)
        if not dlg.result:
            return

        try:
            GprLaborDictionaryService.update_labor_norm(
                int(norm["id"]),
                **dlg.result,
                user_id=_user_id(self.app_ref),
            )
        except Exception as exc:
            messagebox.showerror(
                "Нормы ЗТР",
                (
                    "Не удалось изменить норматив.\n\n"
                    f"{exc}\n\n"
                    "Проверьте период действия нормы."
                ),
                parent=self,
            )
            return

        self._labor_load_items()

    def _labor_toggle_norm(self):
        norm = self._labor_norm_sel()
        if not norm:
            messagebox.showinfo(
                "Нормы ЗТР",
                "Выберите норматив.",
                parent=self,
            )
            return

        action = "отключить" if norm.get("is_active") else "включить"

        if not messagebox.askyesno(
            "Нормы ЗТР",
            (
                f"{action.capitalize()} норматив "
                f"{_fmt_qty(norm.get('labor_hours_per_unit'))} чел.-ч/ед.?"
            ),
            parent=self,
        ):
            return

        try:
            GprLaborDictionaryService.toggle_labor_norm(
                int(norm["id"]),
                _user_id(self.app_ref),
            )
        except Exception as exc:
            messagebox.showerror(
                "Нормы ЗТР",
                f"Не удалось изменить статус норматива:\n{exc}",
                parent=self,
            )
            return

        self._labor_load_items()

    # ══════════════════════════════════════════════════════
    #  ШАБЛОНЫ + ЗАДАЧИ ШАБЛОНОВ (professional)
    # ══════════════════════════════════════════════════════
    def _build_tpl_tab(self, parent):
        self._tpl_data: List[Dict] = []
        self._tt_data: List[Dict] = []
        self._wt_cache: List[Dict] = []
        self._uom_cache: List[Dict] = []

        search_bar = tk.Frame(parent, bg=C["panel"])
        search_bar.pack(fill="x", padx=8, pady=(8, 4))

        tk.Label(search_bar, text="Поиск шаблона:", bg=C["panel"]).pack(side="left")
        self.var_tpl_search = tk.StringVar()
        ent_search = ttk.Entry(search_bar, textvariable=self.var_tpl_search, width=32)
        ent_search.pack(side="left", padx=(6, 8))
        ent_search.bind("<KeyRelease>", lambda _e: self._tpl_load())

        ttk.Button(search_bar, text="🔃 Обновить", command=self._tpl_load).pack(side="left", padx=2)

        pw = tk.PanedWindow(parent, orient="horizontal", sashrelief="raised", bg=C["bg"])
        pw.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        left = tk.Frame(pw, bg=C["panel"])
        right = tk.Frame(pw, bg=C["panel"])
        pw.add(left, minsize=320)
        pw.add(right, minsize=560)

        # Левая панель
        top_left = tk.LabelFrame(left, text=" Шаблоны ", bg=C["panel"], padx=8, pady=6)
        top_left.pack(fill="both", expand=True)

        bar = tk.Frame(top_left, bg=C["panel"])
        bar.pack(fill="x")
        ttk.Button(bar, text="➕ Добавить", command=self._tpl_add).pack(side="left", padx=2)
        ttk.Button(bar, text="✏️ Редактировать", command=self._tpl_rename).pack(side="left", padx=2)
        ttk.Button(bar, text="📄 Дублировать", command=self._tpl_duplicate).pack(side="left", padx=2)
        ttk.Button(bar, text="🔄 Вкл/Выкл", command=self._tpl_toggle).pack(side="left", padx=2)
        ttk.Button(bar, text="🗑 Удалить", command=self._tpl_delete).pack(side="left", padx=2)

        cols_t = ("name", "category", "tasks", "active")
        self.tpl_tree = ttk.Treeview(
            top_left, columns=cols_t, show="headings",
            selectmode="browse", height=16
        )
        self.tpl_tree.heading("name", text="Наименование")
        self.tpl_tree.heading("category", text="Категория")
        self.tpl_tree.heading("tasks", text="Задач")
        self.tpl_tree.heading("active", text="Статус")
        self.tpl_tree.column("name", width=180, anchor="w")
        self.tpl_tree.column("category", width=100, anchor="w")
        self.tpl_tree.column("tasks", width=60, anchor="center")
        self.tpl_tree.column("active", width=70, anchor="center")

        vsb_t = ttk.Scrollbar(top_left, orient="vertical", command=self.tpl_tree.yview)
        self.tpl_tree.configure(yscrollcommand=vsb_t.set)
        self.tpl_tree.pack(side="left", fill="both", expand=True, pady=(6, 0))
        vsb_t.pack(side="right", fill="y", pady=(6, 0))

        self.tpl_tree.tag_configure("inactive", foreground="#bbb")
        self.tpl_tree.bind("<<TreeviewSelect>>", lambda _e: self._tt_load())
        self.tpl_tree.bind("<Double-1>", lambda _e: self._tpl_rename())

        # Правая панель
        info = tk.LabelFrame(right, text=" Карточка шаблона ", bg=C["panel"], padx=10, pady=8)
        info.pack(fill="x", pady=(0, 6))

        self.lbl_tpl_info = tk.Label(
            info,
            text="Выберите шаблон",
            bg=C["panel"],
            fg=C["text2"],
            justify="left",
            anchor="w",
        )
        self.lbl_tpl_info.pack(fill="x")

        tasks_box = tk.LabelFrame(right, text=" Состав шаблона ", bg=C["panel"], padx=8, pady=6)
        tasks_box.pack(fill="both", expand=True)

        bar2 = tk.Frame(tasks_box, bg=C["panel"])
        bar2.pack(fill="x")
        ttk.Button(bar2, text="➕ Задача", command=lambda: self._tt_add("task")).pack(side="left", padx=2)
        ttk.Button(bar2, text="📁 Группа", command=lambda: self._tt_add("group")).pack(side="left", padx=2)
        ttk.Button(bar2, text="🟦 Титул", command=lambda: self._tt_add("title")).pack(side="left", padx=2)
        ttk.Button(bar2, text="✏️ Редактировать", command=self._tt_edit).pack(side="left", padx=2)
        ttk.Button(bar2, text="📄 Дублировать", command=self._tt_duplicate).pack(side="left", padx=2)
        ttk.Button(bar2, text="⬆ Вверх", command=lambda: self._tt_move(-1)).pack(side="left", padx=2)
        ttk.Button(bar2, text="⬇ Вниз", command=lambda: self._tt_move(1)).pack(side="left", padx=2)
        ttk.Button(bar2, text="↻ Пересчитать порядок", command=self._tt_reindex).pack(side="left", padx=2)
        ttk.Button(bar2, text="🗑 Удалить", command=self._tt_del).pack(side="left", padx=2)

        cols_tt = ("sort", "work_type", "name", "uom", "qty", "milestone")
        self.tt_tree = ttk.Treeview(
            tasks_box, columns=cols_tt, show="tree headings",
            selectmode="browse", height=18
        )
        self.tt_tree.heading("#0", text="Иерархия")
        self.tt_tree.column("#0", width=220, anchor="w")

        for c, t, w, a in [
            ("sort", "№", 50, "center"),
            ("work_type", "Тип работ", 160, "w"),
            ("name", "Название", 220, "w"),
            ("uom", "Ед.", 60, "center"),
            ("qty", "Объём", 80, "e"),
            ("milestone", "Веха", 55, "center"),
        ]:
            self.tt_tree.heading(c, text=t)
            self.tt_tree.column(c, width=w, anchor=a)

        vsb_tt = ttk.Scrollbar(tasks_box, orient="vertical", command=self.tt_tree.yview)
        self.tt_tree.configure(yscrollcommand=vsb_tt.set)
        self.tt_tree.pack(side="left", fill="both", expand=True, pady=(6, 0))
        vsb_tt.pack(side="right", fill="y", pady=(6, 0))

        self.tt_tree.bind("<Double-1>", lambda _e: self._tt_edit())

        self.tt_tree.tag_configure("group", font=("Segoe UI", 9, "bold"))
        self.tt_tree.tag_configure("title", font=("Segoe UI", 9, "bold"), background="#e3f2fd")
        self.tt_tree.tag_configure("task", font=("Segoe UI", 9))

        self._tpl_load()

    def _tpl_load(self):
        sel_id = None
        sel = self.tpl_tree.selection() if hasattr(self, "tpl_tree") else ()
        if sel:
            cur_tpl = self._tpl_sel()
            sel_id = cur_tpl["id"] if cur_tpl else None

        self.tpl_tree.delete(*self.tpl_tree.get_children())
        self._tpl_data = []

        try:
            self._tpl_data = GprTemplateService.load_templates_full(
                self.var_tpl_search.get() if hasattr(self, "var_tpl_search") else ""
            )
        except Exception as e:
            messagebox.showerror("Шаблоны", f"Ошибка загрузки шаблонов:\n{e}", parent=self)
            return

        selected_iid = None
        for i, t in enumerate(self._tpl_data):
            active = "Да" if t.get("is_active") else "Нет"
            tags = () if t.get("is_active") else ("inactive",)
            iid = self.tpl_tree.insert(
                "", "end",
                values=(
                    t["name"],
                    t.get("category") or "",
                    t.get("task_count", 0),
                    active,
                ),
                tags=tags
            )
            if sel_id and int(t["id"]) == int(sel_id):
                selected_iid = iid

        if selected_iid:
            self.tpl_tree.selection_set(selected_iid)
            self.tpl_tree.focus(selected_iid)
            self._tt_load()
        else:
            self._tt_clear_view()

    def _tpl_sel(self) -> Optional[Dict]:
        sel = self.tpl_tree.selection()
        if not sel:
            return None
        idx = self.tpl_tree.index(sel[0])
        return self._tpl_data[idx] if 0 <= idx < len(self._tpl_data) else None

    def _tpl_add(self):
        dlg = _TemplateEditDialog(self)
        if not dlg.result:
            return
        try:
            GprTemplateService.create_template(
                dlg.result["name"],
                category=dlg.result.get("category"),
                description=dlg.result.get("description"),
                note=dlg.result.get("note"),
                created_by=_user_id(self.app_ref),
            )
        except Exception as e:
            messagebox.showerror("Шаблоны", f"Ошибка создания шаблона:\n{e}", parent=self)
            return
        self._tpl_load()

    def _tpl_rename(self):
        t = self._tpl_sel()
        if not t:
            messagebox.showinfo("Шаблоны", "Выберите шаблон.", parent=self)
            return

        info = (
            f"Создал: {t.get('creator_name') or '—'}\n"
            f"Обновил: {t.get('updater_name') or '—'}\n"
            f"Задач: {t.get('task_count', 0)}\n"
            f"Создан: {_fmt_dt(t.get('created_at'))}\n"
            f"Обновлён: {_fmt_dt(t.get('updated_at'))}"
        )
        dlg = _TemplateEditDialog(
            self,
            init={
                "name": t["name"],
                "category": t.get("category") or "",
                "description": t.get("description") or "",
                "note": t.get("note") or "",
                "_info": info,
            }
        )
        if not dlg.result:
            return
        try:
            GprTemplateService.update_template(
                t["id"],
                dlg.result["name"],
                category=dlg.result.get("category"),
                description=dlg.result.get("description"),
                note=dlg.result.get("note"),
                updated_by=_user_id(self.app_ref),
            )
        except Exception as e:
            messagebox.showerror("Шаблоны", f"Ошибка изменения шаблона:\n{e}", parent=self)
            return
        self._tpl_load()

    def _tpl_toggle(self):
        t = self._tpl_sel()
        if not t:
            messagebox.showinfo("Шаблоны", "Выберите шаблон.", parent=self)
            return
        try:
            GprTemplateService.toggle_template(t["id"], updated_by=_user_id(self.app_ref))
        except Exception as e:
            messagebox.showerror("Шаблоны", f"Ошибка изменения статуса:\n{e}", parent=self)
            return
        self._tpl_load()

    def _tpl_delete(self):
        t = self._tpl_sel()
        if not t:
            messagebox.showinfo("Шаблоны", "Выберите шаблон.", parent=self)
            return
        if not messagebox.askyesno(
            "Шаблоны",
            f"Удалить шаблон «{t['name']}»?\n\nБудут удалены и все задачи шаблона.",
            parent=self
        ):
            return
        try:
            GprTemplateService.delete_template(t["id"])
        except Exception as e:
            messagebox.showerror("Шаблоны", f"Ошибка удаления шаблона:\n{e}", parent=self)
            return
        self._tpl_load()

    def _tpl_duplicate(self):
        t = self._tpl_sel()
        if not t:
            messagebox.showinfo("Шаблоны", "Выберите шаблон.", parent=self)
            return

        new_name = simpledialog.askstring(
            "Дублирование шаблона",
            "Новое наименование шаблона:",
            initialvalue=f"{t['name']} (копия)",
            parent=self
        )
        if not new_name or not new_name.strip():
            return

        try:
            GprTemplateService.duplicate_template(
                t["id"],
                new_name.strip(),
                created_by=_user_id(self.app_ref),
            )
        except Exception as e:
            messagebox.showerror("Шаблоны", f"Ошибка копирования шаблона:\n{e}", parent=self)
            return
        self._tpl_load()

    def _tt_clear_view(self):
        self.tt_tree.delete(*self.tt_tree.get_children())
        self._tt_data = []
        if hasattr(self, "lbl_tpl_info"):
            self.lbl_tpl_info.config(text="Выберите шаблон")

    def _load_wt_uom_cache(self):
        conn = None
        try:
            conn = _conn()
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("""
                    SELECT id, name
                    FROM public.gpr_work_types
                    WHERE is_active = true
                    ORDER BY sort_order, name
                """)
                self._wt_cache = [dict(r) for r in cur.fetchall()]

                cur.execute("SELECT code, name FROM public.gpr_uom ORDER BY code")
                self._uom_cache = [dict(r) for r in cur.fetchall()]
        finally:
            _release(conn)

    def _tt_load(self):
        self.tt_tree.delete(*self.tt_tree.get_children())
        self._tt_data = []

        tpl = self._tpl_sel()
        if not tpl:
            self._tt_clear_view()
            return

        try:
            self._tt_data = GprTemplateService.load_template_tasks(tpl["id"])
            self._load_wt_uom_cache()
        except Exception as e:
            messagebox.showerror("Шаблоны", f"Ошибка загрузки задач шаблона:\n{e}", parent=self)
            return

        self.lbl_tpl_info.config(
            text=(
                f"Шаблон: {tpl['name']}\n"
                f"Категория: {tpl.get('category') or '—'}\n"
                f"Статус: {'Активен' if tpl.get('is_active') else 'Отключён'}\n"
                f"Задач: {len(self._tt_data)}\n"
                f"Создал: {tpl.get('creator_name') or '—'}\n"
                f"Обновил: {tpl.get('updater_name') or '—'}\n"
                f"Обновлён: {_fmt_dt(tpl.get('updated_at'))}\n"
                f"Описание: {tpl.get('description') or '—'}\n"
                f"Примечание: {tpl.get('note') or '—'}"
            )
        )

        children: Dict[Optional[int], List[Dict]] = {}
        for row in self._tt_data:
            children.setdefault(row.get("parent_id"), []).append(row)

        for key in children:
            children[key].sort(key=lambda x: (x.get("sort_order", 0), x.get("id", 0)))

        def add_node(node: Dict, parent_iid: str = ""):
            ms = "✓" if node.get("is_milestone") else ""
            iid = str(node["id"])
            row_kind = (node.get("row_kind") or "task").strip()
        
            if row_kind == "group":
                text = f"📁 {node['name']}"
                wt_name = "Группа"
                qty = ""
                ms = ""
                tags = ("group",)
            elif row_kind == "title":
                text = f"🟦 {node['name']}"
                wt_name = "Титул"
                qty = ""
                ms = ""
                tags = ("title",)
            else:
                text = node["name"]
                wt_name = node.get("wt_name", "")
                qty = _fmt_qty(node.get("default_qty"))
                tags = ("task",)
        
            self.tt_tree.insert(
                parent_iid,
                "end",
                iid=iid,
                text=text,
                values=(
                    node.get("sort_order", 0),
                    wt_name,
                    node.get("name", ""),
                    node.get("uom_code") or "",
                    qty,
                    ms,
                ),
                tags=tags,
            )
            for child in children.get(node["id"], []):
                add_node(child, iid)

    def _tt_sel(self) -> Optional[Dict]:
        sel = self.tt_tree.selection()
        if not sel:
            return None
        iid = sel[0]
        for x in self._tt_data:
            if str(x["id"]) == str(iid):
                return x
        return None

    def _tt_add(self, row_kind: str = "task"):
        tpl = self._tpl_sel()
        if not tpl:
            messagebox.showinfo("Шаблоны", "Выберите шаблон.", parent=self)
            return
    
        if not self._wt_cache or not self._uom_cache:
            self._load_wt_uom_cache()
    
        if not self._wt_cache:
            messagebox.showwarning("Шаблоны", "Нет доступных типов работ.", parent=self)
            return
    
        dlg = _TemplateTaskDialog(
            self,
            work_types=self._wt_cache,
            uoms=self._uom_cache,
            parents=self._tt_data,
            init={
                "row_kind": row_kind,
                "sort_order": len(self._tt_data) * 10 + 10,
            },
        )
        if not dlg.result:
            return
    
        r = dlg.result
        conn = None
        try:
            conn = _conn()
            with conn, conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO public.gpr_template_tasks
                        (template_id, work_type_id, parent_id, name, uom_code,
                         default_qty, is_milestone, sort_order, row_kind)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    tpl["id"],
                    r["work_type_id"],
                    r.get("parent_id"),
                    r["name"],
                    r.get("uom_code"),
                    r.get("default_qty"),
                    r.get("is_milestone", False),
                    r.get("sort_order", len(self._tt_data) * 10 + 10),
                    r.get("row_kind", "task"),
                ))
        except Exception as e:
            messagebox.showerror("Шаблоны", f"Ошибка добавления строки:\n{e}", parent=self)
            return
        finally:
            _release(conn)
    
        self._tt_load()

    def _tt_edit(self):
        tt = self._tt_sel()
        if not tt:
            messagebox.showinfo("Шаблоны", "Выберите задачу шаблона.", parent=self)
            return

        if not self._wt_cache or not self._uom_cache:
            self._load_wt_uom_cache()

        parents = [p for p in self._tt_data if p["id"] != tt["id"]]

        dlg = _TemplateTaskDialog(
            self,
            work_types=self._wt_cache,
            uoms=self._uom_cache,
            parents=parents,
            init={
                "row_kind": tt.get("row_kind", "task"),
                "work_type_id": tt["work_type_id"],
                "name": tt["name"],
                "uom_code": tt.get("uom_code"),
                "default_qty": tt.get("default_qty"),
                "is_milestone": tt.get("is_milestone", False),
                "parent_id": tt.get("parent_id"),
                "sort_order": tt.get("sort_order", 0),
            },
        )
        if not dlg.result:
            return

        r = dlg.result
        conn = None
        try:
            conn = _conn()
            with conn, conn.cursor() as cur:
                cur.execute("""
                    UPDATE public.gpr_template_tasks
                    SET work_type_id=%s,
                        parent_id=%s,
                        name=%s,
                        uom_code=%s,
                        default_qty=%s,
                        is_milestone=%s,
                        sort_order=%s,
                        row_kind=%s
                    WHERE id=%s
                """, (
                    r["work_type_id"],
                    r.get("parent_id"),
                    r["name"],
                    r.get("uom_code"),
                    r.get("default_qty"),
                    r.get("is_milestone", False),
                    r.get("sort_order", 0),
                    r.get("row_kind", "task"),
                    tt["id"],
                ))
        except Exception as e:
            messagebox.showerror("Шаблоны", f"Ошибка редактирования задачи:\n{e}", parent=self)
            return
        finally:
            _release(conn)

        self._tt_load()

    def _tt_duplicate(self):
        tt = self._tt_sel()
        tpl = self._tpl_sel()
        if not tt or not tpl:
            messagebox.showinfo("Шаблоны", "Выберите задачу шаблона.", parent=self)
            return

        conn = None
        try:
            conn = _conn()
            with conn, conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO public.gpr_template_tasks
                        (template_id, work_type_id, parent_id, name, uom_code,
                         default_qty, is_milestone, sort_order, row_kind)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    tpl["id"],
                    tt["work_type_id"],
                    tt.get("parent_id"),
                    f"{tt['name']} (копия)",
                    tt.get("uom_code"),
                    tt.get("default_qty"),
                    tt.get("is_milestone", False),
                    int(tt.get("sort_order", 0)) + 1,
                    tt.get("row_kind", "task"),
                ))
        except Exception as e:
            messagebox.showerror("Шаблоны", f"Ошибка копирования задачи:\n{e}", parent=self)
            return
        finally:
            _release(conn)

        self._tt_load()

    def _tt_move(self, direction: int):
        tt = self._tt_sel()
        if not tt:
            return

        siblings = [x for x in self._tt_data if x.get("parent_id") == tt.get("parent_id")]
        siblings.sort(key=lambda x: (x.get("sort_order", 0), x.get("id", 0)))

        idx = next((i for i, x in enumerate(siblings) if x["id"] == tt["id"]), None)
        if idx is None:
            return

        new_idx = idx + direction
        if new_idx < 0 or new_idx >= len(siblings):
            return

        other = siblings[new_idx]

        conn = None
        try:
            conn = _conn()
            with conn, conn.cursor() as cur:
                cur.execute(
                    "UPDATE public.gpr_template_tasks SET sort_order=%s WHERE id=%s",
                    (other.get("sort_order", 0), tt["id"])
                )
                cur.execute(
                    "UPDATE public.gpr_template_tasks SET sort_order=%s WHERE id=%s",
                    (tt.get("sort_order", 0), other["id"])
                )
        except Exception as e:
            messagebox.showerror("Шаблоны", f"Ошибка перемещения:\n{e}", parent=self)
            return
        finally:
            _release(conn)

        self._tt_load()

    def _tt_reindex(self):
        tpl = self._tpl_sel()
        if not tpl:
            return

        ordered = sorted(self._tt_data, key=lambda x: (x.get("sort_order", 0), x.get("id", 0)))

        conn = None
        try:
            conn = _conn()
            with conn, conn.cursor() as cur:
                for i, row in enumerate(ordered, start=1):
                    cur.execute("""
                        UPDATE public.gpr_template_tasks
                        SET sort_order=%s
                        WHERE id=%s
                    """, (i * 10, row["id"]))
        except Exception as e:
            messagebox.showerror("Шаблоны", f"Ошибка пересчёта порядка:\n{e}", parent=self)
            return
        finally:
            _release(conn)

        self._tt_load()

    def _tt_del(self):
        tt = self._tt_sel()
        if not tt:
            return
        if not messagebox.askyesno(
            "Шаблоны",
            f"Удалить задачу «{tt['name']}»?\n\nЕсли у неё есть дочерние задачи, у них parent_id станет NULL.",
            parent=self
        ):
            return

        conn = None
        try:
            conn = _conn()
            with conn, conn.cursor() as cur:
                cur.execute("DELETE FROM public.gpr_template_tasks WHERE id=%s", (tt["id"],))
        except Exception as e:
            messagebox.showerror("Шаблоны", f"Ошибка удаления задачи:\n{e}", parent=self)
            return
        finally:
            _release(conn)

        self._tt_load()


# ═══════════════════════════════════════════════════════════════
#  API для main_app
# ═══════════════════════════════════════════════════════════════
def create_gpr_dicts_page(parent, app_ref) -> GprDictionariesPage:
    """Фабричная функция."""
    return GprDictionariesPage(parent, app_ref=app_ref)
