import os
import re
import sys
import configparser
import logging
from dataclasses import dataclass
from datetime import datetime, date, time, timedelta
from pathlib import Path
from typing import List, Tuple, Optional, Any, Dict

import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog

import psycopg2
from psycopg2 import pool
from psycopg2.extras import RealDictCursor, execute_batch

from openpyxl import load_workbook
from urllib.parse import urlparse, parse_qs


# ========================= ЛОГИРОВАНИЕ =========================

logger = logging.getLogger(__name__)
if not logger.handlers:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s"
    )


# ========================= ГЛОБАЛЬНЫЕ КОНСТАНТЫ =========================

APP_TITLE = "Заказ спецтехники"
CONFIG_FILE = "tabel_config.ini"

CONFIG_SECTION_PATHS = "Paths"
CONFIG_SECTION_UI = "UI"
CONFIG_SECTION_INTEGR = "Integrations"
CONFIG_SECTION_ORDERS = "Orders"
CONFIG_SECTION_REMOTE = "Remote"

KEY_SPR = "spravochnik_path"
KEY_SELECTED_DEP = "selected_department"
KEY_PLANNING_ENABLED = "planning_enabled"
KEY_PLANNING_PASSWORD = "planning_password"
KEY_CUTOFF_ENABLED = "cutoff_enabled"
KEY_CUTOFF_HOUR = "cutoff_hour"
KEY_DRIVER_DEPARTMENTS = "driver_departments"
KEY_REMOTE_USE = "use_remote"
KEY_YA_PUBLIC_LINK = "yadisk_public_link"
KEY_YA_PUBLIC_PATH = "yadisk_public_path"

SPRAVOCHNIK_FILE = "Справочник.xlsx"

ORDER_STATUSES = ["Новая", "Назначена", "В работе", "Выполнена", "Отменена"]


# ========================= SETTINGS MANAGER =========================

try:
    import settings_manager as Settings
except Exception:
    Settings = None


# ========================= ПУЛ СОЕДИНЕНИЙ =========================

db_connection_pool = None
USING_SHARED_POOL = False


def set_db_pool(shared_pool):
    """Установка внешнего пула соединений."""
    global db_connection_pool, USING_SHARED_POOL
    db_connection_pool = shared_pool
    USING_SHARED_POOL = True


def release_db_connection(conn):
    """Возврат соединения в пул."""
    global db_connection_pool
    if conn and db_connection_pool:
        db_connection_pool.putconn(conn)


# ========================= УТИЛИТЫ КОНФИГА =========================

def exe_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def config_path() -> Path:
    return exe_dir() / CONFIG_FILE


if Settings:
    ensure_config = Settings.ensure_config
    read_config = Settings.read_config
    write_config = Settings.write_config

    def get_spr_path() -> Path:
        return Settings.get_spr_path_from_config()

    def get_saved_dep() -> str:
        return Settings.get_selected_department_from_config()

    def set_saved_dep(dep: str):
        return Settings.set_selected_department_in_config(dep)

else:
    def ensure_config():
        cp = config_path()
        cfg = configparser.ConfigParser()

        if cp.exists():
            cfg.read(cp, encoding="utf-8")

        changed = False

        if not cfg.has_section(CONFIG_SECTION_PATHS):
            cfg[CONFIG_SECTION_PATHS] = {}
            changed = True
        if KEY_SPR not in cfg[CONFIG_SECTION_PATHS]:
            cfg[CONFIG_SECTION_PATHS][KEY_SPR] = str(exe_dir() / SPRAVOCHNIK_FILE)
            changed = True

        if not cfg.has_section(CONFIG_SECTION_UI):
            cfg[CONFIG_SECTION_UI] = {}
            changed = True
        if KEY_SELECTED_DEP not in cfg[CONFIG_SECTION_UI]:
            cfg[CONFIG_SECTION_UI][KEY_SELECTED_DEP] = "Все"
            changed = True

        if not cfg.has_section(CONFIG_SECTION_INTEGR):
            cfg[CONFIG_SECTION_INTEGR] = {}
            changed = True
        if KEY_PLANNING_ENABLED not in cfg[CONFIG_SECTION_INTEGR]:
            cfg[CONFIG_SECTION_INTEGR][KEY_PLANNING_ENABLED] = "false"
            changed = True
        if KEY_DRIVER_DEPARTMENTS not in cfg[CONFIG_SECTION_INTEGR]:
            cfg[CONFIG_SECTION_INTEGR][KEY_DRIVER_DEPARTMENTS] = "Служба гаража, Автопарк, Транспортный цех"
            changed = True
        if KEY_PLANNING_PASSWORD not in cfg[CONFIG_SECTION_INTEGR]:
            cfg[CONFIG_SECTION_INTEGR][KEY_PLANNING_PASSWORD] = "2025"
            changed = True

        if not cfg.has_section(CONFIG_SECTION_ORDERS):
            cfg[CONFIG_SECTION_ORDERS] = {}
            changed = True
        if KEY_CUTOFF_ENABLED not in cfg[CONFIG_SECTION_ORDERS]:
            cfg[CONFIG_SECTION_ORDERS][KEY_CUTOFF_ENABLED] = "true"
            changed = True
        if KEY_CUTOFF_HOUR not in cfg[CONFIG_SECTION_ORDERS]:
            cfg[CONFIG_SECTION_ORDERS][KEY_CUTOFF_HOUR] = "13"
            changed = True

        if not cfg.has_section(CONFIG_SECTION_REMOTE):
            cfg[CONFIG_SECTION_REMOTE] = {}
            changed = True
        if KEY_REMOTE_USE not in cfg[CONFIG_SECTION_REMOTE]:
            cfg[CONFIG_SECTION_REMOTE][KEY_REMOTE_USE] = "false"
            changed = True
        if KEY_YA_PUBLIC_LINK not in cfg[CONFIG_SECTION_REMOTE]:
            cfg[CONFIG_SECTION_REMOTE][KEY_YA_PUBLIC_LINK] = ""
            changed = True
        if KEY_YA_PUBLIC_PATH not in cfg[CONFIG_SECTION_REMOTE]:
            cfg[CONFIG_SECTION_REMOTE][KEY_YA_PUBLIC_PATH] = ""
            changed = True

        if changed or not cp.exists():
            with open(cp, "w", encoding="utf-8") as f:
                cfg.write(f)

    def read_config() -> configparser.ConfigParser:
        ensure_config()
        cfg = configparser.ConfigParser()
        cfg.read(config_path(), encoding="utf-8")
        return cfg

    def write_config(cfg: configparser.ConfigParser):
        with open(config_path(), "w", encoding="utf-8") as f:
            cfg.write(f)

    def get_spr_path() -> Path:
        cfg = read_config()
        raw = cfg.get(CONFIG_SECTION_PATHS, KEY_SPR, fallback=str(exe_dir() / SPRAVOCHNIK_FILE))
        return Path(os.path.expandvars(raw))

    def get_saved_dep() -> str:
        cfg = read_config()
        return cfg.get(CONFIG_SECTION_UI, KEY_SELECTED_DEP, fallback="Все")

    def set_saved_dep(dep: str):
        cfg = read_config()
        if not cfg.has_section(CONFIG_SECTION_UI):
            cfg[CONFIG_SECTION_UI] = {}
        cfg[CONFIG_SECTION_UI][KEY_SELECTED_DEP] = dep or "Все"
        write_config(cfg)


def get_planning_password() -> str:
    cfg = read_config()
    return cfg.get(CONFIG_SECTION_INTEGR, KEY_PLANNING_PASSWORD, fallback="2025").strip()


def get_planning_enabled() -> bool:
    cfg = read_config()
    v = cfg.get(CONFIG_SECTION_INTEGR, KEY_PLANNING_ENABLED, fallback="false").strip().lower()
    return v in ("1", "true", "yes", "on")


def get_cutoff_enabled() -> bool:
    cfg = read_config()
    v = cfg.get(CONFIG_SECTION_ORDERS, KEY_CUTOFF_ENABLED, fallback="true").strip().lower()
    return v in ("1", "true", "yes", "on")


def get_cutoff_hour() -> int:
    cfg = read_config()
    try:
        h = int(cfg.get(CONFIG_SECTION_ORDERS, KEY_CUTOFF_HOUR, fallback="13").strip())
        return min(23, max(0, h))
    except Exception:
        return 13


# ========================= ПАРСИНГ / ВАЛИДАЦИЯ =========================

def parse_hours_value(v: Any) -> Optional[float]:
    s = str(v or "").strip()
    if not s:
        return None

    if "/" in s:
        total = 0.0
        any_part = False
        for part in s.split("/"):
            n = parse_hours_value(part)
            if isinstance(n, (int, float)):
                total += float(n)
                any_part = True
        return total if any_part else None

    if ":" in s:
        p = s.split(":")
        try:
            hh = float(p[0].replace(",", "."))
            mm = float((p[1] if len(p) > 1 else "0").replace(",", "."))
            ss = float((p[2] if len(p) > 2 else "0").replace(",", "."))
            return hh + mm / 60.0 + ss / 3600.0
        except Exception:
            pass

    s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def parse_time_str(s: str) -> Optional[str]:
    s = (s or "").strip()
    if not s:
        return None
    m = re.match(r"^\s*(\d{1,2}):(\d{2})\s*$", s)
    if not m:
        return None
    hh = int(m.group(1))
    mm = int(m.group(2))
    if not (0 <= hh <= 23 and 0 <= mm <= 59):
        return None
    return f"{hh:02d}:{mm:02d}"


def parse_date_any(s: str) -> Optional[date]:
    s = (s or "").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def normalize_plate(plate: str) -> str:
    return re.sub(r"\s+", "", (plate or "").strip().upper())


def safe_str(v: Any) -> str:
    return str(v or "").strip()


def validate_phone(phone: str) -> bool:
    digits = re.sub(r"\D+", "", phone or "")
    return len(digits) >= 5


def validate_future_order_date(req: date) -> bool:
    return req > date.today()


def make_interval(req_date_str: str, req_time_str: str, hours_val: Any) -> Optional[Tuple[datetime, datetime]]:
    req_date = parse_date_any(req_date_str)
    req_time = parse_time_str(req_time_str)
    hours_num = parse_hours_value(hours_val)

    if not req_date or not req_time or hours_num is None or hours_num <= 0:
        return None

    hh, mm = map(int, req_time.split(":"))
    dt_from = datetime.combine(req_date, time(hour=hh, minute=mm))
    dt_to = dt_from + timedelta(hours=float(hours_num))
    return dt_from, dt_to


def intervals_intersect(a_from: datetime, a_to: datetime, b_from: datetime, b_to: datetime) -> bool:
    return max(a_from, b_from) < min(a_to, b_to)


# ========================= DB =========================

def get_db_connection():
    global db_connection_pool

    if db_connection_pool:
        return db_connection_pool.getconn()

    if USING_SHARED_POOL:
        raise RuntimeError("Общий пул соединений не был передан в модуль.")

    if not Settings:
        raise RuntimeError("settings_manager не доступен, не могу прочитать параметры БД")

    provider = Settings.get_db_provider().strip().lower()
    if provider != "postgres":
        raise RuntimeError(f"Ожидался provider=postgres, а в настройках: {provider!r}")

    db_url = Settings.get_database_url().strip()
    if not db_url:
        raise RuntimeError("В настройках не указана строка подключения (DATABASE_URL)")

    url = urlparse(db_url)
    user = url.username
    password = url.password
    host = url.hostname or "localhost"
    port = url.port or 5432
    dbname = url.path.lstrip("/")
    q = parse_qs(url.query)
    sslmode = (q.get("sslmode", [Settings.get_db_sslmode()])[0] or "require")

    db_connection_pool = pool.SimpleConnectionPool(
        minconn=1,
        maxconn=5,
        host=host,
        port=port,
        dbname=dbname,
        user=user,
        password=password,
        sslmode=sslmode,
    )
    return db_connection_pool.getconn()


def get_or_create_object(cur, excel_id: str, address: str) -> Optional[int]:
    excel_id = safe_str(excel_id)
    address = safe_str(address)

    if not (excel_id or address):
        return None

    if excel_id:
        cur.execute("SELECT id, COALESCE(address, '') FROM objects WHERE excel_id = %s", (excel_id,))
        row = cur.fetchone()
        if row:
            object_id = row[0]
            current_address = row[1] or ""
            if address and address != current_address:
                cur.execute("UPDATE objects SET address = %s WHERE id = %s", (address, object_id))
            return object_id

        cur.execute(
            "INSERT INTO objects (excel_id, address) VALUES (%s, %s) RETURNING id",
            (excel_id, address)
        )
        return cur.fetchone()[0]

    cur.execute("SELECT id FROM objects WHERE address = %s", (address,))
    row = cur.fetchone()
    if row:
        return row[0]

    cur.execute(
        "INSERT INTO objects (excel_id, address) VALUES (NULL, %s) RETURNING id",
        (address,)
    )
    return cur.fetchone()[0]


def validate_order_payload(data: dict):
    if not isinstance(data, dict):
        raise ValueError("Неверный формат заявки")

    dep = safe_str(data.get("department"))
    fio = safe_str(data.get("requester_fio"))
    phone = safe_str(data.get("requester_phone"))
    comment = safe_str(data.get("comment"))
    positions = data.get("positions") or []

    req_date = parse_date_any(data.get("date"))
    if not dep:
        raise ValueError("Не указано подразделение")
    if not fio:
        raise ValueError("Не указано ФИО заявителя")
    if not validate_phone(phone):
        raise ValueError("Некорректный телефон заявителя")
    if not req_date or not validate_future_order_date(req_date):
        raise ValueError("Дата заявки должна быть позже текущей даты")
    if not safe_str((data.get("object") or {}).get("address")):
        raise ValueError("Не указан адрес объекта")
    if not comment:
        raise ValueError("Не указан комментарий")
    if not positions:
        raise ValueError("В заявке должна быть хотя бы одна позиция")

    for idx, p in enumerate(positions, start=1):
        tech = safe_str(p.get("tech"))
        qty_raw = p.get("qty")
        time_str = parse_time_str(p.get("time"))
        hours = parse_hours_value(p.get("hours"))

        try:
            qty = int(qty_raw)
        except Exception:
            qty = 0

        if not tech:
            raise ValueError(f"Позиция {idx}: не указана техника")
        if qty <= 0:
            raise ValueError(f"Позиция {idx}: количество должно быть больше 0")
        if not time_str:
            raise ValueError(f"Позиция {idx}: неверное время подачи")
        if hours is None or hours <= 0:
            raise ValueError(f"Позиция {idx}: неверное количество часов")


def save_transport_order_to_db(data: dict, edit_order_id: Optional[int] = None) -> int:
    """
    Сохраняет заявку.
    При edit_order_id выполняется UPDATE заголовка и полная замена позиций,
    но сама заявка не удаляется.
    """
    validate_order_payload(data)

    conn = None
    try:
        conn = get_db_connection()
        with conn:
            with conn.cursor() as cur:
                obj = data.get("object") or {}
                object_id = get_or_create_object(cur, obj.get("id", ""), obj.get("address", ""))

                created_at = datetime.strptime(data["created_at"], "%Y-%m-%dT%H:%M:%S")
                order_date = datetime.strptime(data["date"], "%Y-%m-%d").date()
                user_id = data.get("user_id")

                if edit_order_id:
                    cur.execute("SELECT id FROM transport_orders WHERE id = %s", (edit_order_id,))
                    existing = cur.fetchone()
                    if not existing:
                        raise ValueError(f"Редактируемая заявка ID={edit_order_id} не найдена")

                    cur.execute(
                        """
                        UPDATE transport_orders
                        SET date = %s,
                            department = %s,
                            requester_fio = %s,
                            requester_phone = %s,
                            object_id = %s,
                            comment = %s,
                            user_id = COALESCE(%s, user_id)
                        WHERE id = %s
                        """,
                        (
                            order_date,
                            safe_str(data.get("department")),
                            safe_str(data.get("requester_fio")),
                            safe_str(data.get("requester_phone")),
                            object_id,
                            safe_str(data.get("comment")),
                            user_id,
                            edit_order_id,
                        ),
                    )
                    order_id = edit_order_id

                    cur.execute("DELETE FROM transport_order_positions WHERE order_id = %s", (order_id,))
                else:
                    cur.execute(
                        """
                        INSERT INTO transport_orders
                            (created_at, date, department, requester_fio, requester_phone, object_id, comment, user_id)
                        VALUES
                            (%s, %s, %s, %s, %s, %s, %s, %s)
                        RETURNING id
                        """,
                        (
                            created_at,
                            order_date,
                            safe_str(data.get("department")),
                            safe_str(data.get("requester_fio")),
                            safe_str(data.get("requester_phone")),
                            object_id,
                            safe_str(data.get("comment")),
                            user_id,
                        ),
                    )
                    order_id = cur.fetchone()[0]

                positions_payload = []
                for p in data.get("positions", []):
                    time_str = parse_time_str(p.get("time"))
                    tval = datetime.strptime(time_str, "%H:%M").time() if time_str else None
                    positions_payload.append(
                        (
                            order_id,
                            safe_str(p.get("tech")),
                            int(p.get("qty") or 0),
                            tval,
                            float(parse_hours_value(p.get("hours")) or 0.0),
                            safe_str(p.get("note")),
                            "Новая" if not edit_order_id else safe_str(p.get("status")) or "Новая",
                        )
                    )

                execute_batch(
                    cur,
                    """
                    INSERT INTO transport_order_positions
                        (order_id, tech, qty, time, hours, note, status)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """,
                    positions_payload
                )

        return order_id
    finally:
        if conn:
            release_db_connection(conn)


def load_user_transport_orders(user_id: int) -> List[Dict[str, Any]]:
    if not user_id:
        return []

    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT
                    o.id,
                    o.date,
                    o.created_at,
                    COALESCE(o.department, '') AS department,
                    COALESCE(o.requester_fio, '') AS requester_fio,
                    COALESCE(obj.address, '') AS object_address,
                    (SELECT COUNT(p.id) FROM transport_order_positions p WHERE p.order_id = o.id) AS positions_count
                FROM transport_orders o
                LEFT JOIN objects obj ON o.object_id = obj.id
                WHERE o.user_id = %s
                ORDER BY o.date DESC, o.id DESC
                """,
                (user_id,),
            )
            return [dict(r) for r in cur.fetchall()]
    finally:
        if conn:
            release_db_connection(conn)


def get_transport_order_with_positions_from_db(order_id: int) -> Dict[str, Any]:
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT
                    o.id, o.created_at, o.date::text, o.department, o.requester_fio,
                    o.requester_phone, o.comment,
                    COALESCE(obj.address, '') AS object_address,
                    COALESCE(obj.excel_id, '') AS object_id
                FROM transport_orders o
                LEFT JOIN objects obj ON o.object_id = obj.id
                WHERE o.id = %s
                """,
                (order_id,),
            )
            order_header = cur.fetchone()
            if not order_header:
                raise ValueError(f"Заявка на транспорт с ID={order_id} не найдена")

            cur.execute(
                """
                SELECT
                    tech,
                    qty,
                    to_char(time, 'HH24:MI') AS time,
                    hours,
                    note,
                    COALESCE(status, 'Новая') AS status
                FROM transport_order_positions
                WHERE order_id = %s
                ORDER BY id
                """,
                (order_id,),
            )
            positions = cur.fetchall()

            return {
                "id": order_header["id"],
                "created_at": order_header["created_at"].strftime("%Y-%m-%dT%H:%M:%S"),
                "date": order_header["date"],
                "department": order_header["department"],
                "requester_fio": order_header["requester_fio"],
                "requester_phone": order_header["requester_phone"],
                "comment": order_header["comment"],
                "object": {
                    "id": order_header["object_id"],
                    "address": order_header["object_address"]
                },
                "positions": [dict(p) for p in positions]
            }
    finally:
        if conn:
            release_db_connection(conn)


def fetch_all_vehicles() -> List[Dict[str, Any]]:
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT id, type, name, plate, department, note
                FROM vehicles
                ORDER BY type, name, plate
                """
            )
            return [dict(r) for r in cur.fetchall()]
    finally:
        if conn:
            release_db_connection(conn)


def insert_vehicle(v_type: str, name: str, plate: str, department: str = "", note: str = "") -> int:
    v_type = safe_str(v_type)
    name = safe_str(name)
    plate = normalize_plate(plate)
    department = safe_str(department)
    note = safe_str(note)

    if not v_type or not name or not plate:
        raise ValueError("Тип, наименование и госномер обязательны")

    conn = None
    try:
        conn = get_db_connection()
        with conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT id
                    FROM vehicles
                    WHERE type = %s AND name = %s AND plate = %s
                    """,
                    (v_type, name, plate),
                )
                row = cur.fetchone()
                if row:
                    raise ValueError("Такое транспортное средство уже существует")

                cur.execute(
                    """
                    INSERT INTO vehicles (type, name, plate, department, note)
                    VALUES (%s, %s, %s, %s, %s)
                    RETURNING id
                    """,
                    (v_type, name, plate, department, note),
                )
                return cur.fetchone()[0]
    finally:
        if conn:
            release_db_connection(conn)


def delete_vehicle(vehicle_id: int) -> None:
    conn = None
    try:
        conn = get_db_connection()
        with conn:
            with conn.cursor() as cur:
                cur.execute("SELECT id, type, name, plate FROM vehicles WHERE id = %s", (vehicle_id,))
                row = cur.fetchone()
                if not row:
                    return

                _, v_type, name, plate = row
                assigned_value = " | ".join([safe_str(v_type), safe_str(name), safe_str(plate)]).strip(" |")

                cur.execute(
                    """
                    SELECT COUNT(*)
                    FROM transport_order_positions
                    WHERE COALESCE(assigned_vehicle, '') = %s
                    """,
                    (assigned_value,),
                )
                used_count = int(cur.fetchone()[0] or 0)
                if used_count > 0:
                    raise ValueError(
                        f"Нельзя удалить транспорт: он уже используется в назначениях ({used_count} записей)"
                    )

                cur.execute("DELETE FROM vehicles WHERE id = %s", (vehicle_id,))
    finally:
        if conn:
            release_db_connection(conn)


def bulk_insert_vehicles(rows: List[Tuple[str, str, str, str, str]]) -> Tuple[int, int]:
    """
    rows: [(type, name, plate, department, note), ...]
    """
    prepared = []
    skipped = 0
    seen = set()

    for row in rows:
        v_type, name, plate, dep, note = [safe_str(x) for x in row]
        plate = normalize_plate(plate)

        if not v_type or not name or not plate:
            skipped += 1
            continue

        key = (v_type, name, plate)
        if key in seen:
            skipped += 1
            continue
        seen.add(key)
        prepared.append((v_type, name, plate, dep, note))

    if not prepared:
        return 0, skipped

    conn = None
    try:
        conn = get_db_connection()
        with conn:
            with conn.cursor() as cur:
                cur.execute("SELECT type, name, plate FROM vehicles")
                existing = {(safe_str(r[0]), safe_str(r[1]), normalize_plate(r[2])) for r in cur.fetchall()}

                to_insert = [r for r in prepared if (r[0], r[1], r[2]) not in existing]
                skipped += len(prepared) - len(to_insert)

                if to_insert:
                    execute_batch(
                        cur,
                        """
                        INSERT INTO vehicles (type, name, plate, department, note)
                        VALUES (%s, %s, %s, %s, %s)
                        """,
                        to_insert,
                    )
                return len(to_insert), skipped
    finally:
        if conn:
            release_db_connection(conn)


def load_employees_for_transport() -> List[Dict[str, Any]]:
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT e.fio, e.tbn, e.position, d.name AS dep
                FROM employees e
                LEFT JOIN departments d ON d.id = e.department_id
                WHERE COALESCE(e.is_fired, FALSE) = FALSE
                ORDER BY e.fio
                """
            )
            return [
                {"fio": r[0] or "", "tbn": r[1] or "", "pos": r[2] or "", "dep": r[3] or ""}
                for r in cur.fetchall()
            ]
    finally:
        if conn:
            release_db_connection(conn)


def load_objects_for_transport() -> List[Tuple[str, str]]:
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT COALESCE(NULLIF(excel_id, ''), '') AS code, address
                FROM objects
                ORDER BY address
                """
            )
            return [(r[0] or "", r[1] or "") for r in cur.fetchall()]
    finally:
        if conn:
            release_db_connection(conn)


def load_vehicles_for_transport() -> List[Dict[str, Any]]:
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT type, name, plate, department AS dep, note
                FROM vehicles
                ORDER BY type, name, plate
                """
            )
            return [dict(r) for r in cur.fetchall()]
    finally:
        if conn:
            release_db_connection(conn)


def get_transport_orders_for_planning(
    filter_date: Optional[str],
    filter_department: Optional[str],
    filter_status: Optional[str]
) -> List[Dict[str, Any]]:
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            params = []
            where = []

            if filter_date:
                where.append("o.date = %s")
                params.append(filter_date)

            if filter_department and filter_department.lower() != "все":
                where.append("o.department = %s")
                params.append(filter_department)

            if filter_status and filter_status.lower() != "все":
                where.append("COALESCE(p.status, 'Новая') = %s")
                params.append(filter_status)

            where_sql = "WHERE " + " AND ".join(where) if where else ""

            sql = f"""
                SELECT
                    p.id,
                    to_char(o.created_at, 'YYYY-MM-DD"T"HH24:MI:SS') AS created_at,
                    o.date::text AS date,
                    COALESCE(o.department, '') AS department,
                    COALESCE(o.requester_fio, '') AS requester_fio,
                    COALESCE(obj.address, '') AS object_address,
                    COALESCE(obj.excel_id, '') AS object_id,
                    COALESCE(p.tech, '') AS tech,
                    COALESCE(p.qty, 0) AS qty,
                    COALESCE(to_char(p.time, 'HH24:MI'), '') AS time,
                    COALESCE(p.hours, 0) AS hours,
                    COALESCE(p.assigned_vehicle, '') AS assigned_vehicle,
                    COALESCE(p.driver, '') AS driver,
                    COALESCE(p.status, 'Новая') AS status,
                    COALESCE(o.comment, '') AS comment,
                    COALESCE(p.note, '') AS position_note
                FROM transport_order_positions p
                JOIN transport_orders o ON o.id = p.order_id
                LEFT JOIN objects obj ON obj.id = o.object_id
                {where_sql}
                ORDER BY o.date, o.created_at, p.id
            """
            cur.execute(sql, params)
            return [dict(r) for r in cur.fetchall()]
    finally:
        if conn:
            release_db_connection(conn)


def save_transport_assignments(assignments: List[Dict[str, Any]]) -> int:
    if not assignments:
        return 0

    payload = []
    for a in assignments:
        pos_id = a.get("id")
        if not pos_id:
            continue

        status = safe_str(a.get("status")) or "Новая"
        if status not in ORDER_STATUSES:
            status = "Новая"

        payload.append((
            safe_str(a.get("assigned_vehicle")),
            safe_str(a.get("driver")),
            status,
            pos_id,
        ))

    if not payload:
        return 0

    conn = None
    try:
        conn = get_db_connection()
        with conn:
            with conn.cursor() as cur:
                execute_batch(
                    cur,
                    """
                    UPDATE transport_order_positions
                    SET assigned_vehicle = %s,
                        driver = %s,
                        status = %s
                    WHERE id = %s
                    """,
                    payload,
                )
        return len(payload)
    finally:
        if conn:
            release_db_connection(conn)


# ========================= UI WIDGETS =========================

class AutoCompleteCombobox(ttk.Combobox):
    def __init__(self, master=None, **kw):
        super().__init__(master, **kw)
        self._all_values: List[str] = []
        self.bind("<KeyRelease>", self._on_keyrelease)
        self.bind("<Control-BackSpace>", self._clear_all)

    def set_completion_list(self, values: List[str]):
        self._all_values = list(values)
        self["values"] = self._all_values

    def _clear_all(self, _=None):
        self.delete(0, tk.END)
        self["values"] = self._all_values

    def _on_keyrelease(self, event):
        if event.keysym in ("Up", "Down", "Left", "Right", "Home", "End", "Return", "Escape", "Tab"):
            return

        typed = self.get().strip()
        if not typed:
            self["values"] = self._all_values
            return

        self["values"] = [x for x in self._all_values if typed.lower() in x.lower()]


class PositionRow:
    def __init__(self, parent, idx: int, tech_values: List[str], on_delete):
        self.parent = parent
        self.idx = idx
        self.on_delete = on_delete
        self.tech_values = tech_values

        self.frame = tk.Frame(parent)

        self.cmb_tech = ttk.Combobox(self.frame, values=tech_values, width=46)
        self.cmb_tech.grid(row=0, column=0, padx=2, pady=1, sticky="w")

        self.ent_qty = ttk.Entry(self.frame, width=6, justify="center")
        self.ent_qty.grid(row=0, column=1, padx=2)
        self.ent_qty.insert(0, "1")

        self.time_var = tk.StringVar()
        self.time_var.trace_add("write", self._on_time_changed)
        self._formatting_time = False
        self._format_timer = None

        self.ent_time = ttk.Entry(self.frame, width=8, justify="center", textvariable=self.time_var)
        self.ent_time.grid(row=0, column=2, padx=2)
        self.ent_time.bind("<FocusOut>", self._format_immediately)
        self.ent_time.bind("<Return>", self._format_immediately)

        self.ent_hours = ttk.Entry(self.frame, width=8, justify="center")
        self.ent_hours.grid(row=0, column=3, padx=2)
        self.ent_hours.insert(0, "4")

        self.ent_note = ttk.Entry(self.frame, width=34)
        self.ent_note.grid(row=0, column=4, padx=2, sticky="w")

        self.btn_del = ttk.Button(self.frame, text="Удалить", width=9, command=self._delete)
        self.btn_del.grid(row=0, column=5, padx=2)

        for i, minsize in enumerate([380, 50, 70, 70, 280, 80]):
            self.frame.grid_columnconfigure(i, minsize=minsize)

    def _on_time_changed(self, *_):
        if self._formatting_time:
            return
        if self._format_timer:
            self.ent_time.after_cancel(self._format_timer)
        self._format_timer = self.ent_time.after(500, self._do_format)

    def _format_immediately(self, event=None):
        if self._format_timer:
            self.ent_time.after_cancel(self._format_timer)
            self._format_timer = None
        self._do_format()
        return None

    def _do_format(self):
        if self._formatting_time:
            return
        current = self.time_var.get()
        formatted = self._auto_format_time_input(current)
        if formatted != current:
            self._formatting_time = True
            try:
                self.time_var.set(formatted)
                self.ent_time.icursor(tk.END)
            finally:
                self._formatting_time = False

    def _auto_format_time_input(self, raw: str) -> str:
        if not raw:
            return ""
        digits = "".join(c for c in raw if c.isdigit())
        if not digits:
            return ""

        if len(digits) == 1:
            hh = int(digits)
            return f"{hh:02d}:00"
        if len(digits) == 2:
            hh = min(int(digits), 23)
            return f"{hh:02d}:00"
        if len(digits) == 3:
            hh = int(digits[0])
            mm = min(int(digits[1:3]), 59)
            return f"{hh:02d}:{mm:02d}"

        hh = min(int(digits[:2]), 23)
        mm = min(int(digits[2:4]), 59)
        return f"{hh:02d}:{mm:02d}"

    def grid(self, row: int):
        self.frame.grid(row=row, column=0, sticky="w")

    def destroy(self):
        self.frame.destroy()

    def apply_zebra(self, row0: int):
        bg = "#f6f8fa" if (row0 % 2 == 1) else "#ffffff"
        self.frame.configure(bg=bg)

    def _delete(self):
        self.on_delete(self)

    def validate(self) -> bool:
        val = safe_str(self.cmb_tech.get())
        if not val:
            return False

        try:
            qty = int(safe_str(self.ent_qty.get()) or "0")
            if qty <= 0:
                return False
        except Exception:
            return False

        tstr = safe_str(self.ent_time.get())
        if not tstr or parse_time_str(tstr) is None:
            return False

        hv = parse_hours_value(self.ent_hours.get())
        if hv is None or hv <= 0:
            return False

        return True

    def get_dict(self) -> Dict[str, Any]:
        return {
            "tech": safe_str(self.cmb_tech.get()),
            "qty": int(safe_str(self.ent_qty.get()) or 0),
            "time": parse_time_str(self.ent_time.get()) or "",
            "hours": float(parse_hours_value(self.ent_hours.get()) or 0.0),
            "note": safe_str(self.ent_note.get()),
        }


class AddVehicleDialog(simpledialog.Dialog):
    def __init__(self, parent, title="Добавить транспортное средство"):
        self.result = None
        super().__init__(parent, title=title)

    def body(self, master):
        tk.Label(master, text="Тип:*").grid(row=0, column=0, sticky="e", padx=4, pady=4)
        tk.Label(master, text="Наименование:*").grid(row=1, column=0, sticky="e", padx=4, pady=4)
        tk.Label(master, text="Гос№:*").grid(row=2, column=0, sticky="e", padx=4, pady=4)
        tk.Label(master, text="Подразделение:").grid(row=3, column=0, sticky="e", padx=4, pady=4)
        tk.Label(master, text="Примечание:").grid(row=4, column=0, sticky="ne", padx=4, pady=4)

        self.ent_type = ttk.Entry(master, width=40)
        self.ent_type.grid(row=0, column=1, sticky="w", pady=4)

        self.ent_name = ttk.Entry(master, width=40)
        self.ent_name.grid(row=1, column=1, sticky="w", pady=4)

        self.ent_plate = ttk.Entry(master, width=20)
        self.ent_plate.grid(row=2, column=1, sticky="w", pady=4)

        self.ent_dep = ttk.Entry(master, width=40)
        self.ent_dep.grid(row=3, column=1, sticky="w", pady=4)

        self.txt_note = tk.Text(master, width=40, height=3)
        self.txt_note.grid(row=4, column=1, sticky="w", pady=4)

        return self.ent_type

    def validate(self):
        v_type = safe_str(self.ent_type.get())
        name = safe_str(self.ent_name.get())
        plate = normalize_plate(self.ent_plate.get())

        if not v_type or not name or not plate:
            messagebox.showwarning(
                "Добавление транспорта",
                "Поля Тип, Наименование и Гос№ обязательны.",
                parent=self,
            )
            return False
        return True

    def apply(self):
        self.result = {
            "type": safe_str(self.ent_type.get()),
            "name": safe_str(self.ent_name.get()),
            "plate": normalize_plate(self.ent_plate.get()),
            "department": safe_str(self.ent_dep.get()),
            "note": self.txt_note.get("1.0", "end").strip(),
        }


# ========================= SPECIAL ORDERS PAGE =========================

class SpecialOrdersPage(tk.Frame):
    def __init__(self, master, existing_data: dict = None, order_id: int = None, on_saved=None):
        super().__init__(master, bg="#f7f7f7")
        ensure_config()

        self.base_dir = exe_dir()
        self.edit_order_id = order_id
        self.on_saved = on_saved
        self.pos_rows: List[PositionRow] = []

        self._load_spr()
        self._build_ui()

        if existing_data:
            self._fill_from_existing(existing_data)
        else:
            self._update_fio_list()
            self._update_tomorrow_hint()
            self.add_position()

    def _load_spr(self):
        self.emps = load_employees_for_transport()
        self.objects = load_objects_for_transport()
        self.techs = load_vehicles_for_transport()

        tech_types = set()
        for v in self.techs:
            tp = safe_str(v.get("type"))
            if tp:
                tech_types.add(tp)
        self.tech_values = sorted(tech_types)

        self.deps = ["Все"] + sorted({safe_str(r["dep"]) for r in self.emps if safe_str(r["dep"])})

        self.addr_to_ids: Dict[str, List[str]] = {}
        for oid, addr in self.objects:
            addr = safe_str(addr)
            oid = safe_str(oid)
            if not addr:
                continue
            self.addr_to_ids.setdefault(addr, [])
            if oid and oid not in self.addr_to_ids[addr]:
                self.addr_to_ids[addr].append(oid)

        addresses_set = set(self.addr_to_ids.keys())
        addresses_set.update(addr for _, addr in self.objects if safe_str(addr))
        self.addresses = sorted(addresses_set)

    def _build_ui(self):
        top = tk.Frame(self, bg="#f7f7f7")
        top.pack(fill="x", padx=10, pady=8)

        tk.Label(top, text="Подразделение*:", bg="#f7f7f7").grid(row=0, column=0, sticky="w")
        self.cmb_dep = ttk.Combobox(top, state="readonly", values=self.deps, width=48)
        saved_dep = get_saved_dep()
        self.cmb_dep.set(saved_dep if saved_dep in self.deps else self.deps[0])
        self.cmb_dep.grid(row=0, column=1, sticky="w", padx=(4, 12))
        self.cmb_dep.bind("<<ComboboxSelected>>", lambda e: (set_saved_dep(self.cmb_dep.get()), self._update_fio_list()))

        tk.Label(top, text="ФИО*:", bg="#f7f7f7").grid(row=0, column=2, sticky="w")
        self.fio_var = tk.StringVar()
        self.cmb_fio = AutoCompleteCombobox(top, textvariable=self.fio_var, width=36)
        self.cmb_fio.grid(row=0, column=3, sticky="w", padx=(4, 12))

        tk.Label(top, text="Телефон*:", bg="#f7f7f7").grid(row=0, column=4, sticky="w")
        self.ent_phone = ttk.Entry(top, width=18)
        self.ent_phone.grid(row=0, column=5, sticky="w", padx=(4, 12))

        tk.Label(top, text="Дата*:", bg="#f7f7f7").grid(row=0, column=6, sticky="w")
        self.ent_date = ttk.Entry(top, width=12)
        self.ent_date.grid(row=0, column=7, sticky="w", padx=(4, 0))
        self.ent_date.insert(0, (date.today() + timedelta(days=1)).strftime("%Y-%m-%d"))
        self.ent_date.bind("<KeyRelease>", lambda e: self._update_tomorrow_hint())
        self.ent_date.bind("<FocusOut>", lambda e: self._update_tomorrow_hint())

        tk.Label(top, text="Адрес*:", bg="#f7f7f7").grid(row=1, column=0, sticky="w", pady=(8, 0))
        self.cmb_address = AutoCompleteCombobox(top, width=56)
        self.cmb_address.set_completion_list(self.addresses)
        self.cmb_address.grid(row=1, column=1, columnspan=3, sticky="w", padx=(4, 12), pady=(8, 0))
        self.cmb_address.bind("<<ComboboxSelected>>", lambda e: self._sync_ids_by_address())
        self.cmb_address.bind("<FocusOut>", lambda e: self._sync_ids_by_address())
        self.cmb_address.bind("<Return>", lambda e: self._sync_ids_by_address())

        tk.Label(top, text="ID объекта:", bg="#f7f7f7").grid(row=1, column=4, sticky="w", pady=(8, 0))
        self.cmb_object_id = ttk.Combobox(top, state="readonly", values=[], width=20)
        self.cmb_object_id.grid(row=1, column=5, sticky="w", padx=(4, 12), pady=(8, 0))

        self.lbl_date_hint = tk.Label(top, text="", fg="#555", bg="#f7f7f7")
        self.lbl_date_hint.grid(row=1, column=6, columnspan=2, sticky="w", pady=(8, 0))

        tk.Label(top, text="Комментарий*:", bg="#f7f7f7").grid(row=2, column=0, sticky="nw", pady=(8, 0))
        self.txt_comment = tk.Text(top, height=3, width=96)
        self.txt_comment.grid(row=2, column=1, columnspan=7, sticky="we", padx=(4, 0), pady=(8, 0))

        pos_wrap = tk.LabelFrame(self, text="Позиции")
        pos_wrap.pack(fill="both", expand=True, padx=10, pady=(6, 8))

        hdr = tk.Frame(pos_wrap)
        hdr.pack(fill="x")
        tk.Label(hdr, text="Техника*", width=52, anchor="w").grid(row=0, column=0, padx=2)
        tk.Label(hdr, text="Кол-во*", width=6, anchor="center").grid(row=0, column=1, padx=2)
        tk.Label(hdr, text="Подача (чч:мм)*", width=12, anchor="center").grid(row=0, column=2, padx=2)
        tk.Label(hdr, text="Часы*", width=10, anchor="center").grid(row=0, column=3, padx=2)
        tk.Label(hdr, text="Примечание", width=38, anchor="w").grid(row=0, column=4, padx=2)
        tk.Label(hdr, text="Действие", width=10, anchor="center").grid(row=0, column=5, padx=2)

        wrap = tk.Frame(pos_wrap)
        wrap.pack(fill="both", expand=True)

        self.cv = tk.Canvas(wrap, borderwidth=0, highlightthickness=0)
        self.rows_holder = tk.Frame(self.cv)
        self.cv.create_window((0, 0), window=self.rows_holder, anchor="nw")
        self.cv.pack(side="left", fill="both", expand=True)

        self.vscroll = ttk.Scrollbar(wrap, orient="vertical", command=self.cv.yview)
        self.vscroll.pack(side="right", fill="y")
        self.cv.configure(yscrollcommand=self.vscroll.set)

        self.rows_holder.bind("<Configure>", lambda e: self.cv.configure(scrollregion=self.cv.bbox("all")))
        self.cv.bind("<MouseWheel>", lambda e: (self.cv.yview_scroll(int(-1 * (e.delta / 120)), "units"), "break"))

        btns = tk.Frame(pos_wrap)
        btns.pack(fill="x")
        ttk.Button(btns, text="Добавить позицию", command=self.add_position).pack(side="left", padx=2, pady=4)

        bottom = tk.Frame(self)
        bottom.pack(fill="x", padx=10, pady=(0, 10))
        ttk.Button(bottom, text="Сохранить заявку", command=self.save_order).pack(side="left", padx=4)
        ttk.Button(bottom, text="Очистить форму", command=self.clear_form).pack(side="left", padx=4)

        self._update_fio_list()
        self._update_tomorrow_hint()

    def _fill_from_existing(self, data: dict):
        self.cmb_dep.set(data.get("department", "Все"))
        self._update_fio_list()
        self.fio_var.set(data.get("requester_fio", ""))
        self.ent_phone.delete(0, "end")
        self.ent_phone.insert(0, data.get("requester_phone", ""))
        self.ent_date.delete(0, "end")
        self.ent_date.insert(0, data.get("date", ""))
        self.txt_comment.delete("1.0", "end")
        self.txt_comment.insert("1.0", data.get("comment", ""))

        obj = data.get("object", {})
        self.cmb_address.set(obj.get("address", ""))
        self._sync_ids_by_address()
        if obj.get("id"):
            self.cmb_object_id.set(obj.get("id"))

        for row in self.pos_rows:
            row.destroy()
        self.pos_rows.clear()

        positions_data = data.get("positions", [])
        if not positions_data:
            self.add_position()
        else:
            for pos_data in positions_data:
                self.add_position()
                row = self.pos_rows[-1]
                row.cmb_tech.set(pos_data.get("tech", ""))
                row.ent_qty.delete(0, "end")
                row.ent_qty.insert(0, str(pos_data.get("qty", "1")))
                row.ent_time.delete(0, "end")
                row.ent_time.insert(0, pos_data.get("time", ""))
                row.ent_hours.delete(0, "end")
                row.ent_hours.insert(0, str(pos_data.get("hours", "4")))
                row.ent_note.delete(0, "end")
                row.ent_note.insert(0, pos_data.get("note", ""))

        self._update_tomorrow_hint()

    def _update_fio_list(self):
        dep = safe_str(self.cmb_dep.get()) or "Все"
        if dep == "Все":
            names = [r["fio"] for r in self.emps]
        else:
            names = [r["fio"] for r in self.emps if safe_str(r["dep"]) == dep]

        seen = set()
        filtered = []
        for n in names:
            if n not in seen:
                seen.add(n)
                filtered.append(n)

        if not filtered and dep != "Все":
            filtered = [r["fio"] for r in self.emps]

        self.cmb_fio.set_completion_list(filtered)

    def _update_tomorrow_hint(self):
        req = parse_date_any(self.ent_date.get())
        today = date.today()
        tomorrow = today + timedelta(days=1)

        if req is None:
            self.lbl_date_hint.config(
                text="Укажите дату в формате YYYY-MM-DD или DD.MM.YYYY",
                fg="#b00020"
            )
        elif req <= today:
            self.lbl_date_hint.config(
                text=f"Дата должна быть не ранее {tomorrow.strftime('%Y-%m-%d')}",
                fg="#b00020"
            )
        else:
            self.lbl_date_hint.config(
                text="Ок: заявка на будущую дату",
                fg="#2e7d32"
            )

    def _sync_ids_by_address(self):
        addr = safe_str(self.cmb_address.get())
        ids = sorted(self.addr_to_ids.get(addr, []))
        if ids:
            self.cmb_object_id.config(state="readonly", values=ids)
            if self.cmb_object_id.get() not in ids:
                self.cmb_object_id.set(ids[0])
        else:
            self.cmb_object_id.config(state="normal", values=[])
            self.cmb_object_id.set("")

    def add_position(self):
        row = PositionRow(self.rows_holder, len(self.pos_rows) + 1, self.tech_values, self.delete_position)
        row.grid(len(self.pos_rows))
        row.apply_zebra(len(self.pos_rows))
        self.pos_rows.append(row)

    def delete_position(self, prow: PositionRow):
        try:
            self.pos_rows.remove(prow)
        except Exception:
            pass
        prow.destroy()
        for i, r in enumerate(self.pos_rows):
            r.grid(i)
            r.apply_zebra(i)

    def _validate_form(self) -> bool:
        if not safe_str(self.cmb_dep.get()):
            messagebox.showwarning("Заявка", "Выберите подразделение.", parent=self)
            return False

        if not safe_str(self.cmb_fio.get()):
            messagebox.showwarning("Заявка", "Укажите ФИО.", parent=self)
            return False

        phone = safe_str(self.ent_phone.get())
        if not validate_phone(phone):
            messagebox.showwarning("Заявка", "Укажите корректный номер телефона (минимум 5 цифр).", parent=self)
            return False

        req = parse_date_any(self.ent_date.get())
        tomorrow = date.today() + timedelta(days=1)
        if req is None or req <= date.today():
            messagebox.showwarning(
                "Заявка",
                f"Заявка возможна на даты, начиная с {tomorrow.strftime('%Y-%m-%d')}.",
                parent=self,
            )
            return False

        addr = safe_str(self.cmb_address.get())
        if not addr:
            messagebox.showwarning("Заявка", "Укажите адрес.", parent=self)
            return False

        comment = self.txt_comment.get("1.0", "end").strip()
        if not comment:
            messagebox.showwarning("Заявка", "Добавьте комментарий к заявке.", parent=self)
            return False

        if not self.pos_rows:
            messagebox.showwarning("Заявка", "Добавьте хотя бы одну позицию.", parent=self)
            return False

        for idx, r in enumerate(self.pos_rows, start=1):
            if not r.validate():
                messagebox.showwarning(
                    "Заявка",
                    f"Исправьте данные в позиции №{idx}.",
                    parent=self
                )
                return False

        return True

    def _build_order_dict(self) -> Dict[str, Any]:
        created_at = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
        req_date = parse_date_any(self.ent_date.get()) or date.today()
        addr = safe_str(self.cmb_address.get())
        oid = safe_str(self.cmb_object_id.get())
        comment = self.txt_comment.get("1.0", "end").strip()
        positions = [r.get_dict() for r in self.pos_rows]

        user_id = None
        app_ref = getattr(self, "app_ref", None)
        if app_ref is not None and hasattr(app_ref, "current_user"):
            try:
                user_id = int((app_ref.current_user or {}).get("id") or 0) or None
            except (ValueError, TypeError):
                user_id = None

        data = {
            "created_at": created_at,
            "date": req_date.strftime("%Y-%m-%d"),
            "department": safe_str(self.cmb_dep.get()),
            "requester_fio": safe_str(self.cmb_fio.get()),
            "requester_phone": safe_str(self.ent_phone.get()),
            "object": {"id": oid, "address": addr},
            "comment": comment,
            "positions": positions,
        }
        if user_id:
            data["user_id"] = user_id
        return data

    def save_order(self):
        if not self._validate_form():
            return

        data = self._build_order_dict()
        try:
            order_db_id = save_transport_order_to_db(data, edit_order_id=self.edit_order_id)
        except Exception as e:
            logger.exception("Ошибка сохранения заявки")
            messagebox.showerror("Сохранение", f"Не удалось сохранить заявку:\n{e}", parent=self)
            return

        messagebox.showinfo(
            "Сохранение",
            f"Заявка {'обновлена' if self.edit_order_id else 'сохранена'}.\nID: {order_db_id}",
            parent=self,
        )

        if self.on_saved:
            self.on_saved()
            if self.edit_order_id:
                self.winfo_toplevel().destroy()

    def clear_form(self):
        self.fio_var.set("")
        self.ent_phone.delete(0, "end")
        self.ent_date.delete(0, "end")
        self.ent_date.insert(0, (date.today() + timedelta(days=1)).strftime("%Y-%m-%d"))
        self.cmb_address.set("")
        self.cmb_object_id.config(values=[])
        self.cmb_object_id.set("")
        self.txt_comment.delete("1.0", "end")

        for r in self.pos_rows:
            r.destroy()
        self.pos_rows.clear()

        self.add_position()
        self._update_tomorrow_hint()


# ========================= TRANSPORT PLANNING =========================

class TransportPlanningPage(tk.Frame):
    def __init__(self, master):
        super().__init__(master, bg="#f7f7f7")
        self.spr_path = get_spr_path()
        self.authenticated = False
        self.row_meta: Dict[str, Dict[str, str]] = {}

        self._load_spr()
        self._build_ui()
        self.load_orders(silent=True)

    def _load_spr(self):
        self.vehicles = load_vehicles_for_transport()
        self.vehicle_types = sorted(
            {(safe_str(v.get("type"))) for v in self.vehicles if safe_str(v.get("type"))}
        )

        employees_raw = load_employees_for_transport()
        cfg = read_config()
        driver_depts_str = cfg.get(CONFIG_SECTION_INTEGR, KEY_DRIVER_DEPARTMENTS, fallback="Служба гаража")
        driver_departments = [d.strip() for d in driver_depts_str.split(",") if d.strip()]

        self.drivers = [e for e in employees_raw if safe_str(e.get("dep")) in driver_departments]
        self.drivers.sort(key=lambda x: x["fio"])

        self.departments = ["Все"] + sorted({safe_str(e.get("dep")) for e in employees_raw if safe_str(e.get("dep"))})

    def _build_ui(self):
        top = tk.Frame(self, bg="#f7f7f7")
        top.pack(fill="x", padx=10, pady=8)

        tk.Label(top, text="Дата:", bg="#f7f7f7").grid(row=0, column=0, sticky="w")
        self.ent_filter_date = ttk.Entry(top, width=12)
        self.ent_filter_date.grid(row=0, column=1, padx=4)
        self.ent_filter_date.insert(0, date.today().strftime("%Y-%m-%d"))

        tk.Label(top, text="Подразделение:", bg="#f7f7f7").grid(row=0, column=2, sticky="w", padx=(12, 0))
        self.cmb_filter_dep = ttk.Combobox(top, state="readonly", values=self.departments, width=20)
        self.cmb_filter_dep.set("Все")
        self.cmb_filter_dep.grid(row=0, column=3, padx=4)

        tk.Label(top, text="Статус:", bg="#f7f7f7").grid(row=0, column=4, sticky="w", padx=(12, 0))
        self.cmb_filter_status = ttk.Combobox(top, state="readonly", values=["Все"] + ORDER_STATUSES, width=15)
        self.cmb_filter_status.set("Все")
        self.cmb_filter_status.grid(row=0, column=5, padx=4)

        ttk.Button(top, text="Обновить", command=self.load_orders).grid(row=0, column=6, padx=12)
        ttk.Button(top, text="Сохранить назначения", command=self.save_assignments).grid(row=0, column=7, padx=4)

        table_frame = tk.Frame(self)
        table_frame.pack(fill="both", expand=True, padx=10, pady=8)

        columns = (
            "id", "created", "date", "dept", "requester",
            "object", "tech", "qty", "time", "hours",
            "assigned_vehicle", "driver", "status"
        )

        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=20)

        headers = {
            "id": "ID",
            "created": "Создано",
            "date": "Дата",
            "dept": "Подразделение",
            "requester": "Заявитель",
            "object": "Объект/Адрес",
            "tech": "Техника",
            "qty": "Кол-во",
            "time": "Подача",
            "hours": "Часы",
            "assigned_vehicle": "Назначен авто",
            "driver": "Водитель",
            "status": "Статус",
        }

        widths = {
            "id": 80,
            "created": 130,
            "date": 90,
            "dept": 120,
            "requester": 150,
            "object": 220,
            "tech": 180,
            "qty": 50,
            "time": 60,
            "hours": 60,
            "assigned_vehicle": 220,
            "driver": 150,
            "status": 100,
        }

        for col in columns:
            self.tree.heading(col, text=headers.get(col, col))
            self.tree.column(col, width=widths.get(col, 100), anchor="w")

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        self.tree.bind("<Double-1>", self.on_row_double_click)

        self.tree.tag_configure("Новая", background="#fff3cd")
        self.tree.tag_configure("Назначена", background="#d1ecf1")
        self.tree.tag_configure("В работе", background="#d4edda")
        self.tree.tag_configure("Выполнена", background="#e2e3e5")
        self.tree.tag_configure("Отменена", background="#f8d7da")

    def load_orders(self, silent: bool = False):
        try:
            filter_date = safe_str(self.ent_filter_date.get())
            filter_dept = safe_str(self.cmb_filter_dep.get())
            filter_status = safe_str(self.cmb_filter_status.get())

            orders = get_transport_orders_for_planning(
                filter_date=filter_date or None,
                filter_department=filter_dept or None,
                filter_status=filter_status or None,
            )
            self._populate_tree(orders)
            if not silent:
                logger.info("Загружено заявок для планирования: %s", len(orders))
        except Exception as e:
            logger.exception("Ошибка загрузки заявок")
            messagebox.showerror("Ошибка", f"Не удалось загрузить заявки из БД:\n{e}", parent=self)

    def _check_vehicle_conflict(self, vehicle_full: str, req_date: str, req_time: str, req_hours: Any, current_id: str) -> List[Dict]:
        if not vehicle_full or not req_date:
            return []

        current_interval = make_interval(req_date, req_time, req_hours)
        conflicts = []

        for item_id in self.tree.get_children():
            values = self.tree.item(item_id)["values"]
            other_id = str(values[0])

            if other_id == str(current_id):
                continue

            other_date = safe_str(values[2])
            other_vehicle = safe_str(values[10])
            other_time = safe_str(values[8])
            other_hours = values[9]
            other_requester = safe_str(values[4])
            other_object = safe_str(values[5])
            other_status = safe_str(values[12])

            if other_vehicle != vehicle_full:
                continue
            if other_date != req_date:
                continue
            if other_status in ["Выполнена", "Отменена"]:
                continue

            if current_interval is None:
                conflicts.append({
                    "time": other_time or "не указано",
                    "requester": other_requester,
                    "object": other_object,
                    "status": other_status
                })
                continue

            other_interval = make_interval(other_date, other_time, other_hours)
            if other_interval is None:
                conflicts.append({
                    "time": other_time or "не указано",
                    "requester": other_requester,
                    "object": other_object,
                    "status": other_status
                })
                continue

            if intervals_intersect(current_interval[0], current_interval[1], other_interval[0], other_interval[1]):
                conflicts.append({
                    "time": other_time or "не указано",
                    "requester": other_requester,
                    "object": other_object,
                    "status": other_status
                })

        return conflicts

    def _populate_tree(self, orders: List[Dict]):
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.row_meta = {}

        for order in orders:
            obj_display = order.get("object_address", "") or order.get("object_id", "")
            status = order.get("status", "Новая")

            item_id = self.tree.insert(
                "",
                "end",
                values=(
                    order.get("id", ""),
                    order.get("created_at", ""),
                    order.get("date", ""),
                    order.get("department", ""),
                    order.get("requester_fio", ""),
                    obj_display,
                    order.get("tech", ""),
                    order.get("qty", ""),
                    order.get("time", ""),
                    order.get("hours", ""),
                    order.get("assigned_vehicle", ""),
                    order.get("driver", ""),
                    status
                ),
                tags=(status,),
            )

            self.row_meta[item_id] = {
                "comment": order.get("comment") or "",
                "note": order.get("position_note") or "",
            }

    def on_row_double_click(self, event):
        selection = self.tree.selection()
        if not selection:
            return
        item = self.tree.item(selection[0])
        values = item["values"]
        self._show_assignment_dialog(selection[0], values)

    def _show_assignment_dialog(self, item_id, values):
        dialog = tk.Toplevel(self)
        dialog.title("Назначение транспорта")
        dialog.geometry("640x700")
        dialog.resizable(True, True)
        dialog.transient(self)
        dialog.grab_set()

        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (640 // 2)
        y = (dialog.winfo_screenheight() // 2) - (700 // 2)
        dialog.geometry(f"640x700+{x}+{y}")

        scroll_container = tk.Frame(dialog)
        scroll_container.pack(fill="both", expand=True, padx=0, pady=0)

        canvas = tk.Canvas(scroll_container, highlightthickness=0)
        scrollbar = ttk.Scrollbar(scroll_container, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)

        def update_scroll_region(event=None):
            canvas.configure(scrollregion=canvas.bbox("all"))

        scrollable_frame.bind("<Configure>", update_scroll_region)

        canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        def on_canvas_configure(event):
            canvas.itemconfig(canvas_window, width=event.width)

        canvas.bind("<Configure>", on_canvas_configure)

        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        def bind_mousewheel(event=None):
            canvas.bind_all("<MouseWheel>", _on_mousewheel)

        def unbind_mousewheel(event=None):
            canvas.unbind_all("<MouseWheel>")

        canvas.bind("<Enter>", bind_mousewheel)
        canvas.bind("<Leave>", unbind_mousewheel)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        info_frame = tk.LabelFrame(scrollable_frame, text="Информация о заявке", padx=12, pady=10)
        info_frame.pack(fill="x", padx=15, pady=10)

        info_data = [
            ("Дата:", values[2]),
            ("Время подачи:", values[8] or "не указано"),
            ("Заявитель:", values[4]),
            ("Объект:", values[5]),
        ]

        for label, value in info_data:
            row = tk.Frame(info_frame)
            row.pack(fill="x", pady=2)
            tk.Label(row, text=label, width=15, anchor="w").pack(side="left")
            tk.Label(row, text=value, anchor="w").pack(side="left", fill="x", expand=True)

        tech_frame = tk.Frame(info_frame, bg="#e3f2fd", relief="solid", borderwidth=1)
        tech_frame.pack(fill="x", pady=(8, 2), padx=5)
        tk.Label(
            tech_frame,
            text=f"Техника: {values[6]} x {values[7]} ({values[9]} ч.)",
            font=("Arial", 10, "bold"),
            fg="#0066cc",
            bg="#e3f2fd",
            padx=8,
            pady=8
        ).pack(anchor="w")

        meta = self.row_meta.get(item_id, {})
        order_comment = safe_str(meta.get("comment"))
        position_note = safe_str(meta.get("note"))

        texts_frame = tk.LabelFrame(scrollable_frame, text="Тексты заявки", padx=12, pady=10)
        texts_frame.pack(fill="x", padx=15, pady=(0, 8))

        row_c = tk.Frame(texts_frame)
        row_c.pack(fill="x", pady=2)
        tk.Label(row_c, text="Комментарий:", width=15, anchor="w").pack(side="left")
        tk.Label(row_c, text=order_comment or "—", anchor="w", justify="left", wraplength=560).pack(side="left", fill="x", expand=True)

        row_n = tk.Frame(texts_frame)
        row_n.pack(fill="x", pady=2)
        tk.Label(row_n, text="Примечание:", width=15, anchor="w").pack(side="left")
        tk.Label(row_n, text=position_note or "—", anchor="w", justify="left", wraplength=560).pack(side="left", fill="x", expand=True)

        warning_frame = tk.Frame(scrollable_frame, bg="#fff3cd", relief="solid", borderwidth=1)
        warning_label = tk.Label(
            warning_frame,
            text="",
            bg="#fff3cd",
            fg="#856404",
            wraplength=580,
            justify="left"
        )
        warning_label.pack(padx=10, pady=8)

        assign_frame = tk.LabelFrame(scrollable_frame, text="Назначение транспорта", padx=15, pady=15)
        assign_frame.pack(fill="both", expand=True, padx=15, pady=5)

        current_assignment = safe_str(values[10])
        current_type = ""
        current_name = ""
        current_plate = ""

        if current_assignment and " | " in current_assignment:
            parts = current_assignment.split(" | ")
            current_type = parts[0].strip() if len(parts) > 0 else ""
            current_name = parts[1].strip() if len(parts) > 1 else ""
            current_plate = parts[2].strip() if len(parts) > 2 else ""
        elif current_assignment:
            current_type = current_assignment

        tk.Label(assign_frame, text="Тип техники:", font=("Arial", 9, "bold")).grid(row=0, column=0, sticky="w", pady=(5, 2))
        vehicle_type_var = tk.StringVar(value=current_type)
        cmb_vehicle_type = ttk.Combobox(assign_frame, textvariable=vehicle_type_var, values=self.vehicle_types, state="readonly", width=55)
        cmb_vehicle_type.grid(row=1, column=0, pady=(0, 12), sticky="we")

        tk.Label(assign_frame, text="Наименование:", font=("Arial", 9, "bold")).grid(row=2, column=0, sticky="w", pady=(5, 2))
        vehicle_name_var = tk.StringVar(value="")
        cmb_vehicle_name = ttk.Combobox(assign_frame, textvariable=vehicle_name_var, values=[], state="readonly", width=55)
        cmb_vehicle_name.grid(row=3, column=0, pady=(0, 12), sticky="we")

        tk.Label(assign_frame, text="Гос. номер:", font=("Arial", 9, "bold")).grid(row=4, column=0, sticky="w", pady=(5, 2))
        vehicle_plate_var = tk.StringVar(value="")
        cmb_vehicle_plate = ttk.Combobox(assign_frame, textvariable=vehicle_plate_var, values=[], state="readonly", width=55)
        cmb_vehicle_plate.grid(row=5, column=0, pady=(0, 12), sticky="we")

        selection_info = tk.Label(assign_frame, text="Выберите тип, затем наименование и гос. номер", fg="#666")
        selection_info.grid(row=6, column=0, sticky="w", pady=(0, 10))

        def get_full_vehicle_string() -> str:
            parts = []
            if vehicle_type_var.get():
                parts.append(vehicle_type_var.get())
            if vehicle_name_var.get():
                parts.append(vehicle_name_var.get())
            if vehicle_plate_var.get():
                parts.append(vehicle_plate_var.get())
            return " | ".join(parts) if parts else ""

        def update_names(*args):
            selected_type = vehicle_type_var.get()
            vehicle_name_var.set("")
            vehicle_plate_var.set("")

            if not selected_type:
                cmb_vehicle_name["values"] = []
                cmb_vehicle_plate["values"] = []
                cmb_vehicle_name.state(["disabled"])
                cmb_vehicle_plate.state(["disabled"])
                selection_info.config(text="Выберите тип техники", fg="#666")
                return

            names = sorted(set(
                safe_str(v["name"]) for v in self.vehicles
                if safe_str(v.get("type")) == selected_type and safe_str(v.get("name"))
            ))

            cmb_vehicle_name["values"] = names
            cmb_vehicle_name.state(["!disabled"])
            cmb_vehicle_plate["values"] = []
            cmb_vehicle_plate.state(["disabled"])

            if len(names) == 0:
                selection_info.config(text="Нет доступных наименований для этого типа", fg="#dc3545")
            elif len(names) == 1:
                vehicle_name_var.set(names[0])
            else:
                selection_info.config(text=f"Доступно наименований: {len(names)}", fg="#666")

        def update_plates(*args):
            selected_type = vehicle_type_var.get()
            selected_name = vehicle_name_var.get()
            vehicle_plate_var.set("")

            if not selected_type or not selected_name:
                cmb_vehicle_plate["values"] = []
                cmb_vehicle_plate.state(["disabled"])
                return

            plates = sorted(set(
                safe_str(v["plate"]) for v in self.vehicles
                if safe_str(v.get("type")) == selected_type
                and safe_str(v.get("name")) == selected_name
                and safe_str(v.get("plate"))
            ))

            cmb_vehicle_plate["values"] = plates
            cmb_vehicle_plate.state(["!disabled"])

            if len(plates) == 0:
                selection_info.config(text="Нет доступных гос. номеров", fg="#dc3545")
            elif len(plates) == 1:
                vehicle_plate_var.set(plates[0])
            else:
                selection_info.config(text=f"Доступно гос. номеров: {len(plates)}", fg="#666")

        driver_var = tk.StringVar(value=safe_str(values[11]))
        driver_display_list = []
        for d in self.drivers:
            display = safe_str(d["fio"])
            if safe_str(d.get("dep")):
                display += f" ({safe_str(d.get('dep'))})"
            driver_display_list.append(display)

        ttk.Separator(assign_frame, orient="horizontal").grid(row=7, column=0, sticky="ew", pady=15)

        tk.Label(assign_frame, text="Водитель:", font=("Arial", 9, "bold")).grid(row=8, column=0, sticky="w", pady=(5, 2))
        tk.Label(assign_frame, text=f"(доступно: {len(self.drivers)} чел.)", fg="#666").grid(row=8, column=0, sticky="e", pady=(5, 2))

        cmb_driver = ttk.Combobox(assign_frame, textvariable=driver_var, values=driver_display_list, width=55)
        cmb_driver.grid(row=9, column=0, pady=(0, 12), sticky="we")

        tk.Label(assign_frame, text="Статус:", font=("Arial", 9, "bold")).grid(row=10, column=0, sticky="w", pady=(5, 2))
        status_var = tk.StringVar(value=safe_str(values[12]) or "Новая")
        cmb_status = ttk.Combobox(assign_frame, textvariable=status_var, values=ORDER_STATUSES, state="readonly", width=55)
        cmb_status.grid(row=11, column=0, pady=(0, 15), sticky="we")

        assign_frame.grid_columnconfigure(0, weight=1)

        def on_vehicle_or_driver_change(*args):
            if get_full_vehicle_string() and status_var.get() == "Новая":
                status_var.set("Назначена")
            check_conflicts()

        def check_conflicts(*args):
            selected_vehicle = get_full_vehicle_string()
            if not selected_vehicle:
                warning_frame.pack_forget()
                return

            req_date = safe_str(values[2])
            req_time = safe_str(values[8])
            req_hours = values[9]
            current_id = safe_str(values[0])

            conflicts = self._check_vehicle_conflict(selected_vehicle, req_date, req_time, req_hours, current_id)
            if conflicts:
                warning_text = (
                    f"ВНИМАНИЕ! Автомобиль '{selected_vehicle}' уже назначен на {len(conflicts)} "
                    f"заявк(у/и) с пересечением по времени:\n\n"
                )
                for i, conf in enumerate(conflicts, 1):
                    warning_text += f"{i}. {conf['time']} — {conf['requester']} ({conf['object']}) [{conf['status']}]\n"
                warning_text += "\nПроверьте возможность выполнения заявок."
                warning_label.config(text=warning_text)
                warning_frame.pack(fill="x", padx=15, pady=(0, 5))
            else:
                warning_frame.pack_forget()

        vehicle_type_var.trace_add("write", update_names)
        vehicle_name_var.trace_add("write", update_plates)
        vehicle_plate_var.trace_add("write", on_vehicle_or_driver_change)
        driver_var.trace_add("write", on_vehicle_or_driver_change)

        button_container = tk.Frame(dialog, bg="#f0f0f0", relief="raised", borderwidth=1)
        button_container.pack(fill="x", side="bottom", padx=0, pady=0)

        def save_and_close():
            if not get_full_vehicle_string():
                messagebox.showwarning("Назначение", "Выберите транспорт!", parent=dialog)
                return

            driver_name = safe_str(driver_var.get())
            if " (" in driver_name:
                driver_name = driver_name.split(" (")[0].strip()

            new_values = list(values)
            new_values[10] = get_full_vehicle_string()
            new_values[11] = driver_name
            new_values[12] = safe_str(status_var.get()) or "Новая"
            self.tree.item(item_id, values=new_values, tags=(new_values[12],))

            unbind_mousewheel()
            dialog.destroy()

        def cancel_and_close():
            unbind_mousewheel()
            dialog.destroy()

        ttk.Button(button_container, text="Сохранить", command=save_and_close, width=20).pack(side="left", padx=15, pady=12)
        ttk.Button(button_container, text="Отмена", command=cancel_and_close, width=20).pack(side="left", padx=5, pady=12)

        dialog.update_idletasks()

        if current_type:
            vehicle_type_var.set(current_type)
            dialog.update_idletasks()
            if current_name:
                vehicle_name_var.set(current_name)
                dialog.update_idletasks()
                if current_plate:
                    vehicle_plate_var.set(current_plate)
                    dialog.update_idletasks()

        canvas.configure(scrollregion=canvas.bbox("all"))
        canvas.yview_moveto(0)

        cmb_vehicle_type.focus_set()
        dialog.bind("<Return>", lambda e: save_and_close())
        dialog.bind("<Escape>", lambda e: cancel_and_close())

        check_conflicts()

    def save_assignments(self):
        assignments = []
        for item in self.tree.get_children():
            values = self.tree.item(item)["values"]
            assignments.append({
                "id": values[0],
                "assigned_vehicle": values[10],
                "driver": values[11],
                "status": values[12],
            })

        if not assignments:
            messagebox.showwarning("Сохранение", "Нет данных для сохранения", parent=self)
            return

        try:
            updated = save_transport_assignments(assignments)
            messagebox.showinfo("Сохранение", f"Назначения успешно сохранены.\nОбновлено записей: {updated}", parent=self)
        except Exception as e:
            logger.exception("Ошибка сохранения назначений")
            messagebox.showerror("Ошибка", f"Ошибка сохранения в БД:\n{e}", parent=self)


# ========================= TRANSPORT REGISTRY =========================

class TransportRegistryPage(tk.Frame):
    def __init__(self, master):
        super().__init__(master, bg="#f7f7f7")

        top = tk.Frame(self, bg="#f7f7f7")
        top.pack(fill="x", padx=10, pady=8)

        ttk.Button(top, text="Добавить транспортное средство", command=self.add_vehicle).pack(side="left", padx=(0, 8))
        ttk.Button(top, text="Загрузить из Excel", command=self.import_from_excel).pack(side="left", padx=(0, 8))
        ttk.Button(top, text="Обновить", command=self.reload_data).pack(side="left", padx=(0, 8))

        table_frame = tk.Frame(self)
        table_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        columns = ("id", "type", "name", "plate", "department", "note")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=20)

        headers = {
            "id": "ID",
            "type": "Тип",
            "name": "Наименование",
            "plate": "Гос№",
            "department": "Подразделение",
            "note": "Примечание",
        }
        widths = {
            "id": 60,
            "type": 120,
            "name": 180,
            "plate": 100,
            "department": 160,
            "note": 260,
        }

        for col in columns:
            self.tree.heading(col, text=headers[col])
            self.tree.column(col, width=widths[col], anchor="w")

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        self.menu = tk.Menu(self, tearoff=0)
        self.menu.add_command(label="Удалить", command=self.delete_selected)

        self.tree.bind("<Button-3>", self._on_right_click)
        self.tree.bind("<Delete>", lambda e: self.delete_selected())

        self.reload_data()

    def reload_data(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        try:
            vehicles = fetch_all_vehicles()
        except Exception as e:
            messagebox.showerror("Реестр транспорта", f"Ошибка загрузки из БД:\n{e}", parent=self)
            return

        for v in vehicles:
            self.tree.insert(
                "",
                "end",
                values=(
                    v["id"],
                    v["type"],
                    v["name"],
                    v["plate"],
                    v.get("department", ""),
                    v.get("note", ""),
                ),
            )

    def add_vehicle(self):
        dlg = AddVehicleDialog(self)
        if not dlg.result:
            return

        data = dlg.result
        try:
            insert_vehicle(
                v_type=data["type"],
                name=data["name"],
                plate=data["plate"],
                department=data.get("department", ""),
                note=data.get("note", ""),
            )
            self.reload_data()
        except Exception as e:
            messagebox.showerror("Добавление транспорта", f"Ошибка записи в БД:\n{e}", parent=self)

    def delete_selected(self):
        sel = self.tree.selection()
        if not sel:
            return

        ids = []
        for item in sel:
            vals = self.tree.item(item)["values"]
            if vals:
                ids.append(int(vals[0]))

        if not ids:
            return

        if not messagebox.askyesno(
            "Удаление транспорта",
            f"Удалить выбранные транспортные средства ({len(ids)} шт.)?",
            parent=self,
        ):
            return

        try:
            for vid in ids:
                delete_vehicle(vid)
            self.reload_data()
        except Exception as e:
            messagebox.showerror("Удаление транспорта", f"Ошибка при удалении из БД:\n{e}", parent=self)

    def _on_right_click(self, event):
        row_id = self.tree.identify_row(event.y)
        if row_id:
            self.tree.selection_set(row_id)
            self.menu.tk_popup(event.x_root, event.y_root)

    def import_from_excel(self):
        path = filedialog.askopenfilename(
            parent=self,
            title="Выберите Excel-файл с транспортом",
            filetypes=[("Excel файлы", "*.xlsx *.xlsm *.xltx *.xltm"), ("Все файлы", "*.*")],
        )
        if not path:
            return

        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            messagebox.showerror("Импорт из Excel", f"Не удалось открыть файл:\n{e}", parent=self)
            return

        sheet_name = "Техника"
        if sheet_name not in wb.sheetnames:
            messagebox.showerror("Импорт из Excel", f"В файле нет листа '{sheet_name}'.", parent=self)
            return

        ws = wb[sheet_name]

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            v_type = safe_str(row[0]) if row and len(row) > 0 else ""
            name = safe_str(row[1]) if row and len(row) > 1 else ""
            plate = safe_str(row[2]) if row and len(row) > 2 else ""
            dep = safe_str(row[3]) if row and len(row) > 3 else ""
            note = safe_str(row[4]) if row and len(row) > 4 else ""

            if not v_type and not name and not plate:
                continue

            rows.append((v_type, name, plate, dep, note))

        try:
            added, skipped = bulk_insert_vehicles(rows)
            self.reload_data()
            messagebox.showinfo(
                "Импорт из Excel",
                f"Загружено записей: {added}\nПропущено: {skipped}",
                parent=self,
            )
        except Exception as e:
            messagebox.showerror("Импорт из Excel", f"Ошибка загрузки данных:\n{e}", parent=self)


# ========================= MY ORDERS PAGE =========================

class MyTransportOrdersPage(tk.Frame):
    def __init__(self, master, app_ref=None):
        super().__init__(master, bg="#f7f7f7")
        self.app_ref = app_ref
        self.tree = None
        self._orders: List[Dict[str, Any]] = []
        self._build_ui()
        self._load_data()

    def _get_current_user_id(self) -> Optional[int]:
        if self.app_ref and hasattr(self.app_ref, "current_user"):
            try:
                return int((self.app_ref.current_user or {}).get("id") or 0) or None
            except (ValueError, TypeError):
                return None
        return None

    def _build_ui(self):
        top = tk.Frame(self, bg="#f7f7f7")
        top.pack(fill="x", padx=8, pady=(8, 4))
        tk.Label(top, text="Мои заявки на транспорт", font=("Segoe UI", 12, "bold"), bg="#f7f7f7").pack(side="left")
        ttk.Button(top, text="Обновить", command=self._load_data).pack(side="right", padx=4)

        frame = tk.Frame(self, bg="#f7f7f7")
        frame.pack(fill="both", expand=True, padx=8, pady=(4, 8))

        cols = ("date", "object", "department", "requester", "count", "created_at")
        self.tree = ttk.Treeview(frame, columns=cols, show="headings", selectmode="browse")

        self.tree.heading("date", text="Дата")
        self.tree.column("date", width=90, anchor="center")
        self.tree.heading("object", text="Объект")
        self.tree.column("object", width=280)
        self.tree.heading("department", text="Подразделение")
        self.tree.column("department", width=180)
        self.tree.heading("requester", text="Заявитель")
        self.tree.column("requester", width=220)
        self.tree.heading("count", text="Позиций")
        self.tree.column("count", width=80, anchor="center")
        self.tree.heading("created_at", text="Создана")
        self.tree.column("created_at", width=140, anchor="center")

        vsb = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        self.tree.bind("<Double-1>", self._on_open)
        self.tree.bind("<Return>", self._on_open)

        bottom = tk.Frame(self, bg="#f7f7f7")
        bottom.pack(fill="x", padx=8, pady=(0, 8))
        tk.Label(
            bottom,
            text="Двойной щелчок или Enter — открыть для редактирования или копирования.",
            font=("Segoe UI", 9),
            fg="#555",
            bg="#f7f7f7"
        ).pack(side="left")

    def _load_data(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        self._orders.clear()

        user_id = self._get_current_user_id()
        if not user_id:
            messagebox.showwarning("Мои заявки", "Не удалось определить текущего пользователя.", parent=self)
            return

        try:
            self._orders = load_user_transport_orders(user_id)
        except Exception as e:
            messagebox.showerror("Мои заявки", f"Ошибка загрузки списка заявок:\n{e}", parent=self)
            return

        for o in self._orders:
            created_str = o["created_at"].strftime("%d.%m.%Y %H:%M") if isinstance(o.get("created_at"), datetime) else ""
            date_str = o["date"].strftime("%Y-%m-%d") if isinstance(o.get("date"), date) else str(o.get("date", ""))
            self.tree.insert(
                "",
                "end",
                iid=str(o["id"]),
                values=(
                    date_str,
                    o.get("object_address", ""),
                    o.get("department", ""),
                    o.get("requester_fio", ""),
                    o.get("positions_count", 0),
                    created_str
                )
            )

    def _get_selected_order_id(self) -> Optional[int]:
        sel = self.tree.selection()
        return int(sel[0]) if sel else None

    def _on_open(self, event=None):
        order_id = self._get_selected_order_id()
        if not order_id:
            return

        try:
            order_data = get_transport_order_with_positions_from_db(order_id)
        except Exception as e:
            messagebox.showerror("Мои заявки", f"Не удалось загрузить данные заявки ID={order_id}:\n{e}", parent=self)
            return

        choice = messagebox.askyesnocancel(
            "Открыть заявку",
            "Нажмите «Да» для РЕДАКТИРОВАНИЯ заявки.\n"
            "Нажмите «Нет» для СОЗДАНИЯ КОПИИ (на другой день).\n"
            "Отмена — закрыть.",
            parent=self
        )

        if choice is None:
            return

        if choice is False:
            try:
                old_date = datetime.strptime(order_data["date"], "%Y-%m-%d").date()
                order_data["date"] = (old_date + timedelta(days=1)).strftime("%Y-%m-%d")
            except Exception:
                pass
            edit_id = None
            title = f"Новая заявка на транспорт (копия #{order_id})"
        else:
            edit_id = order_id
            title = f"Редактирование заявки на транспорт #{order_id}"

        win = tk.Toplevel(self)
        win.title(title)
        win.geometry("1180x720")

        page = SpecialOrdersPage(
            win,
            existing_data=order_data,
            order_id=edit_id,
            on_saved=self._load_data
        )
        page.app_ref = self.app_ref
        page.pack(fill="both", expand=True)


# ========================= API =========================

def create_page(parent, app_ref=None) -> tk.Frame:
    ensure_config()
    try:
        page = SpecialOrdersPage(parent)
        page.app_ref = app_ref
        return page
    except Exception:
        logger.exception("Ошибка создания страницы заявок")
        messagebox.showerror("Заявка — ошибка", "Не удалось открыть страницу заявок.", parent=parent)
        return tk.Frame(parent)


def create_my_transport_orders_page(parent, app_ref=None) -> tk.Frame:
    ensure_config()
    try:
        return MyTransportOrdersPage(parent, app_ref=app_ref)
    except Exception:
        logger.exception("Ошибка создания страницы моих заявок")
        messagebox.showerror("Мои заявки (транспорт)", "Не удалось открыть страницу.", parent=parent)
        return tk.Frame(parent)


def create_planning_page(parent) -> tk.Frame:
    ensure_config()
    try:
        return TransportPlanningPage(parent)
    except Exception:
        logger.exception("Ошибка создания страницы планирования")
        messagebox.showerror("Планирование — ошибка", "Не удалось открыть страницу планирования.", parent=parent)
        return tk.Frame(parent)


def create_transport_registry_page(parent) -> tk.Frame:
    ensure_config()
    try:
        return TransportRegistryPage(parent)
    except Exception:
        logger.exception("Ошибка создания страницы реестра транспорта")
        messagebox.showerror("Реестр транспорта — ошибка", "Не удалось открыть реестр транспорта.", parent=parent)
        return tk.Frame(parent)


# ========================= STANDALONE =========================

class SpecialOrdersApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1180x720")
        self.resizable(True, True)
        page = SpecialOrdersPage(self)
        page.pack(fill="both", expand=True)

    def destroy(self):
        global db_connection_pool, USING_SHARED_POOL
        if not USING_SHARED_POOL and db_connection_pool:
            logger.info("Closing local DB connection pool for SpecialOrders...")
            db_connection_pool.closeall()
            db_connection_pool = None
        super().destroy()


def open_special_orders(parent=None):
    if parent is None:
        app = SpecialOrdersApp()
        app.mainloop()
        return app

    win = tk.Toplevel(parent)
    win.title(APP_TITLE)
    win.geometry("1180x720")
    page = SpecialOrdersPage(win)
    page.pack(fill="both", expand=True)
    return win


def safe_filename(s: str, maxlen: int = 60) -> str:
    if not s:
        return "NOID"
    s = re.sub(r'[<>:"/\\|?*\n\r\t]+', "_", str(s)).strip()
    s = re.sub(r"_+", "_", s)
    return s[:maxlen] if len(s) > maxlen else s


if __name__ == "__main__":
    ensure_config()
    try:
        conn = get_db_connection()
        release_db_connection(conn)
    except Exception as e:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("Критическая ошибка", f"Не удалось подключиться к базе данных:\n{e}")
        root.destroy()
        sys.exit(1)

    app = SpecialOrdersApp()
    app.mainloop()
