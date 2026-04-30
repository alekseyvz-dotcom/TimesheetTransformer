from __future__ import annotations

import logging
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Sequence, Tuple

import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from timesheet_common import (
    calc_row_totals,
    calc_rows_summary,
    format_summary_value,
    month_days,
    normalize_hours_list,
    normalize_spaces,
    normalize_tbn,
    parse_hours_value,
    safe_filename,
)
from timesheet_db import (
    find_fired_employees_in_timesheet,
    load_employees_from_db,
    load_objects_short_for_timesheet,
)
from timesheet_dialogs import (
    AutoCompleteCombobox,
    SelectEmployeesDialog,
    SelectObjectIdDialog,
)
from trip_period_dialog import TripPeriodDialog, EmployeeTripsDialog
from trip_timesheet_db import (
    find_duplicate_employees_for_trip_timesheet,
    find_trip_timesheet_header_id,
    load_trip_timesheet_rows_from_db,
    replace_trip_timesheet_rows,
    upsert_trip_timesheet_header,
)
from virtual_timesheet_grid import VirtualTimesheetGrid

logger = logging.getLogger(__name__)

MONTH_NAMES = {
    1: "Январь",
    2: "Февраль",
    3: "Март",
    4: "Апрель",
    5: "Май",
    6: "Июнь",
    7: "Июль",
    8: "Август",
    9: "Сентябрь",
    10: "Октябрь",
    11: "Ноябрь",
    12: "Декабрь",
}

UI = {
    "bg": "#edf1f5",
    "panel": "#f7f9fb",
    "panel2": "#eef3f8",
    "line": "#c9d3df",
    "accent": "#2f74c0",
    "warning": "#c97a20",
    "text": "#1f2937",
    "muted": "#5b6776",
    "white": "#ffffff",
    "btn_save_bg": "#2f74c0",
    "btn_save_fg": "#ffffff",
}

PRINT_TITLE_FILL = PatternFill("solid", fgColor="DCE6F1")
PRINT_META_FILL = PatternFill("solid", fgColor="EAF2F8")
PRINT_HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
PRINT_TOTAL_FILL = PatternFill("solid", fgColor="E2F0D9")

THIN_BLACK = Side(style="thin", color="000000")
MEDIUM_BLACK = Side(style="medium", color="000000")

BORDER_THIN = Border(left=THIN_BLACK, right=THIN_BLACK, top=THIN_BLACK, bottom=THIN_BLACK)
BORDER_EMPTY = Border()


def _excel_safe_value(value: Any) -> Any:
    return "" if value is None else value


def _apply_print_style(
    cell,
    *,
    bold: bool = False,
    size: int = 9,
    h: str = "center",
    v: str = "center",
    wrap: bool = True,
    border: Border = BORDER_THIN,
    fill=None,
):
    cell.font = Font(name="Segoe UI", size=size, bold=bold)
    cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    cell.border = border
    if fill is not None:
        cell.fill = fill


def _set_outer_medium_border(ws, row1: int, col1: int, row2: int, col2: int):
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            cell = ws.cell(r, c)
            cell.border = Border(
                left=MEDIUM_BLACK if c == col1 else cell.border.left,
                right=MEDIUM_BLACK if c == col2 else cell.border.right,
                top=MEDIUM_BLACK if r == row1 else cell.border.top,
                bottom=MEDIUM_BLACK if r == row2 else cell.border.bottom,
            )


def _setup_print_sheet_params(ws, *, last_col_letter: str, last_row: int, title_rows: str = "$1:$7"):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    ws.page_margins = PageMargins(
        left=0.25,
        right=0.25,
        top=0.4,
        bottom=0.45,
        header=0.2,
        footer=0.2,
    )

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "E8"
    ws.print_title_rows = title_rows
    ws.print_area = f"A1:{last_col_letter}{last_row}"

    ws.oddHeader.center.text = "&\"Segoe UI,Bold\"&12 Командировочный табель"
    ws.oddFooter.left.text = "&\"Segoe UI\"&8 Сформировано автоматически"
    ws.oddFooter.center.text = "&\"Segoe UI\"&8 Страница &[Page] из &N"
    ws.oddFooter.right.text = f"&\"Segoe UI\"&8 {datetime.now().strftime('%d.%m.%Y %H:%M')}"


def build_printable_trip_timesheet_sheet(
    ws,
    *,
    year: int,
    month: int,
    object_addr: str,
    object_id: str,
    rows: list[dict[str, Any]],
    prepared_by: str = "",
):
    days_in_month = month_days(year, month)
    month_ru = MONTH_NAMES.get(month, str(month))

    fixed_headers = ["№", "ФИО", "Таб. №", "Командировка"]
    day_headers = [str(i) for i in range(1, days_in_month + 1)]
    total_headers = ["Дни", "Часы"]
    headers = fixed_headers + day_headers + total_headers

    total_cols = len(headers)
    last_col_letter = get_column_letter(total_cols)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws["A1"]
    c.value = "КОМАНДИРОВОЧНЫЙ ТАБЕЛЬ"
    _apply_print_style(c, bold=True, size=14, h="center", border=BORDER_EMPTY, fill=PRINT_TITLE_FILL)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    c = ws["A2"]
    c.value = f"за {month_ru} {year} г."
    _apply_print_style(c, bold=True, size=11, h="center", border=BORDER_EMPTY)

    meta_split = min(12, total_cols)
    right_meta_start = meta_split + 1

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=meta_split)
    c = ws["A3"]
    c.value = f"Объект: {object_addr or '-'}"
    _apply_print_style(c, bold=True, h="left", fill=PRINT_META_FILL)

    if right_meta_start <= total_cols:
        ws.merge_cells(start_row=3, start_column=right_meta_start, end_row=3, end_column=total_cols)
        c = ws.cell(3, right_meta_start)
        c.value = f"ID объекта: {object_id or '-'}"
        _apply_print_style(c, bold=True, h="left", fill=PRINT_META_FILL)

    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=total_cols)
    c = ws["A4"]
    c.value = f"Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    _apply_print_style(c, h="left", fill=PRINT_META_FILL)

    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=total_cols)
    c = ws["A5"]
    c.value = "В таблице указываются часы по дням и период командировки по сотруднику."
    _apply_print_style(c, size=8, h="left", border=BORDER_EMPTY)

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 16

    header_row = 7
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(header_row, col_idx, title)
        _apply_print_style(cell, bold=True, fill=PRINT_HEADER_FILL)

    ws.row_dimensions[header_row].height = 30

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 11
    ws.column_dimensions["D"].width = 24 # Слегка сузили, чтобы перенос срабатывал красивее

    first_day_col = 5
    last_day_col = first_day_col + days_in_month - 1
    for col_idx in range(first_day_col, last_day_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 4.2

    totals_start_col = last_day_col + 1
    ws.column_dimensions[get_column_letter(totals_start_col)].width = 8
    ws.column_dimensions[get_column_letter(totals_start_col + 1)].width = 10

    current_row = header_row + 1
    normalized_rows: list[dict[str, Any]] = []

    for rec in rows:
        fio = normalize_spaces(rec.get("fio") or "")
        tbn = normalize_tbn(rec.get("tbn"))
        hours = normalize_hours_list(rec.get("hours"), year, month)
        totals = rec.get("_totals") or calc_row_totals(hours, year, month)
        
        # --- ОБРАБОТКА МАССИВА ПЕРИОДОВ ---
        periods = rec.get("trip_periods", [])
        if periods:
            # Сортируем периоды по дате начала
            sorted_periods = sorted(periods, key=lambda x: x["from"])
            # Форматируем: 01.01.2024 - 15.01.2024
            p_strs = [f"{p['from'].strftime('%d.%m.%Y')} - {p['to'].strftime('%d.%m.%Y')}" for p in sorted_periods]
            trip_period = "\n".join(p_strs) # Склеиваем через перенос строки
            lines_count = len(sorted_periods)
        else:
            trip_period = ""
            lines_count = 1

        normalized_rows.append(
            {
                "fio": fio,
                "tbn": tbn,
                "hours": hours,
                "_totals": totals,
                "trip_period": trip_period,
                "lines_count": lines_count, # Сохраняем кол-во строк для расчета высоты ячейки
            }
        )

    for idx, rec in enumerate(normalized_rows, start=1):
        fio = rec["fio"]
        tbn = rec["tbn"]
        hours = rec["hours"]
        totals = rec["_totals"]
        trip_period = rec["trip_period"]

        row_values = [
            idx,
            fio,
            tbn,
            trip_period,
            *[_excel_safe_value(v) for v in hours[:days_in_month]],
            _excel_safe_value(format_summary_value(totals.get("days"))),
            _excel_safe_value(format_summary_value(totals.get("hours"))),
        ]

        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(current_row, col_idx, value)
            if col_idx in (2, 4):
                # wrap=True включен по умолчанию в _apply_print_style
                _apply_print_style(cell, h="left") 
            else:
                _apply_print_style(cell, h="center")

        # --- ДИНАМИЧЕСКАЯ ВЫСОТА СТРОКИ ---
        # Если строк 1, высота 22. Если строк больше, добавляем по 14 пикселей на каждую строку
        calc_height = max(22, 14 * rec["lines_count"])
        ws.row_dimensions[current_row].height = calc_height
        
        current_row += 1

    summary = calc_rows_summary(normalized_rows, year, month)

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    total_cell = ws.cell(current_row, 1)
    total_cell.value = "ИТОГО"
    _apply_print_style(total_cell, bold=True, fill=PRINT_TOTAL_FILL)

    for col_idx in range(5, totals_start_col):
        cell = ws.cell(current_row, col_idx, "")
        _apply_print_style(cell, bold=True, fill=PRINT_TOTAL_FILL)

    summary_values = [
        format_summary_value(summary.get("days")),
        format_summary_value(summary.get("hours")),
    ]
    for offset, value in enumerate(summary_values):
        cell = ws.cell(current_row, totals_start_col + offset, value)
        _apply_print_style(cell, bold=True, fill=PRINT_TOTAL_FILL)

    table_last_row = current_row
    _set_outer_medium_border(ws, header_row, 1, table_last_row, total_cols)

    sign_row = table_last_row + 3
    left_sign_end = min(12, total_cols)
    right_sign_start = left_sign_end + 1

    ws.merge_cells(start_row=sign_row, start_column=1, end_row=sign_row, end_column=left_sign_end)
    c = ws.cell(sign_row, 1)
    c.value = f"Составил: {prepared_by or '__________________'}    Подпись: __________________"
    _apply_print_style(c, h="left", border=BORDER_EMPTY)

    if right_sign_start <= total_cols:
        ws.merge_cells(start_row=sign_row, start_column=right_sign_start, end_row=sign_row, end_column=total_cols)
        c = ws.cell(sign_row, right_sign_start)
        c.value = "Дата: __________________"
        _apply_print_style(c, h="left", border=BORDER_EMPTY)

    _setup_print_sheet_params(
        ws,
        last_col_letter=last_col_letter,
        last_row=sign_row,
        title_rows="$1:$7",
    )

class TripTimeFillDialog(simpledialog.Dialog):
    def __init__(self, parent, max_day: int, title: str = "Проставить время"):
        self.max_day = int(max_day)
        self.result: Optional[Dict[str, Any]] = None
        super().__init__(parent, title=title)

    def body(self, master):
        tk.Label(master, text=f"В текущем месяце дней: {self.max_day}").grid(
            row=0, column=0, columnspan=4, sticky="w", pady=(4, 6)
        )

        self.var_mode = tk.StringVar(value="single")
        ttk.Radiobutton(master, text="Один день", value="single", variable=self.var_mode).grid(
            row=1, column=0, columnspan=2, sticky="w", pady=(2, 2)
        )
        ttk.Radiobutton(master, text="Диапазон дней", value="range", variable=self.var_mode).grid(
            row=1, column=2, columnspan=2, sticky="w", pady=(2, 2)
        )

        tk.Label(master, text="День:").grid(row=2, column=0, sticky="e", padx=(0, 6), pady=(4, 2))
        self.spn_day = tk.Spinbox(master, from_=1, to=self.max_day, width=6)
        self.spn_day.grid(row=2, column=1, sticky="w", pady=(4, 2))
        self.spn_day.delete(0, "end")
        self.spn_day.insert(0, "1")

        tk.Label(master, text="С:").grid(row=3, column=0, sticky="e", padx=(0, 6), pady=(2, 2))
        self.spn_from = tk.Spinbox(master, from_=1, to=self.max_day, width=6)
        self.spn_from.grid(row=3, column=1, sticky="w", pady=(2, 2))
        self.spn_from.delete(0, "end")
        self.spn_from.insert(0, "1")

        tk.Label(master, text="По:").grid(row=3, column=2, sticky="e", padx=(10, 6), pady=(2, 2))
        self.spn_to = tk.Spinbox(master, from_=1, to=self.max_day, width=6)
        self.spn_to.grid(row=3, column=3, sticky="w", pady=(2, 2))
        self.spn_to.delete(0, "end")
        self.spn_to.insert(0, str(self.max_day))

        self.var_clear = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            master,
            text="Очистить выбранные дни",
            variable=self.var_clear,
            command=self._toggle_hours_state,
        ).grid(row=4, column=0, columnspan=4, sticky="w", pady=(8, 2))

        tk.Label(master, text="Часы:").grid(row=5, column=0, sticky="e", padx=(0, 6), pady=(6, 2))
        self.ent_hours = ttk.Entry(master, width=18)
        self.ent_hours.grid(row=5, column=1, sticky="w", pady=(6, 2))
        self.ent_hours.insert(0, "8")

        tk.Label(
            master,
            text="Форматы: 8 | 8,25 | 8.5 | 8:30 | 1/7",
            fg="#555555",
        ).grid(row=6, column=0, columnspan=4, sticky="w", pady=(6, 0))

        return self.ent_hours

    def _toggle_hours_state(self):
        state = "disabled" if self.var_clear.get() else "normal"
        self.ent_hours.configure(state=state)

    def validate(self):
        mode = self.var_mode.get()

        try:
            day_single = int(self.spn_day.get())
            day_from = int(self.spn_from.get())
            day_to = int(self.spn_to.get())
        except Exception:
            messagebox.showwarning("Проставить время", "Дни должны быть целыми числами.", parent=self)
            return False

        if not (1 <= day_single <= self.max_day):
            messagebox.showwarning(
                "Проставить время",
                f"День должен быть в диапазоне 1–{self.max_day}.",
                parent=self,
            )
            return False

        if not (1 <= day_from <= self.max_day and 1 <= day_to <= self.max_day):
            messagebox.showwarning(
                "Проставить время",
                f"Диапазон должен быть в пределах 1–{self.max_day}.",
                parent=self,
            )
            return False

        if mode == "range" and day_from > day_to:
            messagebox.showwarning(
                "Проставить время",
                "Начальный день диапазона не может быть больше конечного.",
                parent=self,
            )
            return False

        if mode == "single":
            self._day_from = day_single
            self._day_to = day_single
        else:
            self._day_from = day_from
            self._day_to = day_to

        if self.var_clear.get():
            self._value = None
            return True

        text = normalize_spaces(self.ent_hours.get())
        if not text:
            messagebox.showwarning("Проставить время", "Введите количество часов.", parent=self)
            return False

        parsed = parse_hours_value(text)
        if parsed is None or parsed < 0:
            messagebox.showwarning(
                "Проставить время",
                "Введите корректное значение часов.\nПримеры: 8, 8,25, 8.5, 8:30, 1/7",
                parent=self,
            )
            return False

        self._value = float(parsed)
        return True

    def apply(self):
        self.result = {
            "from": self._day_from,
            "to": self._day_to,
            "value": self._value,
        }

class CopyTripEmployeesFromMonthDialog(simpledialog.Dialog):
    def __init__(
        self,
        parent,
        *,
        current_year: int,
        current_month: int,
        title: str = "Копировать сотрудников из другого месяца",
    ):
        self.current_year = int(current_year)
        self.current_month = int(current_month)
        self.result: Optional[Dict[str, Any]] = None

        prev_year = self.current_year
        prev_month = self.current_month - 1
        if prev_month < 1:
            prev_month = 12
            prev_year -= 1

        self.default_year = prev_year
        self.default_month = prev_month

        super().__init__(parent, title=title)

    def body(self, master):
        tk.Label(
            master,
            text="Выберите месяц, из которого нужно взять список сотрудников:",
            anchor="w",
        ).grid(row=0, column=0, columnspan=2, sticky="w", pady=(4, 10))

        tk.Label(master, text="Месяц:").grid(row=1, column=0, sticky="e", padx=(0, 6), pady=4)

        self.cmb_month = ttk.Combobox(
            master,
            width=20,
            state="readonly",
            values=[f"{m:02d} — {MONTH_NAMES[m]}" for m in range(1, 13)],
        )
        self.cmb_month.grid(row=1, column=1, sticky="w", pady=4)
        self.cmb_month.current(max(0, self.default_month - 1))

        tk.Label(master, text="Год:").grid(row=2, column=0, sticky="e", padx=(0, 6), pady=4)

        current = date.today().year
        self.var_year = tk.IntVar(value=self.default_year)

        self.cmb_year = ttk.Combobox(
            master,
            width=10,
            state="readonly",
            textvariable=self.var_year,
            values=list(range(current - 5, current + 6)),
        )
        self.cmb_year.grid(row=2, column=1, sticky="w", pady=4)

        self.var_replace = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            master,
            text="Заменить текущий список сотрудников",
            variable=self.var_replace,
        ).grid(row=3, column=0, columnspan=2, sticky="w", pady=(10, 2))

        tk.Label(
            master,
            text=(
                "Будут скопированы только ФИО, табельный номер и график.\n"
                "Часы и периоды командировок не копируются."
            ),
            fg="#666666",
            justify="left",
        ).grid(row=4, column=0, columnspan=2, sticky="w", pady=(8, 0))

        return self.cmb_month

    def validate(self):
        month_index = self.cmb_month.current()
        if month_index < 0:
            messagebox.showwarning(
                "Копирование",
                "Выберите месяц.",
                parent=self,
            )
            return False

        try:
            year = int(self.var_year.get())
        except Exception:
            messagebox.showwarning(
                "Копирование",
                "Выберите корректный год.",
                parent=self,
            )
            return False

        month = month_index + 1

        if year == self.current_year and month == self.current_month:
            messagebox.showwarning(
                "Копирование",
                "Нельзя копировать сотрудников из текущего же месяца.",
                parent=self,
            )
            return False

        self._source_year = year
        self._source_month = month
        return True

    def apply(self):
        self.result = {
            "year": self._source_year,
            "month": self._source_month,
            "replace": bool(self.var_replace.get()),
        }

class TripTimesheetPage(tk.Frame):
    def __init__(self, master, app, *args, **kwargs):
        super().__init__(master, bg=UI["bg"], *args, **kwargs)
        self.app = app

        self.current_header_id: Optional[int] = None
        self.rows: List[Dict[str, Any]] = []

        today = date.today()
        self.var_year = tk.IntVar(value=today.year)
        self.var_month = tk.IntVar(value=today.month)

        self.var_status = tk.StringVar(value="Готово.")
        self.var_trip_info = tk.StringVar(value="")
        self.var_filter = tk.StringVar(value="")

        self.objects_full: List[Tuple[str, str, str]] = []
        self.address_options: List[str] = []

        self._building_ui = False
        self._dirty = False
        self._auto_save_job = None
        self._auto_save_delay_ms = 8000
        self._loaded_context: Dict[str, Any] = {}
        self._suppress_events = False

        self._build_ui()
        self._load_reference_data()
        self._bind_hotkeys()
        self._refresh_grid()

    def destroy(self):
        if self._auto_save_job is not None:
            try:
                self.after_cancel(self._auto_save_job)
            except Exception:
                pass
            self._auto_save_job = None
        super().destroy()

    # =========================================================
    # UI
    # =========================================================
    def _build_ui(self) -> None:
        self._building_ui = True

        self._build_header()
        self._build_top_form()
        self._build_toolbar()
        self._build_filter_bar()
        self._build_grid()
        self._build_bottom()

        self._building_ui = False

    def _build_header(self) -> None:
        hdr = tk.Frame(self, bg=UI["accent"], pady=4)
        hdr.pack(fill="x")

        tk.Label(
            hdr,
            text="📋 Командировочный табель",
            font=("Segoe UI", 12, "bold"),
            bg=UI["accent"],
            fg="white",
            padx=12,
        ).pack(side="left")

        tk.Label(
            hdr,
            textvariable=self.var_status,
            font=("Segoe UI", 8),
            bg=UI["accent"],
            fg="#dbeafe",
            padx=10,
        ).pack(side="right")

    def _ts_lbl(self, parent, text: str, row: int, col: int = 0, required: bool = False, **grid_kw):
        display = f"{text}{'  *' if required else ''}:"
        fg = UI["warning"] if required else "#333333"
        tk.Label(
            parent,
            text=display,
            font=("Segoe UI", 9),
            bg=UI["panel"],
            fg=fg,
            anchor="e",
        ).grid(row=row, column=col, sticky="e", padx=(0, 6), pady=3, **grid_kw)

    def _build_top_form(self) -> None:
        outer = tk.Frame(self, bg=UI["bg"])
        outer.pack(fill="x", padx=10, pady=(6, 2))
        outer.grid_columnconfigure(0, weight=1)
        outer.grid_columnconfigure(1, weight=2)

        self._build_period_panel(outer)
        self._build_object_panel(outer)

    def _build_period_panel(self, parent) -> None:
        pnl = tk.LabelFrame(
            parent,
            text=" 📅 Период ",
            font=("Segoe UI", 9, "bold"),
            bg=UI["panel"],
            fg=UI["accent"],
            relief="groove",
            bd=1,
            padx=10,
            pady=8,
        )
        pnl.grid(row=0, column=0, sticky="nsew", padx=(0, 4), pady=2)

        self._ts_lbl(pnl, "Месяц", 0, required=True)
        self.cmb_month = ttk.Combobox(
            pnl,
            width=16,
            state="readonly",
            values=[f"{m:02d} — {MONTH_NAMES[m]}" for m in range(1, 13)],
        )
        self.cmb_month.grid(row=0, column=1, sticky="w", pady=3)
        self.cmb_month.bind("<<ComboboxSelected>>", lambda _e: self._on_month_combo_changed())
        self.cmb_month.current(max(0, self.var_month.get() - 1))

        tk.Label(
            pnl,
            text="Год  *:",
            font=("Segoe UI", 9),
            bg=UI["panel"],
            fg=UI["warning"],
            anchor="e",
        ).grid(row=0, column=2, sticky="e", padx=(12, 6), pady=3)

        self.cmb_year = ttk.Combobox(
            pnl,
            width=8,
            state="readonly",
            values=self._make_year_values(),
            textvariable=self.var_year,
        )
        self.cmb_year.grid(row=0, column=3, sticky="w", pady=3)
        self.cmb_year.bind("<<ComboboxSelected>>", lambda _e: self._on_period_changed())

    def _build_object_panel(self, parent) -> None:
        pnl = tk.LabelFrame(
            parent,
            text=" 📍 Объект ",
            font=("Segoe UI", 9, "bold"),
            bg=UI["panel"],
            fg=UI["accent"],
            relief="groove",
            bd=1,
            padx=10,
            pady=8,
        )
        pnl.grid(row=0, column=1, sticky="nsew", padx=(4, 0), pady=2)
        pnl.grid_columnconfigure(1, weight=1)

        self._ts_lbl(pnl, "Адрес объекта", 0, required=True)
        self.cmb_address = AutoCompleteCombobox(pnl, width=42, font=("Segoe UI", 9))
        self.cmb_address.grid(row=0, column=1, columnspan=3, sticky="ew", pady=3)
        self.cmb_address.bind("<<ComboboxSelected>>", lambda _e: self._on_address_select())
        self.cmb_address.bind("<Return>", lambda _e: self._on_address_select())

        self._ts_lbl(pnl, "ID объекта", 1)
        id_frame = tk.Frame(pnl, bg=UI["panel"])
        id_frame.grid(row=1, column=1, columnspan=3, sticky="ew", pady=3)

        self.cmb_object_id = ttk.Combobox(id_frame, state="readonly", values=[], width=22)
        self.cmb_object_id.pack(side="left")
        self.cmb_object_id.bind("<<ComboboxSelected>>", lambda _e: self._on_object_id_select())

        tk.Label(
            id_frame,
            text="← подставляется автоматически по адресу",
            font=("Segoe UI", 7),
            fg="#888888",
            bg=UI["panel"],
        ).pack(side="left", padx=8)

    def _ts_btn(self, parent, text: str, cmd, side="left", padx=3, pady=0, width=None):
        b = ttk.Button(parent, text=text, command=cmd, width=width)
        b.pack(side=side, padx=padx, pady=pady)
        return b

    def _build_toolbar(self) -> None:
        bar = tk.Frame(self, bg=UI["panel2"], relief="flat")
        bar.pack(fill="x", padx=10, pady=(4, 0))

        left = tk.Frame(bar, bg=UI["panel2"])
        left.pack(side="left", fill="x", expand=True)

        actions = tk.Frame(bar, bg=UI["panel2"])
        actions.pack(side="right", anchor="n", padx=(10, 4), pady=4)

        row1 = tk.Frame(left, bg=UI["panel2"])
        row1.pack(fill="x", pady=(4, 4))

        self._ts_btn(row1, "Открыть", self._open_timesheet, side="left", padx=(4, 3))
        self._ts_btn(row1, "Добавить сотрудников", self._add_employees, side="left", padx=3)
        self._ts_btn(row1, "Копировать из месяца", self._copy_employees_from_month, side="left", padx=3)
        self._ts_btn(row1, "Период выбранным", self._set_trip_period_for_selected, side="left", padx=3)
        self._ts_btn(row1, "Время выбранным", self._fill_hours_for_selected, side="left", padx=3)
        self._ts_btn(row1, "Время всем", self._fill_hours_for_all, side="left", padx=3)
        self._ts_btn(row1, "Удалить выбранных", self._delete_selected_rows, side="left", padx=3)
        self._ts_btn(row1, "Очистить часы", self._clear_hours_for_selected, side="left", padx=3)
        self._ts_btn(row1, "Проверить дубли", self._check_duplicates, side="left", padx=3)

        btn_save = tk.Button(
            actions,
            text="Сохранить",
            font=("Segoe UI", 9, "bold"),
            bg=UI["btn_save_bg"],
            fg=UI["btn_save_fg"],
            activebackground="#0d47a1",
            activeforeground="white",
            relief="flat",
            cursor="hand2",
            padx=14,
            pady=4,
            command=self._save_timesheet,
            width=14,
        )
        btn_save.pack(fill="x", pady=(0, 4))
        btn_save.bind("<Enter>", lambda _e: btn_save.config(bg="#0d47a1"))
        btn_save.bind("<Leave>", lambda _e: btn_save.config(bg=UI["btn_save_bg"]))

        self.btn_export = ttk.Button(
            actions,
            text="Выгрузить Excel",
            command=self._export_to_excel,
            width=16,
        )
        self.btn_export.pack(fill="x")

    def _build_filter_bar(self) -> None:
        bar = tk.Frame(self, bg=UI["bg"], pady=2)
        bar.pack(fill="x", padx=10, pady=(4, 0))

        tk.Label(
            bar,
            text="🔍 Поиск (ФИО / таб. №):",
            font=("Segoe UI", 9),
            bg=UI["bg"],
        ).pack(side="left")

        ent_filter = ttk.Entry(bar, textvariable=self.var_filter, width=36)
        ent_filter.pack(side="left", padx=(4, 8))
        ent_filter.bind("<KeyRelease>", lambda _e: self._apply_filter())

        ttk.Button(bar, text="Очистить", command=self._clear_filter).pack(side="left")

    def _build_grid(self) -> None:
        grid_wrap = tk.Frame(
            self,
            bg=UI["panel"],
            highlightbackground=UI["line"],
            highlightthickness=1,
            bd=0,
        )
        grid_wrap.pack(fill="both", expand=True, padx=10, pady=(4, 4))

        self.grid_widget = VirtualTimesheetGrid(
            grid_wrap,
            get_year_month=self._get_year_month,
            on_change=self._on_grid_change,
            on_delete_row=self._on_delete_row,
            on_selection_change=self._on_selection_change,
            on_trip_period_click=self._on_trip_period_click,
            show_trip_period=True,
            allow_row_select=True,
            read_only=False,
        )
        self.grid_widget.pack(fill="both", expand=True)

    def _build_bottom(self) -> None:
        bottom = tk.Frame(self, bg=UI["panel2"], pady=5)
        bottom.pack(fill="x", padx=10, pady=(0, 8))

        self.lbl_totals = tk.Label(
            bottom,
            text="Сотрудников: 0 | Дней: 0 | Часов: 0",
            font=("Segoe UI", 9, "bold"),
            fg=UI["accent"],
            bg=UI["panel2"],
        )
        self.lbl_totals.pack(side="left", padx=10)

        self.lbl_trip_info = tk.Label(
            bottom,
            textvariable=self.var_trip_info,
            font=("Segoe UI", 9),
            fg=UI["muted"],
            bg=UI["panel2"],
        )
        self.lbl_trip_info.pack(side="right", padx=10)

    def _make_year_values(self) -> List[int]:
        current = date.today().year
        return list(range(current - 3, current + 4))

    def _bind_hotkeys(self) -> None:
        self.bind_all("<Control-s>", self._hotkey_save, add="+")
        self.bind_all("<Control-S>", self._hotkey_save, add="+")
        self.bind_all("<Control-e>", self._hotkey_export, add="+")
        self.bind_all("<Control-E>", self._hotkey_export, add="+")
        self.bind_all("<F5>", self._hotkey_reload, add="+")

    def _hotkey_save(self, event=None):
        if not self.winfo_exists():
            return
        if self.focus_displayof() is None:
            return
        self._save_timesheet()
        return "break"

    def _hotkey_export(self, event=None):
        if not self.winfo_exists():
            return
        if self.focus_displayof() is None:
            return
        self._export_to_excel()
        return "break"

    def _hotkey_reload(self, event=None):
        if not self.winfo_exists():
            return
        if self.focus_displayof() is None:
            return
        self._open_timesheet()
        return "break"

    def _hotkey_delete(self, event=None):
        if not self.winfo_exists():
            return
        if self.focus_displayof() is None:
            return
    
        # Delete обрабатывается самим гридом и очищает только текущую ячейку.
        return "break"

    # =========================================================
    # Статус / dirty / автосохранение
    # =========================================================
    def _set_status_text(self, text: str) -> None:
        self.var_status.set(text)

    def _mark_dirty(self) -> None:
        self._dirty = True
        self._set_status_text("Есть несохранённые изменения")

    def _mark_saved(self, auto: bool = False) -> None:
        self._dirty = False
        now = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
        if auto:
            self._set_status_text(f"Последнее авто‑сохранение: {now}")
        else:
            self._set_status_text(f"Сохранено: {now}")

    def _mark_save_error(self, text: str) -> None:
        self._set_status_text(text)

    def _schedule_auto_save(self) -> None:
        if self._auto_save_job is not None:
            try:
                self.after_cancel(self._auto_save_job)
            except Exception:
                pass
            self._auto_save_job = None

        self._auto_save_job = self.after(self._auto_save_delay_ms, self._auto_save_callback)

    def _auto_save_callback(self) -> None:
        self._auto_save_job = None
        if not self._dirty:
            return
        self._save_timesheet_internal(show_messages=False, is_auto=True)

    def _capture_current_context(self) -> Dict[str, Any]:
        object_id, object_addr = self._parse_selected_object()
        year, month = self._get_year_month()
        return {
            "year": year,
            "month": month,
            "object_id": object_id,
            "object_addr": object_addr,
        }

    def _restore_controls_to_loaded_context(self) -> None:
        if not self._loaded_context:
            return

        self._suppress_events = True
        try:
            year = int(self._loaded_context.get("year") or date.today().year)
            month = int(self._loaded_context.get("month") or date.today().month)
            object_id = normalize_spaces(self._loaded_context.get("object_id") or "")
            object_addr = normalize_spaces(self._loaded_context.get("object_addr") or "")

            self.var_year.set(year)
            self.var_month.set(month)
            self.cmb_month.current(max(0, month - 1))

            self.cmb_address.set(object_addr)
            self._sync_object_id_values_silent()

            if object_id:
                values = list(self.cmb_object_id.cget("values") or [])
                if object_id not in values:
                    values.append(object_id)
                    self.cmb_object_id.config(values=values)
                self.cmb_object_id.set(object_id)
            else:
                self.cmb_object_id.set("")
        finally:
            self._suppress_events = False

    def _confirm_leave_with_unsaved(self) -> bool:
        if not self._dirty:
            return True

        answer = messagebox.askyesnocancel(
            "Несохранённые изменения",
            "Есть несохранённые изменения.\n\nСохранить перед переключением?",
            parent=self,
        )
        if answer is None:
            return False

        if answer is True:
            return self._save_timesheet_internal(show_messages=True, is_auto=False)

        return True

    # =========================================================
    # Справочники
    # =========================================================
    def _load_reference_data(self) -> None:
        self.objects_full = load_objects_short_for_timesheet()
        self.address_options = sorted(
            {
                normalize_spaces(addr)
                for _object_id, addr, _short_name in self.objects_full
                if normalize_spaces(addr)
            }
        )
        self.cmb_address.set_completion_list(self.address_options)

    def _parse_selected_object(self) -> Tuple[str, str]:
        object_addr = normalize_spaces(self.cmb_address.get() or "")
        object_id = normalize_spaces(self.cmb_object_id.get() or "")
        return object_id, object_addr

    def _sync_object_id_values_silent(self) -> None:
        addr = normalize_spaces(self.cmb_address.get() or "")
        objects_for_addr = [
            (code, a, short_name)
            for (code, a, short_name) in self.objects_full
            if normalize_spaces(a) == addr
        ]

        if not objects_for_addr:
            self.cmb_object_id.config(state="normal", values=[])
            self.cmb_object_id.set("")
            return

        ids = sorted({normalize_spaces(code) for code, _, _ in objects_for_addr if normalize_spaces(code)})
        cur = normalize_spaces(self.cmb_object_id.get() or "")

        self.cmb_object_id.config(state="readonly", values=ids)

        if cur and cur in ids:
            return
        if len(ids) == 1:
            self.cmb_object_id.set(ids[0])
        else:
            if cur not in ids:
                self.cmb_object_id.set("")

    def _auto_open_current_context(self) -> None:
        object_id, object_addr = self._parse_selected_object()
        if not object_addr:
            self._set_rows([])
            self.current_header_id = None
            self._loaded_context = self._capture_current_context()
            self.var_status.set("Выберите объект.")
            self._update_trip_info_from_selection()
            return
    
        self._open_timesheet()
    
    # =========================================================
    # Период / объект
    # =========================================================
    def _get_year_month(self) -> Tuple[int, int]:
        return int(self.var_year.get()), int(self.var_month.get())

    def _on_month_combo_changed(self) -> None:
        idx = self.cmb_month.current()
        if idx < 0:
            return
        self.var_month.set(idx + 1)
        self._on_period_changed()

    def _on_period_changed(self) -> None:
        if self._suppress_events:
            return
    
        if not self._confirm_leave_with_unsaved():
            self._restore_controls_to_loaded_context()
            return
    
        object_id, object_addr = self._parse_selected_object()
        if not object_addr:
            self._set_rows([])
            self.current_header_id = None
            self._loaded_context = self._capture_current_context()
            self.var_status.set("Выберите объект.")
            self._update_trip_info_from_selection()
            return
    
        self._open_timesheet()

    def _on_address_select(self) -> None:
        if self._suppress_events:
            return
    
        if not self._confirm_leave_with_unsaved():
            self._restore_controls_to_loaded_context()
            return
    
        self._sync_object_id_values_silent()
    
        addr = normalize_spaces(self.cmb_address.get() or "")
        objects_for_addr = [
            (normalize_spaces(code), normalize_spaces(a), normalize_spaces(short_name))
            for (code, a, short_name) in self.objects_full
            if normalize_spaces(a) == addr
        ]
    
        ids = sorted({code for code, _, _ in objects_for_addr if code})
        if len(ids) > 1:
            dlg = SelectObjectIdDialog(self, objects_for_addr, addr)
            self.wait_window(dlg)
    
            selected_id = normalize_spaces(dlg.result or "")
            if selected_id and selected_id in ids:
                self.cmb_object_id.set(selected_id)
    
        object_id, object_addr = self._parse_selected_object()
        if not object_addr:
            self._set_rows([])
            self.current_header_id = None
            self._loaded_context = self._capture_current_context()
            self.var_status.set("Выберите объект.")
            self._update_trip_info_from_selection()
            return
    
        self._open_timesheet()

    def _on_object_id_select(self) -> None:
        if self._suppress_events:
            return
    
        if not self._confirm_leave_with_unsaved():
            self._restore_controls_to_loaded_context()
            return
    
        object_id, object_addr = self._parse_selected_object()
        if not object_addr:
            self._set_rows([])
            self.current_header_id = None
            self._loaded_context = self._capture_current_context()
            self.var_status.set("Выберите объект.")
            self._update_trip_info_from_selection()
            return
    
        self._open_timesheet()

    # =========================================================
    # Работа со строками
    # =========================================================
    def _empty_row(self, fio: str = "", tbn: str = "") -> Dict[str, Any]:
        year, month = self._get_year_month()
        hours = normalize_hours_list([], year, month)
        totals = calc_row_totals(hours, year, month)

        return {
            "fio": normalize_spaces(fio),
            "tbn": normalize_tbn(tbn),
            "hours": hours,
            "trip_periods": [],  # НОВОЕ: Теперь это список периодов
            "_totals": totals,
            "work_schedule": "",
        }

    def _normalize_trip_row(self, rec: Dict[str, Any]) -> Dict[str, Any]:
        year, month = self._get_year_month()

        fio = normalize_spaces(rec.get("fio") or "")
        tbn = normalize_tbn(rec.get("tbn"))
        hours = normalize_hours_list(rec.get("hours"), year, month)

        # НОВОЕ: Поддержка старого формата и преобразование в список периодов
        periods = rec.get("trip_periods", [])
        if not periods and rec.get("trip_date_from") and rec.get("trip_date_to"):
            periods = [{"from": rec["trip_date_from"], "to": rec["trip_date_to"]}]

        totals = rec.get("_totals")
        if not isinstance(totals, dict):
            totals = calc_row_totals(hours, year, month)

        out = {
            "fio": fio,
            "tbn": tbn,
            "hours": hours,
            "trip_periods": periods,  # НОВОЕ: сохраняем массив
            "_totals": totals,
            "work_schedule": normalize_spaces(rec.get("work_schedule") or ""),
        }
        return out

    def _recalc_all_totals(self) -> None:
        year, month = self._get_year_month()
        for rec in self.rows:
            rec["hours"] = normalize_hours_list(rec.get("hours"), year, month)
            rec["_totals"] = calc_row_totals(rec["hours"], year, month)

    def _refresh_grid(self) -> None:
        self._recalc_all_totals()
        visible_rows = self._get_filtered_rows()
        self.grid_widget.set_rows(visible_rows)
        self._recalc_bottom_totals()

    def _set_rows(self, rows: Sequence[Dict[str, Any]]) -> None:
        self.rows = [self._normalize_trip_row(dict(r)) for r in rows]
        self._refresh_grid()

    def _get_filtered_rows(self) -> List[Dict[str, Any]]:
        q = normalize_spaces(self.var_filter.get() or "").lower()
        if not q:
            return self.rows

        filtered = []
        for rec in self.rows:
            fio = normalize_spaces(rec.get("fio") or "").lower()
            tbn = normalize_tbn(rec.get("tbn")).lower()
            if q in fio or q in tbn:
                filtered.append(rec)
        return filtered

    def _apply_filter(self) -> None:
        self._refresh_grid()
        self._update_trip_info_from_selection()

    def _clear_filter(self) -> None:
        self.var_filter.set("")
        self._apply_filter()

    def _recalc_bottom_totals(self) -> None:
        year, month = self._get_year_month()
        summary = calc_rows_summary(self.rows, year, month)
        txt = (
            f"Сотрудников: {summary['employees']}  |  "
            f"Дней: {summary['days']}  |  "
            f"Часов: {format_summary_value(summary['hours'])}"
        )
        self.lbl_totals.config(text=txt)

    # =========================================================
    # Открытие / загрузка / сохранение
    # =========================================================
    def _open_timesheet(self) -> None:
        object_id, object_addr = self._parse_selected_object()
        year, month = self._get_year_month()

        if not object_addr:
            messagebox.showwarning("Внимание", "Выберите объект.", parent=self)
            return

        if not self._confirm_leave_with_unsaved():
            self._restore_controls_to_loaded_context()
            return

        try:
            rows = load_trip_timesheet_rows_from_db(
                object_id=object_id or None,
                object_addr=object_addr,
                year=year,
                month=month,
            )
            self.current_header_id = find_trip_timesheet_header_id(
                object_id=object_id or None,
                object_addr=object_addr,
                year=year,
                month=month,
            )
        except Exception as exc:
            messagebox.showerror("Ошибка", f"Не удалось открыть табель:\n{exc}", parent=self)
            return

        self._set_rows(rows)
        self._loaded_context = self._capture_current_context()
        self._dirty = False
        self.var_status.set(f"Открыт командировочный табель: {object_addr}, {month:02d}.{year}.")
        self._update_trip_info_from_selection()

    def _save_timesheet(self) -> None:
        self._save_timesheet_internal(show_messages=True, is_auto=False)

    def _save_timesheet_internal(self, show_messages: bool = True, is_auto: bool = False) -> bool:
        object_id, object_addr = self._parse_selected_object()
        year, month = self._get_year_month()

        if not object_addr:
            if show_messages:
                messagebox.showwarning("Внимание", "Выберите объект.", parent=self)
            if is_auto:
                self._mark_save_error("Ошибка авто‑сохранения: не выбран объект")
            return False

        if not self.rows:
            if show_messages and not is_auto:
                if not messagebox.askyesno(
                    "Сохранение",
                    "В табеле нет строк. Всё равно создать/сохранить пустой табель?",
                    parent=self,
                ):
                    return False

        errors = self._validate_before_save()
        if errors:
            if show_messages:
                messagebox.showerror("Ошибка", "\n".join(errors), parent=self)
            if is_auto:
                self._mark_save_error("Ошибка авто‑сохранения: есть ошибки в данных")
            return False

        try:
            header_id = upsert_trip_timesheet_header(
                object_id=object_id or None,
                object_addr=object_addr,
                year=year,
                month=month,
            )

            self._recalc_all_totals()
            replace_trip_timesheet_rows(
                header_id=header_id,
                rows=self.rows,
                year=year,
                month=month,
            )

            self.current_header_id = header_id
            self._loaded_context = self._capture_current_context()
        except Exception as exc:
            if show_messages:
                messagebox.showerror("Ошибка", f"Не удалось сохранить табель:\n{exc}", parent=self)
            if is_auto:
                self._mark_save_error("Ошибка авто‑сохранения")
            return False

        self._mark_saved(auto=is_auto)
        self._update_trip_info_from_selection()
        return True

    def _validate_before_save(self) -> List[str]:
        errors: List[str] = []

        for i, rec in enumerate(self.rows, start=1):
            fio = normalize_spaces(rec.get("fio") or "")
            tbn = normalize_tbn(rec.get("tbn"))
            periods = rec.get("trip_periods") or []
            hours = rec.get("hours") or []

            if not fio and not tbn:
                errors.append(f"Строка {i}: не заполнены ФИО и табельный номер.")

            for p in periods:
                if p["from"] > p["to"]:
                    errors.append(f"Строка {i}: дата начала командировки ({p['from']}) позже даты окончания ({p['to']}).")

            has_hours = any(v is not None and str(v).strip() != "" for v in hours)
            if has_hours and not periods:
                errors.append(f"Строка {i}: есть часы, но не задан ни один период командировки.")

        return errors

    # =========================================================
    # Выбор сотрудников
    # =========================================================
    def _employee_key(self, fio: str, tbn: str) -> Tuple[str, str]:
        fio_norm = normalize_spaces(fio or "").lower()
        tbn_norm = normalize_tbn(tbn)

        if tbn_norm:
            return ("tbn", tbn_norm)

        return ("fio", fio_norm)

    def _load_source_rows_for_copy(
        self,
        *,
        object_id: str,
        object_addr: str,
        year: int,
        month: int,
    ) -> Tuple[List[Dict[str, Any]], str]:
        """
        Загружает строки табеля-источника для копирования сотрудников.

        Пробует несколько вариантов поиска, потому что старые табели могли быть
        сохранены без object_id или с другим object_id, но с тем же адресом.
        """
        object_id = normalize_spaces(object_id or "")
        object_addr = normalize_spaces(object_addr or "")

        attempts: List[Tuple[Optional[str], str, str]] = []
        seen = set()

        def add_attempt(oid: Optional[str], addr: str, label: str) -> None:
            key = (normalize_spaces(oid or ""), normalize_spaces(addr or ""))
            if key in seen:
                return
            seen.add(key)
            attempts.append((oid or None, addr, label))

        # 1. Основной вариант — текущий ID + адрес
        if object_id:
            add_attempt(
                object_id,
                object_addr,
                f"по ID объекта {object_id} и адресу",
            )

        # 2. Старые табели могли быть сохранены без ID объекта
        add_attempt(
            None,
            object_addr,
            "только по адресу объекта",
        )

        # 3. Если у этого адреса несколько ID, пробуем все
        for code, addr, _short_name in self.objects_full:
            code_norm = normalize_spaces(code or "")
            addr_norm = normalize_spaces(addr or "")

            if addr_norm == object_addr and code_norm:
                add_attempt(
                    code_norm,
                    object_addr,
                    f"по альтернативному ID объекта {code_norm}",
                )

        errors: List[str] = []

        for oid, addr, label in attempts:
            try:
                rows = load_trip_timesheet_rows_from_db(
                    object_id=oid,
                    object_addr=addr,
                    year=year,
                    month=month,
                )

                if rows:
                    return rows, label

            except Exception as exc:
                errors.append(f"{label}: {exc}")

        if errors:
            logger.warning(
                "Ошибки при поиске табеля-источника для копирования: %s",
                "; ".join(errors),
            )

        return [], ", ".join(label for _oid, _addr, label in attempts)

    def _copy_employees_from_month(self) -> None:
        object_id, object_addr = self._parse_selected_object()
        year, month = self._get_year_month()

        if not object_addr:
            messagebox.showwarning(
                "Копирование",
                "Сначала выберите объект.",
                parent=self,
            )
            return

        if self._dirty:
            if not messagebox.askyesno(
                "Копирование",
                (
                    "В текущем табеле есть несохранённые изменения.\n\n"
                    "Копирование изменит текущий табель.\n"
                    "Продолжить?"
                ),
                parent=self,
            ):
                return

        dlg = CopyTripEmployeesFromMonthDialog(
            self,
            current_year=year,
            current_month=month,
        )
        self.wait_window(dlg)

        params = getattr(dlg, "result", None)
        if not params:
            return

        source_year = int(params["year"])
        source_month = int(params["month"])
        replace_current = bool(params["replace"])

        try:
            source_rows, source_search_info = self._load_source_rows_for_copy(
                object_id=object_id,
                object_addr=object_addr,
                year=source_year,
                month=source_month,
            )
        except Exception as exc:
            messagebox.showerror(
                "Копирование",
                f"Не удалось загрузить табель-источник:\n{exc}",
                parent=self,
            )
            return

        if not source_rows:
            messagebox.showinfo(
                "Копирование",
                (
                    f"Не найдены сотрудники в табеле за {source_month:02d}.{source_year}.\n\n"
                    f"Объект:\n{object_addr}\n\n"
                    f"Проверялись варианты:\n{source_search_info}\n\n"
                    "Возможные причины:\n"
                    "1. Табель за этот месяц не был сохранён.\n"
                    "2. Табель за апрель создан в другом модуле табеля.\n"
                    "3. У апрельского табеля другой адрес объекта.\n"
                    "4. В базе старый табель сохранён с другим ID объекта."
                ),
                parent=self,
            )
            return

        source_label = f"{source_month:02d}.{source_year}"
        target_label = f"{month:02d}.{year}"

        if replace_current:
            if self.rows:
                if not messagebox.askyesno(
                    "Копирование",
                    (
                        f"Текущий список сотрудников за {target_label} будет очищен "
                        f"и заменён списком из {source_label}.\n\n"
                        "Продолжить?"
                    ),
                    parent=self,
                ):
                    return

            self.rows = []
            existing_keys = set()
        else:
            existing_keys = {
                self._employee_key(
                    normalize_spaces(r.get("fio") or ""),
                    normalize_tbn(r.get("tbn")),
                )
                for r in self.rows
            }

        added = 0
        skipped_duplicates = 0
        skipped_empty = 0

        for src in source_rows:
            fio = normalize_spaces(
                src.get("fio")
                or src.get("full_name")
                or src.get("employee_name")
                or ""
            )

            tbn = normalize_tbn(
                src.get("tbn")
                or src.get("tab_no")
                or src.get("tab_number")
                or src.get("personnel_number")
                or ""
            )

            if not fio and not tbn:
                skipped_empty += 1
                continue

            key = self._employee_key(fio, tbn)

            if key in existing_keys:
                skipped_duplicates += 1
                continue

            new_row = self._empty_row(fio=fio, tbn=tbn)

            new_row["work_schedule"] = normalize_spaces(src.get("work_schedule") or "")

            # ВАЖНО:
            # Часы и периоды командировок специально не копируем,
            # потому что они относятся к датам другого месяца.
            new_row["hours"] = normalize_hours_list([], year, month)
            new_row["trip_periods"] = []
            new_row["_totals"] = calc_row_totals(new_row["hours"], year, month)

            self.rows.append(new_row)
            existing_keys.add(key)
            added += 1

        self._refresh_grid()
        self._update_trip_info_from_selection()

        if added > 0 or replace_current:
            self._mark_dirty()
            self._schedule_auto_save()

        msg = (
            f"Копирование из {source_label} в {target_label} завершено.\n\n"
            f"Добавлено сотрудников: {added}\n"
            f"Пропущено дублей: {skipped_duplicates}"
        )

        if skipped_empty:
            msg += f"\nПропущено пустых строк: {skipped_empty}"

        messagebox.showinfo(
            "Копирование",
            msg,
            parent=self,
        )

        if added > 0:
            self.var_status.set(
                f"Скопировано сотрудников из {source_label}: {added}"
            )
            self._check_fired_employees_after_add()
            self._check_duplicates(silent_if_empty=True)
        else:
            self.var_status.set(
                f"Из {source_label} нечего добавить: все сотрудники уже есть."
            )

    def _add_employees(self) -> None:
        try:
            employees = load_employees_from_db()
        except Exception as exc:
            messagebox.showerror("Ошибка", f"Не удалось загрузить сотрудников:\n{exc}", parent=self)
            return

        existing_keys = {
            (normalize_spaces(r.get("fio") or "").lower(), normalize_tbn(r.get("tbn")))
            for r in self.rows
        }

        try:
            dlg = SelectEmployeesDialog(self, employees=employees, current_dep="Все")
            self.wait_window(dlg)
            selected = getattr(dlg, "result", None)
        except Exception as exc:
            messagebox.showerror("Ошибка", f"Не удалось открыть выбор сотрудников:\n{exc}", parent=self)
            return

        if not selected:
            return

        added = 0
        for item in selected:
            if len(item) >= 2:
                fio = normalize_spaces(item[0] or "")
                tbn = normalize_tbn(item[1])
            else:
                continue

            key = (fio.lower(), tbn)
            if key in existing_keys:
                continue

            row = self._empty_row(fio=fio, tbn=tbn)
            if len(item) >= 5:
                row["work_schedule"] = normalize_spaces(item[4] or "")

            self.rows.append(row)
            existing_keys.add(key)
            added += 1

        self._refresh_grid()
        if added > 0:
            self._mark_dirty()
            self._schedule_auto_save()
            self.var_status.set(f"Добавлено сотрудников: {added}")
        else:
            self.var_status.set("Все выбранные сотрудники уже есть в табеле.")

        self._check_fired_employees_after_add()
        self._check_duplicates(silent_if_empty=True)

    def _check_fired_employees_after_add(self) -> None:
        employees = []
        for rec in self.rows:
            fio = normalize_spaces(rec.get("fio") or "")
            tbn = normalize_tbn(rec.get("tbn"))
            if fio or tbn:
                employees.append((fio, tbn))

        if not employees:
            return

        year, month = self._get_year_month()

        try:
            fired = find_fired_employees_in_timesheet(
                employees=employees,
                year=year,
                month=month,
            )
        except Exception:
            return

        if not fired:
            return

        lines = ["В табеле обнаружены уволенные сотрудники:"]
        for item in fired[:20]:
            fio = item.get("fio") or ""
            tbn = item.get("tbn") or ""
            dismissal_date = item.get("dismissal_date")
            if dismissal_date:
                lines.append(f"• {fio} ({tbn}) — увольнение: {dismissal_date}")
            else:
                lines.append(f"• {fio} ({tbn}) — отмечен как уволенный")

        if len(fired) > 20:
            lines.append(f"... и ещё {len(fired) - 20}")

        messagebox.showwarning("Проверка сотрудников", "\n".join(lines), parent=self)

    # =========================================================
    # Дубли
    # =========================================================
    def _collect_employee_pairs(self) -> List[Tuple[str, str]]:
        out: List[Tuple[str, str]] = []
        for rec in self.rows:
            fio = normalize_spaces(rec.get("fio") or "")
            tbn = normalize_tbn(rec.get("tbn"))
            if fio or tbn:
                out.append((fio, tbn))
        return out

    def _check_duplicates(self, silent_if_empty: bool = False) -> None:
        object_id, object_addr = self._parse_selected_object()
        year, month = self._get_year_month()
    
        if not object_addr:
            if not silent_if_empty:
                messagebox.showwarning("Внимание", "Сначала выберите объект.", parent=self)
            return
    
        employees = self._collect_employee_pairs()
        if not employees:
            if not silent_if_empty:
                messagebox.showinfo("Проверка дублей", "В табеле нет сотрудников для проверки.", parent=self)
            return
    
        try:
            duplicates = find_duplicate_employees_for_trip_timesheet(
                object_id=object_id or None,
                object_addr=object_addr,
                year=year,
                month=month,
                employees=employees,
            )
        except Exception as exc:
            if not silent_if_empty:
                messagebox.showerror("Ошибка", f"Не удалось проверить дубли:\n{exc}", parent=self)
            return
    
        if not duplicates:
            if not silent_if_empty:
                messagebox.showinfo("Проверка дублей", "Дубликаты не найдены.", parent=self)
            return
    
        lines = ["В текущем табеле обнаружены дубли сотрудников:"]
        for item in duplicates[:30]:
            fio = item.get("fio") or ""
            tbn = item.get("tbn") or ""
            count = item.get("count")
    
            if count and count > 1:
                lines.append(f"• {fio} ({tbn}) — повторений: {count}")
            else:
                lines.append(f"• {fio} ({tbn})")
    
        if len(duplicates) > 30:
            lines.append(f"... и ещё {len(duplicates) - 30}")
    
        messagebox.showwarning("Найдены дубликаты", "\n".join(lines), parent=self)

    # =========================================================
    # События грида
    # =========================================================
    def _get_visible_rows(self) -> List[Dict[str, Any]]:
        return self._get_filtered_rows()

    def _visible_to_real_index(self, visible_index: int) -> Optional[int]:
        visible_rows = self._get_visible_rows()
        if not (0 <= visible_index < len(visible_rows)):
            return None

        rec = visible_rows[visible_index]
        for i, row in enumerate(self.rows):
            if row is rec:
                return i
        return None

    def _on_grid_change(self, row_index: int, col_index: int) -> None:
        real_index = self._visible_to_real_index(row_index)
        if real_index is None:
            return

        year, month = self._get_year_month()
        rec = self.rows[real_index]
        rec["hours"] = normalize_hours_list(rec.get("hours"), year, month)
        rec["_totals"] = calc_row_totals(rec["hours"], year, month)

        self._refresh_grid()
        self._update_trip_info_from_selection()
        self._mark_dirty()
        self._schedule_auto_save()

    def _on_delete_row(self, row_index: int) -> None:
        real_index = self._visible_to_real_index(row_index)
        if real_index is None:
            return

        rec = self.rows[real_index]
        fio = normalize_spaces(rec.get("fio") or "")
        tbn = normalize_tbn(rec.get("tbn"))

        if not messagebox.askyesno(
            "Удаление строки",
            f"Удалить строку сотрудника:\n{fio} ({tbn})?",
            parent=self,
        ):
            return

        del self.rows[real_index]
        self._refresh_grid()
        self._update_trip_info_from_selection()
        self._mark_dirty()
        self._schedule_auto_save()

    def _on_selection_change(self, selected_rows) -> None:
        self._update_trip_info_from_selection()

    def _on_trip_period_click(self, row_index: int) -> None:
        real_index = self._visible_to_real_index(row_index)
        if real_index is None:
            return

        rec = self.rows[real_index]
        year, month = self._get_year_month()

        # НОВОЕ: Вызываем новый менеджер периодов
        result = EmployeeTripsDialog.show(
            self,
            periods=rec.get("trip_periods", []),
            year=year,
            month=month,
        )
        if result is None:
            return

        rec["trip_periods"] = result

        self._refresh_grid()
        self._update_trip_info_from_selection()
        self._mark_dirty()
        self._schedule_auto_save()

    # =========================================================
    # Действия над выбранными строками
    # =========================================================
    def _get_selected_row_indexes(self) -> List[int]:
        try:
            selected = list(self.grid_widget.get_selected_indices())
        except Exception:
            selected = []
        return sorted(i for i in selected if 0 <= i < len(self._get_visible_rows()))

    def _ask_fill_hours_params(self) -> Optional[Dict[str, Any]]:
        year, month = self._get_year_month()
        max_day = month_days(year, month)
    
        dlg = TripTimeFillDialog(self, max_day=max_day, title="Проставить время")
        return dlg.result

    def _row_has_trip_period(self, rec: Dict[str, Any]) -> bool:
        return bool(rec.get("trip_periods"))
    
    def _fill_hours_for_row(
        self,
        rec: Dict[str, Any],
        value: Optional[float],
        day_from: int,
        day_to: int,
        year: int,
        month: int,
    ) -> None:
        days_in_month = month_days(year, month)
        periods = rec.get("trip_periods", [])
    
        if not periods:
            return
    
        hours = normalize_hours_list(rec.get("hours"), year, month)
        day_from = max(1, min(day_from, days_in_month))
        day_to = max(1, min(day_to, days_in_month))
    
        for day in range(day_from, day_to + 1):
            current_day = date(year, month, day)
            # НОВОЕ: Проверяем, попадает ли день хотя бы в один период из списка
            if any(p["from"] <= current_day <= p["to"] for p in periods):
                hours[day - 1] = value
    
        rec["hours"] = hours
        rec["_totals"] = calc_row_totals(hours, year, month)
    
    def _apply_hour_value_to_indexes(
        self,
        real_indexes: List[int],
        value: Optional[float],
        day_from: int,
        day_to: int,
    ) -> int:
        if not real_indexes:
            return 0
    
        year, month = self._get_year_month()
        skipped_without_period = []
    
        applied = 0
        for idx in real_indexes:
            if not (0 <= idx < len(self.rows)):
                continue
    
            rec = self.rows[idx]
            if not self._row_has_trip_period(rec):
                fio = normalize_spaces(rec.get("fio") or "")
                tbn = normalize_tbn(rec.get("tbn"))
                skipped_without_period.append(f"{fio} ({tbn})".strip())
                continue
    
            self._fill_hours_for_row(rec, value, day_from, day_to, year, month)
            applied += 1
    
        self._refresh_grid()
        self._update_trip_info_from_selection()
    
        if applied:
            self._mark_dirty()
            self._schedule_auto_save()
    
        if skipped_without_period:
            messagebox.showwarning(
                "Проставить время",
                "Часть строк пропущена, потому что у них не задан период командировки.\n\n"
                f"Пропущено строк: {len(skipped_without_period)}",
                parent=self,
            )
    
        return applied
    
    def _fill_hours_for_selected(self) -> None:
        indexes = self._get_selected_row_indexes()
        if not indexes:
            messagebox.showinfo("Проставить время", "Не выбраны строки.", parent=self)
            return
    
        params = self._ask_fill_hours_params()
        if not params:
            return
    
        day_from = int(params["from"])
        day_to = int(params["to"])
        value = params["value"]
    
        real_indexes = [self._visible_to_real_index(i) for i in indexes]
        real_indexes = [i for i in real_indexes if i is not None]
    
        applied = self._apply_hour_value_to_indexes(real_indexes, value, day_from, day_to)
        if applied:
            if value is None:
                if day_from == day_to:
                    self.var_status.set(f"Очищен день {day_from} у выбранных строк: {applied}")
                else:
                    self.var_status.set(f"Очищены дни {day_from}-{day_to} у выбранных строк: {applied}")
            else:
                if day_from == day_to:
                    self.var_status.set(f"Проставлено {value} ч. в день {day_from} у выбранных строк: {applied}")
                else:
                    self.var_status.set(f"Проставлено {value} ч. за дни {day_from}-{day_to} у выбранных строк: {applied}")
    
    def _fill_hours_for_all(self) -> None:
        if not self.rows:
            messagebox.showinfo("Проставить время", "В табеле нет строк.", parent=self)
            return
    
        params = self._ask_fill_hours_params()
        if not params:
            return
    
        day_from = int(params["from"])
        day_to = int(params["to"])
        value = params["value"]
    
        real_indexes = list(range(len(self.rows)))
        applied = self._apply_hour_value_to_indexes(real_indexes, value, day_from, day_to)
        if applied:
            if value is None:
                if day_from == day_to:
                    self.var_status.set(f"Очищен день {day_from} у всех строк: {applied}")
                else:
                    self.var_status.set(f"Очищены дни {day_from}-{day_to} у всех строк: {applied}")
            else:
                if day_from == day_to:
                    self.var_status.set(f"Проставлено {value} ч. в день {day_from} у всех строк: {applied}")
                else:
                    self.var_status.set(f"Проставлено {value} ч. за дни {day_from}-{day_to} у всех строк: {applied}")

    def _delete_selected_rows(self) -> None:
        indexes = self._get_selected_row_indexes()
        if not indexes:
            messagebox.showinfo("Удаление", "Не выбраны строки.", parent=self)
            return

        if not messagebox.askyesno(
            "Удаление",
            f"Удалить выбранные строки: {len(indexes)} шт.?",
            parent=self,
        ):
            return

        real_indexes = [self._visible_to_real_index(i) for i in indexes]
        real_indexes = sorted([i for i in real_indexes if i is not None], reverse=True)

        for idx in real_indexes:
            del self.rows[idx]

        self._refresh_grid()
        self._update_trip_info_from_selection()
        self.var_status.set(f"Удалено строк: {len(real_indexes)}")

        if real_indexes:
            self._mark_dirty()
            self._schedule_auto_save()

    def _clear_hours_for_selected(self) -> None:
        indexes = self._get_selected_row_indexes()
        if not indexes:
            messagebox.showinfo("Очистка часов", "Не выбраны строки.", parent=self)
            return

        year, month = self._get_year_month()
        real_indexes = [self._visible_to_real_index(i) for i in indexes]
        real_indexes = [i for i in real_indexes if i is not None]

        for idx in real_indexes:
            self.rows[idx]["hours"] = [None] * 31
            self.rows[idx]["_totals"] = calc_row_totals(self.rows[idx]["hours"], year, month)

        self._refresh_grid()
        self.var_status.set(f"Очищены часы у строк: {len(real_indexes)}")

        if real_indexes:
            self._mark_dirty()
            self._schedule_auto_save()

    def _set_trip_period_for_selected(self) -> None:
        # Для массовой установки логично просто добавлять период к существующим
        # или переопределять? Сделаем вызов стандартного диалога для добавления ОДНОГО общего периода всем выделенным
        indexes = self._get_selected_row_indexes()
        if not indexes:
            messagebox.showinfo("Период выбранным", "Не выбраны строки.", parent=self)
            return

        year, month = self._get_year_month()

        result = TripPeriodDialog.show(
            self,
            initial_date_from=None,
            initial_date_to=None,
            year=year,
            month=month,
        )
        if result is None or result[0] is None:
            return

        trip_date_from, trip_date_to = result
        new_period = {"from": trip_date_from, "to": trip_date_to}

        real_indexes = [self._visible_to_real_index(i) for i in indexes]
        real_indexes = [i for i in real_indexes if i is not None]

        for idx in real_indexes:
            # Если нужно заменять все периоды: self.rows[idx]["trip_periods"] = [new_period]
            # Если нужно добавлять (логичнее):
            if "trip_periods" not in self.rows[idx] or not self.rows[idx]["trip_periods"]:
                self.rows[idx]["trip_periods"] = []
            self.rows[idx]["trip_periods"].append(new_period)

        self._refresh_grid()
        self._update_trip_info_from_selection()
        self.var_status.set(f"Период добавлен строкам: {len(real_indexes)}")

        if real_indexes:
            self._mark_dirty()
            self._schedule_auto_save()

    # =========================================================
    # Экспорт
    # =========================================================
    def _export_to_excel(self) -> None:
        try:
            year, month = self._get_year_month()
            object_id, object_addr = self._parse_selected_object()

            if not self.rows:
                messagebox.showinfo("Выгрузка", "В табеле нет строк для выгрузки.", parent=self)
                return

            obj_part = object_id or object_addr or "без_объекта"

            try:
                prepared_by = normalize_spaces(
                    (getattr(self.app, "current_user", {}) or {}).get("full_name")
                    or (getattr(self.app, "current_user", {}) or {}).get("username")
                    or ""
                )
            except Exception:
                prepared_by = ""

            path = filedialog.asksaveasfilename(
                parent=self,
                title="Сохранить командировочный табель в Excel",
                defaultextension=".xlsx",
                initialfile=f"Командировочный_табель_{safe_filename(obj_part)}_{year}_{month:02d}.xlsx",
                filetypes=[("Excel", "*.xlsx"), ("Все", "*.*")],
            )
            if not path:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Командировочный табель"

            build_printable_trip_timesheet_sheet(
                ws,
                year=year,
                month=month,
                object_addr=object_addr,
                object_id=object_id,
                rows=self.rows,
                prepared_by=prepared_by,
            )

            wb.save(path)

            messagebox.showinfo(
                "Выгрузка",
                f"Готово.\nСтрок: {len(self.rows)}\nФайл: {path}",
                parent=self,
            )
        except Exception as exc:
            messagebox.showerror("Выгрузка", f"Ошибка выгрузки:\n{exc}", parent=self)

    # =========================================================
    # Нижняя инфо-строка
    # =========================================================
    def _update_trip_info_from_selection(self) -> None:
        indexes = self._get_selected_row_indexes()
        visible_rows = self._get_visible_rows()

        if len(indexes) != 1:
            self.var_trip_info.set("")
            return

        idx = indexes[0]
        if not (0 <= idx < len(visible_rows)):
            self.var_trip_info.set("")
            return

        rec = visible_rows[idx]
        fio = normalize_spaces(rec.get("fio") or "")
        tbn = normalize_tbn(rec.get("tbn"))
        periods = rec.get("trip_periods", [])

        # НОВОЕ: Форматируем строку со списком периодов
        if periods:
            p_strs = [f"{p['from'].strftime('%d.%m')} - {p['to'].strftime('%d.%m')}" for p in periods]
            period_str = "Командировки: " + ", ".join(p_strs)
        else:
            period_str = "Период командировки не задан"

        suffix = fio
        if tbn:
            suffix = f"{fio} ({tbn})"

        self.var_trip_info.set(f"{suffix}: {period_str}")

    # =========================================================
    # Внешние helpers
    # =========================================================
    def open_by_context(
        self,
        *,
        object_id: Optional[str],
        object_addr: str,
        year: int,
        month: int,
    ) -> None:
        self.var_year.set(int(year))
        self.var_month.set(int(month))
        self.cmb_month.current(max(0, int(month) - 1))

        object_id_norm = normalize_spaces(object_id or "")
        object_addr_norm = normalize_spaces(object_addr or "")

        self.cmb_address.set(object_addr_norm)
        self._sync_object_id_values_silent()

        if object_id_norm:
            values = list(self.cmb_object_id.cget("values") or [])
            if object_id_norm not in values:
                values.append(object_id_norm)
                self.cmb_object_id.config(values=values)
            self.cmb_object_id.set(object_id_norm)
        else:
            self.cmb_object_id.set("")

        self._open_timesheet()

    def add_empty_row(self) -> None:
        self.rows.append(self._empty_row())
        self._refresh_grid()
        self.var_status.set("Добавлена пустая строка.")
        self._mark_dirty()
        self._schedule_auto_save()
