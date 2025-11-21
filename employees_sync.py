```python
from pathlib import Path
from typing import Optional
from openpyxl import load_workbook
from datetime import datetime
from settings_manager import get_db_connection  # уже есть


def _s(val) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        val = int(val)
    return str(val).strip()


def import_employees_from_excel(path: Path) -> int:
    """
    Импортирует сотрудников из Excel-файла (как в примере "ШТАТ на ноябрь ...").
    Обновляет таблицы departments и employees.
    Возвращает количество обработанных строк (сотрудников).
    """
    if not path.exists():
        raise FileNotFoundError(f"Файл не найден: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    # Читаем заголовки
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    hdr = [_s(c).lower() for c in header_row]

    def col_idx(name_substr: str) -> Optional[int]:
        name_substr = name_substr.lower()
        for i, h in enumerate(hdr):
            if name_substr in h:
                return i
        return None

    idx_tbn = col_idx("табельный")
    idx_fio = col_idx("сотрудник")
    idx_pos = col_idx("должность")
    idx_dep = col_idx("подразделение")
    idx_dismissal = col_idx("увольн")  # "Дата увольнения"

    if idx_fio is None or idx_tbn is None:
        raise RuntimeError("Не найдены обязательные колонки 'Табельный номер' и/или 'Сотрудник'")

    conn = get_db_connection()
    processed = 0

    try:
        with conn:
            with conn.cursor() as cur:
                # Пройдём по строкам, начиная со 2-й
                for row in ws.iter_rows(min_row=2, values_only=True):
                    fio = _s(row[idx_fio]) if idx_fio < len(row) else ""
                    tbn = _s(row[idx_tbn]) if idx_tbn < len(row) else ""
                    pos = _s(row[idx_pos]) if idx_pos is not None and idx_pos < len(row) else ""
                    dep_name = _s(row[idx_dep]) if idx_dep is not None and idx_dep < len(row) else ""
                    dismissal_raw = row[idx_dismissal] if idx_dismissal is not None and idx_dismissal < len(row) else None

                    if not fio and not tbn:
                        continue  # пустая строка, пропускаем

                    # Признак увольнения по дате
                    is_fired = False
                    if dismissal_raw:
                        # openpyxl может дать дату как datetime, date или строку
                        if isinstance(dismissal_raw, (datetime, )):
                            is_fired = True
                        else:
                            s = _s(dismissal_raw)
                            if s:
                                is_fired = True

                    # 1. Подразделение
                    department_id = None
                    if dep_name:
                        cur.execute("SELECT id FROM departments WHERE name = %s", (dep_name,))
                        r = cur.fetchone()
                        if r:
                            department_id = r[0]
                        else:
                            cur.execute(
                                "INSERT INTO departments (name) VALUES (%s) RETURNING id",
                                (dep_name,)
                            )
                            department_id = cur.fetchone()[0]

                    # 2. Сотрудник — ищем по табельному номеру, если есть, иначе по ФИО
                    if tbn:
                        cur.execute("SELECT id FROM employees WHERE tbn = %s", (tbn,))
                        r = cur.fetchone()
                    else:
                        cur.execute("SELECT id FROM employees WHERE fio = %s", (fio,))
                        r = cur.fetchone()

                    if r:
                        emp_id = r[0]
                        cur.execute(
                            """
                            UPDATE employees
                               SET fio = %s,
                                   tbn = %s,
                                   position = %s,
                                   department_id = %s,
                                   is_fired = %s
                             WHERE id = %s
                            """,
                            (fio or None,
                             tbn or None,
                             pos or None,
                             department_id,
                             is_fired,
                             emp_id)
                        )
                    else:
                        cur.execute(
                            """
                            INSERT INTO employees (fio, tbn, position, department_id, is_fired)
                            VALUES (%s, %s, %s, %s, %s)
                            """,
                            (fio or None,
                             tbn or None,
                             pos or None,
                             department_id,
                             is_fired)
                        )

                    processed += 1
    finally:
        conn.close()

    return processed
