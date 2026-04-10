from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook
from psycopg2.extras import RealDictCursor, execute_values

# Используем тот же пул/подключение, что и в остальных модулях
db_connection_pool = None


def set_db_pool(pool) -> None:
    global db_connection_pool
    db_connection_pool = pool


def get_db_connection():
    if db_connection_pool:
        return db_connection_pool.getconn()
    raise RuntimeError("Пул соединений не был установлен из главного приложения.")


def release_db_connection(conn) -> None:
    if conn is None:
        return
    try:
        try:
            if not conn.closed:
                conn.rollback()
        except Exception:
            pass
    finally:
        if db_connection_pool:
            db_connection_pool.putconn(conn)
        else:
            try:
                conn.close()
            except Exception:
                pass


# ------------------------------------------------------------
# Helpers
# ------------------------------------------------------------

MONTHS_RU = {
    "январь": 1,
    "февраль": 2,
    "март": 3,
    "апрель": 4,
    "май": 5,
    "июнь": 6,
    "июль": 7,
    "август": 8,
    "сентябрь": 9,
    "октябрь": 10,
    "ноябрь": 11,
    "декабрь": 12,
}


def _s(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        val = int(val)
    return str(val).strip()


def _normalize_spaces(s: str) -> str:
    return " ".join(str(s or "").replace("\xa0", " ").split())


def _parse_float_hours(raw: Any) -> Optional[float]:
    s = _normalize_spaces(_s(raw)).lower()
    if not s:
        return None

    if s in ("х", "x", "в", "в/в", "-", "вых", "выходной"):
        return None

    s = s.replace(",", ".")
    try:
        val = float(s)
        if val > 0:
            return round(val, 2)
    except Exception:
        pass
    return None


def _extract_schedule_name_and_year(title: str) -> Tuple[str, Optional[int]]:
    raw = _normalize_spaces(title)
    if not raw:
        return "", None

    m = re.search(r"\bза\s+(\d{4})\s+год\b", raw, flags=re.IGNORECASE)
    year = int(m.group(1)) if m else None

    if m:
        name = _normalize_spaces(raw[:m.start()])
    else:
        name = raw

    return name, year


def _find_month_rows(ws) -> Dict[int, int]:
    found: Dict[int, int] = {}
    max_scan = min(ws.max_row, 60)

    for r in range(1, max_scan + 1):
        first_val = _normalize_spaces(_s(ws.cell(r, 1).value)).lower()
        if first_val in MONTHS_RU:
            found[MONTHS_RU[first_val]] = r

    return found


def _parse_schedule_excel(path: Path) -> Tuple[str, int, List[Tuple[date, bool, Optional[float], str]]]:
    if not path.exists():
        raise FileNotFoundError(f"Файл не найден: {path}")

    wb = load_workbook(path, data_only=True, read_only=False)
    ws = wb.active

    title = ""
    for row_idx in range(1, 6):
        candidate = _normalize_spaces(_s(ws.cell(row_idx, 1).value))
        if not candidate:
            continue
        if "год" in candidate.lower():
            title = candidate
            break
        if not title:
            title = candidate

    schedule_name, year = _extract_schedule_name_and_year(title)

    if not schedule_name:
        raise RuntimeError("Не удалось определить название графика из A1:A5.")
    if not year:
        raise RuntimeError("Не удалось определить год графика из A1:A5 (ожидалось '... за 2026 год').")

    month_rows = _find_month_rows(ws)
    if len(month_rows) < 12:
        raise RuntimeError("Не удалось распознать строки месяцев в файле графика.")

    items: List[Tuple[date, bool, Optional[float], str]] = []

    for month_num in range(1, 13):
        row_idx = month_rows.get(month_num)
        if not row_idx:
            continue

        # В файлах графиков дни идут со 2-й колонки:
        # B=1, C=2, ... AF=31
        for day in range(1, 32):
            try:
                work_dt = date(year, month_num, day)
            except ValueError:
                continue

            cell_val = ws.cell(row_idx, day + 1).value
            raw_value = _normalize_spaces(_s(cell_val))
            planned_hours = _parse_float_hours(cell_val)

            is_workday = planned_hours is not None
            if not raw_value:
                is_workday = False
            elif raw_value.lower() in ("х", "x", "в", "в/в", "вых", "выходной", "-"):
                is_workday = False

            items.append((work_dt, is_workday, planned_hours, raw_value))

    return schedule_name, year, items

# ------------------------------------------------------------
# DB API
# ------------------------------------------------------------

def get_unique_employee_schedule_names() -> List[str]:
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT DISTINCT TRIM(work_schedule) AS schedule_name
                FROM public.employees
                WHERE work_schedule IS NOT NULL
                  AND TRIM(work_schedule) <> ''
                ORDER BY TRIM(work_schedule)
                """
            )
            return [_normalize_spaces(r[0]) for r in cur.fetchall() if _normalize_spaces(r[0])]
    finally:
        release_db_connection(conn)


def get_work_schedules_list() -> List[Dict[str, Any]]:
    conn = get_db_connection()
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT
                    ws.id,
                    ws.schedule_name,
                    ws.year,
                    ws.source_filename,
                    ws.created_at,
                    ws.updated_at,
                    COUNT(wsd.id) AS days_count
                FROM public.work_schedules ws
                LEFT JOIN public.work_schedule_days wsd
                    ON wsd.schedule_id = ws.id
                GROUP BY ws.id, ws.schedule_name, ws.year, ws.source_filename, ws.created_at, ws.updated_at
                ORDER BY ws.schedule_name, ws.year
                """
            )
            return list(cur.fetchall())
    finally:
        release_db_connection(conn)


def import_work_schedule_from_excel(path: Path, forced_schedule_name: Optional[str] = None) -> Dict[str, Any]:
    schedule_name, year, items = _parse_schedule_excel(path)

    if forced_schedule_name and _normalize_spaces(forced_schedule_name):
        schedule_name = _normalize_spaces(forced_schedule_name)

    if not items:
        raise RuntimeError("В файле графика не найдено данных по дням.")

    conn = get_db_connection()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO public.work_schedules(schedule_name, year, source_filename, updated_at)
                    VALUES (%s, %s, %s, now())
                    ON CONFLICT (schedule_name, year)
                    DO UPDATE SET
                        source_filename = EXCLUDED.source_filename,
                        updated_at = now()
                    RETURNING id
                    """,
                    (schedule_name, year, path.name),
                )
                schedule_id = int(cur.fetchone()[0])

                cur.execute(
                    "DELETE FROM public.work_schedule_days WHERE schedule_id = %s",
                    (schedule_id,),
                )

                values = [
                    (schedule_id, work_dt, is_workday, planned_hours, raw_value or None)
                    for work_dt, is_workday, planned_hours, raw_value in items
                ]

                execute_values(
                    cur,
                    """
                    INSERT INTO public.work_schedule_days
                        (schedule_id, work_date, is_workday, planned_hours, raw_value)
                    VALUES %s
                    """,
                    values,
                )

        return {
            "schedule_name": schedule_name,
            "year": year,
            "days_loaded": len(items),
        }
    finally:
        release_db_connection(conn)


def delete_work_schedule(schedule_id: int) -> None:
    conn = get_db_connection()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM public.work_schedules WHERE id = %s", (int(schedule_id),))
    finally:
        release_db_connection(conn)


def get_schedule_days_map(schedule_name: str, year: int, month: int) -> Dict[int, Dict[str, Any]]:
    conn = get_db_connection()
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT
                    EXTRACT(DAY FROM d.work_date)::int AS day_num,
                    d.is_workday,
                    d.planned_hours,
                    d.raw_value
                FROM public.work_schedules s
                JOIN public.work_schedule_days d
                    ON d.schedule_id = s.id
                WHERE s.schedule_name = %s
                  AND s.year = %s
                  AND EXTRACT(MONTH FROM d.work_date) = %s
                ORDER BY d.work_date
                """,
                (_normalize_spaces(schedule_name), int(year), int(month)),
            )
            result: Dict[int, Dict[str, Any]] = {}
            for row in cur.fetchall():
                result[int(row["day_num"])] = {
                    "is_workday": bool(row["is_workday"]),
                    "planned_hours": float(row["planned_hours"]) if row["planned_hours"] is not None else None,
                    "raw_value": row["raw_value"] or "",
                }
            return result
    finally:
        release_db_connection(conn)


# ------------------------------------------------------------
# UI
# ------------------------------------------------------------

class WorkSchedulesPage(tk.Frame):
    def __init__(self, master):
        super().__init__(master, bg="#f7f9fb")
        self._unique_employee_schedules: List[str] = []
        self._loaded_rows: List[Dict[str, Any]] = []
        self._loaded_by_name: Dict[str, List[Dict[str, Any]]] = {}
        self._left_filtered_items: List[str] = []

        self.var_search = tk.StringVar()

        self._build_ui()
        self.reload_all()

    def _build_ui(self):
        top = tk.Frame(self, bg="#f7f9fb")
        top.pack(fill="x", padx=8, pady=8)

        ttk.Button(top, text="Обновить", command=self.reload_all).pack(side="left")
        ttk.Button(top, text="Импорт Excel...", command=self._on_import_excel).pack(side="left", padx=4)
        ttk.Button(top, text="Удалить", command=self._on_delete).pack(side="left", padx=4)

        self.lbl_info = tk.Label(
            top,
            text="",
            bg="#f7f9fb",
            fg="#4b5563",
            font=("Segoe UI", 9),
        )
        self.lbl_info.pack(side="right")

        mid = tk.PanedWindow(self, orient="horizontal", sashrelief="flat", bg="#d5dde6")
        mid.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        # -------------------------------------------------
        # Левая часть
        # -------------------------------------------------
        left = tk.LabelFrame(
            mid,
            text="Графики сотрудников",
            bg="#f7f9fb",
            fg="#1f2937",
            padx=6,
            pady=6,
        )
        mid.add(left, minsize=380)

        search_wrap = tk.Frame(left, bg="#f7f9fb")
        search_wrap.pack(fill="x", pady=(0, 6))

        tk.Label(
            search_wrap,
            text="Поиск:",
            bg="#f7f9fb",
            fg="#374151",
            font=("Segoe UI", 9),
        ).pack(side="left")

        ent = ttk.Entry(search_wrap, textvariable=self.var_search)
        ent.pack(side="left", fill="x", expand=True, padx=(6, 6))
        ent.bind("<KeyRelease>", lambda _e: self._refresh_left_list())

        ttk.Button(search_wrap, text="Очистить", command=self._clear_search).pack(side="left")

        legend = tk.Label(
            left,
            text="✅ загружен хотя бы за один год    ⚠ не загружен",
            bg="#f7f9fb",
            fg="#6b7280",
            font=("Segoe UI", 8),
            anchor="w",
        )
        legend.pack(fill="x", pady=(0, 4))

        list_wrap = tk.Frame(left, bg="#f7f9fb")
        list_wrap.pack(fill="both", expand=True)

        self.lst_employee_schedules = tk.Listbox(left, exportselection=False)
        self.lst_employee_schedules = tk.Listbox(
            list_wrap,
            exportselection=False,
            activestyle="none",
            font=("Segoe UI", 9),
        )
        self.lst_employee_schedules.pack(side="left", fill="both", expand=True)

        left_vsb = ttk.Scrollbar(list_wrap, orient="vertical", command=self.lst_employee_schedules.yview)
        left_vsb.pack(side="right", fill="y")
        self.lst_employee_schedules.configure(yscrollcommand=left_vsb.set)

        self.lst_employee_schedules.bind("<<ListboxSelect>>", lambda _e: self._sync_selected_schedule_name())

        self.lbl_left_details = tk.Label(
            left,
            text="Выберите график слева",
            justify="left",
            anchor="nw",
            bg="#f7f9fb",
            fg="#374151",
            font=("Segoe UI", 9),
            relief="solid",
            bd=1,
            padx=8,
            pady=6,
        )
        self.lbl_left_details.pack(fill="x", pady=(6, 0))

        # -------------------------------------------------
        # Правая часть
        # -------------------------------------------------
        right = tk.LabelFrame(
            mid,
            text="Загруженные графики",
            bg="#f7f9fb",
            fg="#1f2937",
            padx=6,
            pady=6,
        )
        mid.add(right, minsize=760)

        columns = ("id", "schedule_name", "year", "days_count", "source_filename", "updated_at")
        self.tree = ttk.Treeview(right, columns=columns, show="headings", height=16)

        self.tree.heading("id", text="ID")
        self.tree.heading("schedule_name", text="Название графика")
        self.tree.heading("year", text="Год")
        self.tree.heading("days_count", text="Дней")
        self.tree.heading("source_filename", text="Файл")
        self.tree.heading("updated_at", text="Обновлён")

        self.tree.column("id", width=50, anchor="center", stretch=False)
        self.tree.column("schedule_name", width=350, anchor="w")
        self.tree.column("year", width=70, anchor="center", stretch=False)
        self.tree.column("days_count", width=70, anchor="center", stretch=False)
        self.tree.column("source_filename", width=180, anchor="w")
        self.tree.column("updated_at", width=140, anchor="center", stretch=False)

        vsb = ttk.Scrollbar(right, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)

        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        self.tree.bind("<<TreeviewSelect>>", lambda _e: self._sync_tree_to_left())

    def _clear_search(self):
        self.var_search.set("")
        self._refresh_left_list()

    def _format_left_item(self, schedule_name: str) -> str:
        rows = self._loaded_by_name.get(schedule_name, [])
        if rows:
            return f"✅ {schedule_name}"
        return f"⚠ {schedule_name}"

    def _extract_schedule_name_from_left_item(self, item_text: str) -> str:
        text = _normalize_spaces(item_text)
        if text.startswith("✅ "):
            return _normalize_spaces(text[2:])
        if text.startswith("⚠ "):
            return _normalize_spaces(text[2:])
        return text

    def _refresh_left_list(self):
        query = _normalize_spaces(self.var_search.get() or "").lower()

        current_selected_name = self._get_selected_employee_schedule()

        self.lst_employee_schedules.delete(0, "end")
        self._left_filtered_items.clear()

        for name in self._unique_employee_schedules:
            if query and query not in name.lower():
                continue

            item_text = self._format_left_item(name)
            self._left_filtered_items.append(item_text)
            self.lst_employee_schedules.insert("end", item_text)

            idx = self.lst_employee_schedules.size() - 1
            if self._loaded_by_name.get(name):
                self.lst_employee_schedules.itemconfig(idx, fg="#166534")
            else:
                self.lst_employee_schedules.itemconfig(idx, fg="#b45309")

        if current_selected_name:
            for i, item_text in enumerate(self._left_filtered_items):
                if self._extract_schedule_name_from_left_item(item_text) == current_selected_name:
                    self.lst_employee_schedules.selection_clear(0, "end")
                    self.lst_employee_schedules.selection_set(i)
                    self.lst_employee_schedules.see(i)
                    break

        self._update_left_details()

    def reload_all(self):
        try:
            self._unique_employee_schedules = get_unique_employee_schedule_names()
            self._loaded_rows = get_work_schedules_list()

            self._loaded_by_name.clear()
            for row in self._loaded_rows:
                name = _normalize_spaces(row.get("schedule_name") or "")
                if not name:
                    continue
                self._loaded_by_name.setdefault(name, []).append(row)

            self._refresh_left_list()

            for item in self.tree.get_children():
                self.tree.delete(item)

            for row in self._loaded_rows:
                updated_at = row.get("updated_at")
                updated_str = updated_at.strftime("%d.%m.%Y %H:%M") if isinstance(updated_at, datetime) else ""
                self.tree.insert(
                    "",
                    "end",
                    iid=str(row["id"]),
                    values=(
                        row["id"],
                        row["schedule_name"],
                        row["year"],
                        row.get("days_count") or 0,
                        row.get("source_filename") or "",
                        updated_str,
                    ),
                )

            loaded_unique_count = len(self._loaded_by_name)
            not_loaded_count = max(0, len(self._unique_employee_schedules) - loaded_unique_count)

            self.lbl_info.config(
                text=(
                    f"Графиков у сотрудников: {len(self._unique_employee_schedules)} | "
                    f"Загружено уникальных: {loaded_unique_count} | "
                    f"Не загружено: {not_loaded_count}"
                )
            )

            self._update_left_details()
        except Exception as e:
            messagebox.showerror("Графики", f"Ошибка загрузки:\n{e}", parent=self)

    def _get_selected_employee_schedule(self) -> str:
        try:
            sel = self.lst_employee_schedules.curselection()
            if not sel:
                return ""
            raw = self.lst_employee_schedules.get(sel[0])
            return self._extract_schedule_name_from_left_item(raw)
        except Exception:
            return ""

    def _update_left_details(self):
        schedule_name = self._get_selected_employee_schedule()
        if not schedule_name:
            self.lbl_left_details.config(text="Выберите график слева")
            return

        rows = self._loaded_by_name.get(schedule_name, [])
        if not rows:
            text = (
                f"График:\n{schedule_name}\n\n"
                f"Статус: НЕ ЗАГРУЖЕН\n"
                f"Загрузок по годам: нет"
            )
        else:
            years = sorted({int(r.get('year') or 0) for r in rows if r.get("year")})
            years_text = ", ".join(str(y) for y in years) if years else "—"

            text = (
                f"График:\n{schedule_name}\n\n"
                f"Статус: ЗАГРУЖЕН\n"
                f"Загружен за годы: {years_text}\n"
                f"Количество загруженных версий: {len(rows)}"
            )

        self.lbl_left_details.config(text=text)

    def _sync_selected_schedule_name(self):
        schedule_name = self._get_selected_employee_schedule()
        self._update_left_details()

        if not schedule_name:
            return

        matched_iid = None
        matched_rows = self._loaded_by_name.get(schedule_name, [])
        if matched_rows:
            latest_row = sorted(
                matched_rows,
                key=lambda r: (
                    int(r.get("year") or 0),
                    r.get("updated_at") or datetime.min,
                ),
                reverse=True,
            )[0]
            matched_iid = str(latest_row["id"])

        self.tree.selection_remove(*self.tree.selection())
        if matched_iid and self.tree.exists(matched_iid):
            self.tree.selection_set(matched_iid)
            self.tree.focus(matched_iid)
            self.tree.see(matched_iid)

    def _sync_tree_to_left(self):
        sel = self.tree.selection()
        if not sel:
            return

        item_id = sel[0]
        values = self.tree.item(item_id, "values")
        if not values:
            return

        schedule_name = _normalize_spaces(values[1])

        for i, item_text in enumerate(self._left_filtered_items):
            if self._extract_schedule_name_from_left_item(item_text) == schedule_name:
                self.lst_employee_schedules.selection_clear(0, "end")
                self.lst_employee_schedules.selection_set(i)
                self.lst_employee_schedules.see(i)
                break

        self._update_left_details()

    def _on_import_excel(self):
        selected_schedule_name = self._get_selected_employee_schedule()

        path = filedialog.askopenfilename(
            parent=self,
            title="Выберите Excel-файл графика",
            filetypes=[("Excel", "*.xlsx"), ("Все файлы", "*.*")],
        )
        if not path:
            return

        try:
            info = import_work_schedule_from_excel(
                Path(path),
                forced_schedule_name=selected_schedule_name or None,
            )

            bind_info = (
                f"\nПривязка к графику сотрудников: {selected_schedule_name}"
                if selected_schedule_name
                else ""
            )

            messagebox.showinfo(
                "Графики",
                "Импорт завершён.\n\n"
                f"График: {info['schedule_name']}\n"
                f"Год: {info['year']}\n"
                f"Загружено дней: {info['days_loaded']}"
                f"{bind_info}",
                parent=self,
            )
            self.reload_all()
        except Exception as e:
            messagebox.showerror("Графики", f"Ошибка импорта:\n{e}", parent=self)

    def _on_delete(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Графики", "Выберите график в таблице.", parent=self)
            return

        item_id = sel[0]
        values = self.tree.item(item_id, "values")
        if not values:
            return

        schedule_id = int(values[0])
        schedule_name = values[1]
        year = values[2]

        if not messagebox.askyesno(
            "Удалить график",
            f"Удалить график:\n{schedule_name}\nза {year} год?",
            parent=self,
        ):
            return

        try:
            delete_work_schedule(schedule_id)
            self.reload_all()
            messagebox.showinfo("Графики", "График удалён.", parent=self)
        except Exception as e:
            messagebox.showerror("Графики", f"Ошибка удаления:\n{e}", parent=self)


def create_work_schedules_page(parent, app_ref=None):
    return WorkSchedulesPage(parent)


__all__ = [
    "set_db_pool",
    "get_unique_employee_schedule_names",
    "get_work_schedules_list",
    "import_work_schedule_from_excel",
    "delete_work_schedule",
    "get_schedule_days_map",
    "WorkSchedulesPage",
    "create_work_schedules_page",
]
