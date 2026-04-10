from __future__ import annotations

import logging
from datetime import datetime
from typing import List, Dict, Any, Optional

import tkinter as tk
from tkinter import ttk, messagebox, filedialog

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from psycopg2.extras import RealDictCursor

import timesheet_module

# ============================================================
#  Цветовая схема
# ============================================================
WK_COLORS = {
    "bg": "#f0f2f5",
    "panel": "#ffffff",
    "accent": "#1565c0",
    "accent_light": "#e3f2fd",
    "success": "#2e7d32",
    "warning": "#b00020",
    "border": "#dde1e7",
    "btn_save_bg": "#1565c0",
    "btn_save_fg": "#ffffff",
    "row_even": "#ffffff",
    "row_odd": "#f8f9fb",
    "row_ot": "#fff9c4",
    "row_night": "#e8f5e9",
    "row_selected": "#bbdefb",
    "sidebar_bg": "#f5f7fa",
    "sidebar_hdr": "#e3f2fd",
}

# ============================================================
#  Пул соединений
# ============================================================
db_connection_pool = None


def set_db_pool(pool):
    global db_connection_pool
    db_connection_pool = pool


def get_db_connection():
    if not db_connection_pool:
        raise RuntimeError("Пул соединений не установлен для employees.py")
    return db_connection_pool.getconn()


def release_db_connection(conn):
    if db_connection_pool and conn:
        db_connection_pool.putconn(conn)


# ============================================================
#  DB API
# ============================================================
def find_employee_work_summary(
    fio: Optional[str] = None,
    tbn: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    department: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """Свод по сотруднику: объекты, периоды, дни/часы/ночные/переработка."""
    if not fio and not tbn:
        return []

    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            where: List[str] = ["1=1"]
            params: List[Any] = []

            if fio:
                where.append("LOWER(TRIM(r.fio)) = LOWER(TRIM(%s))")
                params.append(fio)
            if tbn:
                where.append("COALESCE(TRIM(r.tbn), '') = TRIM(%s)")
                params.append(tbn)
            if year is not None:
                where.append("h.year = %s")
                params.append(year)
            if month is not None:
                where.append("h.month = %s")
                params.append(month)
            if department:
                where.append("COALESCE(h.department, '') = %s")
                params.append(department)

            cur.execute(
                f"""
                SELECT
                    h.object_id,
                    h.object_addr,
                    h.year,
                    h.month,
                    COALESCE(h.department, '')          AS department,
                    SUM(COALESCE(r.total_days,    0))   AS total_days,
                    SUM(COALESCE(r.total_hours,   0))   AS total_hours,
                    SUM(COALESCE(r.night_hours,   0))   AS night_hours,
                    SUM(COALESCE(r.overtime_day,  0))   AS overtime_day,
                    SUM(COALESCE(r.overtime_night,0))   AS overtime_night
                FROM timesheet_headers h
                JOIN timesheet_rows r ON r.header_id = h.id
                WHERE {" AND ".join(where)}
                GROUP BY
                    h.object_id, h.object_addr,
                    h.year, h.month,
                    COALESCE(h.department, '')
                ORDER BY
                    h.year DESC, h.month DESC,
                    h.object_addr,
                    COALESCE(h.department, '')
                """,
                params,
            )
            return [dict(row) for row in cur.fetchall()]
    finally:
        if conn:
            release_db_connection(conn)


month_name_ru = timesheet_module.month_name_ru
load_employees_from_db = timesheet_module.load_employees_from_db


class WorkersPage(tk.Frame):
    def __init__(self, master, app_ref):
        super().__init__(master, bg=WK_COLORS["bg"])
        self.app_ref = app_ref

        self.employees = load_employees_from_db()
        self.emp_info: Dict[str, Dict[str, str]] = {}

        for fio, tbn, pos, dep, work_schedule in self.employees:
            self.emp_info[fio] = {
                "tbn": tbn or "",
                "pos": pos or "",
                "dep": dep or "",
                "work_schedule": work_schedule or "",
            }

        deps_set = {
            (dep or "").strip()
            for _, _, _, dep, _work_schedule in self.employees
            if (dep or "").strip()
        }
        self.departments = ["Все"] + sorted(deps_set)

        self._selected_fio: str = ""
        self._selected_tbn: str = ""

        self.var_year = tk.StringVar(value=str(datetime.now().year))
        self.var_month = tk.StringVar(value="Все")
        self.var_dep = tk.StringVar(value="Все")

        self._rows: List[Dict[str, Any]] = []

        self._build_ui()

    def _build_ui(self):
        hdr = tk.Frame(self, bg=WK_COLORS["accent"], pady=6)
        hdr.pack(fill="x")
        tk.Label(
            hdr,
            text="👷  Работники — история по объектам",
            font=("Segoe UI", 12, "bold"),
            bg=WK_COLORS["accent"],
            fg="white",
            padx=12,
        ).pack(side="left")

        main = tk.Frame(self, bg=WK_COLORS["bg"])
        main.pack(fill="both", expand=True, padx=0, pady=0)
        main.grid_columnconfigure(1, weight=1)
        main.grid_rowconfigure(0, weight=1)

        self._build_sidebar(main)
        self._build_content(main)

        bottom = tk.Frame(self, bg=WK_COLORS["accent_light"], pady=5)
        bottom.pack(fill="x", padx=0, pady=0)

        self.lbl_total = tk.Label(
            bottom,
            text="Выберите сотрудника в списке слева",
            font=("Segoe UI", 9, "bold"),
            fg=WK_COLORS["accent"],
            bg=WK_COLORS["accent_light"],
        )
        self.lbl_total.pack(side="left", padx=10)

        tk.Label(
            bottom,
            text="🟡 Переработка  🟢 Ночные часы",
            font=("Segoe UI", 8),
            fg="#555",
            bg=WK_COLORS["accent_light"],
        ).pack(side="right", padx=10)

    def _build_sidebar(self, parent):
        sidebar = tk.Frame(parent, bg=WK_COLORS["sidebar_bg"], width=260, relief="flat")
        sidebar.grid(row=0, column=0, sticky="nsew", padx=(10, 4), pady=10)
        sidebar.grid_propagate(False)
        sidebar.grid_rowconfigure(2, weight=1)
        sidebar.grid_columnconfigure(0, weight=1)

        hdr_sb = tk.Frame(sidebar, bg=WK_COLORS["sidebar_hdr"], pady=6)
        hdr_sb.grid(row=0, column=0, sticky="ew")
        tk.Label(
            hdr_sb,
            text="👤  Сотрудники",
            font=("Segoe UI", 9, "bold"),
            bg=WK_COLORS["sidebar_hdr"],
            fg=WK_COLORS["accent"],
            padx=8,
        ).pack(side="left")

        self.lbl_emp_count = tk.Label(
            hdr_sb,
            text="",
            font=("Segoe UI", 8),
            fg="#666",
            bg=WK_COLORS["sidebar_hdr"],
        )
        self.lbl_emp_count.pack(side="right", padx=6)

        srch = tk.Frame(sidebar, bg=WK_COLORS["sidebar_bg"], pady=4)
        srch.grid(row=1, column=0, sticky="ew", padx=6)
        srch.grid_columnconfigure(0, weight=1)

        self.var_search = tk.StringVar()
        ent_srch = ttk.Entry(srch, textvariable=self.var_search, font=("Segoe UI", 9))
        ent_srch.grid(row=0, column=0, sticky="ew")
        ent_srch.bind("<KeyRelease>", self._on_search_key)

        tk.Label(
            srch,
            text="🔍",
            bg=WK_COLORS["sidebar_bg"],
            font=("Segoe UI", 10),
        ).grid(row=0, column=1, padx=(4, 0))

        dep_f = tk.Frame(sidebar, bg=WK_COLORS["sidebar_bg"])
        dep_f.grid(row=2, column=0, sticky="ew", padx=6, pady=(2, 0))
        dep_f.grid_columnconfigure(0, weight=1)

        self.var_sidebar_dep = tk.StringVar(value="Все")
        self.cmb_sidebar_dep = ttk.Combobox(
            dep_f,
            state="readonly",
            width=28,
            textvariable=self.var_sidebar_dep,
            values=self.departments,
            font=("Segoe UI", 8),
        )
        self.cmb_sidebar_dep.grid(row=0, column=0, sticky="ew", pady=2)
        self.cmb_sidebar_dep.bind("<<ComboboxSelected>>", lambda e: self._rebuild_sidebar_list())

        list_frame = tk.Frame(sidebar, bg=WK_COLORS["sidebar_bg"])
        list_frame.grid(row=3, column=0, sticky="nsew", padx=6, pady=(4, 6))
        list_frame.grid_rowconfigure(0, weight=1)
        list_frame.grid_columnconfigure(0, weight=1)
        sidebar.grid_rowconfigure(3, weight=1)

        self.lb_employees = tk.Listbox(
            list_frame,
            font=("Segoe UI", 9),
            bg=WK_COLORS["panel"],
            selectbackground=WK_COLORS["accent"],
            selectforeground="white",
            activestyle="none",
            relief="flat",
            bd=1,
            highlightthickness=1,
            highlightcolor=WK_COLORS["border"],
            highlightbackground=WK_COLORS["border"],
        )
        vsb_lb = ttk.Scrollbar(list_frame, orient="vertical", command=self.lb_employees.yview)
        self.lb_employees.configure(yscrollcommand=vsb_lb.set)
        self.lb_employees.grid(row=0, column=0, sticky="nsew")
        vsb_lb.grid(row=0, column=1, sticky="ns")

        self.lb_employees.bind("<<ListboxSelect>>", self._on_emp_select)
        self.lb_employees.bind("<Double-1>", self._on_emp_double)
        self.lb_employees.bind("<Return>", self._on_emp_double)

        self._rebuild_sidebar_list()

    def _build_content(self, parent):
        content = tk.Frame(parent, bg=WK_COLORS["bg"])
        content.grid(row=0, column=1, sticky="nsew", padx=(0, 10), pady=10)
        content.grid_rowconfigure(2, weight=1)
        content.grid_columnconfigure(0, weight=1)

        card_pnl = tk.LabelFrame(
            content,
            text=" 👤 Выбранный сотрудник ",
            font=("Segoe UI", 9, "bold"),
            bg=WK_COLORS["panel"],
            fg=WK_COLORS["accent"],
            relief="groove",
            bd=1,
            padx=10,
            pady=8,
        )
        card_pnl.grid(row=0, column=0, sticky="ew", pady=(0, 4))
        card_pnl.grid_columnconfigure(1, weight=1)

        self.lbl_card_fio = tk.Label(
            card_pnl,
            text="— не выбран —",
            font=("Segoe UI", 11, "bold"),
            fg=WK_COLORS["accent"],
            bg=WK_COLORS["panel"],
            anchor="w",
        )
        self.lbl_card_fio.grid(row=0, column=0, columnspan=4, sticky="ew", pady=(0, 4))

        for col, label in enumerate(["Таб. №:", "Должность:", "Подразделение:"]):
            tk.Label(
                card_pnl,
                text=label,
                font=("Segoe UI", 8),
                fg="#777",
                bg=WK_COLORS["panel"],
            ).grid(row=1, column=col * 2, sticky="e", padx=(0, 4))

        self.lbl_card_tbn = tk.Label(
            card_pnl,
            text="—",
            font=("Segoe UI", 9, "bold"),
            fg="#333",
            bg=WK_COLORS["panel"],
            anchor="w",
        )
        self.lbl_card_tbn.grid(row=1, column=1, sticky="w", padx=(0, 16))

        self.lbl_card_pos = tk.Label(
            card_pnl,
            text="—",
            font=("Segoe UI", 9),
            fg="#333",
            bg=WK_COLORS["panel"],
            anchor="w",
        )
        self.lbl_card_pos.grid(row=1, column=3, sticky="w", padx=(0, 16))

        self.lbl_card_dep = tk.Label(
            card_pnl,
            text="—",
            font=("Segoe UI", 9),
            fg="#333",
            bg=WK_COLORS["panel"],
            anchor="w",
        )
        self.lbl_card_dep.grid(row=1, column=5, sticky="w")

        tk.Label(
            card_pnl,
            text="График:",
            font=("Segoe UI", 8),
            fg="#777",
            bg=WK_COLORS["panel"],
        ).grid(row=2, column=0, sticky="ne", padx=(0, 4), pady=(6, 0))

        self.lbl_card_schedule = tk.Label(
            card_pnl,
            text="—",
            font=("Segoe UI", 9),
            fg="#333",
            bg=WK_COLORS["panel"],
            anchor="w",
            justify="left",
            wraplength=700,
        )
        self.lbl_card_schedule.grid(row=2, column=1, columnspan=5, sticky="ew", pady=(6, 0))

        card_pnl.grid_columnconfigure(1, weight=0)
        card_pnl.grid_columnconfigure(3, weight=1)
        card_pnl.grid_columnconfigure(5, weight=1)

        flt_pnl = tk.LabelFrame(
            content,
            text=" 📅 Фильтр периода (необязательно) ",
            font=("Segoe UI", 9, "bold"),
            bg=WK_COLORS["panel"],
            fg=WK_COLORS["accent"],
            relief="groove",
            bd=1,
            padx=10,
            pady=6,
        )
        flt_pnl.grid(row=1, column=0, sticky="ew", pady=(0, 4))

        tk.Label(flt_pnl, text="Год:", font=("Segoe UI", 9), bg=WK_COLORS["panel"]).grid(
            row=0, column=0, sticky="e", padx=(0, 6), pady=2
        )
        tk.Spinbox(
            flt_pnl,
            from_=2000,
            to=2100,
            width=7,
            textvariable=self.var_year,
            font=("Segoe UI", 9),
        ).grid(row=0, column=1, sticky="w", pady=2)

        tk.Label(flt_pnl, text="Месяц:", font=("Segoe UI", 9), bg=WK_COLORS["panel"]).grid(
            row=0, column=2, sticky="e", padx=(16, 6), pady=2
        )
        ttk.Combobox(
            flt_pnl,
            state="readonly",
            width=13,
            textvariable=self.var_month,
            values=["Все"] + [month_name_ru(i) for i in range(1, 13)],
        ).grid(row=0, column=3, sticky="w", pady=2)

        tk.Label(flt_pnl, text="Подразделение:", font=("Segoe UI", 9), bg=WK_COLORS["panel"]).grid(
            row=0, column=4, sticky="e", padx=(16, 6), pady=2
        )
        ttk.Combobox(
            flt_pnl,
            state="readonly",
            width=28,
            textvariable=self.var_dep,
            values=self.departments,
        ).grid(row=0, column=5, sticky="ew", pady=2)
        flt_pnl.grid_columnconfigure(5, weight=1)

        btn_f = tk.Frame(flt_pnl, bg=WK_COLORS["panel"])
        btn_f.grid(row=0, column=6, sticky="e", padx=(14, 0))

        ttk.Button(btn_f, text="Сбросить", command=self._reset_filters).pack(side="left", padx=(0, 6))
        ttk.Button(btn_f, text="📊 Excel", command=self._export_excel).pack(side="left")

        tbl_pnl = tk.LabelFrame(
            content,
            text=" 📋 История работы на объектах ",
            font=("Segoe UI", 9, "bold"),
            bg=WK_COLORS["panel"],
            fg=WK_COLORS["accent"],
            relief="groove",
            bd=1,
        )
        tbl_pnl.grid(row=2, column=0, sticky="nsew")
        tbl_pnl.grid_rowconfigure(0, weight=1)
        tbl_pnl.grid_columnconfigure(0, weight=1)

        cols = (
            "period",
            "object",
            "object_id",
            "department",
            "total_days",
            "total_hours",
            "night_hours",
            "overtime_day",
            "overtime_night",
        )
        self.tree = ttk.Treeview(tbl_pnl, columns=cols, show="headings", selectmode="browse")

        heads = {
            "period": ("Период", 100, "center"),
            "object": ("Объект (адрес)", 340, "w"),
            "object_id": ("ID объекта", 90, "center"),
            "department": ("Подразделение", 160, "w"),
            "total_days": ("Дни", 60, "center"),
            "total_hours": ("Часы", 80, "e"),
            "night_hours": ("Ночных ч.", 80, "e"),
            "overtime_day": ("Пер. день", 90, "e"),
            "overtime_night": ("Пер. ночь", 90, "e"),
        }
        for col, (text, width, anchor) in heads.items():
            self.tree.heading(col, text=text)
            self.tree.column(col, width=width, anchor=anchor, stretch=(col == "object"))

        self.tree.tag_configure("even", background=WK_COLORS["row_even"])
        self.tree.tag_configure("odd", background=WK_COLORS["row_odd"])
        self.tree.tag_configure("ot", background=WK_COLORS["row_ot"])
        self.tree.tag_configure("night", background=WK_COLORS["row_night"])
        self.tree.tag_configure(
            "total",
            background=WK_COLORS["accent_light"],
            font=("Segoe UI", 9, "bold"),
        )

        vsb = ttk.Scrollbar(tbl_pnl, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")

    def _rebuild_sidebar_list(self, filter_text: str = ""):
        dep_filter = self.var_sidebar_dep.get().strip()
        ft = filter_text.strip().lower()

        self.lb_employees.delete(0, "end")
        self._lb_data: List[Dict[str, str]] = []

        for fio, tbn, pos, dep, work_schedule in self.employees:
            if dep_filter and dep_filter != "Все":
                if (dep or "").strip() != dep_filter:
                    continue

            if ft:
                if (
                    ft not in fio.lower()
                    and ft not in (tbn or "").lower()
                    and ft not in (pos or "").lower()
                    and ft not in (work_schedule or "").lower()
                ):
                    continue

            display = f"{fio}"
            if tbn:
                display += f"  [{tbn}]"

            self.lb_employees.insert("end", display)
            self._lb_data.append(
                {
                    "fio": fio,
                    "tbn": tbn or "",
                    "pos": pos or "",
                    "dep": dep or "",
                    "work_schedule": work_schedule or "",
                }
            )

        count = self.lb_employees.size()
        try:
            self.lbl_emp_count.config(text=f"{count} чел.")
        except Exception:
            pass

        if self._selected_fio:
            for i, d in enumerate(self._lb_data):
                if d["fio"] == self._selected_fio:
                    self.lb_employees.selection_clear(0, "end")
                    self.lb_employees.selection_set(i)
                    self.lb_employees.see(i)
                    break

    def _on_search_key(self, _event=None):
        self._rebuild_sidebar_list(self.var_search.get())

    def _on_emp_select(self, _event=None):
        sel = self.lb_employees.curselection()
        if not sel:
            return
        emp = self._lb_data[sel[0]]
        self._select_employee(emp)
        self._search()

    def _on_emp_double(self, _event=None):
        sel = self.lb_employees.curselection()
        if not sel:
            return
        emp = self._lb_data[sel[0]]
        self._select_employee(emp)
        self._search()

    def _select_employee(self, emp: Dict[str, str]):
        self._selected_fio = emp["fio"]
        self._selected_tbn = emp["tbn"]

        try:
            self.lbl_card_fio.config(text=emp["fio"] or "—")
            self.lbl_card_tbn.config(text=emp["tbn"] or "—")
            self.lbl_card_pos.config(text=emp["pos"] or "—")
            self.lbl_card_dep.config(text=emp["dep"] or "—")
            self.lbl_card_schedule.config(text=emp.get("work_schedule") or "—")
        except Exception:
            pass

        if emp["dep"] and emp["dep"] in self.departments:
            self.var_dep.set(emp["dep"])

        self._clear_table()

    def _clear_table(self):
        if self.tree:
            self.tree.delete(*self.tree.get_children())
        self._rows = []
        try:
            self.lbl_total.config(text=f"Выбран: {self._selected_fio}")
        except Exception:
            pass

    def _reset_filters(self):
        self.var_year.set("")
        self.var_month.set("Все")
        self.var_dep.set("Все")
        self._rows.clear()
        if self.tree:
            self.tree.delete(*self.tree.get_children())
        try:
            self.lbl_total.config(text="")
        except Exception:
            pass

    def _search(self):
        fio = self._selected_fio
        tbn = self._selected_tbn

        if not fio and not tbn:
            messagebox.showwarning("Работники", "Выберите сотрудника в списке слева.")
            return

        year = None
        y_str = self.var_year.get().strip()
        if y_str:
            try:
                y = int(y_str)
                if not (2000 <= y <= 2100):
                    raise ValueError
                year = y
            except ValueError:
                messagebox.showwarning("Работники", "Год введён некорректно (ожидается 2000–2100).")
                return

        month = None
        m_name = self.var_month.get().strip()
        if m_name and m_name != "Все":
            try:
                month = [month_name_ru(i) for i in range(1, 13)].index(m_name) + 1
            except ValueError:
                pass

        dep_val = self.var_dep.get().strip()
        dep = dep_val if (dep_val and dep_val != "Все") else None

        try:
            rows = find_employee_work_summary(
                fio=fio or None,
                tbn=tbn or None,
                year=year,
                month=month,
                department=dep,
            )
        except Exception as e:
            logging.exception("Ошибка поиска работника")
            messagebox.showerror("Работники", f"Ошибка при обращении к БД:\n{e}")
            return

        self._rows = rows
        self._fill_tree()

        if not rows:
            messagebox.showinfo("Работники", f"Для «{fio}» нет записей по заданным условиям.")

    def _fmt(self, v) -> str:
        if v is None:
            return ""
        if isinstance(v, float):
            return f"{v:.2f}".rstrip("0").rstrip(".")
        return str(v)

    def _fill_tree(self):
        self.tree.delete(*self.tree.get_children())

        if not self._rows:
            try:
                self.lbl_total.config(text="Ничего не найдено")
            except Exception:
                pass
            return

        sum_days = sum_hours = sum_night = sum_otd = sum_otn = 0.0

        for idx, r in enumerate(self._rows):
            yr = r.get("year")
            mn = r.get("month")
            period_str = f"{month_name_ru(mn)} {yr}" if yr and mn else ""

            td = float(r.get("total_days", 0) or 0)
            th = float(r.get("total_hours", 0) or 0)
            nh = float(r.get("night_hours", 0) or 0)
            otd = float(r.get("overtime_day", 0) or 0)
            otn = float(r.get("overtime_night", 0) or 0)

            sum_days += td
            sum_hours += th
            sum_night += nh
            sum_otd += otd
            sum_otn += otn

            if otd > 0 or otn > 0:
                tag = "ot"
            elif nh > 0:
                tag = "night"
            elif idx % 2 == 0:
                tag = "even"
            else:
                tag = "odd"

            self.tree.insert(
                "",
                "end",
                iid=str(idx),
                values=(
                    period_str,
                    r.get("object_addr") or "",
                    r.get("object_id") or "",
                    r.get("department") or "",
                    self._fmt(td) if td else "",
                    self._fmt(th) if th else "",
                    self._fmt(nh) if nh else "",
                    self._fmt(otd) if otd else "",
                    self._fmt(otn) if otn else "",
                ),
                tags=(tag,),
            )

        self.tree.insert(
            "",
            "end",
            iid="__total__",
            values=(
                "ИТОГО",
                f"Записей: {len(self._rows)}",
                "",
                "",
                self._fmt(sum_days),
                self._fmt(sum_hours),
                self._fmt(sum_night),
                self._fmt(sum_otd),
                self._fmt(sum_otn),
            ),
            tags=("total",),
        )

        parts = [f"Записей: {len(self._rows)}"]
        if sum_days:
            parts.append(f"Дней: {self._fmt(sum_days)}")
        if sum_hours:
            parts.append(f"Часов: {self._fmt(sum_hours)}")
        if sum_night:
            parts.append(f"Ночных: {self._fmt(sum_night)}")
        if sum_otd or sum_otn:
            parts.append(f"Переработка: {self._fmt(sum_otd)} / {self._fmt(sum_otn)}")

        try:
            self.lbl_total.config(text="  |  ".join(parts))
        except Exception:
            pass

    def _export_excel(self):
        if not self._rows:
            messagebox.showinfo("Экспорт", "Нет данных для выгрузки.")
            return

        who = self._selected_fio or self._selected_tbn or "работник"

        path = filedialog.asksaveasfilename(
            title="Сохранить историю работника",
            defaultextension=".xlsx",
            initialfile=f"История_{who.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Все", "*.*")],
        )
        if not path:
            return

        try:
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "История работы"

            info = self.emp_info.get(self._selected_fio, {})
            ws.append([f"Сотрудник: {who}"])
            ws.append(
                [
                    f"Таб. №: {info.get('tbn', '—')}  |  "
                    f"Должность: {info.get('pos', '—')}  |  "
                    f"Подразделение: {info.get('dep', '—')}"
                ]
            )
            ws.append([f"График работы: {info.get('work_schedule', '—')}"])
            ws.append([f"Экспорт: {datetime.now().strftime('%d.%m.%Y %H:%M')}"])
            ws.append([])

            header = [
                "Период",
                "Объект (адрес)",
                "ID объекта",
                "Подразделение",
                "Дни",
                "Часы",
                "Ночных ч.",
                "Пер. день",
                "Пер. ночь",
            ]
            ws.append(header)
            hdr_row = ws.max_row

            fill_hdr = PatternFill("solid", fgColor="1565C0")
            fill_ot = PatternFill("solid", fgColor="FFF9C4")
            fill_night = PatternFill("solid", fgColor="E8F5E9")
            fill_total = PatternFill("solid", fgColor="E3F2FD")
            font_hdr = Font(bold=True, color="FFFFFF")
            font_total = Font(bold=True)

            for c in range(1, len(header) + 1):
                cell = ws.cell(hdr_row, c)
                cell.font = font_hdr
                cell.fill = fill_hdr
                cell.alignment = Alignment(horizontal="center")

            sum_days = sum_hours = sum_night = sum_otd = sum_otn = 0.0

            for r in self._rows:
                yr = r.get("year")
                mn = r.get("month")
                period_str = f"{month_name_ru(mn)} {yr}" if yr and mn else ""

                td = float(r.get("total_days", 0) or 0)
                th = float(r.get("total_hours", 0) or 0)
                nh = float(r.get("night_hours", 0) or 0)
                otd = float(r.get("overtime_day", 0) or 0)
                otn = float(r.get("overtime_night", 0) or 0)

                sum_days += td
                sum_hours += th
                sum_night += nh
                sum_otd += otd
                sum_otn += otn

                ws.append(
                    [
                        period_str,
                        r.get("object_addr") or "",
                        r.get("object_id") or "",
                        r.get("department") or "",
                        td or None,
                        th or None,
                        nh or None,
                        otd or None,
                        otn or None,
                    ]
                )

                cur_r = ws.max_row
                if otd > 0 or otn > 0:
                    for c in range(1, len(header) + 1):
                        ws.cell(cur_r, c).fill = fill_ot
                elif nh > 0:
                    for c in range(1, len(header) + 1):
                        ws.cell(cur_r, c).fill = fill_night

            ws.append(
                [
                    "ИТОГО",
                    f"Записей: {len(self._rows)}",
                    "",
                    "",
                    sum_days or None,
                    sum_hours or None,
                    sum_night or None,
                    sum_otd or None,
                    sum_otn or None,
                ]
            )
            tot = ws.max_row
            for c in range(1, len(header) + 1):
                ws.cell(tot, c).fill = fill_total
                ws.cell(tot, c).font = font_total

            widths = [14, 44, 12, 22, 8, 10, 10, 12, 12]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

            ws.freeze_panes = f"A{hdr_row + 1}"
            wb.save(path)

            messagebox.showinfo("Экспорт", f"Файл сохранён:\n{path}\nЗаписей: {len(self._rows)}")
        except Exception as e:
            logging.exception("Ошибка экспорта")
            messagebox.showerror("Экспорт", f"Ошибка:\n{e}")


def create_workers_page(parent, app_ref) -> WorkersPage:
    return WorkersPage(parent, app_ref=app_ref)
