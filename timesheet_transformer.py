import argparse
import ctypes
import re
import sys
from datetime import time, datetime
from pathlib import Path
from typing import List, Optional, Tuple, Any

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import column_index_from_string, get_column_letter

# GUI (tkinter)
import tkinter as tk
from tkinter import filedialog, ttk

# ===== БРЕНДИНГ (жёстко вшито) =====
APP_NAME = "Табель‑конвертер"
WELCOME_HEADER = "Табель‑конвертер"
WELCOME_SUBTITLE = "Преобразование табеля 1С ЗУП в удобную таблицу.\nВыберите режим:"
WELCOME_COPYRIGHT = "Разработал Алексей Зезюкин, 2025"

# ===== Настройки =====
START_ROW = 21
HOURS_OFFSET = 2
RESULT_SHEET_NAME = "Результат"

# Под 3000+ сотрудников (4 строки на сотрудника ≈ 12000 строк)
MAX_SCAN_ROWS = 20000       # максимум строк смотреть от START_ROW
NO_GOOD_BREAK = 120         # обрывать поиск после стольких «пустых» строк подряд
PROGRESS_EVERY = 200        # обновлять прогресс каждые N шагов

# Полу-«ломаные» колонки дней
DAY_COLS_HALF1_LETTERS = ["I", "K", "M", "N", "P", "R", "T", "V", "X", "Z", "AB", "AD", "AF", "AH", "AK"]          # 1..15
DAY_COLS_HALF2_LETTERS = ["I", "K", "M", "N", "P", "R", "T", "V", "X", "Z", "AB", "AD", "AF", "AH", "AK", "AL"]     # 16..31
AO_COL_LETTER = "AO"

NON_WORKING_CODES = {"В", "НН", "ОТ", "ОД", "У", "УД", "Б", "ДО", "К", "ПР", "ОЖ", "ОЗ", "НС", "Н", "НВ"}

# ===== ЛОГ И СООБЩЕНИЯ =====
def exe_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path.cwd()

def safe_name(name: str) -> str:
    return re.sub(r'[<>:"/\\|?*]+', "_", str(name)).strip()

LOG_PATH = exe_dir() / f"{safe_name(APP_NAME)}.log"

def log(msg: str):
    try:
        with open(LOG_PATH, "a", encoding="utf-8") as f:
            f.write(msg.rstrip() + "\n")
    except Exception:
        pass

def clog(msg: str):
    log(msg)
    try:
        print(msg)
    except Exception:
        pass

def msg_info(title: str, text: str):
    try:
        ctypes.windll.user32.MessageBoxW(0, text, title, 0x40)  # MB_ICONINFORMATION
    except Exception:
        pass

def msg_error(title: str, text: str):
    try:
        ctypes.windll.user32.MessageBoxW(0, text, title, 0x10)  # MB_ICONERROR
    except Exception:
        pass

# ===== GUI: Приветствие и Прогресс =====
class WelcomeUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title(APP_NAME)
        self.root.geometry("440x240")
        self.root.resizable(False, False)

        title = tk.Label(self.root, text=WELCOME_HEADER, font=("Segoe UI", 14, "bold"))
        title.pack(pady=(12, 4))

        desc = tk.Label(self.root, text=WELCOME_SUBTITLE, font=("Segoe UI", 10), justify="center")
        desc.pack(pady=(0, 12))

        btn_frame = tk.Frame(self.root)
        btn_frame.pack(pady=4)

        self.choice = None  # ("file", path) | ("latest", folder) | None

        b1 = tk.Button(btn_frame, text="Выбрать файл…", width=22, command=self.choose_file)
        b1.grid(row=0, column=0, padx=6, pady=6)

        b2 = tk.Button(btn_frame, text="Последний в папке…", width=22, command=self.choose_folder)
        b2.grid(row=0, column=1, padx=6, pady=6)

        b3 = tk.Button(self.root, text="Отмена", width=16, command=self.cancel)
        b3.pack(pady=(6, 2))

        copy = tk.Label(self.root, text=WELCOME_COPYRIGHT, font=("Segoe UI", 8), fg="#666666")
        copy.pack(side="bottom", pady=(0, 8))

        self.root.attributes("-topmost", True)
        self.root.after(200, lambda: self.root.attributes("-topmost", False))

    def choose_file(self):
        fp = filedialog.askopenfilename(
            title="Выберите файл табеля",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")]
        )
        if fp:
            self.choice = ("file", fp)
            self.root.destroy()

    def choose_folder(self):
        folder = filedialog.askdirectory(title="Выберите папку с табелями")
        if folder:
            self.choice = ("latest", folder)
            self.root.destroy()

    def cancel(self):
        self.choice = None
        self.root.destroy()

    def run(self):
        self.root.mainloop()
        return self.choice


class ProgressUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title(f"{APP_NAME} — выполняется")
        self.root.geometry("500x170")
        self.root.resizable(False, False)

        self.label = tk.Label(self.root, text="Подготовка…", font=("Segoe UI", 10))
        self.label.pack(pady=(16, 6))

        self.progress = ttk.Progressbar(self.root, mode="determinate")
        self.progress.pack(fill="x", padx=16, pady=(0, 8))

        self.percent = tk.Label(self.root, text="0%", font=("Segoe UI", 9))
        self.percent.pack()

        self.cancelled = False
        self.btn_cancel = tk.Button(self.root, text="Отмена", width=12, command=self._cancel)
        self.btn_cancel.pack(pady=(8, 8))

        self.total = 100
        self.value = 0
        self._set_total(100)
        self._update()

        self.root.attributes("-topmost", True)
        self.root.after(200, lambda: self.root.attributes("-topmost", False))

    def _set_total(self, total: int):
        self.total = max(1, int(total))
        self.progress.configure(maximum=self.total)

    def set_phase(self, text: str, total: int):
        self.label.config(text=text)
        self._set_total(total)
        self.value = 0
        self.progress.configure(value=0)
        self._update()

    def set_progress(self, current: int):
        self.value = max(0, min(self.total, int(current)))
        self.progress.configure(value=self.value)
        pct = int(self.value * 100 / self.total) if self.total else 0
        self.percent.config(text=f"{pct}%")
        self._update()

    def inc(self, step: int = 1):
        self.set_progress(self.value + step)

    def _update(self):
        try:
            self.root.update_idletasks()
            self.root.update()
        except Exception:
            pass

    def _cancel(self):
        self.cancelled = True
        self.label.config(text="Отмена…")
        self._update()

    def is_cancelled(self) -> bool:
        return self.cancelled

    def close(self):
        try:
            self.root.destroy()
        except Exception:
            pass

class CancelledError(Exception):
    pass

# ===== Утилиты =====
def only_digits(s: str) -> str:
    return "".join(ch for ch in str(s or "") if ch.isdigit())

def has_letters(s: str) -> bool:
    return re.search(r"[A-Za-zА-Яа-яЁё]", str(s) or "") is not None

def clean_spaces(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    for cp in (0x00A0, 0x202F, 0x2009, 0x200A, 0x200B, 0x2060, 0xFEFF, 0x200E, 0x200F):
        s = s.replace(chr(cp), " ")
    s = s.replace("\t", " ").replace("\r\n", "\n").replace("\r", "\n")
    return s.strip()

def get_first_paren_content(s: str) -> str:
    m = re.search(r"\((.*?)\)", s)
    return m.group(1) if m else ""

def split_fio_and_title(raw: Any) -> Tuple[str, str]:
    s = clean_spaces(raw).replace("\r\n", "\n").replace("\r", "\n")
    parts = [p.strip() for p in s.split("\n") if p.strip()]
    fio = parts[0] if parts else ""
    title = ""
    if len(parts) > 1:
        inside = get_first_paren_content(parts[1])
        title = inside.strip() if inside else parts[1].strip()
    return fio, title

def extract_code_token(s: Any) -> str:
    txt = str(s) if s is not None else ""
    m = re.search(r"([A-Za-zА-Яа-яЁё]+)", txt)
    return m.group(1).upper() if m else ""

def is_non_working_code(code: str) -> bool:
    return code.upper().strip() in NON_WORKING_CODES

# ===== Парсинг чисел/времени и 'a/b' =====
def token_to_number(t: str) -> Optional[float]:
    t = clean_spaces(t)
    if not t:
        return None
    # Время "h:mm(:ss)"
    if ":" in t:
        parts = t.split(":")
        if len(parts) >= 2:
            try:
                hh = float(re.sub(r"[^\d.+-]", "", parts[0]) or 0)
                mm = float(re.sub(r"[^\d.+-]", "", parts[1]) or 0)
                ss = float(re.sub(r"[^\d.+-]", "", parts[2])) if len(parts) >= 3 else 0.0
                return hh + mm / 60.0 + ss / 3600.0
            except Exception:
                pass
    # Нормализуем разделители
    t = (t.replace("\uFF0C", ",").replace("\uFF0E", ".").replace("\u201A", ",").replace(" ", ""))
    while t and t[-1] in ",.":
        t = t[:-1]
    t = t.replace(",", ".")
    m = re.search(r"[-+]?\d+(?:\.\d+)?", t)
    if m:
        try:
            return float(m.group(0))
        except Exception:
            return None
    try:
        return float(t)
    except Exception:
        return None

def sum_slash_parts(s: str) -> Optional[float]:
    if "/" not in (s or ""):
        return None
    total = 0.0
    cnt = 0
    for part in s.split("/"):
        n = token_to_number(part)
        if n is not None:
            total += n
            cnt += 1
    if cnt >= 1:
        return total
    return None

def to_number_value(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        try:
            if float(v) < 1.0:
                return float(v) * 24.0
        except Exception:
            pass
        return float(v)
    if isinstance(v, time):
        return v.hour + v.minute / 60.0 + v.second / 3600.0
    if isinstance(v, datetime):
        return v.hour + v.minute / 60.0 + v.second / 3600.0
    s = str(v)
    sumv = sum_slash_parts(s)
    if sumv is not None:
        return float(sumv)
    n = token_to_number(s)
    return float(n) if n is not None else None

def day_hours_from_values(code_val: Any, hours_val: Any) -> Optional[float]:
    n = to_number_value(hours_val)
    if n is not None:
        return n
    code = extract_code_token(code_val)
    if not code:
        return None
    if is_non_working_code(code):
        return 0.0
    return None

# ===== Поиск конца данных (быстрый, с прогрессом) =====
def find_last_data_row(ws, start_row: int, ui: Optional[ProgressUI]) -> int:
    ao_col = column_index_from_string(AO_COL_LETTER)
    limit = min((ws.max_row or (start_row + MAX_SCAN_ROWS - 1)), start_row + MAX_SCAN_ROWS - 1)
    total = max(1, limit - start_row + 1)
    if ui:
        ui.set_phase("Поиск конца данных…", total)

    rows_iter = ws.iter_rows(min_row=start_row, max_row=limit,
                             min_col=2, max_col=ao_col, values_only=True)

    last_good = start_row - 1
    no_good = 0
    r = start_row - 1
    idx_C = 1
    idx_E = 3
    idx_AO = ao_col - 2

    for row in rows_iter:
        r += 1
        c_val = row[idx_C]
        e_val = row[idx_E]
        ao_val = row[idx_AO]

        good = False
        if has_letters(c_val):
            if (str(e_val or "").strip()) or (isinstance(ao_val, (int, float)) or re.search(r"\d", str(ao_val or ""))):
                good = True

        if good:
            last_good = r
            no_good = 0
        else:
            no_good += 1

        if ui:
            ui.set_progress(r - start_row + 1)
            if ui.is_cancelled():
                raise CancelledError()

        if (last_good >= start_row and no_good >= NO_GOOD_BREAK) or (r >= limit):
            break

    if last_good < start_row:
        last_good = start_row
    clog(f"find_last_data_row -> {last_good}")
    return last_good

# ===== Трансформация (быстро, с прогрессом) =====
def transform_sheet(ws, ui: Optional[ProgressUI]) -> Tuple[List[str], List[List[Any]]]:
    ao_col = column_index_from_string(AO_COL_LETTER)
    day_cols_h1 = [column_index_from_string(x) for x in DAY_COLS_HALF1_LETTERS]
    day_cols_h2 = [column_index_from_string(x) for x in DAY_COLS_HALF2_LETTERS]

    def idx_in_slice(col_num: int) -> int:
        return col_num - 2

    idx_B = idx_in_slice(2)
    idx_C = idx_in_slice(3)
    idx_E = idx_in_slice(5)
    idx_AO = idx_in_slice(ao_col)
    day_idx_h1 = [idx_in_slice(cn) for cn in day_cols_h1]
    day_idx_h2 = [idx_in_slice(cn) for cn in day_cols_h2]

    header = ["№", "ФИО", "Должность", "Табельный №", "ID объекта"] + [str(i) for i in range(1, 32)] + ["Отработано дней", "Отработано часов"]

    last_row = find_last_data_row(ws, START_ROW, ui)
    end_fetch = min(ws.max_row, last_row + 3)
    fetch_total = max(1, end_fetch - START_ROW + 1)

    # Префетч B..AO блоком (с прогрессом)
    if ui:
        ui.set_phase("Загрузка данных…", fetch_total)

    rows_values: List[Tuple] = []
    r_counter = 0
    for row in ws.iter_rows(min_row=START_ROW, max_row=end_fetch,
                            min_col=2, max_col=ao_col, values_only=True):
        rows_values.append(row)
        r_counter += 1
        if ui:
            ui.set_progress(r_counter)
            if ui.is_cancelled():
                raise CancelledError()

    total_rows = len(rows_values)
    clog(f"prefetched rows: {total_rows}")

    out_rows: List[List[Any]] = []
    last_i = (last_row - START_ROW)
    if ui:
        ui.set_phase("Обработка сотрудников…", max(1, last_i + 1))

    for i in range(0, last_i + 1):
        if ui:
            ui.set_progress(i + 1)
            if ui.is_cancelled():
                raise CancelledError()

        row = rows_values[i]
        row_p1 = rows_values[i + 1] if i + 1 < total_rows else None
        row_p2 = rows_values[i + 2] if i + 2 < total_rows else None
        row_p3 = rows_values[i + 3] if i + 3 < total_rows else None

        raw_num = only_digits(row[idx_B] if row else "")
        if not raw_num:
            continue

        fio_raw = row[idx_C] if row else ""
        fio, title = split_fio_and_title(fio_raw)
        tbn = clean_spaces(row[idx_E] if row else "")

        days_num = to_number_value(row[idx_AO] if row else None)
        hrs_num = None
        if i + HOURS_OFFSET < total_rows:
            hrs_num = to_number_value(rows_values[i + HOURS_OFFSET][idx_AO])
        if hrs_num is None and row_p1 is not None:
            hrs_num = to_number_value(row_p1[idx_AO])

        if not (has_letters(fio) and (len(tbn) > 0 or isinstance(days_num, (int, float)))):
            continue

        out = [int(raw_num), fio, title, tbn, ""]  # ID объекта пусто

        # 1..15: коды — row, часы — row+1
        for dj in day_idx_h1:
            code_val = row[dj] if row else None
            hours_val = row_p1[dj] if row_p1 else None
            daily = day_hours_from_values(code_val, hours_val)
            out.append(daily if daily is not None else "")

        # 16..31: коды — row+2, часы — row+3
        for dj in day_idx_h2:
            code_val = row_p2[dj] if row_p2 else None
            hours_val = row_p3[dj] if row_p3 else None
            daily = day_hours_from_values(code_val, hours_val)
            out.append(daily if daily is not None else "")

        out.append(days_num if days_num is not None else "")
        out.append(hrs_num if hrs_num is not None else "")

        out_rows.append(out)

    return header, out_rows

# ===== Постобработка чисел и оформление =====
def parse_num_relaxed(v: Any) -> Optional[float]:
    import re
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        try:
            if float(v) < 1.0:
                return float(v) * 24.0
        except Exception:
            pass
        return float(v)

    s = str(v)

    # a/b[/c...] → сумма
    if "/" in s:
        total = 0.0
        got = False
        for part in s.split("/"):
            n = to_number_value(part)
            if isinstance(n, (int, float)):
                total += float(n)
                got = True
        if got:
            return total

    # h:mm(:ss) → часы
    if ":" in s:
        p = s.split(":")
        if len(p) >= 2:
            try:
                hh = float(re.sub(r"[^\d.+-]", "", p[0]) or 0)
                mm = float(re.sub(r"[^\d.+-]", "", p[1]) or 0)
                ss = float(re.sub(r"[^\d.+-]", "", p[2])) if len(p) >= 3 else 0.0
                return hh + mm/60.0 + ss/3600.0
            except Exception:
                pass

    # нормализуем экзотику и срезаем ЛЮБОЙ хвост из , . и пробелов (включая fullwidth/low-9)
    s = (s.replace("\uFF0C", ",").replace("\uFF0E", ".").replace("\u201A", ",")
           .replace("\u00A0", " ").replace("\u202F", " ").replace("\u2009", " ").replace("\u200A", " ")
           .replace("\u200B", "").replace("\u2060", "").replace("\uFEFF", "")
           .replace("\u200E", "").replace("\u200F", "")
           .replace("\u202A", "").replace("\u202B", "").replace("\u202C", "").replace("\u202D", "").replace("\u202E", "")
           .replace("\u2066", "").replace("\u2067", "").replace("\u2068", "").replace("\u2069", ""))
    s = re.sub(r'[\s,.\uFF0C\uFF0E\u201A]+$', '', s)  # <<< вот эта строка срезает хвост

    if not s:
        return None

    s = s.replace(",", ".")
    m = re.search(r"[-+]?\d+(?:\.\d+)?", s)
    if m:
        try:
            return float(m.group(0))
        except Exception:
            return None
    try:
        return float(s)
    except Exception:
        return None

def fix_numeric_range_py(ws, r1: int, r2: int, c1: int, c2: int):
    """Переписывает текстовые числа в float (убирает '7,', поддерживает 'h:mm', 'a/b')."""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(r, c)
            v = cell.value
            if v is None or v == "":
                continue
            if isinstance(v, (int, float)):
                cell.value = float(v)
                # Сброс формата на General перед финальным форматированием
                cell.number_format = "General"
                continue
            n = parse_num_relaxed(v)
            if isinstance(n, (int, float)):
                cell.value = float(n)
                cell.number_format = "General"

def apply_borders(ws, min_row: int, max_row: int, min_col: int, max_col: int):
    thin = Side(style="thin", color="D9D9D9")
    medium = Side(style="medium", color="808080")
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(r, c)
            left = medium if c == min_col else thin
            right = medium if c == max_col else thin
            top = medium if r == min_row else thin
            bottom = medium if r == max_row else thin
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

def save_result(header: List[str], rows: List[List[Any]], out_path: str):
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = RESULT_SHEET_NAME

    # Заголовок
    ws_out.append(header)
    for cell in ws_out[1]:
        cell.font = Font(bold=True)

    # Данные
    for row in rows:
        ws_out.append(row)

    # Индексы
    day_start_col = 6  # после: №, ФИО, Должность, Табельный №, ID объекта
    total_days_col = day_start_col + 31
    total_hours_col = total_days_col + 1
    last_row = ws_out.max_row
    last_col = total_hours_col

    # ПРЕ-нормализация: сразу «ломаем» текстовые числа (убираем '7,', '40,' и т.п.)
    fix_numeric_range_py(ws_out, 2, last_row, day_start_col, day_start_col + 31 - 1)
    fix_numeric_range_py(ws_out, 2, last_row, total_days_col, total_hours_col)

    # Ширины
    for col_idx in range(1, 6):
        ws_out.column_dimensions[get_column_letter(col_idx)].width = 16 if col_idx in (2, 3) else 12
    for col_idx in range(day_start_col, day_start_col + 31):
        ws_out.column_dimensions[get_column_letter(col_idx)].width = 4.25
    ws_out.column_dimensions[get_column_letter(total_days_col)].width = 12
    ws_out.column_dimensions[get_column_letter(total_hours_col)].width = 14

    # Центрирование везде
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(1, last_row + 1):
        for c in range(1, last_col + 1):
            ws_out.cell(r, c).alignment = center

    # Заморозка шапки
    ws_out.freeze_panes = "A2"

    # Таблица без полос (чтобы дни остались белыми)
    last_col_letter = get_column_letter(last_col)
    table_ref = f"A1:{last_col_letter}{last_row}"
    table = Table(displayName="ResultTable", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws_out.add_table(table)

    # Белая заливка для столбцов дней
    white = PatternFill(fill_type="solid", fgColor="FFFFFF")
    for c in range(day_start_col, day_start_col + 31):
        for r in range(1, last_row + 1):
            ws_out.cell(r, c).fill = white

    # Границы по всей таблице
    apply_borders(ws_out, 1, last_row, 1, last_col)

    # ФИНАЛЬНАЯ нормализация (ещё раз, после таблицы/границ)
    fix_numeric_range_py(ws_out, 2, last_row, day_start_col, day_start_col + 31 - 1)
    fix_numeric_range_py(ws_out, 2, last_row, total_days_col, total_hours_col)

    # Форматы: часы — до 2 знаков, нули НЕ показывать; дни — целые, нули НЕ показывать
    # часы по дням 1–31
    for c in range(day_start_col, day_start_col + 31):
        for r in range(2, last_row + 1):
            ws_out.cell(r, c).number_format = "0.##;-0.##;;"
    # итоги
    for r in range(2, last_row + 1):
        ws_out.cell(r, total_days_col).number_format  = "0;-0;;"
        ws_out.cell(r, total_hours_col).number_format = "0.##;-0.##;;"

    # Сохранение
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(out_path)

def safe_save_result(header, rows, primary_out: Path) -> Path:
    try:
        save_result(header, rows, str(primary_out))
        return primary_out
    except Exception as e:
        log(f"Primary save failed: {e}")
        alt_dir = Path.home() / "Desktop" / "Табель‑конвертер_Результаты"
        alt_dir.mkdir(parents=True, exist_ok=True)
        alt_path = alt_dir / primary_out.name
        save_result(header, rows, str(alt_path))
        return alt_path

# ===== Сервис =====
def latest_file_in_folder(folder: str) -> Optional[str]:
    folder = Path(folder)
    if not folder.exists():
        return None
    cand = []
    for ext in ("*.xlsx", "*.xlsm"):
        cand.extend(folder.glob(ext))
    if not cand:
        return None
    cand.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return str(cand[0])

def pick_candidate_sheet(wb) -> Optional[Any]:
    for ws in wb.worksheets:
        if "табел" in ws.title.lower():
            return ws
    for ws in wb.worksheets:
        if ws.sheet_state == "visible":
            return ws
    return wb.worksheets[0] if wb.worksheets else None

# ===== Запуск =====
def transform_file(file_path: str, out_path: Optional[str] = None):
    ui = None
    try:
        p = Path(file_path)
        ext = p.suffix.lower()
        if ext not in (".xlsx", ".xlsm"):
            msg_error(APP_NAME, f"Выбран файл: {p.name}\nПоддерживаются только .xlsx и .xlsm.\nСохраните исходник как .xlsx.")
            return

        clog(f"Open workbook: {file_path}")
        ui = ProgressUI()
        ui.set_phase("Открытие книги…", 100)
        ui.set_progress(10)

        wb = load_workbook(file_path, data_only=True, read_only=True)
        ui.set_progress(30)

        ws = pick_candidate_sheet(wb)
        if ws is None:
            msg_error(APP_NAME, "Не найден лист для обработки.")
            ui.close()
            return
        clog(f"Sheet: {ws.title}")

        header, rows = transform_sheet(ws, ui)
        if ui.is_cancelled():
            ui.close()
            msg_info(APP_NAME, "Операция отменена пользователем.")
            return

        out_path = out_path or str(p.with_name(p.stem + "_result.xlsx"))

        ui.set_phase("Сохранение результата…", 100)
        saved_to = safe_save_result(header, rows, Path(out_path))
        ui.set_progress(100)
        ui.close()

        msg_info(APP_NAME, f"Результат сохранён:\n{saved_to}\n\nЛог: {LOG_PATH}")
        clog("Done.")

    except CancelledError:
        if ui:
            ui.close()
        msg_info(APP_NAME, "Операция отменена пользователем.")
    except Exception as e:
        if ui:
            ui.close()
        import traceback
        tb = traceback.format_exc()
        log(tb)
        msg_error(APP_NAME, f"{e}\n\nПодробности в логе:\n{LOG_PATH}")

def main():
    parser = argparse.ArgumentParser(description="Преобразование табеля (1С ЗУП) в читаемую таблицу")
    g = parser.add_mutually_exclusive_group(required=False)
    g.add_argument("--file", help="Путь к файлу табеля (xlsx/xlsm)")
    g.add_argument("--pick", action="store_true", help="Выбрать файл через диалог")
    g.add_argument("--latest", help="Взять самый свежий файл из указанной папки")
    parser.add_argument("--out", help="Путь для сохранения результата (xlsx)")
    args = parser.parse_args()

    # Если аргументов нет — показываем приветственное окно
    if not any([args.file, args.pick, args.latest]):
        welcome = WelcomeUI()
        choice = welcome.run()
        if not choice:
            msg_info(APP_NAME, "Файл не выбран.")
            return
        mode, payload = choice
        if mode == "file":
            transform_file(payload, args.out)
        elif mode == "latest":
            fp = latest_file_in_folder(payload)
            if not fp:
                msg_error(APP_NAME, "В папке не найден подходящий файл (*.xlsx, *.xlsm).")
                return
            transform_file(fp, args.out)
        return

    # CLI‑режимы
    if args.file:
        transform_file(args.file, args.out)
    elif args.pick:
        fp = filedialog.askopenfilename(
            title="Выберите файл табеля",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")]
        )
        if not fp:
            msg_info(APP_NAME, "Файл не выбран.")
            return
        transform_file(fp, args.out)
    elif args.latest:
        fp = latest_file_in_folder(args.latest)
        if not fp:
            msg_error(APP_NAME, "В папке не найден подходящий файл (*.xlsx, *.xlsm).")
            return
        transform_file(fp, args.out)

if __name__ == "__main__":
    main()


