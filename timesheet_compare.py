# timesheet_compare.py
from __future__ import annotations

import os
import tempfile
import logging
import threading
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

import timesheet_transformer

from timesheet_module import (
    load_all_timesheet_headers,
    load_timesheet_rows_by_header_id,
    month_name_ru,
    month_days,
    set_db_pool as _set_db_pool_from_timesheet,
)

# Глобальный пул БД для прямых запросов командировок
DB_POOL = None

def set_db_pool(pool):
    global DB_POOL
    DB_POOL = pool
    _set_db_pool_from_timesheet(pool)


# ============================================================
#  Цветовая схема
# ============================================================
CMP_COLORS = {
    "bg":            "#f0f2f5",
    "panel":         "#ffffff",
    "accent":        "#1565c0",
    "accent_light":  "#e3f2fd",
    "warning":       "#b00020",
    "border":        "#dde1e7",
    "btn_save_bg":   "#1565c0",
    "btn_save_fg":   "#ffffff",
    "row_obj_ok":    "#ffffff",
    "row_obj_diff":  "#ffcdd2",   # красноватый — расхождение
    "row_1c":        "#f5f5f5",   # серый — строка 1С
    "row_only_obj":  "#fff9c4",   # жёлтый — только в объекте
    "row_only_1c":   "#ffe0b2",   # оранжевый — только в 1С
    "cell_diff":     "#ef9a9a",   # ячейка с расхождением
}


# ============================================================
#  Утилиты
# ============================================================

def normalize_tbn(val: Any) -> str:
    s = str(val or "").strip()
    return s.lstrip("0") if s else ""

def fio_sort_key(fio: Any) -> str:
    s = str(fio or "").strip().lower()
    s = " ".join(s.split())
    s = s.replace("ё", "е")
    return s

def get_number_value(val: Any) -> Optional[float]:
    """Пытается получить числовое значение (часы) из ячейки."""
    if val is None or str(val).strip() == "":
        return None
    try:
        n = timesheet_transformer.to_number_value(val)
        return float(n) if n is not None else None
    except Exception:
        return None

def normalize_text(val: Any) -> str:
    """Возвращает строку (например 'В', 'К', 'ОТ') в нижнем регистре."""
    s = str(val or "").strip().lower().replace(",", ".")
    s = " ".join(s.split())
    return "" if s == "none" else s


# ============================================================
#  Главная страница
# ============================================================

class TimesheetComparePage(tk.Frame):

    def __init__(self, master, app_ref):
        super().__init__(master, bg=CMP_COLORS["bg"])
        self.app_ref = app_ref

        self._headers:       List[Dict[str, Any]] = []
        self._obj_rows:      List[Dict[str, Any]] = []
        self._hr_rows:       List[Dict[str, Any]] = []
        self._merged_groups: List[Dict[str, Any]] = []
        self._agg_headers:   List[Dict[str, Any]] = []

        # Состояние текущего месяца
        self._current_month_days = 31

        # Статистика
        self._stat_total    = 0
        self._stat_diff     = 0
        self._stat_only_obj = 0
        self._stat_only_1c  = 0

        self._build_ui()
        self._load_headers()

    # ──────────────────────────────────────────────────────────
    #  UI
    # ──────────────────────────────────────────────────────────

    def _build_ui(self):
        # ── Заголовок ─────────────────────────────────────────
        hdr = tk.Frame(self, bg=CMP_COLORS["accent"], pady=6)
        hdr.pack(fill="x")
        
        tk.Label(
            hdr, 
            text="🔍  Сравнение табелей (Объекты + Командировки vs 1С)",
            font=("Segoe UI", 12, "bold"),
            bg=CMP_COLORS["accent"], 
            fg="white", 
            padx=12
        ).pack(side="left")

        self.lbl_hr_status = tk.Label(
            hdr, 
            text="1С: не загружен",
            font=("Segoe UI", 9),
            bg=CMP_COLORS["accent"], 
            fg="#ffcc80", 
            padx=10
        )
        self.lbl_hr_status.pack(side="right")

        # ── Панель фильтров ───────────────────────────────────
        filter_pnl = tk.LabelFrame(
            self, 
            text=" 🔍 Фильтры ",
            font=("Segoe UI", 9, "bold"),
            bg=CMP_COLORS["panel"], 
            fg=CMP_COLORS["accent"],
            relief="groove", 
            bd=1, 
            padx=10, 
            pady=8
        )
        filter_pnl.pack(fill="x", padx=10, pady=(8, 4))
        filter_pnl.grid_columnconfigure(1, weight=0)
        filter_pnl.grid_columnconfigure(3, weight=0)
        filter_pnl.grid_columnconfigure(5, weight=1)

        # Год
        tk.Label(
            filter_pnl, 
            text="Год:", 
            font=("Segoe UI", 9), 
            bg=CMP_COLORS["panel"]
        ).grid(row=0, column=0, sticky="e", padx=(0, 6), pady=3)
        
        self.var_year = tk.StringVar(value=str(datetime.now().year))
        tk.Spinbox(
            filter_pnl, 
            from_=2000, 
            to=2100, 
            width=7,
            textvariable=self.var_year, 
            font=("Segoe UI", 9)
        ).grid(row=0, column=1, sticky="w", pady=3)

        # Месяц
        tk.Label(
            filter_pnl, 
            text="Месяц:", 
            font=("Segoe UI", 9), 
            bg=CMP_COLORS["panel"]
        ).grid(row=0, column=2, sticky="e", padx=(16, 6), pady=3)
        
        self.var_month = tk.StringVar(value="Все")
        cmb_m = ttk.Combobox(
            filter_pnl, 
            state="readonly", 
            width=14,
            textvariable=self.var_month,
            values=["Все"] + [month_name_ru(i) for i in range(1, 13)]
        )
        cmb_m.grid(row=0, column=3, sticky="w", pady=3)
        cmb_m.bind("<<ComboboxSelected>>", lambda e: self._load_headers())

        # Подразделение
        tk.Label(
            filter_pnl, 
            text="Подразделение:", 
            font=("Segoe UI", 9), 
            bg=CMP_COLORS["panel"]
        ).grid(row=0, column=4, sticky="e", padx=(16, 6), pady=3)
        
        self.var_dep = tk.StringVar(value="Все")
        self.cmb_dep = ttk.Combobox(
            filter_pnl, 
            state="readonly", 
            width=36,
            textvariable=self.var_dep, 
            values=["Все"]
        )
        self.cmb_dep.grid(row=0, column=5, sticky="ew", pady=3)
        self.cmb_dep.bind("<<ComboboxSelected>>", lambda e: self._load_headers())

        # Кнопки панели фильтров
        btn_f = tk.Frame(filter_pnl, bg=CMP_COLORS["panel"])
        btn_f.grid(row=0, column=6, sticky="e", padx=(16, 0))

        tk.Button(
            btn_f, 
            text="🔄 Обновить",
            font=("Segoe UI", 9, "bold"),
            bg=CMP_COLORS["btn_save_bg"], 
            fg=CMP_COLORS["btn_save_fg"],
            activebackground="#0d47a1", 
            activeforeground="white",
            relief="flat", 
            cursor="hand2", 
            padx=10, 
            pady=3,
            command=self._load_headers
        ).pack(side="left", padx=(0, 4))

        ttk.Button(
            btn_f, 
            text="Сбросить", 
            command=self._reset_filters
        ).pack(side="left", padx=(0, 4))

        tk.Button(
            btn_f, 
            text="📂 Загрузить 1С (xlsx)…",
            font=("Segoe UI", 9, "bold"), 
            bg="#2e7d32", 
            fg="white",
            activebackground="#1b5e20", 
            activeforeground="white",
            relief="flat", 
            cursor="hand2", 
            padx=10, 
            pady=3,
            command=self._load_hr_from_1c
        ).pack(side="left")

        # ── Секция 1: выбор периода/подразделения ─────────────
        sel_pnl = tk.LabelFrame(
            self, 
            text=" 1️⃣  Выбор табеля для сравнения ",
            font=("Segoe UI", 9, "bold"), 
            bg=CMP_COLORS["panel"], 
            fg=CMP_COLORS["accent"],
            relief="groove", 
            bd=1
        )
        sel_pnl.pack(fill="x", padx=10, pady=(4, 2))

        sel_top = tk.Frame(sel_pnl, bg=CMP_COLORS["panel"])
        sel_top.pack(fill="x", padx=8, pady=(4, 2))

        tk.Label(
            sel_top, 
            text="Двойной щелчок или кнопка «Сравнить» — загрузить данные:",
            font=("Segoe UI", 8), 
            fg="#555", 
            bg=CMP_COLORS["panel"]
        ).pack(side="left")

        tk.Button(
            sel_top, 
            text="▶  Сравнить выбранное",
            font=("Segoe UI", 9, "bold"), 
            bg=CMP_COLORS["btn_save_bg"], 
            fg=CMP_COLORS["btn_save_fg"],
            activebackground="#0d47a1", 
            activeforeground="white",
            relief="flat", 
            cursor="hand2", 
            padx=10, 
            pady=3,
            command=self._on_select_header
        ).pack(side="right", padx=(0, 4))

        self.btn_export = tk.Button(
            sel_top, 
            text="📊 Экспорт в Excel",
            font=("Segoe UI", 9, "bold"), 
            bg="#2e7d32", 
            fg="white",
            activebackground="#1b5e20", 
            activeforeground="white",
            relief="flat", 
            cursor="hand2", 
            padx=10, 
            pady=3,
            command=self._start_export_thread
        )
        self.btn_export.pack(side="right", padx=(0, 8))

        # Таблица доступных периодов
        hdr_tbl = tk.Frame(sel_pnl, bg=CMP_COLORS["panel"])
        hdr_tbl.pack(fill="x", padx=8, pady=(0, 6))

        cols_h = ("year", "month", "department", "obj_count")
        self.tree_headers = ttk.Treeview(
            hdr_tbl, 
            columns=cols_h, 
            show="headings", 
            height=4, 
            selectmode="browse"
        )
        self.tree_headers.heading("year",       text="Год")
        self.tree_headers.heading("month",      text="Месяц")
        self.tree_headers.heading("department", text="Подразделение")
        self.tree_headers.heading("obj_count",  text="Объектов")

        self.tree_headers.column("year",       width=60,  anchor="center", stretch=False)
        self.tree_headers.column("month",      width=110, anchor="center", stretch=False)
        self.tree_headers.column("department", width=360, anchor="w")
        self.tree_headers.column("obj_count",  width=80,  anchor="center", stretch=False)

        vsb_h = ttk.Scrollbar(hdr_tbl, orient="vertical", command=self.tree_headers.yview)
        self.tree_headers.configure(yscrollcommand=vsb_h.set)
        
        self.tree_headers.pack(side="left", fill="x", expand=True)
        vsb_h.pack(side="right", fill="y")
        
        self.tree_headers.bind("<Double-1>", lambda e: self._on_select_header())

        # ── Секция 2: результат + статистика ──────────────────
        cmp_outer = tk.Frame(self, bg=CMP_COLORS["bg"])
        cmp_outer.pack(fill="both", expand=True, padx=10, pady=(4, 4))
        cmp_outer.grid_rowconfigure(0, weight=1)
        cmp_outer.grid_columnconfigure(0, weight=1)

        cmp_pnl = tk.LabelFrame(
            cmp_outer, 
            text=" 2️⃣  Результат сравнения ",
            font=("Segoe UI", 9, "bold"), 
            bg=CMP_COLORS["panel"], 
            fg=CMP_COLORS["accent"],
            relief="groove", 
            bd=1
        )
        cmp_pnl.grid(row=0, column=0, sticky="nsew")

        cmp_tool = tk.Frame(cmp_pnl, bg=CMP_COLORS["accent_light"], pady=4)
        cmp_tool.pack(fill="x")

        # Выбор режима (периода)
        tk.Label(
            cmp_tool, 
            text="Период:", 
            font=("Segoe UI", 9, "bold"),
            bg=CMP_COLORS["accent_light"], 
            fg=CMP_COLORS["accent"]
        ).pack(side="left", padx=(8, 2))

        self.var_compare_mode = tk.StringVar(value="Весь месяц")
        cmb_mode = ttk.Combobox(
            cmp_tool, 
            state="readonly", 
            width=22,
            textvariable=self.var_compare_mode,
            values=["Весь месяц", "Первая половина (1-15)"]
        )
        cmb_mode.pack(side="left", padx=(0, 6))
        cmb_mode.bind("<<ComboboxSelected>>", lambda e: self._rebuild_comparison())

        tk.Frame(cmp_tool, bg=CMP_COLORS["border"], width=1).pack(side="left", fill="y", padx=6)

        # Чекбокс расхождений
        self.var_only_diff = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            cmp_tool, 
            text="Только расхождения", 
            variable=self.var_only_diff,
            command=self._rebuild_comparison
        ).pack(side="left", padx=(6, 4))

        tk.Frame(cmp_tool, bg=CMP_COLORS["border"], width=1).pack(side="left", fill="y", padx=6)

        # Легенда
        for color, label in [
            ("#ffcdd2", "Расхождение"), 
            ("#ffe0b2", "Только в 1С"),
            ("#fff9c4", "Только в объекте"), 
            ("#f5f5f5", "Строка 1С")
        ]:
            dot = tk.Label(cmp_tool, text="  ", bg=color, relief="solid", bd=1)
            dot.pack(side="left", padx=(4, 2))
            tk.Label(
                cmp_tool, 
                text=label, 
                font=("Segoe UI", 8), 
                bg=CMP_COLORS["accent_light"]
            ).pack(side="left", padx=(0, 6))

        # Статистика
        self.lbl_stat = tk.Label(
            cmp_tool, 
            text="", 
            font=("Segoe UI", 9, "bold"),
            fg=CMP_COLORS["accent"], 
            bg=CMP_COLORS["accent_light"]
        )
        self.lbl_stat.pack(side="right", padx=12)

        # Таблица результата
        tree_cont = tk.Frame(cmp_pnl, bg=CMP_COLORS["panel"])
        tree_cont.pack(fill="both", expand=True)
        tree_cont.grid_rowconfigure(0, weight=1)
        tree_cont.grid_columnconfigure(0, weight=1)

        self.tree_compare = ttk.Treeview(
            tree_cont, 
            show="headings", 
            selectmode="browse"
        )
        
        vsb_c = ttk.Scrollbar(tree_cont, orient="vertical", command=self.tree_compare.yview)
        hsb_c = ttk.Scrollbar(tree_cont, orient="horizontal", command=self.tree_compare.xview)
        
        self.tree_compare.configure(yscrollcommand=vsb_c.set, xscrollcommand=hsb_c.set)
        
        self.tree_compare.grid(row=0, column=0, sticky="nsew")
        vsb_c.grid(row=0, column=1, sticky="ns")
        hsb_c.grid(row=1, column=0, sticky="ew")

        # Теги для цветов строк
        self.tree_compare.tag_configure("obj_ok",   background=CMP_COLORS["row_obj_ok"])
        self.tree_compare.tag_configure("obj_diff", background=CMP_COLORS["row_obj_diff"])
        self.tree_compare.tag_configure("hr_row",   background=CMP_COLORS["row_1c"])
        self.tree_compare.tag_configure("only_obj", background=CMP_COLORS["row_only_obj"])
        self.tree_compare.tag_configure("only_1c",  background=CMP_COLORS["row_only_1c"])

        self._configure_compare_columns(31)

        # ── Статус-бар ────────────────────────────────────────
        status_bar = tk.Frame(self, bg=CMP_COLORS["panel"], relief="sunken", bd=1)
        status_bar.pack(side="bottom", fill="x")

        self.var_status = tk.StringVar(value="Готов к работе")
        tk.Label(
            status_bar, 
            textvariable=self.var_status, 
            anchor="w", 
            font=("Segoe UI", 9),
            bg=CMP_COLORS["panel"]
        ).pack(side="left", padx=6)

        self.progress_bar = ttk.Progressbar(
            status_bar, 
            orient="horizontal", 
            mode="determinate", 
            length=240
        )
        self.progress_bar.pack(side="right", padx=6, pady=2)

    # ──────────────────────────────────────────────────────────
    #  Вспомогательные методы UI
    # ──────────────────────────────────────────────────────────

    def _configure_compare_columns(self, days_in_month: int):
        cols = (
            ["status", "fio", "tbn", "object", "kind"]
            + [f"d{i}" for i in range(1, days_in_month + 1)]
            + ["total_obj", "total_1c"]
        )
        self.tree_compare["columns"] = cols

        self.tree_compare.heading("status", text="")
        self.tree_compare.column("status", width=26, anchor="center", stretch=False)
        
        self.tree_compare.heading("fio", text="ФИО")
        self.tree_compare.column("fio", width=200, minwidth=140)
        
        self.tree_compare.heading("tbn", text="Таб.№")
        self.tree_compare.column("tbn", width=65, anchor="center", stretch=False)
        
        self.tree_compare.heading("object", text="Объект")
        self.tree_compare.column("object", width=240)
        
        self.tree_compare.heading("kind", text="Источник")
        self.tree_compare.column("kind", width=90, anchor="center", stretch=False)

        for i in range(1, days_in_month + 1):
            col = f"d{i}"
            self.tree_compare.heading(col, text=str(i))
            self.tree_compare.column(col, width=34, anchor="center", stretch=False)

        self.tree_compare.heading("total_obj", text="∑ Об.")
        self.tree_compare.column("total_obj", width=52, anchor="center", stretch=False)
        
        self.tree_compare.heading("total_1c", text="∑ 1С")
        self.tree_compare.column("total_1c", width=52, anchor="center", stretch=False)

    def _update_stat_label(self):
        parts = [f"Всего: {self._stat_total}"]
        if self._stat_diff: 
            parts.append(f"⚠ Расхождений: {self._stat_diff}")
        if self._stat_only_obj: 
            parts.append(f"Только объект: {self._stat_only_obj}")
        if self._stat_only_1c: 
            parts.append(f"Только 1С: {self._stat_only_1c}")
        
        self.lbl_stat.config(text="  |  ".join(parts))

    def _fill_departments_combo(self, headers):
        deps = sorted({
            (h.get("department") or "").strip() 
            for h in headers if h.get("department")
        })
        vals = ["Все"] + deps
        self.cmb_dep.configure(values=vals)
        if self.var_dep.get() not in vals:
            self.var_dep.set("Все")

    # ──────────────────────────────────────────────────────────
    #  Загрузка данных
    # ──────────────────────────────────────────────────────────

    def _reset_filters(self):
        self.var_year.set(str(datetime.now().year))
        self.var_month.set("Все")
        self.var_dep.set("Все")
        self._load_headers()

    def _load_headers(self):
        self.tree_headers.delete(*self.tree_headers.get_children())
        self._headers.clear()

        try:
            y = int(self.var_year.get().strip())
        except Exception:
            y = None

        m_name = self.var_month.get().strip()
        m = None
        if m_name and m_name != "Все":
            for i in range(1, 13):
                if month_name_ru(i) == m_name:
                    m = i
                    break

        d = self.var_dep.get().strip()
        dep = d if d and d != "Все" else None

        try:
            headers = load_all_timesheet_headers(
                year=y, month=m, department=dep,
                object_addr_substr=None, object_id_substr=None
            )
        except Exception as e:
            logging.exception("Load headers error")
            messagebox.showerror("Ошибка", f"Не удалось загрузить список:\n{e}", parent=self)
            return

        self._headers = headers
        self._fill_departments_combo(headers)

        agg_map: Dict[Tuple, Dict] = {}
        for h in headers:
            key = (int(h["year"]), int(h["month"]), (h.get("department") or "").strip())
            if key not in agg_map:
                agg_map[key] = {
                    "year": key[0], 
                    "month": key[1], 
                    "department": key[2], 
                    "headers": []
                }
            agg_map[key]["headers"].append(h)

        self._agg_headers = sorted(
            agg_map.values(), 
            key=lambda a: (a["year"], a["month"], a["department"]), 
            reverse=True
        )

        for agg in self._agg_headers:
            iid = f"{agg['year']}:{agg['month']}:{agg['department']}"
            self.tree_headers.insert(
                "", "end", 
                iid=iid, 
                values=(
                    agg["year"], 
                    month_name_ru(agg["month"]), 
                    agg["department"], 
                    len(agg["headers"])
                )
            )

        self.var_status.set(f"Загружено периодов: {len(self._agg_headers)}")

    def _load_hr_from_1c(self):
        path = filedialog.askopenfilename(
            title="Табель 1С (xlsx)", 
            filetypes=[("Excel", "*.xlsx *.xlsm")]
        )
        if not path: 
            return

        try:
            fd, temp_path = tempfile.mkstemp(suffix=".xlsx", prefix="1c_converted_")
            os.close(fd)

            self.var_status.set("Конвертация файла 1С…")
            self.update_idletasks()
            
            timesheet_transformer.transform_file(path, temp_path, parent=self)

            wb = load_workbook(temp_path, data_only=True)
            ws = wb.active
            rows = []
            
            for r in range(2, ws.max_row + 1):
                fio = str(ws.cell(r, 2).value or "").strip()
                if not fio: 
                    continue
                tbn  = str(ws.cell(r, 4).value or "").strip()
                days = [ws.cell(r, c).value for c in range(6, 6 + 31)]
                
                rows.append({"fio": fio, "tbn": tbn, "days": days})

            self._hr_rows = rows
            
            try: 
                os.remove(temp_path)
            except Exception: 
                pass

            self.var_status.set(f"Загружен табель 1С: {len(rows)} сотрудников")
            self.lbl_hr_status.config(text=f"1С: {len(rows)} чел. ✓", fg="#a5d6a7")
            
            self._rebuild_comparison()

        except Exception as e:
            logging.exception("1C Load Error")
            messagebox.showerror("Ошибка", f"Сбой загрузки 1С:\n{e}", parent=self)
            self.var_status.set("Ошибка загрузки 1С")

    def _on_select_header(self):
        sel = self.tree_headers.selection()
        if not sel: 
            return

        agg = next(
            (a for a in self._agg_headers 
             if f"{a['year']}:{a['month']}:{a['department']}" == sel[0]), 
            None
        )
        
        if not agg: 
            return

        self.var_status.set("Загрузка данных из БД…")
        self.update_idletasks()

        obj_rows = []
        try:
            # 1. Загрузка стандартных (объектных) табелей
            for h in agg["headers"]:
                hid = int(h["id"])
                obj_name = (h.get("object_addr") or "").strip()
                oid = (h.get("object_id") or "").strip()
                if oid: 
                    obj_name = f"[{oid}] {obj_name}"

                for r in load_timesheet_rows_by_header_id(hid):
                    raw  = r.get("hours_raw") or []
                    obj_rows.append({
                        "fio": (r["fio"] or "").strip(),
                        "tbn": (r["tbn"] or "").strip(),
                        "object_display": obj_name,
                        "days": list(raw[:31]),
                    })

            # 2. Загрузка КОМАНДИРОВОЧНЫХ табелей напрямую через SQL
            global DB_POOL
            if DB_POOL:
                conn = DB_POOL.getconn()
                try:
                    with conn.cursor() as cur:
                        dep_param = agg["department"]
                        params = [agg["year"], agg["month"]]
                        dep_sql = ""
                        
                        # Если выбрано конкретное подразделение, фильтруем по нему через таблицу employees
                        if dep_param and dep_param != "Все":
                            dep_sql = " AND d.name = %s "
                            params.append(dep_param)

                        sql = f"""
                            SELECT 
                                th.object_addr, 
                                th.object_id, 
                                tr.fio, 
                                tr.tbn, 
                                tr.hours_raw
                            FROM trip_timesheet_headers th
                            JOIN trip_timesheet_rows tr ON th.id = tr.header_id
                            LEFT JOIN employees e ON e.tbn = tr.tbn
                            LEFT JOIN departments d ON d.id = e.department_id
                            WHERE th.year = %s 
                              AND th.month = %s 
                              {dep_sql}
                        """
                        cur.execute(sql, tuple(params))
                        
                        for row in cur.fetchall():
                            obj_addr, obj_id, fio, tbn, hours_raw = row
                            
                            obj_name = (obj_addr or "").strip()
                            if obj_id:
                                obj_name = f"[{obj_id}] {obj_name}"
                            
                            obj_rows.append({
                                "fio": (fio or "").strip(),
                                "tbn": (tbn or "").strip(),
                                "object_display": f"✈ [Командировка] {obj_name}",
                                "days": list(hours_raw[:31]) if hours_raw else [],
                            })
                except Exception as e:
                    logging.exception("Ошибка загрузки командировочных табелей")
                finally:
                    DB_POOL.putconn(conn)

        except Exception as e:
            messagebox.showerror("Ошибка БД", str(e), parent=self)
            return

        self._obj_rows = obj_rows
        self._current_month_days = month_days(agg["year"], agg["month"])
        self._rebuild_comparison()

    # ──────────────────────────────────────────────────────────
    #  Сравнение
    # ──────────────────────────────────────────────────────────

    def _sum_days(self, days: List, count: int) -> float:
        """Суммирует только числовые значения за указанное количество дней."""
        total = 0.0
        for d in days[:count]:
            val = get_number_value(d)
            if val is not None:
                total += val
        return total

    def _render_compare_from_groups(self):
        """Отрисовывает таблицу результатов на основе подготовленных групп."""
        self.tree_compare.delete(*self.tree_compare.get_children())
        
        # Получаем количество дней (столбцов)
        days_count = len(self.tree_compare["columns"]) - 7
        
        def _fmt(v: float) -> str:
            return f"{v:.1f}".rstrip("0").rstrip(".") if v else ""
    
        for grp in self._merged_groups:
            situation   = grp.get("situation", "both")
            has_diff    = grp.get("has_diff", False)
            daily_diffs = grp.get("daily_diffs", [])
            hr_row      = grp.get("hr_row")
            obj_rows    = grp.get("obj_rows", [])

            # Определение цвета и иконки для группы
            if situation == "only_obj":
                tag_obj = "only_obj"
                icon    = "⚪"
            elif situation == "only_1c":
                tag_obj = "only_1c"
                icon    = "🟠"
            elif has_diff:
                tag_obj = "obj_diff"
                icon    = "🔴"
            else:
                tag_obj = "obj_ok"
                icon    = "🟢"

            # Суммы по группе
            sum_obj = sum(self._sum_days(o["days"], days_count) for o in obj_rows)
            sum_1c  = self._sum_days(hr_row["days"], days_count) if hr_row else 0.0
    
            # Отрисовка строк объектов/командировок
            first = True
            for o_row in obj_rows:
                vals = [
                    icon if first else "", 
                    grp["display_fio"] if first else "", 
                    grp["display_tbn"] if first else "", 
                    o_row["object_display"], 
                    "Объект"
                ]
                
                day_vals = []
                for i in range(days_count):
                    raw_v = o_row["days"][i] if i < len(o_row["days"]) else None
                    display = str(raw_v) if raw_v is not None and str(raw_v).strip() != "None" else ""
                    
                    # Если в этот день есть расхождение между 1С и агрегированной суммой объектов
                    if situation == "both" and i < len(daily_diffs) and daily_diffs[i] and display:
                        display = f"≠{display}"
                    
                    day_vals.append(display)
    
                vals += day_vals + [
                    _fmt(sum_obj) if first else "", 
                    _fmt(sum_1c) if first else ""
                ]
                
                self.tree_compare.insert("", "end", values=vals, tags=(tag_obj,))
                first = False
    
            # Отрисовка строки 1С
            if hr_row:
                hr_days_disp = []
                for i in range(days_count):
                    raw_h = hr_row["days"][i] if i < len(hr_row["days"]) else None
                    display = str(raw_h) if raw_h is not None and str(raw_h).strip() != "None" else ""
                    
                    if situation == "both" and i < len(daily_diffs) and daily_diffs[i] and display:
                        display = f"≠{display}"
                    
                    hr_days_disp.append(display)
    
                tag_1c = "only_1c" if situation == "only_1c" else "hr_row"
                self.tree_compare.insert(
                    "", 
                    "end",
                    values=(
                        ["", grp["display_fio"], grp["display_tbn"], "", "1С Кадры"] 
                        + hr_days_disp 
                        + ["", _fmt(sum_1c)]
                    ),
                    tags=(tag_1c,)
                )

    def _rebuild_comparison(self):
        """Пересобирает логику сравнения с учетом агрегации часов по дням."""
        self._merged_groups.clear()

        # Настраиваем колонки: Весь месяц или 1-15
        mode = self.var_compare_mode.get()
        active_days = 15 if mode == "Первая половина (1-15)" else self._current_month_days
        self._configure_compare_columns(active_days)

        if not self._obj_rows and not self._hr_rows: 
            return

        only_diff = self.var_only_diff.get()
        days_count = active_days

        # Словарь 1С по табельному номеру
        hr_map  = {normalize_tbn(r["tbn"]): r for r in self._hr_rows}
        
        # Словарь объектов по табельному номеру
        obj_map = {}
        for r in self._obj_rows:
            tbn_key = normalize_tbn(r["tbn"])
            if tbn_key not in obj_map:
                obj_map[tbn_key] = []
            obj_map[tbn_key].append(r)

        stat_total = 0
        stat_diff = 0
        stat_only_obj = 0
        stat_only_1c = 0

        # Вспомогательная функция для проверки пустых строк
        def check_has_data_in_period(days_list):
            for i in range(days_count):
                if i < len(days_list) and normalize_text(days_list[i]): 
                    return True
            return False

        all_tbns = sorted(set(hr_map.keys()) | set(obj_map.keys()))

        for tbn in all_tbns:
            hr_row = hr_map.get(tbn)
            obj_rows_lst = sorted(
                obj_map.get(tbn, []), 
                key=lambda x: x.get("object_display", "")
            )

            # Проверяем, есть ли вообще данные у человека в выбранном диапазоне дней (1-15 или 1-31)
            has_data = False
            if hr_row and check_has_data_in_period(hr_row["days"]): 
                has_data = True
            
            if not has_data:
                for o in obj_rows_lst:
                    if check_has_data_in_period(o["days"]):
                        has_data = True
                        break
            
            if not has_data: 
                continue # Человек полностью пуст в данном периоде

            has_obj = bool(obj_rows_lst)
            has_1c  = bool(hr_row)

            # Определение глобальной ситуации
            if has_obj and not has_1c: 
                situation = "only_obj"
            elif has_1c and not has_obj: 
                situation = "only_1c"
            else: 
                situation = "both"

            # ── АГРЕГАЦИЯ ДНЕЙ (УМНОЕ СРАВНЕНИЕ) ──────────────────
            # Складываем часы со всех объектов (и командировок) сотрудника по каждому дню
            merged_obj_days = []
            for i in range(days_count):
                num_sum = 0.0
                has_num = False
                first_txt = ""
                
                for o in obj_rows_lst:
                    v = o["days"][i] if i < len(o["days"]) else None
                    n = get_number_value(v)
                    
                    if n is not None:
                        num_sum += n
                        has_num = True
                    elif not first_txt:
                        t = normalize_text(v)
                        if t: 
                            first_txt = t
                
                # Если были числовые часы, записываем сумму, иначе текст (К, В, ОТ и т.д.)
                if has_num:
                    if num_sum.is_integer():
                        merged_obj_days.append(str(int(num_sum)))
                    else:
                        merged_obj_days.append(str(num_sum).rstrip("0").rstrip("."))
                else:
                    merged_obj_days.append(first_txt)

            # Нормализуем строку 1С
            merged_hr_days = []
            for i in range(days_count):
                v = hr_row["days"][i] if hr_row and i < len(hr_row["days"]) else None
                n = get_number_value(v)
                
                if n is not None:
                    if n.is_integer():
                        merged_hr_days.append(str(int(n)))
                    else:
                        merged_hr_days.append(str(n).rstrip("0").rstrip("."))
                else:
                    merged_hr_days.append(normalize_text(v))

            # Сравниваем агрегированные данные объектов с 1С по каждому дню
            group_has_diff = False
            daily_diffs = [False] * days_count
            
            if situation == "both":
                for i in range(days_count):
                    if merged_obj_days[i] != merged_hr_days[i]:
                        daily_diffs[i] = True
                        group_has_diff = True

            # Фильтрация «Только расхождения»
            if only_diff and not (situation in ("only_obj", "only_1c") or group_has_diff):
                continue

            # Подсчет статистики
            stat_total += 1
            if situation == "only_obj": 
                stat_only_obj += 1
            elif situation == "only_1c": 
                stat_only_1c += 1
            elif group_has_diff: 
                stat_diff += 1

            self._merged_groups.append({
                "tbn_key": tbn,
                "display_fio": obj_rows_lst[0]["fio"] if obj_rows_lst else hr_row["fio"],
                "display_tbn": obj_rows_lst[0]["tbn"] if obj_rows_lst else tbn,
                "hr_row": hr_row,
                "obj_rows": obj_rows_lst,
                "situation": situation,
                "has_diff": group_has_diff,
                "daily_diffs": daily_diffs
            })

        # Сортировка и рендер
        self._merged_groups.sort(
            key=lambda g: (
                fio_sort_key(g.get("display_fio")), 
                normalize_tbn(g.get("display_tbn"))
            )
        )
        self._render_compare_from_groups()        
        
        # Обновление статистики
        self._stat_total = stat_total
        self._stat_diff = stat_diff
        self._stat_only_obj = stat_only_obj
        self._stat_only_1c = stat_only_1c
        self._update_stat_label()

        self.var_status.set(
            f"Сравнение завершено — "
            f"сотрудников: {stat_total} | "
            f"расхождений: {stat_diff} | "
            f"только объект: {stat_only_obj} | "
            f"только 1С: {stat_only_1c}"
        )

    # ──────────────────────────────────────────────────────────
    #  Экспорт в Excel
    # ──────────────────────────────────────────────────────────

    def _start_export_thread(self):
        if not self._merged_groups: 
            messagebox.showwarning("Экспорт", "Нет данных для экспорта.", parent=self)
            return

        fpath = filedialog.asksaveasfilename(
            title="Сохранить", 
            defaultextension=".xlsx",
            initialfile=f"Сравнение_табелей_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )
        if not fpath: 
            return

        self.btn_export.configure(state="disabled")
        self.var_status.set("Подготовка к экспорту…")
        self.progress_bar["value"] = 0
        
        threading.Thread(target=self._export_process, args=(fpath,), daemon=True).start()

    def _export_process(self, fpath: str):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Сравнение"

            days_cnt  = len(self.tree_compare["columns"]) - 7

            # Заголовок файла
            ws.append([
                f"Сравнение табелей (вкл. командировки). "
                f"Период: {self.var_compare_mode.get()}. "
                f"Экспорт: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
            ])
            ws.append([
                f"Сотрудников: {self._stat_total} | "
                f"Расхождений: {self._stat_diff} | "
                f"Только объект: {self._stat_only_obj} | "
                f"Только 1С: {self._stat_only_1c}"
            ])
            ws.append([])

            # Шапка
            hdr_row = (
                ["Статус", "ФИО", "Таб.№", "Объект", "Источник"] 
                + [str(i) for i in range(1, days_cnt + 1)] 
                + ["∑ Объект", "∑ 1С"]
            )
            ws.append(hdr_row)
            hdr_excel_row = ws.max_row

            # Стили Excel
            fill_diff     = PatternFill("solid", fgColor="FF9999")
            fill_hr       = PatternFill("solid", fgColor="EFEFEF")
            fill_only_obj = PatternFill("solid", fgColor="FFF9C4")
            fill_only_1c  = PatternFill("solid", fgColor="FFE0B2")
            fill_cell_diff = PatternFill("solid", fgColor="EF9A9A")

            font_bold   = Font(bold=True)
            font_header = Font(bold=True, color="FFFFFF")
            fill_header = PatternFill("solid", fgColor="1565C0")
            
            border_all_thin = Border(
                left=Side(style="thin", color="A0A0A0"), 
                right=Side(style="thin", color="A0A0A0"), 
                top=Side(style="thin", color="A0A0A0"), 
                bottom=Side(style="thin", color="A0A0A0")
            )
            
            align_center = Alignment(horizontal="center", vertical="center")

            # Применение стилей к шапке
            for c in range(1, len(hdr_row) + 1):
                cell = ws.cell(hdr_excel_row, c)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = border_all_thin

            total_grp = len(self._merged_groups)
            
            def _to_excel_val(val):
                if val is None or str(val).strip() == "None" or str(val).strip() == "": 
                    return ""
                try:
                    f = float(str(val).replace(',', '.'))
                    return int(f) if f.is_integer() else f
                except ValueError: 
                    return str(val)

            for idx, grp in enumerate(self._merged_groups):
                # Обновление прогресс-бара
                if total_grp:
                    pct = int(idx / total_grp * 100)
                    self.after(0, lambda v=pct: self.progress_bar.configure(value=v))
                
                situation = grp.get("situation")
                has_diff  = grp.get("has_diff")
                hr_row    = grp.get("hr_row")
                obj_rows  = grp.get("obj_rows", [])
                is_diff   = grp.get("daily_diffs", [])
                
                # Иконки
                if situation == "only_obj":
                    icon = "⚪ Только объект"
                elif situation == "only_1c":
                    icon = "🟠 Только 1С"
                elif has_diff:
                    icon = "🔴 Расхождение"
                else:
                    icon = "🟢 OK"

                sum_obj = sum(self._sum_days(o["days"], days_cnt) for o in obj_rows)
                sum_1c  = self._sum_days(hr_row["days"], days_cnt) if hr_row else 0.0

                start_row = ws.max_row + 1
                first = True
                
                # Экспорт строк объектов
                for o_row in obj_rows:
                    days_raw = (o_row["days"][:days_cnt] + [None] * days_cnt)[:days_cnt]
                    
                    row_data = (
                        [
                            icon if first else "", 
                            grp["display_fio"] if first else "", 
                            grp["display_tbn"] if first else "", 
                            o_row["object_display"], 
                            "Объект"
                        ]
                        + [_to_excel_val(v) for v in days_raw] 
                        + [
                            _to_excel_val(sum_obj) if first else "", 
                            _to_excel_val(sum_1c) if first else ""
                        ]
                    )
                    
                    ws.append(row_data)
                    cur = ws.max_row

                    # Рамки и выравнивание
                    for c_idx in range(1, len(row_data) + 1):
                        cell = ws.cell(cur, c_idx)
                        cell.border = border_all_thin
                        if c_idx >= 6: 
                            cell.alignment = align_center

                    # Подсветка конкретных ячеек-дней
                    if situation == "both":
                        for i in range(days_cnt):
                            if i < len(is_diff) and is_diff[i] and str(days_raw[i]).strip() and str(days_raw[i]) != 'None':
                                c = ws.cell(cur, 6 + i)
                                c.fill = fill_cell_diff
                                c.font = font_bold

                    # Подсветка строки целиком
                    if situation == "only_obj":
                        for c_idx in range(1, len(row_data) + 1): 
                            ws.cell(cur, c_idx).fill = fill_only_obj
                    elif situation == "both" and has_diff:
                        for c_idx in range(1, 6): 
                            ws.cell(cur, c_idx).fill = fill_diff
                            
                    first = False

                # Экспорт строки 1С
                if hr_row:
                    hr_days = (hr_row["days"][:days_cnt] + [None] * days_cnt)[:days_cnt]
                    
                    row_data = (
                        ["", grp["display_fio"], grp["display_tbn"], "", "1С Кадры"] 
                        + [_to_excel_val(v) for v in hr_days] 
                        + ["", _to_excel_val(sum_1c)]
                    )
                    ws.append(row_data)
                    cur = ws.max_row

                    fill_1c = fill_only_1c if situation == "only_1c" else fill_hr
                    for c_idx in range(1, len(row_data) + 1):
                        cell = ws.cell(cur, c_idx)
                        cell.fill = fill_1c
                        cell.border = border_all_thin
                        if c_idx >= 6: 
                            cell.alignment = align_center

                    if situation == "both":
                        for i in range(days_cnt):
                            if i < len(is_diff) and is_diff[i] and str(hr_days[i]).strip() and str(hr_days[i]) != 'None':
                                c = ws.cell(cur, 6 + i)
                                c.fill = fill_cell_diff
                                c.font = font_bold

                # Жирная рамка вокруг всей группы
                end_row = ws.max_row
                if end_row >= start_row:
                    for c_idx in range(1, len(hdr_row) + 1):
                        top_cell = ws.cell(start_row, c_idx)
                        bot_cell = ws.cell(end_row, c_idx)
                        
                        top_cell.border = Border(
                            left=top_cell.border.left, 
                            right=top_cell.border.right, 
                            top=Side(style="medium"), 
                            bottom=top_cell.border.bottom
                        )
                        bot_cell.border = Border(
                            left=bot_cell.border.left, 
                            right=bot_cell.border.right, 
                            top=bot_cell.border.top, 
                            bottom=Side(style="medium")
                        )
                        
                    # Объединение ячеек ФИО и Таб
                    if end_row > start_row:
                        for col_n in (1, 2, 3):
                            ws.merge_cells(
                                start_row=start_row, 
                                start_column=col_n, 
                                end_row=end_row, 
                                end_column=col_n
                            )
                            ws.cell(start_row, col_n).alignment = Alignment(
                                horizontal="left" if col_n == 2 else "center", 
                                vertical="center", 
                                wrap_text=True
                            )

            # Ширина колонок
            ws.column_dimensions["A"].width = 16
            ws.column_dimensions["B"].width = 34
            ws.column_dimensions["C"].width = 10
            ws.column_dimensions["D"].width = 44
            ws.column_dimensions["E"].width = 12
            
            for i in range(days_cnt): 
                ws.column_dimensions[get_column_letter(6 + i)].width = 4.5
                
            ws.column_dimensions[get_column_letter(6 + days_cnt)].width = 9.5
            ws.column_dimensions[get_column_letter(7 + days_cnt)].width = 9.5

            # Автофильтр и закрепление
            ws.auto_filter.ref = f"A{hdr_excel_row}:{get_column_letter(len(hdr_row))}{ws.max_row}"
            ws.freeze_panes = f"A{hdr_excel_row + 1}"

            wb.save(fpath)
            self.after(0, lambda: self._export_finished(True, fpath))

        except Exception as e:
            self.after(0, lambda: self._export_finished(False, str(e)))

    def _export_finished(self, success: bool, msg: str):
        self.btn_export.configure(state="normal")
        self.progress_bar["value"] = 100 if success else 0
        if success:
            self.var_status.set("✅ Экспорт завершён")
            messagebox.showinfo("Готово", f"Файл сохранён:\n{msg}", parent=self)
        else:
            self.var_status.set("❌ Ошибка экспорта")
            messagebox.showerror("Ошибка", f"Не удалось сохранить файл:\n{msg}", parent=self)


def create_timesheet_compare_page(parent, app_ref) -> TimesheetComparePage:
    return TimesheetComparePage(parent, app_ref=app_ref)
